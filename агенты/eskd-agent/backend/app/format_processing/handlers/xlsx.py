"""Извлечение текста из XLSX."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, Decimal):
        return str(value)
    return str(value).strip()


def extract_xlsx_text(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet in wb.worksheets:
            parts.append(f"=== Лист: {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                cells = [_cell_value(v) for v in row]
                if any(cells):
                    parts.append(" | ".join(cells))
    finally:
        wb.close()
    return "\n".join(parts).strip()
