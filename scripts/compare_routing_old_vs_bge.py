"""Сравнение маршрутизации: keyword-RAG (старая) vs BGE (новая) vs факт из 1С."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import selectinload  # noqa: E402

from agent_pochta.config import PROJECT_ROOT, get_settings  # noqa: E402
from agent_pochta.db.message_filters import (  # noqa: E402
    load_payload_dict,
    resolved_turbo_recipient,
)
from agent_pochta.db.models import EmailMessageRow  # noqa: E402
from agent_pochta.db.session import get_session_factory  # noqa: E402
from agent_pochta.routing import RouteEngine, route_email  # noqa: E402
from agent_pochta.routing.recipients import build_routing_search_text  # noqa: E402
from agent_pochta.services.department_knowledge import (  # noqa: E402
    dept_name,
    load_department_display_names,
)
from agent_pochta.services.email_indexing import build_indexing_text_from_row  # noqa: E402
from agent_pochta.services.embedding_client import EmbeddingClientError, embed_texts  # noqa: E402
from agent_pochta.services.email_rag_qdrant import search_department_corrections  # noqa: E402
from agent_pochta.services.rag import score_department_keywords  # noqa: E402
from agent_pochta.services.rag_qdrant import build_rag_service  # noqa: E402

_ILLEGAL_XLSX = re.compile(r"[\000-\010]|\013|\014|\016-\037")


def _xlsx_safe(value: object) -> str:
    text = "" if value is None else str(value)
    cleaned: list[str] = []
    for ch in text:
        code = ord(ch)
        if code in (9, 10, 13) or 32 <= code <= 0xD7FF or 0xE000 <= code <= 0xFFFD:
            cleaned.append(ch)
    return "".join(cleaned)[:32000]


def _format_dept(code: str | None, name: str | None, names: dict[str, str]) -> str:
    code = (code or "").strip()
    if not code:
        return ""
    label = dept_name(names, code, name)
    if label and label != code:
        return f"{code} — {label}"
    return code


def _email_content(row: EmailMessageRow, payload: dict) -> str:
    stored = str(payload.get("embedding_source_text") or "").strip()
    if stored:
        return stored
    return build_indexing_text_from_row(row)


def _old_dept_keyword_rag(
    *,
    rag,
    search_text: str,
    recipient: str,
    row: EmailMessageRow,
    payload: dict,
    names: dict[str, str],
) -> tuple[str, str, str]:
    """Keyword-RAG top-1; fallback — RuleRouter primary."""
    depts = list(getattr(rag, "_departments", {}).values())
    if not depts and hasattr(rag, "refresh_departments_cache"):
        rag.refresh_departments_cache()
        depts = list(getattr(rag, "_departments", {}).values())

    scored = [
        (score_department_keywords(d, search_text, recipient=recipient or None), d)
        for d in depts
    ]
    scored = [(s, d) for s, d in scored if s > 0]
    scored.sort(key=lambda x: x[0], reverse=True)
    if scored:
        dept = scored[0][1]
        return (
            dept.department_id,
            dept.department_name or "",
            f"keyword_rag score={scored[0][0]}",
        )

    from agent_pochta.schemas import EmailMessage

    received = row.received_at
    email = EmailMessage(
        message_id=row.message_id or f"<db-{row.id}>",
        mailbox=row.mailbox or "",
        sender_email=row.sender_email or "",
        sender_name=row.sender_name,
        subject=row.subject or "",
        body_text=str(payload.get("body_text") or "")[:8000],
        received_at=received,
        routing_recipient=str(payload.get("routing_recipient") or row.mailbox or ""),
    )
    decision = route_email(
        email,
        combined_text=search_text,
        recipient=recipient or row.mailbox,
        engine=RouteEngine.load(),
    )
    primary = decision.services[0] if decision.services else None
    if primary:
        return (
            primary.code,
            primary.name or "",
            f"rules {decision.match_source} score={decision.confidence_score}",
        )
    return "", "", "no_match"


def _new_dept_bge(
    embed_text: str,
    *,
    settings,
    names: dict[str, str],
    recipient: str = "",
    top_k: int = 3,
) -> tuple[str, str, float | None, str]:
    """BGE: top-k по department_corrections_bge + голосование dept_correct_id."""
    text = embed_text.strip()
    if not text or not settings.embedding_base_url:
        return "", "", None, "no_embed_url"
    try:
        vector = embed_texts([text], settings=settings)[0]
        hits = search_department_corrections(
            url=settings.qdrant_url,
            query_vector=vector,
            limit=top_k,
            recipient=recipient or None,
        )
    except EmbeddingClientError as exc:
        return "", "", None, f"bge_error: {exc}"

    if not hits:
        return "", "", None, "no_hits"

    votes = Counter(
        str(h["dept_correct_id"])
        for h in hits
        if h.get("dept_correct_id")
    )
    if votes:
        winner_id, count = votes.most_common(1)[0]
        best_hit = next(h for h in hits if str(h.get("dept_correct_id")) == winner_id)
        score = best_hit.get("score")
        name = str(best_hit.get("dept_correct_name") or "")
        return winner_id, name, score, f"vote {count}/{len(hits)}"

    hit = hits[0]
    return (
        str(hit.get("dept_correct_id") or ""),
        str(hit.get("dept_correct_name") or ""),
        hit.get("score"),
        "top1",
    )


def export_comparison(
    *,
    limit: int = 100,
    out_path: Path | None = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    settings = get_settings()
    names = load_department_display_names()
    rag = build_rag_service(settings)
    factory = get_session_factory()

    with factory() as session:
        rows = session.scalars(
            select(EmailMessageRow)
            .where(EmailMessageRow.status == "done")
            .where(EmailMessageRow.erp_document_number.isnot(None))
            .where(EmailMessageRow.erp_document_number != "SKIP-ERP")
            .where(EmailMessageRow.department_id.isnot(None))
            .where(EmailMessageRow.is_spam.is_(False))
            .order_by(EmailMessageRow.processed_at.desc())
            .limit(limit)
            .options(selectinload(EmailMessageRow.attachments))
        ).all()

    out_dir = PROJECT_ROOT / "data" / "stats"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    out_path = out_path or (out_dir / f"routing_old_vs_bge_{stamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение маршрутизации"
    headers = [
        "Письмо (содержимое + вложения)",
        "Отправитель",
        "Получатель (@turbo-don.ru)",
        "Отдел (старая система, keyword-RAG/правила)",
        "Отдел (новая система, BGE)",
        "Отдел (фактически, 1С)",
        "Тема",
        "Номер 1С",
        "Старый = факт",
        "Новый = факт",
        "BGE score",
        "Источник старого",
    ]
    ws.append(headers)

    old_ok = 0
    new_ok = 0
    both_ok = 0
    n = 0

    wrap = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        payload = load_payload_dict(row.raw_payload_json) or {}
        recipient = resolved_turbo_recipient(mailbox=row.mailbox, payload=payload)
        if not recipient:
            recipient = str(payload.get("routing_recipient") or row.mailbox or "")

        content = _email_content(row, payload)
        body = str(payload.get("body_text") or "")
        search_text = build_routing_search_text(
            recipient=recipient,
            subject=row.subject or "",
            body=body or content,
        )

        old_id, old_name, old_src = _old_dept_keyword_rag(
            rag=rag,
            search_text=search_text,
            recipient=recipient,
            row=row,
            payload=payload,
            names=names,
        )
        new_id, new_name, bge_score, _bge_src = _new_dept_bge(
            content, settings=settings, names=names, recipient=recipient
        )

        actual_id = (row.department_id or "").strip()
        old_match = old_id == actual_id if old_id and actual_id else False
        new_match = new_id == actual_id if new_id and actual_id else False
        n += 1
        if old_match:
            old_ok += 1
        if new_match:
            new_ok += 1
        if old_match and new_match:
            both_ok += 1

        ws.append(
            [
                _xlsx_safe(content),
                _xlsx_safe(row.sender_email or ""),
                _xlsx_safe(recipient),
                _xlsx_safe(_format_dept(old_id, old_name, names)),
                _xlsx_safe(_format_dept(new_id, new_name, names)),
                _xlsx_safe(_format_dept(actual_id, row.department_name, names)),
                _xlsx_safe(row.subject or ""),
                _xlsx_safe(row.erp_document_number or ""),
                "да" if old_match else "нет",
                "да" if new_match else "нет",
                round(bge_score, 4) if bge_score is not None else "",
                _xlsx_safe(old_src),
            ]
        )

    for col in ("A", "G"):
        ws.column_dimensions[col].width = 60
    for col in ("B", "C", "D", "E", "F", "H"):
        ws.column_dimensions[col].width = 28
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).alignment = wrap

    ws_sum = wb.create_sheet("Сводка")
    ws_sum.append(["Метрика", "Значение"])
    ws_sum.append(["Писем с 1С (done, не spam)", n])
    ws_sum.append(["Старая система = факт", old_ok])
    ws_sum.append(["Новая BGE = факт", new_ok])
    ws_sum.append(["Обе совпали с фактом", both_ok])
    ws_sum.append(
        ["Accuracy старая %", round(old_ok / n * 100, 1) if n else 0]
    )
    ws_sum.append(
        ["Accuracy новая BGE %", round(new_ok / n * 100, 1) if n else 0]
    )
    ws_sum.append(["QDRANT_URL", settings.qdrant_url])
    ws_sum.append(["EMBEDDING_BASE_URL", settings.embedding_base_url])

    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Compare old keyword-RAG vs BGE routing")
    parser.add_argument("--limit", type=int, default=100)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    path = export_comparison(limit=args.limit, out_path=args.output)
    print(
        json.dumps(
            {"ok": True, "path": str(path), "limit": args.limit},
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
