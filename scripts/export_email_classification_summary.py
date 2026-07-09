"""Сводная таблица: email → организация → направление → отдел.

Источники (объединяются, приоритет у явного файла классификации):
  - Классификация писем.xlsx (если доступен)
  - data/routing_rules.json
  - data/tz_department_topics.json
  - data/enterprise_positions.json (структура 1С, email из ТЗ)

Листы результата:
  - Сводная — email → организация (кратко) → направление → код отдела → название отдела
  - По организациям — группировка по организациям
  - По направлениям — группировка по направлениям

Пример:
  python scripts/export_email_classification_summary.py
  python scripts/export_email_classification_summary.py --source "data/Классификация писем.xlsx"
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent_pochta.services.routing_departments import (  # noqa: E402
    directions_by_code_from_rules,
    load_routing_rules,
    load_tz_department_topics,
    resolve_enterprise_positions_path,
)

DEFAULT_OUTPUT = ROOT / "data" / "email_classification_summary.xlsx"
DEFAULT_SOURCE = Path(
    r"\\192.168.1.198\Files\22.Служба развития\7. Рабочая группа экспертов по ИИ"
    r"\ИИ агенты\2. Входящая корреспонденция\для алгорitma\Классификация писем.xlsx"
)

LOCAL_SOURCE_CANDIDATES = (
    ROOT / "data" / "Классификация писем.xlsx",
    ROOT / "data" / "klassifikatsiya_pisem.xlsx",
    ROOT / "data" / "tz_source" / "Классификация писем.xlsx",
)

# Общие константы и функции из существующих экспортов
_dept_export_path = ROOT / "scripts" / "export_departments_excel.py"
_spec = importlib.util.spec_from_file_location("export_departments_excel", _dept_export_path)
_dept_export = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_dept_export)

ORG_FULL_NAMES: dict[str, str] = _dept_export.ORG_FULL_NAMES
ORG_ORDER: tuple[str, ...] = _dept_export.ORG_ORDER
build_export_rows = _dept_export.build_export_rows
detect_organization = _dept_export.detect_organization

_org_mapping_path = ROOT / "scripts" / "export_organizations_mapping.py"
_org_spec = importlib.util.spec_from_file_location("export_organizations_mapping", _org_mapping_path)
_org_mapping = importlib.util.module_from_spec(_org_spec)
assert _org_spec.loader is not None
_org_spec.loader.exec_module(_org_mapping)

DIRECTION_LABELS: dict[str, str] = _org_mapping.DIRECTION_LABELS
_resolve_org_from_direction = _org_mapping._resolve_org_from_direction
_resolve_org_from_email = _org_mapping._resolve_org_from_email

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "email": ("email", "e-mail", "e mail", "почта", "адрес", "mailbox", "ящик", "электронная почта"),
    "direction": ("direction", "направление", "напр", "сектор"),
    "dept_code": ("код отдела", "код подразделения", "код", "code", "dept_code"),
    "dept_name": (
        "отдел",
        "подразделение",
        "department",
        "dept_name",
        "название отдела",
        "название подразделения",
    ),
    "org": ("организация", "organization", "org", "орг", "код организации"),
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text)


def _match_column(header: str) -> str | None:
    for field, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in header or header == alias:
                return field
    return None


def _extract_emails(text: str) -> list[str]:
    return list(dict.fromkeys(match.group(0).lower() for match in _EMAIL_RE.finditer(text or "")))


def resolve_source_path(explicit: Path | None) -> tuple[Path | None, str]:
    """Возвращает (путь, описание источника)."""
    if explicit and explicit.is_file():
        return explicit, f"явный: {explicit}"

    for candidate in LOCAL_SOURCE_CANDIDATES:
        if candidate.is_file():
            return candidate, f"локальная копия: {candidate}"

    if DEFAULT_SOURCE.is_file():
        return DEFAULT_SOURCE, f"сеть: {DEFAULT_SOURCE}"

    return None, "файл «Классификация писем.xlsx» недоступен — используются routing_rules + tz_topics + 1С"


def parse_classification_xlsx(source_path: Path) -> list[dict]:
    """Читает строки из Excel-классификации с автоопределением колонок."""
    wb = load_workbook(source_path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet in wb.worksheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue

        col_map: dict[int, str] = {}
        for idx, cell in enumerate(header_row):
            field = _match_column(_normalize_header(cell))
            if field:
                col_map[idx] = field

        if "email" not in col_map.values():
            continue

        for row_values in sheet.iter_rows(min_row=2, values_only=True):
            if not row_values:
                continue
            parsed: dict[str, str] = {"source_sheet": sheet.title}
            for idx, field in col_map.items():
                if idx < len(row_values) and row_values[idx] is not None:
                    parsed[field] = str(row_values[idx]).strip()

            emails = _extract_emails(parsed.get("email", ""))
            if not emails and parsed.get("email"):
                local = parsed["email"].lower().strip()
                if local and "@" not in local:
                    emails = [local]

            for email in emails:
                item = dict(parsed)
                item["email"] = email
                rows.append(item)

    wb.close()
    return rows


def _resolve_direction(
    *,
    code: str,
    explicit: str,
    directions: dict[str, str],
    active_by_code: dict[str, dict],
) -> str:
    direction = (explicit or "").strip()
    if direction:
        return direction
    if code in directions:
        return directions[code]
    active = active_by_code.get(code) or {}
    return str(active.get("direction") or "").strip()


def _resolve_org(
    *,
    email: str,
    code: str,
    explicit_org: str,
    direction: str,
    path: str,
    onec_name: str,
) -> str:
    org = (explicit_org or "").strip().upper()
    if org in ORG_FULL_NAMES:
        return org

    email_org = _resolve_org_from_email(email)
    if email_org:
        return email_org

    if direction in ORG_FULL_NAMES:
        return direction

    return detect_organization(path=path, name=onec_name, code=code, direction=direction or None, emails=[email])


def _resolve_dept_name(
    code: str,
    explicit: str,
    dept_names: dict[str, str],
    active_by_code: dict[str, dict],
    *,
    tz_topics: dict[str, dict],
    onec_name: str = "",
) -> str:
    name = (explicit or "").strip()
    if name and name != code:
        return name
    active = active_by_code.get(code) or {}
    if active.get("name") and str(active["name"]) != code:
        return str(active["name"])
    if onec_name:
        return onec_name
    topic = tz_topics.get(code) or {}
    for field in ("names", "topics"):
        values = topic.get(field) or []
        if values:
            return str(values[0])
    return dept_names.get(code, code)


def build_summary_rows(
    *,
    rules: dict,
    source_rows: list[dict] | None = None,
    enterprise_path: Path | None = None,
    tz_topics_path: Path | None = None,
    rules_path: Path | None = None,
) -> tuple[list[dict], dict[str, str]]:
    """Собирает уникальные строки сводной таблицы."""
    dept_names = {str(k): str(v) for k, v in rules.get("department_names", {}).items()}
    directions = directions_by_code_from_rules(rules)
    tz_topics = load_tz_department_topics(tz_topics_path)

    active_rows, _liquidated = build_export_rows(enterprise_path=enterprise_path, rules_path=rules_path)
    active_by_code = {row["code"]: row for row in active_rows}

    enterprise_file = resolve_enterprise_positions_path(enterprise_path)
    enterprise = json.loads(enterprise_file.read_text(encoding="utf-8"))
    structure_by_code: dict[str, dict] = {}
    for row in enterprise.get("structure_departments_with_codes") or []:
        code = str(row.get("code") or "").strip()
        if code:
            structure_by_code[code] = row

    merged: dict[tuple[str, str], dict] = {}
    meta: dict[str, str] = {"sources_used": []}

    def upsert(
        *,
        email: str,
        code: str,
        dept_name: str = "",
        direction: str = "",
        org: str = "",
        data_source: str,
        note: str = "",
    ) -> None:
        email = email.lower().strip()
        code = (code or "").strip()
        if not email or not code:
            return

        structure = structure_by_code.get(code, {})
        path = str(structure.get("path") or "")
        onec_name = str(structure.get("name") or "")
        active = active_by_code.get(code, {})

        resolved_direction = _resolve_direction(
            code=code,
            explicit=direction,
            directions=directions,
            active_by_code=active_by_code,
        )
        resolved_org = _resolve_org(
            email=email,
            code=code,
            explicit_org=org,
            direction=resolved_direction,
            path=path or str(active.get("path") or ""),
            onec_name=onec_name or str(active.get("onec_name") or ""),
        )
        resolved_name = _resolve_dept_name(
            code,
            dept_name,
            dept_names,
            active_by_code,
            tz_topics=tz_topics,
            onec_name=onec_name or str(active.get("onec_name") or ""),
        )

        key = (email, code)
        existing = merged.get(key)
        if existing:
            if not existing.get("direction") and resolved_direction:
                existing["direction"] = resolved_direction
            if existing.get("org_code") == "НП" and resolved_org != "НП":
                existing["org_code"] = resolved_org
                existing["org_name"] = ORG_FULL_NAMES.get(resolved_org, resolved_org)
            if note and note not in existing.get("note", ""):
                existing["note"] = "; ".join(filter(None, [existing.get("note"), note]))
            if data_source not in existing.get("data_sources", []):
                existing.setdefault("data_sources", []).append(data_source)
            return

        merged[key] = {
            "email": email,
            "org_code": resolved_org,
            "org_name": ORG_FULL_NAMES.get(resolved_org, resolved_org),
            "direction": resolved_direction,
            "dept_code": code,
            "dept_name": resolved_name,
            "note": note,
            "data_sources": [data_source],
        }

    # 1. Файл классификации (если есть)
    if source_rows:
        meta["sources_used"].append("Классификация писем.xlsx")
        for row in source_rows:
            code = str(row.get("dept_code") or "").strip()
            if not re.match(r"00-\d{6}", code):
                continue
            upsert(
                email=row["email"],
                code=code,
                dept_name=str(row.get("dept_name") or ""),
                direction=str(row.get("direction") or ""),
                org=str(row.get("org") or ""),
                data_source="classification_xlsx",
                note=f"лист: {row.get('source_sheet', '')}",
            )

    # 2. exact_email_rules
    meta["sources_used"].append("routing_rules.json (exact_email_rules)")
    for rule in rules.get("exact_email_rules") or []:
        email = str(rule.get("email", "")).strip()
        code = str(rule.get("code", "")).strip()
        if not email or not code:
            continue
        direction = str(rule.get("direction", "")).strip()
        upsert(
            email=email,
            code=code,
            dept_name=str(rule.get("name") or dept_names.get(code, code)),
            direction=direction,
            org=_resolve_org_from_direction(direction),
            data_source="exact_email_rule",
            note=str(rule.get("about") or ""),
        )

    # 3. email_keyword_rules (шаблоны)
    meta["sources_used"].append("routing_rules.json (email_keyword_rules)")
    for rule in rules.get("email_keyword_rules") or []:
        keyword = str(rule.get("keyword", "")).strip()
        code = str(rule.get("code", "")).strip()
        if not keyword or not code:
            continue
        direction = str(rule.get("direction", "")).strip()
        upsert(
            email=f"*{keyword}*@turbo-don.ru",
            code=code,
            dept_name=str(rule.get("name") or dept_names.get(code, code)),
            direction=direction,
            org=_resolve_org_from_direction(direction),
            data_source="email_keyword_rule",
            note=f"ключевое слово «{keyword}» в локальной части",
        )

    # 4. tz_department_topics + 1С
    meta["sources_used"].append("tz_department_topics.json + enterprise_positions.json")
    for code, entry in sorted(tz_topics.items()):
        structure = structure_by_code.get(code, {})
        path = str(structure.get("path") or "")
        onec_name = str(structure.get("name") or "")
        for email in entry.get("emails") or []:
            email = str(email).strip()
            if not email:
                continue
            topic_name = ""
            for field in ("names", "topics"):
                values = entry.get(field) or []
                if values:
                    topic_name = str(values[0])
                    break
            upsert(
                email=email,
                code=code,
                dept_name=topic_name or dept_names.get(code, code),
                direction=directions.get(code, ""),
                org="",
                data_source="tz_topics",
                note="; ".join(entry.get("topics") or [])[:120],
            )
            # уточняем org по структуре 1С
            key = (email.lower(), code)
            if key in merged:
                org = detect_organization(path=path, name=onec_name, code=code, emails=[email])
                merged[key]["org_code"] = org
                merged[key]["org_name"] = ORG_FULL_NAMES.get(org, org)

    # 5. Активные подразделения с email из build_export_rows (заполнение пробелов)
    meta["sources_used"].append("build_export_rows (1С + routing)")
    for row in active_rows:
        code = row["code"]
        emails_raw = str(row.get("email") or "")
        for email in _extract_emails(emails_raw):
            upsert(
                email=email,
                code=code,
                dept_name=str(row.get("name") or ""),
                direction=str(row.get("direction") or ""),
                org=str(row.get("org_code") or ""),
                data_source="export_rows",
            )

    result = list(merged.values())
    result.sort(
        key=lambda item: (
            ORG_ORDER.index(item["org_code"]) if item["org_code"] in ORG_ORDER else 99,
            item["direction"] or "ZZ",
            item["dept_code"],
            item["email"],
        )
    )
    return result, meta


def build_by_org_rows(summary_rows: list[dict]) -> list[dict]:
    by_org: dict[str, list[dict]] = defaultdict(list)
    for row in summary_rows:
        by_org[row["org_code"]].append(row)

    output: list[dict] = []
    for org in ORG_ORDER:
        org_rows = by_org.get(org, [])
        if not org_rows:
            continue
        emails = sorted({r["email"] for r in org_rows})
        depts = sorted({f"{r['dept_code']} — {r['dept_name']}" for r in org_rows})
        directions = sorted({r["direction"] for r in org_rows if r["direction"]})
        output.append(
            {
                "org_code": org,
                "org_name": ORG_FULL_NAMES.get(org, org),
                "email_count": len(emails),
                "dept_count": len(depts),
                "directions": ", ".join(directions),
                "emails_sample": "; ".join(emails[:8]) + (f" … (+{len(emails) - 8})" if len(emails) > 8 else ""),
                "departments": "; ".join(depts[:6]) + (f" … (+{len(depts) - 6})" if len(depts) > 6 else ""),
            }
        )

    for org in sorted(set(by_org) - set(ORG_ORDER)):
        org_rows = by_org[org]
        emails = sorted({r["email"] for r in org_rows})
        depts = sorted({f"{r['dept_code']} — {r['dept_name']}" for r in org_rows})
        directions = sorted({r["direction"] for r in org_rows if r["direction"]})
        output.append(
            {
                "org_code": org,
                "org_name": org,
                "email_count": len(emails),
                "dept_count": len(depts),
                "directions": ", ".join(directions),
                "emails_sample": "; ".join(emails[:8]),
                "departments": "; ".join(depts[:6]),
            }
        )
    return output


def build_by_direction_rows(summary_rows: list[dict]) -> list[dict]:
    by_dir: dict[str, list[dict]] = defaultdict(list)
    for row in summary_rows:
        direction = row["direction"] or "—"
        by_dir[direction].append(row)

    direction_order = ["КС", "ПР", "СС", "МС", "АЛ", "МГ", "АМ", "МИ", "БМ", "—"]
    output: list[dict] = []

    def append_group(direction: str, rows: list[dict]) -> None:
        if not rows:
            return
        orgs = sorted({r["org_code"] for r in rows})
        emails = sorted({r["email"] for r in rows})
        depts = sorted({f"{r['dept_code']} — {r['dept_name']}" for r in rows})
        output.append(
            {
                "direction": direction,
                "description": DIRECTION_LABELS.get(direction, direction),
                "org_codes": ", ".join(orgs),
                "email_count": len(emails),
                "dept_count": len(depts),
                "emails_sample": "; ".join(emails[:8]) + (f" … (+{len(emails) - 8})" if len(emails) > 8 else ""),
                "departments": "; ".join(depts[:6]) + (f" … (+{len(depts) - 6})" if len(depts) > 6 else ""),
            }
        )

    for direction in direction_order:
        append_group(direction, by_dir.pop(direction, []))

    for direction in sorted(by_dir):
        append_group(direction, by_dir[direction])

    return output


def _autosize_columns(sheet, *, min_width: int = 10, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[letter].width = width


def _style_header(sheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in sheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(
    output_path: Path,
    *,
    summary_rows: list[dict],
    by_org_rows: list[dict],
    by_direction_rows: list[dict],
    source_note: str,
) -> dict[str, int]:
    wb = Workbook()

    # Лист «Сводная»
    main = wb.active
    main.title = "Сводная"
    headers = [
        "Email",
        "Организация (кратко)",
        "Организация (полное название)",
        "Направление",
        "Код отдела",
        "Название отдела",
        "Источник данных",
        "Примечание",
    ]
    main.append(headers)
    for row in summary_rows:
        main.append(
            [
                row["email"],
                row["org_code"],
                row["org_name"],
                row["direction"],
                row["dept_code"],
                row["dept_name"],
                ", ".join(row.get("data_sources") or []),
                row.get("note") or "",
            ]
        )
    _style_header(main)
    _autosize_columns(main, max_width=55)
    for cell in main["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист «По организациям»
    org_sheet = wb.create_sheet("По организациям")
    org_headers = [
        "Организация (кратко)",
        "Организация (полное название)",
        "Кол-во email",
        "Кол-во отделов",
        "Направления",
        "Примеры email",
        "Отделы",
    ]
    org_sheet.append(org_headers)
    for row in by_org_rows:
        org_sheet.append(
            [
                row["org_code"],
                row["org_name"],
                row["email_count"],
                row["dept_count"],
                row["directions"],
                row["emails_sample"],
                row["departments"],
            ]
        )
    _style_header(org_sheet)
    _autosize_columns(org_sheet, max_width=80)
    for col in ("F", "G"):
        for cell in org_sheet[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист «По направлениям»
    dir_sheet = wb.create_sheet("По направлениям")
    dir_headers = [
        "Направление",
        "Описание",
        "Организации",
        "Кол-во email",
        "Кол-во отделов",
        "Примеры email",
        "Отделы",
    ]
    dir_sheet.append(dir_headers)
    for row in by_direction_rows:
        dir_sheet.append(
            [
                row["direction"],
                row["description"],
                row["org_codes"],
                row["email_count"],
                row["dept_count"],
                row["emails_sample"],
                row["departments"],
            ]
        )
    _style_header(dir_sheet)
    _autosize_columns(dir_sheet, max_width=80)
    for col in ("F", "G"):
        for cell in dir_sheet[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист «Мета»
    meta_sheet = wb.create_sheet("Мета")
    meta_sheet.append(["Параметр", "Значение"])
    meta_sheet.append(["Источник классификации", source_note])
    meta_sheet.append(["Всего строк (Сводная)", len(summary_rows)])
    meta_sheet.append(["Уникальных email", len({r["email"] for r in summary_rows})])
    meta_sheet.append(["Уникальных отделов", len({r["dept_code"] for r in summary_rows})])
    _style_header(meta_sheet)
    _autosize_columns(meta_sheet, max_width=90)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "Сводная": main.max_row - 1,
        "По организациям": org_sheet.max_row - 1,
        "По направлениям": dir_sheet.max_row - 1,
        "Мета": meta_sheet.max_row - 1,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Сводная таблица email → организация → направление → отдел")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Путь к Excel (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=None,
        help="Путь к «Классификация писем.xlsx» (если не задан — сеть или локальная копия)",
    )
    parser.add_argument("--enterprise", type=Path, help="Путь к enterprise_positions.json")
    parser.add_argument("--rules", type=Path, help="Путь к routing_rules.json")
    parser.add_argument("--tz-topics", type=Path, help="Путь к tz_department_topics.json")
    args = parser.parse_args()

    source_path, source_note = resolve_source_path(args.source)
    source_rows: list[dict] | None = None
    if source_path:
        source_rows = parse_classification_xlsx(source_path)
        source_note = f"{source_note}; строк из файла: {len(source_rows)}"

    rules = load_routing_rules(args.rules)
    summary_rows, build_meta = build_summary_rows(
        rules=rules,
        source_rows=source_rows,
        enterprise_path=args.enterprise,
        tz_topics_path=args.tz_topics,
        rules_path=args.rules,
    )

    by_org_rows = build_by_org_rows(summary_rows)
    by_direction_rows = build_by_direction_rows(summary_rows)
    counts = write_excel(
        args.output,
        summary_rows=summary_rows,
        by_org_rows=by_org_rows,
        by_direction_rows=by_direction_rows,
        source_note=source_note,
    )

    print(f"Файл: {args.output.resolve()}")
    print(f"Источник классификации: {source_note}")
    print("Использованные источники:")
    for src in build_meta.get("sources_used", []):
        print(f"  - {src}")
    print("Строки по листам:")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}")
    print(f"Уникальных email: {len({r['email'] for r in summary_rows})}")
    print(f"Уникальных отделов: {len({r['dept_code'] for r in summary_rows})}")


if __name__ == "__main__":
    main()
