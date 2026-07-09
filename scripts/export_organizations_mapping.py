"""Экспорт маппинга организаций и направлений в Excel.

Источники:
  - data/routing_rules.json (organization_keywords, email/content rules)
  - data/enterprise_positions.json (структура 1С)
  - data/tz_department_topics.json (email по подразделениям ТЗ)
  - scripts/export_departments_excel.py (определение организации по структуре)

Пример:
  python scripts/export_organizations_mapping.py
  python scripts/export_organizations_mapping.py -o data/organizations_mapping.xlsx
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import sys
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent_pochta.services.routing_departments import (  # noqa: E402
    load_routing_rules,
    load_tz_department_topics,
    resolve_enterprise_positions_path,
)

DEFAULT_OUTPUT = ROOT / "data" / "organizations_mapping.xlsx"

# Подгружаем общие константы и detect_organization из export_departments_excel
_dept_export_path = ROOT / "scripts" / "export_departments_excel.py"
_spec = importlib.util.spec_from_file_location("export_departments_excel", _dept_export_path)
_dept_export = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_dept_export)

ORG_FULL_NAMES: dict[str, str] = _dept_export.ORG_FULL_NAMES
ORG_ORDER: tuple[str, ...] = _dept_export.ORG_ORDER
build_export_rows = _dept_export.build_export_rows
detect_organization = _dept_export.detect_organization

ORG_DIRECTIONS: dict[str, list[str]] = {
    "НП": ["КС", "ПР", "СС", "МС"],
    "АЛ": ["АЛ"],
    "МГ": ["МГ"],
    "АМ": ["АМ"],
    "МИ": ["МИ"],
    "БМ": ["БМ"],
}

DIRECTION_LABELS: dict[str, str] = {
    "КС": "Коммерческая служба (продажи, тендеры, ОДП, ОРКК)",
    "ПР": "Производственно-ресурсный сектор (бухгалтерия, юристы, кадры, ОТК)",
    "СС": "Сервисная служба (ремонт, обслуживание)",
    "МС": "Метрологический сектор (ОПМУ)",
    "АЛ": "Организация = направление (ООО «Алмаз»)",
    "МГ": "Организация = направление (ООО «Метрогазсервис»)",
    "АМ": "Организация = направление (ООО «Амурская легенда»)",
    "МИ": "Организация = направление (ООО «МИЛАКА»)",
    "БМ": "Блочно-модульные изделия",
}

STRUCTURE_MARKERS: dict[str, str] = {
    "НП": "По умолчанию, если не распознаны другие организации",
    "АЛ": "алмаз, гранд spi, spu-5, бытовых счетчик; email: almaz",
    "МГ": "метрогаз; email: mgs_, mgs@",
    "АМ": "амурск, акваген, легенд",
    "МИ": "милака",
    "БМ": "бми, блочно-модульн; direction=БМ",
}

ORG_FROM_RECIPIENT = (
    ("almaz", "АЛ"),
    ("mgs_", "МГ"),
    ("mgs@", "МГ"),
)

ORG_CODES = frozenset({"АЛ", "МГ", "АМ", "МИ", "БМ"})


def _resolve_org_from_direction(direction: str) -> str:
    direction = (direction or "").strip()
    if direction in ORG_CODES:
        return direction
    return "НП"


def _resolve_org_from_email(email: str) -> str | None:
    local = email.lower().split("@", 1)[0]
    for marker, org in ORG_FROM_RECIPIENT:
        if marker in local:
            return org
    if local.startswith("npo_") or local.startswith("td_"):
        return "НП"
    return None


def build_organizations_rows(rules: dict) -> list[dict]:
    org_keywords = rules.get("organization_keywords") or {}
    rows: list[dict] = []

    for org in ORG_ORDER:
        keywords = list(org_keywords.get(org, []))
        if org == "НП":
            keyword_cell = "— (организация по умолчанию)"
        else:
            keyword_cell = "; ".join(keywords) if keywords else "—"

        directions = ORG_DIRECTIONS[org]
        if org == "НП":
            direction_rule = (
                "Направление берётся из правил маршрутизации: "
                "КС / ПР / СС / МС (detect_direction)"
            )
        elif org == "БМ":
            direction_rule = "Направление всегда «БМ»"
        else:
            direction_rule = f"Направление = код организации («{org}»)"

        rows.append(
            {
                "code": org,
                "full_name": ORG_FULL_NAMES[org],
                "keywords": keyword_cell,
                "structure_markers": STRUCTURE_MARKERS.get(org, ""),
                "direction_rule": direction_rule,
                "directions": ", ".join(directions),
            }
        )
    return rows


def build_directions_rows() -> list[dict]:
    rows: list[dict] = []
    for org in ORG_ORDER:
        for direction in ORG_DIRECTIONS[org]:
            note = ""
            if org == "НП":
                note = "Для НП направление определяется правилом маршрутизации"
            elif org in {"АЛ", "МГ", "АМ", "МИ"}:
                note = "Для дочерних организаций direction = organization (ТЗ §12)"
            elif org == "БМ":
                note = "Отдельное направление БМИ внутри холдинга"

            rows.append(
                {
                    "org_code": org,
                    "org_name": ORG_FULL_NAMES[org],
                    "direction": direction,
                    "description": DIRECTION_LABELS.get(direction, direction),
                    "note": note,
                }
            )
    return rows


def build_departments_by_org_rows(active_rows: list[dict]) -> list[dict]:
    """Упрощённый список активных подразделений, сгруппированный по организации."""
    simplified: list[dict] = []
    for row in active_rows:
        simplified.append(
            {
                "org_code": row["org_code"],
                "org_name": row["org_name"],
                "code": row["code"],
                "name": row["name"],
                "direction": row["direction"],
            }
        )
    return simplified


def build_email_org_rows(
    rules: dict,
    *,
    tz_topics_path: Path | None = None,
    enterprise_path: Path | None = None,
) -> list[dict]:
    dept_names = {str(k): str(v) for k, v in rules.get("department_names", {}).items()}
    rows: list[dict] = []
    seen: set[tuple[str, str, str]] = set()

    def add_row(
        *,
        email_or_pattern: str,
        rule_type: str,
        code: str,
        name: str,
        direction: str,
        org: str,
        note: str = "",
    ) -> None:
        key = (email_or_pattern.lower(), rule_type, code)
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "email_or_pattern": email_or_pattern,
                "rule_type": rule_type,
                "org_code": org,
                "org_name": ORG_FULL_NAMES.get(org, org),
                "direction": direction,
                "dept_code": code,
                "dept_name": name,
                "note": note,
            }
        )

    for rule in rules.get("exact_email_rules") or []:
        email = str(rule.get("email", "")).strip()
        if not email:
            continue
        direction = str(rule.get("direction", "КС")).strip()
        code = str(rule["code"])
        org = _resolve_org_from_direction(direction)
        email_org = _resolve_org_from_email(email)
        if email_org:
            org = email_org
        add_row(
            email_or_pattern=email,
            rule_type="exact_email",
            code=code,
            name=str(rule.get("name") or dept_names.get(code, code)),
            direction=direction,
            org=org,
            note=str(rule.get("about") or ""),
        )

    for rule in rules.get("email_keyword_rules") or []:
        keyword = str(rule.get("keyword", "")).strip()
        if not keyword:
            continue
        direction = str(rule.get("direction", "КС")).strip()
        code = str(rule["code"])
        org = _resolve_org_from_direction(direction)
        add_row(
            email_or_pattern=f"*{keyword}*@...",
            rule_type="email_keyword",
            code=code,
            name=str(rule.get("name") or dept_names.get(code, code)),
            direction=direction,
            org=org,
            note=f"Ключевое слово «{keyword}» в локальной части адреса",
        )

    tz_topics = load_tz_department_topics(tz_topics_path)
    for code, entry in sorted(tz_topics.items()):
        dept_name = dept_names.get(code, code)
        for email in entry.get("emails") or []:
            email = str(email).strip()
            if not email:
                continue
            org = _resolve_org_from_email(email) or "НП"
            add_row(
                email_or_pattern=email,
                rule_type="tz_topics",
                code=code,
                name=dept_name,
                direction="",
                org=org,
                note="Email из ТЗ (Прил. Д)",
            )

    enterprise_file = resolve_enterprise_positions_path(enterprise_path)
    enterprise = json.loads(enterprise_file.read_text(encoding="utf-8"))
    structure_by_code: dict[str, dict] = {}
    for row in enterprise.get("structure_departments_with_codes") or []:
        code = str(row.get("code") or "").strip()
        if code:
            structure_by_code[code] = row

    for code, entry in sorted(tz_topics.items()):
        structure = structure_by_code.get(code, {})
        path = str(structure.get("path") or "")
        onec_name = str(structure.get("name") or "")
        for email in entry.get("emails") or []:
            email = str(email).strip()
            if not email:
                continue
            org = detect_organization(
                path=path,
                name=onec_name,
                code=code,
                emails=[email],
            )
            key = (email.lower(), "tz_resolved", code)
            if key in seen:
                for row in rows:
                    if (
                        row["email_or_pattern"].lower() == email.lower()
                        and row["dept_code"] == code
                        and row["rule_type"] == "tz_topics"
                    ):
                        row["org_code"] = org
                        row["org_name"] = ORG_FULL_NAMES.get(org, org)
                continue
            seen.add(key)

    rows.sort(
        key=lambda item: (
            ORG_ORDER.index(item["org_code"]) if item["org_code"] in ORG_ORDER else 99,
            item["email_or_pattern"].lower(),
        )
    )
    return rows


def _autosize_columns(sheet, *, min_width: int = 10, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[letter].width = width


def _style_header(sheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in sheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(
    output_path: Path,
    *,
    org_rows: list[dict],
    dept_rows: list[dict],
    direction_rows: list[dict],
    email_rows: list[dict],
) -> dict[str, int]:
    wb = Workbook()

    # Лист «Организации»
    org_sheet = wb.active
    org_sheet.title = "Организации"
    org_headers = [
        "Код ТЗ",
        "Полное название",
        "Ключевые слова (routing_rules)",
        "Маркеры в структуре 1С / email",
        "Правило направления",
        "Доступные направления",
    ]
    org_sheet.append(org_headers)
    for row in org_rows:
        org_sheet.append(
            [
                row["code"],
                row["full_name"],
                row["keywords"],
                row["structure_markers"],
                row["direction_rule"],
                row["directions"],
            ]
        )
    _style_header(org_sheet)
    _autosize_columns(org_sheet, max_width=70)
    for cell in org_sheet["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in org_sheet["D"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист «Организация → Подразделения»
    dept_sheet = wb.create_sheet("Организация → Подразделения")
    dept_headers = [
        "Организация (код)",
        "Организация (название)",
        "Код подразделения",
        "Название подразделения",
        "Направление",
    ]
    dept_sheet.append(dept_headers)
    for row in dept_rows:
        dept_sheet.append(
            [
                row["org_code"],
                row["org_name"],
                row["code"],
                row["name"],
                row["direction"],
            ]
        )
    _style_header(dept_sheet)
    _autosize_columns(dept_sheet, max_width=55)

    # Лист «Направления»
    dir_sheet = wb.create_sheet("Направления")
    dir_headers = [
        "Организация (код)",
        "Организация (название)",
        "Направление",
        "Описание",
        "Примечание",
    ]
    dir_sheet.append(dir_headers)
    for row in direction_rows:
        dir_sheet.append(
            [
                row["org_code"],
                row["org_name"],
                row["direction"],
                row["description"],
                row["note"],
            ]
        )
    _style_header(dir_sheet)
    _autosize_columns(dir_sheet, max_width=70)

    # Лист «Email → Организация»
    email_sheet = wb.create_sheet("Email → Организация")
    email_headers = [
        "Email / шаблон",
        "Тип правила",
        "Организация (код)",
        "Организация (название)",
        "Направление",
        "Код подразделения",
        "Подразделение",
        "Примечание",
    ]
    email_sheet.append(email_headers)
    for row in email_rows:
        email_sheet.append(
            [
                row["email_or_pattern"],
                row["rule_type"],
                row["org_code"],
                row["org_name"],
                row["direction"],
                row["dept_code"],
                row["dept_name"],
                row["note"],
            ]
        )
    _style_header(email_sheet)
    _autosize_columns(email_sheet, max_width=55)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "Организации": org_sheet.max_row - 1,
        "Организация → Подразделения": dept_sheet.max_row - 1,
        "Направления": dir_sheet.max_row - 1,
        "Email → Организация": email_sheet.max_row - 1,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Экспорт маппинга организаций в Excel")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Путь к Excel-файлу (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    parser.add_argument("--enterprise", type=Path, help="Путь к enterprise_positions.json")
    parser.add_argument("--rules", type=Path, help="Путь к routing_rules.json")
    parser.add_argument("--tz-topics", type=Path, help="Путь к tz_department_topics.json")
    args = parser.parse_args()

    rules = load_routing_rules(args.rules)
    active_rows, _liquidated = build_export_rows(
        enterprise_path=args.enterprise,
        rules_path=args.rules,
    )

    org_rows = build_organizations_rows(rules)
    dept_rows = build_departments_by_org_rows(active_rows)
    direction_rows = build_directions_rows()
    email_rows = build_email_org_rows(
        rules,
        tz_topics_path=args.tz_topics,
        enterprise_path=args.enterprise,
    )

    counts = write_excel(
        args.output,
        org_rows=org_rows,
        dept_rows=dept_rows,
        direction_rows=direction_rows,
        email_rows=email_rows,
    )

    org_dept_counts: dict[str, int] = defaultdict(int)
    for row in dept_rows:
        org_dept_counts[row["org_code"]] += 1

    print(f"Файл: {args.output.resolve()}")
    print("Строки по листам:")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}")
    print("Активные подразделения по организациям:")
    for org in ORG_ORDER:
        print(f"  {org}: {org_dept_counts.get(org, 0)}")


if __name__ == "__main__":
    main()
