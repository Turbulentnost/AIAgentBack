"""Smoke: split schedule Заказ/Опытные/Склад × План/Факт → monthly sheet."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    ROLE_PRODUCTION_SCHEDULE,
    UploadedWorkbook,
    _extract_production_schedule_products,
    _plan_demand_total,
    analyze_aveon_excel_files,
)

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")


async def main() -> None:
    schedule_path = TEST_DIR / "График производства.xlsx"
    assert schedule_path.exists(), schedule_path
    schedule = UploadedWorkbook(
        filename=schedule_path.name, content=schedule_path.read_bytes()
    )
    files, plans = _extract_production_schedule_products(
        [schedule], {schedule.filename: ROLE_PRODUCTION_SCHEDULE}
    )
    print("schedule files", files, "plans", len(plans))
    sokol = next(p for p in plans if "СОКОЛ" in p.product and "И (день)" in p.product)
    july = sokol.monthly_qty["Июль"]
    print("Сокол И Июль", july)
    assert july["заказ"]["план"] == 6800.0, july
    assert july["склад"]["план"] == 40000.0, july
    assert july["заказ"]["факт"] == 0.0
    assert july["опытные"]["план"] == 0.0

    uploaded = [
        UploadedWorkbook(filename=p.name, content=p.read_bytes())
        for p in sorted(TEST_DIR.glob("*.xlsx"))
        if not p.name.startswith("~$")
    ]
    result = await analyze_aveon_excel_files(uploaded)
    assert result.result_xlsx_bytes
    out = ROOT / "scripts" / "_smoke_split_monthly.xlsx"
    out.write_bytes(result.result_xlsx_bytes)
    wb = load_workbook(out)
    assert "1-производственный план (мес.)" in wb.sheetnames
    assert any(name.startswith("2-произв. план (") for name in wb.sheetnames)
    monthly = wb["1-производственный план (мес.)"]
    print("A1", monthly["A1"].value)
    print("G3 month", monthly["G3"].value, "G4", monthly["G4"].value, "G5", monthly["G5"].value)
    # summaryRight: G–L деталь Заказ/Опыт/Склад, M–N Потребность, далее недели, итог поступления, прогноз
    assert monthly["G3"].value == "Июль"
    assert monthly["G4"].value == "Заказ"
    assert monthly["G5"].value == "План"
    assert monthly["M4"].value == "Потребность"
    assert monthly["M5"].value == "План"
    assert monthly["N5"].value == "Факт"
    assert monthly.column_dimensions["G"].hidden is True
    assert monthly.column_dimensions["G"].outline_level == 1
    assert monthly.column_dimensions["M"].hidden is not True
    # недельная группа скрыта; итог поступления виден
    assert monthly["O4"].value == "Поступление по неделям"
    assert monthly.column_dimensions["O"].hidden is True
    # data row 6
    assert monthly["A6"].value
    sum_plan = monthly["M6"].value
    print("sum_plan M6", sum_plan)
    assert isinstance(sum_plan, str) and sum_plan.startswith("=")
    assert "G6" in sum_plan and "I6" in sum_plan and "K6" in sum_plan
    # найти колонку прогноза / поступления по подписи
    receipt_col = None
    forecast_col = None
    for col in range(1, monthly.max_column + 1):
        val = str(monthly.cell(4, col).value or "")
        if val.startswith("В пути"):
            receipt_col = col
        if val.startswith("Прогнозируемый остаток"):
            forecast_col = col
    assert receipt_col is not None and forecast_col is not None
    from openpyxl.utils import get_column_letter

    assert monthly.column_dimensions[get_column_letter(receipt_col)].hidden is not True
    forecast = monthly.cell(6, forecast_col).value
    receipt = monthly.cell(6, receipt_col).value
    print("receipt", receipt, "forecast", forecast)
    assert isinstance(receipt, str) and receipt.startswith("=")
    assert isinstance(forecast, str) and forecast.startswith("=")
    assert "M6" in forecast and "F6" in forecast

    # реконструкция: Σ plan×spec по всем изделиям строки
    matched = next(
        (
            r
            for r in result.merged_nomenclatures
            if any("СОКОЛ" in p and "И (день)" in p for p in r.products)
            and _plan_demand_total(r.monthly_demand.get("Июль")) > 0
        ),
        None,
    )
    assert matched is not None
    july_demand = matched.monthly_demand["Июль"]
    plans_by_key = {p.product: p for p in result.production_schedule_plans}
    recon_order = 0.0
    recon_stock = 0.0
    for product, spec_qty in matched.by_product.items():
        plan = plans_by_key.get(product)
        if plan is None or spec_qty is None:
            continue
        recon_order += float(plan.monthly_qty.get("Июль", {}).get("заказ", {}).get("план", 0)) * float(
            spec_qty
        )
        recon_stock += float(plan.monthly_qty.get("Июль", {}).get("склад", {}).get("план", 0)) * float(
            spec_qty
        )
    print(
        "nomenclature",
        matched.nomenclature[:60],
        "заказ план",
        july_demand["заказ"]["план"],
        "recon",
        recon_order,
        "склад план",
        july_demand["склад"]["план"],
        "recon",
        recon_stock,
    )
    assert abs(july_demand["заказ"]["план"] - recon_order) < 1e-6
    assert abs(july_demand["склад"]["план"] - recon_stock) < 1e-6
    print("OK", out)


if __name__ == "__main__":
    asyncio.run(main())
