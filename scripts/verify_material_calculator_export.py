"""Проверка агрегации и выгрузки калькулятора материалов через API."""
from __future__ import annotations

import json
import sys
from io import BytesIO

import httpx
from openpyxl import load_workbook

BASE = "http://127.0.0.1:5454/api/v1"
EMAIL = "bugata.pavel@local.dev"
PASSWORD = "Bugata2026!"


def main() -> int:
    with httpx.Client(base_url=BASE, timeout=60.0) as client:
        login = client.post("/auth/login", json={"email": EMAIL, "password": PASSWORD})
        login.raise_for_status()
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}

        specs = client.get("/agents/document-analysis/resource-specs?limit=20", headers=headers)
        specs.raise_for_status()
        items = [row for row in specs.json().get("items", []) if row.get("materials_count", 0) > 0]
        if len(items) < 2:
            print("SKIP: need at least 2 specs with materials in DB")
            return 0

        calc_items = [
            {"spec_ref_key": items[0]["ref_key"], "quantity": 5},
            {"spec_ref_key": items[1]["ref_key"], "quantity": 3},
        ]
        calc = client.post(
            "/agents/document-analysis/material-calculator",
            headers=headers,
            json={"items": calc_items},
        )
        calc.raise_for_status()
        payload = calc.json()
        lines = payload.get("lines") or []
        print(f"calculate: {len(lines)} lines, warnings={len(payload.get('warnings') or [])}")

        keys = [((line.get("code") or "").strip().lower(), (line.get("name") or "").strip().lower()) for line in lines]
        nom_keys = [(line.get("nomenclature_key") or "").strip() for line in lines]
        if len(keys) != len(set(keys)):
            print("FAIL: duplicate code+name rows in calculation result")
            return 1
        if len(nom_keys) != len(set(nom_keys)):
            print("FAIL: duplicate nomenclature_key rows in calculation result")
            return 1

        export = client.post(
            "/agents/document-analysis/material-calculator/export",
            headers=headers,
            json={"lines": lines},
        )
        export.raise_for_status()
        content_type = export.headers.get("content-type", "")
        if "spreadsheetml" not in content_type:
            print(f"FAIL: unexpected content-type {content_type}")
            return 1

        wb = load_workbook(BytesIO(export.content))
        ws = wb.active
        headers_row = [ws.cell(1, col).value for col in range(1, 5)]
        if headers_row != ["Код", "Номенклатура", "Количество", "Ед. изм."]:
            print(f"FAIL: bad headers {headers_row}")
            return 1

        excel_rows = []
        for row_idx in range(2, ws.max_row + 1):
            excel_rows.append(
                {
                    "code": (ws.cell(row_idx, 1).value or "").strip(),
                    "name": (ws.cell(row_idx, 2).value or "").strip(),
                    "total_qty": float(ws.cell(row_idx, 3).value or 0),
                    "unit": (ws.cell(row_idx, 4).value or "").strip(),
                }
            )

        if len(excel_rows) != len(lines):
            print(f"FAIL: excel rows {len(excel_rows)} != api lines {len(lines)}")
            return 1

        api_sorted = sorted(lines, key=lambda row: (row.get("name") or "").lower())
        for api_line, xlsx_line in zip(api_sorted, excel_rows, strict=True):
            if (api_line.get("code") or "") != xlsx_line["code"]:
                print("FAIL: code mismatch", api_line, xlsx_line)
                return 1
            if (api_line.get("name") or "") != xlsx_line["name"]:
                print("FAIL: name mismatch", api_line, xlsx_line)
                return 1
            if float(api_line.get("total_qty") or 0) != xlsx_line["total_qty"]:
                print("FAIL: qty mismatch", api_line, xlsx_line)
                return 1
            if (api_line.get("unit") or "—") != xlsx_line["unit"]:
                print("FAIL: unit mismatch", api_line, xlsx_line)
                return 1

        print("OK: aggregation unique, excel export matches modal data")
        print(json.dumps({"lines": len(lines), "specs": [items[0]["ref_key"], items[1]["ref_key"]]}, ensure_ascii=False))
        return 0


if __name__ == "__main__":
    sys.exit(main())
