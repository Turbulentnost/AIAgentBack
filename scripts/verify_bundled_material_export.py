"""Проверка export endpoint на bundled backend sidecar."""
from __future__ import annotations

import subprocess
import sys
import time
from io import BytesIO
from pathlib import Path

import httpx
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BUNDLED = ROOT / "AIAgent-desctop" / "aveon-agent" / "resources" / "backend" / "aveon-backend.exe"
PORT = 18768
EMAIL = "bugata.pavel@local.dev"
PASSWORD = "Bugata2026!"


def main() -> int:
    if not BUNDLED.is_file():
        print(f"FAIL: bundled backend missing: {BUNDLED}")
        return 1

    proc = subprocess.Popen(
        [str(BUNDLED), "--host", "127.0.0.1", "--port", str(PORT)],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    try:
        with httpx.Client(base_url=f"http://127.0.0.1:{PORT}/api/v1", timeout=60.0) as client:
            for _ in range(60):
                try:
                    if client.get("/health").status_code == 200:
                        break
                except httpx.HTTPError:
                    pass
                time.sleep(0.5)
            else:
                print("FAIL: bundled backend health timeout")
                return 1

            openapi = client.get("/openapi.json")
            openapi.raise_for_status()
            if "/agents/document-analysis/material-calculator/export" not in openapi.text:
                print("FAIL: export route missing in bundled openapi")
                return 1

            login = client.post("/auth/login", json={"email": EMAIL, "password": PASSWORD})
            login.raise_for_status()
            token = login.json()["access_token"]
            headers = {"Authorization": f"Bearer {token}"}

            calc = client.post(
                "/agents/document-analysis/material-calculator",
                headers=headers,
                json={
                    "items": [
                        {"spec_ref_key": "2db216e6-f6d3-11f0-aa28-289200eb56b7", "quantity": 2},
                        {"spec_ref_key": "779d7921-4387-11f1-aa38-289200eb56b7", "quantity": 3},
                    ]
                },
            )
            calc.raise_for_status()
            lines = calc.json().get("lines") or []
            if not lines:
                print("FAIL: calculator returned no lines")
                return 1

            export = client.post(
                "/agents/document-analysis/material-calculator/export",
                headers=headers,
                json={"lines": lines[:5]},
            )
            if export.status_code == 404:
                print("FAIL: export endpoint returned 404")
                return 1
            export.raise_for_status()

            wb = load_workbook(BytesIO(export.content))
            ws = wb.active
            if ws.cell(1, 1).value != "Код":
                print(f"FAIL: unexpected header {ws.cell(1, 1).value!r}")
                return 1
            if ws.max_row < 2:
                print("FAIL: export workbook is empty")
                return 1

        print("OK: bundled backend export works")
        return 0
    finally:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()


if __name__ == "__main__":
    sys.exit(main())
