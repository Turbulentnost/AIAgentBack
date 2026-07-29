"""Generate daily sheet and assert Header-like colors."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    UploadedWorkbook,
    analyze_aveon_excel_files,
)

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")


async def main() -> None:
    uploaded = [
        UploadedWorkbook(filename=p.name, content=p.read_bytes())
        for p in sorted(TEST_DIR.glob("*.xlsx"))
        if not p.name.startswith("~$")
        and any(
            key in p.name
            for key in ("остатк", "недел", "производств", "ОТГРУЗ", "отгруз")
        )
    ]
    # fallback: all non-temp
    if len(uploaded) < 3:
        uploaded = [
            UploadedWorkbook(filename=p.name, content=p.read_bytes())
            for p in sorted(TEST_DIR.glob("*.xlsx"))
            if not p.name.startswith("~$")
        ]

    result = await analyze_aveon_excel_files(uploaded)
    assert result.result_xlsx_bytes
    out = ROOT / "scripts" / "_smoke_header_styles.xlsx"
    out.write_bytes(result.result_xlsx_bytes)
    wb = load_workbook(out)
    daily_name = next(name for name in wb.sheetnames if name.startswith("обеспечение ("))
    daily = wb[daily_name]
    def rgb(cell):
        value = cell.fill.fgColor.rgb
        return value[-6:].upper() if value else ""

    def font_rgb(cell):
        color = cell.font.color
        value = getattr(color, "rgb", None) if color else None
        return value[-6:].upper() if value else ""

    assert rgb(daily["A1"]) == "1F4E78", rgb(daily["A1"])
    assert rgb(daily["A2"]) == "D9EAF7"
    assert rgb(daily["A3"]) == "5B9BD5"
    assert rgb(daily["F3"]) == "5B9BD5"
    assert rgb(daily["F4"]) == "D9EAF7"
    assert font_rgb(daily["A1"]) == "FFFFFF"
    assert font_rgb(daily["F4"]) == "1F1F1F"
    print("OK styles", out, "A1", daily["A1"].fill.fgColor.rgb)


if __name__ == "__main__":
    asyncio.run(main())
