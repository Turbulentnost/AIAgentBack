"""Экспорт маппинга «организация (кратко) → отдел» в Excel.

Колонки: Организация (кратко) | Код отдела | Отдел (краткое/полное название)

Использует ту же логику определения организации, что export_departments_excel.py
и build_department_records_for_db (detect_organization + build_export_rows).

Пример:
  python scripts/export_org_departments_short.py
  python scripts/export_org_departments_short.py -o data/org_departments_short.xlsx
"""

from __future__ import annotations

import argparse
import importlib.util
import sys
from collections import Counter, defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_dept_export_path = ROOT / "scripts" / "export_departments_excel.py"
_spec = importlib.util.spec_from_file_location("export_departments_excel", _dept_export_path)
_dept_export = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_dept_export)

ORG_FULL_NAMES: dict[str, str] = _dept_export.ORG_FULL_NAMES
ORG_ORDER: tuple[str, ...] = _dept_export.ORG_ORDER
build_export_rows = _dept_export.build_export_rows

DEFAULT_OUTPUT = ROOT / "data" / "org_departments_short.xlsx"


def build_short_rows(active_rows: list[dict]) -> list[dict]:
    """Одна строка на активное подразделение, отсортировано по org → code."""
    rows: list[dict] = []
    for row in active_rows:
        onec_name = str(row.get("onec_name") or "").strip()
        routing_name = str(row.get("name") or "").strip()
        department_name = onec_name or routing_name or row["code"]
        short_name = routing_name if routing_name and routing_name != onec_name else ""

        rows.append(
            {
                "org_code": row["org_code"],
                "code": row["code"],
                "name": department_name,
                "short_name": short_name,
                "display_name": (
                    f"{short_name} ({onec_name})"
                    if short_name and onec_name and short_name != onec_name
                    else department_name
                ),
            }
        )

    rows.sort(
        key=lambda item: (
            ORG_ORDER.index(item["org_code"]) if item["org_code"] in ORG_ORDER else 99,
            item["code"],
        )
    )
    return rows


def _autosize_columns(sheet, *, min_width: int = 10, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[letter].width = width


def _style_header(sheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in sheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(output_path: Path, short_rows: list[dict]) -> dict[str, int]:
    wb = Workbook()

    # Лист 1 — основной список
    main = wb.active
    main.title = "Отделы по организациям"
    headers = ["Организация (кратко)", "Код отдела", "Отдел (краткое/полное название)"]
    main.append(headers)
    for row in short_rows:
        main.append([row["org_code"], row["code"], row["display_name"]])
    _style_header(main)
    _autosize_columns(main, max_width=70)
    for cell in main["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист 2 — сводка по организациям (группировка)
    summary = wb.create_sheet("Сводка по организациям")
    summary.append(["Организация (кратко)", "Полное название", "Кол-во отделов", "Примеры отделов"])
    by_org: dict[str, list[dict]] = defaultdict(list)
    for row in short_rows:
        by_org[row["org_code"]].append(row)

    for org in ORG_ORDER:
        org_rows = by_org.get(org, [])
        examples = "; ".join(r["display_name"] for r in org_rows[:5])
        if len(org_rows) > 5:
            examples += f" … (+{len(org_rows) - 5})"
        summary.append([org, ORG_FULL_NAMES.get(org, org), len(org_rows), examples])

    extra_orgs = sorted(set(by_org) - set(ORG_ORDER))
    for org in extra_orgs:
        org_rows = by_org[org]
        examples = "; ".join(r["display_name"] for r in org_rows[:5])
        summary.append([org, org, len(org_rows), examples])

    total = len(short_rows)
    summary.append(["ИТОГО", "Все организации", total, ""])
    _style_header(summary)
    _autosize_columns(summary, max_width=80)
    for cell in summary["D"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Лист 3 — pivot-вид: НП → список отделов построчно с пустыми строками-разделителями
    pivot = wb.create_sheet("НП → список отделов")
    pivot.append(["Организация (кратко)", "Код отдела", "Отдел"])
    for org in ORG_ORDER:
        org_rows = by_org.get(org, [])
        if not org_rows:
            continue
        for idx, row in enumerate(org_rows):
            pivot.append(
                [
                    org if idx == 0 else "",
                    row["code"],
                    row["display_name"],
                ]
            )
        pivot.append(["", "", ""])
    _style_header(pivot)
    _autosize_columns(pivot, max_width=70)
    for cell in pivot["C"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "Отделы по организациям": main.max_row - 1,
        "Сводка по организациям": summary.max_row - 1,
        "НП → список отделов": pivot.max_row - 1,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Экспорт отделов по кратким кодам организаций (НП, АЛ, МГ…)"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Путь к Excel-файлу (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    parser.add_argument("--enterprise", type=Path, help="Путь к enterprise_positions.json")
    parser.add_argument("--rules", type=Path, help="Путь к routing_rules.json")
    args = parser.parse_args()

    active_rows, _liquidated = build_export_rows(
        enterprise_path=args.enterprise,
        rules_path=args.rules,
    )
    short_rows = build_short_rows(active_rows)
    counts = write_excel(args.output, short_rows)

    org_counter = Counter(row["org_code"] for row in short_rows)
    print(f"Файл: {args.output.resolve()}")
    print("Строки по листам:")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}")
    print("Активные отделы по организациям:")
    for org in ORG_ORDER:
        print(f"  {org}: {org_counter.get(org, 0)}")


if __name__ == "__main__":
    main()
