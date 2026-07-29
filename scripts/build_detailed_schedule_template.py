"""Пересобрать шаблон детального графика производства (формат отчёта П/ф·ОТК·Склад)."""

from __future__ import annotations

from calendar import monthrange
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = (
    Path(__file__).resolve().parents[1]
    / "app"
    / "agents"
    / "document_analysis_agent"
    / "templates"
    / "шаблон_детальный_график_производства.xlsx"
)

_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_STAGE_FILL = PatternFill("solid", fgColor="FFF2CC")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_header(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = Font(bold=True)
    cell.alignment = _CENTER
    cell.border = _THIN


def main() -> None:
    year, month = 2026, 7
    last_day = monthrange(year, month)[1]
    month_label = "Июль"

    wb = Workbook()
    ws = wb.active
    ws.title = "Детальный график"

    ws["A1"] = (
        f"Детальный график производства  {month_label} {year} "
        f"(ШАБЛОН — заполните своими данными)"
    )
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = _LEFT

    ws["A3"] = "№"
    ws["B3"] = "Модель / изделие"
    ws["C3"] = "Стадия"
    for col in (1, 2, 3):
        _style_header(ws.cell(3, col))
        _style_header(ws.cell(2, col))
        ws.cell(2, col).value = None

    # Каждый день месяца отдельно: дата (дд.мм) merge на план+факт
    col = 4
    day_cols: list[tuple[int, int, int]] = []
    for day in range(1, last_day + 1):
        plan_col = col
        fact_col = col + 1
        ws.merge_cells(
            start_row=2,
            start_column=plan_col,
            end_row=2,
            end_column=fact_col,
        )
        day_cell = ws.cell(2, plan_col, datetime(year, month, day))
        day_cell.number_format = "DD.MM"
        _style_header(day_cell)
        _style_header(ws.cell(2, fact_col))
        ws.cell(3, plan_col, "план")
        ws.cell(3, fact_col, "факт")
        _style_header(ws.cell(3, plan_col))
        _style_header(ws.cell(3, fact_col))
        day_cols.append((day, plan_col, fact_col))
        col += 2

    last_data_col = col - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_data_col)

    for c in range(1, last_data_col + 1):
        cell = ws.cell(2, c)
        cell.fill = _HEADER_FILL
        cell.border = _THIN
        if cell.alignment is None or cell.alignment.horizontal is None:
            cell.alignment = _CENTER
        if c <= 3 or cell.value is not None:
            cell.font = Font(bold=True)

    # Примеры: несколько дней с план/факт
    products = [
        (
            "Сокол И (пример)",
            {1: 200, 2: 200, 11: 100, 12: 100},
            {1: 100, 11: 80},
        ),
        (
            "Сокол ИТ (пример)",
            {1: 100, 2: 100, 11: 50, 12: 50},
            {2: 40, 12: 40},
        ),
    ]

    row = 4
    for idx, (name, day_plans, day_facts) in enumerate(products, start=1):
        block_start = row
        ws.cell(row, 1, idx)
        ws.cell(row, 2, name)
        ws.cell(row, 3, "П/ф")
        ws.cell(row, 3).fill = _STAGE_FILL
        for day, plan_col, fact_col in day_cols:
            if day in day_plans:
                ws.cell(row, plan_col, day_plans[day])
            if day in day_facts:
                ws.cell(row, fact_col, day_facts[day])
        row += 1
        for stage in ("ОТК", "Склад"):
            ws.cell(row, 3, stage)
            ws.cell(row, 3).fill = _STAGE_FILL
            for day, plan_col, fact_col in day_cols:
                if day in day_plans:
                    ws.cell(row, plan_col, day_plans[day])
                if day in day_facts:
                    ws.cell(row, fact_col, day_facts[day])
            row += 1
        block_end = row - 1
        ws.merge_cells(
            start_row=block_start, start_column=1, end_row=block_end, end_column=1
        )
        ws.merge_cells(
            start_row=block_start, start_column=2, end_row=block_end, end_column=2
        )
        ws.cell(block_start, 1).alignment = _CENTER
        ws.cell(block_start, 2).alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

    for idx in (3, 4):
        block_start = row
        ws.cell(row, 1, idx)
        ws.cell(row, 3, "П/ф")
        row += 1
        for stage in ("ОТК", "Склад"):
            ws.cell(row, 3, stage)
            row += 1
        ws.merge_cells(
            start_row=block_start, start_column=1, end_row=row - 1, end_column=1
        )
        ws.cell(block_start, 1).alignment = _CENTER

    ws.cell(row, 2, "Итого выпуск")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for c in range(4, last_data_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 8

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "D4"

    instr = wb.create_sheet("Инструкция")
    lines = [
        "Как заполнять шаблон «Детальный график производства»",
        "1. Лист «Детальный график»: в колонке B укажите модель/изделие, в колонке C — стадию (П/ф, ОТК, Склад).",
        "2. В расчёт ежедневного обеспечения агент берёт только строку стадии «П/ф». ОТК и Склад нужны для формы отчёта.",
        "3. Каждый день месяца — отдельная колонка даты (дд.мм); под ней две ячейки: «план» и «факт».",
        "4. План заполняйте на нужные дни; факт — только за дни, по которым есть данные (пустые ячейки факта допустимы).",
        "5. При необходимости добавьте или уберите дни — сохраните структуру: дата → план/факт.",
        "6. Удалите строки-примеры и подставьте свои данные перед загрузкой в агент.",
    ]
    for idx, text in enumerate(lines, start=1):
        cell = instr.cell(idx, 1, text)
        if idx == 1:
            cell.font = Font(bold=True)
        cell.alignment = _LEFT
    instr.column_dimensions["A"].width = 110

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"saved {OUT} ({OUT.stat().st_size} bytes), days=1..{last_day}")


if __name__ == "__main__":
    main()
