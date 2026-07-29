import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook


def dump_cell(ws, coord: str) -> None:
    c = ws[coord]
    fill = c.fill
    font = c.font
    fg = fill.fgColor
    fg_rgb = None
    fg_theme = None
    fg_tint = None
    if fg is not None:
        fg_rgb = getattr(fg, "rgb", None)
        fg_theme = getattr(fg, "theme", None)
        fg_tint = getattr(fg, "tint", None)
    font_color = None
    if font.color is not None:
        font_color = getattr(font.color, "rgb", None) or getattr(font.color, "theme", None)
    print(f"{ws.title!r} {coord}: val={c.value!r}")
    print(
        f"  fill type={fill.fill_type} rgb={fg_rgb} theme={fg_theme} tint={fg_tint}"
    )
    print(f"  font bold={font.bold} color={font_color} size={font.size}")
    print(
        f"  align h={c.alignment.horizontal} v={c.alignment.vertical} wrap={c.alignment.wrap_text}"
    )


paths = [
    Path(r"c:\Users\uaa\Downloads\result (15).xlsx"),
    Path(r"c:\Users\uaa\Desktop\мусор\AI Platform\AIAgentBack\data\aveon\Header.xlsx"),
]

for path in paths:
    wb = load_workbook(path)
    print("====", path.name, wb.sheetnames)
    for name in wb.sheetnames[:2]:
        ws = wb[name]
        print("---", name)
        for coord in [
            "A1",
            "A2",
            "A3",
            "B3",
            "E3",
            "F3",
            "G3",
            "H3",
            "A4",
            "F4",
            "G4",
            "H4",
            "I4",
        ]:
            dump_cell(ws, coord)
        print("merged sample:", [str(r) for r in list(ws.merged_cells.ranges)[:12]])
        print()
