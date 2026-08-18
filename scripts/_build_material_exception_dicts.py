"""Build consumable/workshop exception JSON dicts from classification workbook."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "aveon" / "классификация_расходники_обеспеченность_изделий.xlsx"
OUT_CONSUMABLE = ROOT / "data" / "aveon" / "material_consumable_exceptions.json"
OUT_WORKSHOP = ROOT / "data" / "aveon" / "material_workshop_exceptions.json"

CONSUMABLE_PATTERNS = [
    r"\bболт\b",
    r"\bгайк",
    r"\bшайб",
    r"\bвинт\b",
    r"\bшпильк",
    r"\bзаклеп",
    r"\bсаморез",
    r"\bанкер",
    r"\bшплинт",
    r"\bштифт",
    r"\bхомут",
    r"\bстяжк",
    r"\bскоб",
    r"шлиф",
    r"наждак",
    r"абразив",
    r"круг\s*шлиф",
    r"липучк",
    r"скотч",
    r"изолент",
    r"электрод",
    r"проволок.*свар",
    r"свароч",
    r"флюс",
    r"клей",
    r"герметик",
    r"силикон",
    r"мастик",
    r"антифрик",
    r"смазк",
    r"\bмасло\b",
    r"солидол",
    r"краск",
    r"грунт",
    r"эмаль",
    r"растворит",
    r"очистит",
    r"обезжир",
    r"антикор",
    r"салфет",
    r"ветош",
    r"тряп",
    r"перчат",
    r"респиратор",
    r"упаков",
    r"стретч",
    r"пленк.*упак",
    r"пакет",
    r"этикет",
    r"бирк",
    r"маркиров",
    r"сверл",
    r"фрез",
    r"метчик",
    r"плашк",
    r"разверт",
    r"оснастк",
    r"\bбита\b",
    r"лента.*клей",
    r"двухсторон",
    r"пено-клей",
    r"антифлюс",
    r"наклейк",
    r"термоусад",
    r"стяжка",
    r"упаковочн",
    r"коробк",
]

WORKSHOP_PATTERNS = [
    r"полуфабрик",
    r"\bп/ф\b",
    r"полуф\b",
    r"\bузел\b",
    r"\bблок\b",
    r"агрегат",
    r"механизм",
    r"редуктор",
    r"двигател",
    r"насос",
    r"контроллер",
    r"\bплат",
    r"плата\b",
    r"датчик",
    r"реле",
    r"концевик",
    r"энкодер",
    r"инвертор",
    r"жгут",
    r"кабель",
    r"разъем",
    r"коннектор",
    r"фишк",
    r"корпус",
    r"рама",
    r"каркас",
    r"стойк",
    r"опор",
    r"балк",
    r"лонжерон",
    r"поперечин",
    r"\bлист\b",
    r"листов",
    r"профил",
    r"труб",
    r"уголок",
    r"швеллер",
    r"двутавр",
    r"полос",
    r"отлив",
    r"поков",
    r"штампов",
    r"заготов",
    r"колес",
    r"шина\b",
    r"камера\b",
    r"сидень",
    r"обивк",
    r"чехол",
    r"стекл",
    r"зеркал",
    r"аккумулятор",
    r"батаре",
    r"комплект\b",
    r"сборк",
    r"\bвал\b",
    r"шестерн",
    r"муфт",
    r"подшипник",
    r"гидро",
    r"цилиндр",
    r"манжет",
    r"сальник",
    r"электромагнит",
    r"втулк",
    r"эксцентрик",
    r"рейка",
    r"фильтр\b",
    r"радиатор",
    r"антенн",
    r"модуль",
    r"платы\b",
    r"литье",
    r"hdmi",
    r"розетк",
    r"cnlinko",
]

STOP_TOKENS = frozenset(
    {
        "2026",
        "2025",
        "2024",
        "для",
        "по",
        "на",
        "из",
        "с",
        "и",
        "в",
        "к",
        "от",
        "до",
        "шт",
        "мм",
        "см",
        "мл",
        "кг",
        "г",
        "л",
        "х",
        "x",
        "the",
        "pro",
        "v2",
        "v3",
        "v4",
        "series",
        "ascent",
        "day",
        "night",
        "male",
        "female",
        "pin",
        "type",
    }
)


def normalize_key(value: str) -> str:
    text = (value or "").strip().lower().replace("ё", "е")
    text = text.replace('"', "").replace("«", "").replace("»", "").replace("'", "")
    return re.sub(r"\s+", " ", text)


def tokenize(name: str) -> list[str]:
    text = re.sub(r"[^\w\s-]", " ", name.lower().replace("ё", "е"))
    text = re.sub(r"\s+", " ", text).strip()
    tokens: list[str] = []
    for word in text.split():
        if word.isdigit() or word in STOP_TOKENS or len(word) < 3:
            continue
        tokens.append(word)
    return tokens


def load_rows(sheet_name: str) -> list[dict]:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(value or "").strip() for value in header_row]
        index = {header: idx for idx, header in enumerate(headers)}
        rows: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            record = {
                header: row[idx] if idx < len(row) else None for header, idx in index.items()
            }
            rows.append(record)
        return rows
    finally:
        wb.close()


def aggregate_materials(rows: list[dict], *, classification_prefix: str) -> dict[str, dict]:
    grouped: dict[str, dict] = {}
    for row in rows:
        name = str(row.get("Номенклатура") or "").strip()
        classification = str(row.get("Классификация") or "").strip()
        if not name:
            continue
        if classification_prefix == "расходник":
            if classification != "расходник":
                continue
        elif not classification.startswith(classification_prefix):
            continue

        product = str(row.get("Изделие") or "").strip()
        confidence = str(row.get("Уверенность") or "").strip()
        reason = str(row.get("Обоснование") or "").strip()
        zero_stock = str(row.get("Нулевой остаток (все месяцы)") or "").strip().lower() == "да"

        item = grouped.setdefault(
            name,
            {
                "name": name,
                "normalized_key": normalize_key(name),
                "confidence": confidence,
                "reason": reason,
                "zero_stock": zero_stock,
                "products": [],
            },
        )
        if confidence and item["confidence"] != "высокая" and confidence == "высокая":
            item["confidence"] = confidence
        if reason and len(reason) > len(item["reason"] or ""):
            item["reason"] = reason
        item["zero_stock"] = item["zero_stock"] or zero_stock
        if product and product not in item["products"]:
            item["products"].append(product)
    return grouped


def keyword_hints(materials: dict[str, dict], *, other_counter: Counter[str], min_count: int = 3) -> list[str]:
    counter: Counter[str] = Counter()
    for name in materials:
        counter.update(tokenize(name))
    hints: list[str] = []
    for token, count in counter.most_common():
        if count < min_count:
            continue
        if other_counter[token] >= count:
            continue
        hints.append(token)
    return hints


def build_payload(
    *,
    kind: str,
    label: str,
    patterns: list[str],
    materials: dict[str, dict],
    keyword_hints_list: list[str],
) -> dict:
    items = sorted(materials.values(), key=lambda item: item["name"].casefold())
    return {
        "version": 1,
        "kind": kind,
        "label": label,
        "source_file": XLSX.name,
        "generated_at": date.today().isoformat(),
        "description": label,
        "patterns": patterns,
        "keyword_hints": keyword_hints_list,
        "materials": items,
        "materials_count": len(items),
        "lookup_by_normalized_key": {
            item["normalized_key"]: item["name"] for item in items if item["normalized_key"]
        },
    }


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Workbook not found: {XLSX}")

    main_rows = load_rows("По изделиям и номенклатурам")
    zero_workshop_rows = load_rows("Нулевые не расходники")

    consumables = aggregate_materials(main_rows, classification_prefix="расходник")
    workshop = aggregate_materials(main_rows, classification_prefix="не расходник")

    # Strengthen workshop dict with dedicated zero-stock sheet.
    for row in zero_workshop_rows:
        name = str(row.get("Номенклатура") or "").strip()
        if not name:
            continue
        product = str(row.get("Изделие") or "").strip()
        item = workshop.setdefault(
            name,
            {
                "name": name,
                "normalized_key": normalize_key(name),
                "confidence": str(row.get("Уверенность") or "средняя").strip(),
                "reason": str(row.get("Обоснование") or "").strip(),
                "zero_stock": True,
                "products": [],
            },
        )
        item["zero_stock"] = True
        if product and product not in item["products"]:
            item["products"].append(product)

    cons_tokens = Counter()
    work_tokens = Counter()
    for name in consumables:
        cons_tokens.update(tokenize(name))
    for name in workshop:
        work_tokens.update(tokenize(name))

    consumable_payload = build_payload(
        kind="consumable",
        label="Исключения: возможные расходники (не блокируют условную обеспеченность)",
        patterns=CONSUMABLE_PATTERNS,
        materials=consumables,
        keyword_hints_list=keyword_hints(consumables, other_counter=work_tokens),
    )
    workshop_payload = build_payload(
        kind="workshop",
        label="Исключения: возможно в цехе / WIP (не блокируют условную обеспеченность)",
        patterns=WORKSHOP_PATTERNS,
        materials=workshop,
        keyword_hints_list=keyword_hints(workshop, other_counter=cons_tokens),
    )

    OUT_CONSUMABLE.write_text(
        json.dumps(consumable_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    OUT_WORKSHOP.write_text(
        json.dumps(workshop_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps({
        "consumable": OUT_CONSUMABLE.name,
        "consumable_count": consumable_payload["materials_count"],
        "workshop": OUT_WORKSHOP.name,
        "workshop_count": workshop_payload["materials_count"],
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
