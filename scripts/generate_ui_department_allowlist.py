"""Собрать data/ui_department_allowlist.json из Excel-списка служб.

Пример:
  python scripts/generate_ui_department_allowlist.py ^
    --excel \"C:\\Users\\...\\Лист Microsoft Excel.xlsx\"

Правила:
  - берутся строки с кодом 00-XXXXXX из Excel;
  - позиции с «директор» в названии исключаются;
  - всегда добавляется 00-000001 Председатель Совета Директоров.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from agent_pochta.config import PROJECT_ROOT
from agent_pochta.services.routing_departments import resolve_enterprise_positions_path

_CHAIRMAN_CODE = "00-000001"
_CODE_RE = re.compile(r"^00-\d{6}$")
_EXPLICIT_DIRECTOR_CODES = frozenset({
    "00-000040",
    "00-000049",
    "00-000058",
    "00-000080",
    "00-000152",
    "00-000163",
    "00-000172",
})
_UI_EXCLUDED_CODES = frozenset({"00-000007", "00-000149"})


def _is_director_role(code: str, name: str) -> bool:
    if code in _EXPLICIT_DIRECTOR_CODES:
        return False
    if code == _CHAIRMAN_CODE:
        return False
    if code in _UI_EXCLUDED_CODES:
        return True
    return "директор" in (name or "").lower().replace("ё", "е")


def _parse_excel(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(values_only=True):
        code = ""
        name = ""
        for cell in row:
            if cell is None:
                continue
            value = str(cell).strip()
            if not value or value in {"+", "-"}:
                continue
            if _CODE_RE.match(value):
                code = value
            elif not value.lower().startswith(("код", "название")):
                name = value
        if code:
            rows.append({"code": code, "name": name})
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, required=True, help="Путь к Excel со службами")
    parser.add_argument(
        "--out",
        type=Path,
        default=PROJECT_ROOT / "data" / "ui_department_allowlist.json",
    )
    parser.add_argument("--enterprise", type=Path, default=None)
    args = parser.parse_args()

    excel_rows = _parse_excel(args.excel)
    enterprise = json.loads(
        resolve_enterprise_positions_path(args.enterprise).read_text(encoding="utf-8")
    )
    onec_names = {
        str(row.get("code") or "").strip(): str(row.get("name") or "").strip()
        for row in enterprise.get("structure_departments_routing_codes") or []
        if str(row.get("code") or "").strip()
    }
    chair_name = onec_names.get(_CHAIRMAN_CODE) or "Председатель Совета Директоров"

    kept: list[dict[str, str]] = []
    excluded: list[dict[str, str]] = []
    for row in excel_rows:
        if row["code"] in _UI_EXCLUDED_CODES:
            excluded.append({**row, "reason": "excluded_ui"})
            continue
        if _is_director_role(row["code"], row["name"]):
            excluded.append(row)
        else:
            kept.append(row)

    departments = [
        {"code": _CHAIRMAN_CODE, "name": chair_name, "source": "special"},
    ]
    seen = {_CHAIRMAN_CODE}
    for code in sorted(_EXPLICIT_DIRECTOR_CODES):
        if code in seen:
            continue
        departments.append(
            {
                "code": code,
                "name": onec_names.get(code) or code,
                "source": "explicit_director",
            }
        )
        seen.add(code)
    for row in kept:
        if row["code"] in seen:
            continue
        departments.append(
            {
                "code": row["code"],
                "name": row["name"] or onec_names.get(row["code"]) or row["code"],
                "source": "excel",
            }
        )
        seen.add(row["code"])

    payload = {
        "version": "1.0",
        "description": (
            "Allowlist служб для UI picker и LLM-кандидатов. "
            "Excel минус директора + явные директора + 00-000001; "
            "без Главного метролога (00-000149) и Главного конструктора (00-000007)."
        ),
        "special_codes": [_CHAIRMAN_CODE],
        "explicit_director_codes": sorted(_EXPLICIT_DIRECTOR_CODES),
        "excluded_ui_codes": sorted(_UI_EXCLUDED_CODES),
        "excluded_directors_from_excel": excluded,
        "departments": departments,
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(departments)} departments to {args.out}")
    print(f"Excluded directors: {len(excluded)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
