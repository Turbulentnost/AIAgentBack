"""Smoke: unit column after supplier, stock in F, forecast uses F."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    UploadedWorkbook,
    analyze_aveon_excel_files,
)

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")


async def main() -> None:
    uploaded = [
        UploadedWorkbook(filename=p.name, content=p.read_bytes())
        for p in sorted(TEST_DIR.glob("*.xlsx"))
        if not p.name.startswith("~$")
    ]
    result = await analyze_aveon_excel_files(uploaded)
    assert result.result_xlsx_bytes
    out = ROOT / "scripts" / "_smoke_unit_column.xlsx"
    out.write_bytes(result.result_xlsx_bytes)
    wb = load_workbook(out)

    monthly = wb["1-производственный план (мес.)"]
    assert monthly["C3"].value == "Поставщик"
    assert monthly["D3"].value == "Ед. изм."
    assert monthly["E3"].value == "Цена, руб./ед."
    assert monthly["F3"].value == "Остаток"
    assert monthly["G3"].value == "Июль"
    formula = monthly["N6"].value
    print("monthly D6 unit", monthly["D6"].value, "N6", formula)
    assert isinstance(formula, str) and formula.startswith("=F6+")
    assert "G6" in formula and "I6" in formula and "K6" in formula

    units_filled = sum(
        1 for r in result.merged_nomenclatures if r.unit and str(r.unit).strip()
    )
    print("units_filled", units_filled, "/", len(result.merged_nomenclatures))
    assert units_filled > 100

    daily_name = next(name for name in wb.sheetnames if name.startswith("обеспечение ("))
    daily = wb[daily_name]
    assert daily["D3"].value == "Ед. изм."
    assert daily["F3"].value == "Остаток"
    assert daily["G3"].value == "01.07" or str(daily["G3"].value).startswith("01")
    d_formula = daily["I5"].value
    print("daily D5 unit", daily["D5"].value, "I5", d_formula)
    assert isinstance(d_formula, str) and d_formula.startswith("=F5+")
    print("OK", out)


if __name__ == "__main__":
    asyncio.run(main())
