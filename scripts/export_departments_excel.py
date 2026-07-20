"""Экспорт маппинга организаций → подразделений в Excel.

Источники:
  - data/enterprise_positions.json (structure_departments_*)
  - data/routing_rules.json (направления, коды организаций ТЗ)
  - agent_pochta.services.routing_departments.build_department_records_for_db

Пример:
  python scripts/export_departments_excel.py
  python scripts/export_departments_excel.py -o data/departments_by_organization.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from agent_pochta.services.routing_departments import (  # noqa: E402
    build_department_records_for_db,
    directions_by_code_from_rules,
    is_liquidated_department,
    load_routing_rules,
    resolve_enterprise_positions_path,
)

DEFAULT_OUTPUT = ROOT / "data" / "departments_by_organization.xlsx"

ORG_FULL_NAMES: dict[str, str] = {
    "НП": "НПО «Турбулентность-ДОН»",
    "АЛ": "ООО «Алмаз»",
    "МГ": "ООО «Метрогазсервис»",
    "АМ": "ООО «Амурская легенда»",
    "МИ": "ООО «МИЛАКА»",
    "БМ": "БМИ (блочно-модульные изделия)",
}

ORG_ORDER = ("НП", "АЛ", "МГ", "АМ", "МИ", "БМ")

_BMI_MARKERS = ("бми", "блочно-модульн", "блочно модульн", "блочно-модульные")
_ALMAZ_MARKERS = ("алмаз", "гранд spi", "spu-5", "spu 5", "бытовых счетчик", "гранд")
_MGS_MARKERS = ("метрогаз", "мгс")
_AM_MARKERS = ("амурск", "акваген", "легенд")
_MI_MARKERS = ("милака",)

_CODE_RE = re.compile(r"00-\d{6}")


def _normalize(text: str) -> str:
    return (text or "").lower().replace("ё", "е")


def is_liquidated_for_export(name: str | None, path: str | None) -> bool:
    """Ликвидированные подразделения (расширенная проверка для отчёта)."""
    if is_liquidated_department(name, path):
        return True
    name_norm = _normalize(name)
    if "ликвидирован" in name_norm:
        return True
    if name_norm.startswith("_") and "ликв" in name_norm:
        return True
    return False


def detect_organization(
    *,
    path: str = "",
    name: str = "",
    code: str = "",
    direction: str | None = None,
    emails: list[str] | None = None,
) -> str:
    """Определяет код организации ТЗ (НП, АЛ, МГ, АМ, МИ, БМ) по структуре 1С."""
    text = _normalize(f"{path} {name}")
    email_text = " ".join(emails or []).lower()

    if direction == "БМ" or any(marker in text for marker in _BMI_MARKERS):
        return "БМ"
    if any(marker in text for marker in _MGS_MARKERS) or "mgs_" in email_text or "mgs@" in email_text:
        return "МГ"
    if any(marker in text for marker in _AM_MARKERS):
        return "АМ"
    if any(marker in text for marker in _MI_MARKERS):
        return "МИ"
    if any(marker in text for marker in _ALMAZ_MARKERS) or "almaz" in email_text:
        return "АЛ"

    parts = [segment.strip() for segment in (path or "").split(" / ") if segment.strip()]
    if len(parts) >= 2:
        branch = _normalize(parts[1])
        if "метрогаз" in branch:
            return "МГ"
        if "амурск" in branch:
            return "АМ"
        if "милака" in branch:
            return "МИ"

    return "НП"


def _load_structure_rows(enterprise_path: Path | None) -> list[dict]:
    file_path = resolve_enterprise_positions_path(enterprise_path)
    enterprise = json.loads(file_path.read_text(encoding="utf-8"))
    rows = enterprise.get("structure_departments_with_codes") or []
    return [row for row in rows if _CODE_RE.match(str(row.get("code") or "").strip())]


def build_export_rows(
    *,
    enterprise_path: Path | None = None,
    rules_path: Path | None = None,
) -> tuple[list[dict], list[dict]]:
    rules = load_routing_rules(rules_path)
    directions = directions_by_code_from_rules(rules)
    records = build_department_records_for_db(rules, enterprise_path=enterprise_path)
    records_by_code = {record.code: record for record in records}

    structure_rows = _load_structure_rows(enterprise_path)
    structure_by_code = {str(row["code"]): row for row in structure_rows}

    all_codes = sorted(set(structure_by_code) | set(records_by_code))

    active_rows: list[dict] = []
    liquidated_rows: list[dict] = []

    for code in all_codes:
        structure = structure_by_code.get(code, {})
        record = records_by_code.get(code)

        path = str(structure.get("path") or (record.metadata.get("path") if record else "") or "").strip()
        onec_name = str(structure.get("name") or (record.metadata.get("onec_name") if record else "") or "").strip()
        routing_name = str(record.name if record else rules.get("department_names", {}).get(code, "")).strip()
        department_name = routing_name or onec_name or code

        emails: list[str] = []
        if record:
            emails = list(record.metadata.get("emails") or [])
            if record.email and record.email not in emails:
                emails.insert(0, record.email)
        primary_email = emails[0] if emails else ""
        email_cell = "; ".join(emails) if emails else ""

        direction = (record.direction if record else None) or directions.get(code) or ""
        liquidated = is_liquidated_for_export(onec_name or department_name, path)
        org_code = detect_organization(
            path=path,
            name=onec_name or department_name,
            code=code,
            direction=direction or None,
            emails=emails,
        )

        row = {
            "org_code": org_code,
            "org_name": ORG_FULL_NAMES.get(org_code, org_code),
            "code": code,
            "name": department_name,
            "onec_name": onec_name,
            "direction": direction,
            "path": path,
            "email": email_cell,
            "primary_email": primary_email,
            "active": not liquidated,
            "active_label": "да" if not liquidated else "нет",
        }

        if liquidated:
            liquidated_rows.append(row)
        else:
            active_rows.append(row)

    active_rows.sort(key=lambda item: (ORG_ORDER.index(item["org_code"]) if item["org_code"] in ORG_ORDER else 99, item["code"]))
    liquidated_rows.sort(key=lambda item: item["code"])
    return active_rows, liquidated_rows


def _autosize_columns(sheet, *, min_width: int = 10, max_width: int = 60) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[letter].width = width


def _style_header(sheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in sheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(
    output_path: Path,
    active_rows: list[dict],
    liquidated_rows: list[dict],
) -> dict[str, int]:
    wb = Workbook()

    # Sheet 1 — summary
    summary = wb.active
    summary.title = "Сводка"
    summary.append(["Организация (код)", "Организация (полное название)", "Подразделений", "Активных", "Ликвидированных"])
    org_stats: dict[str, dict[str, int]] = {
        org: {"total": 0, "active": 0, "liquidated": 0} for org in ORG_ORDER
    }
    for row in active_rows:
        org = row["org_code"]
        org_stats.setdefault(org, {"total": 0, "active": 0, "liquidated": 0})
        org_stats[org]["total"] += 1
        org_stats[org]["active"] += 1
    for row in liquidated_rows:
        org = row["org_code"]
        org_stats.setdefault(org, {"total": 0, "active": 0, "liquidated": 0})
        org_stats[org]["total"] += 1
        org_stats[org]["liquidated"] += 1

    for org in ORG_ORDER:
        stats = org_stats.get(org, {"total": 0, "active": 0, "liquidated": 0})
        summary.append([org, ORG_FULL_NAMES[org], stats["total"], stats["active"], stats["liquidated"]])

    extra_orgs = sorted(set(org_stats) - set(ORG_ORDER))
    for org in extra_orgs:
        stats = org_stats[org]
        summary.append([org, org, stats["total"], stats["active"], stats["liquidated"]])

    total_active = len(active_rows)
    total_liquidated = len(liquidated_rows)
    summary.append(["ИТОГО", "Все организации", total_active + total_liquidated, total_active, total_liquidated])
    _style_header(summary)
    _autosize_columns(summary)

    # Sheet 2 — departments
    departments = wb.create_sheet("Подразделения")
    dept_headers = [
        "Организация (код ТЗ)",
        "Организация (полное название)",
        "Код подразделения",
        "Название подразделения",
        "Направление",
        "Путь в 1С",
        "Email",
        "Активен",
    ]
    departments.append(dept_headers)
    for row in active_rows + liquidated_rows:
        departments.append(
            [
                row["org_code"],
                row["org_name"],
                row["code"],
                row["name"],
                row["direction"],
                row["path"],
                row["email"],
                row["active_label"],
            ]
        )
    _style_header(departments)
    _autosize_columns(departments, max_width=80)
    for cell in departments["G"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 3 — liquidated
    liquidated_sheet = wb.create_sheet("Ликвидированные")
    liquidated_sheet.append(dept_headers)
    for row in liquidated_rows:
        liquidated_sheet.append(
            [
                row["org_code"],
                row["org_name"],
                row["code"],
                row["name"],
                row["direction"],
                row["path"],
                row["email"],
                row["active_label"],
            ]
        )
    _style_header(liquidated_sheet)
    _autosize_columns(liquidated_sheet, max_width=80)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "Сводка": summary.max_row - 1,
        "Подразделения": departments.max_row - 1,
        "Ликвидированные": liquidated_sheet.max_row - 1,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Экспорт подразделений по организациям в Excel")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Путь к Excel-файлу (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    parser.add_argument("--enterprise", type=Path, help="Путь к enterprise_positions.json")
    parser.add_argument("--rules", type=Path, help="Путь к routing_rules.json")
    args = parser.parse_args()

    active_rows, liquidated_rows = build_export_rows(
        enterprise_path=args.enterprise,
        rules_path=args.rules,
    )
    counts = write_excel(args.output, active_rows, liquidated_rows)

    org_counter = Counter(row["org_code"] for row in active_rows)
    print(f"Файл: {args.output.resolve()}")
    print("Строки по листам:")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}")
    print("Активные подразделения по организациям:")
    for org in ORG_ORDER:
        print(f"  {org}: {org_counter.get(org, 0)}")


if __name__ == "__main__":
    main()
