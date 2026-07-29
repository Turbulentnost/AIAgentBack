"""Smoke: result.xlsx содержит лист «приоритет сборки» с окраской плана П/ф."""

from __future__ import annotations

import asyncio
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    UploadedWorkbook,
    _SHEET_DETAILED_PRIORITY,
    analyze_aveon_excel_files,
)
from app.agents.document_analysis_agent.product_coverage import _match_detailed_to_catalog


def main() -> None:
    catalog = ['FPV-перехватчик "СОКОЛ" И (день)', 'FPV-перехватчик "СОКОЛ" Т (ночь)']
    assert _match_detailed_to_catalog("Сокол И", catalog) == catalog[0]
    print("match ok")

    base = Path(r"c:\Users\uaa\Desktop\test")
    names = [
        "Отчет 07_2026_6148.xlsx",
        "Остатки 28.07.2026_4262.xls",
        "ГРАФИК ОТГРУЗОК (расширенный).xlsx",
        "График производства.xlsx",
    ]
    wbs = [UploadedWorkbook(filename=n, content=(base / n).read_bytes()) for n in names]
    result = asyncio.run(analyze_aveon_excel_files(wbs))
    assert result.result_xlsx_bytes
    wb = load_workbook(BytesIO(result.result_xlsx_bytes))
    print("sheets:", wb.sheetnames)
    assert _SHEET_DETAILED_PRIORITY in wb.sheetnames
    ws = wb[_SHEET_DETAILED_PRIORITY]
    green = yellow = red = 0
    for row in ws.iter_rows(
        min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 80)
    ):
        for cell in row:
            fill = cell.fill
            if fill is None or fill.fill_type != "solid":
                continue
            fg = str(getattr(fill.fgColor, "rgb", "") or "")
            if fg.endswith("C6EFCE"):
                green += 1
            elif fg.endswith("FFEB9C"):
                yellow += 1
            elif fg.endswith("FFC7CE"):
                red += 1
    print(f"fills green={green} yellow={yellow} red={red}")
    print(
        "D4",
        getattr(getattr(ws["D4"].fill, "fgColor", None), "rgb", None),
        "D5",
        getattr(getattr(ws["D5"].fill, "fgColor", None), "rgb", None),
    )
    assert green + yellow + red > 0
    out = base / "result_priority_smoke.xlsx"
    out.write_bytes(result.result_xlsx_bytes)
    print("wrote", out)


if __name__ == "__main__":
    main()
