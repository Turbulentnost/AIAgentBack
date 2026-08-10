"""Smoke: detailed schedule → result.xlsx with 2 sheets + demand check."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    ROLE_DETAILED_PRODUCTION_SCHEDULE,
    UploadedWorkbook,
    _extract_detailed_production_schedule,
    _match_detailed_plan_for_product,
    _normalize,
    analyze_aveon_excel_files,
)

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")


async def main() -> None:
    uploaded = [
        UploadedWorkbook(filename=path.name, content=path.read_bytes())
        for path in sorted(TEST_DIR.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]
    print("files:", [u.filename for u in uploaded])

    result = await analyze_aveon_excel_files(uploaded)
    print("roles:", result.roles)
    print("detailed_month:", result.detailed_schedule_month)
    print("detailed_files:", result.detailed_production_schedule_files)
    assert result.result_xlsx_bytes, "no result.xlsx"
    assert result.detailed_schedule_month, "expected selected detailed month"
    # as_of ≈ сегодня (июль 2026 в тестовых данных) → лист «Июль», не «Апрель»
    assert result.detailed_schedule_month == "2026-07", result.detailed_schedule_month

    out = ROOT / "scripts" / "_smoke_result_daily.xlsx"
    out.write_bytes(result.result_xlsx_bytes)
    wb = load_workbook(out)
    print("sheets:", wb.sheetnames)
    assert wb.sheetnames[0] == "1-производственный план (мес.)" or "1-производственный план (мес.)" in wb.sheetnames
    assert "обеспечение (Июль)" in wb.sheetnames

    daily = wb["обеспечение (Июль)"]
    print("daily A1:", daily["A1"].value)
    print(
        "daily A5:",
        daily["A5"].value,
        "F5 demand:",
        daily["F5"].value,
        "H5 formula:",
        daily["H5"].value,
    )
    assert daily["H5"].value and str(daily["H5"].value).startswith("=")

    matched = next(
        (r for r in result.merged_nomenclatures if any(v > 0 for v in r.daily_demand.values())),
        None,
    )
    assert matched is not None, "no nonzero daily demand"
    day_key = next(k for k, v in matched.daily_demand.items() if v > 0)
    print(
        "sample:",
        matched.nomenclature[:70],
        "day",
        day_key,
        "demand",
        matched.daily_demand[day_key],
    )

    detailed_files = [
        u
        for u in uploaded
        if result.roles.get(u.filename) == ROLE_DETAILED_PRODUCTION_SCHEDULE
    ]
    extract = _extract_detailed_production_schedule(
        detailed_files,
        {u.filename: ROLE_DETAILED_PRODUCTION_SCHEDULE for u in detailed_files},
    )
    print(
        "re-extract month",
        extract.year,
        extract.month,
        "plans",
        len(extract.plans),
        [p.product for p in extract.plans[:8]],
    )
    sokol = next((p for p in extract.plans if _normalize(p.product) == _normalize("Сокол И")), None)
    if sokol and extract.month == 7:
        assert sokol.daily_qty.get("2026-07-01") == 500.0
        assert 1900.0 not in {
            qty for day, qty in sokol.daily_qty.items() if day.endswith("-07-04") is False
        } or True
        # «Итог» 1900 must not appear as any day's qty for Сокол И
        assert 1900.0 not in sokol.daily_qty.values(), sokol.daily_qty

    plans_by_key = {_normalize(p.product): p for p in extract.plans}
    plan_names = [p.product for p in extract.plans]
    recon = 0.0
    for product, spec_qty in matched.by_product.items():
        plan = _match_detailed_plan_for_product(product, plans_by_key, plan_names)
        if plan is None or spec_qty is None:
            continue
        recon += float(plan.daily_qty.get(day_key, 0.0)) * float(spec_qty)
    print("reconstructed", recon, "stored", matched.daily_demand[day_key])
    assert abs(recon - matched.daily_demand[day_key]) < 1e-6

    monthly = wb["1-производственный план (мес.)"]
    print("monthly A5:", monthly["A5"].value, "E5:", monthly["E5"].value)
    assert monthly["A5"].value
    nonzero_daily = sum(
        1 for r in result.merged_nomenclatures if any(v > 0 for v in r.daily_demand.values())
    )
    print("daily_demand_nonzero", nonzero_daily)
    print("OK")


if __name__ == "__main__":
    asyncio.run(main())
