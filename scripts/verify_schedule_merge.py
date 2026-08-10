"""Verify universal schedule merge on test files."""
from __future__ import annotations

import asyncio
import base64
import io
import json
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl

from app.agents.document_analysis_agent.temp_schedule_merge import (
    NameDict,
    NomRow,
    _conflicting_models,
    _find_existing_key,
    _is_safe_same_item,
    _item_signature,
    _merge_rows,
    _norm_name,
    _parse_schedule_layout,
    detect_sheet_layout,
    merge_schedule_files,
)

GRAFIK = Path(r"c:\Users\uaa\Desktop\test объединение\ГРАФИК ОТГРУЗОК (расширенный).xlsx")
TAM = Path(r"c:\Users\uaa\Desktop\test объединение\ТАМОЖНЯ.xlsx")
OUT = Path(r"c:\Users\uaa\Downloads\merged_schedule_universal.xlsx")


async def audit() -> int:
    files = [
        (GRAFIK.name, GRAFIK.read_bytes()),
        (TAM.name, TAM.read_bytes()),
    ]
    result = await merge_schedule_files(files)
    print("OK:", result.get("ok"), result.get("message"))
    stats = result.get("stats", {})
    brief = {k: v for k, v in stats.items() if k != "layouts_detected"}
    print("Stats:", json.dumps(brief, ensure_ascii=False, indent=2))
    print("Layouts:")
    for layout in stats.get("layouts_detected", []):
        print(" ", layout)

    if not result.get("ok"):
        return 1

    raw = base64.b64decode(result["file_base64"])
    OUT.write_bytes(raw)
    print("Saved:", OUT)

    wb = openpyxl.load_workbook(GRAFIK, data_only=True)
    source_by_date: dict[date, float] = defaultdict(float)
    source_rows: list[NomRow] = []
    for sn in wb.sheetnames:
        layout = await detect_sheet_layout(wb[sn], sn, GRAFIK.name)
        if layout.kind != "schedule":
            continue
        items, _ = _parse_schedule_layout(wb[sn], layout, NameDict())
        source_rows.extend(items)
        for it in items:
            for d, q in it.schedule.items():
                source_by_date[d] += q

    merged_wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = merged_wb["График"]
    headers = [c.value for c in ws[1]]
    date_cols: dict[int, date] = {}
    for i, h in enumerate(headers):
        if isinstance(h, str) and len(h) == 10 and h[4] == "-":
            date_cols[i + 1] = date.fromisoformat(h)

    merged_by_date: dict[date, float] = defaultdict(float)
    merged_names: list[str] = []
    merged_by_key: dict[str, NomRow] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0])
        merged_names.append(name)
        item = NomRow(name=name)
        for col, d in date_cols.items():
            q = row[col - 1]
            if q:
                item.schedule[d] = float(q)
                merged_by_date[d] += float(q)
        merged_by_key[_norm_name(name)] = item

    source_merged = _merge_rows(source_rows)

    def resolve_merged(it: NomRow) -> NomRow | None:
        key = _find_existing_key(it.name, merged_by_key)
        if key is not None:
            return merged_by_key[key]
        for mr in merged_by_key.values():
            if _is_safe_same_item(it.name, mr.name):
                return mr
        return None

    miss: list[str] = []
    schedule_loss: list[str] = []
    for it in source_merged.values():
        merged_item = resolve_merged(it)
        if merged_item is None:
            miss.append(it.name)
            continue
        for d, q in it.schedule.items():
            got = merged_item.schedule.get(d, 0.0)
            if got + 0.01 < q:
                schedule_loss.append(f"{it.name} @ {d}: src={q} merged={got}")

    date_loss: dict[str, dict[str, float]] = {}
    for d in sorted(set(source_by_date)):
        s = source_by_date.get(d, 0.0)
        m = merged_by_date.get(d, 0.0)
        if m + 0.01 < s:
            date_loss[d.isoformat()] = {"source": s, "merged": m, "delta": m - s}

    sig_map: dict[str, list[str]] = defaultdict(list)
    for n in merged_names:
        sig = _item_signature(n)
        if sig:
            sig_map[sig].append(n)
    wrongly_merged: list[tuple[str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()
    for row in merged_by_key.values():
        sig = _item_signature(row.name)
        if not sig:
            continue
        for other in merged_by_key.values():
            if other is row:
                continue
            if _item_signature(other.name) != sig:
                continue
            pair = tuple(sorted([row.name, other.name]))
            if pair in seen_pairs:
                continue
            if _conflicting_models(*pair) or not _is_safe_same_item(*pair):
                continue
            seen_pairs.add(pair)
            wrongly_merged.append(pair)

    print("--- AUDIT ---")
    print("source_rows:", len(source_rows), "source_unique:", len(source_merged), "merged_noms:", len(merged_names))
    print("miss:", len(miss))
    if miss:
        for m in miss[:10]:
            print("  -", m)
    print("schedule_loss:", len(schedule_loss))
    for s in schedule_loss[:5]:
        print(" ", s)
    print("date_loss (grafik only):", len(date_loss))
    for k, v in list(date_loss.items())[:5]:
        print(" ", k, v)
    print("wrongly_merged_pairs:", len(wrongly_merged))
    for a, b in wrongly_merged[:5]:
        print(" ", a, "||", b)

    ok = len(miss) == 0 and len(schedule_loss) == 0 and len(date_loss) == 0 and len(wrongly_merged) == 0
    print("PASS" if ok else "FAIL")
    return 0 if ok else 2


if __name__ == "__main__":
    raise SystemExit(asyncio.run(audit()))
