"""Импорт правил маршрутизации из data/imports/spiski/ в routing_rules.json и tz_department_topics.json."""

from __future__ import annotations

import json
import re
import shutil
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook

NETWORK_SRC = Path(
    r"\\192.168.1.198\project\управление делами\Орг контр рег вх корр\Списки"
)
IMPORT_DIR = ROOT / "data" / "imports" / "spiski"
RULES_PATH = ROOT / "data" / "routing_rules.json"
TZ_TOPICS_PATH = ROOT / "data" / "tz_department_topics.json"
REPORT_PATH = IMPORT_DIR / "import_report.json"

_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
_CODE_RE = re.compile(r"00-\d{6}")
_CLASSIFICATION_GLOB = "Образец для классификации писем*.xlsx"
_STRUCTURE_XLS = "управ.структура для распределения 21.04.2025.xls"
_IT_XLSX = "Список эл адресов для IT.xlsx"

_IT_SECTION_MAP = {
    "бухгалтер": "00-000002",
    "суп": "00-000063",
    "кадр": "00-000063",
    "юрид": "00-000044",
    "тендер": "00-000054",
    "оркк": "00-000076",
    "опму": "00-000074",
    "омто": "00-000065",
    "сервис": "00-000163",
    "вед": "00-000015",
    "развит": "00-000013",
    "омис": "00-000025",
    "метролог": "00-000025",
    "отк": "00-000100",
    "отп": "00-000099",
    "ахо": "00-000046",
    "бми": "00-000163",
    "одп": "00-000155",
    "управление дел": "00-000066",
    "офис": "00-000066",
}

_CONTENT_SKIP = frozenset(
    {
        "если",
        "кроме",
        "смотрим",
        "направить",
        "рассмотрение",
        "ознакомление",
        "исполнение",
        "подписание",
        "входящий запрос не попадает под правила",
    }
)
_MAX_CONTENT_KEYWORDS_PER_CODE = 12
_MAX_KEYWORD_LEN = 40


def _is_routing_keyword(kw: str) -> bool:
    if len(kw) < 4 or len(kw) > _MAX_KEYWORD_LEN:
        return False
    if kw in _CONTENT_SKIP:
        return False
    if any(ch in kw for ch in "()[]{}"):
        return False
    if kw.startswith("если ") or " смотрим " in f" {kw} ":
        return False
    return True


def _version_key(name: str) -> tuple[int, str]:
    m = re.search(r"версия\s*(\d+)", name, re.I)
    if m:
        return (int(m.group(1)), name)
    if "19_05_2025" in name:
        return (3, name)
    return (0, name)


def find_classification_file() -> Path | None:
    files = sorted(IMPORT_DIR.glob(_CLASSIFICATION_GLOB), key=lambda p: _version_key(p.name))
    return files[-1] if files else None


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace("ё", "е")


def _find_header_row(rows: list[tuple]) -> int | None:
    for idx, row in enumerate(rows):
        texts = [_normalize_header(c) for c in row]
        if any("электронн" in t and "почт" in t for t in texts):
            return idx
    return None


def _map_classification_columns(header: list[str]) -> dict[str, int]:
    col: dict[str, int] = {}
    for j, h in enumerate(header):
        if "электронн" in h and "почт" in h:
            col["email"] = j
        elif h.startswith("о чем"):
            col["about"] = j
        elif "процесс" in h:
            col["process"] = j
        elif "время" in h:
            col["deadline"] = j
        elif "должност" in h:
            col["position"] = j
        elif "код подраздел" in h:
            col["code"] = j
        elif "направлен" in h:
            col["direction"] = j
    return col


def _split_keywords(text: str) -> list[str]:
    keywords: list[str] = []
    for part in re.split(r"[\n,;]+", text or ""):
        kw = part.strip().lower()
        kw = re.sub(r"\s+", " ", kw)
        if _is_routing_keyword(kw):
            keywords.append(kw)
    return keywords


def _select_content_keywords(
    candidates: dict[str, list[str]],
    existing_by_code: dict[str, set[str]],
) -> dict[str, list[str]]:
    """Отбирает короткие частотные ключи; не более N на код."""
    selected: dict[str, list[str]] = {}
    for code, kws in candidates.items():
        existing = existing_by_code.get(code, set())
        freq: dict[str, int] = defaultdict(int)
        for kw in kws:
            if kw not in existing:
                freq[kw] += 1
        ranked = sorted(freq.items(), key=lambda x: (-x[1], len(x[0]), x[0]))
        picked: list[str] = []
        for kw, count in ranked:
            if count < 2 and len(kw) > 20:
                continue
            picked.append(kw)
            if len(picked) >= _MAX_CONTENT_KEYWORDS_PER_CODE:
                break
        if picked:
            selected[code] = picked
    return selected


def parse_classification_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hdr_idx = _find_header_row(rows)
    if hdr_idx is None:
        return []

    header = [_normalize_header(c) for c in rows[hdr_idx]]
    col = _map_classification_columns(header)
    if "email" not in col or "code" not in col:
        return []

    parsed: list[dict] = []
    for row in rows[hdr_idx + 1 :]:
        if not row or all(c is None for c in row):
            continue
        code = str(row[col["code"]] or "").strip()
        if not _CODE_RE.fullmatch(code):
            continue

        emails = _EMAIL_RE.findall(str(row[col["email"]] or ""))
        about = str(row[col["about"]] or "").strip() if "about" in col else ""
        direction = str(row[col["direction"]] or "").strip() if "direction" in col else ""
        position = str(row[col["position"]] or "").strip() if "position" in col else ""
        process = str(row[col["process"]] or "").strip() if "process" in col else ""
        deadline = str(row[col["deadline"]] or "").strip() if "deadline" in col else ""

        for email in emails:
            email = email.lower()
            if "turbo--don" in email:
                continue
            parsed.append(
                {
                    "email": email,
                    "dept_code": code,
                    "dept_name": position,
                    "direction": direction,
                    "about": about,
                    "process": process,
                    "deadline": deadline,
                    "source_file": path.name,
                    "source_type": "classification_xlsx",
                }
            )
    return parsed


def parse_process_sheet(path: Path) -> list[dict]:
    """Лист «процесс»: продукт/услуга → тема → процесс (для content_rules)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.worksheets if "процесс" in s.title.lower()), None)
    if sheet is None:
        wb.close()
        return []

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    parsed: list[dict] = []
    for row in rows:
        if not row:
            continue
        texts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(texts) < 2:
            continue
        product = texts[0] if len(texts) >= 3 else ""
        about = texts[1] if len(texts) >= 3 else texts[0]
        process = texts[2] if len(texts) >= 4 else (texts[-1] if len(texts) >= 2 else "")
        keywords = _split_keywords(about)
        if product:
            keywords.extend(_split_keywords(product))
        if not keywords:
            continue
        parsed.append(
            {
                "keywords": keywords,
                "about": about[:120],
                "process": process,
                "source_type": "process_sheet",
                "source_file": path.name,
            }
        )
    return parsed


def parse_it_email_list(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    section = ""
    parsed: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cell = str(row[1] or "").strip() if len(row) > 1 else ""
        explicit_code = str(row[4] or "").strip() if len(row) > 4 else ""
        if not cell:
            continue
        if not _EMAIL_RE.search(cell) and row[0] is None and len(cell) > 2:
            section = cell.lower()
            continue
        emails = _EMAIL_RE.findall(cell)
        code = explicit_code if _CODE_RE.fullmatch(explicit_code) else ""
        if not code:
            for key, dept_code in _IT_SECTION_MAP.items():
                if key in section:
                    code = dept_code
                    break
        for email in emails:
            parsed.append(
                {
                    "email": email.lower(),
                    "dept_code": code,
                    "dept_name": section,
                    "source_file": path.name,
                    "source_type": "it_email_list",
                }
            )
    wb.close()
    return parsed


def parse_structure_xls(path: Path) -> dict[str, str]:
    try:
        import xlrd
    except ImportError:
        return {}

    book = xlrd.open_workbook(path)
    names: dict[str, str] = {}
    for sheet in book.sheets():
        for i in range(1, sheet.nrows):
            row = sheet.row_values(i)
            if len(row) < 2:
                continue
            code = str(row[0]).strip()
            name = str(row[1]).strip()
            if not _CODE_RE.fullmatch(code) or not name:
                continue
            if "(ликв.)" in name.lower() or name.startswith("_"):
                continue
            names[code] = name
    return names


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def existing_exact_emails(rules: dict) -> dict[str, dict]:
    return {str(r["email"]).lower(): r for r in rules.get("exact_email_rules", [])}


def existing_email_keywords(rules: dict) -> set[str]:
    return {str(r["keyword"]).lower() for r in rules.get("email_keyword_rules", [])}


def content_keywords_by_code(rules: dict) -> dict[str, set[str]]:
    out: dict[str, set[str]] = defaultdict(set)
    for rule in rules.get("content_rules", []):
        code = str(rule.get("code", ""))
        for kw in rule.get("keywords", []):
            out[code].add(str(kw).lower())
    return out


def tz_emails_by_code(tz: dict) -> dict[str, set[str]]:
    out: dict[str, set[str]] = {}
    for code, entry in tz.items():
        out[code] = {e.lower() for e in entry.get("emails", [])}
    return out


def direction_for_code(code: str, rules: dict, row: dict) -> str:
    explicit = str(row.get("direction") or "").strip()
    if explicit:
        return explicit
    for rule_list in ("exact_email_rules", "email_keyword_rules", "content_rules"):
        for rule in rules.get(rule_list, []):
            if str(rule.get("code")) == code and rule.get("direction"):
                return str(rule["direction"])
    return "ПР"


def dept_name_for_code(code: str, rules: dict, row: dict, structure_names: dict[str, str]) -> str:
    for source in (
        str(row.get("dept_name") or "").strip(),
        str(rules.get("department_names", {}).get(code, "")).strip(),
        structure_names.get(code, ""),
    ):
        if source and source != code:
            return source
    return code


def email_covered_by_keyword(email: str, kw_set: set[str]) -> bool:
    local = email.split("@", 1)[0]
    return any(kw in local for kw in kw_set)


def merge_import(
    rules: dict,
    tz: dict,
    *,
    classification_rows: list[dict],
    it_rows: list[dict],
    process_rows: list[dict],
    structure_names: dict[str, str],
    source_label: str,
) -> dict:
    stats: dict = {
        "source": source_label,
        "import_date": date.today().isoformat(),
        "new_exact_email_rules": [],
        "new_content_keywords": [],
        "new_department_names": {},
        "tz_emails_added": [],
        "tz_process_metadata": [],
        "skipped_duplicates": {"exact_email": 0, "email_keyword_covered": 0, "tz_email": 0, "content_keyword": 0},
        "skipped_conflicts": [],
    }

    exact_map = existing_exact_emails(rules)
    kw_set = existing_email_keywords(rules)
    content_by_code = content_keywords_by_code(rules)
    tz_emails = tz_emails_by_code(tz)

    # Конфликты email → несколько отделов в классификации
    email_to_codes: dict[str, set[str]] = defaultdict(set)
    for row in classification_rows:
        if row.get("email"):
            email_to_codes[row["email"]].add(row["dept_code"])
    conflict_emails = {e for e, codes in email_to_codes.items() if len(codes) > 1}

    all_rows = classification_rows + [r for r in it_rows if r.get("dept_code")]

    for row in all_rows:
        email = str(row.get("email") or "").strip().lower()
        code = str(row.get("dept_code") or "").strip()
        if not email or not code:
            continue

        name = dept_name_for_code(code, rules, row, structure_names)
        if code not in rules.get("department_names", {}) and name != code:
            rules.setdefault("department_names", {})[code] = name
            stats["new_department_names"][code] = name

        # tz_department_topics
        tz.setdefault(code, {"topics": [], "names": [], "emails": []})
        if name and name not in tz[code].get("names", []):
            if not tz[code].get("names"):
                tz[code]["names"] = [name]
        process = str(row.get("process") or "").strip()
        deadline = str(row.get("deadline") or "").strip()
        if process or deadline:
            meta = tz[code].setdefault("task_defaults", {})
            if process and not meta.get("process"):
                meta["process"] = process
            if deadline and not meta.get("deadline"):
                meta["deadline"] = deadline
            if row.get("source_type") == "classification_xlsx":
                stats["tz_process_metadata"].append({"code": code, "process": process, "deadline": deadline})

        if email not in tz_emails.get(code, set()):
            tz[code].setdefault("emails", []).append(email)
            tz_emails.setdefault(code, set()).add(email)
            stats["tz_emails_added"].append({"code": code, "email": email, "source": row.get("source_type")})
        else:
            stats["skipped_duplicates"]["tz_email"] += 1

        if email in exact_map:
            stats["skipped_duplicates"]["exact_email"] += 1
            continue
        if email in conflict_emails:
            stats["skipped_conflicts"].append(
                {"email": email, "codes": sorted(email_to_codes[email]), "reason": "multi_dept_in_classification"}
            )
            continue
        if email_covered_by_keyword(email, kw_set):
            stats["skipped_duplicates"]["email_keyword_covered"] += 1
            continue

        about_text = str(row.get("about") or "")[:120]
        if not about_text and name:
            about_text = name[:120]
        new_rule = {
            "email": email,
            "code": code,
            "name": name,
            "direction": direction_for_code(code, rules, row),
            "about": about_text,
        }
        org = str(row.get("direction") or "").strip()
        if org in {"АЛ", "МГ", "АМ", "МИ", "БМ"}:
            new_rule["organization"] = org
        rules.setdefault("exact_email_rules", []).append(new_rule)
        exact_map[email] = new_rule
        stats["new_exact_email_rules"].append({**new_rule, "source": row.get("source_type")})

    # content_rules из «О чем» классификации (консервативно: частотные короткие фразы)
    content_candidates: dict[str, list[str]] = defaultdict(list)
    for row in classification_rows:
        code = row.get("dept_code", "")
        content_candidates[code].extend(_split_keywords(str(row.get("about") or "")))

    selected_content = _select_content_keywords(content_candidates, content_by_code)
    for code, new_kws in selected_content.items():
        row = next((r for r in classification_rows if r.get("dept_code") == code), {})
        existing = content_by_code.get(code, set())
        rule = next((r for r in rules.get("content_rules", []) if str(r.get("code")) == code), None)
        if rule is None:
            rule = {
                "keywords": [],
                "code": code,
                "name": dept_name_for_code(code, rules, row, structure_names),
                "direction": direction_for_code(code, rules, row),
                "about": str(row.get("about") or "")[:80],
            }
            rules.setdefault("content_rules", []).append(rule)
        for kw in new_kws:
            rule["keywords"].append(kw)
            existing.add(kw)
            stats["new_content_keywords"].append({"code": code, "keyword": kw})

    # Полные темы классификации — в отдельный справочник (не в RuleRouter)
    topics_path = IMPORT_DIR / "classification_topics.json"
    topics_export: dict[str, list[dict]] = defaultdict(list)
    for row in classification_rows:
        code = row["dept_code"]
        topics_export[code].append(
            {
                "email": row.get("email"),
                "about": row.get("about"),
                "process": row.get("process"),
                "deadline": row.get("deadline"),
                "direction": row.get("direction"),
                "position": row.get("dept_name"),
            }
        )
    save_json(topics_path, {"source": source_label, "topics_by_code": dict(topics_export)})
    stats["classification_topics_file"] = str(topics_path.relative_to(ROOT))

    referenced_codes = {str(r.get("dept_code") or r.get("code") or "") for r in all_rows}
    referenced_codes.update(content_candidates.keys())
    referenced_codes.discard("")

    for code in referenced_codes:
        name = structure_names.get(code)
        if name and code not in rules.get("department_names", {}):
            rules.setdefault("department_names", {})[code] = name
            stats["new_department_names"][code] = name

    return stats


def write_readme(stats: dict) -> None:
    readme = IMPORT_DIR / "README.md"
    files = [f for f in stats.get("files_analyzed", []) if f not in {"README.md", "import_report.json"}]
    readme.write_text(
        f"""# Импорт из сетевой папки «Списки»

Дата: {stats['import_date']}
Источник: `\\\\192.168.1.198\\project\\управление делами\\Орг контр рег вх корр\\Списки`

## Файлы ({len(files)})
{chr(10).join('- ' + f for f in files)}

## Классификация
Использован: **{stats.get('classification_file_used', '—')}**

## Результат
| Метрика | До | После | Добавлено |
|---------|-----|-------|-----------|
| exact_email_rules | {stats['counts']['exact_email_rules_before']} | {stats['counts']['exact_email_rules_after']} | {stats['counts']['exact_email_rules_added']} |
| content keywords | {stats['counts']['content_keywords_before']} | {stats['counts']['content_keywords_after']} | {stats['counts']['content_keywords_added']} |
| tz emails | {stats['counts']['tz_emails_before']} | {stats['counts']['tz_emails_after']} | {stats['counts']['tz_emails_added']} |
| department_names | {stats['counts']['department_names_before']} | {stats['counts']['department_names_after']} | {stats['counts']['department_names_added']} |

Конфликты (пропущены): {len(stats.get('skipped_conflicts', []))}

Подробности: `import_report.json`
""",
        encoding="utf-8",
    )


def main() -> int:
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)

    if NETWORK_SRC.is_dir():
        for f in NETWORK_SRC.iterdir():
            if f.is_file():
                shutil.copy2(f, IMPORT_DIR / f.name)

    classification_path = find_classification_file()
    if not classification_path:
        print("Нет файлов классификации в", IMPORT_DIR)
        return 1

    print(f"Классификация: {classification_path.name}")
    classification_rows = parse_classification_xlsx(classification_path)
    process_rows = parse_process_sheet(classification_path)
    print(f"  email-строк: {len(classification_rows)}, лист процесс: {len(process_rows)}")

    it_rows: list[dict] = []
    it_path = IMPORT_DIR / _IT_XLSX
    if it_path.is_file():
        it_rows = parse_it_email_list(it_path)
        print(f"  IT-список: {len(it_rows)} email")

    structure_names: dict[str, str] = {}
    structure_path = IMPORT_DIR / _STRUCTURE_XLS
    if structure_path.is_file():
        structure_names = parse_structure_xls(structure_path)
        print(f"  структура 1С: {len(structure_names)} активных отделов")

    rules = load_json(RULES_PATH)
    tz = load_json(TZ_TOPICS_PATH)

    before = {
        "exact": len(rules.get("exact_email_rules", [])),
        "content_kw": sum(len(r.get("keywords", [])) for r in rules.get("content_rules", [])),
        "tz_emails": sum(len(v.get("emails", [])) for v in tz.values()),
        "dept_names": len(rules.get("department_names", {})),
    }

    stats = merge_import(
        rules,
        tz,
        classification_rows=classification_rows,
        it_rows=it_rows,
        process_rows=process_rows,
        structure_names=structure_names,
        source_label=classification_path.name,
    )

    after = {
        "exact": len(rules.get("exact_email_rules", [])),
        "content_kw": sum(len(r.get("keywords", [])) for r in rules.get("content_rules", [])),
        "tz_emails": sum(len(v.get("emails", [])) for v in tz.values()),
        "dept_names": len(rules.get("department_names", {})),
    }

    stats["files_analyzed"] = sorted(
        p.name for p in IMPORT_DIR.iterdir() if p.is_file() and p.name != "import_report.json"
    )
    stats["classification_file_used"] = classification_path.name
    stats["counts"] = {
        "exact_email_rules_before": before["exact"],
        "exact_email_rules_after": after["exact"],
        "exact_email_rules_added": after["exact"] - before["exact"],
        "content_keywords_before": before["content_kw"],
        "content_keywords_after": after["content_kw"],
        "content_keywords_added": after["content_kw"] - before["content_kw"],
        "tz_emails_before": before["tz_emails"],
        "tz_emails_after": after["tz_emails"],
        "tz_emails_added": after["tz_emails"] - before["tz_emails"],
        "department_names_before": before["dept_names"],
        "department_names_after": after["dept_names"],
        "department_names_added": after["dept_names"] - before["dept_names"],
    }

    save_json(RULES_PATH, rules)
    save_json(TZ_TOPICS_PATH, tz)
    save_json(REPORT_PATH, stats)
    write_readme(stats)

    print("\n=== ИТОГ ===")
    print(json.dumps(stats["counts"], ensure_ascii=False, indent=2))
    print(f"Конфликты: {len(stats.get('skipped_conflicts', []))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
