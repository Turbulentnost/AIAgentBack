"""Сводная таблица: отдел → темы писем, которые должны в него приходить.

Источники:
  - data/imports/spiski/Образец для классификации писем версия 6.xlsx (основной)
  - data/imports/spiski/classification_topics.json (если есть)
  - data/tz_department_topics.json
  - data/routing_rules.json (department_names, content_rules)

Листы результата (data/department_topics_summary.xlsx):
  - Сводная по отделам — код, подразделение, темы, email-адреса
  - Детализация — каждая строка: email/тема → отдел
  - По ключевым словам — content_rules и ключи из «О чем»

Пример:
  python scripts/export_department_topics_summary.py
  python scripts/export_department_topics_summary.py -o data/department_topics_summary.xlsx
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_OUTPUT = ROOT / "data" / "department_topics_summary.xlsx"
IMPORT_DIR = ROOT / "data" / "imports" / "spiski"
RULES_PATH = ROOT / "data" / "routing_rules.json"
TZ_TOPICS_PATH = ROOT / "data" / "tz_department_topics.json"
TOPICS_JSON_PATH = IMPORT_DIR / "classification_topics.json"
STRUCTURE_XLS = IMPORT_DIR / "управ.структура для распределения 21.04.2025.xls"

_CODE_RE = re.compile(r"00-\d{6}")


def _load_import_module():
    path = ROOT / "scripts" / "import_spiski_routing.py"
    spec = importlib.util.spec_from_file_location("import_spiski_routing", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def _load_json(path: Path) -> dict:
    if not path.is_file():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _normalize_topic(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _split_about_lines(text: str) -> list[str]:
    """Разбивает поле «О чем» на отдельные темы/фразы."""
    parts: list[str] = []
    for chunk in re.split(r"[\n;,]+", text or ""):
        item = _normalize_topic(chunk)
        if not item:
            continue
        if item.lower().startswith("(кроме"):
            continue
        parts.append(item)
    return parts


def _unique_preserve(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out


def resolve_dept_name(
    code: str,
    *,
    classification_name: str = "",
    rules: dict,
    structure_names: dict[str, str],
    tz_topics: dict,
) -> str:
    for source in (
        classification_name.strip(),
        str(rules.get("department_names", {}).get(code, "")).strip(),
        str((tz_topics.get(code) or {}).get("names", [""])[0]).strip()
        if (tz_topics.get(code) or {}).get("names")
        else "",
        structure_names.get(code, "").strip(),
    ):
        if source and source != code:
            return source
    return code


def load_classification_rows(
    import_mod,
    *,
    classification_path: Path | None = None,
) -> tuple[list[dict], str]:
    path = classification_path or import_mod.find_classification_file()
    if not path or not path.is_file():
        return [], "файл классификации не найден"

    rows = import_mod.parse_classification_xlsx(path)
    return rows, path.name


def load_topics_from_json(path: Path) -> dict[str, list[dict]]:
    data = _load_json(path)
    return data.get("topics_by_code") or {}


def build_detail_rows(classification_rows: list[dict]) -> list[dict]:
    detail: list[dict] = []
    for row in classification_rows:
        code = str(row.get("dept_code") or "").strip()
        if not _CODE_RE.fullmatch(code):
            continue
        about = _normalize_topic(str(row.get("about") or ""))
        detail.append(
            {
                "dept_code": code,
                "dept_name": str(row.get("dept_name") or "").strip(),
                "email": str(row.get("email") or "").strip().lower(),
                "about": about,
                "topic_lines": _split_about_lines(about),
                "direction": str(row.get("direction") or "").strip(),
                "process": str(row.get("process") or "").strip(),
                "deadline": str(row.get("deadline") or "").strip(),
                "source_file": str(row.get("source_file") or "").strip(),
            }
        )
    detail.sort(key=lambda r: (r["dept_code"], r["email"], r["about"]))
    return detail


def build_summary_rows(
    detail_rows: list[dict],
    *,
    rules: dict,
    structure_names: dict[str, str],
    tz_topics: dict,
) -> list[dict]:
    by_code: dict[str, dict] = {}

    for row in detail_rows:
        code = row["dept_code"]
        entry = by_code.setdefault(
            code,
            {
                "dept_code": code,
                "dept_name": "",
                "topics": [],
                "about_texts": [],
                "emails": [],
                "directions": [],
            },
        )

        if row["dept_name"] and not entry["dept_name"]:
            entry["dept_name"] = row["dept_name"]
        if row["email"] and row["email"] not in entry["emails"]:
            entry["emails"].append(row["email"])
        if row["about"] and row["about"] not in entry["about_texts"]:
            entry["about_texts"].append(row["about"])
        for topic in row["topic_lines"]:
            entry["topics"].append(topic)
        if row["direction"] and row["direction"] not in entry["directions"]:
            entry["directions"].append(row["direction"])

    # Дополнить email из tz_department_topics
    for code, tz_entry in tz_topics.items():
        if not _CODE_RE.fullmatch(str(code)):
            continue
        bucket = by_code.setdefault(
            code,
            {
                "dept_code": code,
                "dept_name": "",
                "topics": [],
                "about_texts": [],
                "emails": [],
                "directions": [],
            },
        )
        for email in tz_entry.get("emails") or []:
            email = str(email).strip().lower()
            if email and email not in bucket["emails"]:
                bucket["emails"].append(email)
        for topic in tz_entry.get("topics") or []:
            bucket["topics"].append(str(topic))

    summary: list[dict] = []
    for code in sorted(by_code):
        entry = by_code[code]
        dept_name = resolve_dept_name(
            code,
            classification_name=entry["dept_name"],
            rules=rules,
            structure_names=structure_names,
            tz_topics=tz_topics,
        )
        topics = _unique_preserve(entry["topics"])
        about_texts = _unique_preserve(entry["about_texts"])
        topics_text = "\n".join(about_texts) if about_texts else "\n".join(topics)
        if about_texts and topics:
            extra = [t for t in topics if t not in about_texts]
            if extra:
                topics_text = topics_text + "\n---\n" + "\n".join(extra)

        summary.append(
            {
                "dept_code": code,
                "dept_name": dept_name,
                "topics_text": topics_text,
                "topic_count": len(topics) or len(about_texts),
                "emails_text": "; ".join(sorted(entry["emails"])),
                "email_count": len(entry["emails"]),
                "directions": ", ".join(sorted(entry["directions"])),
            }
        )

    return summary


def build_keyword_rows(
    *,
    rules: dict,
    detail_rows: list[dict],
    rules_mod,
) -> list[dict]:
    rows: list[dict] = []

    content_rules = rules.get("content_rules") or []
    for rule in content_rules:
        code = str(rule.get("code") or "").strip()
        if not code:
            continue
        name = str(rule.get("name") or rules.get("department_names", {}).get(code, code))
        keywords = rule.get("keywords") or []
        for kw in keywords:
            rows.append(
                {
                    "dept_code": code,
                    "dept_name": name,
                    "keyword": str(kw),
                    "source": "content_rules",
                    "about": str(rule.get("about") or ""),
                }
            )

    if not rows:
        candidates: dict[str, list[str]] = defaultdict(list)
        for row in detail_rows:
            code = row["dept_code"]
            for topic in row["topic_lines"]:
                for kw in rules_mod._split_keywords(topic):
                    candidates[code].append(kw)
            for kw in rules_mod._split_keywords(row["about"]):
                candidates[code].append(kw)

        for code in sorted(candidates):
            name = str(rules.get("department_names", {}).get(code, code))
            for kw in _unique_preserve(candidates[code]):
                rows.append(
                    {
                        "dept_code": code,
                        "dept_name": name,
                        "keyword": kw,
                        "source": "classification_about",
                        "about": "",
                    }
                )

    rows.sort(key=lambda r: (r["dept_code"], r["keyword"]))
    return rows


def _autosize_columns(sheet, *, min_width: int = 10, max_width: int = 80) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells:
            if cell.value is None:
                continue
            lines = str(cell.value).split("\n")
            width = max(width, min(max(len(line) for line in lines) + 2, max_width))
        sheet.column_dimensions[letter].width = width


def _style_header(sheet, row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    for cell in sheet[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(
    output_path: Path,
    *,
    summary_rows: list[dict],
    detail_rows: list[dict],
    keyword_rows: list[dict],
    meta: dict,
) -> dict[str, int]:
    wb = Workbook()

    summary = wb.active
    summary.title = "Сводная по отделам"
    summary_headers = [
        "Код отдела",
        "Подразделение",
        "Темы писем",
        "Email-адреса",
        "Кол-во email",
        "Направления",
    ]
    summary.append(summary_headers)
    for row in summary_rows:
        summary.append(
            [
                row["dept_code"],
                row["dept_name"],
                row["topics_text"],
                row["emails_text"],
                row["email_count"],
                row["directions"],
            ]
        )
    _style_header(summary)
    _autosize_columns(summary, max_width=90)
    for col in ("C", "D"):
        for cell in summary[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    detail = wb.create_sheet("Детализация")
    detail_headers = [
        "Код отдела",
        "Подразделение",
        "Email",
        "О чем (тема)",
        "Направление",
        "Процесс",
        "Срок",
        "Источник",
    ]
    detail.append(detail_headers)
    for row in detail_rows:
        detail.append(
            [
                row["dept_code"],
                row["dept_name"],
                row["email"],
                row["about"],
                row["direction"],
                row["process"],
                row["deadline"],
                row["source_file"],
            ]
        )
    _style_header(detail)
    _autosize_columns(detail, max_width=90)
    for col in ("C", "D"):
        for cell in detail[col]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    keywords = wb.create_sheet("По ключевым словам")
    kw_headers = ["Код отдела", "Подразделение", "Ключевое слово", "Источник", "Контекст"]
    keywords.append(kw_headers)
    for row in keyword_rows:
        keywords.append(
            [
                row["dept_code"],
                row["dept_name"],
                row["keyword"],
                row["source"],
                row["about"],
            ]
        )
    _style_header(keywords)
    _autosize_columns(keywords, max_width=70)

    meta_sheet = wb.create_sheet("Мета")
    meta_sheet.append(["Параметр", "Значение"])
    for key, value in meta.items():
        meta_sheet.append([key, value])
    _style_header(meta_sheet)
    _autosize_columns(meta_sheet, max_width=90)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "Сводная по отделам": summary.max_row - 1,
        "Детализация": detail.max_row - 1,
        "По ключевым словам": keywords.max_row - 1,
        "Мета": meta_sheet.max_row - 1,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Сводная таблица отдел → темы писем")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Путь к Excel (по умолчанию: {DEFAULT_OUTPUT.name})",
    )
    parser.add_argument(
        "--classification",
        type=Path,
        help="Путь к файлу «Образец для классификации писем*.xlsx»",
    )
    parser.add_argument("--rules", type=Path, default=RULES_PATH)
    parser.add_argument("--tz-topics", type=Path, default=TZ_TOPICS_PATH)
    args = parser.parse_args()

    import_mod = _load_import_module()
    rules = _load_json(args.rules)
    tz_topics = _load_json(args.tz_topics)

    structure_names: dict[str, str] = {}
    if STRUCTURE_XLS.is_file():
        structure_names = import_mod.parse_structure_xls(STRUCTURE_XLS)

    classification_rows, source_file = load_classification_rows(
        import_mod,
        classification_path=args.classification,
    )
    if not classification_rows and TOPICS_JSON_PATH.is_file():
        topics_by_code = load_topics_from_json(TOPICS_JSON_PATH)
        for code, items in topics_by_code.items():
            for item in items:
                classification_rows.append(
                    {
                        "email": item.get("email"),
                        "dept_code": code,
                        "dept_name": item.get("position"),
                        "direction": item.get("direction"),
                        "about": item.get("about"),
                        "process": item.get("process"),
                        "deadline": item.get("deadline"),
                        "source_file": TOPICS_JSON_PATH.name,
                        "source_type": "classification_topics_json",
                    }
                )
        source_file = f"{TOPICS_JSON_PATH.name} (fallback)"

    detail_rows = build_detail_rows(classification_rows)
    summary_rows = build_summary_rows(
        detail_rows,
        rules=rules,
        structure_names=structure_names,
        tz_topics=tz_topics,
    )
    keyword_rows = build_keyword_rows(rules=rules, detail_rows=detail_rows, rules_mod=import_mod)

    meta = {
        "Файл классификации": source_file,
        "Строк детализации": len(detail_rows),
        "Отделов в сводной": len(summary_rows),
        "Ключевых слов": len(keyword_rows),
        "Структура 1С (отделов)": len(structure_names),
    }

    counts = write_excel(
        args.output,
        summary_rows=summary_rows,
        detail_rows=detail_rows,
        keyword_rows=keyword_rows,
        meta=meta,
    )

    print(f"Файл: {args.output.resolve()}")
    print(f"Источник: {source_file}")
    print("Строки по листам:")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count}")

    print("\nПримеры (Сводная по отделам):")
    for row in summary_rows[:5]:
        preview = row["topics_text"].replace("\n", " / ")[:100]
        print(f"  {row['dept_code']} | {row['dept_name']} | {preview}...")

    return 0 if detail_rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
