"""Call analyze-excel API and verify two sheets in result.xlsx."""
from __future__ import annotations

import base64
import json
import urllib.request
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")
FILES = [
    TEST_DIR / "С остатками.xlsx",
    TEST_DIR / "План по недельно.xlsx",
    TEST_DIR / "График производства.xlsx",
    TEST_DIR / "ГРАФИК ОТГРУЗОК (расширенный).xlsx",
]


def main() -> None:
    boundary = "----cursorboundary"
    body = b""
    for path in FILES:
        body += f"--{boundary}\r\n".encode()
        body += (
            f'Content-Disposition: form-data; name="files"; filename="{path.name}"\r\n'
        ).encode()
        body += b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        body += path.read_bytes()
        body += b"\r\n"
    body += f"--{boundary}--\r\n".encode()

    req = urllib.request.Request(
        "http://127.0.0.1:5454/api/v1/agents/document-analysis/analyze-excel",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=180) as resp:
        data = json.loads(resp.read().decode())

    print("detailed_month", data.get("detailed_schedule_month"))
    print("detailed_files", data.get("detailed_production_schedule_files"))
    print("daily_nonzero", data.get("daily_demand_nonzero_count"))
    raw = base64.b64decode(data["file_base64"])
    out = Path(__file__).resolve().parent / "_api_result_daily.xlsx"
    out.write_bytes(raw)
    wb = load_workbook(BytesIO(raw))
    print("sheets", wb.sheetnames)
    assert "1-производственный план (мес.)" in wb.sheetnames
    daily_name = next(name for name in wb.sheetnames if name.startswith("обеспечение ("))
    print("daily A1", wb[daily_name]["A1"].value)
    print("OK", out)


if __name__ == "__main__":
    main()
