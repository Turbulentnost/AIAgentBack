# -*- coding: utf-8 -*-
"""Анализ листа «4-обеспеченность по изделиям»: классификация расходников."""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"c:\Users\uaa\Downloads\result (57).xlsx")
OUT = Path(__file__).resolve().parents[1] / "data" / "aveon" / "классификация_расходники_обеспеченность_изделий.xlsx"
SHEET = "4-обеспеченность по изделиям"

CONSUMABLE_PATTERNS = [
    r"\bболт\b",
    r"\bгайк",
    r"\bшайб",
    r"\bвинт\b",
    r"\bшпильк",
    r"\bзаклеп",
    r"\bсаморез",
    r"\bанкер",
    r"\bшплинт",
    r"\bштифт",
    r"\bхомут",
    r"\bстяжк",
    r"\bскоб",
    r"шлиф",
    r"наждак",
    r"абразив",
    r"круг\s*шлиф",
    r"липучк",
    r"скотч",
    r"изолент",
    r"электрод",
    r"проволок.*свар",
    r"свароч",
    r"флюс",
    r"клей",
    r"герметик",
    r"силикон",
    r"мастик",
    r"антифрик",
    r"смазк",
    r"\bмасло\b",
    r"солидол",
    r"краск",
    r"грунт",
    r"эмаль",
    r"растворит",
    r"очистит",
    r"обезжир",
    r"антикор",
    r"салфет",
    r"ветош",
    r"тряп",
    r"перчат",
    r"респиратор",
    r"упаков",
    r"стретч",
    r"пленк.*упак",
    r"пакет",
    r"этикет",
    r"бирк",
    r"маркиров",
    r"сверл",
    r"фрез",
    r"метчик",
    r"плашк",
    r"разверт",
    r"оснастк",
    r"\bбита\b",
    r"лента.*клей",
    r"двухсторон",
]

NON_CONSUMABLE_PATTERNS = [
    r"полуфабрик",
    r"\bп/ф\b",
    r"полуф\b",
    r"\bузел\b",
    r"\bблок\b",
    r"агрегат",
    r"механизм",
    r"редуктор",
    r"двигател",
    r"насос",
    r"контроллер",
    r"\bплат",
    r"плата\b",
    r"датчик",
    r"реле",
    r"концевик",
    r"энкодер",
    r"инвертор",
    r"жгут",
    r"кабель",
    r"разъем",
    r"коннектор",
    r"фишк",
    r"корпус",
    r"рама",
    r"каркас",
    r"стойк",
    r"опор",
    r"балк",
    r"лонжерон",
    r"поперечин",
    r"\bлист\b",
    r"листов",
    r"профил",
    r"труб",
    r"уголок",
    r"швеллер",
    r"двутавр",
    r"полос",
    r"отлив",
    r"поков",
    r"штампов",
    r"заготов",
    r"колес",
    r"шина\b",
    r"камера\b",
    r"сидень",
    r"обивк",
    r"чехол",
    r"стекл",
    r"зеркал",
    r"аккумулятор",
    r"батаре",
    r"комплект\b",
    r"сборк",
    r"\bвал\b",
    r"шестерн",
    r"муфт",
    r"подшипник",
    r"гидро",
    r"цилиндр",
    r"манжет",
    r"сальник",
    r"сolenoid",
    r"solenoid",
    r"электромагнит",
    r"втулк",
    r"эксцентрик",
    r"рейка",
    r"фильтр\b",
    r"радиатор",
    r"антенн",
    r"модуль",
    r"платы\b",
]

CONSUMABLE_RE = [re.compile(p, re.I) for p in CONSUMABLE_PATTERNS]
NON_CONSUMABLE_RE = [re.compile(p, re.I) for p in NON_CONSUMABLE_PATTERNS]


@dataclass
class MonthMetrics:
    available: float = 0.0
    plan: float = 0.0
    fact: float = 0.0


@dataclass
class NomenclatureRow:
    product: str
    nomenclature: str
    months: dict[str, MonthMetrics] = field(default_factory=dict)

    @property
    def max_plan(self) -> float:
        return max((m.plan for m in self.months.values()), default=0.0)

    @property
    def total_plan(self) -> float:
        return sum(m.plan for m in self.months.values())

    @property
    def all_zero_available(self) -> bool:
        if not self.months:
            return True
        return all(abs(m.available) < 1e-9 for m in self.months.values())


def classify(name: str, *, max_plan: float, all_zero: bool) -> tuple[str, str, str]:
    n = (name or "").strip()
    lower = n.lower()

    for rx in NON_CONSUMABLE_RE:
        if rx.search(lower):
            return (
                "не расходник (комплектующее / возможно в цехе)",
                "высокая",
                "покупная деталь/узел по наименованию; нулевой остаток часто = WIP в цехе или ожидание поставки",
            )

    for rx in CONSUMABLE_RE:
        if rx.search(lower):
            return (
                "расходник",
                "высокая",
                "расходный материал/крепёж/оснастка по наименованию",
            )

    if re.search(r"\bм\d+\b", lower) or re.search(r"din\s*\d+", lower) or re.search(r"iso\s*\d+", lower):
        return ("расходник", "средняя", "стандартный крепёж/метиз (DIN/ISO/M)")

    if re.search(r"\d+[xх×]\d+", lower) and any(w in lower for w in ("лист", "труб", "проф", "полос", "угол")):
        return (
            "не расходник (комплектующее / возможно в цехе)",
            "средняя",
            "материал/заготовка с типоразмером",
        )

    if all_zero and max_plan > 0:
        if any(w in lower for w in ("комплект", "узел", "блок", "корпус", "рама", "вал", "ось")):
            return (
                "не расходник (комплектующее / возможно в цехе)",
                "средняя",
                "нулевой остаток при плане — вероятнее комплектующее/WIP",
            )
        if len(n) > 35:
            return (
                "не расходник (комплектующее / возможно в цехе)",
                "низкая",
                "уникальное длинное наименование — проверить как покупную деталь",
            )
        return (
            "расходник",
            "низкая",
            "нулевой остаток при плане без признаков узла — проверить закупщиком",
        )

    if all_zero:
        return (
            "не расходник (комплектующее / возможно в цехе)",
            "низкая",
            "нулевой остаток без плана — неактивная позиция или учёт в цехе",
        )

    return (
        "не расходник (комплектующее / возможно в цехе)",
        "низкая",
        "остаток > 0 или нет явных признаков расходника",
    )


def parse_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[str], list[NomenclatureRow]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = next(s for s in wb.sheetnames if "обеспеченность" in s.lower())

    ws = wb[sheet_name]
    months: list[str] = []
    col = 2
    while ws.cell(3, col).value is not None:
        months.append(str(ws.cell(3, col).value).strip())
        col += 3
    if not months:
        raise RuntimeError("Не найдены месяцы в строке 3")

    products: list[str] = []
    rows: list[NomenclatureRow] = []
    current_product: str | None = None

    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None or str(name).strip() == "":
            continue
        name_s = str(name).strip()
        outline = ws.row_dimensions[r].outline_level or 0
        is_detail = outline >= 1

        month_data: dict[str, MonthMetrics] = {}
        for mi, month in enumerate(months):
            base = 2 + mi * 3
            month_data[month] = MonthMetrics(
                available=float(ws.cell(r, base).value or 0),
                plan=float(ws.cell(r, base + 1).value or 0),
                fact=float(ws.cell(r, base + 2).value or 0),
            )

        if not is_detail:
            current_product = name_s
            products.append(name_s)
            continue

        if current_product:
            rows.append(
                NomenclatureRow(product=current_product, nomenclature=name_s, months=month_data)
            )

    wb.close()
    return months, products, rows


def write_sheet(ws, data: list[dict], col_widths: list[float] | None = None) -> None:
    if not data:
        ws.cell(1, 1, "нет данных")
        return

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = list(data[0].keys())
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = wrap

    for r, rec in enumerate(data, 2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, rec[h])
            cell.border = border
            cell.alignment = wrap

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else SRC
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else OUT

    months, products, nom_rows = parse_sheet(src, SHEET)

    records: list[dict] = []
    for row in nom_rows:
        cat, conf, rationale = classify(
            row.nomenclature, max_plan=row.max_plan, all_zero=row.all_zero_available
        )
        rec = {
            "Изделие": row.product,
            "Номенклатура": row.nomenclature,
            "Классификация": cat,
            "Уверенность": conf,
            "Обоснование": rationale,
            "Нулевой остаток (все месяцы)": "да" if row.all_zero_available else "нет",
            "Макс. план": round(row.max_plan, 3),
            "Сумма план": round(row.total_plan, 3),
        }
        for m in months:
            rec[f"{m} обесп"] = round(row.months[m].available, 3)
            rec[f"{m} план"] = round(row.months[m].plan, 3)
        records.append(rec)

    zero_rows = [r for r in records if r["Нулевой остаток (все месяцы)"] == "да"]
    zero_non_consumable = [r for r in zero_rows if r["Классификация"].startswith("не расходник")]
    zero_consumable = [r for r in zero_rows if r["Классификация"] == "расходник"]

    by_product: dict[str, dict[str, int]] = defaultdict(
        lambda: {"total": 0, "consumable": 0, "non_consumable": 0, "zero_non_consumable": 0}
    )
    for r in records:
        p = r["Изделие"]
        by_product[p]["total"] += 1
        if r["Классификация"] == "расходник":
            by_product[p]["consumable"] += 1
        else:
            by_product[p]["non_consumable"] += 1
        if r["Нулевой остаток (все месяцы)"] == "да" and r["Классификация"].startswith("не расходник"):
            by_product[p]["zero_non_consumable"] += 1

    summary_records = [
        {
            "Изделие": p,
            "Номенклатур в спеке": by_product[p]["total"],
            "Расходники": by_product[p]["consumable"],
            "Не расходники (комплектующие)": by_product[p]["non_consumable"],
            "Нулевые - не расходник (возможно в цехе)": by_product[p]["zero_non_consumable"],
        }
        for p in products
    ]

    methodology = [
        {
            "Раздел": "Цель",
            "Описание": (
                "Разделить номенклатуры спецификаций на расходники и комплектующие; "
                "среди нулевых остатков выделить позиции, которые вероятнее в цехе/WIP."
            ),
        },
        {
            "Раздел": "Источник",
            "Описание": f"Лист «{SHEET}» файла {src.name}",
        },
        {
            "Раздел": "Расходник",
            "Описание": (
                "Крепёж, метизы, абразив, сварка, клеи, смазки, краски, упаковка, "
                "мелкая оснастка — списываются в процессе."
            ),
        },
        {
            "Раздел": "Не расходник",
            "Описание": (
                "Узлы, электроника, жгуты, корпуса, рамы, заготовки, полуфабрикаты — "
                "при нулевом остатке часто WIP на участке или ожидание поставки."
            ),
        },
        {
            "Раздел": "Ограничение",
            "Описание": "Эвристика по наименованию; строки с «низкая» уверенность — ручная проверка.",
        },
        {
            "Раздел": "Статистика",
            "Описание": (
                f"Изделий: {len(products)}; номенклатур: {len(records)}; "
                f"нулевых: {len(zero_rows)}; нулевых «не расходник»: {len(zero_non_consumable)}; "
                f"нулевых «расходник»: {len(zero_consumable)}"
            ),
        },
    ]

    out.parent.mkdir(parents=True, exist_ok=True)
    wb_out = Workbook()

    ws1 = wb_out.active
    ws1.title = "По изделиям и номенклатурам"
    write_sheet(ws1, records, [28, 48, 34, 12, 52, 14, 12, 12])

    ws2 = wb_out.create_sheet("Нулевые не расходники")
    write_sheet(ws2, zero_non_consumable, [28, 48, 34, 12, 52, 14, 12, 12])

    ws2b = wb_out.create_sheet("Нулевые расходники")
    write_sheet(ws2b, zero_consumable, [28, 48, 34, 12, 52, 14, 12, 12])

    ws3 = wb_out.create_sheet("Сводка по изделиям")
    write_sheet(ws3, summary_records, [28, 16, 12, 22, 28])

    ws4 = wb_out.create_sheet("Методика")
    write_sheet(ws4, methodology, [18, 90])

    wb_out.save(out)

    print(f"months={months}")
    print(f"products={len(products)} nomenclatures={len(records)}")
    print(f"zero={len(zero_rows)} zero_non_consumable={len(zero_non_consumable)} zero_consumable={len(zero_consumable)}")
    print(f"saved={out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
