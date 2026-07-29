"""Smoke: daily plan/fact columns from detailed schedule."""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    UploadedWorkbook,
    analyze_aveon_excel_files,
)

TEST_DIR = Path(r"c:\Users\uaa\Desktop\test")
NAMES = [
    "Отчет 07_2026_6148.xlsx",
    "График производства.xlsx",
    "С остатками.xlsx",
    "ГРАФИК ОТГРУЗОК (расширенный).xlsx",
]


async def main() -> None:
    uploaded = [
        UploadedWorkbook(filename=name, content=(TEST_DIR / name).read_bytes())
        for name in NAMES
    ]
    result = await analyze_aveon_excel_files(uploaded)
    print("month", result.detailed_schedule_month)
    print("detailed_files", result.detailed_production_schedule_files)
    assert result.result_xlsx_bytes
    out = ROOT / "scripts" / "_smoke_daily_plan_fact.xlsx"
    out.write_bytes(result.result_xlsx_bytes)

    wb = load_workbook(out, data_only=False)
    assert "обеспечение (Июль)" in wb.sheetnames
    ws = wb["обеспечение (Июль)"]
    print("H4", ws.cell(4, 8).value)
    print("I4", ws.cell(4, 9).value)
    print("J4", ws.cell(4, 10).value)
    print("K4", ws.cell(4, 11).value)
    assert "план" in str(ws.cell(4, 8).value).lower()
    assert "факт" in str(ws.cell(4, 9).value).lower()
    assert ws.max_column == 7 + 31 * 4

    found = False
    for row in range(5, min(ws.max_row, 200) + 1):
        products = str(ws.cell(row, 2).value or "")
        if "СОКОЛ" in products.upper() and "день" in products.lower():
            print("sample row", row, products[:80])
            print(
                "01.07 plan/fact/receipt",
                ws.cell(row, 8).value,
                ws.cell(row, 9).value,
                ws.cell(row, 10).value,
            )
            print("forecast formula", ws.cell(row, 11).value)
            assert float(ws.cell(row, 8).value or 0) > 0
            assert float(ws.cell(row, 9).value or 0) > 0
            assert str(ws.cell(row, 11).value).startswith("=F")
            found = True
            break
    assert found, "no СОКОЛ day product row"

    nz_plan = sum(
        1 for r in result.merged_nomenclatures if any(v > 0 for v in r.daily_demand.values())
    )
    nz_fact = sum(
        1
        for r in result.merged_nomenclatures
        if any(v > 0 for v in r.daily_demand_fact.values())
    )
    print("nonzero plan rows", nz_plan, "fact rows", nz_fact)
    assert nz_plan > 0 and nz_fact > 0
    print("OK", out)


if __name__ == "__main__":
    asyncio.run(main())
