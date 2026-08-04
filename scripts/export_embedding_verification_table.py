"""Excel-таблица для проверки BGE-индексации на реальных письмах."""

from __future__ import annotations

import argparse
import json
import random
import sys
from datetime import datetime
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import selectinload  # noqa: E402

from agent_pochta.config import PROJECT_ROOT, get_settings  # noqa: E402
from agent_pochta.db.message_filters import load_payload_dict, resolved_turbo_recipient  # noqa: E402
from agent_pochta.db.models import EmailMessageRow  # noqa: E402
from agent_pochta.db.session import get_session_factory  # noqa: E402
from agent_pochta.services.department_knowledge import (  # noqa: E402
    collect_department_correction_records,
    dept_name,
    load_department_display_names,
)
from agent_pochta.services.department_correction_indexing import record_id_for  # noqa: E402
from agent_pochta.services.email_indexing import build_indexing_text_from_row  # noqa: E402
from agent_pochta.services.embedding_client import EmbeddingClientError, embed_texts  # noqa: E402
from agent_pochta.services.email_rag_qdrant import (  # noqa: E402
    DEPARTMENT_CORRECTIONS_COLLECTION,
    search_department_corrections,
)


def _indexed_correction_record_ids(settings) -> set[str]:
    try:
        from qdrant_client import QdrantClient

        client = QdrantClient(url=settings.qdrant_url, prefer_grpc=False)
        indexed: set[str] = set()
        offset = None
        while True:
            points, offset = client.scroll(
                collection_name=DEPARTMENT_CORRECTIONS_COLLECTION,
                limit=256,
                offset=offset,
                with_payload=["record_id"],
                with_vectors=False,
            )
            for point in points:
                rid = (point.payload or {}).get("record_id")
                if rid:
                    indexed.add(str(rid))
            if offset is None:
                break
        client.close()
        return indexed
    except Exception:
        return set()


def _format_dept(code: str, name: str) -> str:
    if code and name and name != code:
        return f"{code} — {name}"
    return code or name or ""


def _top1_dept(embed_text: str, settings) -> tuple[str, float | None]:
    if not embed_text.strip() or not settings.embedding_base_url:
        return "", None
    try:
        vector = embed_texts([embed_text], settings=settings)[0]
        hits = search_department_corrections(
            url=settings.qdrant_url, query_vector=vector, limit=1
        )
        if not hits:
            return "", None
        hit = hits[0]
        label = _format_dept(
            str(hit.get("dept_correct_id") or ""),
            str(hit.get("dept_correct_name") or ""),
        )
        return label, hit.get("score")
    except EmbeddingClientError:
        return "BGE error", None


def export_verification_table(
    *,
    sample_size: int = 50,
    holdout_size: int = 20,
    out_path: Path | None = None,
) -> Path:
    from openpyxl import Workbook

    settings = get_settings()
    names = load_department_display_names()
    factory = get_session_factory()
    indexed_correction_ids = _indexed_correction_record_ids(settings)

    with factory() as session:
        correction_records = collect_department_correction_records(session)

    out_dir = PROJECT_ROOT / "data" / "stats"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    out_path = out_path or (out_dir / f"embedding_verification_{stamp}.xlsx")

    wb = Workbook()
    ws_corr = wb.active
    ws_corr.title = "Коррекции отделов"
    ws_corr.append(
        [
            "Кому (@turbo-don.ru)",
            "От кого",
            "Содержимое для эмбеддинга",
            "Отдел (ошибочный)",
            "Отдел (исправленный)",
            "Проиндексировано",
            "Top-1 поиск",
            "Score",
            "source",
        ]
    )

    indexed_ok = 0
    top1_ok = 0
    holdout_checked = 0

    holdout_pool = [r for r in correction_records if r.dept_correct_id and r.embed_text.strip()]
    holdout_sample = random.sample(holdout_pool, min(holdout_size, len(holdout_pool)))
    holdout_ids = {id(r) for r in holdout_sample}

    for record in correction_records:
        top1, score = _top1_dept(record.embed_text, settings)
        if id(record) in holdout_ids and record.dept_correct_id and top1.startswith(record.dept_correct_id):
            top1_ok += 1
        if id(record) in holdout_ids:
            holdout_checked += 1

        ws_corr.append(
            [
                record.recipient,
                record.sender_email,
                record.embed_text,
                _format_dept(record.dept_wrong_id, record.dept_wrong_name),
                _format_dept(record.dept_correct_id, record.dept_correct_name),
                "да" if record_id_for(record) in indexed_correction_ids else "нет",
                top1,
                score,
                record.source,
            ]
        )

    with factory() as session:
        rows = session.scalars(
            select(EmailMessageRow)
            .order_by(EmailMessageRow.received_at.desc())
            .limit(sample_size)
            .options(selectinload(EmailMessageRow.attachments))
        ).all()

        ws_sample = wb.create_sheet("Все письма sample")
        ws_sample.append(
            [
                "Кому (@turbo-don.ru)",
                "От кого",
                "Содержимое для эмбеддинга",
                "Отдел (ошибочный)",
                "Отдел (исправленный)",
                "Проиндексировано",
            ]
        )

        for row in rows:
            payload = load_payload_dict(row.raw_payload_json) or {}
            recipient = resolved_turbo_recipient(mailbox=row.mailbox, payload=payload)
            embed_text = build_indexing_text_from_row(row)
            indexed = bool(payload.get("qdrant_indexed_at"))
            if indexed:
                indexed_ok += 1
            indexed_label = payload.get("qdrant_indexed_at") or ("да" if indexed else "нет")
            ws_sample.append(
                [
                    recipient,
                    row.sender_email,
                    embed_text,
                    _format_dept(row.department_id or "", dept_name(names, row.department_id, row.department_name)),
                    "",
                    indexed_label,
                ]
            )

    ws_summary = wb.create_sheet("Сводка")
    accuracy = (top1_ok / holdout_checked * 100) if holdout_checked else 0
    ws_summary.append(["Метрика", "Значение"])
    ws_summary.append(["Коррекций в анализе", len(correction_records)])
    ws_summary.append(["Писем в sample", len(rows)])
    ws_summary.append(["Sample проиндексировано", indexed_ok])
    ws_summary.append(["Holdout для top-1", holdout_checked])
    ws_summary.append(["Top-1 совпал с исправленным отделом", top1_ok])
    ws_summary.append(["Top-1 accuracy %", round(accuracy, 1)])
    ws_summary.append(["QDRANT_URL", settings.qdrant_url])
    ws_summary.append(["EMBEDDING_BASE_URL", settings.embedding_base_url])

    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Export embedding verification Excel")
    parser.add_argument("--sample-size", type=int, default=50)
    parser.add_argument("--holdout-size", type=int, default=20)
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()

    path = export_verification_table(
        sample_size=args.sample_size,
        holdout_size=args.holdout_size,
        out_path=args.output,
    )
    print(json.dumps({"ok": True, "path": str(path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
