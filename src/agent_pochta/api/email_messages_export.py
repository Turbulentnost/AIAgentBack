"""Excel-отчёт для режима «Вид 1С» (summary + detail по периоду)."""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from agent_pochta.db.message_filters import MSK, msk_day_end_exclusive_utc, msk_day_start_utc
from agent_pochta.services.erp_attachments import display_erp_document_number
from agent_pochta.stats.classification_log import collect_operator_approvals

ExportPeriod = Literal["day", "week", "month"]

_MSK = MSK

# Колонки детального листа — как в IncomingMailTable, без «Гриф» и «Ответственный».
EXPORT_DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("operator_review_label", "Отметка"),
    ("spam_label", "Спам"),
    ("mail_date_display", "Дата"),
    ("erp_document_number", "Номер"),
    ("attachments_display", "Влож."),
    ("organization_name", "Организация"),
    ("sender_email", "Email отправителя"),
    ("partner_name", "Партнер"),
    ("department_id", "Кому"),
    ("department_name", "Кому (подразделение)"),
    ("payer_direction_label", "Плательщик-направление"),
)

_OPERATOR_REVIEW_LABELS = {
    "pending": "Не проверено",
    "verified": "Проверено",
    "corrected": "Исправлено",
}

_PERIOD_LABELS = {
    "day": "День (сегодня, MSK)",
    "week": "Неделя (7 дней, MSK)",
    "month": "Месяц (30 дней, MSK)",
}


def resolve_export_period(period: ExportPeriod, *, now: datetime | None = None) -> tuple[date, date]:
    """Границы периода по received_at в Europe/Moscow (включительно).

    - day: календарный день «сегодня»;
    - week: последние 7 календарных дней, включая сегодня;
    - month: последние 30 календарных дней, включая сегодня.
    """
    if now is None:
        now = datetime.now(_MSK)
    elif now.tzinfo is None:
        now = now.replace(tzinfo=_MSK)
    else:
        now = now.astimezone(_MSK)

    today = now.date()
    if period == "day":
        return today, today
    if period == "week":
        return today - timedelta(days=6), today
    if period == "month":
        return today - timedelta(days=29), today
    raise ValueError(f"Unknown period: {period}")


def export_filename(period: ExportPeriod, *, now: datetime | None = None) -> str:
    date_from, date_to = resolve_export_period(period, now=now)
    if date_from == date_to:
        suffix = date_from.isoformat()
    else:
        suffix = f"{date_from.isoformat()}_{date_to.isoformat()}"
    return f"vid_1c_report_{period}_{suffix}.xlsx"


def _format_msk_datetime(value: str | None) -> str:
    if not value:
        return "—"
    text = str(value).strip()
    if not text:
        return "—"
    normalized = text.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(normalized)
    except ValueError:
        return text
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return (
        dt.astimezone(_MSK)
        .replace(tzinfo=None)
        .strftime("%d.%m.%Y %H:%M")
    )


def _spam_label(row_dict: dict[str, Any]) -> str:
    status = str(row_dict.get("status") or "").strip()
    is_spam = bool(row_dict.get("is_spam"))
    if status == "spam" or is_spam:
        return "Спам"
    if status == "done" and not is_spam:
        return "Не спам"
    if status:
        return status
    return "—"


def _attachments_display(row_dict: dict[str, Any]) -> str:
    summary = row_dict.get("attachments_summary") or []
    if isinstance(summary, list) and summary:
        names = [
            str(item.get("filename") or "").strip()
            for item in summary
            if isinstance(item, dict) and str(item.get("filename") or "").strip()
        ]
        if names:
            return ", ".join(names)
    count = row_dict.get("attachments_count")
    if isinstance(count, int) and count > 0:
        return str(count)
    return "—"


def row_dict_to_export_detail(row_dict: dict[str, Any]) -> dict[str, Any]:
    state = str(row_dict.get("operator_review_state") or "pending")
    mail_date = row_dict.get("mail_date") or row_dict.get("received_at")
    return {
        "operator_review_label": _OPERATOR_REVIEW_LABELS.get(state, state),
        "spam_label": _spam_label(row_dict),
        "mail_date_display": _format_msk_datetime(
            str(mail_date) if mail_date is not None else None
        ),
        "erp_document_number": display_erp_document_number(
            row_dict.get("erp_document_number")
        )
        or "—",
        "attachments_display": _attachments_display(row_dict),
        "organization_name": str(row_dict.get("organization_name") or row_dict.get("organization") or "").strip() or "—",
        "sender_email": str(row_dict.get("sender_email") or "").strip() or "—",
        "partner_name": str(row_dict.get("partner_name") or "").strip() or "—",
        "department_id": str(row_dict.get("department_id") or "").strip() or "—",
        "department_name": str(row_dict.get("department_name") or "").strip() or "—",
        "payer_direction_label": str(row_dict.get("payer_direction_label") or "").strip() or "—",
    }


def build_export_summary(
    *,
    period: ExportPeriod,
    date_from: date,
    date_to: date,
    by_status: dict[str, int],
    operator_review_counts: dict[str, int],
    operator_approvals: dict[str, Any] | None = None,
    erp_created: int = 0,
    erp_skipped: int = 0,
    total: int,
) -> list[tuple[str, Any]]:
    start_local = datetime.combine(date_from, datetime.min.time()).strftime("%d.%m.%Y")
    end_local = datetime.combine(date_to, datetime.min.time()).strftime("%d.%m.%Y")
    period_text = start_local if date_from == date_to else f"{start_local} — {end_local}"

    spam_count = int(by_status.get("spam", 0))
    done_count = int(by_status.get("done", 0))
    error_count = int(by_status.get("error", 0))
    processing_count = int(by_status.get("processing", 0))
    verified_count = int(operator_review_counts.get("verified", 0))
    corrected_count = int(operator_review_counts.get("corrected", 0))
    pending_count = int(operator_review_counts.get("pending", 0))
    reviewed_count = verified_count + corrected_count

    rows: list[tuple[str, Any]] = [
        ("Период", period_text),
        ("Тип периода", _PERIOD_LABELS[period]),
        ("Часовой пояс", "Europe/Moscow"),
        ("", ""),
        ("Всего писем", total),
        ("Проверок (verified + corrected)", reviewed_count),
        ("Проверено (без правок)", verified_count),
        ("Исправлено оператором", corrected_count),
        ("Не проверено", pending_count),
        ("", ""),
        ("Обработано (done)", done_count),
        ("ERP создано", erp_created),
        ("ERP пропущено", erp_skipped),
        ("В обработке (processing)", processing_count),
        ("Спам", spam_count),
        ("Ошибки (error)", error_count),
    ]

    if operator_approvals:
        rows.extend(
            [
                ("", ""),
                ("Принято без изменений (saved)", int(operator_approvals.get("saved", 0))),
                ("Принято с правками (changed)", int(operator_approvals.get("changed", 0))),
            ]
        )
        rate = operator_approvals.get("rate")
        if isinstance(rate, (int, float)):
            rows.append(("Доля без изменений", f"{round(rate * 100, 1)}%"))

    for status, count in sorted(by_status.items()):
        if status in {"done", "spam", "error", "processing"}:
            continue
        rows.append((f"Статус: {status}", int(count)))

    return rows


def build_export_xlsx(
    *,
    period: ExportPeriod,
    date_from: date,
    date_to: date,
    summary_rows: list[tuple[str, Any]],
    detail_rows: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    ws_summary["A1"] = "Отчёт «Вид 1С»"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"Сформировано: {datetime.now(_MSK).strftime('%d.%m.%Y %H:%M')} MSK"

    header_font = Font(bold=True)
    ws_summary["A4"] = "Показатель"
    ws_summary["B4"] = "Значение"
    ws_summary["A4"].font = header_font
    ws_summary["B4"].font = header_font

    for index, (label, value) in enumerate(summary_rows, start=5):
        ws_summary[f"A{index}"] = label
        ws_summary[f"B{index}"] = value

    ws_summary.column_dimensions["A"].width = 42
    ws_summary.column_dimensions["B"].width = 28

    ws_detail = wb.create_sheet("Письма")
    for col_index, (_, header) in enumerate(EXPORT_DETAIL_COLUMNS, start=1):
        cell = ws_detail.cell(row=1, column=col_index, value=header)
        cell.font = header_font

    for row_index, detail in enumerate(detail_rows, start=2):
        for col_index, (key, _) in enumerate(EXPORT_DETAIL_COLUMNS, start=1):
            ws_detail.cell(row=row_index, column=col_index, value=detail.get(key, "—"))

    for col_index, (_, header) in enumerate(EXPORT_DETAIL_COLUMNS, start=1):
        letter = get_column_letter(col_index)
        max_len = len(header)
        for row_index in range(2, len(detail_rows) + 2):
            value = ws_detail.cell(row=row_index, column=col_index).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws_detail.column_dimensions[letter].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def collect_export_data(
    session,
    *,
    period: ExportPeriod,
    row_to_list_dict,
) -> tuple[bytes, str]:
    """Собирает summary + detail и возвращает (xlsx_bytes, filename)."""
    from agent_pochta.db.repository import EmailRepository

    date_from, date_to = resolve_export_period(period)
    repo = EmailRepository(session)

    by_status = repo.count_by_status(date_from=date_from, date_to=date_to)
    total = sum(by_status.values())
    operator_review_counts = repo.count_operator_review_states(
        date_from=date_from,
        date_to=date_to,
    )

    approvals_start = msk_day_start_utc(date_from)
    approvals_end = msk_day_end_exclusive_utc(date_to)
    operator_approvals = collect_operator_approvals(
        session,
        start_utc=approvals_start,
        end_utc=approvals_end,
    )

    erp_created = repo.count_erp_created(date_from=date_from, date_to=date_to)
    erp_skipped = repo.count_erp_skipped(date_from=date_from, date_to=date_to)

    summary_rows = build_export_summary(
        period=period,
        date_from=date_from,
        date_to=date_to,
        by_status=by_status,
        operator_review_counts=operator_review_counts,
        operator_approvals=operator_approvals,
        erp_created=erp_created,
        erp_skipped=erp_skipped,
        total=total,
    )

    db_rows = repo.list_all_messages(date_from=date_from, date_to=date_to)
    event_hints = repo.batch_operator_review_event_hints([row.id for row in db_rows])
    detail_rows = [
        row_dict_to_export_detail(
            row_to_list_dict(row, operator_event_hints=event_hints.get(row.id))
        )
        for row in db_rows
    ]

    content = build_export_xlsx(
        period=period,
        date_from=date_from,
        date_to=date_to,
        summary_rows=summary_rows,
        detail_rows=detail_rows,
    )
    return content, export_filename(period)
