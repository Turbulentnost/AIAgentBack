"""Unit-тесты обеспеченности изделий (α + добор сверху вниз)."""

from __future__ import annotations

from types import SimpleNamespace

from app.agents.document_analysis_agent.product_coverage import compute_product_coverage


def _plan(product: str, months: dict[str, float], *, facts: dict[str, float] | None = None):
    monthly_qty = {}
    for month, qty in months.items():
        fact_qty = float((facts or {}).get(month, 0.0))
        monthly_qty[month] = {
            "заказ": {"план": qty, "факт": fact_qty},
            "опытные": {"план": 0.0, "факт": 0.0},
            "склад": {"план": 0.0, "факт": 0.0},
        }
    return SimpleNamespace(product=product, monthly_qty=monthly_qty)


def _row(
    name: str,
    by_product: dict[str, float],
    stock: float = 0.0,
    receipts: dict | None = None,
    stock_match: str = "exact",
    *,
    kind: str = "required",
    kind_by_product: dict[str, str] | None = None,
):
    row = SimpleNamespace(
        nomenclature=name,
        by_product=by_product,
        stock=stock,
        monthly_receipts=receipts or {},
        stock_match=stock_match,
        coverage_material_kind=kind,
        coverage_material_label="",
        coverage_material_confidence="",
        coverage_material_reason="",
        coverage_material_kinds_by_product=kind_by_product or {},
        coverage_material_labels_by_product={},
        coverage_material_confidences_by_product={},
        coverage_material_reasons_by_product={},
    )
    return row


def test_proportional_split_on_shared_material():
    """Общий KIT: при дефиците оба изделия > 0 и близки к пропорции планов 2:1."""
    plans = [
        _plan("Изделие А", {"Июль": 200}),
        _plan("Изделие Б", {"Июль": 100}),
    ]
    merged = [_row("KIT", {"Изделие А": 1, "Изделие Б": 1}, stock=150)]
    result = compute_product_coverage(plans, merged, ["Июль"])
    a = result.cell("Изделие А", "Июль")
    b = result.cell("Изделие Б", "Июль")
    assert a.covered + b.covered == 150
    assert a.covered > 0 and b.covered > 0
    assert a.covered == 100
    assert b.covered == 50


def test_top_down_remainder_after_floor():
    """После floor остаток 1 шт уходит верхнему изделию."""
    plans = [
        _plan("Верх", {"Июль": 10}),
        _plan("Низ", {"Июль": 10}),
    ]
    merged = [_row("M", {"Верх": 1, "Низ": 1}, stock=11)]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("Верх", "Июль").covered == 6
    assert result.cell("Низ", "Июль").covered == 5


def test_month2_uses_closing_from_covered():
    """Второй месяц видит closing после сборки первого."""
    plans = [_plan("A", {"Июль": 10, "Август": 10})]
    merged = [_row("M", {"A": 1}, stock=15, receipts={"Август": 0})]
    result = compute_product_coverage(plans, merged, ["Июль", "Август"])
    assert result.cell("A", "Июль").covered == 10
    assert result.cell("A", "Август").covered == 5
    assert result.cell("A", "Август").plan == 10


def test_unmatched_spec_yields_zero_coverage():
    plans = [_plan("Без спеки", {"Июль": 50})]
    merged = [_row("Чужой материал", {"Другое": 1}, stock=1000)]
    result = compute_product_coverage(plans, merged, ["Июль"])
    cell = result.cell("Без спеки", "Июль")
    assert cell.plan == 50
    assert cell.covered == 0
    assert result.boms["Без спеки"].matched is False


def test_zero_plan_stays_zero():
    plans = [_plan("A", {"Июль": 0, "Август": 5})]
    merged = [_row("M", {"A": 1}, stock=100)]
    result = compute_product_coverage(plans, merged, ["Июль", "Август"])
    assert result.cell("A", "Июль").covered == 0
    assert result.cell("A", "Август").covered == 5


def test_exclusive_zero_does_not_kill_shared_alpha():
    plans = [
        _plan("A", {"Июль": 200}),
        _plan("B", {"Июль": 100}),
        _plan("C", {"Июль": 50}),
    ]
    merged = [
        _row("KIT", {"A": 1, "B": 1}, stock=150),
        _row("ONLY_C", {"C": 1}, stock=0.0, receipts={"Август": 100}),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("A", "Июль").covered == 100
    assert result.cell("B", "Июль").covered == 50
    assert result.cell("C", "Июль").covered == 0


def test_zero_stock_without_receipts_blocks_full_bom():
    plans = [_plan("A", {"Июль": 10})]
    merged = [
        _row("KIT", {"A": 1}, stock=10),
        _row("WIRE", {"A": 2}, stock=0.0, stock_match="exact"),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("A", "Июль").covered == 0


def test_zero_supply_optional_does_not_block_conditional_coverage():
    plans = [_plan("A", {"Июль": 10})]
    merged = [
        _row("KIT", {"A": 1}, stock=10),
        _row(
            "Винт M3",
            {"A": 2},
            stock=0.0,
            kind="consumable",
            kind_by_product={"A": "consumable"},
        ),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    cell = result.cell("A", "Июль")
    assert cell.covered == 0
    assert cell.conditional_covered == 10


def test_stocked_optional_still_blocks_conditional_coverage():
    plans = [_plan("A", {"Июль": 10})]
    merged = [
        _row("KIT", {"A": 1}, stock=10),
        _row(
            "Винт M3",
            {"A": 20},
            stock=5.0,
            kind="consumable",
            kind_by_product={"A": "consumable"},
        ),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    cell = result.cell("A", "Июль")
    assert cell.covered == 0
    assert cell.conditional_covered == 0


def test_expected_receipts_keep_optional_material_blocking():
    plans = [_plan("A", {"Июль": 10})]
    merged = [
        _row("KIT", {"A": 1}, stock=10),
        _row(
            "Кабель",
            {"A": 1},
            stock=0.0,
            receipts={"Июль": 5},
            kind="workshop",
            kind_by_product={"A": "workshop"},
        ),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    cell = result.cell("A", "Июль")
    assert cell.covered == 5
    assert cell.conditional_covered == 5


def test_full_bom_covered_when_all_materials_available():
    plans = [_plan("A", {"Июль": 10})]
    merged = [
        _row("KIT", {"A": 1}, stock=10),
        _row("WIRE", {"A": 2}, stock=20),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("A", "Июль").covered == 10


def test_pre_horizon_receipts_seed_opening():
    plans = [_plan("A", {"Июль": 10})]
    merged = [_row("M", {"A": 1}, stock=0.0, receipts={"Июнь": 7, "Июль": 0})]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("A", "Июль").covered == 7


def test_bom_detail_plan_fact_and_available():
    plans = [_plan("A", {"Июль": 10}, facts={"Июль": 6})]
    merged = [
        _row("Винт M3", {"A": 2}, stock=5.0, receipts={"Июль": 3}),
        _row("Корпус", {"A": 1}, stock=100.0),
    ]
    result = compute_product_coverage(plans, merged, ["Июль"])
    assert result.cell("A", "Июль").fact == 6
    assert result.material_plan("A", "Июль", "винт m3") == 20.0
    assert result.material_fact("A", "Июль", "винт m3") == 12.0
    assert result.material_available("Июль", "винт m3") == 8.0


def test_coverage_sheet_has_collapsed_outline_rows():
    from openpyxl import Workbook

    from app.agents.document_analysis_agent.excel_service import _write_product_coverage_sheet

    plans = [_plan("Изделие А", {"Июль": 5})]
    merged = [_row("Деталь X", {"Изделие А": 3}, stock=30)]
    result = compute_product_coverage(plans, merged, ["Июль"])
    wb = Workbook()
    ws = wb.active
    _write_product_coverage_sheet(ws, result)
    assert ws["A5"].value == "Изделие А"
    assert ws["A6"].value == "Деталь X"
    assert ws.row_dimensions[6].outline_level == 1
    assert ws.row_dimensions[6].hidden is True
    assert ws["B6"].value == 30
    assert ws["C6"].value == 15
    assert ws["D6"].value == 0
    assert ws["D4"].value == "Факт"
