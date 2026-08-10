"""Тесты детального графика формата «Отчёт» (П/ф · ОТК · Склад)."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook

from app.agents.document_analysis_agent.excel_service import (
    ROLE_DETAILED_PRODUCTION_SCHEDULE,
    DetailedScheduleExtract,
    UploadedWorkbook,
    _classify_preview_locally,
    _enrich_merged_with_daily_demand,
    _expand_priority_sheet_coverage_columns,
    _extract_detailed_production_schedule,
    _infer_detailed_sheet_year_month,
    _match_detailed_plan_for_product,
    _normalize,
    _parse_detailed_schedule_sheet,
    _preview_looks_like_detailed_production_schedule,
    _sheet_is_pf_stage_report,
    DetailedScheduleProductPlan,
    MergedNomenclatureRow,
)


def _pf_report_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws["B3"] = "Модель / изделие"
    ws["D2"] = "01.07.2026-17.07.2026"
    ws["D3"] = "план"
    ws["E3"] = "факт"
    ws["F2"] = datetime(2026, 7, 18)
    ws["F3"] = "план"
    ws["G3"] = "факт"
    ws["H2"] = "Итог недели"
    ws["H3"] = "План недели"
    ws["I3"] = "Факт недели"
    ws["J2"] = "План месяца"
    ws["K2"] = "Нарастающий"
    ws["K3"] = "план"
    ws["L3"] = "факт"
    ws["M3"] = "откл."

    ws["A4"] = 1
    ws["B4"] = "Сокол И"
    ws["C4"] = "П/ф"
    ws["D4"] = 6800
    ws["E4"] = 1995
    ws["G4"] = 500

    ws["C5"] = "ОТК"
    ws["D5"] = 6800
    ws["E5"] = 1995
    ws["G5"] = 9999  # не должно попасть в план/факт

    ws["C6"] = "Склад"
    ws["D6"] = 1

    ws["A7"] = 2
    ws["B7"] = "Сокол ИТ"
    ws["C7"] = "П/ф"
    ws["D7"] = 1700
    ws["F7"] = 100

    ws["A8"] = 5
    ws["C8"] = "П/ф"  # пустой слот без имени

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_pf_report_parser_uses_only_pf_and_distributes_range() -> None:
    from openpyxl import load_workbook

    content = _pf_report_bytes()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    assert _sheet_is_pf_stage_report(ws)
    assert _infer_detailed_sheet_year_month(ws, 2026) == (2026, 7)

    plans, plan_cells = _parse_detailed_schedule_sheet(ws, 2026, 7)
    assert {p.product for p in plans} == {"Сокол И", "Сокол ИТ"}
    assert plan_cells
    assert all(cell.plan_col >= 1 and cell.row >= 1 for cell in plan_cells)

    sok = next(p for p in plans if p.product == "Сокол И")
    plan_1_17 = sum(sok.daily_qty.get(date(2026, 7, d).isoformat(), 0.0) for d in range(1, 18))
    fact_1_17 = sum(sok.daily_fact.get(date(2026, 7, d).isoformat(), 0.0) for d in range(1, 18))
    assert plan_1_17 == 6800.0
    assert fact_1_17 == 1995.0
    assert sok.daily_qty.get("2026-07-01") == 400.0
    assert sok.daily_qty.get("2026-07-18") == 0.0
    assert sok.daily_fact.get("2026-07-18") == 500.0
    # ОТК 9999 не учитывается
    assert sok.daily_fact.get("2026-07-18") != 9999
    assert "2026-07-26" not in sok.daily_fact


def test_pf_report_extract_fills_plan_all_month_fact_only_existing() -> None:
    content = _pf_report_bytes()
    filename = "Отчет 07_2026_6148.xlsx"
    extract = _extract_detailed_production_schedule(
        [UploadedWorkbook(filename=filename, content=content)],
        {filename: ROLE_DETAILED_PRODUCTION_SCHEDULE},
        as_of=date(2026, 7, 24),
    )
    assert extract.year == 2026 and extract.month == 7
    assert len(extract.day_keys) == 31
    sok = next(p for p in extract.plans if p.product == "Сокол И")
    assert set(sok.daily_qty) == set(extract.day_keys)
    assert sok.daily_qty["2026-07-26"] == 0.0
    assert "2026-07-26" not in sok.daily_fact
    assert "2026-07-18" in sok.daily_fact
    assert set(sok.daily_fact).issubset(set(extract.day_keys))


def test_daily_demand_from_pf_plan_and_fact() -> None:
    content = _pf_report_bytes()
    filename = "Отчет 07_2026_6148.xlsx"
    extract = _extract_detailed_production_schedule(
        [UploadedWorkbook(filename=filename, content=content)],
        {filename: ROLE_DETAILED_PRODUCTION_SCHEDULE},
        as_of=date(2026, 7, 24),
    )
    row = MergedNomenclatureRow(
        nomenclature="Комплектующая А",
        products=["Сокол И полный комплект"],
        quantity=2.0,
        by_product={"Сокол И полный комплект": 2.0},
    )
    _enrich_merged_with_daily_demand([row], extract)
    sok = next(p for p in extract.plans if p.product == "Сокол И")
    # 400 изд. × 2 = 800 на 01.07 (план из диапазона)
    assert row.daily_demand["2026-07-01"] == 800.0
    # день без плана выпуска → 0
    assert row.daily_demand["2026-07-26"] == 0.0
    # план на 18.07 пустой → 0
    assert row.daily_demand["2026-07-18"] == 0.0
    # факт 500 на 18.07 × 2 = 1000
    assert row.daily_demand_fact["2026-07-18"] == 1000.0
    assert row.daily_demand_fact["2026-07-01"] == float(sok.daily_fact["2026-07-01"]) * 2
    assert row.daily_demand_fact["2026-07-26"] == 0.0


def test_preview_detects_otchet_filename() -> None:
    preview = {
        "filename": "Отчет 07_2026_6148.xlsx",
        "sheets": [
            {
                "name": "Лист1",
                "headers": ["Модель / изделие", "план", "факт", "П/ф"],
                "sample_rows": [["1", "Сокол И", "П/ф", 100]],
            }
        ],
    }
    assert _preview_looks_like_detailed_production_schedule(preview)
    assert _classify_preview_locally(preview) == ROLE_DETAILED_PRODUCTION_SCHEDULE


def test_plan_fact_report_without_pf_stage_is_used_for_daily_plan() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Август"
    ws["B3"] = "Модель / изделие"
    ws["D2"] = datetime(2026, 8, 3)
    ws["D3"] = "план"
    ws["E3"] = "факт"
    ws["F2"] = datetime(2026, 8, 4)
    ws["F3"] = "план"
    ws["G3"] = "факт"

    ws["A4"] = 1
    ws["B4"] = "Сокол И"
    ws["C4"] = "Выпуск"
    ws["D4"] = 120
    ws["E4"] = 100
    ws["F4"] = 80
    ws["G4"] = 70

    ws["C5"] = "ОТК"
    ws["D5"] = 999
    ws["E5"] = 999

    assert _sheet_is_pf_stage_report(ws)
    plans, _plan_cells = _parse_detailed_schedule_sheet(ws, 2026, 8)
    assert len(plans) == 1
    assert plans[0].product == "Сокол И"
    assert plans[0].daily_qty == {"2026-08-03": 120.0, "2026-08-04": 80.0}
    assert plans[0].daily_fact == {"2026-08-03": 100.0, "2026-08-04": 70.0}


def test_daily_demand_matches_full_product_to_short_detailed_name() -> None:
    detailed = [
        DetailedScheduleProductPlan(product="Сокол И", daily_qty={"2026-08-03": 120.0}),
        DetailedScheduleProductPlan(product="Сокол Т", daily_qty={"2026-08-03": 80.0}),
        DetailedScheduleProductPlan(product="НСУ 1.0", daily_qty={"2026-08-03": 5.0}),
    ]
    plans_by_key = {_normalize(plan.product): plan for plan in detailed}
    plan_names = [plan.product for plan in detailed]

    assert (
        _match_detailed_plan_for_product(
            'FPV-перехватчик "СОКОЛ" И (день)', plans_by_key, plan_names
        ).product
        == "Сокол И"
    )
    assert (
        _match_detailed_plan_for_product(
            'FPV-перехватчик "СОКОЛ" Т (ночь)', plans_by_key, plan_names
        ).product
        == "Сокол Т"
    )
    assert (
        _match_detailed_plan_for_product("НСУ 1", plans_by_key, plan_names).product
        == "НСУ 1.0"
    )


def test_generic_release_schedule_priority_sheet_expands_day_columns() -> None:
    from types import SimpleNamespace

    wb = Workbook()
    ws = wb.active
    ws["A8"] = "№ п/п"
    ws["B8"] = "Наименование"
    ws["C8"] = "Остаток"
    ws["D8"] = 8
    ws["E8"] = 9
    ws["F8"] = "Итог"
    ws["G8"] = 10
    ws.merge_cells("A9:G9")
    ws["A9"] = "Производство № 1"
    ws["A10"] = 1
    ws["B10"] = "Сокол И"
    ws["D10"] = 120
    ws["E10"] = 80
    ws["G10"] = 60

    detailed = DetailedScheduleExtract(
        files=[],
        plans=[],
        year=2026,
        month=8,
        day_keys=["2026-08-08", "2026-08-09", "2026-08-10"],
    )

    class FakeDailyCoverage:
        products_in_order = ["Сокол И"]

        def cell(self, product: str, day: str):
            values = {
                ("Сокол И", "2026-08-08"): (100.0, 120.0, 90.0),
                ("Сокол И", "2026-08-09"): (0.0, 80.0, 70.0),
                ("Сокол И", "2026-08-10"): (50.0, 60.0, 40.0),
            }
            covered, plan, fact = values.get((product, day), (0.0, 0.0, 0.0))
            return SimpleNamespace(covered=covered, plan=plan, fact=fact)

        def status_for_plan_cell(self, product: str, day_keys: list[str], plan_qty: float):
            return "green"

    stats = _expand_priority_sheet_coverage_columns(
        ws,
        detailed=detailed,
        daily_plan_coverage=FakeDailyCoverage(),
    )

    assert stats["day_slots"] == 3
    assert [ws.cell(9, col).value for col in range(4, 10)] == [
        "обесп",
        "план",
        "факт",
        "обесп",
        "план",
        "факт",
    ]
    assert [ws.cell(9, col).value for col in range(11, 14)] == [
        "обесп",
        "план",
        "факт",
    ]
    assert ws["J8"].value == "Итог"
    assert ws["A10"].value == "Производство № 1"
    assert [ws.cell(11, col).value for col in range(4, 10)] == [
        100,
        120,
        90,
        0,
        80,
        70,
    ]
    assert [ws.cell(11, col).value for col in range(11, 14)] == [50, 60, 40]
    merge_refs = {str(item) for item in ws.merged_cells.ranges}
    assert {"A8:A9", "B8:B9", "C8:C9", "D8:F8", "G8:I8", "K8:M8"}.issubset(
        merge_refs
    )
    assert ws["A8"].fill.fill_type == "solid"
    assert ws["D8"].fill.fill_type == "solid"
    assert ws["D9"].fill.fill_type == "solid"
    assert ws["E11"].border.left.style == "thin"
    assert ws["F11"].border.left.style == "thin"
    assert ws["D11"].border.left.style == "medium"
    assert ws["F11"].border.right.style == "medium"
    assert ws["G11"].border.left.style == "medium"
    assert ws["I11"].border.right.style == "medium"
    assert ws["K11"].border.left.style == "medium"
    assert ws["M11"].border.right.style == "medium"
