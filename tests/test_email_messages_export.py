"""Тесты Excel-выгрузки «Таняфикация»."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from agent_pochta.api.email_messages_export import (
    EXPORT_DETAIL_COLUMNS,
    build_export_summary,
    build_export_xlsx,
    export_filename,
    resolve_export_period,
    row_dict_to_export_detail,
)

_MSK = ZoneInfo("Europe/Moscow")


def test_resolve_export_period_day():
    now = datetime(2026, 7, 22, 15, 0, tzinfo=_MSK)
    date_from, date_to = resolve_export_period("day", now=now)
    assert date_from == date_to == now.date()


def test_resolve_export_period_week():
    now = datetime(2026, 7, 22, 15, 0, tzinfo=_MSK)
    date_from, date_to = resolve_export_period("week", now=now)
    assert date_to == now.date()
    assert (date_to - date_from).days == 6


def test_resolve_export_period_month():
    now = datetime(2026, 7, 22, 15, 0, tzinfo=_MSK)
    date_from, date_to = resolve_export_period("month", now=now)
    assert date_to == now.date()
    assert (date_to - date_from).days == 29


def test_export_detail_columns_exclude_access_and_responsible():
    keys = {key for key, _ in EXPORT_DETAIL_COLUMNS}
    assert "access_label" not in keys
    assert "responsible_label" not in keys
    headers = [header for _, header in EXPORT_DETAIL_COLUMNS]
    assert "Гриф" not in headers
    assert "Ответственный" not in headers


def test_row_dict_to_export_detail_maps_table_fields():
    detail = row_dict_to_export_detail(
        {
            "operator_review_state": "verified",
            "status": "done",
            "is_spam": False,
            "received_at": "2026-07-22T07:37:41+00:00",
            "erp_document_number": "КБ00-000028",
            "attachments_summary": [{"index": 0, "filename": "scan.pdf"}],
            "organization_name": "ООО НПО «Турбулентность-ДОН»",
            "sender_email": "client@example.ru",
            "partner_name": "ООО «Газпром»",
            "department_id": "00-000076",
            "department_name": "Отдел договорной работы",
            "payer_direction_label": "ООО НПО «Турбулентность-ДОН» пр-во1",
        }
    )
    assert detail["operator_review_label"] == "Проверено"
    assert detail["spam_label"] == "Не спам"
    assert detail["erp_document_number"] == "КБ00-000028"
    assert detail["attachments_display"] == "scan.pdf"
    assert detail["organization_name"] == "ООО НПО «Турбулентность-ДОН»"


def test_build_export_summary_includes_review_counts():
    date_from, date_to = resolve_export_period("day", now=datetime(2026, 7, 22, tzinfo=_MSK))
    rows = build_export_summary(
        period="day",
        date_from=date_from,
        date_to=date_to,
        by_status={"done": 8, "spam": 2, "processing": 1},
        operator_review_counts={"verified": 5, "corrected": 2, "pending": 4, "all": 11},
        operator_approvals={"saved": 3, "changed": 2, "rate": 0.6},
        erp_created=7,
        erp_skipped=2,
        total=11,
    )
    labels = dict(rows)
    assert labels["Всего писем"] == 11
    assert labels["Проверок (verified + corrected)"] == 7
    assert labels["Проверено (без правок)"] == 5
    assert labels["Исправлено оператором"] == 2
    assert labels["Не проверено"] == 4
    assert labels["ERP создано"] == 7
    assert labels["ERP пропущено"] == 2
    assert labels["Принято без изменений (saved)"] == 3


def test_build_export_xlsx_has_two_sheets():
    date_from, date_to = resolve_export_period("day", now=datetime(2026, 7, 22, tzinfo=_MSK))
    content = build_export_xlsx(
        period="day",
        date_from=date_from,
        date_to=date_to,
        summary_rows=[("Всего писем", 1)],
        detail_rows=[
            row_dict_to_export_detail(
                {
                    "operator_review_state": "pending",
                    "status": "done",
                    "is_spam": False,
                    "mail_date": "2026-07-22T10:00:00",
                    "sender_email": "a@b.ru",
                }
            )
        ],
    )
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == ["Сводка", "Письма"]
    assert wb["Сводка"]["B5"].value == 1
    assert wb["Письма"]["A2"].value == "Не проверено"


def test_export_filename():
    now = datetime(2026, 7, 22, tzinfo=_MSK)
    assert export_filename("day", now=now) == "tanyafication_report_day_2026-07-22.xlsx"
    assert export_filename("week", now=now) == (
        "tanyafication_report_week_2026-07-16_2026-07-22.xlsx"
    )


def test_export_api_endpoint_returns_xlsx():
    from fastapi.testclient import TestClient

    from agent_pochta.api.app import app

    client = TestClient(app)
    response = client.get("/api/v1/email-messages/export", params={"period": "day"})
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "content-disposition" in response.headers
    assert response.content[:2] == b"PK"

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["Сводка", "Письма"]
