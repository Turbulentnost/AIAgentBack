"""Unit-тесты плана заказов (lead + скользящий дефицит)."""

from __future__ import annotations

from datetime import date
from types import SimpleNamespace

from app.agents.document_analysis_agent.order_plan import (
    compute_order_plan,
    first_day_of_month,
    lead_days_for,
    plan_demand_for_month,
)


def _row(
    name: str,
    stock: float = 0.0,
    demand: dict | None = None,
    receipts: dict | None = None,
):
    monthly_demand = {}
    for month, qty in (demand or {}).items():
        monthly_demand[month] = {
            "заказ": {"план": qty, "факт": 0.0},
            "опытные": {"план": 0.0, "факт": 0.0},
            "склад": {"план": 0.0, "факт": 0.0},
        }
    return SimpleNamespace(
        nomenclature=name,
        stock=stock,
        monthly_demand=monthly_demand,
        monthly_receipts=receipts or {},
    )


def test_lead_default_21():
    assert lead_days_for("unknown", None) == 21
    assert lead_days_for("unknown", {}) == 21


def test_lead_from_logistics_index():
    # 14 + 2 + 10 = 26
    assert lead_days_for("kit", {"kit": (14, 10)}) == 26


def test_first_day_and_order_date():
    assert first_day_of_month(2026, "Август") == date(2026, 8, 1)
    result = compute_order_plan(
        [_row("A", stock=100, demand={"Август": 0})],
        ["Август"],
        2026,
        {"a": (14, 10)},
    )
    cell = result.cell("A", "Август")
    assert cell is not None
    assert cell.order_date == date(2026, 8, 1) - __import__("datetime").timedelta(days=26)


def test_rolling_deficit_covers_shortage():
    """stock 10, Jul Jul 50 → order 40; Aug demand 30, receipt 5 → opening 0+5, order 25."""
    row = _row(
        "M",
        stock=10,
        demand={"Июль": 50, "Август": 30},
        receipts={"Август": 5},
    )
    result = compute_order_plan([row], ["Июль", "Август"], 2026, None)
    july = result.cell("M", "Июль")
    aug = result.cell("M", "Август")
    assert july is not None and aug is not None
    assert july.qty == 40
    assert july.order_date == date(2026, 7, 1) - __import__("datetime").timedelta(days=21)
    # after July: opening = 10+40-50 = 0; Aug available = 0+5 = 5; order = 25
    assert aug.qty == 25


def test_no_order_when_stock_covers():
    row = _row("M", stock=100, demand={"Июль": 40})
    result = compute_order_plan([row], ["Июль"], 2026, None)
    assert result.cell("M", "Июль").qty == 0


def test_plan_demand_sums_categories():
    row = SimpleNamespace(
        monthly_demand={
            "Июль": {
                "заказ": {"план": 10, "факт": 1},
                "опытные": {"план": 2, "факт": 0},
                "склад": {"план": 5, "факт": 0},
            }
        }
    )
    assert plan_demand_for_month(row, "Июль") == 17


def test_order_plan_sheet_writes_dates():
    from openpyxl import Workbook

    from app.agents.document_analysis_agent.excel_service import _write_order_plan_sheet

    row = _row("Деталь", stock=0, demand={"Июль": 10})
    plan = compute_order_plan([row], ["Июль"], 2026, {"деталь": (14, 10)})
    wb = Workbook()
    ws = wb.active
    _write_order_plan_sheet(ws, plan)
    assert ws.title == "план заказов"
    assert ws["A5"].value == "Деталь"
    assert ws["B5"].value == date(2026, 7, 1) - __import__("datetime").timedelta(days=26)
    assert ws["C5"].value == 10


def test_order_plan_qty_linked_to_monthly_receipt():
    """Количество на плане заказов — формула на поступление помесячного листа."""
    from openpyxl import Workbook

    from app.agents.document_analysis_agent.excel_service import (
        _SHEET_MONTHLY_ASSURANCE,
        _order_plan_qty_formula,
        _write_order_plan_sheet,
    )

    layout = {
        "Июль": {"sum_plan": 10, "receipt": 12},
        "Август": {"sum_plan": 20, "receipt": 22},
    }
    formula_july = _order_plan_qty_formula(
        monthly_sheet=_SHEET_MONTHLY_ASSURANCE,
        monthly_row=6,
        order_row=5,
        months=["Июль", "Август"],
        month_index=0,
        monthly_layout=layout,
        qty_cols_by_month=[3, 5],
    )
    assert formula_july is not None
    assert "MAX(0," in formula_july
    assert f"'{_SHEET_MONTHLY_ASSURANCE}'!J6" in formula_july  # sum_plan col 10
    assert f"'{_SHEET_MONTHLY_ASSURANCE}'!L6" in formula_july  # receipt col 12
    assert f"'{_SHEET_MONTHLY_ASSURANCE}'!F6" in formula_july

    formula_aug = _order_plan_qty_formula(
        monthly_sheet=_SHEET_MONTHLY_ASSURANCE,
        monthly_row=6,
        order_row=5,
        months=["Июль", "Август"],
        month_index=1,
        monthly_layout=layout,
        qty_cols_by_month=[3, 5],
    )
    assert formula_aug is not None
    assert "C5" in formula_aug  # previous order qty on this sheet

    row = _row("Деталь", stock=0, demand={"Июль": 10})
    plan = compute_order_plan([row], ["Июль"], 2026, None)
    wb = Workbook()
    ws = wb.active
    _write_order_plan_sheet(
        ws,
        plan,
        monthly_sheet_title=_SHEET_MONTHLY_ASSURANCE,
        monthly_layout={"Июль": {"sum_plan": 10, "receipt": 12}},
        monthly_data_start_row=6,
    )
    assert isinstance(ws["C5"].value, str) and ws["C5"].value.startswith("=MAX(0,")


def test_monthly_receipt_total_is_editable_value_not_week_sum():
    """Итог поступления — число, чтобы можно было вставить qty из плана заказов."""
    from openpyxl import Workbook

    from app.agents.document_analysis_agent.excel_service import (
        MergedNomenclatureRow,
        _write_monthly_assurance_sheet,
    )

    row = MergedNomenclatureRow(
        nomenclature="Деталь X",
        products=["A"],
        quantity=1.0,
        stock=0.0,
        monthly_demand={
            "Июль": {
                "заказ": {"план": 100.0, "факт": 0.0},
                "опытные": {"план": 0.0, "факт": 0.0},
                "склад": {"план": 0.0, "факт": 0.0},
            }
        },
        monthly_receipts={"Июль": 40.0},
        weekly_receipts={"Июль": {"2026-07-01_2026-07-07": 10.0, "2026-07-08_2026-07-14": 30.0}},
    )
    wb = Workbook()
    ws = wb.active
    layout = _write_monthly_assurance_sheet(ws, [row])
    receipt_col = layout["Июль"]["receipt"]
    cell = ws.cell(6, receipt_col)
    assert isinstance(cell.value, (int, float))
    assert float(cell.value) == 40.0
    assert not (isinstance(cell.value, str) and str(cell.value).startswith("="))
