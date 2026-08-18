"""Сборщик пакета Cursor: upload + роли → input/ и manifest.json."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from app.agents.document_analysis_agent.cursor_cloud.job_pack import (
    STOCK_1C_FILENAME,
    load_job_manifest,
    pack_aveon_cursor_job,
)
from app.agents.document_analysis_agent.excel_service import (
    ROLE_PRODUCTION_SCHEDULE,
    ROLE_STOCK,
    UploadedWorkbook,
)


def _xlsx_bytes(headers: list[str], rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows or []:
        sheet.append(row)
    from io import BytesIO

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.mark.asyncio
async def test_pack_uploads_writes_manifest(tmp_path: Path) -> None:
    pack = await pack_aveon_cursor_job(
        [
            UploadedWorkbook(filename="план.xlsx", content=_xlsx_bytes(["Наименования изделий", "Август"])),
        ],
        role_map={"план.xlsx": ROLE_PRODUCTION_SCHEDULE},
        jobs_dir=tmp_path,
        include_db_exports=False,
        as_of="2026-08-18",
        job_id="job-test-1",
    )

    assert pack.job_id == "job-test-1"
    assert (pack.job_dir / "input" / "план.xlsx").is_file()
    assert (pack.job_dir / "output").is_dir()
    assert (pack.job_dir / "manifest.json").is_file()
    assert (pack.job_dir / "analysis_result.schema.json").is_file()

    manifest = load_job_manifest("job-test-1", tmp_path)
    assert manifest is not None
    assert manifest["as_of"] == "2026-08-18"
    assert manifest["files"][0]["role"] == ROLE_PRODUCTION_SCHEDULE
    assert manifest["files"][0]["source"] == "upload"
    assert manifest["output_path"] == "output/analysis_result.json"


@pytest.mark.asyncio
async def test_pack_exports_1c_stock(tmp_path: Path) -> None:
    async def _fake_export(_db):
        return (
            [UploadedWorkbook(filename=STOCK_1C_FILENAME, content=_xlsx_bytes(["Номенклатура", "Остаток"], [["Болт", 5]]))],
            [ROLE_STOCK],
            [],
        )

    import app.agents.document_analysis_agent.cursor_cloud.job_pack as job_pack

    original = job_pack._export_db_workbooks
    job_pack._export_db_workbooks = _fake_export
    try:
        pack = await pack_aveon_cursor_job(
            [UploadedWorkbook(filename="план.xlsx", content=_xlsx_bytes(["A"]))],
            role_map={"план.xlsx": ROLE_PRODUCTION_SCHEDULE},
            db=SimpleNamespace(),
            jobs_dir=tmp_path,
            include_db_exports=True,
            job_id="job-1c",
        )
    finally:
        job_pack._export_db_workbooks = original

    names = {item.filename: item for item in pack.files}
    assert STOCK_1C_FILENAME in names
    assert names[STOCK_1C_FILENAME].source == "1c"
    assert names[STOCK_1C_FILENAME].role == ROLE_STOCK
    sheet = load_workbook(pack.job_dir / "input" / STOCK_1C_FILENAME).active
    assert sheet["A2"].value == "Болт"


@pytest.mark.asyncio
async def test_pack_empty_rejected(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Нет файлов"):
        await pack_aveon_cursor_job([], jobs_dir=tmp_path, include_db_exports=False)
