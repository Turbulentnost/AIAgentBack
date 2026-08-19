from __future__ import annotations

from io import BytesIO
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from app.services.material_calculator import (
    MaterialCalculatorInputItem,
    MaterialCalculatorLine,
    MaterialCalculatorResult,
    build_material_calculator_xlsx,
    calculate_material_requirements,
    material_calculator_lines_to_result,
)


@pytest.mark.asyncio
async def test_calculate_material_requirements_aggregates_and_scales():
    spec_a = {
        "ref_key": "spec-a",
        "code": "001",
        "description": "Спека A",
        "main_product": {"name": "Изделие A", "qty": 2.0},
        "materials": [
            {
                "nomenclature_key": "nom-1",
                "code": "M1",
                "name": "Болт M6",
                "qty": 4.0,
                "unit": "шт",
                "produced_in_process": False,
            }
        ],
    }
    spec_b = {
        "ref_key": "spec-b",
        "code": "002",
        "description": "Спека B",
        "main_product": {"name": "Изделие B", "qty": 1.0},
        "materials": [
            {
                "nomenclature_key": "nom-1",
                "code": "M1",
                "name": "Болт M6",
                "qty": 2.0,
                "unit": "шт",
                "produced_in_process": False,
            }
        ],
    }

    async def fake_get_spec(_db, ref_key: str):
        return {"spec-a": spec_a, "spec-b": spec_b}.get(ref_key)

    with patch(
        "app.services.material_calculator.get_resource_spec_from_db",
        new=AsyncMock(side_effect=fake_get_spec),
    ):
        result = await calculate_material_requirements(
            AsyncMock(),
            [
                MaterialCalculatorInputItem(spec_ref_key="spec-a", quantity=4),
                MaterialCalculatorInputItem(spec_ref_key="spec-b", quantity=3),
            ],
        )

    assert result.ok is True
    assert len(result.lines) == 1
    line = result.lines[0]
    assert line.name == "Болт M6"
    assert line.unit == "шт"
    # spec-a: 4 изделия / 2 на партию * 4 болта = 8
    # spec-b: 3 изделия / 1 * 2 болта = 6
    assert line.total_qty == 14
    assert len(line.breakdown) == 2


@pytest.mark.asyncio
async def test_calculate_material_requirements_merges_by_code_without_nom_key():
    spec_a = {
        "ref_key": "spec-a",
        "code": "001",
        "description": "Спека A",
        "main_product": {"name": "Изделие A", "qty": 1.0},
        "materials": [
            {
                "nomenclature_key": "nom-1",
                "code": "M1",
                "name": "Болт M6",
                "qty": 2.0,
                "unit": "шт",
                "produced_in_process": False,
            }
        ],
    }
    spec_b = {
        "ref_key": "spec-b",
        "code": "002",
        "description": "Спека B",
        "main_product": {"name": "Изделие B", "qty": 1.0},
        "materials": [
            {
                "nomenclature_key": "",
                "code": "M1",
                "name": "Болт M6",
                "qty": 3.0,
                "unit": "шт",
                "produced_in_process": False,
            }
        ],
    }

    async def fake_get_spec(_db, ref_key: str):
        return {"spec-a": spec_a, "spec-b": spec_b}.get(ref_key)

    with patch(
        "app.services.material_calculator.get_resource_spec_from_db",
        new=AsyncMock(side_effect=fake_get_spec),
    ):
        result = await calculate_material_requirements(
            AsyncMock(),
            [
                MaterialCalculatorInputItem(spec_ref_key="spec-a", quantity=2),
                MaterialCalculatorInputItem(spec_ref_key="spec-b", quantity=4),
            ],
        )

    assert len(result.lines) == 1
    assert result.lines[0].total_qty == 16
    assert len(result.lines[0].breakdown) == 2


def test_build_material_calculator_xlsx_matches_lines():
    result = MaterialCalculatorResult(
        ok=True,
        warnings=[],
        lines=[
            MaterialCalculatorLine(
                nomenclature_key="nom-1",
                code="M1",
                name="Болт M6",
                unit="шт",
                total_qty=14,
            ),
            MaterialCalculatorLine(
                nomenclature_key="nom-2",
                code="M2",
                name="Гайка M6",
                unit="шт",
                total_qty=7.5,
            ),
        ],
    )

    workbook = load_workbook(BytesIO(build_material_calculator_xlsx(result)))
    sheet = workbook.active
    assert sheet.title == "Потребность"
    assert [sheet.cell(1, col).value for col in range(1, 5)] == [
        "Код",
        "Номенклатура",
        "Количество",
        "Ед. изм.",
    ]
    assert sheet.cell(2, 1).value == "M1"
    assert sheet.cell(2, 2).value == "Болт M6"
    assert sheet.cell(2, 3).value == 14
    assert sheet.cell(2, 4).value == "шт"
    assert sheet.cell(3, 1).value == "M2"
    assert sheet.cell(3, 3).value == 7.5


def test_material_calculator_lines_to_result_sorts_and_skips_empty_names():
    result = material_calculator_lines_to_result(
        [
            {"code": "B", "name": "Болт", "unit": "шт", "total_qty": 3},
            {"code": "", "name": "   ", "unit": "шт", "total_qty": 1},
            {"code": "A", "name": "Анкер", "unit": "шт", "total_qty": 2},
        ]
    )
    assert [line.name for line in result.lines] == ["Анкер", "Болт"]
    assert [line.total_qty for line in result.lines] == [2, 3]
