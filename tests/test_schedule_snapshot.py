from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from app.agents.document_analysis_agent.excel_service import (
    ROLE_PRODUCTION_SCHEDULE,
    UploadedWorkbook,
    _detailed_diff_message,
    _schedule_diff_message,
)
from app.agents.document_analysis_agent.production_schedule_diff import (
    compare_production_schedule_with_snapshot,
)
from app.agents.document_analysis_agent.schedule_snapshot import (
    detailed_month_key,
    get_saved_detailed_file,
    get_saved_production_file,
    list_saved_detailed_schedules,
    load_schedule_snapshot,
    save_schedule_snapshot,
    schedule_snapshot_status,
)


def _minimal_workbook_bytes(label: str = "v1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = f"Версия {label}"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture
def snapshot_user_id() -> str:
    return "test-schedule-snapshot-user"


@pytest.fixture(autouse=True)
def cleanup_snapshot(snapshot_user_id: str, tmp_path, monkeypatch):
    monkeypatch.setattr(
        "app.agents.document_analysis_agent.schedule_snapshot._SNAPSHOT_ROOT",
        tmp_path,
    )
    yield


def test_first_analysis_saves_baseline_without_comparison(snapshot_user_id: str):
    content = _minimal_workbook_bytes("1")
    workbooks = [UploadedWorkbook(filename="graph.xlsx", content=content)]
    role_map = {"graph.xlsx": ROLE_PRODUCTION_SCHEDULE}

    workbooks, role_map, diff, compared = compare_production_schedule_with_snapshot(
        workbooks,
        role_map,
        saved_file=None,
    )

    assert compared is False
    assert diff is None
    assert len(workbooks) == 1

    save_schedule_snapshot(
        snapshot_user_id,
        production=("graph.xlsx", content),
    )

    saved = get_saved_production_file(snapshot_user_id)
    assert saved is not None
    assert saved[0] == "graph.xlsx"
    assert saved[1] == content
    assert load_schedule_snapshot(snapshot_user_id) is not None


def test_second_analysis_compares_with_saved_snapshot(snapshot_user_id: str):
    old_content = _minimal_workbook_bytes("1")
    new_content = _minimal_workbook_bytes("2")
    save_schedule_snapshot(snapshot_user_id, production=("old.xlsx", old_content))

    workbooks = [UploadedWorkbook(filename="new.xlsx", content=new_content)]
    role_map = {"new.xlsx": ROLE_PRODUCTION_SCHEDULE}
    saved = get_saved_production_file(snapshot_user_id)

    _workbooks, _role_map, diff, compared = compare_production_schedule_with_snapshot(
        workbooks,
        role_map,
        saved_file=saved,
    )

    assert compared is True
    assert diff is not None
    assert diff.old_filename == "old.xlsx"
    assert diff.new_filename == "new.xlsx"


def test_compare_keeps_baseline_as_old_when_its_version_is_higher(snapshot_user_id: str):
    baseline_content = _minimal_workbook_bytes("2")
    upload_content = _minimal_workbook_bytes("1")
    save_schedule_snapshot(snapshot_user_id, production=("baseline.xlsx", baseline_content))

    workbooks = [UploadedWorkbook(filename="upload.xlsx", content=upload_content)]
    role_map = {"upload.xlsx": ROLE_PRODUCTION_SCHEDULE}
    saved = get_saved_production_file(snapshot_user_id)

    _workbooks, _role_map, diff, compared = compare_production_schedule_with_snapshot(
        workbooks,
        role_map,
        saved_file=saved,
    )

    assert compared is True
    assert diff is not None
    assert diff.old_filename == "baseline.xlsx"
    assert diff.new_filename == "upload.xlsx"
    assert diff.old_version_label == "2"
    assert diff.new_version_label == "1"


def test_detailed_snapshots_are_stored_per_month(snapshot_user_id: str):
    aug = _minimal_workbook_bytes("aug")
    jul = _minimal_workbook_bytes("jul")

    save_schedule_snapshot(snapshot_user_id, detailed=(2025, 8, "aug.xlsx", aug))
    assert get_saved_detailed_file(snapshot_user_id, 2025, 8) == ("aug.xlsx", aug)
    assert get_saved_detailed_file(snapshot_user_id, 2025, 7) is None

    save_schedule_snapshot(snapshot_user_id, detailed=(2025, 7, "jul.xlsx", jul))
    assert get_saved_detailed_file(snapshot_user_id, 2025, 7) == ("jul.xlsx", jul)
    assert get_saved_detailed_file(snapshot_user_id, 2025, 8) == ("aug.xlsx", aug)

    months = [item["month"] for item in list_saved_detailed_schedules(snapshot_user_id)]
    assert months == ["2025-07", "2025-08"]


def test_schedule_diff_messages_for_baseline_and_saved_compare():
    assert "базовая версия" in _schedule_diff_message(
        None,
        compared_with_saved=False,
        baseline_saved=True,
    ).lower()
    assert "базовая версия" in _detailed_diff_message(
        None,
        compared_with_saved=False,
        baseline_saved=True,
        baseline_month=detailed_month_key(2025, 7),
    ).lower()
    assert "2025-07" in _detailed_diff_message(
        None,
        compared_with_saved=False,
        baseline_saved=True,
        baseline_month=detailed_month_key(2025, 7),
    )


def test_schedule_snapshot_status(snapshot_user_id: str):
    assert schedule_snapshot_status(snapshot_user_id)["has_production"] is False
    save_schedule_snapshot(
        snapshot_user_id,
        production=("graph.xlsx", _minimal_workbook_bytes("3")),
    )
    status = schedule_snapshot_status(snapshot_user_id)
    assert status["has_production"] is True
    assert status["production_filename"] == "graph.xlsx"
    assert status["detailed_schedules"] == []
