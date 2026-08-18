from io import BytesIO

from openpyxl import load_workbook

from app.agents.document_analysis_agent.excel_service import (
    MergedNomenclatureRow,
    _build_result_xlsx,
    _enrich_snapshot_rows_units,
    _write_daily_assurance_sheet_from_snapshot,
)
from app.agents.document_analysis_agent.onec_db_sources import DbNomenclatureUnitEntry


def test_enrich_snapshot_rows_units_from_merged() -> None:
    merged = [
        MergedNomenclatureRow(
            nomenclature="Болт М6",
            products=["Изделие 1"],
            quantity=1.0,
            unit="шт",
        ),
    ]
    rows_data = [
        {"nomenclature": "Болт М6", "products": ["Изделие 1"], "unit": None},
    ]
    filled = _enrich_snapshot_rows_units(rows_data, merged)
    assert filled == 1
    assert rows_data[0]["unit"] == "шт"


def test_enrich_snapshot_rows_units_from_db_when_merged_missing() -> None:
    db_index = {
        "гайка м6": DbNomenclatureUnitEntry(nomenclature="Гайка М6", unit="компл"),
    }
    rows_data = [{"nomenclature": "Гайка М6", "products": ["X"], "unit": ""}]
    filled = _enrich_snapshot_rows_units(rows_data, [], db_index)
    assert filled == 1
    assert rows_data[0]["unit"] == "компл"


def test_daily_snapshot_sheet_writes_unit_column() -> None:
    from openpyxl import Workbook

    snapshot = {
        "year": 2026,
        "month": 7,
        "period_key": "2026-07",
        "day_keys": ["2026-07-01"],
        "rows": [
            {
                "nomenclature": "Болт М6",
                "products": ["Изделие 1"],
                "supplier": None,
                "country_of_origin": None,
                "unit": None,
                "price": 10.0,
                "stock": 5.0,
                "ordered": 0.0,
                "daily_demand": {"2026-07-01": 1.0},
                "daily_demand_fact": {"2026-07-01": 0.0},
                "daily_receipts": {"2026-07-01": 0.0},
                "daily_forecast": {"2026-07-01": 4.0},
            },
        ],
    }
    merged = [
        MergedNomenclatureRow(
            nomenclature="Болт М6",
            products=["Изделие 1"],
            quantity=1.0,
            unit="шт",
        ),
    ]
    workbook = Workbook()
    worksheet = workbook.active
    _write_daily_assurance_sheet_from_snapshot(
        worksheet,
        snapshot,
        merged_rows=merged,
    )
    assert worksheet["E5"].value == "шт"


def test_build_result_xlsx_enriches_historical_daily_units() -> None:
    from unittest.mock import patch

    snapshot = {
        "year": 2026,
        "month": 6,
        "period_key": "2026-06",
        "day_keys": ["2026-06-01"],
        "rows": [
            {
                "nomenclature": "Болт М6",
                "products": ["Изделие 1"],
                "supplier": None,
                "country_of_origin": None,
                "unit": None,
                "price": 10.0,
                "stock": 5.0,
                "ordered": 0.0,
                "daily_demand": {"2026-06-01": 1.0},
                "daily_demand_fact": {"2026-06-01": 0.0},
                "daily_receipts": {"2026-06-01": 0.0},
                "daily_forecast": {"2026-06-01": 4.0},
            },
        ],
    }
    merged = [
        MergedNomenclatureRow(
            nomenclature="Болт М6",
            products=["Изделие 1"],
            quantity=1.0,
            unit="шт",
        ),
    ]
    from app.agents.document_analysis_agent.excel_service import DetailedScheduleExtract

    detailed = DetailedScheduleExtract(
        files=[],
        plans=[],
        year=2026,
        month=7,
        day_keys=["2026-07-01"],
    )
    with patch(
        "app.agents.document_analysis_agent.daily_plan_snapshot.list_daily_plan_snapshots",
        return_value=[snapshot],
    ):
        result_bytes = _build_result_xlsx(merged, detailed)

    workbook = load_workbook(BytesIO(result_bytes))
    june_sheet = next(
        name
        for name in workbook.sheetnames
        if name.startswith("2-произв. план (") and "Июн" in name
    )
    assert workbook[june_sheet]["E5"].value == "шт"
