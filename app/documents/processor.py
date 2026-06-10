from __future__ import annotations
import io
import re

def extract_text(data: bytes, mime_type: str) -> str:
    if "pdf" in mime_type:
        import fitz
        with fitz.open(stream=data, filetype="pdf") as doc:
            return "\n".join(page.get_text() for page in doc)
    if "word" in mime_type or "docx" in mime_type:
        from docx import Document
        document = Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs)
    if "sheet" in mime_type or "xlsx" in mime_type or "excel" in mime_type:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return "\n".join("\t".join("" if c is None else str(c) for c in row) for ws in wb.worksheets for row in ws.iter_rows(values_only=True))
    return data.decode("utf-8", errors="ignore")


# Граница предложения: конец на .!?…, за которым идёт пробел и начало нового
# предложения (заглавная буква, цифра или кавычка/тире списка).
_SENTENCE_SPLIT = re.compile(r"(?<=[.!?…])\s+(?=[\"«(\[]?[A-ZА-ЯЁ0-9])")


def _split_sentences(text: str) -> list[str]:
    """Разбивает текст на предложения, сохраняя их целиком.

    Сначала режем по абзацам (переводы строк), затем внутри абзаца — по
    границам предложений. Слова и предложения не разрываются.
    """
    sentences: list[str] = []
    for paragraph in re.split(r"\n+", text or ""):
        paragraph = re.sub(r"[ \t]+", " ", paragraph).strip()
        if not paragraph:
            continue
        for part in _SENTENCE_SPLIT.split(paragraph):
            part = part.strip()
            if part:
                sentences.append(part)
    return sentences


def _split_long_sentence(sentence: str, chunk_size: int) -> list[str]:
    """Делит слишком длинное предложение по границам слов (без разрыва слов)."""
    pieces: list[str] = []
    current: list[str] = []
    current_len = 0
    for word in sentence.split():
        addition = len(word) + (1 if current else 0)
        if current and current_len + addition > chunk_size:
            pieces.append(" ".join(current))
            current = []
            current_len = 0
        current.append(word)
        current_len += len(word) + (1 if len(current) > 1 else 0)
    if current:
        pieces.append(" ".join(current))
    return pieces


def _tail_overlap(sentences: list[str], overlap: int) -> list[str]:
    """Возвращает последние целые предложения в пределах бюджета overlap.

    За счёт переноса целых предложений следующий фрагмент всегда начинается
    с начала предложения.
    """
    if overlap <= 0:
        return []
    tail: list[str] = []
    length = 0
    for sentence in reversed(sentences):
        addition = len(sentence) + (1 if tail else 0)
        if tail and length + addition > overlap:
            break
        tail.insert(0, sentence)
        length += addition
    return tail


def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 150) -> list[str]:
    """Разбивает текст на чанки по границам предложений.

    Гарантии:
    - слова никогда не разрываются;
    - предложения не обрываются на середине;
    - каждый фрагмент начинается с начала предложения.
    """
    sentences = _split_sentences(text)
    if not sentences:
        stripped = (text or "").strip()
        return [stripped] if stripped else []

    units: list[str] = []
    for sentence in sentences:
        if len(sentence) <= chunk_size:
            units.append(sentence)
        else:
            units.extend(_split_long_sentence(sentence, chunk_size))

    def joined_len(items: list[str]) -> int:
        if not items:
            return 0
        return sum(len(item) for item in items) + (len(items) - 1)

    chunks: list[str] = []
    current: list[str] = []
    for unit in units:
        if current and joined_len(current) + 1 + len(unit) > chunk_size:
            chunks.append(" ".join(current))
            current = _tail_overlap(current, overlap)
        current.append(unit)
    if current:
        chunks.append(" ".join(current))
    return chunks
