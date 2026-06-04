from __future__ import annotations

import io
import json
import time
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.documents.storage import object_storage
from app.models.document import Document, DocumentChunk, DocumentVersion
from app.models.enums import DocumentProcessingStatus, TextExtractStatus


class XlsxParsingError(RuntimeError):
    pass


@dataclass
class XlsxSheetResult:
    sheet_index: int
    sheet_name: str
    is_hidden: bool
    range_ref: str | None
    rows_count: int
    columns_count: int
    headers: list[str]
    rows: list[list[Any]]
    merged_ranges: list[str]
    formulas_count: int
    comments_count: int
    text: str
    error: str | None = None


@dataclass
class XlsxParseResult:
    document_id: uuid.UUID
    document_version_id: uuid.UUID
    text: str
    sheets: list[XlsxSheetResult]
    extracted_text_object_name: str
    extraction_method: str
    characters_count: int
    sheets_count: int
    tables_count: int
    failed_sheets: list[str]
    duration_ms: int


class XlsxParsingService:
    """Parse XLSX documents stored in MinIO and persist sheet/table chunks."""

    SUPPORTED_CONTENT_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }

    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    async def parse_document(
        self,
        *,
        document_id: uuid.UUID | None = None,
        document_version_id: uuid.UUID | None = None,
    ) -> XlsxParseResult:
        started = time.perf_counter()
        document, document_version = await self._load_document(
            document_id=document_id,
            document_version_id=document_version_id,
        )

        content_type = document.content_type or document_version.content_type
        if not document.object_name:
            raise XlsxParsingError("У документа нет object_name MinIO")
        if content_type not in self.SUPPORTED_CONTENT_TYPES:
            raise XlsxParsingError(f"Документ не является поддерживаемым XLSX: {content_type}")

        try:
            xlsx_data = object_storage.get_object(document.object_name)
            sheets = self._extract_sheets(xlsx_data)
            if not any(sheet.rows for sheet in sheets):
                raise XlsxParsingError("В XLSX не найдено заполненных данных")

            full_text = "\n\n".join(sheet.text for sheet in sheets if sheet.text.strip())
            extracted_text_object_name = self._save_extraction_result(
                document=document,
                document_version=document_version,
                sheets=sheets,
                full_text=full_text,
            )
            result = XlsxParseResult(
                document_id=document.id,
                document_version_id=document_version.id,
                text=full_text,
                sheets=sheets,
                extracted_text_object_name=extracted_text_object_name,
                extraction_method="openpyxl",
                characters_count=len(full_text),
                sheets_count=len(sheets),
                tables_count=sum(1 for sheet in sheets if sheet.rows),
                failed_sheets=[sheet.sheet_name for sheet in sheets if sheet.error],
                duration_ms=int((time.perf_counter() - started) * 1000),
            )
            await self._persist_result(document, document_version, result)
            return result
        except Exception as exc:
            await self._mark_failed(document, document_version, exc, started)
            if isinstance(exc, XlsxParsingError):
                raise
            raise XlsxParsingError("Не удалось обработать XLSX-документ") from exc

    async def _load_document(
        self,
        *,
        document_id: uuid.UUID | None,
        document_version_id: uuid.UUID | None,
    ) -> tuple[Document, DocumentVersion]:
        if document_version_id:
            document_version = await self.db.get(DocumentVersion, document_version_id)
            if not document_version:
                raise XlsxParsingError("Версия документа не найдена")
            document = await self.db.get(Document, document_version.document_id)
            if not document:
                raise XlsxParsingError("Документ не найден")
            return document, document_version

        if not document_id:
            raise XlsxParsingError("Нужно передать document_id или document_version_id")

        document = await self.db.get(Document, document_id)
        if not document:
            raise XlsxParsingError("Документ не найден")

        result = await self.db.execute(
            select(DocumentVersion)
            .where(DocumentVersion.document_id == document.id)
            .order_by(DocumentVersion.is_current.desc(), DocumentVersion.version_number.desc())
            .limit(1)
        )
        document_version = result.scalar_one_or_none()
        if not document_version:
            raise XlsxParsingError("У документа нет версий")
        return document, document_version

    def _extract_sheets(self, xlsx_data: bytes) -> list[XlsxSheetResult]:
        workbook = load_workbook(io.BytesIO(xlsx_data), read_only=False, data_only=False)
        sheets: list[XlsxSheetResult] = []
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            try:
                sheets.append(self._extract_sheet(index, worksheet))
            except Exception as exc:
                sheets.append(
                    XlsxSheetResult(
                        sheet_index=index,
                        sheet_name=worksheet.title,
                        is_hidden=worksheet.sheet_state != "visible",
                        range_ref=None,
                        rows_count=0,
                        columns_count=0,
                        headers=[],
                        rows=[],
                        merged_ranges=[],
                        formulas_count=0,
                        comments_count=0,
                        text="",
                        error=str(exc),
                    )
                )
        workbook.close()
        return sheets

    def _extract_sheet(self, sheet_index: int, worksheet: Worksheet) -> XlsxSheetResult:
        raw_rows = [
            [self._normalize_cell(cell.value) for cell in row]
            for row in worksheet.iter_rows()
        ]
        trimmed_rows, min_row, min_col = self._trim_empty_rows_and_columns(raw_rows)
        rows_count = len(trimmed_rows)
        columns_count = max((len(row) for row in trimmed_rows), default=0)
        max_row = min_row + rows_count - 1 if rows_count else None
        max_col = min_col + columns_count - 1 if columns_count else None
        range_ref = self._range_ref(min_row, min_col, max_row, max_col)
        headers = self._detect_headers(trimmed_rows)
        merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
        formulas_count = sum(
            1
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        comments_count = sum(1 for row in worksheet.iter_rows() for cell in row if cell.comment)
        text = self._sheet_to_text(
            sheet_name=worksheet.title,
            range_ref=range_ref,
            headers=headers,
            rows=trimmed_rows,
            merged_ranges=merged_ranges,
        )
        return XlsxSheetResult(
            sheet_index=sheet_index,
            sheet_name=worksheet.title,
            is_hidden=worksheet.sheet_state != "visible",
            range_ref=range_ref,
            rows_count=rows_count,
            columns_count=columns_count,
            headers=headers,
            rows=trimmed_rows,
            merged_ranges=merged_ranges,
            formulas_count=formulas_count,
            comments_count=comments_count,
            text=text,
        )

    def _trim_empty_rows_and_columns(
        self,
        rows: list[list[Any]],
    ) -> tuple[list[list[Any]], int, int]:
        non_empty_row_indexes = [
            index for index, row in enumerate(rows) if any(self._has_value(value) for value in row)
        ]
        if not non_empty_row_indexes:
            return [], 1, 1

        min_row_index = min(non_empty_row_indexes)
        max_row_index = max(non_empty_row_indexes)
        relevant_rows = rows[min_row_index : max_row_index + 1]
        non_empty_col_indexes = [
            index
            for index in range(max(len(row) for row in relevant_rows))
            if any(index < len(row) and self._has_value(row[index]) for row in relevant_rows)
        ]
        if not non_empty_col_indexes:
            return [], min_row_index + 1, 1

        min_col_index = min(non_empty_col_indexes)
        max_col_index = max(non_empty_col_indexes)
        trimmed = [
            [
                row[index] if index < len(row) else None
                for index in range(min_col_index, max_col_index + 1)
            ]
            for row in relevant_rows
        ]
        return trimmed, min_row_index + 1, min_col_index + 1

    def _detect_headers(self, rows: list[list[Any]]) -> list[str]:
        for row in rows[:5]:
            values = ["" if value is None else str(value).strip() for value in row]
            filled = [value for value in values if value]
            if len(filled) >= 2:
                return values
        return []

    def _sheet_to_text(
        self,
        *,
        sheet_name: str,
        range_ref: str | None,
        headers: list[str],
        rows: list[list[Any]],
        merged_ranges: list[str],
    ) -> str:
        lines = [f"Лист: {sheet_name}"]
        if range_ref:
            lines.append(f"Диапазон: {range_ref}")
        if headers:
            lines.append("Колонки: " + " | ".join(header or "-" for header in headers))
        if merged_ranges:
            lines.append("Объединённые ячейки: " + ", ".join(merged_ranges))

        for row_number, row in enumerate(rows, start=1):
            row_values = ["" if value is None else str(value) for value in row]
            if any(value.strip() for value in row_values):
                lines.append(f"Строка {row_number}: " + " | ".join(row_values))
        return "\n".join(lines)

    def _save_extraction_result(
        self,
        *,
        document: Document,
        document_version: DocumentVersion,
        sheets: list[XlsxSheetResult],
        full_text: str,
    ) -> str:
        object_name = f"documents/{document.id}/extracted_text/xlsx_v{document_version.version_number}.json"
        payload = {
            "document_id": str(document.id),
            "document_version_id": str(document_version.id),
            "text": full_text,
            "sheets": [sheet.__dict__ for sheet in sheets],
        }
        object_storage.put_object(
            object_name,
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json",
        )
        return object_name

    async def _persist_result(
        self,
        document: Document,
        document_version: DocumentVersion,
        result: XlsxParseResult,
    ) -> None:
        status = (
            DocumentProcessingStatus.FAILED
            if result.failed_sheets and not result.text.strip()
            else DocumentProcessingStatus.TEXT_EXTRACTED
        )
        text_status = (
            TextExtractStatus.FAILED
            if result.failed_sheets and not result.text.strip()
            else TextExtractStatus.EXTRACTED
        )
        document.pages_count = result.sheets_count
        document.processing_status = status
        document.text_extract_status = text_status
        document.extracted_text_object_name = result.extracted_text_object_name
        document.metadata_ = self._merge_metadata(document.metadata_, result)

        document_version.pages_count = result.sheets_count
        document_version.processing_status = status
        document_version.text_extract_status = text_status
        document_version.extracted_text_object_name = result.extracted_text_object_name
        document_version.metadata_ = self._merge_metadata(document_version.metadata_, result)

        await self.db.execute(delete(DocumentChunk).where(DocumentChunk.document_version_id == document_version.id))
        chunk_index = 0
        for sheet in result.sheets:
            if not sheet.text.strip():
                continue
            self.db.add(
                DocumentChunk(
                    document_id=document.id,
                    document_version_id=document_version.id,
                    chunk_index=chunk_index,
                    text=sheet.text,
                    section_title=sheet.sheet_name,
                    token_count=len(sheet.text.split()),
                    qdrant_collection=settings.QDRANT_COLLECTION,
                    embedding_model=settings.LLM_EMBEDDING_MODEL,
                    is_indexed=False,
                    metadata_={
                        "source": "xlsx_parser",
                        "extraction_method": result.extraction_method,
                        "sheet_name": sheet.sheet_name,
                        "sheet_index": sheet.sheet_index,
                        "range_ref": sheet.range_ref,
                        "rows_count": sheet.rows_count,
                        "columns_count": sheet.columns_count,
                        "headers": sheet.headers,
                        "merged_ranges": sheet.merged_ranges,
                        "formulas_count": sheet.formulas_count,
                        "comments_count": sheet.comments_count,
                        "is_hidden": sheet.is_hidden,
                    },
                    content=sheet.text,
                    chunk_metadata={
                        "source": "xlsx_parser",
                        "sheet_name": sheet.sheet_name,
                        "range_ref": sheet.range_ref,
                    },
                )
            )
            chunk_index += 1
        await self.db.flush()

    async def _mark_failed(
        self,
        document: Document,
        document_version: DocumentVersion,
        exc: Exception,
        started: float,
    ) -> None:
        metadata = {
            **(document.metadata_ or {}),
            "xlsx_parsing": {
                "status": "failed",
                "error": str(exc),
                "duration_ms": int((time.perf_counter() - started) * 1000),
            },
        }
        document.processing_status = DocumentProcessingStatus.FAILED
        document.text_extract_status = TextExtractStatus.FAILED
        document.metadata_ = metadata
        document_version.processing_status = DocumentProcessingStatus.FAILED
        document_version.text_extract_status = TextExtractStatus.FAILED
        document_version.metadata_ = {
            **(document_version.metadata_ or {}),
            "xlsx_parsing": metadata["xlsx_parsing"],
        }
        await self.db.flush()

    def _merge_metadata(self, current: dict | None, result: XlsxParseResult) -> dict:
        return {
            **(current or {}),
            "xlsx_parsing": {
                "status": "partial" if result.failed_sheets else "completed",
                "extraction_method": result.extraction_method,
                "characters_count": result.characters_count,
                "sheets_count": result.sheets_count,
                "tables_count": result.tables_count,
                "failed_sheets": result.failed_sheets,
                "duration_ms": result.duration_ms,
                "sheets": [
                    {
                        "sheet_name": sheet.sheet_name,
                        "range_ref": sheet.range_ref,
                        "rows_count": sheet.rows_count,
                        "columns_count": sheet.columns_count,
                        "headers": sheet.headers,
                        "merged_ranges": sheet.merged_ranges,
                        "formulas_count": sheet.formulas_count,
                        "comments_count": sheet.comments_count,
                        "is_hidden": sheet.is_hidden,
                        "error": sheet.error,
                    }
                    for sheet in result.sheets
                ],
            },
        }

    def _range_ref(
        self,
        min_row: int,
        min_col: int,
        max_row: int | None,
        max_col: int | None,
    ) -> str | None:
        if max_row is None or max_col is None:
            return None
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    def _normalize_cell(self, value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _has_value(self, value: Any) -> bool:
        return value is not None and str(value).strip() != ""
