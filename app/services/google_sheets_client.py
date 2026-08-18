"""Google Sheets API v4 through Service Account (Aveon)."""

from __future__ import annotations

import json
import time
from functools import lru_cache
from pathlib import Path
from typing import Any

from app.core.config import get_settings

SCOPES = (
    "https://www.googleapis.com/auth/spreadsheets.readonly",
    "https://www.googleapis.com/auth/drive.readonly",
)

DEFAULT_SHEET_TITLE = "ИТЦ В РАБОТЕ"

CHINA_WORKSHEET_ALIASES = (
    "ИТЦ В РАБОТЕ",
    "Гонконг В РАБОТЕ",
    "ГОНКОНГ В РАБОТЕ",
    "ITC В РАБОТЕ",
)


def _normalize_sheet_title(value: str) -> str:
    return str(value or "").strip().casefold()


def is_china_worksheet_title(title: str) -> bool:
    normalized = _normalize_sheet_title(title)
    if not normalized:
        return False
    if normalized in {_normalize_sheet_title(item) for item in CHINA_WORKSHEET_ALIASES}:
        return True
    return "в работе" in normalized


def _sheet_props_list(spreadsheet: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for sheet in spreadsheet.get("sheets") or []:
        props = sheet.get("properties") or {}
        title = str(props.get("title") or "").strip()
        if not title:
            continue
        items.append(
            {
                "title": title,
                "gid": props.get("sheetId"),
                "index": props.get("index"),
            }
        )
    return items


def _resolve_preferred_sheet(
    sheets: list[dict[str, Any]],
    *,
    sheet_gid: str | None = None,
    sheet_title: str | None = None,
) -> dict[str, Any] | None:
    if not sheets:
        return None
    if sheet_gid:
        for item in sheets:
            if str(item.get("gid")) == str(sheet_gid):
                return item
    if sheet_title:
        target = _normalize_sheet_title(sheet_title)
        for item in sheets:
            if _normalize_sheet_title(str(item.get("title") or "")) == target:
                return item
    for alias in CHINA_WORKSHEET_ALIASES:
        target = _normalize_sheet_title(alias)
        for item in sheets:
            if _normalize_sheet_title(str(item.get("title") or "")) == target:
                return item
    for item in sheets:
        if is_china_worksheet_title(str(item.get("title") or "")):
            return item
    return sheets[0]


class GoogleSheetsConfigError(Exception):
    """Service Account is not configured or JSON is invalid."""


def is_configured() -> bool:
    settings = get_settings()
    return bool(
        (settings.GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON or "").strip()
        or (settings.GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE or "").strip()
    )


def _load_service_account_info() -> dict[str, Any]:
    settings = get_settings()
    raw_json = (settings.GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON or "").strip()
    if raw_json:
        try:
            return json.loads(raw_json)
        except json.JSONDecodeError as exc:
            raise GoogleSheetsConfigError(
                f"GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON: invalid JSON ({exc})"
            ) from exc

    file_path = (settings.GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE or "").strip()
    if file_path:
        path = Path(file_path)
        if not path.is_file():
            raise GoogleSheetsConfigError(
                f"GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE not found: {file_path}"
            )
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise GoogleSheetsConfigError(
                f"Service Account file is not valid JSON: {file_path}"
            ) from exc

    raise GoogleSheetsConfigError(
        "Google Sheets Service Account is not configured. "
        "Set GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE or GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON."
    )


@lru_cache(maxsize=1)
def _credentials_bundle() -> tuple[Any, str | None]:
    from google.oauth2 import service_account

    info = _load_service_account_info()
    credentials = service_account.Credentials.from_service_account_info(
        info,
        scopes=list(SCOPES),
    )
    client_email = info.get("client_email")
    return credentials, str(client_email) if client_email else None


def get_service_account_email() -> str | None:
    if not is_configured():
        return None
    try:
        return _credentials_bundle()[1]
    except Exception:
        return None


def _sheets_service():
    from googleapiclient.discovery import build

    credentials, _ = _credentials_bundle()
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def _find_sheet(
    spreadsheet: dict[str, Any],
    *,
    sheet_gid: str | None = None,
    sheet_title: str | None = None,
) -> dict[str, Any] | None:
    sheets = spreadsheet.get("sheets") or []
    if sheet_title:
        title_norm = sheet_title.strip().casefold()
        for sheet in sheets:
            props = sheet.get("properties") or {}
            if str(props.get("title") or "").strip().casefold() == title_norm:
                return sheet
    if sheet_gid:
        for sheet in sheets:
            props = sheet.get("properties") or {}
            if str(props.get("sheetId")) == str(sheet_gid):
                return sheet
    return None


def _preview_values(values: list[list[Any]], max_rows: int = 5, max_cols: int = 10) -> list[list[Any]]:
    preview: list[list[Any]] = []
    for row in values[:max_rows]:
        preview.append([(cell if cell is not None else "") for cell in row[:max_cols]])
    return preview


def _normalize_matrix(values: list[list[Any]]) -> list[list[str]]:
    width = max((len(row) for row in values), default=0)
    matrix: list[list[str]] = []
    for row in values:
        cells = [("" if cell is None else str(cell)) for cell in row]
        if len(cells) < width:
            cells.extend([""] * (width - len(cells)))
        matrix.append(cells)
    return matrix


def fetch_spreadsheet_all_sheets(
    spreadsheet_id: str,
    *,
    sheet_gid: str | None = None,
    preferred_sheet_title: str | None = None,
    include_values: bool = True,
    preview_rows: int = 5,
) -> dict[str, Any]:
    """Читает все листы таблицы через Google Sheets API v4."""
    started = time.perf_counter()
    record: dict[str, Any] = {
        "name": "service_account_api_all_sheets",
        "ok": False,
        "spreadsheet_id": spreadsheet_id,
        "sheet_gid": sheet_gid,
        "sheet_title": preferred_sheet_title or DEFAULT_SHEET_TITLE,
        "service_account_email": get_service_account_email(),
        "elapsed_ms": None,
        "error": None,
        "hint": None,
        "parsed": None,
    }

    if not is_configured():
        record["error"] = "Service Account is not configured on backend"
        record["hint"] = (
            "Set GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE or GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON in .env"
        )
        return record

    try:
        service = _sheets_service()
        meta = (
            service.spreadsheets()
            .get(spreadsheetId=spreadsheet_id, includeGridData=False)
            .execute()
        )
        sheet_items = _sheet_props_list(meta)
        if not sheet_items:
            record["error"] = "Spreadsheet has no worksheets"
            record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
            return record

        spreadsheet_title = (meta.get("properties") or {}).get("title")
        loaded_sheets: list[dict[str, Any]] = []
        for item in sheet_items:
            title = str(item["title"])
            sheet_record: dict[str, Any] = {
                "title": title,
                "gid": item.get("gid"),
                "index": item.get("index"),
                "ok": False,
                "row_count": 0,
                "column_count": 0,
                "values": [],
                "preview_rows": [],
                "error": None,
            }
            try:
                values_result = (
                    service.spreadsheets()
                    .values()
                    .get(spreadsheetId=spreadsheet_id, range=f"'{title}'")
                    .execute()
                )
                values = values_result.get("values") or []
                matrix = _normalize_matrix(values)
                sheet_record.update(
                    {
                        "ok": True,
                        "row_count": len(matrix),
                        "column_count": max((len(row) for row in matrix), default=0),
                        "preview_rows": _preview_values(matrix, max_rows=preview_rows),
                    }
                )
                if include_values:
                    sheet_record["values"] = matrix
            except Exception as exc:
                sheet_record["error"] = f"{type(exc).__name__}: {exc}"
            loaded_sheets.append(sheet_record)

        preferred = _resolve_preferred_sheet(
            sheet_items,
            sheet_gid=sheet_gid,
            sheet_title=preferred_sheet_title or DEFAULT_SHEET_TITLE,
        )
        preferred_title = str((preferred or {}).get("title") or DEFAULT_SHEET_TITLE)
        preferred_payload = next(
            (item for item in loaded_sheets if item.get("title") == preferred_title),
            loaded_sheets[0],
        )
        any_ok = any(item.get("ok") for item in loaded_sheets)
        record["ok"] = any_ok
        record["sheet_gid"] = (
            str(preferred_payload.get("gid"))
            if preferred_payload.get("gid") is not None
            else sheet_gid
        )
        record["sheet_title"] = preferred_title
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        record["hint"] = "Access via Google Sheets API v4 (Service Account)"
        record["parsed"] = {
            "format": "sheets_api_v4_all",
            "spreadsheet_title": spreadsheet_title,
            "sheet_title": preferred_title,
            "sheet_gid": preferred_payload.get("gid"),
            "row_count": preferred_payload.get("row_count", 0),
            "column_count": preferred_payload.get("column_count", 0),
            "preview_rows": preferred_payload.get("preview_rows") or [],
            "values": preferred_payload.get("values") or [],
            "sheets": loaded_sheets,
            "preferred_sheet_title": preferred_title,
            "available_sheets": [
                {"gid": item.get("gid"), "title": item.get("title")} for item in sheet_items
            ],
        }
        if not any_ok:
            record["error"] = "Failed to read all worksheets"
        return record
    except GoogleSheetsConfigError as exc:
        record["error"] = str(exc)
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        return record
    except ImportError as exc:
        record["error"] = f"Google API libraries not installed: {exc}"
        record["hint"] = "pip install google-auth google-api-python-client"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        return record
    except Exception as exc:
        message = str(exc)
        record["error"] = f"{type(exc).__name__}: {message}"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        lowered = message.lower()
        if "403" in message or "permission" in lowered or "forbidden" in lowered:
            email = get_service_account_email() or "<service-account-email>"
            record["hint"] = f"Share the spreadsheet with {email} (Viewer access)"
        elif "404" in message or "not found" in lowered:
            record["hint"] = "Check GOOGLE_SHEETS_SPREADSHEET_ID"
        elif "invalid_grant" in lowered:
            record["hint"] = "Check Service Account JSON and that Sheets API is enabled"
        return record


def fetch_sheet_via_api(
    spreadsheet_id: str,
    sheet_gid: str | None = None,
    *,
    sheet_title: str | None = None,
    include_values: bool = False,
    preview_rows: int = 5,
) -> dict[str, Any]:
    """Read a sheet via Google Sheets API v4 (Service Account)."""
    started = time.perf_counter()
    target_title = (sheet_title or DEFAULT_SHEET_TITLE).strip()
    record: dict[str, Any] = {
        "name": "service_account_api",
        "ok": False,
        "spreadsheet_id": spreadsheet_id,
        "sheet_gid": sheet_gid,
        "sheet_title": target_title,
        "service_account_email": get_service_account_email(),
        "elapsed_ms": None,
        "error": None,
        "hint": None,
        "parsed": None,
    }

    if not is_configured():
        record["error"] = "Service Account is not configured on backend"
        record["hint"] = (
            "Set GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE or GOOGLE_SHEETS_SERVICE_ACCOUNT_JSON in .env"
        )
        return record

    try:
        service = _sheets_service()
        meta = (
            service.spreadsheets()
            .get(spreadsheetId=spreadsheet_id, includeGridData=False)
            .execute()
        )
        sheet = _find_sheet(meta, sheet_gid=sheet_gid, sheet_title=target_title)
        if sheet is None:
            available = [
                {
                    "gid": (item.get("properties") or {}).get("sheetId"),
                    "title": (item.get("properties") or {}).get("title"),
                }
                for item in meta.get("sheets") or []
            ]
            record["error"] = f'Sheet "{target_title}" was not found'
            record["parsed"] = {"available_sheets": available}
            record["hint"] = "Check sheet title or GOOGLE_SHEETS_SHEET_GID"
            record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
            return record

        props = sheet.get("properties") or {}
        resolved_title = props.get("title") or target_title
        spreadsheet_title = (meta.get("properties") or {}).get("title")
        resolved_gid = props.get("sheetId")

        values_result = (
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=f"'{resolved_title}'")
            .execute()
        )
        values = values_result.get("values") or []
        matrix = _normalize_matrix(values)
        elapsed_ms = round((time.perf_counter() - started) * 1000)

        parsed: dict[str, Any] = {
            "format": "sheets_api_v4",
            "spreadsheet_title": spreadsheet_title,
            "sheet_title": resolved_title,
            "sheet_gid": resolved_gid,
            "row_count": len(matrix),
            "column_count": max((len(row) for row in matrix), default=0),
            "preview_rows": _preview_values(matrix, max_rows=preview_rows),
        }
        if include_values:
            parsed["values"] = matrix

        record["ok"] = True
        record["sheet_gid"] = str(resolved_gid) if resolved_gid is not None else sheet_gid
        record["sheet_title"] = resolved_title
        record["elapsed_ms"] = elapsed_ms
        record["hint"] = "Access via Google Sheets API v4 (Service Account)"
        record["parsed"] = parsed
        return record
    except GoogleSheetsConfigError as exc:
        record["error"] = str(exc)
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        return record
    except ImportError as exc:
        record["error"] = f"Google API libraries not installed: {exc}"
        record["hint"] = "pip install google-auth google-api-python-client"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        return record
    except Exception as exc:
        message = str(exc)
        record["error"] = f"{type(exc).__name__}: {message}"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
        lowered = message.lower()
        if "403" in message or "permission" in lowered or "forbidden" in lowered:
            email = get_service_account_email() or "<service-account-email>"
            record["hint"] = (
                f"Share the spreadsheet with {email} (Viewer access)"
            )
        elif "404" in message or "not found" in lowered:
            record["hint"] = "Check GOOGLE_SHEETS_SPREADSHEET_ID"
        elif "invalid_grant" in lowered:
            record["hint"] = "Check Service Account JSON and that Sheets API is enabled"
        return record


def get_default_spreadsheet_target() -> tuple[str, str]:
    settings = get_settings()
    spreadsheet_id = (settings.GOOGLE_SHEETS_SPREADSHEET_ID or "").strip()
    sheet_gid = (settings.GOOGLE_SHEETS_SHEET_GID or "").strip()
    if not spreadsheet_id:
        raise GoogleSheetsConfigError("Set GOOGLE_SHEETS_SPREADSHEET_ID in .env")
    return spreadsheet_id, sheet_gid


def _values_matrix_to_xlsx_bytes(sheet_title: str, values: list[list[str]]) -> bytes:
    import io

    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]
    for row_index, row in enumerate(values, start=1):
        for col_index, cell in enumerate(row, start=1):
            if cell != "":
                worksheet.cell(row_index, col_index, cell)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _multi_sheet_values_to_xlsx_bytes(
    spreadsheet_title: str,
    sheets: list[dict[str, Any]],
) -> bytes:
    import io

    import openpyxl

    workbook = openpyxl.Workbook()
    default_ws = workbook.active
    workbook.remove(default_ws)
    used_titles: set[str] = set()
    for sheet in sheets:
        if not sheet.get("ok"):
            continue
        values = sheet.get("values") or []
        if not values:
            continue
        raw_title = str(sheet.get("title") or "Sheet")[:31]
        title = raw_title
        suffix = 2
        while title in used_titles:
            tail = f" ({suffix})"
            title = f"{raw_title[: max(1, 31 - len(tail))]}{tail}"
            suffix += 1
        used_titles.add(title)
        worksheet = workbook.create_sheet(title=title)
        for row_index, row in enumerate(values, start=1):
            for col_index, cell in enumerate(row, start=1):
                if cell != "":
                    worksheet.cell(row_index, col_index, cell)
    if not workbook.sheetnames:
        worksheet = workbook.create_sheet(title="Sheet1")
        worksheet.cell(1, 1, spreadsheet_title or "")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def fetch_itc_sheet_workbook_payload() -> tuple[str, bytes] | None:
    """Скачивает все листы Google Sheets (Китай) как .xlsx для merge."""
    if not is_configured():
        return None

    try:
        spreadsheet_id, sheet_gid = get_default_spreadsheet_target()
    except GoogleSheetsConfigError:
        return None

    result = fetch_spreadsheet_all_sheets(
        spreadsheet_id,
        sheet_gid=sheet_gid or None,
        preferred_sheet_title=DEFAULT_SHEET_TITLE,
        include_values=True,
    )
    if not result.get("ok"):
        return None

    parsed = result.get("parsed") or {}
    sheets = [item for item in (parsed.get("sheets") or []) if item.get("ok")]
    if not sheets:
        return None

    spreadsheet_title = str(parsed.get("spreadsheet_title") or "google_sheets")
    safe_name = spreadsheet_title.replace("/", "-").strip() or "google_sheets"
    filename = f"google_sheets_{safe_name}.xlsx"
    return filename, _multi_sheet_values_to_xlsx_bytes(spreadsheet_title, sheets)
