"""Executive procurement summary for Aveon document analysis."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_TITLE_FILL = PatternFill("solid", fgColor="FF0B3040")
_HEADER_FILL = PatternFill("solid", fgColor="FFD9EAF7")
_WARN_FILL = PatternFill("solid", fgColor="FFFFF2CC")
_DANGER_FILL = PatternFill("solid", fgColor="FFF8CBAD")
_OK_FILL = PatternFill("solid", fgColor="FFE2F0D9")
_THIN = Side(style="thin", color="FFB0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

_MAX_ACTION_ITEMS = 25
_MAX_LOGISTICS_ITEMS = 12
_MAX_RED_PRODUCTS = 10
_MAX_SHORTAGES_PER_ROW = 3
_TEXT_LIMIT = 160
_ACTION_TEXT_LIMIT = 120


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _num(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _short(value: Any, limit: int = _TEXT_LIMIT) -> str:
    text = " ".join(_text(value).split())
    if len(text) <= limit:
        return text
    return f"{text[: limit - 1].rstrip()}…"


def _format_now() -> str:
    return datetime.now().astimezone().strftime("%d.%m.%Y %H:%M")


def _format_iso_datetime(value: Any) -> str:
    raw = _text(value)
    if not raw:
        return "—"
    try:
        normalized = raw.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone()
        return parsed.strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return raw[:16]


def _status_label(status: str) -> str:
    return {
        "submitted": "смена закрыта",
        "in_progress": "в работе",
        "missing": "нет отчёта",
        "resolved": "выполнено",
        "partial": "частично",
        "not_resolved": "не выполнено",
        "active": "активно",
    }.get(status, status or "—")


def _sheet_title(ws, title: str, subtitle: str = "", columns: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = ws.cell(1, 1, title)
    cell.fill = _TITLE_FILL
    cell.font = Font(bold=True, color="FFFFFFFF", size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24
    for col in range(2, columns + 1):
        ws.cell(1, col).fill = _TITLE_FILL

    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
        cell = ws.cell(2, 1, subtitle)
        cell.fill = PatternFill("solid", fgColor="FFF2F6FA")
        cell.font = Font(color="FF304050", size=10)
        cell.alignment = _WRAP
        ws.row_dimensions[2].height = 34


def _write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    *,
    widths: list[float] | None = None,
) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = _HEADER_FILL
        cell.font = Font(bold=True)
        cell.border = _BORDER
        cell.alignment = _CENTER

    for row_idx, values in enumerate(rows, start_row + 1):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col, value)
            cell.border = _BORDER
            cell.alignment = _WRAP
            if isinstance(value, (int, float)):
                cell.alignment = _CENTER

    if widths:
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    return start_row + len(rows) + 2


def _coverage_periods(snapshot: dict[str, Any] | None) -> dict[str, Any]:
    coverage = (snapshot or {}).get("coverage_dashboard")
    if not isinstance(coverage, dict):
        return {}
    periods = coverage.get("periods")
    return periods if isinstance(periods, dict) else {}


def _coverage_tiles(snapshot: dict[str, Any] | None, period_key: str, side_key: str) -> dict[str, Any]:
    period = _coverage_periods(snapshot).get(period_key)
    if not isinstance(period, dict):
        return {}
    side = period.get(side_key)
    if not isinstance(side, dict):
        return {}
    tiles = side.get("tiles")
    return tiles if isinstance(tiles, dict) else {}


def _coverage_summary_compact(snapshot: dict[str, Any] | None) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for period_key, default_label in (("day", "День"), ("week", "Неделя")):
        period = _coverage_periods(snapshot).get(period_key)
        if not isinstance(period, dict):
            continue
        label = _text(period.get("label") or default_label)
        for side_key, side_label in (("products", "Изделия"), ("nomenclatures", "Номенклатуры")):
            tiles = _coverage_tiles(snapshot, period_key, side_key)
            if not tiles:
                continue
            all_count = int(_num(tiles.get("all")))
            green = int(_num(tiles.get("green")))
            yellow = int(_num(tiles.get("yellow")))
            red = int(_num(tiles.get("red")))
            covered_pct = round((green / all_count) * 100) if all_count else 0
            rows.append(
                [
                    label,
                    side_label,
                    all_count,
                    green,
                    yellow,
                    red,
                    f"{covered_pct}%",
                    round(_num(tiles.get("plan_total")), 1),
                ]
            )
    return rows


def _top_red_products(snapshot: dict[str, Any] | None, *, limit: int = _MAX_RED_PRODUCTS) -> list[list[Any]]:
    period = _coverage_periods(snapshot).get("week") or _coverage_periods(snapshot).get("day")
    if not isinstance(period, dict):
        return []
    side = period.get("products")
    if not isinstance(side, dict):
        return []

    candidates: list[tuple[float, list[Any]]] = []
    for row in side.get("rows") or []:
        if not isinstance(row, dict):
            continue
        if _text(row.get("status")) != "red":
            continue
        shortages = row.get("shortages") or []
        top_shortages = "; ".join(
            f"{_short(item.get('name'), 48)} (−{round(_num(item.get('shortage')), 1)})"
            for item in shortages[:_MAX_SHORTAGES_PER_ROW]
            if isinstance(item, dict)
        )
        shortage_total = sum(_num(item.get("shortage")) for item in shortages if isinstance(item, dict))
        candidates.append(
            (
                shortage_total,
                [
                    _short(row.get("name"), 56),
                    round(_num(row.get("plan")), 1),
                    round(_num(row.get("covered")), 1),
                    _short(top_shortages, 120) or "—",
                ],
            )
        )

    candidates.sort(key=lambda item: item[0], reverse=True)
    return [row for _, row in candidates[:limit]]


def _top_logistics_risks(snapshot: dict[str, Any] | None, *, limit: int = _MAX_LOGISTICS_ITEMS) -> list[list[Any]]:
    risks = (snapshot or {}).get("logistics_risks")
    if not isinstance(risks, dict):
        return []

    rows: list[tuple[int, list[Any]]] = []
    for stage in risks.get("stages") or []:
        if not isinstance(stage, dict):
            continue
        stage_label = _short(stage.get("label") or stage.get("key"), 24)
        for item in stage.get("items") or []:
            if not isinstance(item, dict):
                continue
            risk = _text(item.get("risk_level") or "critical").lower()
            if risk not in {"critical", "high", "urgent"}:
                continue
            rows.append(
                (
                    0 if risk == "critical" else 1,
                    [
                        stage_label,
                        _short(item.get("nomenclature"), 56),
                        _short(item.get("supplier"), 28) or "—",
                        _text(item.get("quantity")) or "—",
                        _text(item.get("moscow_date")) or "—",
                        _text(item.get("milestone_date")) or "—",
                    ],
                )
            )

    rows.sort(key=lambda item: (item[0], item[1][4]))
    return [row for _, row in rows[:limit]]


def _priority_score(task: dict[str, Any]) -> int:
    status = _text(task.get("status"))
    status_score = {
        "not_resolved": 100,
        "partial": 80,
        "active": 70,
        "resolved": 0,
    }.get(status, 50)
    priority = _text(task.get("priority")).lower()
    priority_score = {"urgent": 30, "today": 20, "high": 15}.get(priority, 5)
    deficit_score = min(int(_num(task.get("deficit")) // 1000), 20)
    return status_score + priority_score + deficit_score


def _collect_open_tasks(completion: dict[str, Any]) -> list[dict[str, Any]]:
    period_mode = _text(completion.get("period_mode") or "day")
    deduped: dict[tuple[str, str], dict[str, Any]] = {}

    for manager in completion.get("managers") or []:
        if not isinstance(manager, dict):
            continue
        manager_name = _text(manager.get("manager_name"))
        for task in manager.get("tasks") or []:
            if not isinstance(task, dict):
                continue
            if _text(task.get("status")) == "resolved":
                continue
            nomenclature = _text(task.get("nomenclature")) or _text(task.get("task_type")) or "—"
            shift_date = _text(task.get("shift_date") or manager.get("report_date"))
            key = (manager_name, nomenclature.casefold())
            enriched = {
                **task,
                "manager_name": manager_name,
                "shift_date": shift_date,
                "report_status": _text(manager.get("report_status")),
                "_score": _priority_score(task),
            }
            existing = deduped.get(key)
            if existing is None or enriched["_score"] > existing["_score"]:
                deduped[key] = enriched
            elif enriched["_score"] == existing["_score"] and shift_date > _text(existing.get("shift_date")):
                deduped[key] = enriched

    tasks = list(deduped.values())
    if period_mode == "day":
        report_date = _text(completion.get("report_date"))
        tasks = [task for task in tasks if _text(task.get("shift_date")) == report_date or not task.get("shift_date")]

    tasks.sort(key=lambda item: (-int(item.get("_score") or 0), _text(item.get("shift_date"))))
    return tasks


def _action_rows(completion: dict[str, Any], snapshot: dict[str, Any] | None) -> list[list[Any]]:
    rows: list[tuple[int, list[Any]]] = []

    for task in _collect_open_tasks(completion):
        rows.append(
            (
                int(task.get("_score") or 0),
                [
                    "Менеджер",
                    _short(task.get("nomenclature"), 56),
                    _short(task.get("problem"), _ACTION_TEXT_LIMIT),
                    _text(task.get("manager_name")),
                    _status_label(_text(task.get("status"))),
                    _short(task.get("solution") or task.get("reason") or task.get("eval_comment"), 90),
                ],
            )
        )

    for product_row in _top_red_products(snapshot, limit=8):
        rows.append(
            (
                95,
                [
                    "Обеспеченность",
                    product_row[0],
                    f"Не обеспечено: план {product_row[1]}, есть {product_row[2]}",
                    "Производство",
                    "критично",
                    _short(product_row[3], 90),
                ],
            )
        )

    for logistics_row in _top_logistics_risks(snapshot, limit=6):
        rows.append(
            (
                90,
                [
                    "Логистика",
                    logistics_row[1],
                    f"Риск поставки, контроль {logistics_row[5]}",
                    logistics_row[2],
                    "критично",
                    f"Этап: {logistics_row[0]}",
                ],
            )
        )

    rows.sort(key=lambda item: -item[0])
    return [row for _, row in rows[:_MAX_ACTION_ITEMS]]


def _manager_scorecard_rows(completion: dict[str, Any]) -> list[list[Any]]:
    period_mode = _text(completion.get("period_mode") or "day")
    include_days = period_mode != "day"
    rows: list[list[Any]] = []

    for manager in completion.get("managers") or []:
        if not isinstance(manager, dict):
            continue
        stats = manager.get("stats") or {}
        open_tasks = [
            task
            for task in manager.get("tasks") or []
            if isinstance(task, dict) and _text(task.get("status")) != "resolved"
        ]
        critical_open = sum(
            1
            for task in open_tasks
            if _text(task.get("status")) in {"not_resolved", "partial"}
            or _text(task.get("priority")).lower() in {"urgent", "today"}
        )
        row = [
            _text(manager.get("manager_name")),
            _text(manager.get("region_label")),
            _status_label(_text(manager.get("report_status"))),
        ]
        if include_days:
            row.append(int(_num(manager.get("days_with_reports"))))
        row.extend(
            [
                int(_num(stats.get("total"))),
                int(_num(stats.get("resolved"))),
                int(_num(stats.get("incomplete"))),
                int(_num(stats.get("resolved_percent"))),
                len(open_tasks),
                critical_open,
                _format_iso_datetime(manager.get("email_sent_at") or manager.get("live_updated_at")),
            ]
        )
        rows.append(row)
    return rows


def _manager_scorecard_headers(completion: dict[str, Any]) -> list[str]:
    headers = ["Менеджер", "Зона", "Статус смены"]
    if _text(completion.get("period_mode") or "day") != "day":
        headers.append("Дней с отчётом")
    headers.extend(
        [
            "Задач",
            "Закрыто",
            "Открыто",
            "%",
            "Открытых",
            "Срочных",
            "Обновлено",
        ]
    )
    return headers


def _kpi_rows(completion: dict[str, Any], snapshot: dict[str, Any] | None) -> list[list[Any]]:
    summary = completion.get("summary") or {}
    roster = completion.get("roster") or {}
    period_mode = _text(completion.get("period_mode") or "day")

    week_products = _coverage_tiles(snapshot, "week", "products")
    day_products = _coverage_tiles(snapshot, "day", "products")
    red_week = int(_num(week_products.get("red")))
    yellow_week = int(_num(week_products.get("yellow")))
    all_week = int(_num(week_products.get("all")))
    coverage_pct = round((int(_num(week_products.get("green"))) / all_week) * 100) if all_week else 0

    open_tasks = _collect_open_tasks(completion)
    logistics_count = len(_top_logistics_risks(snapshot, limit=999))

    manager_label = "Задачи менеджеров за день" if period_mode == "day" else "Задачи менеджеров за период"
    manager_value = (
        f"{int(_num(summary.get('resolved')))}/{int(_num(summary.get('total')))} закрыто "
        f"({int(_num(summary.get('resolved_percent')))}%)"
    )

    if period_mode == "day":
        discipline_value = (
            f"Смен сдано {int(_num(roster.get('submitted')))}/{int(_num(roster.get('total')))}, "
            f"в работе {int(_num(roster.get('in_progress')))}"
        )
    else:
        discipline_value = (
            f"Закрыто {int(_num(completion.get('submitted_shift_days')))} смен из "
            f"{int(_num(completion.get('expected_shift_days')))}"
        )

    return [
        [manager_label, manager_value, "Зелёный — норма, жёлтый — внимание, красный — блокер"],
        ["Открытые задачи", str(len(open_tasks)), "Срочные и незакрытые позиции для планёрки"],
        ["Дисциплина смен", discipline_value, "Контроль сдачи отчётов менеджерами"],
        [
            "Обеспеченность (неделя, изделия)",
            f"{coverage_pct}% обеспечено, красных {red_week}, частичных {yellow_week}",
            "Текущий срез производственного плана",
        ],
        [
            "Не обеспечено сегодня (изделия)",
            str(int(_num(day_products.get("red")))),
            "Позиции с нулевым покрытием на сегодня",
        ],
        ["Логистические риски", str(logistics_count), "Критичные поставки без подтверждения"],
        ["Пунктов «требует решения»", str(min(len(_action_rows(completion, snapshot)), _MAX_ACTION_ITEMS)), "Сводный приоритетный список"],
    ]


def _management_focus(completion: dict[str, Any], snapshot: dict[str, Any] | None) -> list[list[Any]]:
    actions = _action_rows(completion, snapshot)
    if not actions:
        return [["Статус", "Критичных блокеров не выявлено", "Поддерживать текущий темп закрытия задач"]]

    focus: list[list[Any]] = []
    manager_items = [row for row in actions if row[0] == "Менеджер"]
    coverage_items = [row for row in actions if row[0] == "Обеспеченность"]
    logistics_items = [row for row in actions if row[0] == "Логистика"]

    if manager_items:
        focus.append(
            [
                "1. Задачи менеджеров",
                f"{len(manager_items)} открытых позиций, топ: {_short(manager_items[0][1], 50)}",
                "Разобрать на планёрке с ответственными менеджерами",
            ]
        )
    if coverage_items:
        focus.append(
            [
                "2. Обеспеченность",
                f"{len(coverage_items)} изделий без покрытия на неделе",
                "Согласовать закупку/перераспределение по TOP-дефицитам",
            ]
        )
    if logistics_items:
        focus.append(
            [
                "3. Логистика",
                f"{len(logistics_items)} рисков по поставкам",
                "Подтвердить даты или включить альтернативных поставщиков",
            ]
        )
    return focus[:4]


def _apply_status_colors(wb: Workbook) -> None:
    for ws in wb.worksheets:
        ws.freeze_panes = "A4"
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        for row_cells in ws.iter_rows(min_row=1):
            for cell in row_cells:
                if cell.value in {"критично", "не выполнено", "нет отчёта"}:
                    cell.fill = _DANGER_FILL
                elif cell.value in {"частично", "активно", "в работе"}:
                    cell.fill = _WARN_FILL
                elif cell.value in {"выполнено", "смена закрыта"}:
                    cell.fill = _OK_FILL


def build_executive_procurement_report(
    *,
    snapshot: dict[str, Any] | None,
    completion_dashboard: dict[str, Any],
    report_date: str,
    period_mode: str = "day",
    period_label: str = "",
    date_from: str | None = None,
    date_to: str | None = None,
) -> bytes:
    """Return compact .xlsx bytes for executive production/procurement control."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Резюме"

    analyzed_at = _text((snapshot or {}).get("analyzed_at"))
    dashboard_date = _text((snapshot or {}).get("dashboard_date_msk"))
    generated_at = _format_now()
    normalized_mode = period_mode or "day"
    normalized_label = period_label or report_date
    mode_label = {"day": "За день", "range": "За период", "all": "За всё время"}.get(normalized_mode, normalized_mode)

    _sheet_title(
        ws,
        "Управленческая сводка: производство и закупки",
        "Краткий срез для руководителя: KPI, фокус решений, обеспеченность и риски на текущий момент.",
        columns=4,
    )

    meta_rows = [
        ["Сформировано", generated_at, ""],
        ["Режим", mode_label, ""],
        ["Период работы менеджеров", normalized_label, "Агрегация смен менеджеров"],
        ["Срез обеспеченности", dashboard_date or report_date, "Текущее состояние производства"],
        ["Последний анализ", analyzed_at, ""],
    ]
    if normalized_mode != "day":
        meta_rows.insert(3, ["Диапазон", f"{date_from or report_date} — {date_to or report_date}", ""])

    row = _write_table(ws, 4, ["Параметр", "Значение", "Комментарий"], meta_rows, widths=[28, 34, 52])
    row = _write_table(
        ws,
        row,
        ["KPI", "Значение", "Как читать"],
        _kpi_rows(completion_dashboard, snapshot),
        widths=[28, 34, 52],
    )
    _write_table(
        ws,
        row,
        ["Фокус руководителя", "Ситуация", "Действие"],
        _management_focus(completion_dashboard, snapshot),
        widths=[22, 48, 44],
    )

    ws = wb.create_sheet("Менеджеры")
    subtitle = (
        "Сводная эффективность за период без детализации каждой задачи."
        if normalized_mode != "day"
        else "Сводная эффективность смены."
    )
    _sheet_title(ws, "Работа менеджеров", subtitle, columns=11)
    manager_headers = _manager_scorecard_headers(completion_dashboard)
    manager_widths = [22, 18, 16]
    if _text(completion_dashboard.get("period_mode") or "day") != "day":
        manager_widths.append(12)
    manager_widths.extend([10, 10, 10, 8, 10, 10, 18])
    _write_table(
        ws,
        4,
        manager_headers,
        _manager_scorecard_rows(completion_dashboard),
        widths=manager_widths,
    )

    ws = wb.create_sheet("Требует решения")
    _sheet_title(
        ws,
        "Приоритетные решения",
        f"Не более {_MAX_ACTION_ITEMS} позиций: открытые задачи, не обеспеченные изделия и логистика.",
        columns=6,
    )
    _write_table(
        ws,
        4,
        ["Источник", "Объект", "Проблема", "Ответственный", "Статус", "Что сделать"],
        _action_rows(completion_dashboard, snapshot),
        widths=[16, 34, 44, 20, 14, 36],
    )

    ws = wb.create_sheet("Обеспеченность")
    _sheet_title(
        ws,
        "Обеспеченность производства",
        "Сводка по дню и неделе + TOP не обеспеченных изделий.",
        columns=8,
    )
    row = _write_table(
        ws,
        4,
        ["Период", "Раздел", "С планом", "Обеспечено", "Частично", "Не обеспечено", "% OK", "План"],
        _coverage_summary_compact(snapshot),
        widths=[14, 16, 10, 12, 10, 14, 10, 12],
    )
    _write_table(
        ws,
        row,
        ["Изделие", "План", "Есть", "Ключевые дефициты"],
        _top_red_products(snapshot),
        widths=[40, 12, 12, 56],
    )

    logistics_rows = _top_logistics_risks(snapshot)
    if logistics_rows:
        ws = wb.create_sheet("Логистика")
        _sheet_title(
            ws,
            "Критичные поставки",
            f"TOP-{len(logistics_rows)} логистических рисков.",
            columns=6,
        )
        _write_table(
            ws,
            4,
            ["Этап", "Номенклатура", "Поставщик", "Кол-во", "Дата Москва", "Контроль"],
            logistics_rows,
            widths=[18, 42, 24, 12, 14, 14],
        )

    _apply_status_colors(wb)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
