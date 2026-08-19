"""Расчёт потребности в материалах по ресурсным спецификациям 1С."""
from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.onec_resource_spec_sync import get_resource_spec_from_db

MATERIAL_CALCULATOR_XLSX_FILENAME = "potrebnost_materialov.xlsx"

_HEADER_FILL = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _normalize_name(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def _normalize_code(value: str) -> str:
    return _normalize_name(value)


def _resolve_aggregate_key(material: dict[str, Any], code_index: dict[str, str]) -> str:
    """Ключ агрегации: один код/номенклатура — одна строка итога."""
    nom_key = (material.get("nomenclature_key") or "").strip()
    code = (material.get("code") or "").strip()
    name = (material.get("name") or "").strip()
    code_norm = _normalize_code(code) if code else ""

    if code_norm and code_norm in code_index:
        return code_index[code_norm]

    if nom_key:
        key = f"nom:{nom_key}"
    elif code_norm:
        key = f"code:{code_norm}"
    else:
        key = f"name:{_normalize_name(name)}"

    if code_norm:
        code_index[code_norm] = key

    return key


def _round_qty(value: float) -> float:
    if abs(value - round(value)) < 1e-9:
        return float(round(value))
    return round(value, 4)


@dataclass
class MaterialCalculatorInputItem:
    spec_ref_key: str
    quantity: float


@dataclass
class MaterialCalculatorBreakdown:
    spec_ref_key: str
    spec_label: str
    product_qty: float
    material_qty: float


@dataclass
class MaterialCalculatorLine:
    nomenclature_key: str
    code: str
    name: str
    unit: str
    total_qty: float
    breakdown: list[MaterialCalculatorBreakdown] = field(default_factory=list)


@dataclass
class MaterialCalculatorResult:
    ok: bool
    lines: list[MaterialCalculatorLine]
    warnings: list[str]


def _spec_label(spec: dict[str, Any]) -> str:
    parts = [spec.get("code") or "", spec.get("description") or ""]
    main = (spec.get("main_product") or {}).get("name") or ""
    if main:
        parts.append(f"→ {main}")
    return " · ".join(part for part in parts if part)


async def calculate_material_requirements(
    db: AsyncSession,
    items: list[MaterialCalculatorInputItem],
) -> MaterialCalculatorResult:
    warnings: list[str] = []
    aggregate: dict[str, MaterialCalculatorLine] = {}
    code_index: dict[str, str] = {}

    for entry in items:
        ref_key = (entry.spec_ref_key or "").strip()
        product_qty = float(entry.quantity or 0)
        if not ref_key:
            warnings.append("Пропущена строка без спецификации")
            continue
        if product_qty <= 0:
            continue

        spec = await get_resource_spec_from_db(db, ref_key)
        if spec is None:
            warnings.append(f"Спецификация не найдена: {ref_key}")
            continue

        materials = spec.get("materials") or []
        if not materials:
            warnings.append(f"У спецификации «{_spec_label(spec)}» нет материалов")
            continue

        main_product_qty = float((spec.get("main_product") or {}).get("qty") or 0)
        batch_factor = product_qty / main_product_qty if main_product_qty > 0 else product_qty
        label = _spec_label(spec)

        for material in materials:
            name = (material.get("name") or "").strip()
            if not name:
                continue
            if material.get("produced_in_process"):
                continue

            per_batch = float(material.get("qty") or 0)
            if per_batch <= 0:
                continue

            need = _round_qty(batch_factor * per_batch)
            if need <= 0:
                continue

            aggregate_key = _resolve_aggregate_key(material, code_index)
            code = (material.get("code") or "").strip()
            unit = (material.get("unit") or "").strip() or "—"
            nom_key = (material.get("nomenclature_key") or "").strip() or aggregate_key

            line = aggregate.get(aggregate_key)
            if line is None:
                line = MaterialCalculatorLine(
                    nomenclature_key=nom_key,
                    code=code,
                    name=name,
                    unit=unit,
                    total_qty=0.0,
                )
                aggregate[aggregate_key] = line
            elif not line.code and code:
                line.code = code
            if line.unit in {"", "—"} and unit not in {"", "—"}:
                line.unit = unit

            line.total_qty = _round_qty(line.total_qty + need)
            line.breakdown.append(
                MaterialCalculatorBreakdown(
                    spec_ref_key=ref_key,
                    spec_label=label,
                    product_qty=product_qty,
                    material_qty=need,
                )
            )

    lines = sorted(aggregate.values(), key=lambda row: _normalize_name(row.name))
    return MaterialCalculatorResult(ok=True, lines=lines, warnings=warnings)


def material_calculator_to_dict(result: MaterialCalculatorResult) -> dict[str, Any]:
    return {
        "ok": result.ok,
        "warnings": result.warnings,
        "lines": [
            {
                "nomenclature_key": line.nomenclature_key,
                "code": line.code,
                "name": line.name,
                "unit": line.unit,
                "total_qty": line.total_qty,
                "breakdown": [
                    {
                        "spec_ref_key": part.spec_ref_key,
                        "spec_label": part.spec_label,
                        "product_qty": part.product_qty,
                        "material_qty": part.material_qty,
                    }
                    for part in line.breakdown
                ],
            }
            for line in result.lines
        ],
    }


def build_material_calculator_xlsx(result: MaterialCalculatorResult) -> bytes:
    """Формирует Excel-файл с итоговой потребностью в материалах."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Потребность"

    headers = ("Код", "Номенклатура", "Количество", "Ед. изм.")
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    for row_idx, line in enumerate(result.lines, start=2):
        ws.cell(row=row_idx, column=1, value=line.code or "")
        ws.cell(row=row_idx, column=2, value=line.name)
        qty_cell = ws.cell(row=row_idx, column=3, value=line.total_qty)
        qty_cell.alignment = _CENTER
        ws.cell(row=row_idx, column=4, value=line.unit or "—").alignment = _CENTER

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=2).alignment = _LEFT

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def material_calculator_lines_to_result(lines: list[dict[str, Any]]) -> MaterialCalculatorResult:
    """Преобразует строки расчёта из API в результат для выгрузки в Excel."""
    parsed: list[MaterialCalculatorLine] = []
    for row in lines:
        name = (row.get("name") or "").strip()
        if not name:
            continue
        parsed.append(
            MaterialCalculatorLine(
                nomenclature_key=(row.get("nomenclature_key") or name).strip(),
                code=(row.get("code") or "").strip(),
                name=name,
                unit=(row.get("unit") or "").strip() or "—",
                total_qty=float(row.get("total_qty") or 0),
            )
        )
    parsed.sort(key=lambda item: _normalize_name(item.name))
    return MaterialCalculatorResult(ok=True, lines=parsed, warnings=[])
