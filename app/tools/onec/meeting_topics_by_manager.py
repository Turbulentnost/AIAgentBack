"""
Темы совещаний по руководителю из Catalog_ТД_ТемыСовещаний (1С OData).

Справочник содержит поле «Руководитель» — для каждого руководителя свой набор тем.
"""

from __future__ import annotations

from typing import Any
from urllib.parse import quote

import requests

from app.integrations.onec_odata import fetch_all
from app.tools.onec.connection import CONFIG, ODataConfig, create_session
from app.tools.onec.lookup_user_ref import normalize_name, resolve_user_by_fio
from app.tools.onec.meeting_topics_registry import (
    CATALOG_ENTITY,
    build_filter_parts,
    build_list_url,
    normalize_topic,
)

CSV_COLUMNS = (
    ("manager_fio", "Руководитель"),
    ("topic_description", "Тема совещания"),
    ("topic_code", "Код"),
    ("meeting_type", "Вид совещания"),
    ("priority", "Приоритет"),
    ("is_active", "Активна"),
    ("topic_ref_key", "Ref_Key"),
)


def manager_fio_from_topic(topic: dict[str, Any]) -> str:
    manager = (topic.get("manager") or "").strip()
    return manager or "—"


def topic_matches_manager_fio(topic: dict[str, Any], manager_fio: str) -> bool:
    query = normalize_name(manager_fio)
    if not query:
        return True
    candidate = normalize_name(manager_fio_from_topic(topic))
    if not candidate or candidate == normalize_name("—"):
        return False
    return candidate == query or query in candidate or candidate in query


def group_topics_by_manager(topics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}

    for topic in topics:
        manager_fio = manager_fio_from_topic(topic)
        manager_ref = (topic.get("keys") or {}).get("manager")
        group_key = manager_ref or manager_fio

        if group_key not in groups:
            groups[group_key] = {
                "manager_fio": manager_fio,
                "manager_ref_key": manager_ref,
                "topics_count": 0,
                "topics": [],
            }

        groups[group_key]["topics"].append(
            {
                "ref_key": topic.get("ref_key"),
                "code": topic.get("code"),
                "description": topic.get("description"),
                "meeting_type": topic.get("meeting_type"),
                "priority": topic.get("priority"),
                "is_active": topic.get("is_active"),
                "schedule_defined": topic.get("schedule_defined"),
                "department": topic.get("department"),
                "room": topic.get("room"),
            }
        )
        groups[group_key]["topics_count"] += 1

    return sorted(
        groups.values(),
        key=lambda item: (item["manager_fio"] == "—", item["manager_fio"].casefold()),
    )


def flatten_topics_for_export(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for group in groups:
        manager_fio = group.get("manager_fio") or "—"
        for topic in group.get("topics") or []:
            rows.append(
                {
                    "manager_fio": manager_fio,
                    "manager_ref_key": group.get("manager_ref_key"),
                    "topic_description": topic.get("description"),
                    "topic_code": topic.get("code"),
                    "meeting_type": topic.get("meeting_type"),
                    "priority": topic.get("priority"),
                    "is_active": topic.get("is_active"),
                    "topic_ref_key": topic.get("ref_key"),
                }
            )
    return rows


def fetch_all_meeting_topics(
    session: requests.Session,
    config: ODataConfig,
    *,
    active_only: bool = True,
    expand_related: bool = True,
    manager_ref_key: str | None = None,
    page_size: int = 500,
) -> list[dict[str, Any]]:
    filters = build_filter_parts(
        query=None,
        code=None,
        meeting_type=None,
        active_only=active_only,
        ref_key=None,
    )
    if manager_ref_key:
        filters.append(f"Руководитель_Key eq guid'{manager_ref_key}'")

    odata_filter = " and ".join(filters)
    url = build_list_url(
        config,
        odata_filter=odata_filter,
        limit=page_size,
        expand_related=expand_related,
    )
    # build_list_url задаёт $top=limit; для полной выгрузки убираем жёсткий top из URL.
    url = url.replace(f"&$top={page_size}", "")

    rows = fetch_all(session, url, page=page_size, timeout=config.timeout)
    return [normalize_topic(row, expand_related=expand_related) for row in rows]


def export_meeting_topics_by_manager(
    *,
    manager_fio: str | None = None,
    active_only: bool = True,
    expand_related: bool = True,
    config: ODataConfig = CONFIG,
) -> dict[str, Any]:
    session = create_session(config)
    manager_ref_key: str | None = None
    resolved_manager_fio: str | None = None

    if manager_fio and manager_fio.strip():
        try:
            manager_ref_key, resolved_manager_fio, _ = resolve_user_by_fio(
                session,
                manager_fio.strip(),
                config=config,
            )
        except ValueError:
            resolved_manager_fio = manager_fio.strip()

    topics = fetch_all_meeting_topics(
        session,
        config,
        active_only=active_only,
        expand_related=expand_related,
        manager_ref_key=manager_ref_key,
    )

    if manager_fio and manager_fio.strip() and manager_ref_key is None:
        topics = [topic for topic in topics if topic_matches_manager_fio(topic, manager_fio)]

    groups = group_topics_by_manager(topics)
    flat_rows = flatten_topics_for_export(groups)

    return {
        "catalog_entity": CATALOG_ENTITY,
        "filter": {
            "manager_fio": manager_fio,
            "resolved_manager_fio": resolved_manager_fio,
            "manager_ref_key": manager_ref_key,
            "active_only": active_only,
        },
        "managers_count": len(groups),
        "topics_count": len(topics),
        "managers": groups,
        "rows": flat_rows,
    }


def write_topics_csv(path: str, rows: list[dict[str, Any]]) -> None:
    import csv
    from pathlib import Path

    fieldnames = [key for key, _title in CSV_COLUMNS]
    headers = {key: title for key, title in CSV_COLUMNS}

    with Path(path).open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, extrasaction="ignore")
        writer.writerow(headers)
        for row in rows:
            writer.writerow(
                {
                    key: (
                        "Да" if key == "is_active" and row.get(key) is True
                        else "Нет" if key == "is_active" and row.get(key) is False
                        else row.get(key)
                    )
                    for key in fieldnames
                }
            )


def write_topics_xlsx(path: str, rows: list[dict[str, Any]]) -> None:
    from pathlib import Path

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Темы по руководителям"

    headers = [title for _key, title in CSV_COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(
            [
                row.get("manager_fio"),
                row.get("topic_description"),
                row.get("topic_code"),
                row.get("meeting_type"),
                row.get("priority"),
                "Да" if row.get("is_active") else "Нет",
                row.get("topic_ref_key"),
            ]
        )

    for index, (_key, title) in enumerate(CSV_COLUMNS, start=1):
        column = get_column_letter(index)
        max_len = max(len(title), *(len(str(sheet[f"{column}{row}"].value or "")) for row in range(2, sheet.max_row + 1)) if sheet.max_row > 1 else [0])
        sheet.column_dimensions[column].width = min(max_len + 2, 60)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
