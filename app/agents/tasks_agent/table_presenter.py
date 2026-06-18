from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

PORUCHENIYA_TASK_TABLE_COLUMNS: list[dict[str, str]] = [
    {"key": "document_number", "title": "№ протокола / решения / поручения"},
    {"key": "document_date", "title": "Дата"},
    {"key": "task_text", "title": "Задача / решение"},
    {"key": "assignee", "title": "Исполнитель"},
    {"key": "reviewer", "title": "Проверяющий"},
    {"key": "department", "title": "Подразделение"},
    {"key": "due_date", "title": "Срок"},
    {"key": "status", "title": "Статус"},
    {"key": "artifact", "title": "Артефакт"},
    {"key": "overdue_days", "title": "Просрочка, дней"},
    {"key": "overdue_reason", "title": "Причина просрочки"},
    {"key": "postponement_request", "title": "Запрос переноса"},
    {"key": "postponement_basis", "title": "Основание переноса"},
    {"key": "controller_action", "title": "Действие контролера"},
    {"key": "rk_required", "title": "Требуется РК"},
]

_STATUS_LABELS = {
    "ВРаботе": "В работе",
    "Выполнено": "Выполнено",
    "Закрыто": "Закрыто",
    "Черновик": "Черновик",
}


def format_display_date(value: str | date | datetime | None) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime(value.year, value.month, value.day)
    else:
        normalized = str(value).strip()
        if not normalized or normalized.startswith("0001-01-01"):
            return ""
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00")[:19])
    return parsed.strftime("%d.%m.%Y")


def compute_overdue_days(due_date: str | None, *, now: datetime) -> int | None:
    if not due_date:
        return None
    normalized = str(due_date).strip()
    if not normalized or normalized.startswith("0001-01-01"):
        return None
    parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00")[:19])
    delta = (now.date() - parsed.date()).days
    return delta if delta > 0 else None


def format_task_status(document_status: str | None, *, overdue_days: int | None) -> str:
    if overdue_days is not None:
        return "Просрочено"
    if document_status:
        return _STATUS_LABELS.get(document_status, document_status)
    return ""


def build_porucheniya_task_row(
    document: dict[str, Any],
    task: dict[str, Any],
    *,
    now: datetime | None = None,
) -> dict[str, Any]:
    current = (now or datetime.now()).replace(microsecond=0)
    overdue_days = compute_overdue_days(task.get("due_date"), now=current)

    return {
        "document_number": document.get("document_number") or "",
        "document_date": format_display_date(document.get("document_date")),
        "task_text": task.get("activity") or "",
        "assignee": task.get("responsible") or "",
        "reviewer": document.get("reviewer") or "",
        "department": task.get("department") or document.get("department") or "",
        "due_date": format_display_date(task.get("due_date")),
        "status": format_task_status(document.get("status"), overdue_days=overdue_days),
        "artifact": task.get("has_file") or "",
        "overdue_days": overdue_days if overdue_days is not None else "",
        "overdue_reason": task.get("overdue_reason") or "",
        "postponement_request": task.get("postponement_request") or "",
        "postponement_basis": task.get("postponement_basis") or "",
        "controller_action": task.get("controller_action") or "",
        "rk_required": task.get("rk_required") or "",
        "_meta": {
            "document_ref": document.get("document_ref"),
            "line_number": task.get("line_number"),
            "priority": task.get("priority"),
            "manager": document.get("manager"),
        },
    }


def build_tasks_table(
    porucheniya: list[dict[str, Any]],
    protocols: list[dict[str, Any]] | None = None,
    *,
    now: datetime | None = None,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for document in porucheniya:
        for task in document.get("tasks") or []:
            rows.append(build_porucheniya_task_row(document, task, now=now))
    for document in protocols or []:
        for task in document.get("tasks") or []:
            rows.append(build_porucheniya_task_row(document, task, now=now))

    return {
        "columns": PORUCHENIYA_TASK_TABLE_COLUMNS,
        "rows": rows,
        "row_count": len(rows),
    }


def build_porucheniya_tasks_table(
    porucheniya: list[dict[str, Any]],
    *,
    now: datetime | None = None,
) -> dict[str, Any]:
    return build_tasks_table(porucheniya, protocols=None, now=now)


def write_tasks_table_xlsx(path: Path | str, table: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    columns = table.get("columns") or PORUCHENIYA_TASK_TABLE_COLUMNS
    rows = table.get("rows") or []
    keys = [column["key"] for column in columns]
    titles = [column["title"] for column in columns]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Задачи"

    header_font = Font(bold=True)
    for col_idx, title in enumerate(titles, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=title)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            value = row.get(key, "")
            sheet.cell(row=row_idx, column=col_idx, value="" if value is None else value)

    for col_idx, key in enumerate(keys, start=1):
        max_len = len(titles[col_idx - 1])
        for row in rows[:200]:
            max_len = max(max_len, len(str(row.get(key, ""))))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    sheet.freeze_panes = "A2"
    workbook.save(Path(path))
