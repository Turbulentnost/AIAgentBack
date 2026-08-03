from __future__ import annotations

import asyncio
import os
import time
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Literal

from fastapi.encoders import jsonable_encoder
from langgraph.checkpoint.memory import MemorySaver
from langgraph.types import Command
from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.agents.procurement_manager_agent.allocation import allocate_materials_by_deadline
from app.agents.procurement_manager_agent.batches import sync_batches_workspace
from app.agents.procurement_manager_agent.delivery_schedule import compute_schedule
from app.agents.procurement_manager_agent.documents import (
    render_purchase_order_draft,
    render_rfq_draft,
)
from app.agents.procurement_manager_agent.fulfillment import (
    FULFILLMENT_LABELS,
    fulfillment_payload,
)
from app.agents.procurement_manager_agent.orchestrator_coverage import (
    coverage_snapshot_from_metadata,
    merge_order_coverage_with_orchestrator,
    search_fields_from_coverage,
)
from app.agents.procurement_manager_agent.graph import build_graph
from app.agents.procurement_manager_agent.material_bank import get_material_bank
from app.agents.procurement_manager_agent.operations import MutationGate
from app.agents.procurement_manager_agent.pricing import (
    AMOUNT_FORMULA,
    estimate_nomenclature_amount,
    supplier_price_bounds,
)
from app.agents.procurement_manager_agent.schemas import (
    AgentResumeRequest,
    AgentRunRequest,
    AgentStatus,
    AllPositionsResponse,
    AllPositionsRow,
    ApprovalRecord,
    ApprovalRequest,
    ComparisonWeights,
    FulfillmentStatusUpdateRequest,
    LineAmountEntry,
    LineAmountsUpdateRequest,
    LineScheduleUpdateRequest,
    MaterialBankResponse,
    NomenclatureSearchItem,
    NomenclatureSupplierResult,
    NonconformityRequest,
    OperationStatus,
    PurchaseOrderDraft,
    PurchaseOrderDraftRequest,
    QuoteComparison,
    QuoteSubmission,
    RecommendationRecord,
    RecommendationRequest,
    RFQDraft,
    RFQDraftRequest,
    RFQLine,
    ShipmentEventRequest,
    StrategyResumeRequest,
    StrategyRunRequest,
    StrategyStatus,
    Supplier,
    SupplierOffersResponse,
    SupplierQuote,
    SupplierSearchRequest,
    SupplierSearchResult,
    TopSupplierOffer,
    UsedSupplierPart,
    WorkspaceSummary,
)
from app.agents.procurement_manager_agent.strategy_graph import build_strategy_graph
from app.agents.procurement_manager_agent.supplier_ranking import (
    SCORE_FORMULA,
    collect_supplier_offers,
    price_bounds_from_offers,
    rank_supplier_offers,
)
from app.agents.procurement_manager_agent.scoring import compare_quotes
from app.agents.procurement_manager_agent.suppliers import (
    MIN_SUPPLIERS_BEFORE_SKIP,
    WEB_LIMIT_PER_NOMENCLATURE,
    HybridSupplierSearchService,
    qualifying_suppliers_for_skip,
)
from app.agents.procurement_manager_agent.search_progress import (
    emit_progress,
    finish_progress,
    get_progress,
    get_progress_meta,
    progress_scope,
    truncate_query,
)
from app.agents.procurement_manager_agent.web_page_enrichment import enrich_web_suppliers
from app.agents.procurement_manager_agent.web_qwen import qwen_agent_enabled
from app.agents.procurement_role_agents.schemas import (
    ProcurementRoleAgentRequest,
    ProcurementRoleAgentResult,
)
from app.core.logging import get_logger
from app.models.enums import ConfidenceLevel, ProcurementCaseStatus, TaskStatus
from app.models.procurement import ProcurementCase, ProcurementCaseEvent
from app.models.task import Task
from app.services.procurement_case_statuses import ACTIVE_CASE_STATUSES

logger = get_logger(__name__)

AGENT_ID = "purchase_manager_agent"
METADATA_KEY = "procurement_manager"

# Same statuses as the left-queue filter in /dashboard.
# Must NOT include engineer-stage statuses (human_required, agent_waiting, …) —
# those inflated KPI to hundreds of DB cases outside the manager list.
MANAGER_QUEUE_STATUSES = frozenset(
    {
        ProcurementCaseStatus.PURCHASE_DRAFT.value,
        ProcurementCaseStatus.APPROVAL_REQUIRED.value,
        ProcurementCaseStatus.ORDERED.value,
        ProcurementCaseStatus.PAYMENT_PENDING.value,
        ProcurementCaseStatus.IN_TRANSIT.value,
        ProcurementCaseStatus.RECEIVING.value,
        ProcurementCaseStatus.POSTING_REQUIRED.value,
        ProcurementCaseStatus.POSTED.value,
        ProcurementCaseStatus.NONCONFORMITY.value,
    }
)


def case_in_manager_queue(
    *,
    current_agent_id: str | None,
    status: str | None,
    purchase_manager_invoked_at: str | None = None,
    supplier_coverage_status: str | None = None,
) -> bool:
    """Queue = cases already in purchasing (supplier coverage partial/full).

    Status-only MANAGER_QUEUE_STATUSES is intentionally not used: on main the
    orchestrator hands off via supplier_order_coverage, not engineer route.
    """
    _ = status
    if supplier_coverage_status in {"partial", "full"}:
        return True
    # Keep current PM owner briefly if coverage snapshot is mid-refresh.
    return current_agent_id == AGENT_ID and bool(purchase_manager_invoked_at)


# In-app graph with MemorySaver for HITL interrupt/resume (not used by Studio).
_runtime_graph = build_graph(checkpointer=MemorySaver())
_strategy_graph = build_strategy_graph(checkpointer=MemorySaver())

# Process-local last strategy artifact (also persisted onto each case workspace).
_STRATEGY_ARTIFACT: dict[str, Any] = {}
STRATEGY_METADATA_KEY = "supply_policy"
STRATEGY_GRAPH_KEY = "strategy_graph"

# Coalesce dashboard / summary / case-detail / all-positions allocation work.
# Without this, page load fires 3× full-queue allocate + DB case loads and can
# exhaust the asyncpg pool (UI stuck on «Загрузка позиций...»).
_ALLOCATION_TTL_SEC = 20.0
_allocation_lock = asyncio.Lock()
_allocation_cache: dict[str, Any] = {"ts": 0.0, "payload": None}


def invalidate_allocation_cache() -> None:
    """Drop cached queue allocation (call after coverage-affecting mutations)."""
    _allocation_cache["payload"] = None
    _allocation_cache["ts"] = 0.0


class ProcurementManagerService:
    def __init__(
        self,
        db: AsyncSession,
        *,
        supplier_search: HybridSupplierSearchService | None = None,
        use_graph: bool = True,
    ) -> None:
        self.db = db
        self.supplier_search = supplier_search or HybridSupplierSearchService()
        self.use_graph = use_graph

    @staticmethod
    def _env_or_setting_float(env_key: str, setting_attr: str, default: float) -> float:
        raw = os.environ.get(env_key)
        if raw is not None and str(raw).strip():
            try:
                return float(raw)
            except ValueError:
                pass
        try:
            from app.core.config import settings

            value = getattr(settings, setting_attr, None)
            if value is not None and str(value).strip() != "":
                return float(value)
        except Exception:
            pass
        return default

    @classmethod
    def _search_timeout_seconds(cls, *, default: float = 90.0) -> float:
        return cls._env_or_setting_float(
            "PROCUREMENT_MANAGER_SEARCH_TIMEOUT_SECONDS",
            "PROCUREMENT_MANAGER_SEARCH_TIMEOUT_SECONDS",
            default,
        )

    @classmethod
    def _force_web_timeout_seconds(cls, n_items: int) -> float:
        """Outer wait_for for manual «Найти поставщиков»."""
        configured_web = cls._env_or_setting_float(
            "PROCUREMENT_MANAGER_WEB_SEARCH_TIMEOUT_SECONDS",
            "PROCUREMENT_MANAGER_WEB_SEARCH_TIMEOUT_SECONDS",
            0.0,
        )
        if configured_web > 0:
            return configured_web
        configured = cls._search_timeout_seconds(default=180.0)
        items = max(1, n_items)
        if qwen_agent_enabled():
            # SERP + select URLs + page fetch + Qwen extract per nomenclature.
            scaled = 90.0 + 55.0 * items
            return max(configured, 240.0, min(480.0, scaled))
        scaled = 60.0 + 40.0 * items
        return max(configured, 180.0, min(300.0, scaled))

    async def run_role(self, payload: dict[str, Any]) -> ProcurementRoleAgentResult:
        request = ProcurementRoleAgentRequest.model_validate(payload)
        case = await self._case(uuid.UUID(request.case_id))
        if case is None:
            return ProcurementRoleAgentResult(
                agent_id=AGENT_ID,
                status="failed",
                summary="Закупочный кейс не найден.",
                data_confidence=ConfidenceLevel.LOW,
                requires_human_review=True,
                case_id=request.case_id,
                correlation_id=request.correlation_id,
                role_status="failed",
                output_data={},
            )

        metadata = self._workspace(case)
        metadata["lifecycle_state"] = "agent_running"
        metadata["handoff_received_at"] = datetime.now(UTC).isoformat()
        metadata.setdefault("payment_document_draft", None)
        metadata.setdefault("recommendation_audit", [])
        metadata.setdefault("purchase_order_drafts", [])
        self._save_workspace(case, metadata)
        await self._event(
            case,
            "procurement_manager_handoff_received",
            f"{request.idempotency_key}:manager-handoff",
            {"requested_operation": case.requested_operation},
        )
        try:
            agent_status = await self.agent_run(
                case.id,
                AgentRunRequest(
                    idempotency_key=f"auto-agent-run:{request.idempotency_key}"[:255],
                    allow_web_fallback=True,
                ),
            )
            lifecycle = agent_status.stage or "agent_running"
        except Exception as exc:  # noqa: BLE001 — handoff must not fail on agent start
            logger.warning(
                "procurement_manager_auto_agent_run_failed case_id=%s error=%s",
                case.id,
                str(exc)[:500],
            )
            agent_status = None
            lifecycle = "agent_running"
        metadata = self._workspace(case)
        metadata["lifecycle_state"] = (
            "approval_required"
            if agent_status and agent_status.paused_for_human
            else lifecycle
        )
        self._save_workspace(case, metadata)
        return ProcurementRoleAgentResult(
            agent_id=AGENT_ID,
            status="waiting_human",
            summary=(
                "Дефицит передан менеджеру по закупкам. "
                "Агент запущен: поиск, оценка и черновик заказа с HITL."
            ),
            data_confidence=ConfidenceLevel.MEDIUM,
            requires_human_review=True,
            case_id=request.case_id,
            correlation_id=request.correlation_id,
            role_status="waiting_human",
            wait_reason="Требуется подтверждение shortlist/заказа агентом менеджера.",
            output_data={
                "next_status": ProcurementCaseStatus.PURCHASE_DRAFT.value,
                "lifecycle_state": metadata["lifecycle_state"],
                "agent_stage": agent_status.stage if agent_status else None,
                "paused_for_human": bool(agent_status.paused_for_human)
                if agent_status
                else False,
                "payment_execution_allowed": False,
            },
        )

    async def search_suppliers(
        self,
        case_id: uuid.UUID,
        request: SupplierSearchRequest,
    ) -> SupplierSearchResult:
        """Compatibility wrapper (holds DB for the whole call). Prefer phased API."""
        prepared = await self.prepare_supplier_search(case_id, request)
        if isinstance(prepared, SupplierSearchResult):
            return prepared
        result = await self.execute_supplier_search_web(prepared)
        return await self.finalize_supplier_search(prepared, result)

    async def prepare_supplier_search(
        self,
        case_id: uuid.UUID,
        request: SupplierSearchRequest,
    ) -> SupplierSearchResult | dict[str, Any]:
        """Mark search running and return a DB-free prep payload (or replay result)."""
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        targets = request.nomenclatures or self._nomenclature_targets_from_case(
            case, metadata
        )
        query = request.query or self._supplier_query_from_case(case)
        if targets:
            query = ", ".join(
                dict.fromkeys(
                    (item.query or item.nomenclature_name or item.nomenclature_id or "")
                    for item in targets
                    if (item.query or item.nomenclature_name or item.nomenclature_id)
                )
            )[:500] or query
        category = request.category or self._supplier_category_from_case(case)
        force_web = request.is_manual_web
        # Skip only when ≥3 qualifying suppliers exist (manual: 1c/web+URL;
        # bank seeds without links never block «Найти поставщиков»).
        # force_web always searches (web-only enriched cards).
        to_search: list[NomenclatureSearchItem] = []
        skipped: list[NomenclatureSupplierResult] = []
        for item in targets:
            if force_web:
                to_search.append(item)
                continue
            qualifying = qualifying_suppliers_for_skip(
                list(item.existing_suppliers or []),
                force_web=force_web,
            )
            if len(qualifying) >= MIN_SUPPLIERS_BEFORE_SKIP:
                skipped.append(
                    NomenclatureSupplierResult(
                        nomenclature_id=item.nomenclature_id,
                        nomenclature_name=item.nomenclature_name,
                        query=self._clean_supplier_search_query(item.query or item.nomenclature_name) or (item.query or item.nomenclature_name or "поставщик"),
                        suppliers=item.existing_suppliers[: request.limit],
                        sources_used=["existing"],
                        web_fallback_used=False,
                    )
                )
            else:
                to_search.append(item)
        idem_prefix = "supplier-search-manual" if force_web else "supplier-search"
        effective_request = request.model_copy(
            update={
                "query": query,
                "category": category,
                "nomenclatures": to_search,
                "force_web": force_web,
                "mode": "manual_web" if force_web else request.mode,
                "allow_web_fallback": True if force_web else request.allow_web_fallback,
                "idempotency_key": request.idempotency_key
                or f"{idem_prefix}:{case.id}:{query.casefold()}"[:255],
            }
        )
        operation_id = str(effective_request.idempotency_key)
        previous = metadata.get("supplier_searches") or []
        replay = next(
            (
                item
                for item in previous
                if item.get("idempotency_key") == effective_request.idempotency_key
            ),
            None,
        )
        if replay:
            return SupplierSearchResult.model_validate(replay["result"])

        self._upsert_operation(
            case,
            operation_id=operation_id,
            operation="supplier_search",
            status="running",
        )
        if force_web:
            timeout_seconds = self._force_web_timeout_seconds(len(to_search))
        else:
            timeout_seconds = self._search_timeout_seconds(default=90.0)
        # Client alarm-clock budget shortens the outer wait_for (never extends past server max).
        if request.timeout_seconds is not None:
            client_budget = max(30.0, min(600.0, float(request.timeout_seconds)))
            timeout_seconds = min(timeout_seconds, client_budget)
            effective_request = effective_request.model_copy(
                update={"timeout_seconds": client_budget}
            )
        return {
            "case_id": str(case.id),
            "operation_id": operation_id,
            "effective_request": effective_request,
            "timeout_seconds": timeout_seconds,
            "to_search": to_search,
            "skipped": skipped,
            "targets": targets,
            "query": query,
            "force_web": force_web,
        }

    async def execute_supplier_search_web(
        self,
        prepared: dict[str, Any],
    ) -> SupplierSearchResult:
        """Run browser/Qwen search without holding a DB session."""
        operation_id = str(prepared["operation_id"])
        case_id = str(prepared["case_id"])
        query = str(prepared["query"] or "")
        timeout_seconds = float(prepared["timeout_seconds"])
        force_web = bool(prepared["force_web"])
        to_search: list[NomenclatureSearchItem] = list(prepared["to_search"] or [])
        skipped: list[NomenclatureSupplierResult] = list(prepared["skipped"] or [])
        targets: list[NomenclatureSearchItem] = list(prepared["targets"] or [])
        effective_request: SupplierSearchRequest = prepared["effective_request"]

        with progress_scope(operation_id, case_id=case_id):
            emit_progress(
                f"Ищу поставщиков: {truncate_query(query or 'номенклатура')}"
            )
            try:
                if to_search:
                    result = await asyncio.wait_for(
                        self.supplier_search.search(effective_request),
                        timeout=timeout_seconds,
                    )
                else:
                    result = SupplierSearchResult(
                        query=query,
                        suppliers=[],
                        sources_used=["existing"] if skipped else [],
                        web_fallback_used=False,
                        nomenclature_results=[],
                    )
                result = self._merge_nomenclature_search_results(
                    result,
                    skipped=skipped,
                    targets=targets,
                    query=query,
                )
                failed = (
                    result.status == "failed"
                    or (
                        force_web
                        and not result.suppliers
                        and not any(row.suppliers for row in result.nomenclature_results)
                    )
                )
                return result.model_copy(
                    update={
                        "operation_id": operation_id,
                        "pending": False,
                        "status": "failed" if failed else "completed",
                        "message": result.message
                        or (
                            "Веб-поиск не вернул поставщиков"
                            if failed and force_web
                            else result.message
                        ),
                    }
                )
            except TimeoutError:
                emit_progress(
                    f"Время поиска истекло ({timeout_seconds:.0f}с) — останавливаю"
                )
                finish_progress(operation_id, status="failed")
                return SupplierSearchResult(
                    query=query,
                    suppliers=[],
                    sources_used=[],
                    web_fallback_used=False,
                    nomenclature_results=[],
                    operation_id=operation_id,
                    pending=False,
                    status="failed",
                    message=(
                        f"Время поиска истекло ({timeout_seconds:.0f}с). "
                        "Повторите поиск (кнопка «Найти поставщиков»)."
                    ),
                    diagnostics={
                        "timeout_seconds": timeout_seconds,
                        "status": "timeout",
                    },
                )
            except Exception as exc:
                finish_progress(operation_id, status="failed")
                prepared["execute_error"] = str(exc)[:1000]
                raise

    async def finalize_supplier_search(
        self,
        prepared: dict[str, Any],
        result: SupplierSearchResult | None = None,
        *,
        execute_error: str | None = None,
    ) -> SupplierSearchResult:
        """Persist search outcome on a fresh short-lived DB session."""
        case = await self.require_case(uuid.UUID(str(prepared["case_id"])))
        operation_id = str(prepared["operation_id"])
        effective_request: SupplierSearchRequest = prepared["effective_request"]
        timeout_seconds = float(prepared["timeout_seconds"])
        to_search: list[NomenclatureSearchItem] = list(prepared["to_search"] or [])
        skipped: list[NomenclatureSupplierResult] = list(prepared["skipped"] or [])
        err = execute_error or prepared.get("execute_error")
        if err:
            self._upsert_operation(
                case,
                operation_id=operation_id,
                operation="supplier_search",
                status="failed",
                error=str(err)[:1000],
            )
            raise RuntimeError(str(err))

        assert result is not None
        is_timeout = (result.diagnostics or {}).get("status") == "timeout"
        if is_timeout:
            self._upsert_operation(
                case,
                operation_id=operation_id,
                operation="supplier_search",
                status="failed",
                error=(
                    f"Supplier search exceeded {timeout_seconds:.0f}s request timeout; "
                    "use GET operations/{operation_id} and retry with a new idempotency key."
                ),
            )
            metadata = self._workspace(case)
            previous = list(metadata.get("supplier_searches") or [])
            previous.append(
                {
                    "idempotency_key": effective_request.idempotency_key,
                    "at": datetime.now(UTC).isoformat(),
                    "result": result.model_dump(mode="json"),
                }
            )
            metadata["supplier_searches"] = previous[-20:]
            metadata["lifecycle_state"] = "supplier_search_timeout"
            self._save_workspace(case, metadata)
            await self._event(
                case,
                "supplier_search_timeout",
                f"supplier-search-timeout:{operation_id}",
                {"operation_id": operation_id, "timeout_seconds": timeout_seconds},
            )
            return result

        metadata = self._workspace(case)
        previous = list(metadata.get("supplier_searches") or [])
        previous.append(
            {
                "idempotency_key": effective_request.idempotency_key,
                "at": datetime.now(UTC).isoformat(),
                "result": result.model_dump(mode="json"),
            }
        )
        metadata["supplier_searches"] = previous
        metadata["nomenclature_results"] = [
            item.model_dump(mode="json") for item in result.nomenclature_results
        ]
        metadata["suppliers"] = [
            item.model_dump(mode="json") for item in result.suppliers
        ]
        metadata["lifecycle_state"] = "suppliers_identified"
        self._save_workspace(case, metadata)
        self._upsert_operation(
            case,
            operation_id=operation_id,
            operation="supplier_search",
            status="completed",
        )
        finish_progress(
            operation_id,
            status="failed" if result.status == "failed" else "completed",
        )
        await self._event(
            case,
            "supplier_search_completed",
            str(effective_request.idempotency_key),
            {
                "count": len(result.suppliers),
                "nomenclature_count": len(result.nomenclature_results),
                "searched_count": len(to_search),
                "skipped_count": len(skipped),
                "web_limit_per_nomenclature": WEB_LIMIT_PER_NOMENCLATURE,
                "sources": result.sources_used,
                "operation_id": operation_id,
            },
        )
        return result

    async def enrich_web_supplier_cards(
        self,
        case_id: uuid.UUID,
    ) -> SupplierSearchResult:
        """Re-fetch product pages for web suppliers already stored in nomenclature_results."""
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        rows_raw = list(metadata.get("nomenclature_results") or [])
        if not rows_raw:
            latest = (metadata.get("supplier_searches") or [{}])[-1]
            rows_raw = list((latest.get("result") or {}).get("nomenclature_results") or [])
        provider = getattr(self.supplier_search, "_fetch_provider", lambda: None)()
        enriched_rows: list[NomenclatureSupplierResult] = []
        flat: list[Supplier] = []
        for row in rows_raw:
            if not isinstance(row, dict):
                continue
            suppliers = [
                Supplier.model_validate(item)
                for item in (row.get("suppliers") or [])
                if isinstance(item, dict)
            ]
            web_only = [item for item in suppliers if item.source == "web"]
            row_query = str(row.get("query") or row.get("nomenclature_name") or "поставщик")
            if provider is not None and web_only:
                web_only = await enrich_web_suppliers(
                    web_only,
                    provider,
                    product_query=row_query,
                )
            result_row = NomenclatureSupplierResult(
                nomenclature_id=row.get("nomenclature_id"),
                nomenclature_name=row.get("nomenclature_name"),
                query=row_query,
                suppliers=web_only,
                sources_used=["web"] if web_only else list(row.get("sources_used") or []),
                web_fallback_used=True,
            )
            enriched_rows.append(result_row)
            flat.extend(web_only)
        result = SupplierSearchResult(
            query=", ".join(
                dict.fromkeys(row.query for row in enriched_rows if row.query)
            )[:500]
            or "поставщик",
            suppliers=flat,
            sources_used=["web"] if flat else [],
            web_fallback_used=True,
            nomenclature_results=enriched_rows,
            status="completed",
        )
        metadata["nomenclature_results"] = [
            item.model_dump(mode="json") for item in enriched_rows
        ]
        metadata["suppliers"] = [item.model_dump(mode="json") for item in flat]
        self._save_workspace(case, metadata)
        return result

    async def agent_run(
        self,
        case_id: uuid.UUID,
        request: AgentRunRequest | None = None,
    ) -> AgentStatus:
        """Idempotent start of the full search → evaluate → PO draft graph."""
        case = await self.require_case(case_id)
        payload = request or AgentRunRequest()
        idempotency_key = (
            payload.idempotency_key or f"agent-run:{case.id}:{datetime.now(UTC).date()}"
        )[:255]
        workspace = self._workspace(case)
        if workspace.get("agent_run_idempotency_key") == idempotency_key and workspace.get(
            "agent_stage"
        ):
            return await self.agent_status(case_id)

        targets = self._nomenclature_targets_from_case(case, workspace)
        query = payload.query or self._supplier_query_from_case(case)
        if targets:
            query = ", ".join(
                dict.fromkeys(
                    (item.query or item.nomenclature_name or item.nomenclature_id or "")
                    for item in targets
                    if (item.query or item.nomenclature_name or item.nomenclature_id)
                )
            )[:500] or query
        # Pass all targets; HybridSupplierSearchService skips items with ≥3 existing.
        search_request = SupplierSearchRequest(
            query=query,
            category=self._supplier_category_from_case(case),
            allow_web_fallback=payload.allow_web_fallback,
            idempotency_key=f"agent-search:{idempotency_key}"[:255],
            nomenclatures=targets,
        )
        positions = [
            {
                "line_id": position.line_id,
                "nomenclature_id": position.nomenclature_id,
                "nomenclature_name": position.nomenclature_name,
                "quantity": str(position.quantity),
                "unit": position.unit or "шт",
                "required_date": (
                    position.required_date.isoformat() if position.required_date else None
                ),
            }
            for position in case.positions or []
            if not position.cancelled
        ]
        thread_id = f"procurement-manager:{case.id}:{idempotency_key}"
        config = {
            "configurable": {
                "thread_id": thread_id,
                "runtime": self.supplier_search,
            }
        }
        case_context = {
            **workspace,
            "positions": positions,
            "lines": positions,
            "required_date": case.required_date.isoformat() if case.required_date else None,
        }
        result = await _runtime_graph.ainvoke(
            {
                "case_id": str(case.id),
                "case_number": case.source_number or str(case.id),
                "case_context": case_context,
                "positions": positions,
                "request": search_request.model_dump(mode="json"),
            },
            config=config,
        )
        workspace = self._workspace(case)
        workspace["agent_run_idempotency_key"] = idempotency_key
        workspace["agent_thread_id"] = thread_id
        self._save_workspace(case, workspace)
        await self._persist_graph_state(case, result, "agent_run_paused")
        return await self.agent_status(case_id)

    async def agent_resume(
        self,
        case_id: uuid.UUID,
        request: AgentResumeRequest | dict[str, Any],
    ) -> AgentStatus:
        """HITL resume for shortlist / order draft approval (or reject)."""
        case = await self.require_case(case_id)
        if isinstance(request, AgentResumeRequest):
            decision = request.model_dump(mode="json")
        else:
            decision = dict(request)
            AgentResumeRequest.model_validate(decision)
        action = str(decision.get("action") or "")
        workspace = self._workspace(case)
        resume_key = decision.get("idempotency_key") or (
            f"agent-resume:{case.id}:{action}:{workspace.get('agent_stage')}"
        )
        resume_key = str(resume_key)[:255]
        prior = workspace.get("agent_resume_keys") or []
        if resume_key in prior:
            return await self.agent_status(case_id)

        thread_id = workspace.get("agent_thread_id") or f"procurement-manager:{case.id}"
        config = {
            "configurable": {
                "thread_id": thread_id,
                "runtime": self.supplier_search,
            }
        }
        result = await self._invoke_agent_resume(case, decision, config)
        workspace = self._workspace(case)
        prior = list(workspace.get("agent_resume_keys") or [])
        prior.append(resume_key)
        workspace["agent_resume_keys"] = prior[-50:]
        self._save_workspace(case, workspace)
        await self._persist_graph_state(case, result, "agent_resumed")
        return await self.agent_status(case_id)

    async def _invoke_agent_resume(
        self,
        case: ProcurementCase,
        decision: dict[str, Any],
        config: dict[str, Any],
    ) -> dict[str, Any]:
        """Resume HITL via live MemorySaver checkpoint, or rehydrate from workspace.

        MemorySaver is process-local: after a backend restart the UI still shows the
        persisted interrupt, but Command(resume=...) hits an empty thread and the graph
        restarts from entry → KeyError('request'). Rehydrate from supplier_graph.
        """
        live = await _runtime_graph.aget_state(config)
        if live.values and live.next:
            try:
                return await _runtime_graph.ainvoke(Command(resume=decision), config=config)
            except KeyError as exc:
                if exc.args != ("request",):
                    raise
                logger.warning(
                    "procurement_manager agent_resume live checkpoint missing request; "
                    "rehydrating from workspace case_id=%s",
                    case.id,
                )
        else:
            workspace = self._workspace(case)
            if not workspace.get("paused_for_human") and not (
                workspace.get("agent_interrupt") or {}
            ):
                raise ValueError("Нет активного HITL-прерывания для возобновления агента")
            logger.info(
                "procurement_manager agent_resume rehydrate after lost checkpoint "
                "case_id=%s thread_id=%s",
                case.id,
                (config.get("configurable") or {}).get("thread_id"),
            )
        return await self._resume_rehydrated(case, decision, config)

    async def _resume_rehydrated(
        self,
        case: ProcurementCase,
        decision: dict[str, Any],
        config: dict[str, Any],
    ) -> dict[str, Any]:
        """Apply HITL decision using persisted supplier_graph, then continue the graph."""
        workspace = self._workspace(case)
        interrupt_payload = workspace.get("agent_interrupt") or {}
        interrupt_type = str(
            interrupt_payload.get("type")
            if isinstance(interrupt_payload, dict)
            else ""
        )
        action = str(decision.get("action") or "")
        graph_state = {
            key: value
            for key, value in dict(workspace.get("supplier_graph") or {}).items()
            if key not in {"paused_for_human", "runtime", "__interrupt__"}
        }
        if not graph_state.get("case_id"):
            graph_state["case_id"] = str(case.id)
        if not graph_state.get("case_number"):
            graph_state["case_number"] = case.source_number or str(case.id)
        if not graph_state.get("request"):
            raise ValueError(
                "Состояние агента потеряно после перезапуска сервера "
                "(нет checkpoint и нет request в workspace). Запустите агента заново."
            )

        is_order = (
            interrupt_type == "procurement_order_approval" or "order" in interrupt_type
        )
        is_shortlist = (
            interrupt_type == "procurement_shortlist_approval"
            or "shortlist" in interrupt_type
            or "rfq" in interrupt_type
        )
        if is_order:
            if action not in {"approve_order_draft", "reject"}:
                raise ValueError(
                    "Для подтверждения заказа нужно approve_order_draft или reject"
                )
            approved = action == "approve_order_draft"
            patch = {
                **graph_state,
                "order_approval": dict(decision),
                "status": "order_draft_approved" if approved else "order_rejected",
                "stage": "await_order_hitl",
            }
            as_node = "await_order_hitl"
        elif is_shortlist:
            if action not in {"approve_shortlist", "approve_rfq_draft", "reject"}:
                raise ValueError(
                    "Для shortlist нужно approve_shortlist, approve_rfq_draft или reject"
                )
            approved = action in {"approve_shortlist", "approve_rfq_draft"}
            flags = dict(graph_state.get("kpi_flags") or {})
            if approved:
                flags["supplier_confirmed"] = True
            patch = {
                **graph_state,
                "shortlist_approval": dict(decision),
                "kpi_flags": flags,
                "status": "shortlist_approved" if approved else "rejected",
                "stage": "await_supplier_hitl",
            }
            as_node = "await_supplier_hitl"
        else:
            raise ValueError(
                f"Неизвестный тип HITL для восстановления: {interrupt_type or '—'}"
            )

        await _runtime_graph.aupdate_state(config, patch, as_node=as_node)
        return await _runtime_graph.ainvoke(None, config=config)

    async def resume_supplier_graph(
        self,
        case_id: uuid.UUID,
        decision: dict[str, Any],
    ) -> dict[str, Any]:
        """Legacy wrapper: resume via agent_resume and return graph snapshot."""
        status = await self.agent_resume(case_id, decision)
        case = await self.require_case(case_id)
        snapshot = dict(self._workspace(case).get("supplier_graph") or {})
        snapshot["agent_status"] = status.model_dump(mode="json")
        return snapshot

    async def agent_status(self, case_id: uuid.UUID) -> AgentStatus:
        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        graph = dict(workspace.get("supplier_graph") or {})
        interrupt_payload = workspace.get("agent_interrupt") or {}
        po_drafts = workspace.get("purchase_order_drafts") or []
        latest_po = None
        if po_drafts:
            latest = po_drafts[-1]
            latest_po = latest.get("draft") if isinstance(latest, dict) else latest
        rfq_drafts = workspace.get("rfq_drafts") or []
        latest_rfq = None
        if rfq_drafts:
            latest = rfq_drafts[-1]
            latest_rfq = latest.get("draft") if isinstance(latest, dict) else latest
        if graph.get("rfq_draft"):
            latest_rfq = graph.get("rfq_draft")
        if graph.get("purchase_order_draft"):
            latest_po = graph.get("purchase_order_draft")
        return AgentStatus(
            case_id=str(case.id),
            stage=workspace.get("agent_stage") or graph.get("stage"),
            status=workspace.get("lifecycle_state") or graph.get("status"),
            paused_for_human=bool(workspace.get("paused_for_human") or graph.get("paused_for_human")),
            interrupt_type=(
                interrupt_payload.get("type")
                if isinstance(interrupt_payload, dict)
                else None
            ),
            recommendation=graph.get("recommendation") or workspace.get("recommendation"),
            evaluation=workspace.get("evaluation") or graph.get("evaluation"),
            cost_estimate=workspace.get("cost_estimate") or graph.get("cost_estimate"),
            rfq_draft=latest_rfq,
            purchase_order_draft=latest_po,
            comparison=workspace.get("comparison") or graph.get("comparison"),
            kpi_flags=dict(workspace.get("kpi_flags") or graph.get("kpi_flags") or {}),
            candidates_count=len(graph.get("candidates") or workspace.get("suppliers") or []),
            payment_execution_allowed=False,
        )

    @staticmethod
    def _case_to_strategy_payload(case: ProcurementCase) -> dict[str, Any]:
        positions = []
        for position in case.positions or []:
            if position.cancelled:
                continue
            positions.append(
                {
                    "line_id": position.line_id,
                    "id": position.line_id,
                    "nomenclature_id": position.nomenclature_id,
                    "nomenclature_name": position.nomenclature_name,
                    "quantity": str(position.quantity),
                    "unit": position.unit or "шт",
                    "required_date": (
                        position.required_date.isoformat()
                        if position.required_date
                        else None
                    ),
                    "cancelled": False,
                }
            )
        return {
            "id": str(case.id),
            "case_id": str(case.id),
            "source_number": case.source_number,
            "required_date": (
                case.required_date.isoformat() if case.required_date else None
            ),
            "positions": positions,
        }

    async def strategy_run(
        self,
        request: StrategyRunRequest | None = None,
    ) -> StrategyStatus:
        """Idempotent queue-level supply strategy (waves → optimize → HITL → multi-PO)."""
        invalidate_allocation_cache()
        payload = request or StrategyRunRequest()
        cases = await self._manager_cases()
        if payload.case_ids:
            wanted = {str(cid).strip().casefold() for cid in payload.case_ids if cid}
            cases = [case for case in cases if str(case.id).casefold() in wanted]
        if not cases:
            raise LookupError("Очередь менеджера пуста — нет кейсов для стратегии")

        idempotency_key = (
            payload.idempotency_key
            or f"strategy-run:{AGENT_ID}:{datetime.now(UTC).date()}"
        )[:255]
        existing = dict(_STRATEGY_ARTIFACT)
        if (
            existing.get("strategy_run_idempotency_key") == idempotency_key
            and existing.get("stage")
        ):
            return await self.strategy_status()

        case_payloads = [self._case_to_strategy_payload(case) for case in cases]
        case_ids = [str(case.id) for case in cases]
        thread_id = f"procurement-strategy:{AGENT_ID}:{idempotency_key}"
        config = {
            "configurable": {
                "thread_id": thread_id,
                "runtime": self.supplier_search,
            }
        }
        search_request = SupplierSearchRequest(
            query=payload.query or "поставщик",
            allow_web_fallback=payload.allow_web_fallback,
            idempotency_key=f"strategy-search:{idempotency_key}"[:255],
        )
        result = await _strategy_graph.ainvoke(
            {
                "manager_id": AGENT_ID,
                "cases": case_payloads,
                "case_ids": case_ids,
                "request": search_request.model_dump(mode="json"),
                "today": date.today().isoformat(),
            },
            config=config,
        )
        await self._persist_strategy_state(
            cases,
            result,
            thread_id=thread_id,
            idempotency_key=idempotency_key,
            event_type="strategy_run_paused",
        )
        return await self.strategy_status()

    async def strategy_resume(
        self,
        request: StrategyResumeRequest | dict[str, Any],
    ) -> StrategyStatus:
        """HITL resume for policy/shortlist or multi-PO order drafts."""
        if isinstance(request, StrategyResumeRequest):
            decision = request.model_dump(mode="json")
        else:
            decision = dict(request)
            StrategyResumeRequest.model_validate(decision)
        action = str(decision.get("action") or "")
        artifact = dict(_STRATEGY_ARTIFACT)
        if not artifact.get("thread_id") and not artifact.get("stage"):
            # Try rehydrate from any manager case workspace.
            artifact = await self._load_strategy_artifact_from_cases()
        if not artifact.get("thread_id"):
            raise ValueError("Нет активного прогона стратегии для resume")

        resume_key = decision.get("idempotency_key") or (
            f"strategy-resume:{artifact.get('run_id')}:{action}:{artifact.get('stage')}"
        )
        resume_key = str(resume_key)[:255]
        prior = list(artifact.get("resume_keys") or [])
        if resume_key in prior:
            return await self.strategy_status()

        thread_id = str(artifact.get("thread_id"))
        config = {
            "configurable": {
                "thread_id": thread_id,
                "runtime": self.supplier_search,
            }
        }
        result = await self._invoke_strategy_resume(decision, config, artifact)
        prior.append(resume_key)
        artifact["resume_keys"] = prior[-50:]
        _STRATEGY_ARTIFACT.update(artifact)
        cases = await self._strategy_cases_from_ids(list(artifact.get("case_ids") or []))
        await self._persist_strategy_state(
            cases,
            result,
            thread_id=thread_id,
            idempotency_key=str(artifact.get("strategy_run_idempotency_key") or ""),
            event_type="strategy_resumed",
            resume_keys=prior[-50:],
        )
        return await self.strategy_status()

    async def _invoke_strategy_resume(
        self,
        decision: dict[str, Any],
        config: dict[str, Any],
        artifact: dict[str, Any],
    ) -> dict[str, Any]:
        live = await _strategy_graph.aget_state(config)
        if live.values and live.next:
            try:
                return await _strategy_graph.ainvoke(Command(resume=decision), config=config)
            except Exception as exc:
                logger.warning(
                    "strategy_resume live checkpoint failed (%s); rehydrating",
                    exc,
                )
        if not artifact.get("paused_for_human") and not artifact.get("interrupt"):
            raise ValueError("Нет активного HITL-прерывания стратегии")
        return await self._resume_strategy_rehydrated(decision, config, artifact)

    async def _resume_strategy_rehydrated(
        self,
        decision: dict[str, Any],
        config: dict[str, Any],
        artifact: dict[str, Any],
    ) -> dict[str, Any]:
        interrupt_payload = artifact.get("interrupt") or {}
        interrupt_type = str(
            interrupt_payload.get("type") if isinstance(interrupt_payload, dict) else ""
        )
        action = str(decision.get("action") or "")
        graph_state = {
            key: value
            for key, value in dict(artifact.get("graph") or {}).items()
            if key not in {"runtime", "__interrupt__", "paused_for_human"}
        }
        if not graph_state.get("cases"):
            raise ValueError(
                "Состояние стратегии потеряно после перезапуска. Запустите strategy/run заново."
            )

        is_order = (
            interrupt_type == "procurement_order_approval" or "order" in interrupt_type
        )
        is_policy = (
            interrupt_type == "procurement_policy_approval"
            or "policy" in interrupt_type
            or "shortlist" in interrupt_type
        )
        if is_order:
            if action not in {"approve_order_draft", "reject"}:
                raise ValueError(
                    "Для подтверждения заказов нужно approve_order_draft или reject"
                )
            approved = action == "approve_order_draft"
            drafts = list(graph_state.get("purchase_order_drafts") or [])
            if approved:
                for draft in drafts:
                    if isinstance(draft, dict):
                        draft["status"] = "approved_draft"
                        draft["payment_execution_allowed"] = False
                        draft["executed"] = False
            patch = {
                **graph_state,
                "order_approval": dict(decision),
                "purchase_order_drafts": drafts,
                "status": "order_draft_approved" if approved else "order_rejected",
                "stage": "await_order_hitl",
            }
            as_node = "await_order_hitl"
        elif is_policy:
            if action not in {
                "approve_shortlist",
                "approve_policy",
                "approve_rfq_draft",
                "reject",
            }:
                raise ValueError(
                    "Для политики нужно approve_shortlist / approve_policy / reject"
                )
            approved = action in {
                "approve_shortlist",
                "approve_policy",
                "approve_rfq_draft",
            }
            flags = dict(graph_state.get("kpi_flags") or {})
            if approved:
                flags["supplier_confirmed"] = True
            patch = {
                **graph_state,
                "policy_approval": dict(decision),
                "kpi_flags": flags,
                "status": "policy_approved" if approved else "rejected",
                "stage": "await_policy_hitl",
            }
            as_node = "await_policy_hitl"
        else:
            raise ValueError(
                f"Неизвестный тип HITL стратегии: {interrupt_type or '—'}"
            )

        await _strategy_graph.aupdate_state(config, patch, as_node=as_node)
        return await _strategy_graph.ainvoke(None, config=config)

    async def strategy_status(self) -> StrategyStatus:
        artifact = dict(_STRATEGY_ARTIFACT)
        if not artifact.get("stage"):
            artifact = await self._load_strategy_artifact_from_cases()
        graph = dict(artifact.get("graph") or {})
        interrupt_payload = artifact.get("interrupt") or {}
        supply_policy = (
            artifact.get("supply_policy")
            or graph.get("supply_policy")
            or {}
        )
        drafts = (
            artifact.get("purchase_order_drafts")
            or graph.get("purchase_order_drafts")
            or supply_policy.get("purchase_order_drafts")
            or []
        )
        queue_plan = graph.get("queue_plan") or {}
        return StrategyStatus(
            run_id=artifact.get("run_id"),
            stage=artifact.get("stage") or graph.get("stage"),
            status=artifact.get("status") or graph.get("status"),
            paused_for_human=bool(artifact.get("paused_for_human")),
            interrupt_type=(
                interrupt_payload.get("type")
                if isinstance(interrupt_payload, dict)
                else None
            ),
            case_ids=list(artifact.get("case_ids") or graph.get("case_ids") or []),
            waves=artifact.get("waves") or graph.get("waves") or supply_policy.get("waves"),
            supply_policy=supply_policy or None,
            explanation=artifact.get("explanation")
            or graph.get("explanation")
            or (supply_policy.get("explanation") if isinstance(supply_policy, dict) else None),
            cost_estimate=artifact.get("cost_estimate")
            or graph.get("cost_estimate")
            or (supply_policy.get("cost_estimate") if isinstance(supply_policy, dict) else None),
            purchase_order_drafts=[
                item if isinstance(item, dict) else {} for item in drafts
            ],
            queue_plan_summary=(
                queue_plan.get("summary")
                if isinstance(queue_plan, dict)
                else None
            )
            or (
                supply_policy.get("queue_summary")
                if isinstance(supply_policy, dict)
                else None
            ),
            supplier_diversity=list(
                (queue_plan.get("supplier_diversity") if isinstance(queue_plan, dict) else None)
                or (supply_policy.get("supplier_diversity") if isinstance(supply_policy, dict) else None)
                or []
            ),
            kpi_flags=dict(artifact.get("kpi_flags") or graph.get("kpi_flags") or {}),
            candidates_count=len(graph.get("candidates") or []),
            payment_execution_allowed=False,
        )

    async def _strategy_cases_from_ids(self, case_ids: list[str]) -> list[ProcurementCase]:
        if not case_ids:
            return await self._manager_cases()
        out: list[ProcurementCase] = []
        for raw in case_ids:
            try:
                case = await self._case(uuid.UUID(str(raw)))
            except Exception:
                case = None
            if case is not None:
                out.append(case)
        return out

    async def _load_strategy_artifact_from_cases(self) -> dict[str, Any]:
        cases = await self._manager_cases()
        best: dict[str, Any] = {}
        best_ts = ""
        for case in cases:
            workspace = self._workspace(case)
            artifact = workspace.get("strategy_artifact")
            if not isinstance(artifact, dict) or not artifact.get("stage"):
                continue
            ts = str(artifact.get("updated_at") or "")
            if ts >= best_ts:
                best_ts = ts
                best = dict(artifact)
        if best:
            _STRATEGY_ARTIFACT.clear()
            _STRATEGY_ARTIFACT.update(best)
        return best

    async def _persist_strategy_state(
        self,
        cases: list[ProcurementCase],
        state: dict[str, Any],
        *,
        thread_id: str,
        idempotency_key: str,
        event_type: str,
        resume_keys: list[str] | None = None,
    ) -> None:
        interrupt_raw = state.get("__interrupt__") or ()
        interrupt_payload: dict[str, Any] | None = None
        if interrupt_raw:
            first = interrupt_raw[0]
            value = getattr(first, "value", first)
            interrupt_payload = dict(value) if isinstance(value, dict) else {"value": value}

        snapshot = {
            key: value
            for key, value in state.items()
            if key not in {"runtime", "__interrupt__"}
        }
        paused = bool(interrupt_raw)
        snapshot["paused_for_human"] = paused
        supply_policy = snapshot.get("supply_policy") or {}
        if isinstance(supply_policy, dict):
            supply_policy = dict(supply_policy)
            supply_policy["payment_execution_allowed"] = False

        case_ids = list(snapshot.get("case_ids") or [str(case.id) for case in cases])
        artifact = {
            "run_id": idempotency_key or thread_id,
            "strategy_run_idempotency_key": idempotency_key,
            "thread_id": thread_id,
            "case_ids": case_ids,
            "stage": snapshot.get("stage"),
            "status": snapshot.get("status"),
            "paused_for_human": paused,
            "interrupt": interrupt_payload,
            "graph": snapshot,
            "waves": snapshot.get("waves"),
            "supply_policy": supply_policy,
            "explanation": snapshot.get("explanation"),
            "cost_estimate": snapshot.get("cost_estimate"),
            "purchase_order_drafts": list(snapshot.get("purchase_order_drafts") or []),
            "kpi_flags": dict(snapshot.get("kpi_flags") or {}),
            "resume_keys": list(
                resume_keys
                if resume_keys is not None
                else (_STRATEGY_ARTIFACT.get("resume_keys") or [])
            ),
            "updated_at": datetime.now(UTC).isoformat(),
            "payment_execution_allowed": False,
        }
        _STRATEGY_ARTIFACT.clear()
        _STRATEGY_ARTIFACT.update(artifact)

        drafts = list(artifact.get("purchase_order_drafts") or [])
        for case in cases:
            workspace = self._workspace(case)
            workspace[STRATEGY_METADATA_KEY] = supply_policy
            workspace[STRATEGY_GRAPH_KEY] = {
                "stage": artifact.get("stage"),
                "status": artifact.get("status"),
                "waves": artifact.get("waves"),
                "case_ids": case_ids,
            }
            workspace["strategy_artifact"] = artifact
            workspace["strategy_stage"] = artifact.get("stage")
            workspace["strategy_paused_for_human"] = paused
            if drafts:
                po_entries = list(workspace.get("purchase_order_drafts") or [])
                for draft in drafts:
                    if not isinstance(draft, dict) or not draft.get("po_id"):
                        continue
                    existing = next(
                        (
                            item
                            for item in po_entries
                            if (item.get("draft") or {}).get("po_id") == draft.get("po_id")
                        ),
                        None,
                    )
                    entry = {
                        "idempotency_key": f"strategy-po:{draft.get('po_id')}",
                        "draft": {
                            **draft,
                            "payment_execution_allowed": False,
                            "executed": False,
                        },
                        "executed": False,
                    }
                    if existing is None:
                        po_entries.append(entry)
                    else:
                        existing.update(entry)
                workspace["purchase_order_drafts"] = po_entries
                self._sync_line_amounts_from_po_drafts(workspace)
            if snapshot.get("cost_estimate") is not None:
                workspace["cost_estimate"] = snapshot.get("cost_estimate")
            if paused:
                workspace["lifecycle_state"] = "approval_required"
            self._save_workspace(case, workspace)
            if case.status not in {
                ProcurementCaseStatus.ORDERED.value,
                ProcurementCaseStatus.IN_TRANSIT.value,
                ProcurementCaseStatus.RECEIVING.value,
            }:
                case.status = ProcurementCaseStatus.PURCHASE_DRAFT.value
                case.control_point = "purchase"
            await self._event(
                case,
                event_type,
                f"{event_type}:{case.id}:{artifact.get('stage')}:{paused}",
                {
                    "status": artifact.get("status"),
                    "stage": artifact.get("stage"),
                    "run_id": artifact.get("run_id"),
                    "case_ids": case_ids,
                },
            )

    async def create_purchase_order_draft(
        self,
        case_id: uuid.UUID,
        request: PurchaseOrderDraftRequest,
        *,
        approval_id: str | None = None,
    ) -> PurchaseOrderDraft:
        """Create a draft-only PO; never executes order/payment/1C write."""
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        drafts = metadata.get("purchase_order_drafts") or []
        replay = next(
            (
                item
                for item in drafts
                if item.get("idempotency_key") == request.idempotency_key
            ),
            None,
        )
        if replay:
            return PurchaseOrderDraft.model_validate(replay["draft"])

        approvals = [
            ApprovalRecord.model_validate(item["approval"])
            for item in metadata.get("approvals", [])
        ]
        approved = False
        if approval_id:
            MutationGate.authorize("create_supplier_order", approval_id, approvals)
            approved = True

        supplier_name = request.supplier_id
        for supplier in metadata.get("suppliers") or []:
            if supplier.get("supplier_id") == request.supplier_id:
                supplier_name = str(supplier.get("name") or supplier_name)
                break
        draft = render_purchase_order_draft(
            supplier_id=request.supplier_id,
            supplier_name=supplier_name,
            lines=request.lines,
            case_number=case.source_number or str(case.id),
            source_quote_id=request.source_quote_id,
        )
        draft_payload = draft.model_dump(mode="json")
        draft_payload["status"] = "approved_draft" if approved else "draft"
        draft_payload["payment_execution_allowed"] = False
        draft_payload["executed"] = False
        drafts.append(
            {
                "idempotency_key": request.idempotency_key,
                "draft": draft_payload,
                "approval_id": approval_id,
                "executed": False,
            }
        )
        metadata["purchase_order_drafts"] = drafts
        metadata["lifecycle_state"] = "purchase_order_draft"
        self._sync_line_amounts_from_po_drafts(metadata)
        self._save_workspace(case, metadata)
        case.status = ProcurementCaseStatus.PURCHASE_DRAFT.value
        case.control_point = "purchase"
        await self._event(
            case,
            "purchase_order_draft_created",
            request.idempotency_key,
            {
                "po_id": draft.po_id,
                "executed": False,
                "payment_execution_allowed": False,
                "approval_id": approval_id,
            },
        )
        return PurchaseOrderDraft.model_validate(draft_payload)

    async def list_purchase_order_drafts(
        self,
        case_id: uuid.UUID,
    ) -> list[PurchaseOrderDraft]:
        case = await self.require_case(case_id)
        return [
            PurchaseOrderDraft.model_validate(
                item["draft"] if isinstance(item, dict) and "draft" in item else item
            )
            for item in self._workspace(case).get("purchase_order_drafts", [])
        ]

    async def get_purchase_order_draft(
        self,
        case_id: uuid.UUID,
        po_id: str,
    ) -> PurchaseOrderDraft:
        for draft in await self.list_purchase_order_drafts(case_id):
            if draft.po_id == po_id:
                return draft
        raise LookupError("Purchase order draft not found")

    async def _persist_graph_state(
        self,
        case: ProcurementCase,
        state: dict[str, Any],
        event_type: str,
    ) -> None:
        workspace = self._workspace(case)
        interrupt_raw = state.get("__interrupt__") or ()
        interrupt_payload: dict[str, Any] | None = None
        if interrupt_raw:
            first = interrupt_raw[0]
            value = getattr(first, "value", first)
            interrupt_payload = dict(value) if isinstance(value, dict) else {"value": value}

        snapshot = {
            key: value
            for key, value in state.items()
            if key not in {"runtime", "__interrupt__"}
        }
        paused = bool(interrupt_raw)
        snapshot["paused_for_human"] = paused

        workspace["supplier_graph"] = snapshot
        workspace["agent_stage"] = state.get("stage") or workspace.get("agent_stage")
        workspace["paused_for_human"] = paused
        workspace["agent_interrupt"] = interrupt_payload
        workspace["kpi_flags"] = dict(state.get("kpi_flags") or workspace.get("kpi_flags") or {})
        if state.get("evaluation") is not None:
            workspace["evaluation"] = state.get("evaluation")
        if state.get("cost_estimate") is not None:
            workspace["cost_estimate"] = state.get("cost_estimate")
        if state.get("comparison") is not None:
            workspace["comparison"] = state.get("comparison")
        if state.get("recommendation") is not None:
            workspace["recommendation"] = state.get("recommendation")
        if state.get("candidates"):
            workspace["suppliers"] = list(state.get("candidates") or [])
        if state.get("nomenclature_results"):
            workspace["nomenclature_results"] = list(state.get("nomenclature_results") or [])
        if state.get("quotes"):
            workspace["quotes"] = [
                {"idempotency_key": f"graph-quote:{item.get('quote_id')}", "quote": item}
                for item in state.get("quotes") or []
                if isinstance(item, dict)
            ]

        rfq_draft = state.get("rfq_draft")
        if isinstance(rfq_draft, dict) and rfq_draft.get("rfq_id"):
            drafts = list(workspace.get("rfq_drafts") or [])
            existing = next(
                (
                    item
                    for item in drafts
                    if (item.get("draft") or {}).get("rfq_id") == rfq_draft.get("rfq_id")
                ),
                None,
            )
            entry = {
                "idempotency_key": f"agent-rfq:{rfq_draft.get('rfq_id')}",
                "draft": rfq_draft,
            }
            if existing is None:
                drafts.append(entry)
            else:
                existing.update(entry)
            workspace["rfq_drafts"] = drafts

        po_draft = state.get("purchase_order_draft")
        if isinstance(po_draft, dict) and po_draft.get("po_id"):
            po_payload = dict(po_draft)
            po_payload["payment_execution_allowed"] = False
            po_payload["executed"] = False
            if state.get("status") == "order_draft_approved":
                po_payload["status"] = "approved_draft"
            else:
                po_payload.setdefault("status", "draft")
            drafts = list(workspace.get("purchase_order_drafts") or [])
            existing = next(
                (
                    item
                    for item in drafts
                    if (item.get("draft") or {}).get("po_id") == po_payload.get("po_id")
                ),
                None,
            )
            entry = {
                "idempotency_key": f"agent-po:{po_payload.get('po_id')}",
                "draft": po_payload,
                "executed": False,
            }
            if existing is None:
                drafts.append(entry)
            else:
                existing.update(entry)
            workspace["purchase_order_drafts"] = drafts
            self._sync_line_amounts_from_po_drafts(workspace)

        if paused:
            workspace["lifecycle_state"] = "approval_required"
        else:
            workspace["lifecycle_state"] = state.get("status") or workspace.get(
                "lifecycle_state"
            ) or "agent_running"
        self._save_workspace(case, workspace)
        if case.status not in {
            ProcurementCaseStatus.ORDERED.value,
            ProcurementCaseStatus.IN_TRANSIT.value,
            ProcurementCaseStatus.RECEIVING.value,
        }:
            case.status = ProcurementCaseStatus.PURCHASE_DRAFT.value
            case.control_point = "purchase"
        await self._event(
            case,
            event_type,
            f"{event_type}:{case.id}:{workspace.get('agent_stage')}:{paused}",
            {
                "status": snapshot.get("status"),
                "stage": snapshot.get("stage"),
                "candidate_count": len(snapshot.get("candidates") or []),
                "web_fallback_used": snapshot.get("web_fallback_used", False),
                "paused_for_human": paused,
                "interrupt_type": (
                    interrupt_payload.get("type") if interrupt_payload else None
                ),
                "payment_execution_allowed": False,
            },
        )

    async def list_suppliers(self, case_id: uuid.UUID) -> list[Supplier]:
        case = await self.require_case(case_id)
        return [
            Supplier.model_validate(item)
            for item in self._workspace(case).get("suppliers", [])
        ]

    async def create_rfq_draft(
        self,
        case_id: uuid.UUID,
        request: RFQDraftRequest,
    ) -> RFQDraft:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        drafts = metadata.get("rfq_drafts") or []
        replay = next(
            (item for item in drafts if item.get("idempotency_key") == request.idempotency_key),
            None,
        )
        if replay:
            return RFQDraft.model_validate(replay["draft"])
        suppliers = [
            Supplier.model_validate(item)
            for item in metadata.get("suppliers", [])
            if item.get("supplier_id") in request.supplier_ids
        ]
        draft = render_rfq_draft(
            request,
            suppliers,
            case_number=case.source_number or str(case.id),
        )
        drafts.append(
            {"idempotency_key": request.idempotency_key, "draft": draft.model_dump(mode="json")}
        )
        metadata["rfq_drafts"] = drafts
        metadata["lifecycle_state"] = "rfq_draft"
        self._save_workspace(case, metadata)
        case.status = ProcurementCaseStatus.PURCHASE_DRAFT.value
        case.control_point = "purchase"
        await self._event(
            case,
            "rfq_draft_created",
            request.idempotency_key,
            {"rfq_id": draft.rfq_id},
        )
        return draft

    async def list_rfq_drafts(self, case_id: uuid.UUID) -> list[RFQDraft]:
        case = await self.require_case(case_id)
        return [
            RFQDraft.model_validate(item["draft"])
            for item in self._workspace(case).get("rfq_drafts", [])
        ]

    async def submit_quote(
        self,
        case_id: uuid.UUID,
        submission: QuoteSubmission,
    ) -> SupplierQuote:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        quotes = metadata.get("quotes") or []
        existing = next(
            (
                item
                for item in quotes
                if item.get("idempotency_key") == submission.idempotency_key
            ),
            None,
        )
        if existing:
            return SupplierQuote.model_validate(existing["quote"])
        quotes.append(
            {
                "idempotency_key": submission.idempotency_key,
                "quote": submission.quote.model_dump(mode="json"),
            }
        )
        metadata["quotes"] = quotes
        metadata["lifecycle_state"] = "quotes_received"
        self._save_workspace(case, metadata)
        await self._event(
            case,
            "supplier_quote_recorded",
            submission.idempotency_key,
            {"quote_id": submission.quote.quote_id, "supplier_id": submission.quote.supplier_id},
        )
        return submission.quote

    async def list_quotes(self, case_id: uuid.UUID) -> list[SupplierQuote]:
        case = await self.require_case(case_id)
        return [
            SupplierQuote.model_validate(item["quote"])
            for item in self._workspace(case).get("quotes", [])
        ]

    async def comparison(
        self,
        case_id: uuid.UUID,
        weights: ComparisonWeights | None = None,
    ) -> QuoteComparison:
        case = await self.require_case(case_id)
        comparison = compare_quotes(await self.list_quotes(case_id), weights)
        metadata = self._workspace(case)
        metadata["comparison"] = comparison.model_dump(mode="json")
        metadata["lifecycle_state"] = "comparison_ready"
        self._save_workspace(case, metadata)
        await self.db.flush()
        return comparison

    async def recommendation(
        self,
        case_id: uuid.UUID,
        request: RecommendationRequest,
    ) -> RecommendationRecord:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        audit = metadata.get("recommendation_audit") or []
        replay = next(
            (
                item
                for item in audit
                if item.get("idempotency_key") == request.idempotency_key
            ),
            None,
        )
        if replay:
            return RecommendationRecord.model_validate(replay["recommendation"])

        quotes = await self.list_quotes(case_id)
        quote = next((item for item in quotes if item.quote_id == request.quote_id), None)
        if quote is None:
            raise ValueError("Quote not found")
        if quote.supplier_id != request.supplier_id:
            raise ValueError("Quote does not belong to selected supplier")
        comparison = await self.comparison(case_id)
        score = next(
            (item.final_score for item in comparison.scores if item.quote_id == quote.quote_id),
            None,
        )
        approvals = [
            ApprovalRecord.model_validate(item["approval"])
            for item in metadata.get("approvals", [])
        ]
        supplier_approved = False
        price_approved = False
        if request.supplier_selection_approval_id:
            MutationGate.authorize(
                "select_supplier",
                request.supplier_selection_approval_id,
                approvals,
            )
            supplier_approved = True
        if request.price_approval_id:
            MutationGate.authorize("approve_price", request.price_approval_id, approvals)
            price_approved = True

        recommendation = RecommendationRecord(
            recommendation_id=str(uuid.uuid4()),
            supplier_id=request.supplier_id,
            quote_id=request.quote_id,
            total=quote.total,
            currency=quote.currency,
            score=score,
            rationale=request.rationale,
            status="approved" if supplier_approved and price_approved else "approval_required",
            supplier_selection_approval_id=request.supplier_selection_approval_id,
            price_approval_id=request.price_approval_id,
            created_at=datetime.now(UTC),
        )
        serialized = recommendation.model_dump(mode="json")
        metadata["recommendation"] = serialized
        audit.append(
            {
                "idempotency_key": request.idempotency_key,
                "at": datetime.now(UTC).isoformat(),
                "recommendation": serialized,
            }
        )
        metadata["recommendation_audit"] = audit
        metadata["lifecycle_state"] = recommendation.status
        self._save_workspace(case, metadata)
        case.status = ProcurementCaseStatus.APPROVAL_REQUIRED.value
        await self._event(
            case,
            "procurement_recommendation_recorded",
            request.idempotency_key,
            serialized,
        )
        return recommendation

    async def record_approval(
        self,
        case_id: uuid.UUID,
        request: ApprovalRequest,
        *,
        actor_user_id: str,
    ) -> ApprovalRecord:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        approvals = metadata.get("approvals") or []
        replay = next(
            (
                item
                for item in approvals
                if item.get("idempotency_key") == request.idempotency_key
            ),
            None,
        )
        if replay:
            return ApprovalRecord.model_validate(replay["approval"])
        approval_id = request.approval_id or str(uuid.uuid4())
        approval = ApprovalRecord(
            approval_id=approval_id,
            operation=request.operation,
            status=request.status,
            comment=request.comment,
            actor_user_id=actor_user_id,
            created_at=datetime.now(UTC),
        )
        existing_approval = next(
            (
                item
                for item in reversed(approvals)
                if (item.get("approval") or {}).get("approval_id") == approval_id
                and (item.get("approval") or {}).get("operation") == request.operation
            ),
            None,
        )
        approval_payload = {
            "idempotency_key": request.idempotency_key,
            "approval": approval.model_dump(mode="json"),
        }
        if existing_approval is None:
            approvals.append(approval_payload)
        else:
            existing_approval.update(approval_payload)
        metadata["approvals"] = approvals
        operations = metadata.get("operations") or []
        operation = next(
            (item for item in operations if item.get("operation_id") == approval_id),
            None,
        )
        operation_status: Literal["approval_required", "approved", "rejected"]
        if request.status == "requested":
            operation_status = "approval_required"
        elif request.status == "approved":
            operation_status = "approved"
        else:
            operation_status = "rejected"
        operation_payload = OperationStatus(
            operation_id=approval_id,
            case_id=str(case.id),
            operation=request.operation,
            status=operation_status,
            approval_id=approval_id,
            updated_at=datetime.now(UTC),
        ).model_dump(mode="json")
        if operation is None:
            operations.append(operation_payload)
        else:
            operation.update(operation_payload)
        metadata["operations"] = operations
        metadata["lifecycle_state"] = (
            "approved" if approval.status == "approved" else "approval_required"
        )
        self._save_workspace(case, metadata)
        case.status = (
            ProcurementCaseStatus.ORDERED.value
            if approval.status == "approved" and approval.operation == "create_supplier_order"
            else ProcurementCaseStatus.APPROVAL_REQUIRED.value
        )
        await self._event(
            case,
            "procurement_approval_recorded",
            request.idempotency_key,
            approval.model_dump(mode="json"),
        )
        return approval

    async def list_approvals(self, case_id: uuid.UUID) -> list[ApprovalRecord]:
        case = await self.require_case(case_id)
        return [
            ApprovalRecord.model_validate(item["approval"])
            for item in self._workspace(case).get("approvals", [])
        ]

    async def record_shipment(
        self,
        case_id: uuid.UUID,
        request: ShipmentEventRequest,
    ) -> dict[str, Any]:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        approvals = [
            ApprovalRecord.model_validate(item["approval"])
            for item in metadata.get("approvals", [])
        ]
        MutationGate.authorize("record_shipment", request.approval_id, approvals)
        events = metadata.get("shipment_events") or []
        replay = next(
            (item for item in events if item.get("idempotency_key") == request.idempotency_key),
            None,
        )
        if replay:
            return replay["event"]
        event = request.event.model_dump(mode="json")
        events.append({"idempotency_key": request.idempotency_key, "event": event})
        metadata["shipment_events"] = events
        metadata["lifecycle_state"] = request.event.event_type
        self._save_workspace(case, metadata)
        status_map = {
            "ordered": ProcurementCaseStatus.ORDERED.value,
            "dispatched": ProcurementCaseStatus.IN_TRANSIT.value,
            "in_transit": ProcurementCaseStatus.IN_TRANSIT.value,
            "delayed": ProcurementCaseStatus.IN_TRANSIT.value,
            "received": ProcurementCaseStatus.RECEIVING.value,
        }
        case.status = status_map[request.event.event_type]
        case.control_point = "receipt" if request.event.event_type == "received" else "delivery"
        await self._event(case, "shipment_event_recorded", request.idempotency_key, event)
        return event

    async def list_shipment_events(self, case_id: uuid.UUID) -> list[dict[str, Any]]:
        case = await self.require_case(case_id)
        return [
            dict(item["event"])
            for item in self._workspace(case).get("shipment_events", [])
        ]

    async def handoff_nonconformity(
        self,
        case_id: uuid.UUID,
        request: NonconformityRequest,
    ) -> dict[str, Any]:
        case = await self.require_case(case_id)
        metadata = self._workspace(case)
        rows = metadata.get("nonconformities") or []
        replay = next(
            (item for item in rows if item.get("idempotency_key") == request.idempotency_key),
            None,
        )
        if replay:
            return replay["nonconformity"]
        item = request.nonconformity.model_dump(mode="json")
        rows.append({"idempotency_key": request.idempotency_key, "nonconformity": item})
        metadata["nonconformities"] = rows
        metadata["lifecycle_state"] = "nonconformity"
        self._save_workspace(case, metadata)
        root_metadata = dict(case.case_metadata or {})
        root_metadata["next_quality_agent"] = "otk_head_agent"
        root_metadata["quality_stage"] = ProcurementCaseStatus.NONCONFORMITY.value
        root_metadata["quality_context"] = {"nonconformity": item}
        case.case_metadata = root_metadata
        if case.current_task_id:
            task = await self.db.get(Task, case.current_task_id)
            if task is not None:
                task.status = TaskStatus.COMPLETED
                task.finished_at = datetime.now(UTC)
        case.current_task_id = None
        case.status = ProcurementCaseStatus.NONCONFORMITY.value
        case.current_agent_id = "otk_head_agent"
        case.assigned_agents = ["otk_head_agent"]
        case.control_point = "receipt"
        await self._event(
            case,
            "nonconformity_handed_to_quality",
            request.idempotency_key,
            {**item, "next_agent": "otk_head_agent"},
        )
        return {**item, "next_agent": "otk_head_agent"}

    def _upsert_operation(
        self,
        case: ProcurementCase,
        *,
        operation_id: str,
        operation: str,
        status: Literal[
            "draft",
            "running",
            "completed",
            "approval_required",
            "approved",
            "executed",
            "rejected",
            "failed",
        ],
        approval_id: str | None = None,
        error: str | None = None,
    ) -> OperationStatus:
        metadata = self._workspace(case)
        operations = metadata.get("operations") or []
        payload = OperationStatus(
            operation_id=operation_id,
            case_id=str(case.id),
            operation=operation,
            status=status,
            approval_id=approval_id,
            error=error,
            updated_at=datetime.now(UTC),
        ).model_dump(mode="json")
        existing = next(
            (item for item in operations if item.get("operation_id") == operation_id),
            None,
        )
        if existing is None:
            operations.append(payload)
        else:
            existing.update(payload)
        metadata["operations"] = operations
        self._save_workspace(case, metadata)
        return OperationStatus.model_validate(payload)

    @staticmethod
    def supplier_search_progress(
        *,
        case_id: str,
        operation_id: str,
    ) -> dict[str, Any]:
        """Lightweight poll payload from the in-memory progress buffer."""
        try:
            meta = get_progress_meta(operation_id)
        except Exception:
            meta = None
        if meta is None:
            return {
                "operation_id": operation_id,
                "case_id": case_id,
                "status": "unknown",
                "thoughts": [],
            }
        buffer_case = meta.get("case_id")
        if buffer_case and str(buffer_case) != str(case_id):
            return {
                "operation_id": operation_id,
                "case_id": case_id,
                "status": "unknown",
                "thoughts": [],
            }
        return {
            "operation_id": operation_id,
            "case_id": case_id,
            "status": meta.get("status") or "running",
            "thoughts": list(meta.get("thoughts") or []),
        }

    @staticmethod
    def _attach_operation_thoughts(
        payload: OperationStatus,
        *,
        operation_id: str,
    ) -> OperationStatus:
        """Merge live progress buffer into OperationStatus (soft, never fails)."""
        try:
            thoughts = get_progress(operation_id)
        except Exception:
            thoughts = []
        if not thoughts:
            return payload
        return payload.model_copy(update={"thoughts": thoughts})

    @staticmethod
    def _operation_from_progress_buffer(
        operation_id: str,
        *,
        case_id: str | None = None,
    ) -> OperationStatus | None:
        """Synthetic running status when DB has not committed yet."""
        try:
            meta = get_progress_meta(operation_id)
        except Exception:
            return None
        if not meta:
            return None
        buffer_case = meta.get("case_id")
        if case_id and buffer_case and str(buffer_case) != str(case_id):
            return None
        allowed = {
            "draft",
            "running",
            "completed",
            "approval_required",
            "approved",
            "executed",
            "rejected",
            "failed",
        }
        status_raw = str(meta.get("status") or "running")
        op_status = status_raw if status_raw in allowed else "running"
        return OperationStatus(
            operation_id=operation_id,
            case_id=str(buffer_case or case_id or "") or None,
            operation="supplier_search",
            status=op_status,  # type: ignore[arg-type]
            updated_at=datetime.now(UTC),
            thoughts=list(meta.get("thoughts") or []),
        )

    async def operation_status(
        self,
        case_id: uuid.UUID,
        operation_id: str,
    ) -> OperationStatus | None:
        case = await self.require_case(case_id)
        for item in self._workspace(case).get("operations", []):
            if item.get("operation_id") == operation_id:
                return self._attach_operation_thoughts(
                    OperationStatus.model_validate(item),
                    operation_id=operation_id,
                )
        return self._operation_from_progress_buffer(
            operation_id, case_id=str(case_id)
        )

    async def global_operation_status(
        self,
        operation_id: str,
    ) -> OperationStatus | None:
        cases = (await self.db.execute(select(ProcurementCase))).scalars().all()
        for case in cases:
            for item in self._workspace(case).get("operations", []):
                if item.get("operation_id") == operation_id:
                    payload = {**item, "case_id": str(case.id)}
                    return self._attach_operation_thoughts(
                        OperationStatus.model_validate(payload),
                        operation_id=operation_id,
                    )
        return self._operation_from_progress_buffer(operation_id)

    async def workspace_payload(self, case_id: uuid.UUID) -> dict[str, Any]:
        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        # Heal older agent runs: PO draft prices were not mapped into line_amounts.
        dirty = self._sync_line_amounts_from_po_drafts(workspace)
        payload = self._public_workspace(workspace)
        meta = dict(case.case_metadata or {})
        # Surface project/demo titles for the manager workspace header.
        if meta.get("need_title"):
            payload["need_title"] = meta.get("need_title")
        if meta.get("project_code"):
            payload["project_code"] = meta.get("project_code")
        if meta.get("project_name"):
            payload["project_name"] = meta.get("project_name")
        allocation = await self.allocate_coverage()
        case_coverage = (allocation.get("case_index") or {}).get(str(case_id))
        payload["coverage"] = case_coverage
        base_order_coverage = (
            {
                "tone": case_coverage.get("tone"),
                "label": case_coverage.get("label"),
                "covered_count": case_coverage.get("covered_count") or 0,
                "positions_count": case_coverage.get("positions_count") or 0,
                "uncovered_positions_count": case_coverage.get(
                    "uncovered_positions_count"
                )
                or 0,
                "has_suppliers": any(
                    Decimal(str(line.get("from_supplier") or 0)) > 0
                    or line.get("coverage_source")
                    in {"supplier", "mixed", "supplier_order"}
                    for line in case_coverage.get("lines") or []
                ),
                "lines": case_coverage.get("lines") or [],
            }
            if case_coverage
            else {
                "tone": "uncovered",
                "label": "Полностью необеспечен",
                "covered_count": 0,
                "positions_count": 0,
                "uncovered_positions_count": 0,
                "has_suppliers": False,
                "lines": [],
            }
        )
        order_coverage = merge_order_coverage_with_orchestrator(
            base_order_coverage,
            metadata=meta,
            bucket_reason=(
                meta.get("purchase_manager_bucket_reason")
                if isinstance(meta.get("purchase_manager_bucket_reason"), str)
                else None
            ),
        )
        payload["order_coverage"] = order_coverage
        payload["coverage"] = order_coverage
        payload["supplier_orders"] = list(order_coverage.get("supplier_orders") or [])
        payload["material_allocation"] = {"summary": allocation.get("summary")}
        search = search_fields_from_coverage(
            source_number=getattr(case, "source_number", None),
            order_coverage=order_coverage,
            positions=list(case.positions or []),
        )
        payload.update(search)
        cov_lines = list((payload.get("order_coverage") or {}).get("lines") or [])
        prev_batches = workspace.get("batches")
        batches = sync_batches_workspace(
            workspace,
            positions=list(case.positions or []),
            coverage_lines=cov_lines,
        )
        # Avoid write+commit on every GET — that amplified pool pressure under polling.
        if dirty or prev_batches != batches:
            self._save_workspace(case, workspace)
            await self.db.flush()
        payload["batches"] = batches
        payload["line_schedules"] = dict(workspace.get("line_schedules") or {})
        ful_workspace = {
            **workspace,
            "supplier_orders": payload["supplier_orders"],
            "order_coverage": order_coverage,
        }
        ful = fulfillment_payload(
            case_status=str(case.status or ""),
            workspace=ful_workspace,
        )
        payload.update(ful)
        payload["fulfillment_status"] = ful["fulfillment_status"]
        if not payload.get("summary"):
            payload["summary"] = order_coverage.get("label")
        return payload

    async def update_fulfillment_status(
        self,
        case_id: uuid.UUID,
        payload: FulfillmentStatusUpdateRequest,
    ) -> dict[str, Any]:
        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        status = payload.fulfillment_status
        if status not in FULFILLMENT_LABELS:
            raise ValueError(f"Неизвестный статус: {status}")
        workspace["fulfillment_status"] = status
        workspace["fulfillment_status_manual"] = True
        self._save_workspace(case, workspace)
        await self._event(
            case,
            "fulfillment_status_updated",
            payload.idempotency_key or f"fulfillment:{case_id}:{status}",
            {"fulfillment_status": status},
        )
        return fulfillment_payload(case_status=str(case.status or ""), workspace=workspace)

    async def create_otk_presentation_from_case(
        self,
        case_id: uuid.UUID,
    ) -> dict[str, Any]:
        """Create OTK presentation card from case positions (manager UI button)."""
        from app.agents.quality_engineer_agent.otk_service import OtkPresentationService

        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        meta = dict(case.case_metadata or {})
        drafts = workspace.get("purchase_order_drafts") or []
        supplier_name = ""
        for item in drafts:
            draft = item.get("draft") if isinstance(item, dict) and "draft" in item else item
            if isinstance(draft, dict) and draft.get("supplier_name"):
                supplier_name = str(draft.get("supplier_name"))
                break
        lines = []
        for position in case.positions or []:
            if position.cancelled:
                continue
            lines.append(
                {
                    "code": position.nomenclature_id or "",
                    "nomenclature": position.nomenclature_name or position.nomenclature_id or "",
                    "storage_unit": position.unit or "шт",
                    "qty_upd": float(position.quantity or 0),
                    "qty_fact": float(position.quantity or 0),
                    "category": "other",
                }
            )
        card = OtkPresentationService().create_presentation(
            {
                "purchase_order": case.source_number or str(case.id)[:8],
                "supplier": supplier_name or "Поставщик",
                "counterparty": supplier_name or "Поставщик",
                "project_code": meta.get("project_code"),
                "project_name": meta.get("project_name") or meta.get("need_title"),
                "lines": lines,
            }
        )
        workspace["otk_presentation_id"] = card.id
        workspace["fulfillment_status"] = "otk_presentation"
        workspace["fulfillment_status_manual"] = True
        self._save_workspace(case, workspace)
        await self._event(
            case,
            "otk_presentation_created",
            f"otk-pres:{case.id}:{card.id}",
            {"presentation_id": card.id},
        )
        return {
            "presentation_id": card.id,
            "presentation": card.model_dump(mode="json"),
            **fulfillment_payload(case_status=str(case.status or ""), workspace=workspace),
        }

    async def update_line_schedule(
        self,
        case_id: uuid.UUID,
        line_id: str,
        payload: LineScheduleUpdateRequest,
    ) -> dict[str, Any]:
        invalidate_allocation_cache()
        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        position = next(
            (
                p
                for p in (case.positions or [])
                if str(p.line_id) == str(line_id) and not p.cancelled
            ),
            None,
        )
        if position is None:
            raise LookupError(f"Позиция {line_id} не найдена")
        required = payload.required_date or getattr(position, "required_date", None)
        schedule = compute_schedule(
            required_date=required,
            lead_days=payload.lead_days,
            ship_date=payload.ship_date,
        )
        schedules = dict(workspace.get("line_schedules") or {})
        schedules[str(line_id)] = {
            **schedule,
            "batch_no": payload.batch_no,
            "updated_at": datetime.now(UTC).isoformat(),
        }
        workspace["line_schedules"] = schedules
        if payload.required_date is not None:
            position.required_date = payload.required_date
        cov_lines = list((workspace.get("order_coverage") or {}).get("lines") or [])
        sync_batches_workspace(
            workspace,
            positions=list(case.positions or []),
            coverage_lines=cov_lines,
        )
        self._save_workspace(case, workspace)
        await self._event(
            case,
            "line_schedule_updated",
            payload.idempotency_key or f"schedule:{case_id}:{line_id}",
            {"line_id": line_id, **schedule},
        )
        return {
            "line_id": line_id,
            "schedule": schedules[str(line_id)],
            "batches": workspace.get("batches") or [],
            "formula": schedule.get("formula"),
        }

    async def material_bank(self) -> MaterialBankResponse:
        bank = get_material_bank()
        public = bank.to_public()
        return MaterialBankResponse.model_validate(public)

    async def allocate_coverage(
        self,
        *,
        cases: list[ProcurementCase] | None = None,
    ) -> dict[str, Any]:
        """Deadline allocation for the manager queue (short TTL + single-flight)."""
        now = time.monotonic()
        cached = _allocation_cache.get("payload")
        if cached is not None and (now - float(_allocation_cache["ts"])) < _ALLOCATION_TTL_SEC:
            return cached
        async with _allocation_lock:
            now = time.monotonic()
            cached = _allocation_cache.get("payload")
            if cached is not None and (now - float(_allocation_cache["ts"])) < _ALLOCATION_TTL_SEC:
                return cached
            queue = cases if cases is not None else await self._manager_cases()
            result = allocate_materials_by_deadline(queue, bank=get_material_bank())
            _allocation_cache["payload"] = result
            _allocation_cache["ts"] = time.monotonic()
            return result

    async def all_positions(self) -> AllPositionsResponse:
        """Aggregate queue nomenclature with supplier price_min/max and estimate."""
        cases = await self._manager_cases()
        bank = get_material_bank()
        allocation = await self.allocate_coverage(cases=cases)
        bounds = supplier_price_bounds(bank)
        coverage_by_nom = {
            str(row.get("nomenclature_id") or "").strip().casefold(): row
            for row in allocation.get("by_nomenclature") or []
            if row.get("nomenclature_id")
        }

        buckets: dict[str, dict[str, Any]] = {}
        for case in cases:
            workspace = self._workspace(case)
            line_amounts = dict(workspace.get("line_amounts") or {})
            for position in case.positions or []:
                if position.cancelled:
                    continue
                qty = Decimal(str(position.quantity or 0))
                if qty <= 0:
                    continue
                nom_id = str(position.nomenclature_id or "").strip()
                nom_name = position.nomenclature_name
                key = nom_id.casefold() if nom_id else (nom_name or "").strip().casefold()
                if not key:
                    key = f"line:{position.line_id}"
                manual = line_amounts.get(position.line_id) or {}
                override = manual.get("unit_price")
                if override is None and manual.get("amount") is not None and qty > 0:
                    override = Decimal(str(manual["amount"])) / qty
                elif override is not None:
                    override = Decimal(str(override))
                currency = str(manual.get("currency") or "RUB").upper()
                line_required = getattr(position, "required_date", None) or getattr(
                    case, "required_date", None
                )
                bucket = buckets.get(key)
                if bucket is None:
                    buckets[key] = {
                        "nomenclature_id": nom_id or None,
                        "nomenclature_name": nom_name or nom_id or "Без названия",
                        "unit": position.unit or "шт",
                        "quantity": qty,
                        "line_overrides": [(qty, override)],
                        "currency": currency,
                        "has_manual_override": override is not None,
                        "positions_count": 1,
                        "required_date": line_required,
                    }
                    continue
                bucket["quantity"] += qty
                bucket["line_overrides"].append((qty, override))
                bucket["positions_count"] += 1
                if override is not None:
                    bucket["has_manual_override"] = True
                if not bucket.get("nomenclature_name") and nom_name:
                    bucket["nomenclature_name"] = nom_name
                if line_required is not None:
                    current_req = bucket.get("required_date")
                    if current_req is None or line_required < current_req:
                        bucket["required_date"] = line_required

        rows: list[AllPositionsRow] = []
        total = Decimal("0")
        any_total = False
        offers_by_nom: dict[str, list[Any]] = {}
        for key, bucket in buckets.items():
            bound = bounds.get(key)
            if bound is None and bucket.get("nomenclature_id"):
                bound = bounds.get(str(bucket["nomenclature_id"]).casefold())
            price_min = bound["price_min"] if bound else None
            price_max = bound["price_max"] if bound else None
            cov = coverage_by_nom.get(key) or {}
            from_supplier = None
            raw_from_supplier = cov.get("from_supplier")
            if raw_from_supplier is not None and raw_from_supplier != "":
                from_supplier = Decimal(str(raw_from_supplier))
            from_warehouse = None
            raw_from_warehouse = cov.get("from_warehouse")
            if raw_from_warehouse is not None and raw_from_warehouse != "":
                from_warehouse = Decimal(str(raw_from_warehouse))
            nom_id = bucket.get("nomenclature_id")
            if nom_id:
                nom_key = str(nom_id)
                if nom_key not in offers_by_nom:
                    offers_by_nom[nom_key] = collect_supplier_offers(nom_key, bank=bank)
                offers = offers_by_nom[nom_key]
            else:
                offers = []
            estimate = estimate_nomenclature_amount(
                bucket["quantity"],
                price_min=price_min,
                line_overrides=bucket["line_overrides"],
                coverage_source=cov.get("coverage_source"),
                from_supplier=from_supplier,
                offers=offers,
            )
            if estimate.amount is not None:
                total += estimate.amount
                any_total = True
            # Only suppliers that allocation actually used (not bank top-N ranking).
            coverage_source = cov.get("coverage_source")
            used_suppliers: list[UsedSupplierPart] = []
            if coverage_source != "warehouse" and Decimal(str(cov.get("from_supplier") or 0)) > 0:
                raw_used = cov.get("used_suppliers") or cov.get("supplier_parts") or []
                for part in raw_used:
                    if not isinstance(part, dict):
                        continue
                    sid = str(part.get("supplier_id") or "").strip()
                    if not sid:
                        continue
                    try:
                        qty = Decimal(str(part.get("quantity") or 0))
                    except Exception:
                        qty = Decimal("0")
                    if qty <= 0:
                        continue
                    part_price = None
                    raw_price = part.get("unit_price")
                    if raw_price is not None and raw_price != "":
                        try:
                            part_price = Decimal(str(raw_price))
                        except Exception:
                            part_price = None
                    used_suppliers.append(
                        UsedSupplierPart(
                            supplier_id=sid,
                            supplier_name=str(part.get("supplier_name") or sid),
                            quantity=qty,
                            unit_price=part_price if part_price is not None and part_price > 0 else None,
                        )
                    )
            req = bucket.get("required_date")
            if isinstance(req, datetime):
                req_out: date | datetime | str | None = req.date()
            else:
                req_out = req
            rows.append(
                AllPositionsRow(
                    nomenclature_id=bucket.get("nomenclature_id"),
                    nomenclature_name=bucket.get("nomenclature_name"),
                    unit=bucket.get("unit") or "шт",
                    quantity=bucket["quantity"],
                    price_min=price_min,
                    price_max=price_max,
                    avg_unit_price=estimate.avg_unit_price,
                    estimated_amount=estimate.amount,
                    amount=estimate.amount,
                    overpay=estimate.overpay,
                    amount_source=estimate.source,
                    amount_formula=AMOUNT_FORMULA,
                    currency=bucket.get("currency") or "RUB",
                    coverage_source=coverage_source,
                    coverage_source_label=cov.get("coverage_source_label"),
                    from_warehouse=from_warehouse,
                    from_supplier=from_supplier,
                    positions_count=int(bucket.get("positions_count") or 0),
                    has_manual_override=bool(bucket.get("has_manual_override")),
                    top_suppliers=[],
                    used_suppliers=used_suppliers,
                    required_date=req_out,
                )
            )
        rows.sort(
            key=lambda item: str(item.nomenclature_name or item.nomenclature_id or ""),
        )
        return AllPositionsResponse(
            rows=rows,
            total_estimated_amount=total if any_total else None,
            currency="RUB",
            amount_formula=AMOUNT_FORMULA,
            price_formula=AMOUNT_FORMULA,
            score_formula=SCORE_FORMULA,
        )

    async def supplier_offers_for_case(
        self,
        case_id: uuid.UUID,
        *,
        nomenclature_id: str,
        need_qty: Decimal | None = None,
        top_n: int = 3,
    ) -> SupplierOffersResponse:
        """Top-N ranked supplier offers for a case nomenclature need."""
        case = await self.require_case(case_id)
        nom = nomenclature_id.strip()
        if not nom:
            raise ValueError("nomenclature is required")

        resolved_need = need_qty
        unit = "шт"
        nomenclature_name: str | None = None
        if resolved_need is None:
            total = Decimal("0")
            for position in case.positions or []:
                if position.cancelled:
                    continue
                pid = str(position.nomenclature_id or "").strip()
                if pid.casefold() != nom.casefold():
                    continue
                total += Decimal(str(position.quantity or 0))
                unit = position.unit or unit
                nomenclature_name = position.nomenclature_name or nomenclature_name
            resolved_need = total

        bank = get_material_bank()
        raw_offers = collect_supplier_offers(nom, bank=bank)
        if nomenclature_name is None and raw_offers:
            nomenclature_name = raw_offers[0].get("nomenclature_name")
        if raw_offers and raw_offers[0].get("unit"):
            unit = str(raw_offers[0]["unit"])
        price_min, price_max = price_bounds_from_offers(raw_offers)
        # Align with global table bounds when offers exist in bank pricing.
        bounds = supplier_price_bounds(bank).get(nom.casefold())
        if bounds:
            price_min = bounds["price_min"]
            price_max = bounds["price_max"]

        earliest_required = None
        for position in case.positions or []:
            if position.cancelled:
                continue
            pid = str(position.nomenclature_id or "").strip()
            if pid.casefold() != nom.casefold():
                continue
            req = position.required_date or case.required_date
            if req is None:
                continue
            if earliest_required is None or req < earliest_required:
                earliest_required = req
        # Prefer agent-selected suppliers from evaluation / cost estimate when present.
        workspace = self._workspace(case)
        evaluation = workspace.get("evaluation") or {}
        cost_estimate = workspace.get("cost_estimate") or evaluation.get("cost_estimate") or {}
        agent_tops: list[dict[str, Any]] = []
        for line in [
            *(cost_estimate.get("lines") or []),
            *(evaluation.get("lines") or []),
        ]:
            if not isinstance(line, dict):
                continue
            line_nom = str(line.get("nomenclature_id") or "").strip()
            if line_nom.casefold() != nom.casefold():
                continue
            tops = line.get("top_suppliers") or []
            if tops:
                agent_tops = [item for item in tops if isinstance(item, dict)]
                break
        if agent_tops:
            ranked = agent_tops[: max(1, top_n)]
        else:
            ranked = rank_supplier_offers(
                nom,
                resolved_need or Decimal("0"),
                bank=bank,
                top_n=top_n,
                required_date=earliest_required,
            )
        return SupplierOffersResponse(
            nomenclature_id=nom,
            nomenclature_name=nomenclature_name,
            need_qty=resolved_need or Decimal("0"),
            unit=unit,
            price_min=price_min,
            price_max=price_max,
            score_formula=SCORE_FORMULA,
            top_suppliers=[TopSupplierOffer.model_validate(item) for item in ranked],
        )

    async def workspace_summary(self) -> WorkspaceSummary:
        allocation = await self.allocate_coverage()
        summary = allocation.get("summary") or {}
        queue_cases = int(summary.get("total_orders_count") or 0)
        positions_count = int(summary.get("positions_count") or 0)
        uncovered_orders = int(summary.get("uncovered_orders_count") or 0)
        uncovered_positions = int(summary.get("uncovered_positions_count") or 0)
        active_suppliers = int(summary.get("active_suppliers_count") or 0)
        # Invariants: KPI must be scoped to the manager queue only.
        if uncovered_orders > queue_cases or uncovered_positions > positions_count:
            logger.warning(
                "procurement_manager KPI invariant violated: "
                "queue_cases=%s uncovered_orders=%s positions=%s uncovered_positions=%s",
                queue_cases,
                uncovered_orders,
                positions_count,
                uncovered_positions,
            )
        else:
            logger.info(
                "procurement_manager workspace_summary: "
                "queue_cases=%s uncovered_orders=%s positions=%s "
                "uncovered_positions=%s suppliers=%s",
                queue_cases,
                uncovered_orders,
                positions_count,
                uncovered_positions,
                active_suppliers,
            )
        return WorkspaceSummary(
            uncovered_orders_count=uncovered_orders,
            active_suppliers_count=active_suppliers,
            uncovered_positions_count=uncovered_positions,
            # Compat: former nomenclature_count now = uncovered position lines
            nomenclature_count=uncovered_positions,
            total_orders_count=queue_cases,
            ready_orders_count=int(summary.get("ready_orders_count") or 0),
            attention_orders_count=int(summary.get("attention_orders_count") or 0),
            positions_count=positions_count,
            need_quantity_total=Decimal(str(summary.get("need_quantity_total") or 0)),
            bank_quantity_total=Decimal(str(summary.get("bank_quantity_total") or 0)),
            warehouses_count=int(summary.get("warehouses_count") or 0),
            generated_at=datetime.now(UTC),
        )

    async def enrich_dashboard_cases(self, payload: dict[str, Any]) -> dict[str, Any]:
        """Attach deadline-based coverage status to dashboard case cards."""
        queue = await self._manager_cases()
        allocation = await self.allocate_coverage(cases=queue)
        case_index = allocation.get("case_index") or {}
        meta_by_id = {
            str(case.id).casefold(): dict(case.case_metadata or {})
            for case in queue
        }
        positions_by_id = {
            str(case.id).casefold(): list(case.positions or [])
            for case in queue
        }
        for group in payload.get("groups", []):
            for item in group.get("cases") or []:
                case_id = str(item.get("id") or "").strip().casefold()
                coverage = case_index.get(case_id)
                if coverage is None:
                    # Also try raw id (case_index keys are str(uuid) lowercased below).
                    coverage = case_index.get(str(item.get("id") or ""))
                positions_count = int(
                    (coverage or {}).get("positions_count")
                    or item.get("positions_count")
                    or 0
                )
                if coverage is None:
                    # Unknown allocation → attention if there are positions, not false uncovered.
                    base_order_coverage = {
                        "tone": "attention" if positions_count > 0 else "uncovered",
                        "label": (
                            "Требуют внимания"
                            if positions_count > 0
                            else "Полностью необеспечен"
                        ),
                        "covered_count": 0,
                        "positions_count": positions_count,
                        "uncovered_positions_count": positions_count,
                        "has_suppliers": False,
                        "lines": [],
                    }
                else:
                    base_order_coverage = {
                        "tone": coverage.get("tone"),
                        "label": coverage.get("label"),
                        "covered_count": coverage.get("covered_count") or 0,
                        "positions_count": coverage.get("positions_count") or 0,
                        "uncovered_positions_count": coverage.get(
                            "uncovered_positions_count"
                        )
                        or 0,
                        "has_suppliers": any(
                            (
                                line.get("from_supplier")
                                and Decimal(str(line["from_supplier"])) > 0
                            )
                            or line.get("coverage_source")
                            in {"supplier", "mixed", "supplier_order"}
                            for line in coverage.get("lines") or []
                        ),
                        "needed_quantity": coverage.get("needed_quantity"),
                        "covered_quantity": coverage.get("covered_quantity"),
                        "deficit_quantity": coverage.get("deficit_quantity"),
                        "lines": coverage.get("lines") or [],
                    }
                pm_meta = meta_by_id.get(case_id) or (
                    item.get("case_metadata")
                    if isinstance(item.get("case_metadata"), dict)
                    else {}
                )
                snap = coverage_snapshot_from_metadata(pm_meta)
                if not snap:
                    snap = {
                        "coverage_status": item.get("supplier_coverage_status"),
                        "positions": [],
                        "supplier_orders": item.get("supplier_orders") or [],
                    }
                order_coverage = merge_order_coverage_with_orchestrator(
                    base_order_coverage,
                    snapshot=snap,
                    metadata=pm_meta,
                    bucket_reason=item.get("purchase_manager_bucket_reason"),
                )
                item["order_coverage"] = order_coverage
                item["coverage"] = order_coverage
                item["supplier_orders"] = list(
                    order_coverage.get("supplier_orders") or []
                )
                search = search_fields_from_coverage(
                    source_number=item.get("source_number"),
                    order_coverage=order_coverage,
                    positions=positions_by_id.get(case_id) or item.get("positions") or [],
                )
                item.update(search)
                if not item.get("summary"):
                    item["summary"] = order_coverage.get("label")
                # Surface earliest line deadline when case-level required_date is empty.
                if not item.get("required_date"):
                    line_dates = [
                        line.get("required_date")
                        for line in order_coverage.get("lines") or []
                        if line.get("required_date")
                    ]
                    if line_dates:
                        item["required_date"] = min(str(value) for value in line_dates)
                # Nested copy so clients reading procurement_manager.* still see bank tone.
                pm = dict(item.get("procurement_manager") or {})
                pm["order_coverage"] = order_coverage
                pm["supplier_orders"] = item["supplier_orders"]
                ful = fulfillment_payload(
                    case_status=str(item.get("status") or ""),
                    workspace=pm,
                )
                pm.update(ful)
                item["procurement_manager"] = pm
                item["fulfillment_status"] = ful["fulfillment_status"]
                item["fulfillment_label"] = ful["fulfillment_label"]
                item["fulfillment_tone"] = ful["fulfillment_tone"]
                item["show_otk_button"] = ful["show_otk_button"]
                item["is_completed"] = ful["is_completed"]
            # Urgency order within each dashboard group: earlier required_date first.
            group_cases = list(group.get("cases") or [])
            group_cases.sort(
                key=lambda case: (
                    1 if not case.get("required_date") else 0,
                    str(case.get("required_date") or ""),
                )
            )
            group["cases"] = group_cases
        payload["material_allocation_summary"] = allocation.get("summary")
        return payload

    async def save_line_amounts(
        self,
        case_id: uuid.UUID,
        payload: LineAmountsUpdateRequest,
    ) -> dict[str, Any]:
        invalidate_allocation_cache()
        case = await self.require_case(case_id)
        workspace = self._workspace(case)
        current = dict(workspace.get("line_amounts") or {})
        valid_line_ids = {
            position.line_id
            for position in case.positions or []
            if not position.cancelled
        }
        for entry in payload.lines:
            if entry.line_id not in valid_line_ids:
                continue
            amount = entry.amount
            unit_price = entry.unit_price
            quantity = next(
                (
                    Decimal(str(position.quantity))
                    for position in case.positions or []
                    if position.line_id == entry.line_id
                ),
                Decimal("0"),
            )
            if amount is None and unit_price is not None and quantity > 0:
                amount = (unit_price * quantity).quantize(Decimal("0.01"))
            if unit_price is None and amount is not None and quantity > 0:
                unit_price = (amount / quantity).quantize(Decimal("0.0001"))
            if amount is None and unit_price is None:
                current.pop(entry.line_id, None)
                continue
            current[entry.line_id] = LineAmountEntry(
                line_id=entry.line_id,
                unit_price=unit_price,
                amount=amount,
                currency=(entry.currency or "RUB").upper(),
                source=(entry.source or "manual"),
            ).model_dump(mode="json")
        workspace["line_amounts"] = current
        self._save_workspace(case, workspace)
        if payload.idempotency_key:
            await self._event(
                case,
                "procurement_manager_line_amounts_updated",
                payload.idempotency_key,
                {"lines": list(current.keys())},
            )
        await self.db.flush()
        return {"line_amounts": current}

    @staticmethod
    def _estimate_decimal(value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        try:
            return Decimal(str(value))
        except Exception:
            return None

    def build_estimate_xlsx(self, case: ProcurementCase) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font

        workspace = self._workspace(case)
        line_amounts = dict(workspace.get("line_amounts") or {})
        po_portions = self._po_priced_portions(workspace)
        quote_offers = self._quote_priced_offers(workspace)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Смета"
        headers = [
            "№",
            "Номенклатура",
            "Код",
            "Количество",
            "Ед.",
            "Цена",
            "Сумма",
            "Валюта",
            "Источник цены",
        ]
        for col, title in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=title)
            cell.font = Font(bold=True)

        total = Decimal("0")
        currency = "RUB"
        row_idx = 2
        for index, position in enumerate(case.positions or [], start=1):
            if position.cancelled:
                continue
            quantity = self._estimate_decimal(position.quantity) or Decimal("0")
            manual = line_amounts.get(position.line_id) or {}
            if not isinstance(manual, dict):
                manual = {}
            unit_price = self._estimate_decimal(manual.get("unit_price"))
            amount = self._estimate_decimal(manual.get("amount"))
            manual_source = str(manual.get("source") or "").strip().casefold()
            source = (
                "вручную"
                if manual_source == "manual"
                or (
                    (unit_price is not None or amount is not None)
                    and manual_source not in {"po", "quote"}
                )
                else ""
            )
            if manual_source != "manual":
                portions = po_portions.get(position.line_id) or []
                if portions:
                    amount = sum(
                        (qty * price for qty, price, _cur in portions), Decimal("0")
                    ).quantize(Decimal("0.01"))
                    if quantity > 0:
                        unit_price = (amount / quantity).quantize(Decimal("0.0001"))
                    else:
                        purchase_qty = sum((qty for qty, _p, _c in portions), Decimal("0"))
                        unit_price = (
                            (amount / purchase_qty).quantize(Decimal("0.0001"))
                            if purchase_qty > 0
                            else unit_price
                        )
                    source = "PO"
                elif amount is None and unit_price is None:
                    covered = self._greedy_cover_amount(
                        quantity, quote_offers.get(position.line_id) or []
                    )
                    if covered is not None:
                        amount, _covered_qty = covered
                        unit_price = (
                            (amount / quantity).quantize(Decimal("0.0001"))
                            if quantity > 0
                            else None
                        )
                        source = "КП"
            if amount is None and unit_price is not None:
                amount = (unit_price * quantity).quantize(Decimal("0.01"))
            if amount is not None:
                total += amount
            line_currency = str(manual.get("currency") or currency) or "RUB"
            currency = line_currency
            sheet.cell(row=row_idx, column=1, value=index)
            sheet.cell(
                row=row_idx,
                column=2,
                value=position.nomenclature_name or position.nomenclature_id,
            )
            sheet.cell(row=row_idx, column=3, value=position.nomenclature_id)
            sheet.cell(row=row_idx, column=4, value=float(quantity))
            sheet.cell(row=row_idx, column=5, value=position.unit or "шт")
            sheet.cell(
                row=row_idx,
                column=6,
                value=float(unit_price) if unit_price is not None else None,
            )
            sheet.cell(
                row=row_idx,
                column=7,
                value=float(amount) if amount is not None else None,
            )
            sheet.cell(row=row_idx, column=8, value=line_currency)
            sheet.cell(row=row_idx, column=9, value=source or "—")
            row_idx += 1

        total_label = sheet.cell(row=row_idx + 1, column=6, value="Итого")
        total_label.font = Font(bold=True)
        total_value = sheet.cell(row=row_idx + 1, column=7, value=float(total))
        total_value.font = Font(bold=True)
        sheet.cell(row=row_idx + 1, column=8, value=currency)
        sheet.cell(row=row_idx + 3, column=1, value="Заказ")
        sheet.cell(
            row=row_idx + 3,
            column=2,
            value=case.source_number or case.source_1c_ref,
        )
        sheet.cell(row=row_idx + 4, column=1, value="Подразделение")
        sheet.cell(row=row_idx + 4, column=2, value=case.department_name or "")
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    async def require_case(self, case_id: uuid.UUID) -> ProcurementCase:
        case = await self._case(case_id)
        if case is None:
            raise LookupError("Procurement case not found")
        return case

    async def _case(self, case_id: uuid.UUID) -> ProcurementCase | None:
        return await self.db.scalar(
            select(ProcurementCase)
            .options(selectinload(ProcurementCase.positions))
            .where(ProcurementCase.id == case_id)
        )

    @staticmethod
    def _workspace(case: ProcurementCase) -> dict[str, Any]:
        return dict((case.case_metadata or {}).get(METADATA_KEY) or {})

    @staticmethod
    def _save_workspace(case: ProcurementCase, workspace: dict[str, Any]) -> None:
        metadata = dict(case.case_metadata or {})
        # Graph evaluation / ranking embeds Decimal; JSONB requires JSON-safe values.
        metadata[METADATA_KEY] = jsonable_encoder(workspace)
        case.case_metadata = metadata

    async def _manager_cases(self) -> list[ProcurementCase]:
        """Cases in the manager left-hand queue (same set as /dashboard).

        Previous bug: status filter included engineer-stage statuses
        (human_required, agent_waiting), so KPI counted hundreds of cases
        outside the visible queue (e.g. 503 uncovered orders / 3487 lines).
        """
        manager_visible_statuses = list(
            ACTIVE_CASE_STATUSES | MANAGER_QUEUE_STATUSES | {ProcurementCaseStatus.ORDERED.value}
        )
        result = await self.db.scalars(
            select(ProcurementCase)
            .options(selectinload(ProcurementCase.positions))
            .where(
                ProcurementCase.status.in_(manager_visible_statuses),
                or_(
                    ProcurementCase.current_agent_id == AGENT_ID,
                    ProcurementCase.case_metadata.has_key("purchase_manager_invoked_at"),
                    ProcurementCase.case_metadata.has_key("supplier_order_coverage"),
                ),
                ProcurementCase.closed_at.is_(None),
            )
            .order_by(ProcurementCase.updated_at.desc())
        )
        return [
            case
            for case in result
            if case_in_manager_queue(
                current_agent_id=case.current_agent_id,
                status=case.status,
                purchase_manager_invoked_at=(case.case_metadata or {}).get(
                    "purchase_manager_invoked_at"
                ),
                supplier_coverage_status=(
                    ((case.case_metadata or {}).get("supplier_order_coverage") or {}).get(
                        "coverage_status"
                    )
                    if isinstance(
                        (case.case_metadata or {}).get("supplier_order_coverage"), dict
                    )
                    else None
                ),
            )
            and not (case.case_metadata or {}).get("purchase_manager_workspace_archived_at")
        ]

    @staticmethod
    def _quote_unit_prices(workspace: dict[str, Any]) -> dict[str, Decimal]:
        prices: dict[str, Decimal] = {}
        for quote in workspace.get("quotes") or []:
            if not isinstance(quote, dict):
                continue
            body = quote.get("quote") if isinstance(quote.get("quote"), dict) else quote
            if not isinstance(body, dict):
                continue
            for line in body.get("lines") or []:
                if not isinstance(line, dict):
                    continue
                line_id = str(line.get("line_id") or "").strip()
                unit_price = line.get("unit_price")
                if not line_id or unit_price is None:
                    continue
                prices.setdefault(line_id, Decimal(str(unit_price)))
        return prices

    @staticmethod
    def _iter_po_draft_payloads(workspace: dict[str, Any]) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for item in workspace.get("purchase_order_drafts") or []:
            if not isinstance(item, dict):
                continue
            draft = item.get("draft") if isinstance(item.get("draft"), dict) else item
            if isinstance(draft, dict) and draft.get("po_id"):
                out.append(draft)
        return out

    @staticmethod
    def _po_priced_portions(
        workspace: dict[str, Any],
    ) -> dict[str, list[tuple[Decimal, Decimal, str]]]:
        """line_id → list of (quantity, unit_price, currency) from all PO drafts."""
        portions: dict[str, list[tuple[Decimal, Decimal, str]]] = {}
        for draft in ProcurementManagerService._iter_po_draft_payloads(workspace):
            currency = str(draft.get("currency") or "RUB").upper() or "RUB"
            for line in draft.get("lines") or []:
                if not isinstance(line, dict):
                    continue
                line_id = str(line.get("line_id") or "").strip()
                if not line_id:
                    continue
                try:
                    unit_price = Decimal(str(line.get("unit_price")))
                    quantity = Decimal(str(line.get("quantity") or 0))
                except Exception:
                    continue
                if unit_price <= 0 or quantity <= 0:
                    continue
                portions.setdefault(line_id, []).append((quantity, unit_price, currency))
        return portions

    @staticmethod
    def _quote_priced_offers(
        workspace: dict[str, Any],
    ) -> dict[str, list[tuple[Decimal, Decimal]]]:
        """line_id → list of (available_qty, unit_price) from quotes (offers, not yet allocated)."""
        offers: dict[str, list[tuple[Decimal, Decimal]]] = {}
        for quote in workspace.get("quotes") or []:
            if not isinstance(quote, dict):
                continue
            body = quote.get("quote") if isinstance(quote.get("quote"), dict) else quote
            if not isinstance(body, dict):
                continue
            for line in body.get("lines") or []:
                if not isinstance(line, dict):
                    continue
                line_id = str(line.get("line_id") or "").strip()
                if not line_id:
                    continue
                try:
                    unit_price = Decimal(str(line.get("unit_price")))
                    quantity = Decimal(str(line.get("quantity") or 0))
                except Exception:
                    continue
                if unit_price <= 0 or quantity <= 0:
                    continue
                offers.setdefault(line_id, []).append((quantity, unit_price))
        return offers

    @staticmethod
    def _greedy_cover_amount(
        need_qty: Decimal,
        offers: list[tuple[Decimal, Decimal]],
    ) -> tuple[Decimal, Decimal] | None:
        """Return (amount, covered_qty) covering need cheapest-first from (qty, price) offers."""
        if need_qty <= 0 or not offers:
            return None
        ranked = sorted(offers, key=lambda row: (row[1], -row[0]))
        left = need_qty
        total = Decimal("0")
        covered = Decimal("0")
        for available, price in ranked:
            if left <= 0:
                break
            take = min(left, available)
            if take <= 0:
                continue
            total += take * price
            covered += take
            left -= take
        if covered <= 0:
            return None
        return total.quantize(Decimal("0.01")), covered

    @classmethod
    def _line_amount_protected_from_po_sync(
        cls,
        existing: dict[str, Any] | None,
        portions: list[tuple[Decimal, Decimal, str]],
    ) -> bool:
        """True when an existing line_amounts entry must not be overwritten by PO sync."""
        if not existing or not isinstance(existing, dict):
            return False
        source = str(existing.get("source") or "").strip().casefold()
        if source == "manual":
            return True
        if source == "po":
            return False
        # Legacy (no source): allow refresh when value matches PO-derived patterns,
        # including the bug «first_price × sum(qty)» so multi-supplier can heal.
        try:
            price = Decimal(str(existing.get("unit_price") or 0))
            amount = Decimal(str(existing.get("amount") or 0))
        except Exception:
            return bool(existing.get("unit_price"))
        if price <= 0:
            return False
        po_qty = sum((qty for qty, _price, _cur in portions), Decimal("0"))
        po_amount = sum((qty * p for qty, p, _cur in portions), Decimal("0")).quantize(
            Decimal("0.01")
        )
        po_prices = {p for _qty, p, _cur in portions}
        if amount == po_amount:
            return False
        if price in po_prices and amount == (price * po_qty).quantize(Decimal("0.01")):
            return False
        for qty, p, _cur in portions:
            if price == p and amount == (p * qty).quantize(Decimal("0.01")):
                return False
        return True

    @classmethod
    def _sync_line_amounts_from_po_drafts(cls, workspace: dict[str, Any]) -> bool:
        """Copy PO draft prices into line_amounts for position-row UI.

        Multi-supplier lines: amount = Σ(qty_i × price_i), unit_price = amount / Σ qty_i
        (purchase-weighted). Does not overwrite an explicit manual source. Returns whether
        ``workspace['line_amounts']`` changed.
        """
        current = dict(workspace.get("line_amounts") or {})
        portions_by_line = cls._po_priced_portions(workspace)
        changed = False
        for line_id, portions in portions_by_line.items():
            existing = current.get(line_id) if isinstance(current.get(line_id), dict) else None
            if cls._line_amount_protected_from_po_sync(existing, portions):
                continue
            purchase_qty = sum((qty for qty, _price, _cur in portions), Decimal("0"))
            amount = sum((qty * price for qty, price, _cur in portions), Decimal("0")).quantize(
                Decimal("0.01")
            )
            if purchase_qty <= 0 or amount <= 0:
                continue
            unit_price = (amount / purchase_qty).quantize(Decimal("0.0001"))
            currency = portions[0][2] if portions else "RUB"
            entry = LineAmountEntry(
                line_id=line_id,
                unit_price=unit_price,
                amount=amount,
                currency=currency,
                source="po",
            ).model_dump(mode="json")
            if current.get(line_id) != entry:
                current[line_id] = entry
                changed = True
        if changed:
            workspace["line_amounts"] = current
        return changed

    @staticmethod
    def _public_workspace(workspace: dict[str, Any]) -> dict[str, Any]:
        keys = (
            "lifecycle_state",
            "agent_stage",
            "paused_for_human",
            "agent_interrupt",
            "kpi_flags",
            "evaluation",
            "suppliers",
            "supplier_searches",
            "nomenclature_results",
            "quotes",
            "comparison",
            "rfq_drafts",
            "purchase_order_drafts",
            "approvals",
            "shipment_events",
            "payment_document_draft",
            "recommendation",
            "recommendation_audit",
            "operations",
            "nonconformities",
            "supplier_graph",
            "supply_policy",
            "strategy_graph",
            "strategy_artifact",
            "strategy_stage",
            "strategy_paused_for_human",
            "cost_estimate",
            "line_amounts",
        )
        list_keys = {
            "suppliers",
            "supplier_searches",
            "nomenclature_results",
            "quotes",
            "rfq_drafts",
            "purchase_order_drafts",
            "approvals",
            "shipment_events",
            "recommendation_audit",
            "operations",
            "nonconformities",
        }
        dict_keys = {
            "line_amounts",
            "kpi_flags",
            "evaluation",
            "agent_interrupt",
            "supply_policy",
            "strategy_graph",
            "strategy_artifact",
            "cost_estimate",
        }
        payload = {
            key: workspace.get(
                key,
                []
                if key in list_keys
                else {}
                if key in dict_keys
                else False
                if key == "paused_for_human"
                else None,
            )
            for key in keys
        }
        return payload


    @staticmethod
    def _clean_supplier_search_query(raw: str | None) -> str:
        """Tighten web/SERP query: drop designation noise, keep searchable name."""
        import re

        value = str(raw or "").strip()
        if not value:
            return ""
        # Strip trailing stock markers like (П), (М).
        value = re.sub(r"\s*[\(（]\s*[ПпMmМм]\s*[\)）]\s*$", "", value)
        value = re.sub(r"\s+", " ", value).strip(" -·,;")
        if len(value) > 90:
            value = value[:90].rstrip()
        return value[:500]

    @staticmethod
    def _supplier_query_from_case(case: ProcurementCase) -> str:
        values = [
            position.nomenclature_name or position.nomenclature_id
            for position in case.positions or []
            if not position.cancelled
        ]
        query = ", ".join(dict.fromkeys(value for value in values if value))
        return query[:500] or case.source_number or str(case.id)

    @staticmethod
    def _supplier_category_from_case(case: ProcurementCase) -> str | None:
        for position in case.positions or []:
            raw = position.raw_payload or {}
            category = raw.get("category") or raw.get("Категория")
            if category:
                return str(category)[:255]
        return None

    def _nomenclature_targets_from_case(
        self,
        case: ProcurementCase,
        metadata: dict[str, Any] | None = None,
    ) -> list[NomenclatureSearchItem]:
        """One search target per unique nomenclature, with existing bank/prior matches."""
        workspace = metadata or {}
        prior_by_key = self._prior_nomenclature_suppliers(workspace)
        seen: set[str] = set()
        targets: list[NomenclatureSearchItem] = []
        for position in case.positions or []:
            if position.cancelled:
                continue
            nom_id = str(position.nomenclature_id or "").strip()
            nom_name = str(position.nomenclature_name or nom_id or "").strip()
            query = self._clean_supplier_search_query(nom_name) or nom_name or nom_id
            if len(query) < 2:
                continue
            key = (nom_id or nom_name).casefold()
            if key in seen:
                continue
            seen.add(key)
            existing = self._existing_suppliers_for_nomenclature(
                nomenclature_id=nom_id or None,
                nomenclature_name=nom_name or None,
                prior_by_key=prior_by_key,
            )
            targets.append(
                NomenclatureSearchItem(
                    nomenclature_id=nom_id or None,
                    nomenclature_name=nom_name or query,
                    query=query[:500],
                    existing_suppliers=existing,
                )
            )
        return targets

    def _prior_nomenclature_suppliers(
        self, workspace: dict[str, Any]
    ) -> dict[str, list[Supplier]]:
        prior: dict[str, list[Supplier]] = {}
        for row in workspace.get("nomenclature_results") or []:
            if not isinstance(row, dict):
                continue
            key = str(
                row.get("nomenclature_id") or row.get("nomenclature_name") or row.get("query") or ""
            ).strip().casefold()
            if not key:
                continue
            suppliers: list[Supplier] = []
            for item in row.get("suppliers") or []:
                if isinstance(item, dict):
                    try:
                        suppliers.append(Supplier.model_validate(item))
                    except Exception:
                        continue
            if suppliers:
                prior[key] = suppliers
        # Also fold latest supplier_searches payload if present.
        for search in reversed(workspace.get("supplier_searches") or []):
            result = (search or {}).get("result") or {}
            for row in result.get("nomenclature_results") or []:
                if not isinstance(row, dict):
                    continue
                key = str(
                    row.get("nomenclature_id")
                    or row.get("nomenclature_name")
                    or row.get("query")
                    or ""
                ).strip().casefold()
                if not key or key in prior:
                    continue
                suppliers = []
                for item in row.get("suppliers") or []:
                    if isinstance(item, dict):
                        try:
                            suppliers.append(Supplier.model_validate(item))
                        except Exception:
                            continue
                if suppliers:
                    prior[key] = suppliers
            break
        return prior

    def _existing_suppliers_for_nomenclature(
        self,
        *,
        nomenclature_id: str | None,
        nomenclature_name: str | None,
        prior_by_key: dict[str, list[Supplier]],
    ) -> list[Supplier]:
        """Bank matches + prior search results for skip-if-≥3 logic."""
        by_id: dict[str, Supplier] = {}
        for key in filter(
            None,
            [
                (nomenclature_id or "").strip().casefold(),
                (nomenclature_name or "").strip().casefold(),
            ],
        ):
            for supplier in prior_by_key.get(key) or []:
                by_id.setdefault(supplier.tax_id or supplier.supplier_id, supplier)

        nom_key = (nomenclature_id or "").strip()
        if nom_key:
            try:
                ranked = rank_supplier_offers(
                    nom_key,
                    Decimal("1"),
                    bank=get_material_bank(),
                    top_n=MIN_SUPPLIERS_BEFORE_SKIP,
                )
            except Exception:
                ranked = []
            for offer in ranked:
                sid = str(offer.get("supplier_id") or "")
                if not sid:
                    continue
                unit_price = offer.get("unit_price")
                score = offer.get("score")
                rating = None
                if score is not None:
                    try:
                        # Bank score is 0..1; surface as 0..100 for UI.
                        rating = (Decimal(str(score)) * Decimal("100")).quantize(
                            Decimal("0.01")
                        )
                    except Exception:
                        rating = None
                price = None
                try:
                    price = Decimal(str(unit_price)) if unit_price is not None else None
                except Exception:
                    price = None
                by_id.setdefault(
                    sid,
                    Supplier(
                        supplier_id=sid,
                        name=str(offer.get("supplier_name") or sid),
                        source="internal",
                        evidence=[f"bank:{sid}"],
                        unit_price=price,
                        approx_cost=price,
                        rating=rating,
                        quality_rating=rating or Decimal("0"),
                        commercial_rating=rating or Decimal("0"),
                    ),
                )
        return list(by_id.values())

    @staticmethod
    def _merge_nomenclature_search_results(
        result: SupplierSearchResult,
        *,
        skipped: list[NomenclatureSupplierResult],
        targets: list[NomenclatureSearchItem],
        query: str,
    ) -> SupplierSearchResult:
        by_key: dict[str, NomenclatureSupplierResult] = {}
        for row in [*result.nomenclature_results, *skipped]:
            key = str(
                row.nomenclature_id or row.nomenclature_name or row.query or ""
            ).strip().casefold()
            if key:
                by_key[key] = row
        ordered: list[NomenclatureSupplierResult] = []
        for item in targets:
            key = str(
                item.nomenclature_id or item.nomenclature_name or item.query or ""
            ).strip().casefold()
            if key and key in by_key:
                ordered.append(by_key.pop(key))
        ordered.extend(by_key.values())
        flat: list[Supplier] = []
        sources: list[str] = []
        web_used = False
        for row in ordered:
            flat.extend(row.suppliers)
            sources.extend(row.sources_used)
            web_used = web_used or row.web_fallback_used
        unique: dict[str, Supplier] = {}
        for supplier in flat:
            unique.setdefault(supplier.tax_id or supplier.supplier_id, supplier)
        return result.model_copy(
            update={
                "query": query,
                "suppliers": list(unique.values()),
                "sources_used": list(dict.fromkeys(sources)),
                "web_fallback_used": web_used,
                "nomenclature_results": ordered,
            }
        )

    async def _event(
        self,
        case: ProcurementCase,
        event_type: str,
        idempotency_key: str,
        payload: dict[str, Any],
    ) -> None:
        exists = await self.db.scalar(
            select(ProcurementCaseEvent.id).where(
                ProcurementCaseEvent.case_id == case.id,
                ProcurementCaseEvent.idempotency_key == idempotency_key,
            )
        )
        if exists is None:
            self.db.add(
                ProcurementCaseEvent(
                    case_id=case.id,
                    correlation_id=case.correlation_id,
                    event_type=event_type,
                    agent_id=AGENT_ID,
                    actor_role="procurement_manager",
                    previous_status=case.status,
                    new_status=case.status,
                    idempotency_key=idempotency_key,
                    payload=jsonable_encoder(payload),
                )
            )
        await self.db.flush()


def build_default_rfq_lines(case: ProcurementCase) -> list[RFQLine]:
    return [
        RFQLine(
            line_id=position.line_id,
            nomenclature_id=position.nomenclature_id,
            description=position.nomenclature_name or position.nomenclature_id,
            quantity=position.quantity,
            unit=position.unit or "шт.",
            required_date=position.required_date.date() if position.required_date else None,
        )
        for position in case.positions or []
        if not position.cancelled
    ]


__all__ = [
    "AGENT_ID",
    "MANAGER_QUEUE_STATUSES",
    "ProcurementManagerService",
    "build_default_rfq_lines",
    "case_in_manager_queue",
    "invalidate_allocation_cache",
]
