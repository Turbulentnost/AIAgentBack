from __future__ import annotations

import asyncio
import base64
import hashlib
import json
import re
import time
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass, field
from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any, TypeVar

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TCatalogEntry = TypeVar("TCatalogEntry")

from app.core.config import settings
from app.core.logging import get_logger
from app.agents.document_analysis_agent.material_classification import (
    MATERIAL_KIND_CONSUMABLE,
    MATERIAL_KIND_LABELS,
    MATERIAL_KIND_REQUIRED,
    MATERIAL_KIND_WORKSHOP,
    is_optional_material_kind,
    load_material_classification_index,
    material_classification_for,
)
from app.agents.document_analysis_agent.xls_compat import ensure_openpyxl_bytes

logger = get_logger(__name__)

_AVEON_DATA_DIR = Path(__file__).resolve().parents[3] / "data" / "aveon"
_MAPPING_FILE = _AVEON_DATA_DIR / "Сопоставление номенклатур.xlsx"
_SPECS_FILE = _AVEON_DATA_DIR / "Сокол Спецификация из 1с.xlsx"
_HEADER_FILE = _AVEON_DATA_DIR / "Header.xlsx"
_PRICES_FILE = _AVEON_DATA_DIR / "Цены закупки за 2026_0833.xlsx"
_MATERIAL_CLASSIFICATION_FILE = (
    _AVEON_DATA_DIR / "классификация_расходники_обеспеченность_изделий.xlsx"
)
_RESULT_DATA_START_ROW = 5  # дневной лист
_MONTHLY_DATA_START_ROW = 6  # помесячный: шапка 1–5, данные с 6
_RESULT_GRID_COLS = 23  # устаревший лимит Header.xlsx; помесячный строится динамически
_PRICE_FUZZY_THRESHOLD = 0.78
_SHEET_MONTHLY_ASSURANCE = "1-производственный план (мес.)"
_SHEET_DAILY_ASSURANCE_PREFIX = "2-произв. план ("
_SHEET_DETAILED_PRIORITY = "3-произв. план по обеспеч."
_SHEET_PRODUCT_COVERAGE = "4-обеспеченность по изделиям"
_SHEET_ORDER_PLAN = "план заказов"
_FILL_COVER_GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
_FILL_COVER_YELLOW = PatternFill(fill_type="solid", fgColor="FFEB9C")
_FILL_COVER_RED = PatternFill(fill_type="solid", fgColor="FFC7CE")
_FILL_NONE = PatternFill(fill_type=None)
_PRIORITY_GRID_SIDE = Side(style="thin", color="B0B0B0")
_PRIORITY_DAY_SIDE = Side(style="medium", color="000000")
_PRIORITY_GRID_BORDER = Border(
    left=_PRIORITY_GRID_SIDE,
    right=_PRIORITY_GRID_SIDE,
    top=_PRIORITY_GRID_SIDE,
    bottom=_PRIORITY_GRID_SIDE,
)
_SCHEDULE_CATEGORIES = ("заказ", "опытные", "склад")
_SCHEDULE_METRICS = ("план", "факт")
_CATEGORY_LABELS = {
    "заказ": "Заказ",
    "опытные": "Опытные образцы",
    "склад": "Склад",
}
# Потребность: деталь Заказ/Опыт/Склад (outline) + сводка План/Факт;
# Поступление: недели месяца (outline) + итог; затем прогноз. Число колонок на месяц зависит от недель.
_MONTHLY_SUMMARY_COLS = 2
_MONTHLY_DETAIL_COLS = 6
_MONTHLY_RECEIPT_TOTAL_COLS = 1
_MONTHLY_FORECAST_COLS = 1
_MONTHLY_FIXED_TAIL_COLS = (
    _MONTHLY_DETAIL_COLS
    + _MONTHLY_SUMMARY_COLS
    + _MONTHLY_RECEIPT_TOTAL_COLS
    + _MONTHLY_FORECAST_COLS
)
# A–H: номенклатура, изделия, поставщик, страна, ед. изм., цена, остаток, заказано
_FIXED_RESULT_COLS = 8
# Дневной лист: потребность план + потребность факт + поступление + прогноз
_DAILY_COLS_PER_DAY = 4
_STOCK_COL_LETTER = "G"
# Как в эталоне «Анализ обеспеченности»: дефицит < 0
_FORECAST_DEFICIT_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
_FORECAST_DEFICIT_FONT = Font(color="9C0006")
# Цвета шапки как в Header.xlsx (помесячное / по дням — одинаково; ARGB)
_HEADER_TITLE_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_HEADER_TITLE_FONT = Font(bold=True, color="FFFFFFFF", size=15)
_HEADER_SUBTITLE_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_HEADER_SUBTITLE_FONT = Font(bold=False, color="FF1F1F1F", size=11)
_HEADER_GROUP_FILL = PatternFill(start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid")
_HEADER_GROUP_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_HEADER_METRIC_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_HEADER_METRIC_FONT = Font(bold=True, color="FF1F1F1F", size=11)
_DETAILED_DAY_SKIP_HEADERS = {
    "итог",
    "итого",
    "план",
    "остаток",
    "отклонение",
    "№",
    "п/п",
    "№ п/п",
    "наименование",
    "наимнование",
    "наименования",
}
# Стадии выпуска в отчёте «Модель / изделие»: П/ф приоритетный, но не обязательный.
_PF_STAGE_KEYS = {"п/ф", "пф", "п / ф", "полуфабрикат", "полуфабрикаты"}
_DETAILED_SKIP_STAGE_TOKENS = (
    "отк",
    "склад",
    "итог",
    "остаток",
    "нарастающ",
    "откл",
    "отклон",
)
_DETAILED_RELEASE_STAGE_TOKENS = (
    "выпуск",
    "готов",
    "гп",
    "производ",
    "сборк",
    "план",
)
_PF_SKIP_COLUMN_TOKENS = (
    "итог недели",
    "план недели",
    "факт недели",
    "план месяца",
    "нарастающ",
    "примечан",
    "риск",
    "откл",
)
_PF_DATE_RANGE_RE = re.compile(
    r"(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?\s*[-–—]\s*(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?"
)

_SECTION_NAME_RE = re.compile(
    r"(?i)^(материалы\s+и\s+работы|дерево\s+спецификации|спецификация)\b"
)


@dataclass
class UploadedWorkbook:
    filename: str
    content: bytes


@dataclass
class ProductSpecLink:
    schedule_product: str
    nomenclature: str | None = None
    spec_sheet: str | None = None
    spec_ref_key: str | None = None
    status: str = "unmatched"
    reason: str = ""


@dataclass
class SpecMaterialItem:
    """Позиция спецификации в разрезе одного изделия."""

    nomenclature: str
    quantity: float | None
    product: str
    unit: str | None = None
    spec_sheet: str | None = None


@dataclass
class PurchasePriceEntry:
    """Строка из файла закупочных цен (после агрегации дублей)."""

    nomenclature: str
    supplier: str
    price: float | None
    turnover_qty: float = 0.0


@dataclass
class MergedNomenclatureRow:
    """Уникальная номенклатура после объединения по всем изделиям.

    Структура хранится в памяти результата анализа для дальнейших расчётов.
    """

    nomenclature: str
    products: list[str]
    quantity: float | None
    unit: str | None = None
    by_product: dict[str, float | None] = field(default_factory=dict)
    supplier: str | None = None
    country_of_origin: str | None = None
    price: float | None = None
    price_match: str = ""  # exact | contains | fuzzy | unmatched
    stock: float | None = None
    stock_match: str = ""  # exact | contains | fuzzy | unmatched
    # «Заказано кол-во» из файла остатков (колонка после Остаток в result)
    ordered: float | None = None
    # месяц → категория → {план, факт}
    monthly_demand: dict[str, dict[str, dict[str, float]]] = field(default_factory=dict)
    monthly_receipts: dict[str, float] = field(default_factory=dict)
    monthly_forecast: dict[str, float] = field(default_factory=dict)
    # месяц → ключ недели (start_end ISO) → qty
    weekly_receipts: dict[str, dict[str, float]] = field(default_factory=dict)
    # Ключи ISO-дат YYYY-MM-DD для листа «обеспечение (Месяц)»
    daily_demand: dict[str, float] = field(default_factory=dict)  # план
    daily_demand_fact: dict[str, float] = field(default_factory=dict)  # факт
    daily_receipts: dict[str, float] = field(default_factory=dict)
    daily_forecast: dict[str, float] = field(default_factory=dict)
    # Имя номенклатуры в графике отгрузок (если сопоставлено при обогащении)
    shipment_nomenclature: str | None = None
    coverage_material_kind: str = MATERIAL_KIND_REQUIRED
    coverage_material_label: str = ""
    coverage_material_confidence: str = ""
    coverage_material_reason: str = ""
    coverage_material_kinds_by_product: dict[str, str] = field(default_factory=dict)
    coverage_material_labels_by_product: dict[str, str] = field(default_factory=dict)
    coverage_material_confidences_by_product: dict[str, str] = field(default_factory=dict)
    coverage_material_reasons_by_product: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class _MonthWeekSpan:
    """Календарная неделя (Пн–Вс), обрезанная границами месяца."""

    key: str
    label: str
    start: date
    end: date


@dataclass
class StockEntry:
    nomenclature: str
    quantity: float | None
    ordered_qty: float | None = None


@dataclass
class ShipmentReceiptEntry:
    """Сводка ожидаемых поступлений по номенклатуре (сумма по всем листам/изделиям)."""

    nomenclature: str
    country_of_origin: str | None = None
    monthly_qty: dict[str, float] = field(default_factory=dict)
    daily_qty: dict[str, float] = field(default_factory=dict)  # ISO date → qty


@dataclass
class ShipmentParsedSheet:
    """Кэш распарсенного листа графика отгрузок для рисков логистики."""

    title: str
    header_idx: int
    name_col: int
    msk_col: int | None
    rostov_col: int | None
    date_cols: list[tuple[int, date]]
    rows: list[tuple[Any, ...]]


@dataclass
class ShipmentScheduleBundle:
    """Единый проход по файлам графика отгрузок: поступления, сроки, риски."""

    receipt_index: dict[str, ShipmentReceiptEntry] = field(default_factory=dict)
    shipment_files: list[str] = field(default_factory=list)
    logistics_leads: dict[str, tuple[int, int]] = field(default_factory=dict)
    parsed_sheets: list[ShipmentParsedSheet] = field(default_factory=list)


@dataclass
class LogisticsRiskItem:
    """Позиция на контрольной точке логистики (на расчётную дату = сегодня)."""

    nomenclature: str
    supplier: str | None
    quantity: float
    moscow_date: str
    milestone_date: str  # крайняя дата окна (deadline)
    sheet: str
    window_start: str = ""
    window_end: str = ""
    days_remaining: int = 0
    risk_ratio: float = 0.0  # 0 = красный (риск), 1 = зелёный (запас)
    risk_level: str = "critical"  # low | medium | high | critical


@dataclass
class LogisticsRiskStage:
    key: str
    label: str
    items: list[LogisticsRiskItem] = field(default_factory=list)


@dataclass
class LogisticsRiskBoard:
    """Доска рисков по стадиям логистики (сортировка = порядок цепочки)."""

    as_of: str
    stages: list[LogisticsRiskStage] = field(default_factory=list)


@dataclass
class ScheduleQtyColumn:
    """Колонка qty в графике производства (месяц × категория × план/факт)."""

    col: int
    month: str
    category: str  # заказ | опытные | склад
    metric: str  # план | факт


@dataclass
class ScheduleTableLayout:
    """Распознанная таблица помесячного графика."""

    name_col: int
    data_start_row: int
    columns: list[ScheduleQtyColumn]
    is_split: bool  # True = Заказ/Опытные/Склад × План/Факт


@dataclass
class ScheduleProductPlan:
    """Изделие из графика производства с помесячным планом выпуска.

    monthly_qty: месяц → категория → {план, факт}
    """

    product: str
    monthly_qty: dict[str, dict[str, dict[str, float]]] = field(default_factory=dict)
    spec_name: str = ""
    spec_ref_key: str = ""


@dataclass
class DetailedScheduleProductPlan:
    """Изделие из детального графика с планом выпуска по дням выбранного месяца.

    daily_qty — план (после извлечения заполняется на все дни месяца, пропуски → 0).
    daily_fact — факт только за дни/периоды, которые есть в файле (без автозаполнения нулями).
    """

    product: str
    daily_qty: dict[str, float] = field(default_factory=dict)  # ISO date → план
    daily_fact: dict[str, float] = field(default_factory=dict)  # ISO date → факт
    year: int = 0
    month: int = 0


@dataclass
class DetailedPlanCellRef:
    """Ячейка плана П/ф на исходном листе Отчёта (для окраски копии)."""

    product: str
    row: int
    plan_col: int
    day_keys: list[str] = field(default_factory=list)
    plan_qty: float = 0.0


@dataclass
class DetailedScheduleExtract:
    files: list[str]
    plans: list[DetailedScheduleProductPlan]
    year: int
    month: int
    day_keys: list[str] = field(default_factory=list)
    source_filename: str = ""
    source_sheet_name: str = ""
    source_bytes: bytes | None = None
    plan_cells: list[DetailedPlanCellRef] = field(default_factory=list)


@dataclass
class AveonAnalysisResult:
    roles: dict[str, str]
    source: str
    production_schedule_files: list[str]
    production_schedule_products: list[str]
    production_schedule_plans: list[ScheduleProductPlan] = field(default_factory=list)
    detailed_production_schedule_files: list[str] = field(default_factory=list)
    detailed_schedule_month: str = ""
    detailed_schedule_year: int = 0
    detailed_schedule_month_num: int = 0
    detailed_schedule_day_keys: list[str] = field(default_factory=list)
    detailed_schedule_plans: list[DetailedScheduleProductPlan] = field(default_factory=list)
    product_spec_links: list[ProductSpecLink] = field(default_factory=list)
    material_usages: list[SpecMaterialItem] = field(default_factory=list)
    merged_nomenclatures: list[MergedNomenclatureRow] = field(default_factory=list)
    result_xlsx_bytes: bytes | None = None
    shift_assignment_xlsx_bytes: bytes | None = None
    shift_assignment_file_name: str = "сменное_задание_закупки.xlsx"
    shift_assignment_values: list[list[str]] = field(default_factory=list)
    shift_assignment_row_priorities: list[str | None] = field(default_factory=list)
    shift_assignment_row_kinds: list[str] = field(default_factory=list)
    shift_assignment_meta: dict[str, Any] = field(default_factory=dict)
    stock_files: list[str] = field(default_factory=list)
    shipment_files: list[str] = field(default_factory=list)
    logistics_risks: LogisticsRiskBoard | None = None
    schedule_diff_has_changes: bool = False
    schedule_diff_changed_months: list[str] = field(default_factory=list)
    schedule_diff_changed_cells: int = 0
    schedule_diff_file_name: str = "график_производства_изменения.xlsx"
    schedule_diff_xlsx_bytes: bytes | None = None
    schedule_diff_old_version: str = ""
    schedule_diff_new_version: str = ""
    schedule_diff_message: str = ""
    schedule_baseline_saved: bool = False
    schedule_compared_with_saved: bool = False
    detailed_diff_has_changes: bool = False
    detailed_diff_changed_dates: list[str] = field(default_factory=list)
    detailed_diff_changed_cells: int = 0
    detailed_diff_file_name: str = "детальный_график_изменения.xlsx"
    detailed_diff_xlsx_bytes: bytes | None = None
    detailed_diff_old_version: str = ""
    detailed_diff_new_version: str = ""
    detailed_diff_message: str = ""
    detailed_baseline_saved: bool = False
    detailed_compared_with_saved: bool = False
    coverage_dashboard: dict[str, Any] | None = None
    coverage_rebuild: dict[str, Any] | None = None
    input_sources: dict[str, Any] = field(default_factory=dict)

@dataclass
class _MappingRow:
    nomenclature: str
    contract_nomenclature: str
    contract: str = ""


WorkbookRole = str
ROLE_SPECIFICATION = "specification"
ROLE_STOCK = "stock"
ROLE_PRODUCTION_SCHEDULE = "production_schedule"
ROLE_DETAILED_PRODUCTION_SCHEDULE = "detailed_production_schedule"
ROLE_SHIPMENT_SCHEDULE = "shipment_schedule"
ROLE_OTHER = "other"

# Роли загружаемых пользователем файлов (ровно 4 типа в UI).
UPLOAD_FILE_ROLES = frozenset(
    {
        ROLE_STOCK,
        ROLE_PRODUCTION_SCHEDULE,
        ROLE_DETAILED_PRODUCTION_SCHEDULE,
        ROLE_SHIPMENT_SCHEDULE,
    }
)

KNOWN_ROLES = {
    ROLE_SPECIFICATION,
    ROLE_STOCK,
    ROLE_PRODUCTION_SCHEDULE,
    ROLE_DETAILED_PRODUCTION_SCHEDULE,
    ROLE_SHIPMENT_SCHEDULE,
    ROLE_OTHER,
}

# «Неделя 01.07-03.07» / «План 01.07-03.07» (префикс обязателен — иначе ловим годовые диапазоны)
_DETAILED_PERIOD_HEADER_RE = re.compile(
    r"(?:недел[яи]?|план)\s+\d{1,2}\.\d{1,2}\s*[-–—]\s*\d{1,2}\.\d{1,2}",
    re.IGNORECASE,
)
_WEEK_HEADER_RE = _DETAILED_PERIOD_HEADER_RE
# Дневные колонки в превью (datetime из openpyxl → ISO-строка)
_DAY_DATE_HEADER_RE = re.compile(r"20\d{2}-\d{2}-\d{2}")

_MONTH_LOWER = (
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
)

_MONTH_NOMINATIVE = (
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
)

_MONTH_GENITIVE = (
    "января",
    "февраля",
    "марта",
    "апреля",
    "мая",
    "июня",
    "июля",
    "августа",
    "сентября",
    "октября",
    "ноября",
    "декабря",
)

# Порядок = стадии цепочки для UI. D = дата поставки в Москву из колонки графика.
# «5-10» / «7-14 к.д.» → (короткая, длинная) календарные дни.
# МСК/Ростов — окна [короткая … длинная]; загрузка/таможня — точечные дни.
_LOGISTICS_CUSTOMS_DAYS = 2
_LOGISTICS_STAGE_DEFS: tuple[tuple[str, str], ...] = (
    ("loading_dispatch", "Загрузка и отправка"),
    ("msk_arrival", "Прибытие в МСК"),
    ("customs_clearance", "Таможня"),
    ("rostov_arrival", "Прибытие в Ростов"),
)
_LOGISTICS_RANGE_RE = re.compile(r"(\d+)\s*[-–—]\s*(\d+)")


async def _resolve_role_map_for_analysis(
    workbooks: list[UploadedWorkbook],
) -> tuple[dict[str, WorkbookRole], str]:
    """Роли для analyze: то же быстрое определение, что и при загрузке файлов."""
    return await classify_aveon_excel_files(workbooks, use_lm=False)


def _schedule_diff_message(
    schedule_diff: Any | None,
    *,
    compared_with_saved: bool,
    baseline_saved: bool,
) -> str:
    if baseline_saved:
        return (
            "График производства сохранён как базовая версия. "
            "При следующем анализе изменения будут сравниваться с этой версией."
        )
    if not schedule_diff or not compared_with_saved:
        return ""
    if schedule_diff.has_changes:
        months = (
            ", ".join(schedule_diff.changed_months)
            if schedule_diff.changed_months
            else "есть расхождения"
        )
        return (
            f"Планы изменились относительно сохранённой версии "
            f"{schedule_diff.old_version_label} → новая {schedule_diff.new_version_label}: {months}"
        )
    return (
        f"Сохранённая версия {schedule_diff.old_version_label} и новая "
        f"{schedule_diff.new_version_label} совпадают по планам"
    )


def _detailed_diff_message(
    detailed_diff: Any | None,
    *,
    compared_with_saved: bool,
    baseline_saved: bool,
    baseline_month: str = "",
) -> str:
    month_suffix = f" за {baseline_month}" if baseline_month else ""
    if baseline_saved:
        return (
            f"Детальный график{month_suffix} сохранён как базовая версия. "
            "При следующем анализе изменения будут сравниваться с этой версией."
        )
    if not detailed_diff or not compared_with_saved:
        return ""
    if detailed_diff.has_changes:
        dates = (
            ", ".join(detailed_diff.changed_months)
            if detailed_diff.changed_months
            else "есть расхождения"
        )
        return (
            f"Детальный план изменился относительно сохранённой версии "
            f"{detailed_diff.old_version_label} → новая {detailed_diff.new_version_label}: {dates}"
        )
    return (
        f"Сохранённая версия детального графика {detailed_diff.old_version_label} и новая "
        f"{detailed_diff.new_version_label} совпадают по планам"
    )


_ANALYSIS_SOURCE_LABELS: dict[str, str] = {
    "upload": "Загруженный Excel",
    "1c_db": "1С → БД",
    "mixed": "Смешанный (загрузка + 1С)",
    "upload_merged": "Пользовательский график + Google Sheets",
    "server_merged": "Сервер: 1С (Россия) + Google Sheets (Китай)",
    "none": "Не задано",
}


def _analysis_input_entry(
    *,
    source: str,
    files: list[str],
    detail: str = "",
) -> dict[str, Any]:
    return {
        "source": source,
        "source_label": _ANALYSIS_SOURCE_LABELS.get(source, source),
        "files": files,
        "detail": detail,
    }


def _build_analysis_input_sources(
    *,
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    production_from_upload: bool,
    production_files: list[str],
    detailed_from_upload: bool,
    detailed_files: list[str],
    db_spec_catalog_nonempty: bool,
    db_stock_index: dict[str, StockEntry] | None,
    stock_files: list[str],
    shipment_files: list[str],
) -> dict[str, Any]:
    """Источники данных для анализа: 1С БД, загруженные Excel или смешанный режим."""
    uploaded_by_role: dict[str, list[str]] = {}
    for uploaded in workbooks:
        role = role_map.get(uploaded.filename, ROLE_OTHER)
        if role == ROLE_OTHER or uploaded.filename.lower() in {
            "merged_schedule.xlsx",
            "merged_schedule_uploaded.xlsx",
        }:
            continue
        uploaded_by_role.setdefault(role, []).append(uploaded.filename)

    production_source = "none"
    if production_from_upload:
        production_source = "upload"
    elif production_files:
        production_source = "1c_db"

    detailed_source = "none"
    if detailed_from_upload:
        detailed_source = "upload"
    elif detailed_files:
        detailed_source = "1c_db"

    stock_upload_files = uploaded_by_role.get(ROLE_STOCK, [])
    if db_stock_index:
        stock_source = "1c_db"
        stock_detail = f"{len(db_stock_index)} позиций в БД"
    elif stock_upload_files:
        stock_source = "upload"
        stock_detail = ""
    else:
        stock_source = "none"
        stock_detail = ""

    specs_source = "1c_db" if db_spec_catalog_nonempty else "none"
    specs_detail = "Ресурсные спецификации из PostgreSQL (onec_resource_specs)"

    shipment_upload_files = [
        name
        for name in shipment_files
        if name.lower() != "merged_schedule.xlsx" and "merged_schedule" not in name.lower()
    ]
    has_merged_shipment = any(
        name.lower() == "merged_schedule.xlsx" or "merged_schedule" in name.lower()
        for name in shipment_files
    )
    has_uploaded_merged_shipment = any(
        "uploaded" in name.lower() and "merged_schedule" in name.lower()
        for name in shipment_files
    )
    if has_uploaded_merged_shipment:
        shipment_source = "upload_merged"
        shipment_detail = (
            "Россия — из пользовательского графика, сохранённого в БД после анализа; "
            "Китай — актуальный Google Sheets"
        )
    elif has_merged_shipment and shipment_upload_files:
        shipment_source = "mixed"
        shipment_detail = (
            "merged_schedule.xlsx собран на сервере (Россия — БД 1С, Китай — Google Sheets); "
            f"дополнительно загружено: {', '.join(shipment_upload_files)}"
        )
    elif has_merged_shipment:
        shipment_source = "server_merged"
        shipment_detail = "merged_schedule.xlsx: Россия — БД 1С, Китай — Google Sheets"
    elif shipment_upload_files:
        shipment_source = "upload"
        shipment_detail = ""
    else:
        shipment_source = "none"
        shipment_detail = ""

    entries = {
        "production_schedule": _analysis_input_entry(
            source=production_source,
            files=production_files if production_files else uploaded_by_role.get(ROLE_PRODUCTION_SCHEDULE, []),
            detail="Помесячный план производства",
        ),
        "detailed_production_schedule": _analysis_input_entry(
            source=detailed_source,
            files=detailed_files if detailed_files else uploaded_by_role.get(ROLE_DETAILED_PRODUCTION_SCHEDULE, []),
            detail="Детальный план по дням",
        ),
        "specifications": _analysis_input_entry(
            source=specs_source,
            files=["1С → PostgreSQL (onec_resource_specs)"] if specs_source == "1c_db" else [],
            detail=specs_detail,
        ),
        "stock": _analysis_input_entry(
            source=stock_source,
            files=stock_files if stock_files else stock_upload_files,
            detail=stock_detail,
        ),
        "shipment_schedule": _analysis_input_entry(
            source=shipment_source,
            files=shipment_files,
            detail=shipment_detail,
        ),
    }

    core_sources = {
        entries["production_schedule"]["source"],
        entries["detailed_production_schedule"]["source"],
        entries["stock"]["source"],
    } - {"none"}
    # Спецификации всегда из 1С при наличии каталога.
    if specs_source == "1c_db":
        core_sources.add("1c_db")
    if shipment_source not in ("none",):
        core_sources.add(shipment_source)

    upload_like = {"upload", "upload_merged"}
    db_like = {"1c_db", "server_merged"}
    has_upload = bool(core_sources & upload_like)
    has_db = bool(core_sources & db_like)
    has_mixed_kind = "mixed" in core_sources or shipment_source == "mixed"

    if not core_sources:
        summary_mode = "empty"
        summary_text = "Нет данных для расчёта — загрузите файлы или выполните синхронизацию 1С"
    elif has_mixed_kind or (has_upload and has_db):
        summary_mode = "mixed"
        parts: list[str] = []
        for key, label in (
            ("production_schedule", "План производства"),
            ("detailed_production_schedule", "Детальный план"),
            ("specifications", "Спецификации"),
            ("stock", "Остатки"),
            ("shipment_schedule", "График отгрузок"),
        ):
            item = entries[key]
            if item["source"] == "none":
                continue
            files_label = ", ".join(item["files"]) if item["files"] else "—"
            parts.append(f"{label}: {item['source_label']} ({files_label})")
        summary_text = "; ".join(parts)
    elif has_upload and not has_db:
        summary_mode = "all_upload"
        summary_text = "Все ключевые данные — из загруженных Excel-файлов"
    elif has_db and not has_upload:
        summary_mode = "all_1c"
        summary_text = "Все ключевые данные — из 1С (БД PostgreSQL) и серверных графиков"
    else:
        summary_mode = "partial"
        summary_text = "Часть данных отсутствует — см. детализацию по блокам"

    user_uploads = [
        {"filename": uploaded.filename, "role": role_map.get(uploaded.filename, ROLE_OTHER)}
        for uploaded in workbooks
        if uploaded.filename.lower() != "merged_schedule.xlsx"
    ]
    server_injected = [
        uploaded.filename
        for uploaded in workbooks
        if uploaded.filename.lower() == "merged_schedule.xlsx"
    ]

    return {
        "summary": {"mode": summary_mode, "text": summary_text},
        "uploaded_files": user_uploads,
        "server_injected_files": server_injected,
        **entries,
    }


async def analyze_aveon_excel_files(
    workbooks: list[UploadedWorkbook],
    db: "AsyncSession | None" = None,
    user_id: Any | None = None,
) -> AveonAnalysisResult:
    """Роли → изделия графика → спецификации (БД 1С) → материалы → result.xlsx."""
    from app.agents.document_analysis_agent.onec_db_sources import (
        build_country_index_from_db,
        build_stock_index_from_db,
        build_unit_index_from_db,
        build_product_spec_hints,
        load_latest_detailed_production_schedule_from_db,
        load_plan_product_spec_links_from_db,
        load_latest_production_schedule_from_db,
        load_db_spec_catalog,
        preload_spec_materials_for_links,
        repair_spec_links_without_materials,
        finalize_onec_spec_links,
        products_with_loaded_onec_specs,
        expand_spec_eligible_product_names,
    )

    if db is None:
        raise ValueError("Для анализа нужна сессия БД (остатки и спецификации из 1С)")

    from app.services.onec_db_schema import ensure_onec_agent_tables

    await ensure_onec_agent_tables()

    workbooks = _normalize_uploaded_workbooks(workbooks)
    role_map, source = await _resolve_role_map_for_analysis(workbooks)

    # AsyncSession cannot service concurrent queries on the same connection.
    db_stock_index = await build_stock_index_from_db(db)
    db_country_index = await build_country_index_from_db(db)
    db_unit_index = await build_unit_index_from_db(db)
    db_spec_catalog = await load_db_spec_catalog(db)
    if not db_stock_index:
        logger.warning("document_analysis_agent.db_stock_empty")
    if not db_spec_catalog:
        logger.warning("document_analysis_agent.db_spec_catalog_empty")

    from app.agents.document_analysis_agent.production_schedule_diff import (
        prune_workbooks_to_latest_schedules,
    )
    from app.agents.document_analysis_agent.detailed_schedule_diff import (
        infer_detailed_workbook_month,
        prune_workbooks_to_latest_detailed,
    )
    from app.agents.document_analysis_agent.schedule_snapshot import (
        detailed_month_key,
        get_saved_detailed_file,
        get_saved_production_file,
        save_schedule_snapshot,
    )

    saved_production = get_saved_production_file(user_id)

    detailed_upload_month: tuple[int, int] | None = None
    for uploaded in workbooks:
        if role_map.get(uploaded.filename) != ROLE_DETAILED_PRODUCTION_SCHEDULE:
            continue
        detailed_upload_month = infer_detailed_workbook_month(uploaded.content)
        if detailed_upload_month is not None:
            break

    saved_detailed = (
        get_saved_detailed_file(user_id, detailed_upload_month[0], detailed_upload_month[1])
        if detailed_upload_month is not None
        else None
    )

    workbooks, role_map, schedule_diff, schedule_compared_with_saved = await asyncio.to_thread(
        prune_workbooks_to_latest_schedules,
        workbooks,
        role_map,
        saved_file=saved_production,
    )
    workbooks, role_map, detailed_diff, detailed_compared_with_saved = await asyncio.to_thread(
        prune_workbooks_to_latest_detailed,
        workbooks,
        role_map,
        saved_file=saved_detailed,
    )

    has_production_upload = any(
        role_map.get(wb.filename) == ROLE_PRODUCTION_SCHEDULE for wb in workbooks
    )
    has_detailed_upload = any(
        role_map.get(wb.filename) == ROLE_DETAILED_PRODUCTION_SCHEDULE for wb in workbooks
    )
    schedule_baseline_saved = has_production_upload and not schedule_compared_with_saved
    detailed_baseline_saved = has_detailed_upload and not detailed_compared_with_saved

    (schedule_files, schedule_plans), detailed_extract = await asyncio.gather(
        asyncio.to_thread(_extract_production_schedule_products, workbooks, role_map),
        asyncio.to_thread(_extract_detailed_production_schedule, workbooks, role_map),
    )
    production_from_upload = bool(schedule_plans)
    db_schedule_plans_for_specs: list[ScheduleProductPlan] = []
    if not schedule_plans:
        db_schedule_files, db_schedule_plans = await load_latest_production_schedule_from_db(db)
        if db_schedule_plans:
            schedule_files = db_schedule_files
            schedule_plans = db_schedule_plans
            db_schedule_plans_for_specs = db_schedule_plans
    else:
        _db_schedule_files, db_schedule_plans_for_specs = await load_latest_production_schedule_from_db(db)
    detailed_from_upload = bool(detailed_extract.plans)
    if not detailed_extract.plans:
        db_detailed_extract = await load_latest_detailed_production_schedule_from_db(db)
        if db_detailed_extract.plans:
            detailed_extract = db_detailed_extract
    products = list(
        dict.fromkeys(
            [plan.product for plan in schedule_plans]
            + [plan.product for plan in detailed_extract.plans]
        )
    )
    product_spec_hints = build_product_spec_hints(
        list(schedule_plans) + list(db_schedule_plans_for_specs)
    )
    product_spec_links = await _resolve_schedule_products_to_specs(
        products,
        db_spec_catalog=db_spec_catalog,
        product_spec_hints=product_spec_hints,
    )
    plan_product_spec_links = await load_plan_product_spec_links_from_db(db)
    seen_plan_specs = {
        (
            (link.schedule_product or "").strip().casefold(),
            (link.spec_ref_key or "").strip().lower(),
            (link.spec_sheet or "").strip().casefold(),
        )
        for link in product_spec_links
    }
    for link in plan_product_spec_links:
        dedupe_key = (
            (link.schedule_product or "").strip().casefold(),
            (link.spec_ref_key or "").strip().lower(),
            (link.spec_sheet or "").strip().casefold(),
        )
        if dedupe_key not in seen_plan_specs:
            product_spec_links.append(link)
            seen_plan_specs.add(dedupe_key)
    db_materials_by_ref = await preload_spec_materials_for_links(db, product_spec_links)
    db_materials_by_ref = await repair_spec_links_without_materials(
        db,
        product_spec_links,
        db_materials_by_ref,
        db_spec_catalog,
        product_spec_hints=product_spec_hints,
    )
    finalize_onec_spec_links(product_spec_links, db_materials_by_ref, db_spec_catalog)
    spec_eligible_products = expand_spec_eligible_product_names(
        products_with_loaded_onec_specs(
            product_spec_links,
            db_materials_by_ref,
            db_spec_catalog,
        ),
        list(schedule_plans),
        list(detailed_extract.plans),
    )
    shipment_bundle = await asyncio.to_thread(
        _load_shipment_schedule_bundle, workbooks, role_map
    )
    (
        material_usages,
        merged_nomenclatures,
        result_xlsx,
        stock_files,
        shipment_files,
        logistics_risks,
        coverage_dashboard,
        coverage_rebuild,
    ) = await asyncio.to_thread(
        _collect_and_merge_spec_materials,
        product_spec_links,
        workbooks,
        role_map,
        schedule_plans,
        detailed_extract,
        db_stock_index,
        db_materials_by_ref,
        db_country_index,
        db_unit_index,
        user_id,
        shipment_bundle,
        spec_eligible_products,
    )
    price_matched = sum(
        1 for row in merged_nomenclatures if row.price_match not in ("", "unmatched")
    )
    stock_matched = sum(
        1 for row in merged_nomenclatures if row.stock_match not in ("", "unmatched")
    )
    receipts_nonzero = sum(
        1
        for row in merged_nomenclatures
        if any(value > 0 for value in row.monthly_receipts.values())
    )
    forecast_deficit = sum(
        1
        for row in merged_nomenclatures
        if any(value < 0 for value in row.monthly_forecast.values())
    )
    logger.info(
        "document_analysis_agent.roles_classified",
        source=source,
        roles=role_map,
    )
    logger.info(
        "document_analysis_agent.production_schedule_products",
        files=schedule_files,
        count=len(products),
        products=products[:20],
        plans=[
            {"product": plan.product, "months": plan.monthly_qty}
            for plan in schedule_plans[:5]
        ],
    )
    logger.info(
        "document_analysis_agent.detailed_production_schedule",
        files=detailed_extract.files,
        month=f"{detailed_extract.year:04d}-{detailed_extract.month:02d}"
        if detailed_extract.year and detailed_extract.month
        else "",
        products=len(detailed_extract.plans),
        days=len(detailed_extract.day_keys),
    )
    logger.info(
        "document_analysis_agent.product_spec_links",
        matched=sum(1 for item in product_spec_links if item.status == "matched"),
        spec_eligible=len(spec_eligible_products),
        total=len(product_spec_links),
    )
    from app.agents.document_analysis_agent.shift_assignment import (
        SHIFT_ASSIGNMENT_FILE_NAME,
        build_shift_assignment_bundle,
    )

    schedule_name_index = _schedule_name_index_from_shipment_bundle(shipment_bundle)
    if not schedule_name_index:
        schedule_name_index = await _resolve_shipment_schedule_name_index(
            workbooks, role_map, shipment_bundle=shipment_bundle
        )
    shift_assignment_bundle = await build_shift_assignment_bundle(
        merged_nomenclatures,
        logistics_risks,
        detailed_extract,
        schedule_name_index=schedule_name_index,
    )
    shift_assignment_xlsx = shift_assignment_bundle.xlsx_bytes
    shift_assignment_preview = shift_assignment_bundle.preview
    logger.info(
        "document_analysis_agent.spec_materials_merged",
        usages=len(material_usages),
        unique=len(merged_nomenclatures),
        price_matched=price_matched,
        stock_matched=stock_matched,
        stock_files=stock_files,
        shipment_files=shipment_files,
        receipts_nonzero=receipts_nonzero,
        forecast_deficit_rows=forecast_deficit,
        logistics_risk_items=sum(len(stage.items) for stage in logistics_risks.stages),
        result_bytes=len(result_xlsx) if result_xlsx else 0,
        shift_assignment_bytes=len(shift_assignment_xlsx),
    )
    production_wb = next(
        (wb for wb in workbooks if role_map.get(wb.filename) == ROLE_PRODUCTION_SCHEDULE),
        None,
    )
    detailed_wb = next(
        (
            wb
            for wb in workbooks
            if role_map.get(wb.filename) == ROLE_DETAILED_PRODUCTION_SCHEDULE
        ),
        None,
    )
    if production_wb is not None or detailed_wb is not None:
        detailed_payload: tuple[int, int, str, bytes] | None = None
        if detailed_wb is not None:
            month_info = detailed_upload_month or infer_detailed_workbook_month(
                detailed_wb.content
            )
            if month_info is None and detailed_extract.year > 0 and detailed_extract.month > 0:
                month_info = (detailed_extract.year, detailed_extract.month)
            if month_info is not None:
                detailed_payload = (
                    month_info[0],
                    month_info[1],
                    detailed_wb.filename,
                    detailed_wb.content,
                )
            else:
                logger.warning(
                    "document_analysis_agent.detailed_snapshot_month_unknown",
                    filename=detailed_wb.filename,
                )
        try:
            save_schedule_snapshot(
                user_id,
                production=(
                    (production_wb.filename, production_wb.content) if production_wb else None
                ),
                detailed=detailed_payload,
            )
        except OSError:
            logger.warning("document_analysis_agent.schedule_snapshot_save_failed")

    detailed_baseline_month = ""
    if detailed_baseline_saved and detailed_extract.year > 0 and detailed_extract.month > 0:
        detailed_baseline_month = detailed_month_key(
            detailed_extract.year,
            detailed_extract.month,
        )

    input_sources = _build_analysis_input_sources(
        workbooks=workbooks,
        role_map=role_map,
        production_from_upload=production_from_upload,
        production_files=list(schedule_files),
        detailed_from_upload=detailed_from_upload,
        detailed_files=list(detailed_extract.files),
        db_spec_catalog_nonempty=bool(db_spec_catalog),
        db_stock_index=db_stock_index if db_stock_index else None,
        stock_files=list(stock_files),
        shipment_files=list(shipment_files),
    )

    return AveonAnalysisResult(
        roles=role_map,
        source=source,
        production_schedule_files=schedule_files,
        production_schedule_products=products,
        production_schedule_plans=schedule_plans,
        detailed_production_schedule_files=detailed_extract.files,
        detailed_schedule_month=(
            f"{detailed_extract.year:04d}-{detailed_extract.month:02d}"
            if detailed_extract.year and detailed_extract.month
            else ""
        ),
        detailed_schedule_year=detailed_extract.year,
        detailed_schedule_month_num=detailed_extract.month,
        detailed_schedule_day_keys=list(detailed_extract.day_keys),
        detailed_schedule_plans=list(detailed_extract.plans),
        product_spec_links=product_spec_links,
        material_usages=material_usages,
        merged_nomenclatures=merged_nomenclatures,
        result_xlsx_bytes=result_xlsx,
        shift_assignment_xlsx_bytes=shift_assignment_xlsx,
        shift_assignment_file_name=SHIFT_ASSIGNMENT_FILE_NAME,
        shift_assignment_values=list(shift_assignment_preview.values),
        shift_assignment_row_priorities=list(shift_assignment_preview.row_priorities),
        shift_assignment_row_kinds=list(shift_assignment_preview.row_kinds),
        shift_assignment_meta={
            "as_of": shift_assignment_preview.as_of,
            "week_period": shift_assignment_preview.week_period,
            "week_in_period": shift_assignment_preview.week_in_period,
            "task_count": shift_assignment_preview.task_count,
            "urgent_count": shift_assignment_preview.urgent_count,
            "today_count": shift_assignment_preview.today_count,
            "week_count": shift_assignment_preview.week_count,
        },
        stock_files=stock_files,
        shipment_files=shipment_files,
        logistics_risks=logistics_risks,
        schedule_diff_has_changes=bool(schedule_diff and schedule_diff.has_changes),
        schedule_diff_changed_months=list(schedule_diff.changed_months) if schedule_diff else [],
        schedule_diff_changed_cells=schedule_diff.changed_cells if schedule_diff else 0,
        schedule_diff_file_name=(
            schedule_diff.file_name if schedule_diff else "график_производства_изменения.xlsx"
        ),
        schedule_diff_xlsx_bytes=(
            schedule_diff.file_bytes if schedule_diff and schedule_diff.has_changes else None
        ),
        schedule_diff_old_version=schedule_diff.old_version_label if schedule_diff else "",
        schedule_diff_new_version=schedule_diff.new_version_label if schedule_diff else "",
        schedule_diff_message=_schedule_diff_message(
            schedule_diff,
            compared_with_saved=schedule_compared_with_saved,
            baseline_saved=schedule_baseline_saved,
        ),
        schedule_baseline_saved=schedule_baseline_saved,
        schedule_compared_with_saved=schedule_compared_with_saved,
        detailed_diff_has_changes=bool(detailed_diff and detailed_diff.has_changes),
        detailed_diff_changed_dates=list(detailed_diff.changed_months) if detailed_diff else [],
        detailed_diff_changed_cells=detailed_diff.changed_cells if detailed_diff else 0,
        detailed_diff_file_name=(
            detailed_diff.file_name if detailed_diff else "детальный_график_изменения.xlsx"
        ),
        detailed_diff_xlsx_bytes=(
            detailed_diff.file_bytes if detailed_diff and detailed_diff.has_changes else None
        ),
        detailed_diff_old_version=detailed_diff.old_version_label if detailed_diff else "",
        detailed_diff_new_version=detailed_diff.new_version_label if detailed_diff else "",
        detailed_diff_message=_detailed_diff_message(
            detailed_diff,
            compared_with_saved=detailed_compared_with_saved,
            baseline_saved=detailed_baseline_saved,
            baseline_month=detailed_baseline_month,
        ),
        detailed_baseline_saved=detailed_baseline_saved,
        detailed_compared_with_saved=detailed_compared_with_saved,
        coverage_dashboard=coverage_dashboard,
        coverage_rebuild=coverage_rebuild,
        input_sources=input_sources,
    )


# Классификация ролей: LM только как последний шанс и не дольше пары секунд.
_CLASSIFY_LM_TIMEOUT_SECONDS = 3
_PREVIEW_MAX_ROWS = 12
_PREVIEW_MAX_COLS = 14
# Кэш ролей по содержимому файла (повторная загрузка / debounce на UI).
_ROLE_CACHE_MAX = 64
_role_cache: OrderedDict[str, WorkbookRole] = OrderedDict()


def _workbook_content_key(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _role_cache_get(content_key: str) -> WorkbookRole | None:
    role = _role_cache.get(content_key)
    if role is None:
        return None
    _role_cache.move_to_end(content_key)
    return role


def _role_cache_put(content_key: str, role: WorkbookRole) -> None:
    _role_cache[content_key] = role
    _role_cache.move_to_end(content_key)
    while len(_role_cache) > _ROLE_CACHE_MAX:
        _role_cache.popitem(last=False)


async def classify_aveon_excel_files(
    workbooks: list[UploadedWorkbook],
    *,
    use_lm: bool = False,
) -> tuple[dict[str, WorkbookRole], str]:
    """Роль по содержимому файла: первые строки листов, без полного скана и без LM."""
    started = time.perf_counter()
    role_map: dict[str, WorkbookRole] = {}
    need_preview: list[UploadedWorkbook] = []
    key_by_filename: dict[str, str] = {}

    for wb in workbooks:
        key = _workbook_content_key(wb.content)
        key_by_filename[wb.filename] = key
        cached = _role_cache_get(key)
        if cached is not None:
            role_map[wb.filename] = cached
            continue
        need_preview.append(wb)

    if not need_preview:
        logger.info(
            "document_analysis_agent.roles_classified_only",
            source="cache",
            roles=role_map,
            elapsed_ms=int((time.perf_counter() - started) * 1000),
        )
        return role_map, "cache"

    try:
        normalized = _normalize_uploaded_workbooks(need_preview)
        previews = await _build_workbook_previews_async(normalized)
    except OSError as exc:
        logger.warning(
            "document_analysis_agent.preview_build_failed",
            error=str(exc),
        )
        for wb in need_preview:
            role_map[wb.filename] = ROLE_OTHER
        return role_map, "preview_failed"

    fresh_roles = {
        str(preview["filename"]): _classify_preview_locally(preview)
        for preview in previews
    }
    source = "content_preview"
    if use_lm:
        try:
            fresh_roles, source = await _classify_workbooks_with_lm(
                previews,
                lm_timeout_seconds=_CLASSIFY_LM_TIMEOUT_SECONDS,
            )
        except Exception as exc:
            logger.warning(
                "document_analysis_agent.classify_failed_fallback_local",
                error=str(exc),
            )
            source = "content_preview"

    for filename, role in fresh_roles.items():
        role_map[filename] = role
        key = key_by_filename.get(filename)
        if key is not None and role in UPLOAD_FILE_ROLES:
            _role_cache_put(key, role)

    if role_map.keys() - fresh_roles.keys():
        source = f"cache+{source}"

    logger.info(
        "document_analysis_agent.roles_classified_only",
        source=source,
        roles=role_map,
        elapsed_ms=int((time.perf_counter() - started) * 1000),
    )
    return role_map, source


def _normalize_uploaded_workbooks(workbooks: list[UploadedWorkbook]) -> list[UploadedWorkbook]:
    """Приводит .xls к содержимому .xlsx; имя файла оставляем исходным (роли/UI)."""
    normalized: list[UploadedWorkbook] = []
    for uploaded in workbooks:
        try:
            content = ensure_openpyxl_bytes(uploaded.filename, uploaded.content)
        except ValueError as exc:
            raise ValueError(f"{uploaded.filename}: {exc}") from exc
        if content is not uploaded.content:
            logger.info(
                "document_analysis_agent.xls_converted",
                filename=uploaded.filename,
                source_bytes=len(uploaded.content),
                xlsx_bytes=len(content),
            )
        normalized.append(UploadedWorkbook(filename=uploaded.filename, content=content))
    return normalized


def _classify_filename_locally(filename: str) -> WorkbookRole:
    """Имя файла само по себе роль не задаёт — нужен разбор содержимого."""
    _ = filename
    return ROLE_OTHER


def _coerce_upload_role(role: WorkbookRole | None) -> WorkbookRole | None:
    if role in UPLOAD_FILE_ROLES:
        return role
    return None


def _sample_sheet_rows(sheet: Any) -> list[list[str | None]]:
    """Первые строки листа без sheet.max_row — иначе openpyxl сканирует весь файл."""
    rows: list[list[str | None]] = []
    for index, row in enumerate(
        sheet.iter_rows(
            min_row=1,
            max_row=_PREVIEW_MAX_ROWS,
            max_col=_PREVIEW_MAX_COLS,
            values_only=True,
        )
    ):
        rows.append(
            [_short(_clean_text(value), 120) if value is not None else None for value in row]
        )
        if index + 1 >= _PREVIEW_MAX_ROWS:
            break
    return rows


def _sheet_sample_has_text(rows: list[list[str | None]]) -> bool:
    return any(cell for row in rows for cell in row)


def _load_workbook_for_preview(content: bytes, *, read_only: bool):
    try:
        return load_workbook(BytesIO(content), data_only=True, read_only=read_only)
    except OSError:
        return load_workbook(BytesIO(content), data_only=True, read_only=False)


def _build_one_workbook_preview(uploaded: UploadedWorkbook) -> dict[str, Any]:
    workbook = _load_workbook_for_preview(uploaded.content, read_only=True)
    try:
        sheet_previews: list[dict[str, Any]] = []
        empty_samples = True
        for sheet in workbook.worksheets[:8]:
            rows = _sample_sheet_rows(sheet)
            if _sheet_sample_has_text(rows):
                empty_samples = False
            sheet_previews.append(
                {
                    "sheet": sheet.title,
                    "max_row": len(rows),
                    "max_column": _PREVIEW_MAX_COLS,
                    "sample_rows": rows,
                }
            )
    finally:
        workbook.close()

    if empty_samples and len(uploaded.content) > 12_000:
        workbook = _load_workbook_for_preview(uploaded.content, read_only=False)
        try:
            sheet_previews = []
            for sheet in workbook.worksheets[:8]:
                rows = _sample_sheet_rows(sheet)
                sheet_previews.append(
                    {
                        "sheet": sheet.title,
                        "max_row": len(rows),
                        "max_column": _PREVIEW_MAX_COLS,
                        "sample_rows": rows,
                    }
                )
        finally:
            workbook.close()

    return {"filename": uploaded.filename, "sheets": sheet_previews}


def _build_workbook_previews(workbooks: list[UploadedWorkbook]) -> list[dict[str, Any]]:
    return [_build_one_workbook_preview(uploaded) for uploaded in workbooks]


async def _build_workbook_previews_async(
    workbooks: list[UploadedWorkbook],
) -> list[dict[str, Any]]:
    """Параллельный разбор превью по файлам (openpyxl в thread pool)."""
    if not workbooks:
        return []
    if len(workbooks) == 1:
        return [await asyncio.to_thread(_build_one_workbook_preview, workbooks[0])]
    return list(
        await asyncio.gather(
            *[
                asyncio.to_thread(_build_one_workbook_preview, uploaded)
                for uploaded in workbooks
            ]
        )
    )


async def _classify_workbooks_with_lm(
    previews: list[dict[str, Any]],
    *,
    lm_timeout_seconds: int | float | None = None,
) -> tuple[dict[str, WorkbookRole], str]:
    """Local-first: уверенные локальные роли без LM; LM только для other/сомнительных."""
    resolved: dict[str, WorkbookRole] = {}
    pending: list[dict[str, Any]] = []

    for preview in previews:
        filename = str(preview["filename"])
        local = _classify_preview_locally(preview)
        if local in UPLOAD_FILE_ROLES:
            resolved[filename] = local
        else:
            pending.append(preview)

    lm_role_map: dict[str, WorkbookRole] = {}
    if pending and lm_timeout_seconds and float(lm_timeout_seconds) > 0:
        lm_role_map = (
            await _try_lm_classify_workbooks(pending, timeout=lm_timeout_seconds) or {}
        )
        for preview in pending:
            filename = str(preview["filename"])
            lm_role = _coerce_upload_role(_normalize_lm_role(lm_role_map.get(filename)))
            if lm_role is not None:
                resolved[filename] = lm_role
            else:
                fallback = _classify_preview_locally(preview)
                resolved[filename] = (
                    fallback if fallback in UPLOAD_FILE_ROLES else ROLE_OTHER
                )

    local_confident = len(previews) - len(pending)
    if local_confident > 0 and not pending:
        source = "local_fast"
    elif local_confident > 0 and lm_role_map:
        source = "local+lm"
    elif local_confident > 0 and pending and not lm_role_map:
        source = "local_parser"
    elif local_confident == 0 and lm_role_map:
        source = "lm_studio"
    else:
        source = "local_parser"

    logger.info(
        "document_analysis_agent.classify_local_first",
        local_confident=local_confident,
        lm_pending=len(pending),
        lm_answered=len(lm_role_map),
        source=source,
    )
    return resolved, source


def _preview_text(preview: dict[str, Any]) -> str:
    """Только листы и строки — имя файла в разбор роли не входит."""
    return _normalize(json.dumps(preview.get("sheets") or [], ensure_ascii=False))


def _preview_looks_like_shipment_schedule(preview: dict[str, Any]) -> bool:
    text = _preview_text(preview)
    if "график отгрузок" in text or "график отгрузки" in text:
        return True
    sheet_names = _preview_sheet_names(preview)
    if any("тамож" in name for name in sheet_names):
        return True
    if any("итц" in name for name in sheet_names):
        return True
    if any("реестр" in name and "заказ" in name for name in sheet_names):
        return True
    # расширенный график / ИТЦ: номенклатура + даты поставки / логистика
    if "номенклатура" in text and (
        "примерная дата поставки" in text
        or ("дата заказа" in text and "логистика" in text)
        or ("заказано" in text and "поставк" in text)
    ):
        return True
    if (
        "позици" in text
        and "модел" in text
        and ("отгруз" in text or "спецификац" in text or "итц" in text)
    ):
        return True
    if "тип операции" in text and "основание" in text:
        return True
    return False


def _preview_has_material_stock_table(preview: dict[str, Any]) -> bool:
    """Таблица остатков материалов: «Номенклатура» + «Остаток…» (не остаток ГП в графике выпуска)."""
    text = _preview_text(preview)
    if "ведомость по товарам" in text or "конечный остаток" in text:
        return True
    has_nomenclature = "номенклатура" in text or "наименование тмц" in text
    has_stock_col = (
        "остаток на" in text
        or "остаток," in text
        or " остаток" in text
        or "ост." in text
        or "конечный остаток" in text
        or "начальный остаток" in text
    )
    # Гибрид «план сверху + остатки снизу»: приоритет остатков для агента
    return has_nomenclature and has_stock_col


def _preview_looks_like_detailed_production_schedule(preview: dict[str, Any]) -> bool:
    """Детальный график: по дням/неделям. Не путать с остатками и помесячным графиком."""
    if _preview_looks_like_shipment_schedule(preview):
        return False
    # Файл с таблицей остатков материалов (даже если сверху есть кусок плана) → не детальный
    if _preview_has_material_stock_table(preview):
        return False

    text = _preview_text(preview)

    if "детальный график" in text or "детальный план" in text:
        return True
    if "п/ф" in text and ("факт" in text or "модель" in text):
        return True
    # «План по недельно»: график выпуска ГП с колонками-днями
    if "график выпуска" in text:
        return True
    # Отчёт план/факт по дням: «Модель / изделие» + план/факт
    if ("модель" in text and "изделие" in text) and ("факт" in text):
        return True

    day_date_hits = len(_DAY_DATE_HEADER_RE.findall(text))
    week_word_hits = text.count("неделя") + text.count("недели")
    period_hits = len(_DETAILED_PERIOD_HEADER_RE.findall(text))
    has_product_rows = any(
        marker in text
        for marker in ("наименование", "сокол", "fpv", "модель", "изделие")
    )
    # Много дневных дат в шапке (01.04, 02.04…) при изделиях
    if day_date_hits >= 5 and has_product_rows and "логистика" not in text:
        return True
    # Явные недельные колонки без таблицы материалов
    if (week_word_hits >= 2 or period_hits >= 3) and has_product_rows and "номенклатура" not in text:
        return True
    return False


def _preview_looks_like_production_schedule(preview: dict[str, Any]) -> bool:
    """Обычный (помесячный) график производства."""
    text = _preview_text(preview)
    if "график отгрузок" in text or "график отгрузки" in text:
        return False
    if _preview_looks_like_detailed_production_schedule(preview):
        return False
    if _preview_has_material_stock_table(preview):
        return False
    if "график выпуска" in text:
        return False
    if "неделя" in text or "недели" in text:
        return False
    if "график производства" in text:
        return True
    if "наименования изделий" in text:
        month_hits = sum(1 for month in _MONTH_LOWER if month in text)
        return month_hits >= 2
    return False


def _preview_looks_like_specification(preview: dict[str, Any]) -> bool:
    text = _preview_text(preview)
    return (
        "наименование тмц" in text
        or "цена по спецификации" in text
        or "manufacturer partno" in text
        or ("description" in text and "designator" in text)
        or "спецификация" in text
    )


def _preview_looks_like_stock(preview: dict[str, Any]) -> bool:
    if _preview_looks_like_shipment_schedule(preview):
        return False
    # Сильный детальный график (выпуск/план-факт) важнее слабого слова «остаток» в шапке
    text = _preview_text(preview)
    if "график выпуска" in text or (("модель" in text and "изделие" in text) and "факт" in text):
        return False
    if _preview_has_material_stock_table(preview):
        return True
    return False


def _preview_sheet_names(preview: dict[str, Any]) -> list[str]:
    return [_normalize(str(sheet.get("sheet") or "")) for sheet in preview.get("sheets") or []]


async def _try_lm_classify_workbooks(
    previews: list[dict[str, Any]],
    *,
    timeout: int | float | None = None,
) -> dict[str, WorkbookRole] | None:
    payload = _lm_settings()
    if payload is None:
        return None
    base_url, model = payload
    prompt = (
        "Ты классифицируешь Excel-файлы для агента закупок Авион. "
        "На входе превью: имя файла, названия листов, первые строки таблиц. "
        "Для КАЖДОГО файла выбери РОВНО ОДНУ из четырёх ролей (других нет):\n"
        "1) shipment_schedule — график отгрузок материалов и всё, что к нему относится: "
        "номенклатура + колонки дат поставки, логистика до МСК/Ростов, заказано/остаток; "
        "файлы с листами «ТАМОЖНЯ», «ИТЦ В РАБОТЕ», «Реестр Заказов» (позиция+модель, партии, "
        "таможенные операции) — это тоже shipment_schedule, дополнение к отгрузкам;\n"
        "2) stock — остатки МАТЕРИАЛОВ на складе: «Номенклатура» + «Остаток…» "
        "(даже если сверху есть небольшой план изделий);\n"
        "3) production_schedule — помесячный график производства "
        "(колонки месяцев, «Наименования изделий», Заказ/Опытные/Склад × План/Факт, "
        "без дневной/недельной сетки);\n"
        "4) detailed_production_schedule — детальный график/отчёт выпуска по дням или неделям "
        "(«График выпуска готовой продукции», колонки-даты, отчёт «Модель/изделие» П/ф·ОТК·Склад).\n"
        "Имя файла может вводить в заблуждение — опирайся на структуру листов и заголовков. "
        "Не используй роли specification, customs, other.\n"
        "Верни строго JSON: {\"files\":[{\"filename\":\"...\",\"role\":\"...\",\"reason\":\"...\"}]}.\n\n"
        f"FILES={json.dumps(previews, ensure_ascii=False)}"
    )
    try:
        data = await _post_lm_json(
            base_url,
            model,
            prompt,
            timeout=timeout if timeout is not None else settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS,
        )
        files = data.get("files")
        if not isinstance(files, list):
            return None
        result: dict[str, WorkbookRole] = {}
        for item in files:
            if not isinstance(item, dict):
                continue
            filename = _clean_text(item.get("filename"))
            role = _normalize_lm_role(item.get("role"))
            if filename and role:
                result[filename] = role
        return result or None
    except Exception as exc:
        logger.warning("document_analysis_agent.classification_lm_failed", error=str(exc))
        return None


def _classify_preview_locally(preview: dict[str, Any]) -> WorkbookRole:
    """Fallback если LM недоступна — только по превью содержимого."""
    if _preview_looks_like_shipment_schedule(preview):
        return ROLE_SHIPMENT_SCHEDULE
    if _preview_looks_like_stock(preview):
        return ROLE_STOCK
    if _preview_looks_like_detailed_production_schedule(preview):
        return ROLE_DETAILED_PRODUCTION_SCHEDULE
    if _preview_looks_like_production_schedule(preview):
        return ROLE_PRODUCTION_SCHEDULE
    return ROLE_OTHER


def _normalize_lm_role(value: Any) -> WorkbookRole:
    role = _normalize(value)
    aliases = {
        "stock": ROLE_STOCK,
        "inventory": ROLE_STOCK,
        "остатки": ROLE_STOCK,
        "остаток": ROLE_STOCK,
        "production_schedule": ROLE_PRODUCTION_SCHEDULE,
        "monthly_production_schedule": ROLE_PRODUCTION_SCHEDULE,
        "график производства": ROLE_PRODUCTION_SCHEDULE,
        "detailed_production_schedule": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "detailed_schedule": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "weekly_production_schedule": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "daily_production_schedule": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "production_output_schedule": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "детальный график производства": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "детальный план производства": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "график выпуска": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "план по недельно": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "отчет производства": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "отчёт производства": ROLE_DETAILED_PRODUCTION_SCHEDULE,
        "shipment_schedule": ROLE_SHIPMENT_SCHEDULE,
        "shipping_schedule": ROLE_SHIPMENT_SCHEDULE,
        "delivery_schedule": ROLE_SHIPMENT_SCHEDULE,
        "график отгрузок": ROLE_SHIPMENT_SCHEDULE,
        "график отгрузки": ROLE_SHIPMENT_SCHEDULE,
        "customs_itc": ROLE_SHIPMENT_SCHEDULE,
        "customs": ROLE_SHIPMENT_SCHEDULE,
        "таможня": ROLE_SHIPMENT_SCHEDULE,
        "itc": ROLE_SHIPMENT_SCHEDULE,
        "итц": ROLE_SHIPMENT_SCHEDULE,
        "specs": ROLE_OTHER,
        "spec": ROLE_OTHER,
        "specification": ROLE_OTHER,
        "спецификация": ROLE_OTHER,
        "спецификации": ROLE_OTHER,
        "unknown": ROLE_OTHER,
        "other": ROLE_OTHER,
    }
    mapped = aliases.get(role, ROLE_OTHER)
    return mapped if mapped in UPLOAD_FILE_ROLES or mapped == ROLE_OTHER else ROLE_OTHER


def _lm_settings() -> tuple[str, str] | None:
    base_url = settings.AVEON_LM_STUDIO_BASE_URL.strip().rstrip("/")
    model = settings.AVEON_LM_STUDIO_MODEL.strip()
    if not base_url or not model:
        return None
    # Windows + httpx: localhost → ::1 иногда даёт OSError Errno 22 Invalid argument.
    if "://localhost" in base_url or "://localhost:" in base_url:
        base_url = base_url.replace("://localhost", "://127.0.0.1", 1)
    return base_url, model


async def _post_lm_json(base_url: str, model: str, prompt: str, timeout: int | float) -> dict[str, Any]:
    # connect отдельно: иначе при «мёртвом» LM Studio ждём весь timeout на TCP.
    client_timeout = httpx.Timeout(timeout, connect=min(5.0, float(timeout)))
    async with httpx.AsyncClient(timeout=client_timeout, trust_env=False) as client:
        response = await client.post(
            f"{base_url}/chat/completions",
            json={
                "model": model,
                "messages": [
                    {
                        "role": "system",
                        "content": "Ты возвращаешь только валидный JSON. Не используй markdown и пояснения.",
                    },
                    {"role": "user", "content": prompt},
                ],
                "temperature": 0.1,
                "stream": False,
            },
        )
        response.raise_for_status()
        data = response.json()
        content = data["choices"][0]["message"]["content"]
        try:
            return _parse_json_object(content)
        except json.JSONDecodeError:
            repair_response = await client.post(
                f"{base_url}/chat/completions",
                json={
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "Ты исправляешь ответ в валидный JSON без markdown и пояснений.",
                        },
                        {"role": "user", "content": prompt},
                        {"role": "assistant", "content": content},
                        {
                            "role": "user",
                            "content": (
                                "Предыдущий ответ был невалидным JSON. "
                                "Верни только исправленный JSON с теми же данными."
                            ),
                        },
                    ],
                    "temperature": 0,
                    "stream": False,
                },
            )
            repair_response.raise_for_status()
            repair_data = repair_response.json()
            repair_content = repair_data["choices"][0]["message"]["content"]
            return _parse_json_object(repair_content)


def _parse_json_object(content: str) -> dict[str, Any]:
    text = content.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?", "", text).strip()
        text = re.sub(r"```$", "", text).strip()
    return json.loads(text)


def _empty_month_bucket() -> dict[str, dict[str, float]]:
    return {cat: {metric: 0.0 for metric in _SCHEDULE_METRICS} for cat in _SCHEDULE_CATEGORIES}


def _round_qty(value: float) -> float:
    return round(value, 6) if abs(value - round(value)) > 1e-9 else float(round(value))


def _plan_demand_total(month_bucket: dict[str, dict[str, float]] | None) -> float:
    """Сумма плановых потребностей по категориям (для прогноза)."""
    if not month_bucket:
        return 0.0
    total = 0.0
    for category in _SCHEDULE_CATEGORIES:
        total += float(month_bucket.get(category, {}).get("план", 0.0))
    return total


def _monthly_demand_has_nonzero(demand: dict[str, dict[str, dict[str, float]]]) -> bool:
    return any(
        float(qty) > 0
        for month_bucket in demand.values()
        for metrics in month_bucket.values()
        for qty in metrics.values()
    )


def _classify_schedule_category(value: Any) -> str | None:
    text = _normalize(value)
    if not text:
        return None
    if "опытн" in text:
        return "опытные"
    if text == "заказ" or text.startswith("заказ"):
        return "заказ"
    if text == "склад" or text.startswith("склад"):
        return "склад"
    return None


def _classify_schedule_metric(value: Any) -> str | None:
    text = _normalize(value)
    if text == "план" or text.startswith("план"):
        return "план"
    if text == "факт" or text.startswith("факт"):
        return "факт"
    return None


def _sheet_cell_value(sheet: Worksheet, row: int, col: int) -> Any:
    """Значение ячейки с учётом merge (значение только в левом верхнем углу)."""
    value = sheet.cell(row, col).value
    if value is not None and value != "":
        return value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col).value
    return value


def _safe_set_cell_value(sheet: Worksheet, row: int, col: int, value: Any) -> Any:
    """Записать value, разъединив merge если ячейка read-only (не верхняя левая)."""
    from openpyxl.cell.cell import MergedCell

    cell = sheet.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged in list(sheet.merged_cells.ranges):
            if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
                sheet.unmerge_cells(str(merged))
        cell = sheet.cell(row, col)
    cell.value = value
    return cell


def _filled_month_labels(sheet: Worksheet, row: int, max_col: int) -> dict[int, str]:
    """Месяц для каждой колонки (propagate из merge / влево)."""
    labels: dict[int, str] = {}
    last = ""
    for col in range(1, max_col + 1):
        month_num = _month_number_from_header(_normalize(_sheet_cell_value(sheet, row, col)))
        if month_num is not None:
            last = _MONTH_NOMINATIVE[month_num - 1]
        if last:
            labels[col] = last
    return labels


def _find_production_schedule_layout(sheet: Worksheet) -> ScheduleTableLayout | None:
    """Ищет помесячную таблицу: split (3 уровня) или legacy (1 qty на месяц)."""
    max_col = min(sheet.max_column or 1, 80)
    max_scan = min(sheet.max_row or 1, 40)
    legacy_candidate: ScheduleTableLayout | None = None

    for month_row in range(1, max_scan + 1):
        headers = {
            col: _normalize(_sheet_cell_value(sheet, month_row, col))
            for col in range(1, max_col + 1)
        }
        name_col = _pick_column(
            headers,
            [
                "наименования изделий",
                "наименование изделий",
                "наименование",
                "изделие",
                "контракты / изделия",
                "контракты/изделия",
            ],
        )
        month_labels = _filled_month_labels(sheet, month_row, max_col)
        distinct_months = list(dict.fromkeys(month_labels.values()))
        if name_col is None or len(distinct_months) < 2:
            continue

        # --- split: месяц / категория / план|факт ---
        cat_row = month_row + 1
        metric_row = month_row + 2
        if metric_row <= (sheet.max_row or 0):
            split_cols: list[ScheduleQtyColumn] = []
            last_category: str | None = None
            last_category_month: str | None = None
            for col in range(1, max_col + 1):
                month = month_labels.get(col)
                if not month:
                    last_category = None
                    last_category_month = None
                    continue
                category = _classify_schedule_category(
                    _sheet_cell_value(sheet, cat_row, col)
                )
                if category is not None:
                    last_category = category
                    last_category_month = month
                elif last_category is not None and last_category_month == month:
                    category = last_category
                metric = _classify_schedule_metric(
                    _sheet_cell_value(sheet, metric_row, col)
                )
                if category is None or metric is None:
                    continue
                split_cols.append(
                    ScheduleQtyColumn(
                        col=col, month=month, category=category, metric=metric
                    )
                )
            if len(split_cols) >= 6:
                return ScheduleTableLayout(
                    name_col=name_col,
                    data_start_row=metric_row + 1,
                    columns=split_cols,
                    is_split=True,
                )

        # --- legacy: одна колонка на месяц → заказ/план ---
        month_cols: list[ScheduleQtyColumn] = []
        seen_month: set[str] = set()
        for col, month in month_labels.items():
            header = headers.get(col, "")
            if _month_number_from_header(header) is None:
                continue
            if month in seen_month:
                continue
            seen_month.add(month)
            month_cols.append(
                ScheduleQtyColumn(col=col, month=month, category="заказ", metric="план")
            )
        if len(month_cols) >= 2 and legacy_candidate is None:
            data_cols = [item.col for item in month_cols]
            legacy_candidate = ScheduleTableLayout(
                name_col=name_col,
                data_start_row=month_row + 1,
                columns=month_cols,
                is_split=False,
            )
            # продолжаем искать split на других строках/не берём сразу
            _ = data_cols

    return legacy_candidate


def _extract_production_schedule_products(
    workbooks: list[UploadedWorkbook], role_map: dict[str, WorkbookRole]
) -> tuple[list[str], list[ScheduleProductPlan]]:
    schedule_files = [
        uploaded.filename
        for uploaded in workbooks
        if role_map.get(uploaded.filename) == ROLE_PRODUCTION_SCHEDULE
    ]
    plans: list[ScheduleProductPlan] = []
    seen: set[str] = set()

    for uploaded in workbooks:
        if role_map.get(uploaded.filename) != ROLE_PRODUCTION_SCHEDULE:
            continue
        workbook = load_workbook(BytesIO(uploaded.content), data_only=True)
        try:
            chosen: tuple[Worksheet, ScheduleTableLayout] | None = None
            legacy: tuple[Worksheet, ScheduleTableLayout] | None = None
            for sheet in workbook.worksheets:
                layout = _find_production_schedule_layout(sheet)
                if layout is None:
                    continue
                if layout.is_split:
                    chosen = (sheet, layout)
                    break
                if legacy is None:
                    legacy = (sheet, layout)
            if chosen is None:
                chosen = legacy
            if chosen is None:
                continue

            sheet, layout = chosen
            data_cols = [item.col for item in layout.columns]
            for row_idx in range(layout.data_start_row, sheet.max_row + 1):
                product_name = _clean_text(sheet.cell(row_idx, layout.name_col).value)
                if not _is_schedule_product_name(product_name):
                    continue
                if not any(
                    _cell_has_number(sheet.cell(row_idx, col_idx).value)
                    for col_idx in data_cols
                ):
                    continue
                key = _normalize(product_name)
                if key in seen:
                    continue
                seen.add(key)
                monthly_qty: dict[str, dict[str, dict[str, float]]] = {}
                for item in layout.columns:
                    bucket = monthly_qty.setdefault(item.month, _empty_month_bucket())
                    qty = _to_float(sheet.cell(row_idx, item.col).value)
                    bucket[item.category][item.metric] = float(qty) if qty is not None else 0.0
                plans.append(ScheduleProductPlan(product=product_name, monthly_qty=monthly_qty))
            logger.info(
                "document_analysis_agent.production_schedule_layout",
                file=uploaded.filename,
                sheet=sheet.title,
                split=layout.is_split,
                columns=len(layout.columns),
                products=len(plans),
            )
        finally:
            workbook.close()

    return schedule_files, plans


def _extract_detailed_production_schedule(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    as_of: date | None = None,
) -> DetailedScheduleExtract:
    """Извлекает дневной план изделий из файлов роли detailed_production_schedule."""
    as_of_day = as_of or date.today()
    files = [
        uploaded.filename
        for uploaded in workbooks
        if role_map.get(uploaded.filename) == ROLE_DETAILED_PRODUCTION_SCHEDULE
    ]
    if not files:
        year, month = as_of_day.year, as_of_day.month
        return DetailedScheduleExtract(
            files=[],
            plans=[],
            year=year,
            month=month,
            day_keys=_month_day_keys(year, month),
        )

    sheet_candidates: list[_DetailedSheetCandidate] = []
    for uploaded in workbooks:
        if role_map.get(uploaded.filename) != ROLE_DETAILED_PRODUCTION_SCHEDULE:
            continue
        workbook = load_workbook(BytesIO(uploaded.content), data_only=True)
        try:
            for sheet in workbook.worksheets:
                year, month = _infer_detailed_sheet_year_month(sheet, as_of_day.year)
                if year <= 0 or month <= 0:
                    continue
                plans, plan_cells = _parse_detailed_schedule_sheet(sheet, year, month)
                if not plans and not _sheet_has_daily_day_columns(sheet, year, month):
                    continue
                sheet_candidates.append(
                    _DetailedSheetCandidate(
                        year=year,
                        month=month,
                        filename=uploaded.filename,
                        sheet_name=sheet.title,
                        content=uploaded.content,
                        plans=plans,
                        plan_cells=plan_cells,
                    )
                )
        finally:
            workbook.close()

    if not sheet_candidates:
        year, month = as_of_day.year, as_of_day.month
        logger.warning(
            "document_analysis_agent.detailed_schedule_empty",
            files=files,
            fallback_month=f"{year:04d}-{month:02d}",
        )
        return DetailedScheduleExtract(
            files=files,
            plans=[],
            year=year,
            month=month,
            day_keys=_month_day_keys(year, month),
        )

    chosen = _choose_detailed_schedule_month(sheet_candidates, as_of_day)
    year, month = chosen.year, chosen.month
    # схлопнуть одноимённые изделия с разных блоков/файлов
    merged_plans = _merge_detailed_product_plans(chosen.plans)
    day_keys = _month_day_keys(year, month)
    for plan in merged_plans:
        for day_key in day_keys:
            plan.daily_qty.setdefault(day_key, 0.0)
        plan.year = year
        plan.month = month

    logger.info(
        "document_analysis_agent.detailed_schedule_extracted",
        files=files,
        month=f"{year:04d}-{month:02d}",
        products=len(merged_plans),
        nonzero_days=sum(
            1 for plan in merged_plans for qty in plan.daily_qty.values() if qty > 0
        ),
        source_sheet=chosen.sheet_name,
        plan_cells=len(chosen.plan_cells),
    )
    return DetailedScheduleExtract(
        files=files,
        plans=merged_plans,
        year=year,
        month=month,
        day_keys=day_keys,
        source_filename=chosen.filename,
        source_sheet_name=chosen.sheet_name,
        source_bytes=chosen.content,
        plan_cells=list(chosen.plan_cells),
    )


def _month_day_keys(year: int, month: int) -> list[str]:
    if year <= 0 or month <= 0:
        return []
    days = monthrange(year, month)[1]
    return [date(year, month, day).isoformat() for day in range(1, days + 1)]


def _year_for_sheet_month(month: str, months_ordered: list[str], as_of: date | None = None) -> int:
    """Год подписи месяца на листе (перенос через январь → +1)."""
    as_of = as_of or date.today()
    year = as_of.year
    prev_idx = -1
    for name in months_ordered:
        idx = _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else -1
        if idx >= 0 and prev_idx >= 0 and idx < prev_idx:
            year += 1
        if name == month:
            return year
        if idx >= 0:
            prev_idx = idx
    return as_of.year


def _calendar_weeks_in_month(year: int, month: int) -> list[_MonthWeekSpan]:
    """Недели Пн–Вс внутри месяца (хвосты соседних месяцев обрезаются)."""
    if year <= 0 or not (1 <= month <= 12):
        return []
    first = date(year, month, 1)
    last = date(year, month, monthrange(year, month)[1])
    cursor = first - timedelta(days=first.weekday())  # понедельник
    spans: list[_MonthWeekSpan] = []
    while cursor <= last:
        week_end = cursor + timedelta(days=6)
        seg_start = max(cursor, first)
        seg_end = min(week_end, last)
        if seg_start <= seg_end and seg_start.month == month:
            spans.append(
                _MonthWeekSpan(
                    key=f"{seg_start.isoformat()}_{seg_end.isoformat()}",
                    label=(
                        f"{seg_start.day:02d}.{seg_start.month:02d}"
                        f"–{seg_end.day:02d}.{seg_end.month:02d}"
                    ),
                    start=seg_start,
                    end=seg_end,
                )
            )
        cursor += timedelta(days=7)
    return spans


def _qty_in_week_span(daily_qty: dict[str, float], span: _MonthWeekSpan) -> float:
    total = 0.0
    day = span.start
    while day <= span.end:
        total += float(daily_qty.get(day.isoformat(), 0.0))
        day += timedelta(days=1)
    return total


def _weeks_by_month_labels(months: list[str], as_of: date | None = None) -> dict[str, list[_MonthWeekSpan]]:
    result: dict[str, list[_MonthWeekSpan]] = {}
    for month in months:
        if month not in _MONTH_NOMINATIVE:
            result[month] = []
            continue
        year = _year_for_sheet_month(month, months, as_of)
        month_num = _MONTH_NOMINATIVE.index(month) + 1
        result[month] = _calendar_weeks_in_month(year, month_num)
    return result


def _infer_detailed_sheet_year_month(sheet: Worksheet, default_year: int) -> tuple[int, int]:
    """Год/месяц листа: имя листа, заголовок, либо даты в шапке дневных колонок."""
    month = _month_number_from_header(_normalize(sheet.title))
    year = default_year
    title_year = _year_from_text(sheet.title)
    if title_year:
        year = title_year

    for row_idx in range(1, min(sheet.max_row, 24) + 1):
        for col_idx in range(1, min(sheet.max_column, 16) + 1):
            text = _clean_text(sheet.cell(row_idx, col_idx).value)
            if not text:
                continue
            if month is None:
                month = _month_number_from_header(_normalize(text))
            found_year = _year_from_text(text)
            if found_year:
                year = found_year

    # даты в шапке имеют приоритет для месяца/года (month=0 — без фильтра по месяцу)
    pf_days = _pf_report_anchor_dates(sheet, year, 0)
    if pf_days:
        return pf_days[0].year, pf_days[0].month

    text_year = year
    for header_idx, _name_col, day_cols in _iter_detailed_schedule_tables(sheet, year, month or 0):
        if not day_cols:
            continue
        first_day = day_cols[0][1]
        if month is None:
            month = first_day.month
        if text_year < 2000:
            year = first_day.year
        break

    if month is None:
        return 0, 0
    return year, month


def _year_from_text(value: Any) -> int | None:
    text = _clean_text(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        candidate = int(value)
        if 2020 <= candidate <= 2035:
            return candidate
        return None
    match = re.search(r"\b(20\d{2})\b", text)
    if not match:
        return None
    year = int(match.group(1))
    if year < 2020 or year > 2035:
        return None
    return year


def _sheet_has_daily_day_columns(sheet: Worksheet, year: int, month: int) -> bool:
    if _sheet_is_pf_stage_report(sheet) and _pf_report_anchor_dates(sheet, year, month):
        return True
    for _header_idx, _name_col, day_cols in _iter_detailed_schedule_tables(sheet, year, month):
        if day_cols:
            return True
    return False


@dataclass
class _DetailedSheetCandidate:
    year: int
    month: int
    filename: str
    sheet_name: str
    content: bytes
    plans: list[DetailedScheduleProductPlan]
    plan_cells: list[DetailedPlanCellRef] = field(default_factory=list)


def _choose_detailed_schedule_month(
    candidates: list[_DetailedSheetCandidate],
    as_of: date,
) -> _DetailedSheetCandidate:
    """Месяц asOf → ближайший будущий → первый кандидат."""
    by_month: dict[tuple[int, int], list[_DetailedSheetCandidate]] = {}
    for item in candidates:
        by_month.setdefault((item.year, item.month), []).append(item)

    target = (as_of.year, as_of.month)
    if target in by_month:
        return _flatten_month_candidates(by_month[target])

    future = sorted(
        [(y, m) for (y, m) in by_month if (y, m) > target],
    )
    if future:
        return _flatten_month_candidates(by_month[future[0]])

    past_or_any = sorted(by_month.keys())
    return _flatten_month_candidates(by_month[past_or_any[0]])


def _flatten_month_candidates(
    items: list[_DetailedSheetCandidate],
) -> _DetailedSheetCandidate:
    """Сливает планы всех листов месяца; для копии/окраски берёт первый лист."""
    primary = items[0]
    plans: list[DetailedScheduleProductPlan] = []
    for item in items:
        plans.extend(item.plans)
    return _DetailedSheetCandidate(
        year=primary.year,
        month=primary.month,
        filename=primary.filename,
        sheet_name=primary.sheet_name,
        content=primary.content,
        plans=plans,
        plan_cells=list(primary.plan_cells),
    )


def _merge_detailed_product_plans(
    plans: list[DetailedScheduleProductPlan],
) -> list[DetailedScheduleProductPlan]:
    merged: dict[str, DetailedScheduleProductPlan] = {}
    for plan in plans:
        key = _normalize(plan.product)
        existing = merged.get(key)
        if existing is None:
            merged[key] = DetailedScheduleProductPlan(
                product=plan.product,
                daily_qty=dict(plan.daily_qty),
                daily_fact=dict(plan.daily_fact),
                year=plan.year,
                month=plan.month,
            )
            continue
        for day_key, qty in plan.daily_qty.items():
            existing.daily_qty[day_key] = existing.daily_qty.get(day_key, 0.0) + float(qty)
        for day_key, qty in plan.daily_fact.items():
            existing.daily_fact[day_key] = existing.daily_fact.get(day_key, 0.0) + float(qty)
    return list(merged.values())


def _iter_detailed_schedule_tables(
    sheet: Worksheet,
    year: int,
    month: int,
) -> list[tuple[int, int, list[tuple[int, date]]]]:
    """Все блоки таблицы на листе: (header_row, name_col, [(col, date), …])."""
    tables: list[tuple[int, int, list[tuple[int, date]]]] = []
    max_scan = min(sheet.max_row, 120)
    for header_idx in range(1, max_scan + 1):
        name_col = None
        day_cols: list[tuple[int, date]] = []
        for col_idx in range(1, sheet.max_column + 1):
            value = sheet.cell(header_idx, col_idx).value
            text = _normalize(value)
            if name_col is None and text in {
                "наименование",
                "наимнование",
                "наименования",
                "наименования изделий",
                "наименование изделий",
                "изделие",
            }:
                name_col = col_idx
                continue
            day_date = _detailed_header_to_day(value, year, month)
            if day_date is not None:
                day_cols.append((col_idx, day_date))

        # шапка дат может быть на строке ниже «Наименование» (как в Апреле)
        if name_col is not None and not day_cols and header_idx < max_scan:
            for col_idx in range(1, sheet.max_column + 1):
                day_date = _detailed_header_to_day(
                    sheet.cell(header_idx + 1, col_idx).value, year, month
                )
                if day_date is not None:
                    day_cols.append((col_idx, day_date))
            if day_cols:
                tables.append((header_idx + 1, name_col, day_cols))
                continue

        if name_col is not None and len(day_cols) >= 2:
            tables.append((header_idx, name_col, day_cols))
    return tables


def _detailed_header_to_day(value: Any, year: int, month: int) -> date | None:
    """Колонка дня: дата, либо номер 1..31; skip Итог/План/Отклонение/Остаток."""
    text = _normalize(value)
    if text:
        if text in _DETAILED_DAY_SKIP_HEADERS:
            return None
        if re.fullmatch(r"\d{1,2}", text) is None:
            if any(
                token in text
                for token in ("итог", "факт", "оклон", "отклон", "остаток")
            ):
                return None
            if text == "план" or text.startswith("план "):
                return None

    # datetime/date в шапке (лист «Апрель») — до разбора чисел 1..31
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        parsed = None

    if parsed is not None:
        if year > 0 and month > 0:
            # Excel иногда отдаёт «3» как 2000-08-03 — день месяца важнее года ячейки
            if parsed.month == month and 1 <= parsed.day <= monthrange(year, month)[1]:
                return date(year, month, parsed.day)
            if parsed.year != year or parsed.month != month:
                return None
        return parsed

    # Номер дня месяца: НЕ через from_excel(1) → 1899/1900
    day_num: int | None = None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            candidate = int(value)
            if 1 <= candidate <= 31:
                day_num = candidate
    else:
        text_raw = _clean_text(value)
        if re.fullmatch(r"\d{1,2}", text_raw):
            day_num = int(text_raw)

    if day_num is not None:
        if year <= 0 or month <= 0:
            return None
        max_day = monthrange(year, month)[1]
        if day_num > max_day:
            return None
        return date(year, month, day_num)

    # Строковые даты «2026-04-01» / «01.04.2026»
    parsed = _header_value_to_date(value)
    if parsed is None:
        # «11.07» / «11.07.» без года — день месяца шаблона
        text_raw = _clean_text(value)
        match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.?", text_raw)
        if match and year > 0:
            day_n = int(match.group(1))
            month_n = int(match.group(2))
            use_year = year
            if month > 0 and month_n != month:
                # если месяц в шапке другой — доверяем шапке
                pass
            try:
                parsed = date(use_year, month_n, day_n)
            except ValueError:
                parsed = None
            if parsed is not None and month > 0 and parsed.month != month:
                return None
            return parsed
        return None
    if year > 0 and month > 0 and (parsed.year != year or parsed.month != month):
        return None
    return parsed


def _parse_detailed_schedule_sheet(
    sheet: Worksheet,
    year: int,
    month: int,
) -> tuple[list[DetailedScheduleProductPlan], list[DetailedPlanCellRef]]:
    # Новый формат отчёта: Модель/изделие + стадии П/ф·ОТК·Склад + план/факт
    if _sheet_is_pf_stage_report(sheet):
        pf_plans, pf_cells = _parse_pf_stage_report_sheet(sheet, year, month)
        if pf_plans:
            return pf_plans, pf_cells

    plans: list[DetailedScheduleProductPlan] = []
    plan_cells: list[DetailedPlanCellRef] = []
    seen_rows: set[tuple[int, str]] = set()
    tables = _iter_detailed_schedule_tables(sheet, year, month)
    for header_idx, name_col, day_cols in tables:
        # строки данных до следующего блока шапки или пустой зоны
        next_headers = sorted(h for h, _, _ in tables if h > header_idx)
        end_row = next_headers[0] if next_headers else sheet.max_row + 1
        # если шапка дат на header_idx, а «Наименование» на строке выше — данные с header_idx+1
        start_row = header_idx + 1
        for row_idx in range(start_row, end_row):
            product_name = _clean_text(sheet.cell(row_idx, name_col).value)
            if not _is_schedule_product_name(product_name):
                continue
            # пропуск повторных шапок «Наименование»
            if _normalize(product_name) in {
                "наименование",
                "наимнование",
                "наименования",
            }:
                continue
            row_key = (row_idx, _normalize(product_name))
            if row_key in seen_rows:
                continue
            # есть ли хоть одно число в дневных колонках (или текстовый мусор — skip qty)
            has_qty = False
            daily_qty: dict[str, float] = {}
            for col_idx, day in day_cols:
                raw = sheet.cell(row_idx, col_idx).value
                qty = _to_float(raw)
                if qty is None:
                    # текст вроде «Заказ комплектующих» — 0
                    daily_qty[day.isoformat()] = 0.0
                    plan_cells.append(
                        DetailedPlanCellRef(
                            product=product_name,
                            row=row_idx,
                            plan_col=col_idx,
                            day_keys=[day.isoformat()],
                            plan_qty=0.0,
                        )
                    )
                    continue
                has_qty = has_qty or qty != 0
                daily_qty[day.isoformat()] = float(qty)
                plan_cells.append(
                    DetailedPlanCellRef(
                        product=product_name,
                        row=row_idx,
                        plan_col=col_idx,
                        day_keys=[day.isoformat()],
                        plan_qty=float(qty),
                    )
                )
            if not daily_qty:
                continue
            # строка без чисел вообще (кроме полностью пустой) всё равно берём с нулями,
            # если это похоже на изделие и рядом есть № п/п
            pp_val = sheet.cell(row_idx, 1).value
            if not has_qty and not _cell_has_number(pp_val):
                continue
            seen_rows.add(row_key)
            plans.append(
                DetailedScheduleProductPlan(
                    product=product_name,
                    daily_qty=daily_qty,
                    year=year,
                    month=month,
                )
            )
    return plans, plan_cells


@dataclass(frozen=True)
class _PfMetricPair:
    """Пара колонок план/факт: один день либо диапазон дат."""

    plan_col: int
    fact_col: int | None
    day: date | None = None
    range_start: date | None = None
    range_end: date | None = None

    def days(self) -> list[date]:
        if self.day is not None:
            return [self.day]
        if self.range_start is None or self.range_end is None:
            return []
        if self.range_end < self.range_start:
            return []
        out: list[date] = []
        cursor = self.range_start
        while cursor <= self.range_end:
            out.append(cursor)
            cursor += timedelta(days=1)
        return out


@dataclass(frozen=True)
class _DetailedDayColumn:
    """Одна дневная колонка обычного «Графика выпуска»."""

    col: int
    day: date


@dataclass
class _DetailedStageSeries:
    """Одна строка выпуска изделия из детального отчёта, до выбора рабочей стадии."""

    product: str
    stage: str
    row: int
    priority: int
    daily_qty: dict[str, float] = field(default_factory=dict)
    daily_fact: dict[str, float] = field(default_factory=dict)
    plan_cells: list[DetailedPlanCellRef] = field(default_factory=list)

    @property
    def total_qty(self) -> float:
        return sum(float(value) for value in self.daily_qty.values())

    @property
    def total_fact(self) -> float:
        return sum(float(value) for value in self.daily_fact.values())

    @property
    def has_nonzero(self) -> bool:
        return self.total_qty > 1e-12 or self.total_fact > 1e-12


def _sheet_is_pf_stage_report(sheet: Worksheet) -> bool:
    """Отчёт план/факт по изделиям: со стадиями или без явного П/ф."""
    max_r = min(sheet.max_row, 60)
    max_c = min(sheet.max_column, 12)
    has_pf = False
    has_model = False
    plan_hits = 0
    fact_hits = 0
    for row_idx in range(1, max_r + 1):
        for col_idx in range(1, max_c + 1):
            text = _normalize(sheet.cell(row_idx, col_idx).value)
            if not text:
                continue
            if text in _PF_STAGE_KEYS or text.startswith("п/ф"):
                has_pf = True
            if "модель" in text and "изделие" in text:
                has_model = True
            metric = _classify_schedule_metric(text)
            if metric == "план":
                plan_hits += 1
            elif metric == "факт":
                fact_hits += 1
            if has_pf and has_model:
                return True
    return has_pf or (has_model and plan_hits >= 2 and fact_hits >= 1)


def _pf_header_is_skip(value: Any) -> bool:
    text = _normalize(value)
    if not text:
        return False
    return any(token in text for token in _PF_SKIP_COLUMN_TOKENS)


def _parse_pf_date_range(
    value: Any,
    *,
    default_year: int = 0,
    default_month: int = 0,
) -> tuple[date, date] | None:
    text = _clean_text(value)
    if not text:
        return None
    match = _PF_DATE_RANGE_RE.search(text)
    if not match:
        return None

    def _y(raw: str | None, fallback_month: int) -> int:
        if raw:
            year_n = int(raw)
            return year_n + 2000 if year_n < 100 else year_n
        if default_year > 0:
            # период без года: если месяц диапазона < текущего месяца шаблона — следующий год
            if default_month > 0 and fallback_month < default_month:
                return default_year + 1
            return default_year
        return date.today().year

    try:
        start_month = int(match.group(2))
        end_month = int(match.group(5))
        start = date(_y(match.group(3), start_month), start_month, int(match.group(1)))
        end = date(_y(match.group(6), end_month), end_month, int(match.group(4)))
    except ValueError:
        return None
    return start, end


def _find_pf_report_layout(
    sheet: Worksheet,
    year: int,
    month: int,
) -> tuple[int, int, int, list[_PfMetricPair]] | None:
    """(metric_row, name_col, stage_col, pairs) для отчёта план/факт."""
    metric_row: int | None = None
    plan_hits = 0
    for row_idx in range(1, min(sheet.max_row, 20) + 1):
        hits = 0
        for col_idx in range(1, sheet.max_column + 1):
            if _classify_schedule_metric(sheet.cell(row_idx, col_idx).value) == "план":
                hits += 1
        if hits > plan_hits:
            plan_hits = hits
            metric_row = row_idx
    if metric_row is None or plan_hits < 2:
        return None

    date_row = metric_row - 1 if metric_row > 1 else metric_row
    name_col: int | None = None
    stage_col: int | None = None
    for col_idx in range(1, min(sheet.max_column, 8) + 1):
        header = _normalize(sheet.cell(metric_row, col_idx).value)
        if name_col is None and (
            ("модель" in header and "изделие" in header)
            or header in {"наименование", "изделие", "модель"}
        ):
            name_col = col_idx
    if name_col is None:
        # часто «Модель / изделие» только в строке метрик, col B
        name_col = 2

    # колонка стадии: ищем «П/ф» в первых строках данных
    for row_idx in range(metric_row + 1, min(sheet.max_row, metric_row + 40) + 1):
        for col_idx in range(1, min(sheet.max_column, 6) + 1):
            if _normalize(sheet.cell(row_idx, col_idx).value) in _PF_STAGE_KEYS:
                stage_col = col_idx
                break
        if stage_col is not None:
            break
    if stage_col is None:
        stage_col = 3

    pairs: list[_PfMetricPair] = []
    col_idx = 1
    while col_idx <= sheet.max_column:
        metric = _classify_schedule_metric(sheet.cell(metric_row, col_idx).value)
        if metric != "план":
            col_idx += 1
            continue
        date_header = sheet.cell(date_row, col_idx).value
        # значение даты может быть в merge — только в левой ячейке; иначе смотрим sheet merge
        if date_header is None:
            date_header = _sheet_cell_value(sheet, date_row, col_idx)
        metric_header = sheet.cell(metric_row, col_idx).value
        if _pf_header_is_skip(date_header) or _pf_header_is_skip(metric_header):
            col_idx += 1
            continue

        fact_col: int | None = None
        if col_idx + 1 <= sheet.max_column:
            if _classify_schedule_metric(sheet.cell(metric_row, col_idx + 1).value) == "факт":
                fact_col = col_idx + 1

        range_pair = _parse_pf_date_range(
            date_header, default_year=year, default_month=month
        )
        day = None if range_pair else _detailed_header_to_day(date_header, year, month)
        if range_pair is None and day is None:
            # пустая шапка даты над «план» — не дневная колонка
            col_idx += 1
            continue
        # фильтр месяца: диапазон/день должны пересекаться с целевым месяцем
        if range_pair is not None:
            start, end = range_pair
            if year > 0 and month > 0:
                month_start = date(year, month, 1)
                month_end = date(year, month, monthrange(year, month)[1])
                if end < month_start or start > month_end:
                    col_idx = (fact_col + 1) if fact_col else col_idx + 1
                    continue
            pairs.append(
                _PfMetricPair(
                    plan_col=col_idx,
                    fact_col=fact_col,
                    range_start=start,
                    range_end=end,
                )
            )
        else:
            assert day is not None
            pairs.append(_PfMetricPair(plan_col=col_idx, fact_col=fact_col, day=day))

        col_idx = (fact_col + 1) if fact_col else col_idx + 1

    if not pairs:
        return None
    return metric_row, name_col, stage_col, pairs


def _pf_report_anchor_dates(sheet: Worksheet, year: int, month: int) -> list[date]:
    layout = _find_pf_report_layout(sheet, year, month)
    if layout is None:
        return []
    _metric_row, _name_col, _stage_col, pairs = layout
    dates: list[date] = []
    for pair in pairs:
        days = pair.days()
        if days:
            dates.append(days[0])
    return dates


def _distribute_qty_across_days(total: float, days: list[date]) -> dict[str, float]:
    """Равномерное распределение qty по дням (целые — с остатком на первые дни)."""
    if not days:
        return {}
    n = len(days)
    if abs(total - round(total)) < 1e-9:
        total_i = int(round(total))
        base, rem = divmod(abs(total_i), n)
        sign = 1 if total_i >= 0 else -1
        out: dict[str, float] = {}
        for i, day in enumerate(days):
            out[day.isoformat()] = float(sign * (base + (1 if i < rem else 0)))
        return out
    base = total / n
    return {day.isoformat(): float(base) for day in days}


def _add_distributed_qty(
    target: dict[str, float],
    total: float | None,
    days: list[date],
    *,
    only_if_present: bool,
) -> None:
    if total is None:
        return
    if only_if_present and not days:
        return
    # для факта: дни диапазона «есть в файле»; для плана одиночного дня — даже 0 пишем
    for day_key, qty in _distribute_qty_across_days(float(total), days).items():
        target[day_key] = target.get(day_key, 0.0) + qty


def _detailed_release_stage_priority(stage: str) -> int | None:
    """Приоритет строки выпуска: П/ф лучше, но при его отсутствии берём строку плана."""
    text = _normalize(stage)
    if text in _PF_STAGE_KEYS or text.startswith("п/ф"):
        return 0
    if any(token in text for token in _DETAILED_SKIP_STAGE_TOKENS):
        return None
    if any(token in text for token in _DETAILED_RELEASE_STAGE_TOKENS):
        return 1
    # В детальных планах без стадий колонка рядом с изделием часто пустая/числовая.
    return 2


def _choose_detailed_stage_series(
    series: list[_DetailedStageSeries],
) -> list[_DetailedStageSeries]:
    """Из всех строк изделия выбирает единую рабочую строку выпуска для расчёта."""
    by_product: dict[str, list[_DetailedStageSeries]] = {}
    for item in series:
        if item.has_nonzero or item.plan_cells:
            by_product.setdefault(_normalize(item.product), []).append(item)

    chosen: list[_DetailedStageSeries] = []
    for items in by_product.values():
        best = sorted(
            items,
            key=lambda item: (
                item.priority,
                -item.total_qty,
                -item.total_fact,
                item.row,
            ),
        )[0]
        if best.has_nonzero:
            chosen.append(best)
    return chosen


def _parse_pf_stage_report_sheet(
    sheet: Worksheet,
    year: int,
    month: int,
) -> tuple[list[DetailedScheduleProductPlan], list[DetailedPlanCellRef]]:
    """Отчёт «Модель/изделие»: собираем строки выпуска и выбираем плановую стадию.

    План: все дни из колонок/диапазонов (далее extract дозаполняет месяц нулями).
    Факт: только дни/периоды, для которых в файле есть значение факта.
    Диапазон «01.07–17.07» распределяется равномерно по дням диапазона ∩ месяц.
    """
    layout = _find_pf_report_layout(sheet, year, month)
    if layout is None:
        return [], []
    metric_row, name_col, stage_col, pairs = layout
    month_start = date(year, month, 1) if year > 0 and month > 0 else None
    month_end = (
        date(year, month, monthrange(year, month)[1]) if year > 0 and month > 0 else None
    )

    series: list[_DetailedStageSeries] = []
    current_product = ""
    for row_idx in range(metric_row + 1, sheet.max_row + 1):
        stage = _normalize(sheet.cell(row_idx, stage_col).value)
        name = _clean_text(sheet.cell(row_idx, name_col).value)
        if name and _is_schedule_product_name(name):
            current_product = name
        priority = _detailed_release_stage_priority(stage)
        if priority is None:
            continue
        if not current_product:
            continue

        daily_qty: dict[str, float] = {}
        daily_fact: dict[str, float] = {}
        row_plan_cells: list[DetailedPlanCellRef] = []
        has_any = False
        for pair in pairs:
            days = pair.days()
            if month_start and month_end:
                days = [d for d in days if month_start <= d <= month_end]
            if not days:
                continue
            day_keys = [d.isoformat() for d in days]
            plan_raw = sheet.cell(row_idx, pair.plan_col).value
            plan_qty = _to_float(plan_raw)
            if plan_qty is not None:
                has_any = has_any or plan_qty != 0
                _add_distributed_qty(daily_qty, plan_qty, days, only_if_present=False)
                row_plan_cells.append(
                    DetailedPlanCellRef(
                        product=current_product,
                        row=row_idx,
                        plan_col=pair.plan_col,
                        day_keys=day_keys,
                        plan_qty=float(plan_qty),
                    )
                )
            elif pair.day is not None:
                # явная дневная колонка без числа → план 0 на этот день
                daily_qty[pair.day.isoformat()] = daily_qty.get(pair.day.isoformat(), 0.0)
                row_plan_cells.append(
                    DetailedPlanCellRef(
                        product=current_product,
                        row=row_idx,
                        plan_col=pair.plan_col,
                        day_keys=[pair.day.isoformat()],
                        plan_qty=0.0,
                    )
                )

            if pair.fact_col is None:
                continue
            fact_raw = sheet.cell(row_idx, pair.fact_col).value
            fact_qty = _to_float(fact_raw)
            if fact_qty is None:
                continue
            has_any = has_any or fact_qty != 0
            _add_distributed_qty(daily_fact, fact_qty, days, only_if_present=True)

        if not daily_qty and not daily_fact:
            continue
        # пустые слоты (номер есть, изделия нет) — без чисел выпуска
        if not has_any:
            continue
        series.append(
            _DetailedStageSeries(
                product=current_product,
                stage=stage,
                row=row_idx,
                priority=priority,
                daily_qty=daily_qty,
                daily_fact=daily_fact,
                plan_cells=row_plan_cells,
            )
        )

    chosen = _choose_detailed_stage_series(series)
    plans = [
        DetailedScheduleProductPlan(
            product=item.product,
            daily_qty=item.daily_qty,
            daily_fact=item.daily_fact,
            year=year,
            month=month,
        )
        for item in chosen
    ]
    plan_cells = [cell for item in chosen for cell in item.plan_cells]
    return plans, plan_cells


def _find_schedule_table(
    sheet: Worksheet,
) -> tuple[int, int, list[int], list[tuple[int, str]]] | None:
    for header_idx in range(1, min(sheet.max_row, 40) + 1):
        headers = {
            col_idx: _normalize(sheet.cell(header_idx, col_idx).value)
            for col_idx in range(1, sheet.max_column + 1)
        }
        name_col = _pick_column(
            headers,
            ["наименования изделий", "наименование изделий", "наименование", "изделие"],
        )
        month_cols: list[tuple[int, str]] = []
        for col_idx, header in headers.items():
            month_num = _month_number_from_header(header)
            if month_num is None:
                continue
            month_cols.append((col_idx, _MONTH_NOMINATIVE[month_num - 1]))
        data_cols = [
            col_idx
            for col_idx, header in headers.items()
            if col_idx != name_col
            and (_month_number_from_header(header) or "кол" in header or "итог" in header)
        ]
        if name_col and len(month_cols) >= 2:
            if not data_cols:
                data_cols = [col_idx for col_idx, _ in month_cols]
            return header_idx, name_col, data_cols, month_cols
    return None


def _cell_has_number(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, int | float):
        return True
    text = _clean_text(value).replace(" ", "").replace(",", ".")
    try:
        float(text)
        return True
    except ValueError:
        return False


def _month_number_from_header(header: str) -> int | None:
    normalized = _normalize(header)
    if not normalized:
        return None
    for idx, name in enumerate(_MONTH_NOMINATIVE):
        month_key = _normalize(name)
        if normalized == month_key or normalized.startswith(month_key):
            return idx + 1
    for idx, name in enumerate(_MONTH_LOWER):
        if normalized == name or normalized.startswith(name):
            return idx + 1
    # «График выпуска … Август 2026», «"___" августа 2026 г.»
    tokens = re.findall(r"[a-zа-яё]+", normalized)
    for token in tokens:
        for idx, name in enumerate(_MONTH_LOWER):
            if token == name or token.startswith(name):
                return idx + 1
    for idx, name in enumerate(_MONTH_LOWER):
        if len(name) >= 4 and name in normalized:
            return idx + 1
    return None


def _pick_column(headers: dict[int, str], candidates: list[str]) -> int | None:
    normalized_candidates = [_normalize(candidate) for candidate in candidates]
    for candidate in normalized_candidates:
        for col_idx, header in headers.items():
            if header == candidate:
                return col_idx
    for candidate in normalized_candidates:
        for col_idx, header in headers.items():
            if candidate and candidate in header:
                return col_idx
    return None


def _is_schedule_product_name(value: str) -> bool:
    if not value:
        return False
    lowered = value.lower().strip()
    if lowered.startswith("итого"):
        return False
    if lowered in {"наименование", "наименования изделий", "изделие"}:
        return False
    if re.search(r"производство\s*№", lowered):
        return False
    if any(token in lowered for token in ("руководитель", "начальник", "подпись", "утверждаю")):
        return False
    return len(value) > 1


async def _resolve_schedule_products_to_specs(
    products: list[str],
    *,
    db_spec_catalog: list | None = None,
    product_spec_hints: dict[str, tuple[str, str]] | None = None,
) -> list[ProductSpecLink]:
    from app.agents.document_analysis_agent.onec_db_sources import (
        DbSpecCatalogEntry,
        lookup_product_spec_hint,
        match_product_to_db_spec,
        _valid_spec_ref_key,
    )

    mapping_rows = await asyncio.to_thread(_load_nomenclature_mapping)
    use_db_specs = db_spec_catalog is not None
    catalog: list[DbSpecCatalogEntry] = list(db_spec_catalog or [])

    if not products:
        return []
    if not mapping_rows and not use_db_specs:
        logger.warning("document_analysis_agent.spec_resolve_skipped", mapping_exists=_MAPPING_FILE.exists())
        return [
            ProductSpecLink(
                schedule_product=product,
                status="unmatched",
                reason="Нет сопоставления номенклатур и спецификаций в БД",
            )
            for product in products
        ]
    if use_db_specs and not catalog:
        return [
            ProductSpecLink(
                schedule_product=product,
                status="unmatched",
                reason="Спецификации в БД отсутствуют — выполните синхронизацию из 1С",
            )
            for product in products
        ]

    unique_nomenclatures = list(dict.fromkeys(row.nomenclature for row in mapping_rows)) if mapping_rows else []
    local_nomenclature_map = (
        _match_schedule_to_nomenclatures_locally(products, mapping_rows) if mapping_rows else {}
    )
    ambiguous = [
        product
        for product in products
        if local_nomenclature_map.get(product) is None
        or local_nomenclature_map[product][1] < 0.62
    ]
    lm_nomenclature_map = (
        await _match_schedule_to_nomenclatures_with_lm(ambiguous, mapping_rows) if mapping_rows else {}
    )

    legacy_sheet_names = (
        await asyncio.to_thread(_load_spec_sheet_names) if not use_db_specs else []
    )

    links: list[ProductSpecLink] = []
    db_spec_pending: list[tuple[str, str, str, str, str]] = []
    legacy_spec_pending: list[tuple[str, str, str]] = []
    catalog_by_ref: dict[str, DbSpecCatalogEntry] = {}
    if use_db_specs:
        for entry in catalog:
            catalog_by_ref[entry.ref_key] = entry
            lowered = entry.ref_key.strip().lower()
            if lowered:
                catalog_by_ref[lowered] = entry

    def _catalog_entry_for_ref(ref_key: str) -> DbSpecCatalogEntry | None:
        cleaned = (ref_key or "").strip()
        if not cleaned:
            return None
        return catalog_by_ref.get(cleaned) or catalog_by_ref.get(cleaned.lower())

    for product in products:
        spec_hint, spec_ref_key = lookup_product_spec_hint(product, product_spec_hints)
        spec_hint = (spec_hint or "").strip()
        spec_ref_key = _valid_spec_ref_key(spec_ref_key)

        if use_db_specs and spec_ref_key:
            entry = _catalog_entry_for_ref(spec_ref_key)
            if entry is not None:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=entry.main_product_name or spec_hint or product,
                        spec_sheet=entry.description or entry.label,
                        spec_ref_key=entry.ref_key,
                        status="matched",
                        reason="спецификация из плана производства (Ref_Key)",
                    )
                )
                continue
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    nomenclature=spec_hint or product,
                    spec_sheet=spec_hint or product,
                    spec_ref_key=spec_ref_key,
                    status="unmatched",
                    reason="спецификация из плана производства (Ref_Key) отсутствует в каталоге 1С",
                )
            )
            continue

        if use_db_specs and spec_hint:
            entry, direct_reason = match_product_to_db_spec(
                product,
                spec_hint,
                catalog,
                spec_hint=spec_hint,
            )
            if entry is not None:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=entry.main_product_name or spec_hint,
                        spec_sheet=entry.description or entry.label,
                        spec_ref_key=entry.ref_key,
                        status="matched",
                        reason=f"спецификация из плана производства; {direct_reason}",
                    )
                )
                continue

        nomenclature: str | None = None
        reason = ""
        local = local_nomenclature_map.get(product)
        if local and local[1] >= 0.62:
            nomenclature, score = local[0], local[1]
            reason = f"локальный матч номенклатуры ({score:.2f})"
        elif product in lm_nomenclature_map and lm_nomenclature_map[product]:
            nomenclature = lm_nomenclature_map[product]
            reason = "LM Studio: номенклатура"
        elif local and local[0]:
            nomenclature, score = local[0], local[1]
            reason = f"слабый локальный матч номенклатуры ({score:.2f})"

        if not nomenclature and use_db_specs:
            entry, direct_reason = match_product_to_db_spec(
                product,
                product,
                catalog,
                spec_hint=spec_hint,
            )
            if entry is not None:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=entry.main_product_name or entry.description or product,
                        spec_sheet=entry.description or entry.label,
                        spec_ref_key=entry.ref_key,
                        status="matched",
                        reason=f"прямой матч к БД; {direct_reason}",
                    )
                )
                continue

        if not nomenclature:
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    status="no_mapping",
                    reason="Не найдена номенклатура в таблице сопоставления",
                )
            )
            continue

        if nomenclature not in unique_nomenclatures and unique_nomenclatures:
            best_name, best_score = _best_text_match(nomenclature, unique_nomenclatures)
            if best_name and best_score >= 0.55:
                nomenclature = best_name
                reason = f"{reason}; нормализовано к mapping"

        if use_db_specs:
            entry, db_reason = match_product_to_db_spec(
                product,
                nomenclature,
                catalog,
                spec_hint=spec_hint,
            )
            if entry is not None:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=nomenclature,
                        spec_sheet=entry.description or entry.label,
                        spec_ref_key=entry.ref_key,
                        status="matched",
                        reason=f"{reason}; {db_reason}".strip("; "),
                    )
                )
                continue
            db_spec_pending.append((product, nomenclature, reason, spec_hint, spec_ref_key))
            continue

        if not legacy_sheet_names:
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    nomenclature=nomenclature,
                    status="no_sheet",
                    reason=f"{reason}; нет спецификаций",
                )
            )
            continue

        sheet, sheet_reason = _match_nomenclature_to_sheet(product, nomenclature, legacy_sheet_names)
        if sheet:
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    nomenclature=nomenclature,
                    spec_sheet=sheet,
                    status="matched",
                    reason=f"{reason}; {sheet_reason}".strip("; "),
                )
            )
            continue
        legacy_spec_pending.append((product, nomenclature, reason))

    if use_db_specs and db_spec_pending:
        catalog_labels = [entry.label for entry in catalog]
        lm_inputs = [
            spec_hint or nomenclature
            for _, nomenclature, _, spec_hint, _ in db_spec_pending
        ]
        lm_sheets = await _match_nomenclatures_to_sheets_with_lm(lm_inputs, catalog_labels)
        for product, nomenclature, reason, spec_hint, spec_ref_key in db_spec_pending:
            lm_key = spec_hint or nomenclature
            sheet_name = lm_sheets.get(lm_key)
            entry = None
            db_reason = "спецификация не найдена в БД"
            if spec_ref_key:
                entry = _catalog_entry_for_ref(spec_ref_key)
                if entry is not None:
                    db_reason = "спецификация из плана производства (Ref_Key)"
                else:
                    links.append(
                        ProductSpecLink(
                            schedule_product=product,
                            nomenclature=spec_hint or nomenclature,
                            spec_sheet=spec_hint or nomenclature,
                            spec_ref_key=spec_ref_key,
                            status="unmatched",
                            reason=f"{reason}; спецификация из плана (Ref_Key) отсутствует в каталоге 1С",
                        )
                    )
                    continue
            elif spec_hint:
                entry, hint_reason = match_product_to_db_spec(
                    product,
                    spec_hint,
                    catalog,
                    spec_hint=spec_hint,
                )
                if entry is not None:
                    db_reason = hint_reason
            if entry is None and sheet_name:
                for item in catalog:
                    if item.label == sheet_name or _normalize(item.label) == _normalize(sheet_name):
                        entry = item
                        db_reason = "LM Studio: спецификация БД"
                        break
            if entry is None:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=nomenclature,
                        status="no_sheet",
                        reason=f"{reason}; {db_reason}",
                    )
                )
                continue
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    nomenclature=nomenclature,
                    spec_sheet=entry.description or entry.label,
                    spec_ref_key=entry.ref_key,
                    status="matched",
                    reason=f"{reason}; {db_reason}".strip("; "),
                )
            )

    if legacy_spec_pending and legacy_sheet_names:
        lm_sheets = await _match_nomenclatures_to_sheets_with_lm(
            [nomenclature for _, nomenclature, _ in legacy_spec_pending],
            legacy_sheet_names,
        )
        for product, nomenclature, reason in legacy_spec_pending:
            sheet = lm_sheets.get(nomenclature)
            if not sheet:
                links.append(
                    ProductSpecLink(
                        schedule_product=product,
                        nomenclature=nomenclature,
                        status="no_sheet",
                        reason=f"{reason}; лист не найден",
                    )
                )
                continue
            links.append(
                ProductSpecLink(
                    schedule_product=product,
                    nomenclature=nomenclature,
                    spec_sheet=sheet,
                    status="matched",
                    reason=f"{reason}; LM Studio: лист спецификации",
                )
            )

    return links


def _restrict_merged_rows_to_spec_products(
    rows: list[MergedNomenclatureRow],
    spec_eligible_products: frozenset[str] | None,
) -> None:
    """Убирает из by_product изделия без загруженной ресурсной спецификации 1С."""
    if not spec_eligible_products:
        return

    allowed_keys = {_normalize(product) for product in spec_eligible_products}
    canonical_name = {_normalize(product): product for product in spec_eligible_products}
    trimmed_rows = 0

    for row in rows:
        kept_by_product: dict[str, float | None] = {}
        kept_products: list[str] = []
        for product, qty in row.by_product.items():
            key = _normalize(product)
            if key not in allowed_keys:
                continue
            canonical = canonical_name[key]
            kept_by_product[canonical] = qty
            kept_products.append(canonical)

        if len(kept_by_product) != len(row.by_product):
            trimmed_rows += 1

        row.by_product = kept_by_product
        row.products = sorted(kept_products, key=_normalize)
        known = [value for value in kept_by_product.values() if value is not None]
        if not known:
            row.quantity = None
        elif all(abs(value - known[0]) < 1e-9 for value in known):
            row.quantity = known[0]
        else:
            row.quantity = None

    logger.info(
        "document_analysis_agent.merged_rows_spec_filtered",
        rows=len(rows),
        trimmed_rows=trimmed_rows,
        spec_products=len(spec_eligible_products),
    )


def _material_kind_priority(kind: str) -> int:
    if kind == MATERIAL_KIND_CONSUMABLE:
        return 2
    if kind == MATERIAL_KIND_WORKSHOP:
        return 1
    return 0


def _enrich_merged_with_coverage_material_classification(
    rows: list[MergedNomenclatureRow],
) -> None:
    """Помечает строки спеки расходниками/возможными цеховыми остатками для условного расчёта."""
    if not rows:
        return
    index = load_material_classification_index(str(_MATERIAL_CLASSIFICATION_FILE))
    if not index.by_pair and not index.by_material:
        logger.warning(
            "document_analysis_agent.coverage_material_classification_empty",
            path=str(_MATERIAL_CLASSIFICATION_FILE),
        )
        return

    classified = 0
    optional = 0
    for row in rows:
        row.coverage_material_kinds_by_product = {}
        row.coverage_material_labels_by_product = {}
        row.coverage_material_confidences_by_product = {}
        row.coverage_material_reasons_by_product = {}

        best_kind = MATERIAL_KIND_REQUIRED
        best_label = MATERIAL_KIND_LABELS[MATERIAL_KIND_REQUIRED]
        best_confidence = ""
        best_reason = ""

        products = list(row.by_product) or list(row.products)
        for product in products:
            item = material_classification_for(
                index,
                product=product,
                material=row.nomenclature,
            )
            row.coverage_material_kinds_by_product[product] = item.kind
            row.coverage_material_labels_by_product[product] = item.label
            row.coverage_material_confidences_by_product[product] = item.confidence
            row.coverage_material_reasons_by_product[product] = item.reason
            if _material_kind_priority(item.kind) > _material_kind_priority(best_kind):
                best_kind = item.kind
                best_label = item.label
                best_confidence = item.confidence
                best_reason = item.reason

        row.coverage_material_kind = best_kind
        row.coverage_material_label = best_label
        row.coverage_material_confidence = best_confidence
        row.coverage_material_reason = best_reason
        classified += 1
        if is_optional_material_kind(best_kind):
            optional += 1

    logger.info(
        "document_analysis_agent.coverage_material_classification_loaded",
        rows=classified,
        optional=optional,
        path=str(_MATERIAL_CLASSIFICATION_FILE),
    )


def _collect_and_merge_spec_materials(
    links: list[ProductSpecLink],
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    schedule_plans: list[ScheduleProductPlan],
    detailed_extract: DetailedScheduleExtract | None = None,
    db_stock_index: dict[str, StockEntry] | None = None,
    db_materials_by_ref: dict[str, list[SpecMaterialItem]] | None = None,
    db_country_index: dict[str, Any] | None = None,
    db_unit_index: dict[str, Any] | None = None,
    user_id: Any | None = None,
    shipment_bundle: ShipmentScheduleBundle | None = None,
    spec_eligible_products: frozenset[str] | None = None,
) -> tuple[
    list[SpecMaterialItem],
    list[MergedNomenclatureRow],
    bytes | None,
    list[str],
    list[str],
    LogisticsRiskBoard,
    dict[str, Any] | None,
    dict[str, Any] | None,
]:
    """Разбор листов → merge → цены → остатки → потребность → поступления → риски → result.xlsx."""
    from app.agents.document_analysis_agent.coverage_dashboard import build_coverage_dashboard
    from app.agents.document_analysis_agent.order_plan import compute_order_plan
    from app.agents.document_analysis_agent.product_coverage import (
        compute_daily_plan_coverage,
        compute_product_coverage,
    )

    if detailed_extract is None:
        detailed_extract = DetailedScheduleExtract(files=[], plans=[], year=0, month=0)
    if shipment_bundle is None:
        shipment_bundle = _load_shipment_schedule_bundle(workbooks, role_map)
    usages = _extract_materials_from_matched_specs(links, db_materials_by_ref=db_materials_by_ref)
    merged = _merge_material_usages(usages)
    _restrict_merged_rows_to_schedule_products(
        merged,
        list(schedule_plans) + list(detailed_extract.plans),
    )
    _restrict_merged_rows_to_spec_products(merged, spec_eligible_products)
    _enrich_merged_with_coverage_material_classification(merged)
    _enrich_merged_with_purchase_prices(merged)
    _enrich_merged_with_country_of_origin(merged, db_country_index)
    _enrich_merged_with_units(merged, db_unit_index)
    stock_files = _enrich_merged_with_stock(
        merged,
        workbooks,
        role_map,
        db_stock_index=db_stock_index,
    )
    _enrich_merged_with_monthly_demand(merged, schedule_plans)
    shipment_files = _enrich_merged_with_monthly_receipts(
        merged, workbooks, role_map, shipment_bundle=shipment_bundle
    )
    _enrich_merged_with_shipment_country(merged, shipment_bundle=shipment_bundle)
    _enrich_merged_with_monthly_forecast(merged)
    _enrich_merged_with_daily_demand(merged, detailed_extract)
    _enrich_merged_with_daily_receipts(
        merged, workbooks, role_map, detailed_extract, shipment_bundle=shipment_bundle
    )
    _enrich_merged_with_daily_forecast(merged, detailed_extract)
    logistics_risks = _build_logistics_risk_board(
        merged, workbooks, role_map, shipment_bundle=shipment_bundle
    )
    _strip_excluded_suppliers_from_rows(merged)
    months = _months_for_coverage_sheet(merged, schedule_plans)
    as_of_day = date.today()
    if logistics_risks and logistics_risks.as_of:
        try:
            as_of_day = date.fromisoformat(str(logistics_risks.as_of))
        except ValueError:
            pass
    schedule_month = (
        f"{detailed_extract.year:04d}-{detailed_extract.month:02d}"
        if detailed_extract.year and detailed_extract.month
        else f"{as_of_day.year:04d}-{as_of_day.month:02d}"
    )
    from app.agents.document_analysis_agent.coverage_dashboard import (
        resolve_coverage_target_month,
        resolve_plan_month_keys,
        _filter_day_keys_to_schedule_month,
    )

    _, coverage_month_label = resolve_coverage_target_month(
        schedule_month=schedule_month,
        as_of=as_of_day,
        merged=merged,
    )
    coverage_months = resolve_plan_month_keys(
        schedule_plans,
        schedule_month=schedule_month,
        month_label=coverage_month_label,
    )
    product_coverage_full = compute_product_coverage(schedule_plans, merged, months)
    product_coverage = (
        product_coverage_full
        if list(coverage_months) == list(months)
        else compute_product_coverage(schedule_plans, merged, coverage_months)
    )
    coverage_day_keys = _filter_day_keys_to_schedule_month(
        list(detailed_extract.day_keys),
        schedule_month,
    )
    daily_plan_coverage = compute_daily_plan_coverage(
        detailed_extract.plans,
        merged,
        coverage_day_keys,
    )
    logistics_leads = _load_shipment_logistics_leads(
        workbooks, role_map, shipment_bundle=shipment_bundle
    )
    order_year = detailed_extract.year if detailed_extract.year > 0 else date.today().year
    order_plan = compute_order_plan(merged, months, order_year, logistics_leads)
    schedule_month = (
        f"{detailed_extract.year:04d}-{detailed_extract.month:02d}"
        if detailed_extract.year and detailed_extract.month
        else f"{as_of_day.year:04d}-{as_of_day.month:02d}"
    )
    from app.agents.document_analysis_agent.coverage_dashboard import dump_coverage_rebuild

    coverage_dashboard = build_coverage_dashboard(
        daily_plan_coverage=daily_plan_coverage,
        product_coverage=product_coverage,
        merged=merged,
        day_keys=coverage_day_keys,
        detailed_plans=detailed_extract.plans,
        as_of=as_of_day,
        schedule_month=schedule_month,
        spec_eligible_products=spec_eligible_products,
    )
    coverage_rebuild = dump_coverage_rebuild(
        merged=merged,
        detailed_plans=detailed_extract.plans,
        day_keys=coverage_day_keys,
        as_of=as_of_day,
        schedule_month=schedule_month,
        spec_eligible_products=spec_eligible_products,
    )
    result_bytes = _build_result_xlsx(
        merged,
        detailed_extract,
        product_coverage_full,
        order_plan,
        daily_plan_coverage,
        user_id=user_id,
        spec_eligible_products=spec_eligible_products,
    )
    return (
        usages,
        merged,
        result_bytes,
        stock_files,
        shipment_files,
        logistics_risks,
        coverage_dashboard,
        coverage_rebuild,
    )


def _months_for_coverage_sheet(
    rows: list[MergedNomenclatureRow],
    schedule_plans: list[ScheduleProductPlan],
) -> list[str]:
    """Месяцы для листа обеспеченности изделий — как у помесячного + месяцы из графика."""
    months = _months_for_monthly_sheet(rows)
    for plan in schedule_plans:
        for month in plan.monthly_qty:
            if month not in months:
                months.append(month)
    return sorted(
        months,
        key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
    )


def _extract_materials_from_matched_specs(
    links: list[ProductSpecLink],
    *,
    db_materials_by_ref: dict[str, list[SpecMaterialItem]] | None = None,
) -> list[SpecMaterialItem]:
    if db_materials_by_ref is not None:
        usages: list[SpecMaterialItem] = []
        for link in links:
            if link.status != "matched":
                continue
            ref_key = (link.spec_ref_key or "").strip()
            items = db_materials_by_ref.get(ref_key, [])
            if not items and ref_key:
                items = db_materials_by_ref.get(ref_key.lower(), [])
            product_items = [
                SpecMaterialItem(
                    nomenclature=item.nomenclature,
                    quantity=item.quantity,
                    product=link.schedule_product,
                    unit=item.unit,
                    spec_sheet=link.spec_sheet or item.spec_sheet,
                )
                for item in items
            ]
            usages.extend(product_items)
            logger.info(
                "document_analysis_agent.db_spec_materials_used",
                product=link.schedule_product,
                ref_key=link.spec_ref_key,
                materials=len(product_items),
            )
        return usages

    matched = [link for link in links if link.status == "matched" and link.spec_sheet]
    if not matched:
        return []
    if not _SPECS_FILE.exists():
        logger.warning("document_analysis_agent.specs_file_missing", path=str(_SPECS_FILE))
        return []

    workbook = load_workbook(_SPECS_FILE, data_only=True, read_only=True)
    try:
        sheet_lookup = {name: name for name in workbook.sheetnames}
        sheet_lookup_norm = {_normalize(name): name for name in workbook.sheetnames}

        usages: list[SpecMaterialItem] = []
        for link in matched:
            sheet_name = link.spec_sheet or ""
            resolved = sheet_lookup.get(sheet_name) or sheet_lookup_norm.get(
                _normalize(sheet_name)
            )
            if not resolved:
                logger.warning(
                    "document_analysis_agent.spec_sheet_missing_in_workbook",
                    sheet=sheet_name,
                    product=link.schedule_product,
                )
                continue
            worksheet = workbook[resolved]
            sheet_items = _parse_spec_sheet_materials(
                worksheet,
                product=link.schedule_product,
                sheet_name=resolved,
            )
            usages.extend(sheet_items)
            logger.info(
                "document_analysis_agent.spec_sheet_parsed",
                product=link.schedule_product,
                sheet=resolved,
                materials=len(sheet_items),
            )
        return usages
    finally:
        workbook.close()


def _parse_spec_sheet_materials(
    worksheet: Worksheet,
    *,
    product: str,
    sheet_name: str,
) -> list[SpecMaterialItem]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_idx, name_col, qty_col, unit_col = _detect_spec_header(rows)
    if header_row_idx is None or name_col is None:
        return []

    items: list[SpecMaterialItem] = []
    for row in rows[header_row_idx + 1 :]:
        if name_col >= len(row):
            continue
        name = _clean_text(row[name_col])
        if not name or _normalize(name) in {"наименование", "номенклатура"}:
            continue
        if _SECTION_NAME_RE.match(name):
            continue
        if _is_spec_root_product_row(name, product, sheet_name, row, qty_col):
            continue

        quantity = _to_float(row[qty_col]) if qty_col is not None and qty_col < len(row) else None
        unit = None
        if unit_col is not None and unit_col < len(row):
            unit = _clean_text(row[unit_col]) or None

        items.append(
            SpecMaterialItem(
                nomenclature=name,
                quantity=quantity,
                product=product,
                unit=unit,
                spec_sheet=sheet_name,
            )
        )
    return items


def _detect_spec_header(
    rows: list[tuple[Any, ...]],
) -> tuple[int | None, int | None, int | None, int | None]:
    """Ищет строку заголовка: Наименование + Количество (+ ед. изм.)."""
    for idx, row in enumerate(rows[:20]):
        name_col: int | None = None
        qty_col: int | None = None
        unit_col: int | None = None
        for col_idx, value in enumerate(row):
            text = _normalize(value)
            if not text:
                continue
            if name_col is None and ("наименование" in text or text == "номенклатура"):
                name_col = col_idx
            elif qty_col is None and "количество" in text:
                qty_col = col_idx
            elif unit_col is None and ("ед" in text and "изм" in text):
                unit_col = col_idx
        if name_col is not None and qty_col is not None:
            return idx, name_col, qty_col, unit_col
        if name_col is not None and qty_col is None:
            # редкий случай: количество без явного заголовка рядом — угадываем позже
            inferred_qty = _infer_qty_column(rows, idx, name_col)
            if inferred_qty is not None:
                return idx, name_col, inferred_qty, unit_col
    return None, None, None, None


def _infer_qty_column(
    rows: list[tuple[Any, ...]],
    header_idx: int,
    name_col: int,
) -> int | None:
    scores: dict[int, int] = {}
    for row in rows[header_idx + 1 : header_idx + 40]:
        if name_col >= len(row) or not _clean_text(row[name_col]):
            continue
        for col_idx, value in enumerate(row):
            if col_idx == name_col:
                continue
            if _to_float(value) is not None:
                scores[col_idx] = scores.get(col_idx, 0) + 1
    if not scores:
        return None
    return max(scores.items(), key=lambda item: item[1])[0]


def _is_spec_root_product_row(
    name: str,
    product: str,
    sheet_name: str,
    row: tuple[Any, ...],
    qty_col: int | None,
) -> bool:
    """Корневая строка изделия в выгрузке 1С (НСУ и т.п.), не материал."""
    quantity = _to_float(row[qty_col]) if qty_col is not None and qty_col < len(row) else None
    if quantity not in (1, 1.0):
        return False
    if len(name) > 80:
        return False
    sheet_score = _product_match_score(name, sheet_name)
    product_score = _product_match_score(name, product)
    if max(sheet_score, product_score) >= 0.72:
        return True
    # короткие корни вроде «НСУ», «НСУ 2.0»
    compact = _match_key(name)
    if compact in {_match_key(sheet_name), _match_key(product)}:
        return True
    if len(compact) <= 12 and (
        compact in _match_key(sheet_name) or compact in _match_key(product)
    ):
        return True
    return False


def _merge_material_usages(usages: list[SpecMaterialItem]) -> list[MergedNomenclatureRow]:
    """Сводит поиздельные позиции в уникальные номенклатуры.

    - ключ уникальности: нормализованное наименование;
    - изделия перечисляются без повторов;
    - количество: сумма внутри одного изделия, при совпадении между изделиями —
      общее значение, иначе None (разбивка остаётся в by_product).
    """
    # product -> nomenclature_key -> aggregate
    per_product: dict[str, dict[str, dict[str, Any]]] = {}
    display_names: dict[str, str] = {}
    units: dict[str, str | None] = {}

    for item in usages:
        key = _normalize(item.nomenclature)
        if not key:
            continue
        display_names.setdefault(key, item.nomenclature)
        if key not in units and item.unit:
            units[key] = item.unit
        product_bucket = per_product.setdefault(item.product, {})
        bucket = product_bucket.setdefault(key, {"quantity": None})
        if item.quantity is not None:
            current = bucket["quantity"]
            bucket["quantity"] = item.quantity if current is None else float(current) + float(item.quantity)

    merged_map: dict[str, MergedNomenclatureRow] = {}
    for product, materials in per_product.items():
        for key, payload in materials.items():
            qty = payload["quantity"]
            if key not in merged_map:
                merged_map[key] = MergedNomenclatureRow(
                    nomenclature=display_names[key],
                    products=[product],
                    quantity=qty if isinstance(qty, (int, float)) else None,
                    unit=units.get(key),
                    by_product={product: qty if isinstance(qty, (int, float)) else None},
                )
                continue
            row = merged_map[key]
            if product not in row.products:
                row.products.append(product)
            row.by_product[product] = qty if isinstance(qty, (int, float)) else None
            known = [value for value in row.by_product.values() if value is not None]
            if not known:
                row.quantity = None
            elif all(abs(value - known[0]) < 1e-9 for value in known):
                row.quantity = known[0]
            else:
                row.quantity = None

    rows = list(merged_map.values())
    rows.sort(key=lambda item: _normalize(item.nomenclature))
    for row in rows:
        row.products = sorted(row.products, key=_normalize)
    return rows


def _restrict_merged_rows_to_schedule_products(
    rows: list[MergedNomenclatureRow],
    schedule_plans: list[Any],
) -> None:
    """Оставляет в строке только изделия из месячного или детального графика производства."""
    if not schedule_plans:
        return

    allowed_keys = {_normalize(plan.product) for plan in schedule_plans}
    canonical_name = {_normalize(plan.product): plan.product for plan in schedule_plans}
    trimmed_rows = 0

    for row in rows:
        kept_by_product: dict[str, float | None] = {}
        kept_products: list[str] = []
        for product, qty in row.by_product.items():
            key = _normalize(product)
            if key not in allowed_keys:
                continue
            canonical = canonical_name[key]
            kept_by_product[canonical] = qty
            kept_products.append(canonical)

        if len(kept_by_product) != len(row.by_product):
            trimmed_rows += 1

        row.by_product = kept_by_product
        row.products = sorted(kept_products, key=_normalize)
        known = [value for value in kept_by_product.values() if value is not None]
        if not known:
            row.quantity = None
        elif all(abs(value - known[0]) < 1e-9 for value in known):
            row.quantity = known[0]
        else:
            row.quantity = None

    logger.info(
        "document_analysis_agent.merged_rows_schedule_filtered",
        rows=len(rows),
        trimmed_rows=trimmed_rows,
        schedule_products=len(schedule_plans),
    )


# Поставщики, которых нельзя показывать в итоговой таблице (колонка «Поставщик»).
# Цена из их строк прайса может остаться; в колонку C пишем пусто.
_EXCLUDED_SUPPLIER_EXACT = frozenset(
    {
        "итц ооо",
        "ооо итц",
        "магакян екатерина ивановна ип",
        "ип магакян екатерина ивановна",
    }
)
_LEGAL_FORM_TOKENS_RE = re.compile(
    r"\b(ооо|ип|ао|зао|пао|оао|общество|с|ограниченной|ответственностью)\b",
    re.IGNORECASE,
)


def _supplier_match_key(supplier: str | None) -> str:
    text = _normalize(supplier)
    if not text:
        return ""
    text = text.replace('"', "").replace("«", "").replace("»", "").replace("'", "")
    return re.sub(r"\s+", " ", text).strip()


def _is_excluded_supplier(supplier: str | None) -> bool:
    """True для ИТЦ ООО и ИП Магакян Е.И. — в итоговой таблице поставщик пустой."""
    key = _supplier_match_key(supplier)
    if not key:
        return False
    if key in _EXCLUDED_SUPPLIER_EXACT:
        return True
    if "магакян" in key and "екатерина" in key:
        return True
    core = _LEGAL_FORM_TOKENS_RE.sub(" ", key)
    core = re.sub(r"\s+", " ", core).strip()
    return core == "итц"


def _sanitize_result_supplier(supplier: str | None) -> str | None:
    """Возвращает поставщика для итоговой таблицы или None, если он в стоп-листе."""
    cleaned = _clean_text(supplier) if supplier else ""
    if not cleaned or _is_excluded_supplier(cleaned):
        return None
    return cleaned


def _strip_excluded_suppliers_from_rows(rows: list[MergedNomenclatureRow]) -> int:
    """Финальная зачистка стоп-листа перед записью result.xlsx / API."""
    cleared = 0
    for row in rows:
        if row.supplier and _is_excluded_supplier(row.supplier):
            row.supplier = None
            cleared += 1
    if cleared:
        logger.info(
            "document_analysis_agent.excluded_suppliers_stripped",
            cleared=cleared,
            total=len(rows),
        )
    return cleared


def _load_purchase_price_index() -> dict[str, PurchasePriceEntry]:
    """Индекс закупочных цен: ключ = нормализованная номенклатура.

    При нескольких закупках одной позиции берём строку с максимальным оборотом
    (КоличествоОборот) — наиболее представительный контрагент и цена.
    """
    if not _PRICES_FILE.exists():
        logger.warning("document_analysis_agent.prices_file_missing", path=str(_PRICES_FILE))
        return {}

    workbook = load_workbook(_PRICES_FILE, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_idx = None
    name_col = supplier_col = qty_col = price_col = None
    for idx, row in enumerate(rows[:15]):
        normalized = [_normalize(value) for value in row]
        for col_idx, text in enumerate(normalized):
            if not text:
                continue
            if name_col is None and ("номенклатура" in text):
                name_col = col_idx
            elif supplier_col is None and ("контрагент" in text or "поставщик" in text):
                supplier_col = col_idx
            elif qty_col is None and "количество" in text:
                qty_col = col_idx
            elif price_col is None and text.startswith("цена"):
                price_col = col_idx
        if name_col is not None and supplier_col is not None and price_col is not None:
            header_idx = idx
            break

    if header_idx is None or name_col is None or price_col is None:
        logger.warning("document_analysis_agent.prices_header_not_found")
        return {}

    index: dict[str, PurchasePriceEntry] = {}
    for row in rows[header_idx + 1 :]:
        if name_col >= len(row):
            continue
        name = _clean_text(row[name_col]).rstrip("*").strip()
        if not name:
            continue
        supplier = ""
        if supplier_col is not None and supplier_col < len(row):
            supplier = _clean_text(row[supplier_col])
        qty = 0.0
        if qty_col is not None and qty_col < len(row):
            qty = _to_float(row[qty_col]) or 0.0
        price = _to_float(row[price_col]) if price_col < len(row) else None
        key = _normalize(name)
        candidate = PurchasePriceEntry(
            nomenclature=name,
            supplier=supplier,
            price=price,
            turnover_qty=qty,
        )
        previous = index.get(key)
        if previous is None:
            index[key] = candidate
            continue
        prev_excluded = _is_excluded_supplier(previous.supplier)
        cand_excluded = _is_excluded_supplier(candidate.supplier)
        # При дублях предпочитаем строку с «разрешённым» поставщиком.
        if prev_excluded and not cand_excluded:
            index[key] = candidate
        elif not prev_excluded and cand_excluded:
            continue
        elif qty >= previous.turnover_qty:
            index[key] = candidate
    logger.info("document_analysis_agent.prices_loaded", unique=len(index))
    return index


def _enrich_merged_with_purchase_prices(rows: list[MergedNomenclatureRow]) -> None:
    """Дополняет итоговую структуру поставщиком и ценой без НДС за единицу."""
    index = _load_purchase_price_index()
    if not index:
        for row in rows:
            row.price_match = "unmatched"
        return

    candidates = [entry.nomenclature for entry in index.values()]
    matched = 0
    excluded_suppliers = 0
    for row in rows:
        entry, method = _match_catalog_entry(row.nomenclature, index, candidates)
        if entry is None:
            row.supplier = None
            row.price = None
            row.price_match = "unmatched"
            continue
        sanitized = _sanitize_result_supplier(entry.supplier)
        if entry.supplier and sanitized is None:
            excluded_suppliers += 1
        row.supplier = sanitized
        row.price = round(entry.price, 2) if entry.price is not None else None
        row.price_match = method
        matched += 1
    logger.info(
        "document_analysis_agent.prices_enriched",
        matched=matched,
        total=len(rows),
        unmatched=len(rows) - matched,
        excluded_suppliers=excluded_suppliers,
    )


def _enrich_merged_with_country_of_origin(
    rows: list[MergedNomenclatureRow],
    db_country_index: dict[str, Any] | None,
) -> None:
    """Подставляет страну происхождения из onec_nomenclature по названию номенклатуры."""
    if not db_country_index:
        for row in rows:
            row.country_of_origin = None
        return

    candidates = [entry.nomenclature for entry in db_country_index.values()]
    matched = 0
    for row in rows:
        entry, _method = _match_catalog_entry(row.nomenclature, db_country_index, candidates)
        if entry is None:
            row.country_of_origin = None
            continue
        country = _clean_text(getattr(entry, "country_of_origin", None) or "")
        row.country_of_origin = country or None
        if row.country_of_origin:
            matched += 1
    logger.info(
        "document_analysis_agent.country_enriched",
        matched=matched,
        total=len(rows),
        catalog=len(db_country_index),
    )


def _enrich_merged_with_shipment_country(
    rows: list[MergedNomenclatureRow],
    *,
    shipment_bundle: ShipmentScheduleBundle | None,
) -> None:
    """Страна поставщика из объединённого графика отгрузок (Россия / Китай)."""
    if shipment_bundle is None or not shipment_bundle.receipt_index:
        return

    candidates = [entry.nomenclature for entry in shipment_bundle.receipt_index.values()]
    matched = 0
    for row in rows:
        entry, _method = _match_catalog_entry(row.nomenclature, shipment_bundle.receipt_index, candidates)
        if entry is None or not entry.country_of_origin:
            continue
        row.country_of_origin = entry.country_of_origin
        matched += 1

    logger.info(
        "document_analysis_agent.shipment_country_enriched",
        matched=matched,
        total=len(rows),
        files=shipment_bundle.shipment_files,
    )


def _enrich_merged_with_units(
    rows: list[MergedNomenclatureRow],
    db_unit_index: dict[str, Any] | None,
) -> None:
    """Дополняет ед. изм. из onec_nomenclature, если в материалах спецификации пусто."""
    if not db_unit_index:
        return

    candidates = [entry.nomenclature for entry in db_unit_index.values()]
    matched = 0
    for row in rows:
        if row.unit and str(row.unit).strip():
            continue
        entry, _method = _match_catalog_entry(row.nomenclature, db_unit_index, candidates)
        if entry is None:
            continue
        unit = _clean_text(getattr(entry, "unit", None) or "")
        if unit:
            row.unit = unit
            matched += 1
    logger.info(
        "document_analysis_agent.unit_enriched",
        matched=matched,
        total=len(rows),
        catalog=len(db_unit_index),
    )


def _enrich_merged_with_stock(
    rows: list[MergedNomenclatureRow],
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    *,
    db_stock_index: dict[str, StockEntry] | None = None,
) -> list[str]:
    """Дополняет итоговую структуру остатками из БД 1С (или legacy Excel stock)."""
    if db_stock_index is not None:
        index = db_stock_index
        stock_files = ["1С → PostgreSQL (onec_stock_balances)"]
    else:
        index, stock_files = _load_stock_index(workbooks, role_map)
    if not index:
        for row in rows:
            row.stock = 0.0
            row.ordered = 0.0
            row.stock_match = "unmatched"
        return stock_files

    candidates = [entry.nomenclature for entry in index.values()]
    matched = 0
    ordered_filled = 0
    for row in rows:
        entry, method = _match_catalog_entry(row.nomenclature, index, candidates)
        if entry is None:
            row.stock = 0.0
            row.ordered = 0.0
            row.stock_match = "unmatched"
            continue
        row.stock = entry.quantity if entry.quantity is not None else 0.0
        row.ordered = entry.ordered_qty if entry.ordered_qty is not None else 0.0
        row.stock_match = method
        matched += 1
        if entry.ordered_qty is not None:
            ordered_filled += 1
    logger.info(
        "document_analysis_agent.stock_enriched",
        matched=matched,
        ordered_filled=ordered_filled,
        total=len(rows),
        unmatched=len(rows) - matched,
        files=stock_files,
    )
    return stock_files


def _load_stock_index(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
) -> tuple[dict[str, StockEntry], list[str]]:
    """Парсит «Номенклатура» + «Остаток…» + опционально «Заказано…» из файлов остатков."""
    stock_files = [
        uploaded.filename
        for uploaded in workbooks
        if role_map.get(uploaded.filename) == ROLE_STOCK
    ]
    if not stock_files:
        return {}, []

    index: dict[str, StockEntry] = {}
    for uploaded in workbooks:
        if role_map.get(uploaded.filename) != ROLE_STOCK:
            continue
        workbook = load_workbook(BytesIO(uploaded.content), data_only=True, read_only=True)
        try:
            for worksheet in workbook.worksheets:
                _consume_stock_sheet(worksheet, index)
        finally:
            workbook.close()

    logger.info(
        "document_analysis_agent.stock_loaded",
        files=stock_files,
        unique=len(index),
        with_ordered=sum(1 for e in index.values() if e.ordered_qty is not None),
    )
    return index, stock_files


def _consume_stock_sheet(worksheet: Worksheet, index: dict[str, StockEntry]) -> None:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return

    header_idx = name_col = stock_col = None
    ordered_col: int | None = None
    for idx, row in enumerate(rows[:40]):
        normalized = [_normalize(value) for value in row]
        local_name = local_stock = local_ordered = None
        for col_idx, text in enumerate(normalized):
            if not text:
                continue
            if local_name is None and ("номенклатура" in text or text in {"наименование", "материал"}):
                local_name = col_idx
            elif local_ordered is None and "заказано" in text:
                local_ordered = col_idx
            elif local_stock is None and "остаток" in text:
                local_stock = col_idx
        if local_name is not None and local_stock is not None:
            header_idx, name_col, stock_col = idx, local_name, local_stock
            ordered_col = local_ordered
            break

    if header_idx is None or name_col is None or stock_col is None:
        return

    for row in rows[header_idx + 1 :]:
        if name_col >= len(row):
            continue
        name = _clean_text(row[name_col]).rstrip("*").strip()
        if not name or _normalize(name) in {"номенклатура", "наименование", "итого"}:
            continue
        qty = _to_float(row[stock_col]) if stock_col < len(row) else None
        ordered_qty = (
            _to_float(row[ordered_col])
            if ordered_col is not None and ordered_col < len(row)
            else None
        )
        key = _normalize(name)
        previous = index.get(key)
        if previous is None:
            index[key] = StockEntry(
                nomenclature=name, quantity=qty, ordered_qty=ordered_qty
            )
            continue
        if qty is not None:
            if previous.quantity is None:
                previous.quantity = qty
            else:
                previous.quantity = float(previous.quantity) + float(qty)
        if ordered_qty is not None:
            if previous.ordered_qty is None:
                previous.ordered_qty = ordered_qty
            else:
                previous.ordered_qty = float(previous.ordered_qty) + float(ordered_qty)


def _match_catalog_entry(
    nomenclature: str,
    key_to_entry: dict[str, TCatalogEntry],
    candidates: list[str],
) -> tuple[TCatalogEntry | None, str]:
    key = _normalize(nomenclature.rstrip("*"))
    if key in key_to_entry:
        return key_to_entry[key], "exact"

    contains_key = _match_catalog_key_by_containment(key, list(key_to_entry.keys()))
    if contains_key is not None:
        return key_to_entry[contains_key], "contains"

    best_name, score = _best_text_match(nomenclature, candidates)
    if best_name and score >= _PRICE_FUZZY_THRESHOLD:
        return key_to_entry[_normalize(best_name)], "fuzzy"
    return None, "unmatched"


def _match_catalog_key_by_containment(key: str, candidate_keys: list[str]) -> str | None:
    """Сопоставление по вхождению артикула/короткого имени (SM6T15A ⊂ диод SM6T15A…)."""
    if len(key) < 4:
        return None
    best_key: str | None = None
    best_len = -1
    for candidate_key in candidate_keys:
        shorter, longer = (
            (candidate_key, key) if len(candidate_key) <= len(key) else (key, candidate_key)
        )
        if len(shorter) < 6:
            continue
        # короткая сторона должна содержать цифру или быть достаточно длинной
        if not re.search(r"[0-9a-z]*\d[0-9a-z]*", shorter):
            if len(shorter) < 12:
                continue
        if shorter in longer and len(shorter) > best_len:
            best_key = candidate_key
            best_len = len(shorter)
    return best_key


def _enrich_merged_with_monthly_demand(
    rows: list[MergedNomenclatureRow],
    schedule_plans: list[ScheduleProductPlan],
) -> None:
    """Потребность: demand[месяц][кат][план|факт] = Σ (qty_график × qty_спеки)."""
    if not rows:
        return

    plans_by_key = {_normalize(plan.product): plan for plan in schedule_plans}
    plan_names = [plan.product for plan in schedule_plans]
    months: list[str] = []
    for plan in schedule_plans:
        for month in plan.monthly_qty:
            if month not in months:
                months.append(month)

    if not months:
        for row in rows:
            row.monthly_demand = {}
        return

    for row in rows:
        demand = {month: _empty_month_bucket() for month in months}
        for product, spec_qty in row.by_product.items():
            plan = _match_schedule_plan_for_product(product, plans_by_key, plan_names)
            if plan is None:
                continue
            per_unit = float(spec_qty) if spec_qty is not None else 0.0
            if per_unit == 0:
                continue
            for month, month_bucket in plan.monthly_qty.items():
                target = demand.setdefault(month, _empty_month_bucket())
                for category in _SCHEDULE_CATEGORIES:
                    metrics = month_bucket.get(category) or {}
                    for metric in _SCHEDULE_METRICS:
                        product_qty = float(metrics.get(metric, 0.0))
                        if product_qty == 0:
                            continue
                        target[category][metric] = (
                            target[category].get(metric, 0.0) + product_qty * per_unit
                        )
        row.monthly_demand = {
            month: {
                category: {
                    metric: _round_qty(float(qty))
                    for metric, qty in metrics.items()
                }
                for category, metrics in month_bucket.items()
            }
            for month, month_bucket in demand.items()
        }

    logger.info(
        "document_analysis_agent.monthly_demand_enriched",
        nomenclatures=len(rows),
        months=months,
        nonzero=sum(1 for row in rows if _monthly_demand_has_nonzero(row.monthly_demand)),
    )


def _enrich_merged_with_monthly_receipts(
    rows: list[MergedNomenclatureRow],
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    *,
    shipment_bundle: ShipmentScheduleBundle | None = None,
) -> list[str]:
    """Ожидаемое поступление по месяцам и неделям из графика отгрузок."""
    if shipment_bundle is None:
        index, shipment_files = _load_shipment_receipts_index(workbooks, role_map)
    else:
        index = shipment_bundle.receipt_index
        shipment_files = shipment_bundle.shipment_files
    months = sorted(
        {month for entry in index.values() for month in entry.monthly_qty},
        key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
    )
    # месяцы шапки = потребность ∪ отгрузки
    sheet_months = _months_for_monthly_sheet(rows)
    for month in months:
        if month not in sheet_months:
            sheet_months.append(month)
    sheet_months = sorted(
        sheet_months,
        key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
    )
    weeks_by_month = _weeks_by_month_labels(sheet_months)

    def _empty_weekly() -> dict[str, dict[str, float]]:
        return {
            month: {span.key: 0.0 for span in weeks_by_month.get(month, [])}
            for month in sheet_months
        }

    # если в графике дат нет — всё равно обнуляем поля под месяцы потребности
    if not months:
        for row in rows:
            seed = list(row.monthly_demand.keys()) or list(_MONTH_NOMINATIVE[6:12])
            row.monthly_receipts = {month: 0.0 for month in seed}
            row.weekly_receipts = {
                month: {span.key: 0.0 for span in _weeks_by_month_labels(seed).get(month, [])}
                for month in seed
            }
        return shipment_files

    if not index:
        for row in rows:
            row.monthly_receipts = {month: 0.0 for month in months}
            row.weekly_receipts = _empty_weekly()
        return shipment_files

    candidates = [entry.nomenclature for entry in index.values()]
    matched = 0
    weekly_nonzero = 0
    for row in rows:
        receipts = {month: 0.0 for month in months}
        # также покрываем месяцы из потребности, даже если в отгрузках их не было
        for month in row.monthly_demand:
            receipts.setdefault(month, 0.0)
        weekly = _empty_weekly()
        entry, _method = _match_catalog_entry(row.nomenclature, index, candidates)
        if entry is not None:
            matched += 1
            row.shipment_nomenclature = _shipment_schedule_display_name(entry.nomenclature)
            for month, qty in entry.monthly_qty.items():
                receipts[month] = receipts.get(month, 0.0) + float(qty)
            daily = entry.daily_qty
            for month, spans in weeks_by_month.items():
                for span in spans:
                    qty = _qty_in_week_span(daily, span)
                    weekly[month][span.key] = _round_qty(qty)
                    if qty > 0:
                        weekly_nonzero += 1
        row.monthly_receipts = {
            month: round(value, 6) if abs(value - round(value)) > 1e-9 else float(round(value))
            for month, value in receipts.items()
        }
        row.weekly_receipts = weekly

    logger.info(
        "document_analysis_agent.monthly_receipts_enriched",
        matched=matched,
        total=len(rows),
        files=shipment_files,
        months=months,
        week_buckets=sum(len(spans) for spans in weeks_by_month.values()),
        weekly_nonzero_cells=weekly_nonzero,
        nonzero=sum(1 for row in rows if any(v > 0 for v in row.monthly_receipts.values())),
    )
    return shipment_files


def _enrich_merged_with_monthly_forecast(rows: list[MergedNomenclatureRow]) -> None:
    """Прогноз: остаток + поступление − Σ(планы по категориям); факт не участвует."""
    for row in rows:
        months = sorted(
            set(row.monthly_demand) | set(row.monthly_receipts),
            key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
        )
        if not months:
            seed = list(_MONTH_NOMINATIVE[6:12])
            row.monthly_forecast = {month: 0.0 for month in seed}
            continue

        balance = 0.0 if row.stock is None else float(row.stock)
        forecasts: dict[str, float] = {}
        for month in months:
            balance = (
                balance
                + float(row.monthly_receipts.get(month, 0.0))
                - _plan_demand_total(row.monthly_demand.get(month))
            )
            forecasts[month] = _round_qty(balance)
        row.monthly_forecast = forecasts

    deficit_rows = sum(
        1 for row in rows if any(value < 0 for value in row.monthly_forecast.values())
    )
    logger.info(
        "document_analysis_agent.monthly_forecast_enriched",
        nomenclatures=len(rows),
        deficit_rows=deficit_rows,
    )


def _match_plan_product_name(
    product: str,
    plans_by_key: dict[str, Any],
    plan_names: list[str],
) -> Any | None:
    """Сопоставляет изделие из спеки с коротким/полным именем в графике."""
    key = _normalize(product)
    plan = plans_by_key.get(key)
    if plan is not None:
        return plan

    contains_key = _match_catalog_key_by_containment(key, list(plans_by_key.keys()))
    if contains_key is not None:
        return plans_by_key[contains_key]

    best_plan = None
    best_len = -1
    for plan_key, candidate in plans_by_key.items():
        if len(plan_key) < 4:
            continue
        if plan_key in key and len(plan_key) > best_len:
            best_plan = candidate
            best_len = len(plan_key)
        elif key in plan_key and len(key) >= 6 and len(key) > best_len:
            best_plan = candidate
            best_len = len(key)
    if best_plan is not None:
        return best_plan

    best_name, score = _best_text_match(product, plan_names)
    if best_name and score >= 0.72:
        return plans_by_key[_normalize(best_name)]
    return None


def _match_schedule_plan_for_product(
    product: str,
    plans_by_key: dict[str, ScheduleProductPlan],
    plan_names: list[str],
) -> ScheduleProductPlan | None:
    return _match_plan_product_name(product, plans_by_key, plan_names)


def _detailed_plan_name_aliases(key: str) -> list[str]:
    aliases = [key]
    if key.endswith(" ист"):
        base = key[: -len(" ист")].rstrip()
        aliases.extend([f"{base} ис -т", f"{base} ис-т", f"{base} ис т"])
    if key.endswith(" ит"):
        base = key[: -len(" ит")].rstrip()
        aliases.append(f"{base} т")
    seen: set[str] = set()
    ordered: list[str] = []
    for item in aliases:
        normalized = _match_key(item)
        if normalized and normalized not in seen:
            seen.add(normalized)
            ordered.append(normalized)
    return ordered


def _alias_fits_product(alias: str, product_key: str) -> bool:
    if not alias or not product_key:
        return False
    start = 0
    while True:
        idx = product_key.find(alias, start)
        if idx < 0:
            return False
        after = idx + len(alias)
        before_ok = idx == 0 or not product_key[idx - 1].isalnum()
        after_ok = after >= len(product_key) or not product_key[after].isalnum()
        if before_ok and after_ok:
            return True
        start = idx + 1


def _match_detailed_plan_by_alias(
    product: str,
    plans_by_key: dict[str, DetailedScheduleProductPlan],
    plan_names: list[str],
) -> DetailedScheduleProductPlan | None:
    product_key = _match_key(product)
    best_key: str | None = None
    best_rank: tuple[int, float] = (-1, 0.0)
    for plan_name in plan_names:
        plan_key = _normalize(plan_name)
        match_key = _match_key(plan_name)
        for alias in _detailed_plan_name_aliases(match_key):
            if not (
                _alias_fits_product(alias, product_key)
                or _alias_fits_product(product_key, alias)
            ):
                continue
            rank = (len(alias), _product_match_score(product, plan_name))
            if rank > best_rank:
                best_key = plan_key
                best_rank = rank
    if best_key is None:
        return None
    return plans_by_key.get(best_key)


def _match_detailed_plan_for_product(
    product: str,
    plans_by_key: dict[str, DetailedScheduleProductPlan],
    plan_names: list[str],
) -> DetailedScheduleProductPlan | None:
    alias_match = _match_detailed_plan_by_alias(product, plans_by_key, plan_names)
    if alias_match is not None:
        return alias_match
    return _match_plan_product_name(product, plans_by_key, plan_names)


def _enrich_merged_with_daily_demand(
    rows: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract,
) -> None:
    """Потребность по дням из детального графика (строки П/ф): план и факт отдельно.

    demand_plan[day] = Σ (план_изделия[день] × qty_спеки)
    demand_fact[day] = Σ (факт_изделия[день] × qty_спеки) — только дни с фактом в файле.
    """
    day_keys = list(detailed.day_keys) or _month_day_keys(detailed.year, detailed.month)
    if not rows:
        return
    if not day_keys:
        for row in rows:
            row.daily_demand = {}
            row.daily_demand_fact = {}
        return

    plans_by_key = {_normalize(plan.product): plan for plan in detailed.plans}
    plan_names = [plan.product for plan in detailed.plans]
    rows_by_key = {_normalize(row.nomenclature): row for row in rows if _normalize(row.nomenclature)}

    try:
        from app.agents.document_analysis_agent.product_coverage import (
            build_boms_for_detailed_products,
        )

        detailed_boms = build_boms_for_detailed_products(plan_names, rows)
    except Exception:
        detailed_boms = {}

    for row in rows:
        row.daily_demand = {day: 0.0 for day in day_keys}
        row.daily_demand_fact = {day: 0.0 for day in day_keys}

    if detailed_boms:
        for plan in detailed.plans:
            bom = detailed_boms.get(plan.product)
            if bom is None or not getattr(bom, "matched", False):
                continue
            for mat_key, spec_qty in bom.materials.items():
                row = rows_by_key.get(mat_key)
                if row is None:
                    continue
                per_unit = float(spec_qty) if spec_qty is not None else 0.0
                if per_unit == 0:
                    continue
                for day_key, product_qty in plan.daily_qty.items():
                    if day_key not in row.daily_demand:
                        continue
                    row.daily_demand[day_key] = (
                        row.daily_demand.get(day_key, 0.0)
                        + float(product_qty) * per_unit
                    )
                for day_key, product_qty in plan.daily_fact.items():
                    if day_key not in row.daily_demand_fact:
                        continue
                    row.daily_demand_fact[day_key] = (
                        row.daily_demand_fact.get(day_key, 0.0)
                        + float(product_qty) * per_unit
                    )
    else:
        for row in rows:
            for product, spec_qty in row.by_product.items():
                plan = _match_detailed_plan_for_product(product, plans_by_key, plan_names)
                if plan is None:
                    continue
                per_unit = float(spec_qty) if spec_qty is not None else 0.0
                if per_unit == 0:
                    continue
                for day_key, product_qty in plan.daily_qty.items():
                    if day_key not in row.daily_demand:
                        continue
                    row.daily_demand[day_key] = (
                        row.daily_demand.get(day_key, 0.0)
                        + float(product_qty) * per_unit
                    )
                for day_key, product_qty in plan.daily_fact.items():
                    if day_key not in row.daily_demand_fact:
                        continue
                    row.daily_demand_fact[day_key] = (
                        row.daily_demand_fact.get(day_key, 0.0)
                        + float(product_qty) * per_unit
                    )

    def _round_map(src: dict[str, float]) -> dict[str, float]:
        return {
            day: round(value, 6) if abs(value - round(value)) > 1e-9 else float(round(value))
            for day, value in src.items()
        }

    for row in rows:
        row.daily_demand = _round_map(row.daily_demand)
        row.daily_demand_fact = _round_map(row.daily_demand_fact)

    logger.info(
        "document_analysis_agent.daily_demand_enriched",
        nomenclatures=len(rows),
        month=f"{detailed.year:04d}-{detailed.month:02d}",
        days=len(day_keys),
        plans=len(detailed.plans),
        nonzero_plan=sum(1 for row in rows if any(v > 0 for v in row.daily_demand.values())),
        nonzero_fact=sum(
            1 for row in rows if any(v > 0 for v in row.daily_demand_fact.values())
        ),
    )


def _enrich_merged_with_daily_receipts(
    rows: list[MergedNomenclatureRow],
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    detailed: DetailedScheduleExtract,
    *,
    shipment_bundle: ShipmentScheduleBundle | None = None,
) -> None:
    """Ожидаемые поступления по дням выбранного месяца из графика отгрузок."""
    day_keys = list(detailed.day_keys) or _month_day_keys(detailed.year, detailed.month)
    if not rows:
        return
    if not day_keys:
        for row in rows:
            row.daily_receipts = {}
        return

    if shipment_bundle is None:
        index, shipment_files = _load_shipment_receipts_index(workbooks, role_map)
    else:
        index = shipment_bundle.receipt_index
        shipment_files = shipment_bundle.shipment_files
    day_set = set(day_keys)
    if not index:
        for row in rows:
            row.daily_receipts = {day: 0.0 for day in day_keys}
        return

    candidates = [entry.nomenclature for entry in index.values()]
    matched = 0
    for row in rows:
        receipts = {day: 0.0 for day in day_keys}
        entry, _method = _match_catalog_entry(row.nomenclature, index, candidates)
        if entry is not None:
            matched += 1
            if not row.shipment_nomenclature:
                row.shipment_nomenclature = _shipment_schedule_display_name(entry.nomenclature)
            for day_key, qty in entry.daily_qty.items():
                if day_key in day_set:
                    receipts[day_key] = receipts.get(day_key, 0.0) + float(qty)
        row.daily_receipts = {
            day: round(value, 6) if abs(value - round(value)) > 1e-9 else float(round(value))
            for day, value in receipts.items()
        }

    logger.info(
        "document_analysis_agent.daily_receipts_enriched",
        matched=matched,
        total=len(rows),
        files=shipment_files,
        month=f"{detailed.year:04d}-{detailed.month:02d}",
        nonzero=sum(1 for row in rows if any(v > 0 for v in row.daily_receipts.values())),
    )


def _enrich_merged_with_daily_forecast(
    rows: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract,
) -> None:
    """Прогноз остатка по дням: цепочка stock + receipt − demand."""
    day_keys = list(detailed.day_keys) or _month_day_keys(detailed.year, detailed.month)
    for row in rows:
        if not day_keys:
            row.daily_forecast = {}
            continue
        balance = 0.0 if row.stock is None else float(row.stock)
        forecasts: dict[str, float] = {}
        for day_key in day_keys:
            balance = (
                balance
                + float(row.daily_receipts.get(day_key, 0.0))
                - float(row.daily_demand.get(day_key, 0.0))
            )
            forecasts[day_key] = (
                round(balance, 6) if abs(balance - round(balance)) > 1e-9 else float(round(balance))
            )
        row.daily_forecast = forecasts

    deficit_rows = sum(
        1 for row in rows if any(value < 0 for value in row.daily_forecast.values())
    )
    logger.info(
        "document_analysis_agent.daily_forecast_enriched",
        nomenclatures=len(rows),
        days=len(day_keys),
        deficit_rows=deficit_rows,
    )


def _load_shipment_schedule_bundle(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
) -> ShipmentScheduleBundle:
    """Один проход по файлам отгрузок: поступления, сроки логистики, листы для рисков."""
    shipment_files = [
        uploaded.filename
        for uploaded in workbooks
        if role_map.get(uploaded.filename) == ROLE_SHIPMENT_SCHEDULE
    ]
    if not shipment_files:
        return ShipmentScheduleBundle()

    receipt_index: dict[str, ShipmentReceiptEntry] = {}
    logistics_leads: dict[str, tuple[int, int]] = {}
    parsed_sheets: list[ShipmentParsedSheet] = []

    for uploaded in workbooks:
        if role_map.get(uploaded.filename) != ROLE_SHIPMENT_SCHEDULE:
            continue
        workbook = load_workbook(BytesIO(uploaded.content), data_only=True, read_only=True)
        try:
            for worksheet in workbook.worksheets:
                _consume_shipment_sheet(worksheet, receipt_index)
                _consume_shipment_logistics_leads(worksheet, logistics_leads)
                parsed = _parse_shipment_sheet_layout(worksheet)
                if parsed is None:
                    continue
                header_idx, name_col, _country_col, msk_col, rostov_col, date_cols = parsed
                if msk_col is None or rostov_col is None:
                    continue
                parsed_sheets.append(
                    ShipmentParsedSheet(
                        title=worksheet.title,
                        header_idx=header_idx,
                        name_col=name_col,
                        msk_col=msk_col,
                        rostov_col=rostov_col,
                        date_cols=list(date_cols),
                        rows=list(worksheet.iter_rows(values_only=True)),
                    )
                )
        finally:
            workbook.close()

    logger.info(
        "document_analysis_agent.shipment_bundle_loaded",
        files=shipment_files,
        unique=len(receipt_index),
        parsed_sheets=len(parsed_sheets),
        logistics_leads=len(logistics_leads),
    )
    return ShipmentScheduleBundle(
        receipt_index=receipt_index,
        shipment_files=shipment_files,
        logistics_leads=logistics_leads,
        parsed_sheets=parsed_sheets,
    )


def _load_shipment_receipts_index(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
) -> tuple[dict[str, ShipmentReceiptEntry], list[str]]:
    bundle = _load_shipment_schedule_bundle(workbooks, role_map)
    return bundle.receipt_index, bundle.shipment_files


def _shipment_schedule_display_name(value: str) -> str:
    """Каноническое имя номенклатуры как в объединённом графике отгрузок."""
    text = re.sub(r"\s+", " ", (value or "").strip())
    return re.sub(
        r"\s*\(\d+\)(?:\s+[A-Za-zА-Яа-я0-9._-]{1,20})?\s*$",
        "",
        text,
    ).strip()


def _is_merged_shipment_workbook(content: bytes) -> bool:
    workbook = load_workbook(BytesIO(content), read_only=True)
    try:
        return "График" in workbook.sheetnames and "Источник" in workbook.sheetnames
    finally:
        workbook.close()


def _read_merged_shipment_grafik_names(content: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        if "График" not in workbook.sheetnames:
            return []
        worksheet = workbook["График"]
        names: list[str] = []
        seen: set[str] = set()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = _clean_text(row[0]).strip()
            if not name:
                continue
            key = _normalize(name)
            if key in {"номенклатура", "наименование", "итого"} or key in seen:
                continue
            seen.add(key)
            names.append(name)
        return names
    finally:
        workbook.close()


def _schedule_name_index_from_shipment_bundle(
    bundle: ShipmentScheduleBundle,
) -> dict[str, str]:
    """Индекс имён номенклатуры для сменного задания из уже распарсенного графика отгрузок."""
    from app.agents.document_analysis_agent.temp_schedule_merge import build_schedule_name_index

    if not bundle.receipt_index:
        return {}
    names = [entry.nomenclature for entry in bundle.receipt_index.values()]
    index = build_schedule_name_index(names)
    logger.info(
        "document_analysis_agent.shift_schedule_names_from_bundle",
        count=len(index),
    )
    return index


async def _resolve_shipment_schedule_name_index(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    *,
    shipment_bundle: ShipmentScheduleBundle | None = None,
) -> dict[str, str]:
    """Канонические имена номенклатуры из объединённого графика отгрузок."""
    from app.agents.document_analysis_agent.temp_schedule_merge import (
        build_schedule_name_index,
        merge_schedule_files,
    )

    shipment_files = [
        uploaded
        for uploaded in workbooks
        if role_map.get(uploaded.filename) == ROLE_SHIPMENT_SCHEDULE
    ]
    if not shipment_files:
        return {}

    merged_candidates = [
        uploaded for uploaded in shipment_files if _is_merged_shipment_workbook(uploaded.content)
    ]
    if merged_candidates:
        names: list[str] = []
        seen: set[str] = set()
        for uploaded in merged_candidates:
            for name in _read_merged_shipment_grafik_names(uploaded.content):
                key = _normalize(name)
                if key in seen:
                    continue
                seen.add(key)
                names.append(name)
        if names:
            index = build_schedule_name_index(names)
            logger.info(
                "document_analysis_agent.shift_schedule_names_from_merged_file",
                files=[item.filename for item in merged_candidates],
                count=len(index),
            )
            return index

    merge_payload = [
        (uploaded.filename, uploaded.content)
        for uploaded in shipment_files
        if not uploaded.filename.startswith("~$")
        and not _is_merged_shipment_workbook(uploaded.content)
    ]
    if len(merge_payload) > 1:
        merge_result = await merge_schedule_files(merge_payload)
        if merge_result.get("ok"):
            file_base64 = merge_result.get("file_base64")
            if isinstance(file_base64, str) and file_base64:
                names = _read_merged_shipment_grafik_names(base64.b64decode(file_base64))
                if names:
                    index = build_schedule_name_index(names)
                    logger.info(
                        "document_analysis_agent.shift_schedule_names_from_merge",
                        source_files=len(merge_payload),
                        count=len(index),
                    )
                    return index

    if shipment_bundle is not None and shipment_bundle.receipt_index:
        index = _schedule_name_index_from_shipment_bundle(shipment_bundle)
        if index:
            return index

    index, _files = _load_shipment_receipts_index(workbooks, role_map)
    names = [entry.nomenclature for entry in index.values()]
    built = build_schedule_name_index(names)
    logger.info(
        "document_analysis_agent.shift_schedule_names_from_receipts_fallback",
        count=len(built),
    )
    return built


def _consume_shipment_sheet(
    worksheet: Worksheet,
    index: dict[str, ShipmentReceiptEntry],
) -> None:
    """Парсит лист графика отгрузок: qty под колонками дат → сумма по календарным месяцам."""
    parsed = _parse_shipment_sheet_layout(worksheet)
    if parsed is None:
        return
    header_idx, name_col, country_col, _msk_col, _rostov_col, date_cols = parsed

    rows = list(worksheet.iter_rows(values_only=True))
    for row in rows[header_idx + 1 :]:
        if name_col >= len(row):
            continue
        name = _clean_text(row[name_col]).rstrip("*").strip()
        if not name or _normalize(name) in {"номенклатура", "наименование", "итого"}:
            continue
        key = _normalize(name)
        entry = index.get(key)
        if entry is None:
            entry = ShipmentReceiptEntry(nomenclature=name)
            index[key] = entry
        if country_col is not None and country_col < len(row):
            country = _clean_text(row[country_col]).strip()
            if country:
                entry.country_of_origin = country

        for col_idx, delivery_date in date_cols:
            if col_idx >= len(row):
                continue
            qty = _to_float(row[col_idx])
            if qty is None or qty == 0:
                continue
            month_label = _MONTH_NOMINATIVE[delivery_date.month - 1]
            entry.monthly_qty[month_label] = entry.monthly_qty.get(month_label, 0.0) + float(qty)
            day_key = delivery_date.isoformat()
            entry.daily_qty[day_key] = entry.daily_qty.get(day_key, 0.0) + float(qty)


def _parse_shipment_sheet_layout(
    worksheet: Worksheet,
) -> tuple[int, int, int | None, int | None, int | None, list[tuple[int, date]]] | None:
    """Шапка листа отгрузок: номенклатура, страна, логистика МСК / МСК-Ростов, колонки дат."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return None

    for idx, row in enumerate(rows[:15]):
        name_col: int | None = None
        country_col: int | None = None
        msk_col: int | None = None
        rostov_col: int | None = None
        date_cols: list[tuple[int, date]] = []
        for col_idx, value in enumerate(row):
            text = _normalize(value)
            if name_col is None and text in {"номенклатура", "наименование"}:
                name_col = col_idx
                continue
            if country_col is None and text == "страна":
                country_col = col_idx
                continue
            if "логистика" in text:
                if "ростов" in text:
                    rostov_col = col_idx
                    continue
                if "мск" in text or "москв" in text:
                    msk_col = col_idx
                    continue
            delivery_date = _header_value_to_date(value)
            if delivery_date is not None:
                date_cols.append((col_idx, delivery_date))
        if name_col is not None and date_cols:
            return idx, name_col, country_col, msk_col, rostov_col, date_cols
    return None


def _parse_logistics_range(value: Any) -> tuple[int, int] | None:
    """«7-14 к.д.» / «5-10» → (короткая, длинная) в календарных днях."""
    text = _clean_text(value)
    if not text:
        return None
    match = _LOGISTICS_RANGE_RE.search(text)
    if not match:
        return None
    short_days = int(match.group(1))
    long_days = int(match.group(2))
    if short_days > long_days:
        short_days, long_days = long_days, short_days
    return short_days, long_days


def _load_shipment_logistics_leads(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    *,
    shipment_bundle: ShipmentScheduleBundle | None = None,
) -> dict[str, tuple[int, int]]:
    """norm → (long_MSK, long_Ростов); по дубликатам берём max каждого плеча."""
    if shipment_bundle is not None:
        return dict(shipment_bundle.logistics_leads)
    bundle = _load_shipment_schedule_bundle(workbooks, role_map)
    return dict(bundle.logistics_leads)


def _consume_shipment_logistics_leads(
    worksheet: Worksheet,
    index: dict[str, tuple[int, int]],
) -> None:
    """Собирает long_MSK / long_Ростов по номенклатуре с листа отгрузок."""
    parsed = _parse_shipment_sheet_layout(worksheet)
    if parsed is None:
        return
    header_idx, name_col, _country_col, msk_col, rostov_col, _date_cols = parsed
    if msk_col is None and rostov_col is None:
        return

    rows = list(worksheet.iter_rows(values_only=True))
    for row in rows[header_idx + 1 :]:
        if name_col >= len(row):
            continue
        name = _clean_text(row[name_col]).rstrip("*").strip()
        if not name or _normalize(name) in {"номенклатура", "наименование", "итого"}:
            continue
        key = _normalize(name)
        long_msk = 0
        long_rostov = 0
        if msk_col is not None and msk_col < len(row):
            parsed_msk = _parse_logistics_range(row[msk_col])
            if parsed_msk is not None:
                long_msk = parsed_msk[1]
        if rostov_col is not None and rostov_col < len(row):
            parsed_rostov = _parse_logistics_range(row[rostov_col])
            if parsed_rostov is not None:
                long_rostov = parsed_rostov[1]
        if long_msk == 0 and long_rostov == 0:
            continue
        prev = index.get(key)
        if prev is None:
            index[key] = (long_msk, long_rostov)
        else:
            index[key] = (max(prev[0], long_msk), max(prev[1], long_rostov))


def _logistics_stage_windows(
    moscow_date: date,
    short_msk: int,
    long_msk: int,
    short_rostov: int,
    long_rostov: int,
) -> dict[str, tuple[date, date]]:
    """Окна стадий от плановой поставки в Москву (D).

    Цепочка: загрузка → МСК [short…long] → таможня (+2) → Ростов [short…long].

    1. Загрузка: точечно D − long_msk − таможня − long_rostov
    2. МСК: [D − (long_msk − short_msk) … D]
    3. Таможня: точечно D + 2
    4. Ростов: [D + 2 + short_rostov … D + 2 + long_rostov]
    """
    early_msk_offset = max(0, long_msk - short_msk)
    customs = _LOGISTICS_CUSTOMS_DAYS
    cleared = moscow_date + timedelta(days=customs)
    load_day = moscow_date - timedelta(days=long_msk + customs + long_rostov)
    msk_start = moscow_date - timedelta(days=early_msk_offset)
    rostov_start = cleared + timedelta(days=short_rostov)
    rostov_end = cleared + timedelta(days=long_rostov)
    return {
        "loading_dispatch": (load_day, load_day),
        "msk_arrival": (msk_start, moscow_date),
        "customs_clearance": (cleared, cleared),
        "rostov_arrival": (rostov_start, rostov_end),
    }


def _logistics_risk_metrics(
    as_of: date, window_start: date, window_end: date
) -> tuple[int, float, str]:
    """days_remaining, risk_ratio (1=зелёный запас, 0=красный риск), risk_level."""
    days_remaining = (window_end - as_of).days
    span = max(1, (window_end - window_start).days)
    risk_ratio = max(0.0, min(1.0, days_remaining / span))
    if days_remaining <= 0:
        level = "critical"
    elif risk_ratio <= 0.25:
        level = "high"
    elif risk_ratio <= 0.5:
        level = "medium"
    else:
        level = "low"
    return days_remaining, round(risk_ratio, 4), level


def _empty_logistics_risk_board(as_of: date | None = None) -> LogisticsRiskBoard:
    day = as_of or date.today()
    return LogisticsRiskBoard(
        as_of=day.isoformat(),
        stages=[
            LogisticsRiskStage(key=key, label=label) for key, label in _LOGISTICS_STAGE_DEFS
        ],
    )


def _build_logistics_risk_board(
    merged: list[MergedNomenclatureRow],
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, WorkbookRole],
    as_of: date | None = None,
    *,
    shipment_bundle: ShipmentScheduleBundle | None = None,
) -> LogisticsRiskBoard:
    """Собирает номенклатуры+поставщиков на контрольных точках логистики на сегодня."""
    as_of_day = as_of or date.today()
    board = _empty_logistics_risk_board(as_of_day)
    stage_map = {stage.key: stage for stage in board.stages}

    supplier_by_key = {_normalize(row.nomenclature): row.supplier for row in merged}
    supplier_candidates = [row.nomenclature for row in merged]
    merged_index = {_normalize(row.nomenclature): row for row in merged}

    # aggregate: (stage_key, nom_key, moscow_iso) -> item
    buckets: dict[tuple[str, str, str], LogisticsRiskItem] = {}

    parsed_sheets: list[ShipmentParsedSheet] = []
    if shipment_bundle is not None:
        parsed_sheets = list(shipment_bundle.parsed_sheets)
    else:
        bundle = _load_shipment_schedule_bundle(workbooks, role_map)
        parsed_sheets = list(bundle.parsed_sheets)

    for sheet in parsed_sheets:
        header_idx = sheet.header_idx
        name_col = sheet.name_col
        msk_col = sheet.msk_col
        rostov_col = sheet.rostov_col
        date_cols = sheet.date_cols
        rows = sheet.rows
        sheet_name = sheet.title
        if msk_col is None or rostov_col is None:
            continue
        for row in rows[header_idx + 1 :]:
            if name_col >= len(row):
                continue
            name = _clean_text(row[name_col]).rstrip("*").strip()
            if not name or _normalize(name) in {
                "номенклатура",
                "наименование",
                "итого",
            }:
                continue
            msk_range = (
                _parse_logistics_range(row[msk_col]) if msk_col < len(row) else None
            )
            rostov_range = (
                _parse_logistics_range(row[rostov_col])
                if rostov_col < len(row)
                else None
            )
            if msk_range is None or rostov_range is None:
                continue
            short_msk, long_msk = msk_range
            short_rostov, long_rostov = rostov_range

            nom_key = _normalize(name)
            supplier = _sanitize_result_supplier(supplier_by_key.get(nom_key))
            if supplier is None and merged_index:
                matched, _method = _match_catalog_entry(
                    name, merged_index, supplier_candidates
                )
                if matched is not None:
                    supplier = _sanitize_result_supplier(matched.supplier)

            for col_idx, moscow_date in date_cols:
                if col_idx >= len(row):
                    continue
                qty = _to_float(row[col_idx])
                if qty is None or qty == 0:
                    continue
                windows = _logistics_stage_windows(
                    moscow_date,
                    short_msk,
                    long_msk,
                    short_rostov,
                    long_rostov,
                )
                for stage_key, (window_start, window_end) in windows.items():
                    if not (window_start <= as_of_day <= window_end):
                        continue
                    days_remaining, risk_ratio, risk_level = _logistics_risk_metrics(
                        as_of_day, window_start, window_end
                    )
                    moscow_iso = moscow_date.isoformat()
                    bucket_key = (stage_key, nom_key, moscow_iso)
                    existing = buckets.get(bucket_key)
                    if existing is None:
                        buckets[bucket_key] = LogisticsRiskItem(
                            nomenclature=name,
                            supplier=supplier,
                            quantity=float(qty),
                            moscow_date=moscow_iso,
                            milestone_date=window_end.isoformat(),
                            sheet=sheet_name,
                            window_start=window_start.isoformat(),
                            window_end=window_end.isoformat(),
                            days_remaining=days_remaining,
                            risk_ratio=risk_ratio,
                            risk_level=risk_level,
                        )
                    else:
                        existing.quantity = round(existing.quantity + float(qty), 6)
                        if sheet_name not in existing.sheet.split(" · "):
                            existing.sheet = f"{existing.sheet} · {sheet_name}"
                        if not existing.supplier and supplier:
                            existing.supplier = supplier

    stage_order = {key: idx for idx, (key, _) in enumerate(_LOGISTICS_STAGE_DEFS)}
    for (stage_key, _nom_key, _moscow_iso), item in sorted(
        buckets.items(),
        key=lambda pair: (
            stage_order.get(pair[0][0], 99),
            pair[1].days_remaining,
            pair[1].nomenclature.lower(),
            pair[0][2],
        ),
    ):
        stage = stage_map.get(stage_key)
        if stage is not None:
            stage.items.append(item)

    logger.info(
        "document_analysis_agent.logistics_risks_built",
        as_of=board.as_of,
        stages={stage.key: len(stage.items) for stage in board.stages},
        total=sum(len(stage.items) for stage in board.stages),
    )
    return board


def _header_value_to_date(value: Any) -> date | None:
    """Дата в шапке графика отгрузок → date."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            parsed = from_excel(value)
            if isinstance(parsed, datetime):
                return parsed.date()
            if isinstance(parsed, date):
                return parsed
        except Exception:
            return None
        return None
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", text)
    if match:
        day_n = int(match.group(1))
        month_n = int(match.group(2))
        year_n = int(match.group(3))
        if year_n < 100:
            year_n += 2000
        try:
            return date(year_n, month_n, day_n)
        except ValueError:
            return None
    return None


def _header_value_to_month_label(value: Any) -> str | None:
    """Дата в шапке графика отгрузок → именительный месяц (Июль…)."""
    parsed = _header_value_to_date(value)
    if parsed is None:
        return None
    return _MONTH_NOMINATIVE[parsed.month - 1]


def _detect_month_metric_columns(worksheet: Worksheet, keyword: str) -> dict[str, int]:
    """Колонки метрики по месяцам из строки 4 шаблона Header.xlsx."""
    mapping: dict[str, int] = {}
    for col_idx in range(1, _RESULT_GRID_COLS + 1):
        text = _normalize(worksheet.cell(4, col_idx).value)
        if not text or keyword not in text:
            continue
        for month_idx, nominative in enumerate(_MONTH_NOMINATIVE):
            month_key = _normalize(nominative)
            genitive = _MONTH_GENITIVE[month_idx]
            if month_key in text or genitive in text:
                mapping[nominative] = col_idx
                break
    return mapping


def _detect_columns_by_keyword(worksheet: Worksheet, keyword: str) -> list[int]:
    """Индексы колонок строки 4, где заголовок содержит keyword (порядок слева направо)."""
    cols: list[int] = []
    for col_idx in range(1, _RESULT_GRID_COLS + 1):
        text = _normalize(worksheet.cell(4, col_idx).value)
        if text and keyword in text:
            cols.append(col_idx)
    return cols


def _build_forecast_column_chain(
    worksheet: Worksheet,
    demand_columns: dict[str, int],
    receipt_columns: dict[str, int],
) -> list[tuple[str, int, int, int]]:
    """Цепочка (месяц, col_потребность, col_поступление, col_прогноз) по шаблону Header."""
    forecast_cols = _detect_columns_by_keyword(worksheet, "прогнозируемый")
    months = sorted(
        demand_columns.keys(),
        key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
    )
    chain: list[tuple[str, int, int, int]] = []
    for index, month in enumerate(months):
        if index >= len(forecast_cols):
            break
        demand_col = demand_columns[month]
        receipt_col = receipt_columns.get(month, demand_col + 1)
        forecast_col = forecast_cols[index]
        chain.append((month, demand_col, receipt_col, forecast_col))
    return chain


def _apply_forecast_deficit_formatting(
    worksheet: Worksheet,
    forecast_cols: list[int],
    first_row: int,
    last_row: int,
) -> None:
    """Красная заливка при прогнозируемом остатке < 0 (как в эталоне обеспеченности)."""
    if last_row < first_row or not forecast_cols:
        return
    for col_idx in forecast_cols:
        letter = get_column_letter(col_idx)
        cell_range = f"{letter}{first_row}:{letter}{last_row}"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=_FORECAST_DEFICIT_FILL,
                font=_FORECAST_DEFICIT_FONT,
            ),
        )


def _months_for_monthly_sheet(rows: list[MergedNomenclatureRow]) -> list[str]:
    """Месяцы колонок = из потребности графика; receipt-only месяцы не добавляем в шапку."""
    months: list[str] = []
    for row in rows:
        for month in row.monthly_demand:
            if month not in months:
                months.append(month)
    if not months:
        months = list(_MONTH_NOMINATIVE[6:12])
    return sorted(
        months,
        key=lambda name: _MONTH_NOMINATIVE.index(name) if name in _MONTH_NOMINATIVE else 99,
    )


def _next_month_forecast_label(month: str) -> str:
    """Подпись «Прогнозируемый остаток на 01.MM.YYYY» (год = текущий / +1 для января)."""
    if month not in _MONTH_NOMINATIVE:
        return f"Прогнозируемый остаток после {month}"
    idx = _MONTH_NOMINATIVE.index(month)
    next_idx = (idx + 1) % 12
    next_month_num = next_idx + 1
    year = date.today().year
    if next_idx < idx:
        year += 1
    return f"Прогнозируемый остаток на 01.{next_month_num:02d}.{year}"


def _daily_assurance_sheet_title(month: int, *, year: int | None = None) -> str:
    """Имя дневного листа: «2-произв. план (Июль)» или с годом для архивных месяцев."""
    if 1 <= month <= 12:
        month_name = _MONTH_NOMINATIVE[month - 1]
    else:
        today = date.today()
        month_name = _MONTH_NOMINATIVE[today.month - 1]
    if year is not None:
        title = f"{_SHEET_DAILY_ASSURANCE_PREFIX}{month_name} {year})"
        if len(title) > 31:
            title = f"{_SHEET_DAILY_ASSURANCE_PREFIX}{month_name} '{year % 100:02d})"
        return title[:31]
    return f"{_SHEET_DAILY_ASSURANCE_PREFIX}{month_name})"


def _build_daily_plan_snapshot_payload(
    rows: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract,
) -> dict[str, Any]:
    """Сериализует дневной лист за месяц для последующих анализов."""
    year = detailed.year if detailed.year > 0 else date.today().year
    month = detailed.month if detailed.month > 0 else date.today().month
    day_keys = list(detailed.day_keys) or _month_day_keys(year, month)
    snapshot_rows: list[dict[str, Any]] = []
    for row in rows:
        snapshot_rows.append(
            {
                "nomenclature": row.nomenclature,
                "products": list(row.products),
                "supplier": row.supplier,
                "country_of_origin": row.country_of_origin,
                "unit": row.unit,
                "price": row.price,
                "stock": row.stock,
                "ordered": row.ordered,
                "daily_demand": dict(row.daily_demand),
                "daily_demand_fact": dict(row.daily_demand_fact),
                "daily_receipts": dict(row.daily_receipts),
                "daily_forecast": dict(row.daily_forecast),
            }
        )
    return {
        "year": year,
        "month": month,
        "day_keys": day_keys,
        "rows": snapshot_rows,
    }


def _build_result_xlsx(
    rows: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract | None = None,
    product_coverage: Any | None = None,
    order_plan: Any | None = None,
    daily_plan_coverage: Any | None = None,
    user_id: Any | None = None,
    spec_eligible_products: frozenset[str] | None = None,
) -> bytes:
    """Собирает result.xlsx: произв. план (мес.) + дневные листы + по обеспеч. + обеспеченность + план заказов."""
    from app.agents.document_analysis_agent.daily_plan_snapshot import (
        list_daily_plan_snapshots,
        period_key as daily_plan_period_key,
        save_daily_plan_snapshot,
    )

    _strip_excluded_suppliers_from_rows(rows)
    if detailed is None:
        today = date.today()
        detailed = DetailedScheduleExtract(
            files=[],
            plans=[],
            year=today.year,
            month=today.month,
            day_keys=_month_day_keys(today.year, today.month),
        )

    current_year = detailed.year if detailed.year > 0 else date.today().year
    current_month = detailed.month if detailed.month > 0 else date.today().month
    current_period = daily_plan_period_key(current_year, current_month)

    historical_snapshots: list[dict[str, Any]] = []
    if current_year > 0 and current_month > 0:
        # Все дневные листы, кроме текущего месяца детального графика — из снимков (не пересчёт).
        historical_snapshots = list_daily_plan_snapshots(
            user_id,
            exclude_period=current_period,
        )

    workbook = Workbook()
    monthly_ws = workbook.active
    monthly_layout = _write_monthly_assurance_sheet(monthly_ws, rows)

    daily_sheet_titles: list[str] = []
    for snapshot in historical_snapshots:
        snap_month = int(snapshot.get("month") or 0)
        snap_year = int(snapshot.get("year") or 0)
        historical_ws = workbook.create_sheet(
            _daily_assurance_sheet_title(snap_month, year=snap_year)
        )
        _write_daily_assurance_sheet_from_snapshot(historical_ws, snapshot)
        daily_sheet_titles.append(historical_ws.title)

    daily_month = current_month
    daily_ws = workbook.create_sheet(
        _daily_assurance_sheet_title(daily_month, year=current_year)
    )
    _write_daily_assurance_sheet(daily_ws, rows, detailed)
    daily_sheet_titles.append(daily_ws.title)

    if user_id is not None and current_year > 0 and current_month > 0 and rows:
        try:
            save_daily_plan_snapshot(
                user_id,
                _build_daily_plan_snapshot_payload(rows, detailed),
            )
        except Exception as exc:
            logger.warning(
                "document_analysis_agent.daily_plan_snapshot_save_failed",
                user_id=str(user_id),
                period=current_period,
                error=str(exc),
            )

    priority_title: str | None = None
    if detailed.source_bytes and detailed.source_sheet_name:
        priority_ws = workbook.create_sheet(_SHEET_DETAILED_PRIORITY)
        _write_detailed_schedule_priority_sheet(
            priority_ws,
            detailed,
            daily_plan_coverage,
        )
        priority_title = priority_ws.title

    coverage_ws = workbook.create_sheet(_SHEET_PRODUCT_COVERAGE)
    _write_product_coverage_sheet(
        coverage_ws,
        product_coverage,
        spec_eligible_products=spec_eligible_products,
    )

    order_ws = workbook.create_sheet(_SHEET_ORDER_PLAN)
    _write_order_plan_sheet(
        order_ws,
        order_plan,
        monthly_sheet_title=monthly_ws.title,
        monthly_layout=monthly_layout,
        monthly_data_start_row=_MONTHLY_DATA_START_ROW,
    )

    sheet_names = [monthly_ws.title, *daily_sheet_titles]
    if priority_title:
        sheet_names.append(priority_title)
    sheet_names.extend([coverage_ws.title, order_ws.title])

    logger.info(
        "document_analysis_agent.result_xlsx_built",
        rows=len(rows),
        sheets=sheet_names,
        daily_month=current_period,
        daily_days=len(detailed.day_keys),
        daily_historical_sheets=len(historical_snapshots),
        coverage_products=(
            len(product_coverage.products_in_order) if product_coverage is not None else 0
        ),
        order_nomenclatures=(
            len(order_plan.nomenclatures) if order_plan is not None else 0
        ),
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _collect_schedule_total_rows(source: Worksheet) -> set[int]:
    """Строки «Итого …» в детальном графике — сохраняем исходную заливку."""
    rows: set[int] = set()
    max_row = int(source.max_row or 0)
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, min(int(source.max_column or 0), 5) + 1):
            text = _normalize(source.cell(row_idx, col_idx).value)
            if text.startswith("итого") or text.startswith("итог:"):
                rows.add(row_idx)
                break
    return rows


def _collect_schedule_total_columns(source: Worksheet) -> set[int]:
    """Колонки «Итог» в шапке детального графика — сохраняем исходную заливку."""
    total_rows = _collect_schedule_total_rows(source)
    cols: set[int] = set()
    scan_rows = min(int(source.max_row or 0), 15)
    max_col = int(source.max_column or 0)
    for row_idx in range(1, scan_rows + 1):
        if row_idx in total_rows:
            continue
        for col_idx in range(1, max_col + 1):
            text = _normalize(source.cell(row_idx, col_idx).value)
            if text and "итог" in text:
                cols.add(col_idx)
    return cols


def _clone_worksheet_into(
    target: Worksheet,
    source: Worksheet,
    *,
    strip_fills_except_totals: bool = False,
) -> None:
    """Копирует ячейки, стили, merges, размеры в уже созданный лист target."""
    total_rows: set[int] = set()
    total_cols: set[int] = set()
    if strip_fills_except_totals:
        total_rows = _collect_schedule_total_rows(source)
        total_cols = _collect_schedule_total_columns(source)

    target.sheet_view.showGridLines = source.sheet_view.showGridLines
    for row in source.iter_rows():
        for cell in row:
            dest = target.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dest.font = cell.font.copy()
                dest.border = cell.border.copy()
                if strip_fills_except_totals:
                    if cell.row in total_rows or cell.column in total_cols:
                        dest.fill = cell.fill.copy()
                    else:
                        dest.fill = _FILL_NONE
                else:
                    dest.fill = cell.fill.copy()
                dest.number_format = cell.number_format
                dest.protection = cell.protection.copy()
                dest.alignment = cell.alignment.copy()
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))
    for col_letter, col_dim in source.column_dimensions.items():
        target.column_dimensions[col_letter].width = col_dim.width
        target.column_dimensions[col_letter].hidden = col_dim.hidden
    for row_idx, row_dim in source.row_dimensions.items():
        target.row_dimensions[row_idx].height = row_dim.height
        target.row_dimensions[row_idx].hidden = row_dim.hidden


def _priority_fill_for_status(status: str | None) -> PatternFill | None:
    if status == "green":
        return _FILL_COVER_GREEN
    if status == "yellow":
        return _FILL_COVER_YELLOW
    if status == "red":
        return _FILL_COVER_RED
    return None


def _unmerge_header_over_columns(
    worksheet: Worksheet,
    start_col: int,
    end_col: int,
    *,
    max_row: int = 20,
) -> None:
    to_remove: list[str] = []
    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row > max_row:
            continue
        if merged.max_col < start_col or merged.min_col > end_col:
            continue
        to_remove.append(str(merged))
    for ref in to_remove:
        worksheet.unmerge_cells(ref)


def _insert_row_preserving_merges(worksheet: Worksheet, row_idx: int) -> None:
    """openpyxl does not reliably move merged ranges on insert_rows."""
    shifted: list[tuple[int, int, int, int]] = []
    for merged in list(worksheet.merged_cells.ranges):
        if merged.max_row < row_idx:
            continue
        min_row = merged.min_row
        max_row = merged.max_row
        if merged.min_row >= row_idx:
            min_row += 1
            max_row += 1
        else:
            max_row += 1
        shifted.append(
            (
                min_row,
                max_row,
                merged.min_col,
                merged.max_col,
            )
        )
        worksheet.unmerge_cells(str(merged))

    worksheet.insert_rows(row_idx)

    for min_row, max_row, min_col, max_col in shifted:
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def _insert_cols_preserving_merges(worksheet: Worksheet, col_idx: int, amount: int) -> None:
    """insert_cols без потери/съезда объединённых дневных шапок."""
    if amount <= 0:
        return

    shifted: list[tuple[int, int, int, int]] = []
    for merged in list(worksheet.merged_cells.ranges):
        if merged.max_col < col_idx:
            continue
        min_col = merged.min_col
        max_col = merged.max_col
        if merged.min_col >= col_idx:
            min_col += amount
            max_col += amount
        else:
            max_col += amount
        shifted.append(
            (
                merged.min_row,
                merged.max_row,
                min_col,
                max_col,
            )
        )
        worksheet.unmerge_cells(str(merged))

    worksheet.insert_cols(col_idx, amount)

    for min_row, max_row, min_col, max_col in shifted:
        worksheet.merge_cells(
            start_row=min_row,
            start_column=min_col,
            end_row=max_row,
            end_column=max_col,
        )


def _style_priority_metric_header(worksheet: Worksheet, row_idx: int, col_idx: int, label: str) -> None:
    cell = worksheet.cell(row_idx, col_idx, label)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = _HEADER_METRIC_FILL
    cell.border = _PRIORITY_GRID_BORDER


def _style_priority_fixed_headers(
    worksheet: Worksheet,
    *,
    header_row: int,
    metric_row: int,
    end_col: int,
) -> None:
    for col_idx in range(1, end_col + 1):
        _unmerge_header_over_columns(worksheet, col_idx, col_idx, max_row=metric_row)
        worksheet.merge_cells(
            start_row=header_row,
            start_column=col_idx,
            end_row=metric_row,
            end_column=col_idx,
        )
        cell = worksheet.cell(header_row, col_idx)
        cell.font = copy(_HEADER_GROUP_FONT)
        cell.fill = _HEADER_GROUP_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _PRIORITY_GRID_BORDER


def _style_priority_day_header(
    worksheet: Worksheet,
    date_row: int,
    metric_row: int,
    cov_col: int,
    plan_col: int,
    fact_col: int | None,
    day: date,
    *,
    unmerge_max_row: int | None = None,
) -> None:
    end_col = fact_col if fact_col is not None else plan_col
    _unmerge_header_over_columns(
        worksheet,
        cov_col,
        end_col,
        max_row=metric_row if unmerge_max_row is None else unmerge_max_row,
    )
    label = f"{day.day:02d}.{day.month:02d}"
    worksheet.merge_cells(
        start_row=date_row,
        start_column=cov_col,
        end_row=date_row,
        end_column=end_col,
    )
    date_cell = worksheet.cell(date_row, cov_col, label)
    date_cell.font = copy(_HEADER_GROUP_FONT)
    date_cell.fill = _HEADER_GROUP_FILL
    date_cell.border = _PRIORITY_GRID_BORDER
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(cov_col, end_col + 1):
        worksheet.cell(date_row, col_idx).border = _PRIORITY_GRID_BORDER
    worksheet.column_dimensions[get_column_letter(cov_col)].width = 10
    worksheet.column_dimensions[get_column_letter(plan_col)].width = 10
    if fact_col is not None:
        worksheet.column_dimensions[get_column_letter(fact_col)].width = 10
    _style_priority_metric_header(worksheet, metric_row, cov_col, "обесп")
    _style_priority_metric_header(worksheet, metric_row, plan_col, "план")
    if fact_col is not None:
        _style_priority_metric_header(worksheet, metric_row, fact_col, "факт")


def _resolve_pf_row_product(
    worksheet: Worksheet,
    row_idx: int,
    *,
    name_col: int,
    metric_row: int,
) -> str:
    current = ""
    for scan_row in range(metric_row + 1, row_idx + 1):
        name = _clean_text(worksheet.cell(scan_row, name_col).value)
        if name and _is_schedule_product_name(name):
            current = name
    return current


def _style_priority_value_cell(cell: Any) -> None:
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _PRIORITY_GRID_BORDER
    cell.number_format = "0"


def _priority_border(
    *,
    left: Side = _PRIORITY_GRID_SIDE,
    right: Side = _PRIORITY_GRID_SIDE,
    top: Side = _PRIORITY_GRID_SIDE,
    bottom: Side = _PRIORITY_GRID_SIDE,
) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


def _apply_priority_day_grid(
    worksheet: Worksheet,
    *,
    start_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Единая тонкая сетка для всего дневного блока, включая пустые и итоговые строки."""
    for row_idx in range(start_row, int(worksheet.max_row or 0) + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = worksheet.cell(row_idx, col_idx)
            cell.border = _PRIORITY_GRID_BORDER
            if row_idx > start_row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for merged in worksheet.merged_cells.ranges:
        if merged.max_row < start_row:
            continue
        if merged.max_col < start_col or merged.min_col > end_col:
            continue
        worksheet.cell(merged.min_row, merged.min_col).border = _PRIORITY_GRID_BORDER


def _apply_priority_day_side_borders(
    worksheet: Worksheet,
    *,
    start_row: int,
    day_columns: list[tuple[int, date]],
) -> None:
    """Толстые боковые границы вокруг каждого дневного блока."""
    max_row = int(worksheet.max_row or 0)
    for row_idx in range(start_row, max_row + 1):
        for cov_col, _day in day_columns:
            plan_col = cov_col + 1
            fact_col = cov_col + 2
            if row_idx == start_row:
                worksheet.cell(row_idx, cov_col).border = _priority_border(
                    left=_PRIORITY_DAY_SIDE,
                    right=_PRIORITY_DAY_SIDE,
                )
                continue
            worksheet.cell(row_idx, cov_col).border = _priority_border(
                left=_PRIORITY_DAY_SIDE
            )
            worksheet.cell(row_idx, plan_col).border = _PRIORITY_GRID_BORDER
            worksheet.cell(row_idx, fact_col).border = _priority_border(
                right=_PRIORITY_DAY_SIDE
            )


def _fill_priority_coverage_plan_block(
    worksheet: Worksheet,
    *,
    metric_row: int,
    name_col: int,
    stage_col: int,
    base_col: int,
    days: list[date],
    daily_plan_coverage: Any,
) -> int:
    today_iso = date.today().isoformat()
    filled = 0
    for row_idx in range(metric_row + 1, int(worksheet.max_row or 0) + 1):
        stage = _normalize(worksheet.cell(row_idx, stage_col).value)
        if _detailed_release_stage_priority(stage) is None:
            continue
        product = _resolve_pf_row_product(
            worksheet, row_idx, name_col=name_col, metric_row=metric_row
        )
        if not product:
            continue
        for index, day in enumerate(days):
            cov_col = base_col + index * 3
            plan_col = base_col + index * 3 + 1
            fact_col = base_col + index * 3 + 2
            day_iso = day.isoformat()
            day_cell = daily_plan_coverage.cell(product, day_iso)
            covered = float(day_cell.covered)
            plan = float(day_cell.plan)
            fact = float(getattr(day_cell, "fact", 0.0) or 0.0)
            cov_value = _round_qty(covered) if (plan > 1e-12 or covered > 1e-12) else None
            plan_value = _round_qty(plan) if plan > 1e-12 else None
            fact_value = _round_qty(fact) if fact > 1e-12 else None
            cov_cell = worksheet.cell(row_idx, cov_col)
            plan_cell = worksheet.cell(row_idx, plan_col)
            fact_cell = worksheet.cell(row_idx, fact_col)
            _safe_set_cell_value(worksheet, row_idx, cov_col, cov_value)
            _safe_set_cell_value(worksheet, row_idx, plan_col, plan_value)
            _safe_set_cell_value(worksheet, row_idx, fact_col, fact_value)
            _style_priority_value_cell(cov_cell)
            _style_priority_value_cell(plan_cell)
            _style_priority_value_cell(fact_cell)
            if day_iso >= today_iso and plan > 1e-12:
                status = daily_plan_coverage.status_for_plan_cell(
                    product, [day_iso], plan
                )
                fill = _priority_fill_for_status(status)
                if fill is not None:
                    cov_cell.fill = fill
            filled += 1
    return filled


def _fill_priority_generic_day_block(
    worksheet: Worksheet,
    *,
    first_data_row: int,
    name_col: int,
    day_columns: list[tuple[int, date]],
    daily_plan_coverage: Any,
) -> int:
    today_iso = date.today().isoformat()
    filled = 0
    for row_idx in range(first_data_row, int(worksheet.max_row or 0) + 1):
        product = _clean_text(worksheet.cell(row_idx, name_col).value)
        if not _is_schedule_product_name(product):
            continue
        for cov_col, day in day_columns:
            plan_col = cov_col + 1
            fact_col = cov_col + 2
            day_iso = day.isoformat()
            day_cell = daily_plan_coverage.cell(product, day_iso)
            covered = float(day_cell.covered)
            plan = float(day_cell.plan)
            daily_fact = float(getattr(day_cell, "fact", 0.0) or 0.0)
            cov_cell = worksheet.cell(row_idx, cov_col)
            plan_cell = worksheet.cell(row_idx, plan_col)
            fact_cell = worksheet.cell(row_idx, fact_col)
            _safe_set_cell_value(
                worksheet,
                row_idx,
                cov_col,
                _round_qty(covered) if (plan > 1e-12 or covered > 1e-12) else None,
            )
            _safe_set_cell_value(
                worksheet, row_idx, plan_col, _round_qty(plan) if plan > 1e-12 else None
            )
            _safe_set_cell_value(
                worksheet,
                row_idx,
                fact_col,
                _round_qty(daily_fact) if daily_fact > 1e-12 else None,
            )
            _style_priority_value_cell(cov_cell)
            _style_priority_value_cell(plan_cell)
            _style_priority_value_cell(fact_cell)
            if day_iso >= today_iso and plan > 1e-12:
                status = daily_plan_coverage.status_for_plan_cell(product, [day_iso], plan)
                fill = _priority_fill_for_status(status)
                if fill is not None:
                    cov_cell.fill = fill
            filled += 1
    return filled


def _expand_priority_generic_schedule_columns(
    worksheet: Worksheet,
    *,
    detailed: DetailedScheduleExtract,
    daily_plan_coverage: Any,
) -> dict[str, int]:
    """Разворачивает обычный «График выпуска»: день → обесп | план | факт."""
    tables = _iter_detailed_schedule_tables(worksheet, detailed.year, detailed.month)
    if not tables:
        return {"pairs": 0, "day_slots": 0, "cells_filled": 0}

    metric_rows: set[int] = set()
    for header_row, _name_col, _day_cols in tables:
        if header_row not in metric_rows:
            _insert_row_preserving_merges(worksheet, header_row + 1)
            metric_rows.add(header_row)

    shifted_tables: list[tuple[int, int, list[_DetailedDayColumn]]] = []
    for header_row, name_col, day_cols in tables:
        shift = sum(1 for row_idx in metric_rows if row_idx < header_row)
        shifted_tables.append(
            (
                header_row + shift,
                name_col,
                [_DetailedDayColumn(col=col_idx, day=day) for col_idx, day in day_cols],
            )
        )

    month_start = (
        date(detailed.year, detailed.month, 1)
        if detailed.year > 0 and detailed.month > 0
        else None
    )
    month_end = (
        date(detailed.year, detailed.month, monthrange(detailed.year, detailed.month)[1])
        if month_start is not None
        else None
    )

    day_slots = 0
    cells_filled = 0
    for header_row, name_col, day_cols in sorted(
        shifted_tables, key=lambda item: item[0], reverse=True
    ):
        metric_row = header_row + 1
        first_data_row = metric_row + 1
        cols = day_cols
        if month_start and month_end:
            cols = [item for item in cols if month_start <= item.day <= month_end]
        if not cols:
            continue
        cols = sorted(cols, key=lambda item: item.col)
        days = [item.day for item in cols]
        fixed_end_col = max(0, cols[0].col - 1)
        if fixed_end_col:
            _style_priority_fixed_headers(
                worksheet,
                header_row=header_row,
                metric_row=metric_row,
                end_col=fixed_end_col,
            )
        for item in sorted(cols, key=lambda item: item.col, reverse=True):
            _unmerge_header_over_columns(
                worksheet,
                item.col,
                item.col + 2,
                max_row=header_row,
            )
            _insert_cols_preserving_merges(worksheet, item.col, 2)
        day_columns = [
            (item.col + index * 2, item.day)
            for index, item in enumerate(cols)
        ]
        _unmerge_header_over_columns(
            worksheet,
            day_columns[0][0],
            day_columns[-1][0] + 2,
            max_row=header_row,
        )
        for cov_col, day in day_columns:
            _style_priority_day_header(
                worksheet,
                header_row,
                metric_row,
                cov_col,
                cov_col + 1,
                cov_col + 2,
                day,
                unmerge_max_row=header_row,
            )
        cells_filled += _fill_priority_generic_day_block(
            worksheet,
            first_data_row=first_data_row,
            name_col=name_col,
            day_columns=day_columns,
            daily_plan_coverage=daily_plan_coverage,
        )
        _apply_priority_day_grid(
            worksheet,
            start_row=header_row,
            start_col=day_columns[0][0],
            end_col=day_columns[-1][0] + 2,
        )
        _apply_priority_day_side_borders(
            worksheet,
            start_row=header_row,
            day_columns=day_columns,
        )
        day_slots += len(days)

    return {"pairs": len(shifted_tables), "day_slots": day_slots, "cells_filled": cells_filled}


def _expand_priority_sheet_coverage_columns(
    worksheet: Worksheet,
    *,
    detailed: DetailedScheduleExtract,
    daily_plan_coverage: Any,
) -> dict[str, int]:
    """Для каждого дня вставляет колонки обеспеченности, плана и факта."""
    layout = _find_pf_report_layout(worksheet, detailed.year, detailed.month)
    if layout is None:
        return _expand_priority_generic_schedule_columns(
            worksheet,
            detailed=detailed,
            daily_plan_coverage=daily_plan_coverage,
        )

    metric_row, name_col, stage_col, pairs = layout
    date_row = metric_row - 1 if metric_row > 1 else metric_row
    month_start = (
        date(detailed.year, detailed.month, 1)
        if detailed.year > 0 and detailed.month > 0
        else None
    )
    month_end = (
        date(detailed.year, detailed.month, monthrange(detailed.year, detailed.month)[1])
        if month_start is not None
        else None
    )

    day_slots = 0
    cells_filled = 0
    for pair in sorted(pairs, key=lambda item: item.plan_col, reverse=True):
        days = pair.days()
        if month_start and month_end:
            days = [item for item in days if month_start <= item <= month_end]
        if not days:
            continue
        days = sorted(days)
        original_width = 2 if pair.fact_col is not None else 1
        insert_count = max(len(days) * 3 - original_width, 0)
        _unmerge_header_over_columns(
            worksheet,
            pair.plan_col,
            pair.plan_col + insert_count + original_width,
            max_row=metric_row + 2,
        )
        _insert_cols_preserving_merges(worksheet, pair.plan_col, insert_count)
        for index, day in enumerate(days):
            cov_col = pair.plan_col + index * 3
            plan_col = pair.plan_col + index * 3 + 1
            fact_col = pair.plan_col + index * 3 + 2
            _style_priority_day_header(
                worksheet, date_row, metric_row, cov_col, plan_col, fact_col, day
            )
        cells_filled += _fill_priority_coverage_plan_block(
            worksheet,
            metric_row=metric_row,
            name_col=name_col,
            stage_col=stage_col,
            base_col=pair.plan_col,
            days=days,
            daily_plan_coverage=daily_plan_coverage,
        )
        _apply_priority_day_grid(
            worksheet,
            start_row=date_row,
            start_col=pair.plan_col,
            end_col=pair.plan_col + len(days) * 3 - 1,
        )
        _apply_priority_day_side_borders(
            worksheet,
            start_row=date_row,
            day_columns=[
                (pair.plan_col + index * 3, day)
                for index, day in enumerate(days)
            ],
        )
        day_slots += len(days)

    return {"pairs": len(pairs), "day_slots": day_slots, "cells_filled": cells_filled}


def _write_detailed_schedule_priority_sheet(
    worksheet: Worksheet,
    detailed: DetailedScheduleExtract,
    daily_plan_coverage: Any | None,
) -> None:
    """Копия листа Отчёта + колонки обесп/план по каждому дню."""
    worksheet.title = _SHEET_DETAILED_PRIORITY
    if not detailed.source_bytes or not detailed.source_sheet_name:
        worksheet["A1"] = "Нет исходного детального графика для копии"
        return

    source_wb = load_workbook(BytesIO(detailed.source_bytes), data_only=False)
    try:
        if detailed.source_sheet_name in source_wb.sheetnames:
            source_ws = source_wb[detailed.source_sheet_name]
        else:
            source_ws = source_wb.worksheets[0]
        _clone_worksheet_into(worksheet, source_ws, strip_fills_except_totals=True)
    finally:
        source_wb.close()

    # Легенда справа от таблицы (вне merges исходного Отчёта)
    legend_col = int(worksheet.max_column or 10) + 2
    legend_row = 2
    worksheet.cell(legend_row, legend_col, "Обеспеченность плана П/ф")
    worksheet.cell(legend_row, legend_col).font = Font(bold=True)
    legend_items = (
        (legend_row + 1, _FILL_COVER_GREEN, "зелёный — обеспеченность ≥ план"),
        (legend_row + 2, _FILL_COVER_YELLOW, "жёлтый — частично (0 < обесп < план)"),
        (legend_row + 3, _FILL_COVER_RED, "красный — не покрыт / нет спеки"),
        (legend_row + 4, None, "Колонки дня: обесп | план; заливка только в «обесп»"),
        (legend_row + 5, None, "Прошлые дни без заливки"),
    )
    for row_idx, fill, label in legend_items:
        cell = worksheet.cell(row_idx, legend_col, label)
        if fill is not None:
            cell.fill = fill

    # Закрепляем до заливки — даже если coverage нет.
    worksheet.freeze_panes = "D4"

    if daily_plan_coverage is None:
        return

    expand_stats = _expand_priority_sheet_coverage_columns(
        worksheet,
        detailed=detailed,
        daily_plan_coverage=daily_plan_coverage,
    )

    # Колонки A–C (№ / изделие·модель / П·ф) + шапка дат остаются при горизонтальном скролле.
    worksheet.freeze_panes = "D4"

    logger.info(
        "document_analysis_agent.detailed_priority_sheet_written",
        source_file=detailed.source_filename,
        source_sheet=detailed.source_sheet_name,
        plan_cells=len(detailed.plan_cells),
        pairs_expanded=expand_stats.get("pairs", 0),
        day_slots=expand_stats.get("day_slots", 0),
        cells_filled=expand_stats.get("cells_filled", 0),
        as_of=date.today().isoformat(),
    )


def _write_product_coverage_sheet(
    worksheet: Worksheet,
    coverage: Any | None,
    *,
    spec_eligible_products: frozenset[str] | None = None,
) -> None:
    """Лист «обеспеченность»: изделия + раскрываемые номенклатуры спеки (outline)."""
    from openpyxl.worksheet.properties import Outline

    from app.agents.document_analysis_agent.product_coverage import ProductCoverageResult

    worksheet.title = _SHEET_PRODUCT_COVERAGE
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    detail_left = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    thin = Side(style="thin", color="B0B0B0")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    product_fill = PatternFill(start_color="FFF7FBFF", end_color="FFF7FBFF", fill_type="solid")
    detail_fill = PatternFill(start_color="FFF9F9F9", end_color="FFF9F9F9", fill_type="solid")
    consumable_fill = PatternFill(start_color="FFEAF4FF", end_color="FFEAF4FF", fill_type="solid")
    workshop_fill = PatternFill(start_color="FFF2E9FF", end_color="FFF2E9FF", fill_type="solid")
    product_font = Font(bold=True, size=11, color="FF1F1F1F")
    detail_font = Font(bold=False, size=10, color="FF333333")

    if coverage is None or not isinstance(coverage, ProductCoverageResult):
        coverage = ProductCoverageResult(months=[], products_in_order=[], boms={})

    months = list(coverage.months)
    products = list(coverage.products_in_order)
    if spec_eligible_products:
        allowed = {_normalize(product) for product in spec_eligible_products}
        products = [
            product
            for product in products
            if _normalize(product) in allowed
        ]
    last_col = max(1, 1 + len(months) * 3)

    if worksheet.sheet_properties.outlinePr is None:
        worksheet.sheet_properties.outlinePr = Outline(
            summaryBelow=False, summaryRight=True, applyStyles=True
        )
    else:
        worksheet.sheet_properties.outlinePr.summaryBelow = False
        worksheet.sheet_properties.outlinePr.summaryRight = True

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = worksheet.cell(
        1,
        1,
        "Обеспеченность плана производства изделиями (сборка из материалов)",
    )
    _style_header_cell(title, fill=_HEADER_TITLE_FILL, font=_HEADER_TITLE_FONT, alignment=left)
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(1, col_idx),
            fill=_HEADER_TITLE_FILL,
            font=_HEADER_TITLE_FONT,
            alignment=left,
        )

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    subtitle = worksheet.cell(
        2,
        1,
        "Строка изделия: Обеспеченность = сколько изделий можно собрать "
        "(нужны ВСЕ позиции спеки: остаток+приход ≥ qty×число изделий); "
        "План = Σ(Заказ+Опытные+Склад)·План из графика производства; "
        "Факт = Σ(Заказ+Опытные+Склад)·Факт из графика производства. "
        "Раскройте «+» слева — номенклатуры спеки: "
        "Обеспеченность = остаток на начало + поступления месяца; "
        "План = план изделия × qty; Факт = факт изделия × qty.",
    )
    _style_header_cell(
        subtitle, fill=_HEADER_SUBTITLE_FILL, font=_HEADER_SUBTITLE_FONT, alignment=left
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(2, col_idx),
            fill=_HEADER_SUBTITLE_FILL,
            font=_HEADER_SUBTITLE_FONT,
            alignment=left,
        )

    # Row 3: Изделие + month group headers; Row 4: Обеспеченность | План | Факт
    worksheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    for row_idx in (3, 4):
        c = worksheet.cell(row_idx, 1, "Изделие / номенклатура" if row_idx == 3 else None)
        _style_header_cell(c, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center)
        c.border = header_border

    for month_index, month in enumerate(months):
        base = 2 + month_index * 3
        worksheet.merge_cells(start_row=3, start_column=base, end_row=3, end_column=base + 2)
        for offset in range(3):
            cell = worksheet.cell(3, base + offset, month if offset == 0 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border
        for offset, label in enumerate(("Обеспеченность", "План", "Факт")):
            cell = worksheet.cell(4, base + offset, label)
            _style_header_cell(
                cell, fill=_HEADER_METRIC_FILL, font=_HEADER_METRIC_FONT, alignment=center
            )
            cell.border = header_border

    excel_row = 5
    detail_rows = 0
    for product in products:
        product_row = excel_row
        name_cell = worksheet.cell(product_row, 1, product)
        name_cell.alignment = left
        name_cell.border = data_border
        name_cell.font = product_font
        name_cell.fill = product_fill
        for month_index, month in enumerate(months):
            base = 2 + month_index * 3
            cell_data = coverage.cell(product, month)
            strict_covered = float(cell_data.covered or 0.0)
            covered_cell = worksheet.cell(product_row, base, strict_covered)
            plan_cell = worksheet.cell(product_row, base + 1, cell_data.plan)
            fact_cell = worksheet.cell(product_row, base + 2, cell_data.fact)
            covered_cell.alignment = center
            plan_cell.alignment = center
            fact_cell.alignment = center
            covered_cell.border = data_border
            plan_cell.border = data_border
            fact_cell.border = data_border
            covered_cell.font = product_font
            plan_cell.font = product_font
            fact_cell.font = product_font
            covered_cell.fill = product_fill
            plan_cell.fill = product_fill
            fact_cell.fill = product_fill
            if cell_data.plan > 0 and strict_covered + 1e-9 < cell_data.plan:
                covered_cell.fill = warn_fill

        bom = coverage.boms.get(product)
        lines = bom.lines() if bom is not None else []
        excel_row += 1
        for line in lines:
            detail_row = excel_row
            detail_rows += 1
            line_kind = getattr(line, "material_kind", MATERIAL_KIND_REQUIRED)
            if line_kind == MATERIAL_KIND_CONSUMABLE:
                row_fill = consumable_fill
            elif line_kind == MATERIAL_KIND_WORKSHOP:
                row_fill = workshop_fill
            else:
                row_fill = detail_fill
            mat_cell = worksheet.cell(detail_row, 1, line.nomenclature)
            mat_cell.alignment = detail_left
            mat_cell.border = data_border
            mat_cell.font = detail_font
            mat_cell.fill = row_fill
            if is_optional_material_kind(line_kind):
                note = getattr(line, "material_kind_label", "") or MATERIAL_KIND_LABELS.get(
                    line_kind, ""
                )
                confidence = getattr(line, "material_kind_confidence", "")
                reason = getattr(line, "material_kind_reason", "")
                comment_parts = [note]
                if confidence:
                    comment_parts.append(f"уверенность: {confidence}")
                if reason:
                    comment_parts.append(reason)
                mat_cell.comment = Comment(
                    "; ".join(part for part in comment_parts if part),
                    "AI Platform",
                )
            for month_index, month in enumerate(months):
                base = 2 + month_index * 3
                available = coverage.material_available(month, line.norm_key)
                plan_need = coverage.material_plan(product, month, line.norm_key)
                fact_need = coverage.material_fact(product, month, line.norm_key)
                avail_cell = worksheet.cell(detail_row, base, available)
                plan_cell = worksheet.cell(detail_row, base + 1, plan_need)
                fact_cell = worksheet.cell(detail_row, base + 2, fact_need)
                avail_cell.alignment = center
                plan_cell.alignment = center
                fact_cell.alignment = center
                avail_cell.border = data_border
                plan_cell.border = data_border
                fact_cell.border = data_border
                avail_cell.font = detail_font
                plan_cell.font = detail_font
                fact_cell.font = detail_font
                avail_cell.fill = row_fill
                plan_cell.fill = row_fill
                fact_cell.fill = row_fill
                if (
                    plan_need > 0
                    and available + 1e-9 < plan_need
                    and not is_optional_material_kind(line_kind)
                ):
                    avail_cell.fill = warn_fill
            dim = worksheet.row_dimensions[detail_row]
            dim.outline_level = 1
            dim.hidden = True
            excel_row += 1

    worksheet.column_dimensions["A"].width = 56
    for col_idx in range(2, last_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 14
    worksheet.row_dimensions[1].height = 20
    worksheet.row_dimensions[2].height = 62
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[4].height = 30
    worksheet.freeze_panes = "B5"

    legend_row = excel_row + 1
    legend_title = worksheet.cell(legend_row, 1, "Легенда условной обеспеченности")
    legend_title.font = product_font
    legend_title.border = data_border

    legend_items = (
        (
            legend_row + 1,
            consumable_fill,
            "Голубой",
            "Возможно расходник: показан в раскрытии, но не блокирует условную обеспеченность изделия.",
        ),
        (
            legend_row + 2,
            workshop_fill,
            "Сиреневый",
            "Возможно в цехе: показан в раскрытии, но не блокирует условную обеспеченность изделия.",
        ),
    )
    for row_idx, fill, color_name, label in legend_items:
        color_cell = worksheet.cell(row_idx, 1, color_name)
        color_cell.fill = fill
        color_cell.border = data_border
        color_cell.alignment = center
        label_cell = worksheet.cell(row_idx, 2, label)
        label_cell.border = data_border
        label_cell.alignment = left

    logger.info(
        "document_analysis_agent.product_coverage_sheet_written",
        products=len(products),
        months=months,
        detail_rows=detail_rows,
        outline_collapsed=True,
    )


def _excel_sheet_ref(sheet_title: str, cell_ref: str) -> str:
    """Ссылка на ячейку другого листа: 'имя'!A1 (экранирование кавычек в имени)."""
    safe = sheet_title.replace("'", "''")
    return f"'{safe}'!{cell_ref}"


def _order_plan_qty_formula(
    *,
    monthly_sheet: str,
    monthly_row: int,
    order_row: int,
    months: list[str],
    month_index: int,
    monthly_layout: dict[str, dict[str, Any]],
    qty_cols_by_month: list[int],
) -> str | None:
    """MAX(0, demand − opening − receipt); opening тянется с остатка и прошлых заказов."""
    month = months[month_index]
    cols = monthly_layout.get(month)
    if cols is None:
        return None
    demand_ref = _excel_sheet_ref(
        monthly_sheet, f"{get_column_letter(cols['sum_plan'])}{monthly_row}"
    )
    receipt_ref = _excel_sheet_ref(
        monthly_sheet, f"{get_column_letter(cols['receipt'])}{monthly_row}"
    )
    opening = _excel_sheet_ref(monthly_sheet, f"{_STOCK_COL_LETTER}{monthly_row}")
    for j in range(month_index):
        prev = months[j]
        prev_cols = monthly_layout.get(prev)
        if prev_cols is None:
            return None
        prev_r = _excel_sheet_ref(
            monthly_sheet, f"{get_column_letter(prev_cols['receipt'])}{monthly_row}"
        )
        prev_d = _excel_sheet_ref(
            monthly_sheet, f"{get_column_letter(prev_cols['sum_plan'])}{monthly_row}"
        )
        prev_o = f"{get_column_letter(qty_cols_by_month[j])}{order_row}"
        opening = f"({opening}+{prev_r}+{prev_o}-{prev_d})"
    return f"=MAX(0,{demand_ref}-({opening})-{receipt_ref})"


def _write_order_plan_sheet(
    worksheet: Worksheet,
    order_plan: Any | None,
    *,
    monthly_sheet_title: str | None = None,
    monthly_layout: dict[str, dict[str, Any]] | None = None,
    monthly_data_start_row: int = _MONTHLY_DATA_START_ROW,
) -> None:
    """Лист «план заказов»: номенклатура × (дата заказа | количество) по месяцам.

    Количество — Excel-формулы на остаток/потребность/поступление с помесячного листа
    (и на qty прошлых месяцев этого листа), чтобы правка поступления пересчитывала заказ.
    """
    from app.agents.document_analysis_agent.order_plan import OrderPlanResult

    worksheet.title = _SHEET_ORDER_PLAN
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B0B0B0")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if order_plan is None or not isinstance(order_plan, OrderPlanResult):
        order_plan = OrderPlanResult(months=[], nomenclatures=[], year=date.today().year)

    months = list(order_plan.months)
    names = list(order_plan.nomenclatures)
    last_col = max(1, 1 + len(months) * 2)
    qty_cols_by_month = [2 + i * 2 + 1 for i in range(len(months))]
    link_formulas = bool(monthly_sheet_title and monthly_layout)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = worksheet.cell(
        1,
        1,
        "План заказов материалов (к началу месяца минус логистика)",
    )
    _style_header_cell(title, fill=_HEADER_TITLE_FILL, font=_HEADER_TITLE_FONT, alignment=left)
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(1, col_idx),
            fill=_HEADER_TITLE_FILL,
            font=_HEADER_TITLE_FONT,
            alignment=left,
        )

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    link_note = (
        " Как пользоваться: скопируйте «Количество» отсюда → вставьте в "
        "«Ожидаемое поступление» на листе «1-производственный план (мес.)» для той же "
        "номенклатуры и месяца — количество заказа здесь пересчитается "
        "(формула MAX(0; потребность − остаток − поступление))."
        if link_formulas
        else ""
    )
    subtitle = worksheet.cell(
        2,
        1,
        "Дата заказа = 1-е число месяца − (логистика до МСК max + 2 дн. таможня + "
        "логистика МСК–Ростов max) из графика отгрузок; если номенклатуры нет в графике — 21 день. "
        "Количество = скользящий дефицит: max(0, плановая потребность − opening − поступления); "
        "заказ считается прибывшим к 1-му числу и закрывает дыру на текущий/следующие месяцы."
        + link_note,
    )
    _style_header_cell(
        subtitle, fill=_HEADER_SUBTITLE_FILL, font=_HEADER_SUBTITLE_FONT, alignment=left
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(2, col_idx),
            fill=_HEADER_SUBTITLE_FILL,
            font=_HEADER_SUBTITLE_FONT,
            alignment=left,
        )

    worksheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    for row_idx in (3, 4):
        c = worksheet.cell(row_idx, 1, "Номенклатура" if row_idx == 3 else None)
        _style_header_cell(c, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center)
        c.border = header_border

    for month_index, month in enumerate(months):
        base = 2 + month_index * 2
        worksheet.merge_cells(start_row=3, start_column=base, end_row=3, end_column=base + 1)
        for offset in range(2):
            cell = worksheet.cell(3, base + offset, month if offset == 0 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border
        for offset, label in enumerate(("Дата заказа", "Количество, шт")):
            cell = worksheet.cell(4, base + offset, label)
            _style_header_cell(
                cell, fill=_HEADER_METRIC_FILL, font=_HEADER_METRIC_FONT, alignment=center
            )
            cell.border = header_border

    qty_cols_for_cf: list[int] = []
    for row_offset, name in enumerate(names):
        excel_row = 5 + row_offset
        monthly_row = monthly_data_start_row + row_offset
        name_cell = worksheet.cell(excel_row, 1, name)
        name_cell.alignment = left
        name_cell.border = data_border
        for month_index, month in enumerate(months):
            base = 2 + month_index * 2
            cell_data = order_plan.cell(name, month)
            date_cell = worksheet.cell(
                excel_row,
                base,
                cell_data.order_date if cell_data is not None else None,
            )
            formula = None
            if link_formulas and monthly_layout is not None and monthly_sheet_title:
                formula = _order_plan_qty_formula(
                    monthly_sheet=monthly_sheet_title,
                    monthly_row=monthly_row,
                    order_row=excel_row,
                    months=months,
                    month_index=month_index,
                    monthly_layout=monthly_layout,
                    qty_cols_by_month=qty_cols_by_month,
                )
            if formula:
                qty_cell = worksheet.cell(excel_row, base + 1, formula)
            else:
                qty_val = cell_data.qty if cell_data is not None else 0
                qty_cell = worksheet.cell(excel_row, base + 1, qty_val)
            if cell_data is not None:
                date_cell.number_format = "DD.MM.YYYY"
            date_cell.alignment = center
            qty_cell.alignment = center
            date_cell.border = data_border
            qty_cell.border = data_border
            if row_offset == 0:
                qty_cols_for_cf.append(base + 1)

    if names and qty_cols_for_cf:
        last_data_row = 5 + len(names) - 1
        order_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        for col_idx in qty_cols_for_cf:
            letter = get_column_letter(col_idx)
            worksheet.conditional_formatting.add(
                f"{letter}5:{letter}{last_data_row}",
                CellIsRule(operator="greaterThan", formula=["0"], fill=order_fill),
            )

    worksheet.column_dimensions["A"].width = 56
    for col_idx in range(2, last_col + 1):
        letter = get_column_letter(col_idx)
        worksheet.column_dimensions[letter].width = 14 if (col_idx % 2 == 0) else 12
    worksheet.row_dimensions[1].height = 20
    worksheet.row_dimensions[2].height = 72
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[4].height = 30
    worksheet.freeze_panes = "B5"

    logger.info(
        "document_analysis_agent.order_plan_sheet_written",
        nomenclatures=len(names),
        months=months,
        linked_to_monthly=link_formulas,
    )


def _build_monthly_assurance_header(
    worksheet: Worksheet, months: list[str]
) -> tuple[dict[str, dict[str, Any]], int]:
    """Шапка: деталь потребности/недель (outline «+») + сводки + прогноз.

    Порядок на месяц (summaryRight): деталь Заказ/Опыт/Склад → Потребность План/Факт
    → недели поступления → итог поступления → прогноз.
    """
    weeks_by_month = _weeks_by_month_labels(months)
    cols_per_month = [
        _MONTHLY_FIXED_TAIL_COLS + len(weeks_by_month.get(month, [])) for month in months
    ]
    last_col = _FIXED_RESULT_COLS + sum(cols_per_month)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B0B0B0")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Деталь слева от сводки: отдельные «+» у потребности и у поступления
    if worksheet.sheet_properties.outlinePr is None:
        from openpyxl.worksheet.properties import Outline

        worksheet.sheet_properties.outlinePr = Outline(
            summaryRight=True, summaryBelow=True
        )
    else:
        worksheet.sheet_properties.outlinePr.summaryRight = True
        worksheet.sheet_properties.outlinePr.summaryBelow = True

    worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=max(last_col, _FIXED_RESULT_COLS)
    )
    title = worksheet.cell(
        1, 1, "Обеспеченность плана производства «Сокол» материалами"
    )
    _style_header_cell(
        title, fill=_HEADER_TITLE_FILL, font=_HEADER_TITLE_FONT, alignment=center
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(1, col_idx),
            fill=_HEADER_TITLE_FILL,
            font=_HEADER_TITLE_FONT,
            alignment=center,
        )

    worksheet.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=max(last_col, _FIXED_RESULT_COLS)
    )
    subtitle = worksheet.cell(
        2,
        1,
        "Остатки на дату анализа; потребность — сводка План/Факт («+» раскрывает "
        "Заказ / Опытные / Склад); ожидаемое поступление — редактируемый итог за месяц "
        "(можно вставить количество из листа «план заказов»; «+» — справочно по неделям из графика); "
        "прогноз и план заказов пересчитываются от этого итога",
    )
    _style_header_cell(
        subtitle, fill=_HEADER_SUBTITLE_FILL, font=_HEADER_SUBTITLE_FONT, alignment=left
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(2, col_idx),
            fill=_HEADER_SUBTITLE_FILL,
            font=_HEADER_SUBTITLE_FONT,
            alignment=left,
        )

    fixed_headers = [
        "Номенклатура",
        "В каких изделиях используется",
        "Поставщик",
        "Страна",
        "Ед. изм.",
        "Цена, руб./ед.",
        "Остаток",
        "Заказано",
    ]
    for col_idx, label in enumerate(fixed_headers, start=1):
        worksheet.merge_cells(
            start_row=3, start_column=col_idx, end_row=5, end_column=col_idx
        )
        for row_idx in (3, 4, 5):
            cell = worksheet.cell(row_idx, col_idx, label if row_idx == 3 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border

    layout: dict[str, dict[str, Any]] = {}
    cursor = _FIXED_RESULT_COLS + 1
    for month in months:
        week_spans = weeks_by_month.get(month, [])
        week_count = len(week_spans)
        month_width = _MONTHLY_FIXED_TAIL_COLS + week_count
        base = cursor
        # [деталь 6][ΣПлан][ΣФакт][недели…][поступление][прогноз]
        detail_start = base
        detail_end = base + _MONTHLY_DETAIL_COLS - 1
        sum_plan_col = detail_end + 1
        sum_fact_col = detail_end + 2
        week_start = sum_fact_col + 1
        week_end = week_start + week_count - 1 if week_count else week_start - 1
        receipt_col = sum_fact_col + 1 + week_count
        forecast_col = receipt_col + 1

        worksheet.merge_cells(
            start_row=3, start_column=base, end_row=3, end_column=base + month_width - 1
        )
        for offset in range(month_width):
            cell = worksheet.cell(3, base + offset, month if offset == 0 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border

        month_cols: dict[str, Any] = {}
        category_spans = [
            (0, "заказ", _CATEGORY_LABELS["заказ"]),
            (2, "опытные", _CATEGORY_LABELS["опытные"]),
            (4, "склад", _CATEGORY_LABELS["склад"]),
        ]
        for offset, category_key, category_label in category_spans:
            start = detail_start + offset
            worksheet.merge_cells(
                start_row=4, start_column=start, end_row=4, end_column=start + 1
            )
            for sub in range(2):
                cell = worksheet.cell(
                    4, start + sub, category_label if sub == 0 else None
                )
                _style_header_cell(
                    cell,
                    fill=_HEADER_GROUP_FILL,
                    font=_HEADER_GROUP_FONT,
                    alignment=center,
                )
                cell.border = header_border
            for metric_offset, metric in enumerate(_SCHEDULE_METRICS):
                cell = worksheet.cell(
                    5, start + metric_offset, "План" if metric == "план" else "Факт"
                )
                _style_header_cell(
                    cell,
                    fill=_HEADER_METRIC_FILL,
                    font=_HEADER_METRIC_FONT,
                    alignment=center,
                )
                cell.border = header_border
                month_cols[f"{category_key}:{metric}"] = start + metric_offset

        worksheet.merge_cells(
            start_row=4, start_column=sum_plan_col, end_row=4, end_column=sum_fact_col
        )
        for sub, col_idx in enumerate((sum_plan_col, sum_fact_col)):
            cell = worksheet.cell(4, col_idx, "Потребность" if sub == 0 else None)
            _style_header_cell(
                cell,
                fill=_HEADER_GROUP_FILL,
                font=_HEADER_GROUP_FONT,
                alignment=center,
            )
            cell.border = header_border
        for col_idx, label in ((sum_plan_col, "План"), (sum_fact_col, "Факт")):
            cell = worksheet.cell(5, col_idx, label)
            _style_header_cell(
                cell,
                fill=_HEADER_METRIC_FILL,
                font=_HEADER_METRIC_FONT,
                alignment=center,
            )
            cell.border = header_border

        week_cols: dict[str, int] = {}
        if week_count:
            worksheet.merge_cells(
                start_row=4, start_column=week_start, end_row=4, end_column=week_end
            )
            for offset, span in enumerate(week_spans):
                col_idx = week_start + offset
                cell4 = worksheet.cell(
                    4, col_idx, "Поступление по неделям" if offset == 0 else None
                )
                _style_header_cell(
                    cell4,
                    fill=_HEADER_GROUP_FILL,
                    font=_HEADER_GROUP_FONT,
                    alignment=center,
                )
                cell4.border = header_border
                cell5 = worksheet.cell(5, col_idx, span.label)
                _style_header_cell(
                    cell5,
                    fill=_HEADER_METRIC_FILL,
                    font=_HEADER_METRIC_FONT,
                    alignment=center,
                )
                cell5.border = header_border
                week_cols[span.key] = col_idx

        worksheet.merge_cells(
            start_row=4, start_column=receipt_col, end_row=5, end_column=receipt_col
        )
        worksheet.merge_cells(
            start_row=4, start_column=forecast_col, end_row=5, end_column=forecast_col
        )
        for row_idx in (4, 5):
            receipt_cell = worksheet.cell(
                row_idx,
                receipt_col,
                f"Ожидаемое поступление {month.lower()}" if row_idx == 4 else None,
            )
            _style_header_cell(
                receipt_cell,
                fill=_HEADER_METRIC_FILL,
                font=_HEADER_METRIC_FONT,
                alignment=center,
            )
            receipt_cell.border = header_border
            forecast_cell = worksheet.cell(
                row_idx,
                forecast_col,
                _next_month_forecast_label(month) if row_idx == 4 else None,
            )
            _style_header_cell(
                forecast_cell,
                fill=_HEADER_METRIC_FILL,
                font=_HEADER_METRIC_FONT,
                alignment=center,
            )
            forecast_cell.border = header_border

        month_cols["sum_plan"] = sum_plan_col
        month_cols["sum_fact"] = sum_fact_col
        month_cols["receipt"] = receipt_col
        month_cols["forecast"] = forecast_col
        month_cols["week_cols"] = week_cols
        month_cols["week_keys"] = [span.key for span in week_spans]
        month_cols["plan_заказ"] = month_cols["заказ:план"]
        month_cols["plan_опытные"] = month_cols["опытные:план"]
        month_cols["plan_склад"] = month_cols["склад:план"]
        month_cols["fact_заказ"] = month_cols["заказ:факт"]
        month_cols["fact_опытные"] = month_cols["опытные:факт"]
        month_cols["fact_склад"] = month_cols["склад:факт"]
        layout[month] = month_cols

        for col_idx in range(detail_start, detail_end + 1):
            dim = worksheet.column_dimensions[get_column_letter(col_idx)]
            dim.outline_level = 1
            dim.hidden = True
        if week_count:
            for col_idx in range(week_start, week_end + 1):
                dim = worksheet.column_dimensions[get_column_letter(col_idx)]
                dim.outline_level = 1
                dim.hidden = True

        cursor += month_width

    worksheet.column_dimensions["A"].width = 50
    worksheet.column_dimensions["B"].width = 43
    worksheet.column_dimensions["C"].width = 36
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 10
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 12
    worksheet.column_dimensions["H"].width = 12
    for col_idx in range(_FIXED_RESULT_COLS + 1, last_col + 1):
        letter = get_column_letter(col_idx)
        dim = worksheet.column_dimensions[letter]
        if dim.width is None:
            dim.width = 12

    worksheet.row_dimensions[1].height = 20
    worksheet.row_dimensions[2].height = 40
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[4].height = 22
    worksheet.row_dimensions[5].height = 18
    return layout, last_col


def _write_monthly_assurance_sheet(
    worksheet: Worksheet,
    rows: list[MergedNomenclatureRow],
) -> dict[str, dict[str, Any]]:
    """Потребность и поступление с outline-детализацией; прогноз по Σ планов.

    Возвращает layout месяц → индексы колонок (для формул плана заказов).
    """
    worksheet.title = _SHEET_MONTHLY_ASSURANCE
    months = _months_for_monthly_sheet(rows)
    layout, last_col = _build_monthly_assurance_header(worksheet, months)

    data_alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
    thin = Side(style="thin", color="B0B0B0")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    receipt_edit_fill = PatternFill(
        start_color="FFDEEBF7", end_color="FFDEEBF7", fill_type="solid"
    )
    width_a = 50.0
    width_b = 43.0
    width_c = 36.0
    width_d = 16.0
    forecast_cols: list[int] = []
    receipt_cols: list[int] = []

    for offset, row in enumerate(rows):
        excel_row = _MONTHLY_DATA_START_ROW + offset
        values: dict[int, Any] = {
            1: row.nomenclature,
            2: "; ".join(row.products),
            3: _sanitize_result_supplier(row.supplier),
            4: row.country_of_origin,
            5: row.unit,
            6: row.price,
            7: 0.0 if row.stock is None else row.stock,
            8: 0.0 if row.ordered is None else row.ordered,
        }
        for month_index, month in enumerate(months):
            cols = layout[month]
            bucket = row.monthly_demand.get(month) or _empty_month_bucket()
            for category in _SCHEDULE_CATEGORIES:
                for metric in _SCHEDULE_METRICS:
                    col_idx = cols[f"{category}:{metric}"]
                    values[col_idx] = float(bucket.get(category, {}).get(metric, 0.0))

            plan_refs = "+".join(
                f"{get_column_letter(cols[key])}{excel_row}"
                for key in ("plan_заказ", "plan_опытные", "plan_склад")
            )
            fact_refs = "+".join(
                f"{get_column_letter(cols[key])}{excel_row}"
                for key in ("fact_заказ", "fact_опытные", "fact_склад")
            )
            values[cols["sum_plan"]] = f"={plan_refs}"
            values[cols["sum_fact"]] = f"={fact_refs}"

            week_cols: dict[str, int] = cols["week_cols"]
            week_keys: list[str] = cols["week_keys"]
            weekly = row.weekly_receipts.get(month) or {}
            for week_key in week_keys:
                values[week_cols[week_key]] = float(weekly.get(week_key, 0.0))

            receipt_col = cols["receipt"]
            forecast_col = cols["forecast"]
            # Итог поступления — обычное число (редактируемое): в него можно вставить
            # qty из «план заказов». Недели остаются справочной разбивкой из графика.
            week_total = sum(float(weekly.get(key, 0.0) or 0.0) for key in week_keys)
            receipt_value = float(row.monthly_receipts.get(month, 0.0) or 0.0)
            if receipt_value == 0.0 and week_total > 0:
                receipt_value = week_total
            values[receipt_col] = receipt_value

            receipt_letter = get_column_letter(receipt_col)
            sum_plan_letter = get_column_letter(cols["sum_plan"])
            if month_index == 0:
                values[forecast_col] = (
                    f"={_STOCK_COL_LETTER}{excel_row}"
                    f"+{receipt_letter}{excel_row}-{sum_plan_letter}{excel_row}"
                )
            else:
                prev_forecast = get_column_letter(
                    layout[months[month_index - 1]]["forecast"]
                )
                values[forecast_col] = (
                    f"={prev_forecast}{excel_row}"
                    f"+{receipt_letter}{excel_row}-{sum_plan_letter}{excel_row}"
                )
            if offset == 0:
                forecast_cols.append(forecast_col)
                receipt_cols.append(receipt_col)

        for col_idx in range(1, last_col + 1):
            cell = worksheet.cell(excel_row, col_idx, values.get(col_idx))
            cell.alignment = data_alignment
            cell.border = data_border
            if col_idx in receipt_cols:
                cell.fill = receipt_edit_fill

        worksheet.row_dimensions[excel_row].height = _estimate_wrapped_row_height(
            [
                str(values[1] or ""),
                str(values[2] or ""),
                str(values[3] or ""),
                str(values[4] or ""),
            ],
            [width_a, width_b, width_c, width_d],
        )

    if rows and forecast_cols:
        last_data_row = _MONTHLY_DATA_START_ROW + len(rows) - 1
        _apply_forecast_deficit_formatting(
            worksheet,
            forecast_cols,
            _MONTHLY_DATA_START_ROW,
            last_data_row,
        )

    # Как на «план заказов»: колонка A (номенклатура) + шапка остаются при скролле.
    worksheet.freeze_panes = f"B{_MONTHLY_DATA_START_ROW}"
    logger.info(
        "document_analysis_agent.monthly_assurance_sheet_written",
        rows=len(rows),
        months=months,
        last_col=last_col,
        forecast_months=len(forecast_cols),
        outline_detail_collapsed=True,
        outline_weeks_collapsed=True,
    )
    return layout


def _style_header_cell(
    cell: Any,
    *,
    fill: PatternFill,
    font: Font,
    alignment: Alignment,
) -> None:
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment


def _build_daily_assurance_header(worksheet: Worksheet, year: int, month: int) -> list[str]:
    """Шапка дневного листа: A–G + блоки по 4 колонки на день (план/факт/поступление/прогноз)."""
    day_keys = _month_day_keys(year, month)
    if year <= 0 or month <= 0:
        today = date.today()
        year, month = today.year, today.month
        day_keys = _month_day_keys(year, month)

    month_label = _MONTH_NOMINATIVE[month - 1]
    last_col = _FIXED_RESULT_COLS + len(day_keys) * _DAILY_COLS_PER_DAY
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="B0B0B0")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = worksheet.cell(
        1, 1, f"Обеспеченность плана производства «Сокол» материалами — по дням ({month_label} {year})"
    )
    _style_header_cell(
        title_cell, fill=_HEADER_TITLE_FILL, font=_HEADER_TITLE_FONT, alignment=center
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(1, col_idx),
            fill=_HEADER_TITLE_FILL,
            font=_HEADER_TITLE_FONT,
            alignment=center,
        )

    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    subtitle = (
        f"Остатки на дату анализа; потребность план/факт — из детального графика (строки П/ф) "
        f"за {month_label.lower()} {year} (01.{month:02d}.{year}–{len(day_keys):02d}.{month:02d}.{year}); "
        f"прогноз остатка вычитает плановую потребность; поступления — из графика отгрузок"
    )
    subtitle_cell = worksheet.cell(2, 1, subtitle)
    _style_header_cell(
        subtitle_cell,
        fill=_HEADER_SUBTITLE_FILL,
        font=_HEADER_SUBTITLE_FONT,
        alignment=left,
    )
    for col_idx in range(2, last_col + 1):
        _style_header_cell(
            worksheet.cell(2, col_idx),
            fill=_HEADER_SUBTITLE_FILL,
            font=_HEADER_SUBTITLE_FONT,
            alignment=left,
        )

    fixed_headers = [
        "Номенклатура",
        "В каких изделиях используется",
        "Поставщик",
        "Страна",
        "Ед. изм.",
        "Цена, руб./ед.",
        "Остаток",
        "Заказано",
    ]
    for col_idx, title in enumerate(fixed_headers, start=1):
        worksheet.merge_cells(
            start_row=3, start_column=col_idx, end_row=4, end_column=col_idx
        )
        for row_idx in (3, 4):
            cell = worksheet.cell(row_idx, col_idx, title if row_idx == 3 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border

    for day_index, day_key in enumerate(day_keys):
        day = date.fromisoformat(day_key)
        base_col = _FIXED_RESULT_COLS + 1 + day_index * _DAILY_COLS_PER_DAY
        label = f"{day.day:02d}.{day.month:02d}"
        worksheet.merge_cells(
            start_row=3,
            start_column=base_col,
            end_row=3,
            end_column=base_col + _DAILY_COLS_PER_DAY - 1,
        )
        for offset in range(_DAILY_COLS_PER_DAY):
            cell = worksheet.cell(3, base_col + offset, label if offset == 0 else None)
            _style_header_cell(
                cell, fill=_HEADER_GROUP_FILL, font=_HEADER_GROUP_FONT, alignment=center
            )
            cell.border = header_border

        next_day = day + timedelta(days=1)
        next_label = f"{next_day.day:02d}.{next_day.month:02d}.{next_day.year}"
        sub = [
            f"Потребность план {label}",
            f"Потребность факт {label}",
            f"Ожидаемое поступление {label}",
            f"Прогнозируемый остаток на {next_label}",
        ]
        for offset, text in enumerate(sub):
            cell = worksheet.cell(4, base_col + offset, text)
            _style_header_cell(
                cell, fill=_HEADER_METRIC_FILL, font=_HEADER_METRIC_FONT, alignment=center
            )
            cell.border = header_border

    worksheet.column_dimensions["A"].width = 50
    worksheet.column_dimensions["B"].width = 43
    worksheet.column_dimensions["C"].width = 36
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 10
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 12
    worksheet.column_dimensions["H"].width = 12
    for day_index in range(len(day_keys)):
        base_col = _FIXED_RESULT_COLS + 1 + day_index * _DAILY_COLS_PER_DAY
        for offset in range(_DAILY_COLS_PER_DAY):
            worksheet.column_dimensions[get_column_letter(base_col + offset)].width = 12

    worksheet.row_dimensions[1].height = 20
    worksheet.row_dimensions[2].height = 36
    worksheet.row_dimensions[3].height = 18
    worksheet.row_dimensions[4].height = 52
    return day_keys


def _write_daily_assurance_sheet(
    worksheet: Worksheet,
    rows: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract,
) -> None:
    """Лист «обеспечение (Месяц)»: шапка + данные + формулы прогноза + CF."""
    year, month = detailed.year, detailed.month
    if year <= 0 or month <= 0:
        today = date.today()
        year, month = today.year, today.month

    worksheet.title = _daily_assurance_sheet_title(month)
    day_keys = _build_daily_assurance_header(worksheet, year, month)
    data_alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
    thin = Side(style="thin", color="B0B0B0")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    width_a = float(worksheet.column_dimensions["A"].width or 50)
    width_b = float(worksheet.column_dimensions["B"].width or 43)
    width_c = float(worksheet.column_dimensions["C"].width or 36)
    width_d = float(worksheet.column_dimensions["D"].width or 16)
    forecast_cols: list[int] = []

    for offset, row in enumerate(rows):
        excel_row = _RESULT_DATA_START_ROW + offset
        values: dict[int, Any] = {
            1: row.nomenclature,
            2: "; ".join(row.products),
            3: _sanitize_result_supplier(row.supplier),
            4: row.country_of_origin,
            5: row.unit,
            6: row.price,
            7: 0.0 if row.stock is None else row.stock,
            8: 0.0 if row.ordered is None else row.ordered,
        }
        for day_index, day_key in enumerate(day_keys):
            base_col = _FIXED_RESULT_COLS + 1 + day_index * _DAILY_COLS_PER_DAY
            demand_plan_col = base_col
            demand_fact_col = base_col + 1
            receipt_col = base_col + 2
            forecast_col = base_col + 3
            values[demand_plan_col] = float(row.daily_demand.get(day_key, 0.0))
            values[demand_fact_col] = float(row.daily_demand_fact.get(day_key, 0.0))
            values[receipt_col] = float(row.daily_receipts.get(day_key, 0.0))
            demand_plan_letter = get_column_letter(demand_plan_col)
            receipt_letter = get_column_letter(receipt_col)
            # прогноз вычитает только план (как на помесячном листе)
            if day_index == 0:
                values[forecast_col] = (
                    f"={_STOCK_COL_LETTER}{excel_row}"
                    f"+{receipt_letter}{excel_row}-{demand_plan_letter}{excel_row}"
                )
            else:
                prev_forecast_letter = get_column_letter(base_col - 1)
                values[forecast_col] = (
                    f"={prev_forecast_letter}{excel_row}"
                    f"+{receipt_letter}{excel_row}-{demand_plan_letter}{excel_row}"
                )
            if offset == 0:
                forecast_cols.append(forecast_col)

        last_col = _FIXED_RESULT_COLS + len(day_keys) * _DAILY_COLS_PER_DAY
        for col_idx in range(1, last_col + 1):
            cell = worksheet.cell(excel_row, col_idx, values.get(col_idx))
            cell.alignment = data_alignment
            cell.border = data_border

        worksheet.row_dimensions[excel_row].height = _estimate_wrapped_row_height(
            [
                str(values[1] or ""),
                str(values[2] or ""),
                str(values[3] or ""),
                str(values[4] or ""),
            ],
            [width_a, width_b, width_c, width_d],
        )

    if rows and forecast_cols:
        last_data_row = _RESULT_DATA_START_ROW + len(rows) - 1
        _apply_forecast_deficit_formatting(
            worksheet,
            forecast_cols,
            _RESULT_DATA_START_ROW,
            last_data_row,
        )

    # Как на «план заказов»: колонка A (номенклатура) + шапка остаются при скролле.
    worksheet.freeze_panes = f"B{_RESULT_DATA_START_ROW}"
    logger.info(
        "document_analysis_agent.daily_assurance_sheet_written",
        rows=len(rows),
        month=f"{year:04d}-{month:02d}",
        days=len(day_keys),
        cols_per_day=_DAILY_COLS_PER_DAY,
        forecast_cols=len(forecast_cols),
    )


def _write_daily_assurance_sheet_from_snapshot(
    worksheet: Worksheet,
    snapshot: dict[str, Any],
) -> None:
    """Дневной лист из сохранённого снимка (зафиксированные значения, без формул прогноза)."""
    year = int(snapshot.get("year") or 0)
    month = int(snapshot.get("month") or 0)
    day_keys = list(snapshot.get("day_keys") or [])
    rows_data = list(snapshot.get("rows") or [])
    if year <= 0 or month <= 0:
        today = date.today()
        year, month = today.year, today.month
    if not day_keys:
        day_keys = _month_day_keys(year, month)

    worksheet.title = _daily_assurance_sheet_title(month, year=year)
    header_day_keys = _build_daily_assurance_header(worksheet, year, month)
    if header_day_keys != day_keys:
        day_keys = header_day_keys

    data_alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
    thin = Side(style="thin", color="B0B0B0")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    width_a = float(worksheet.column_dimensions["A"].width or 50)
    width_b = float(worksheet.column_dimensions["B"].width or 43)
    width_c = float(worksheet.column_dimensions["C"].width or 36)
    width_d = float(worksheet.column_dimensions["D"].width or 16)
    forecast_cols: list[int] = []

    for offset, row_data in enumerate(rows_data):
        excel_row = _RESULT_DATA_START_ROW + offset
        products = row_data.get("products") or []
        if isinstance(products, str):
            products_text = products
        else:
            products_text = "; ".join(str(item) for item in products if item)

        values: dict[int, Any] = {
            1: row_data.get("nomenclature"),
            2: products_text,
            3: _sanitize_result_supplier(row_data.get("supplier")),
            4: row_data.get("country_of_origin"),
            5: row_data.get("unit"),
            6: row_data.get("price"),
            7: 0.0 if row_data.get("stock") is None else row_data.get("stock"),
            8: 0.0 if row_data.get("ordered") is None else row_data.get("ordered"),
        }
        daily_demand = row_data.get("daily_demand") or {}
        daily_demand_fact = row_data.get("daily_demand_fact") or {}
        daily_receipts = row_data.get("daily_receipts") or {}
        daily_forecast = row_data.get("daily_forecast") or {}

        for day_index, day_key in enumerate(day_keys):
            base_col = _FIXED_RESULT_COLS + 1 + day_index * _DAILY_COLS_PER_DAY
            demand_plan_col = base_col
            demand_fact_col = base_col + 1
            receipt_col = base_col + 2
            forecast_col = base_col + 3
            values[demand_plan_col] = float(daily_demand.get(day_key, 0.0))
            values[demand_fact_col] = float(daily_demand_fact.get(day_key, 0.0))
            values[receipt_col] = float(daily_receipts.get(day_key, 0.0))
            values[forecast_col] = float(daily_forecast.get(day_key, 0.0))
            if offset == 0:
                forecast_cols.append(forecast_col)

        last_col = _FIXED_RESULT_COLS + len(day_keys) * _DAILY_COLS_PER_DAY
        for col_idx in range(1, last_col + 1):
            cell = worksheet.cell(excel_row, col_idx, values.get(col_idx))
            cell.alignment = data_alignment
            cell.border = data_border

        worksheet.row_dimensions[excel_row].height = _estimate_wrapped_row_height(
            [
                str(values[1] or ""),
                str(values[2] or ""),
                str(values[3] or ""),
                str(values[4] or ""),
            ],
            [width_a, width_b, width_c, width_d],
        )

    if rows_data and forecast_cols:
        last_data_row = _RESULT_DATA_START_ROW + len(rows_data) - 1
        _apply_forecast_deficit_formatting(
            worksheet,
            forecast_cols,
            _RESULT_DATA_START_ROW,
            last_data_row,
        )

    worksheet.freeze_panes = f"B{_RESULT_DATA_START_ROW}"
    logger.info(
        "document_analysis_agent.daily_assurance_sheet_from_snapshot",
        rows=len(rows_data),
        month=f"{year:04d}-{month:02d}",
        days=len(day_keys),
        period=snapshot.get("period_key"),
    )


def _estimate_wrapped_row_height(texts: list[str], column_widths: list[float]) -> float:
    """Оценивает высоту строки при wrap_text, чтобы текст не вылезал визуально."""
    lines = 1
    for text, width in zip(texts, column_widths, strict=False):
        content = _clean_text(text)
        if not content:
            continue
        chars_per_line = max(int(width * 1.05), 8)
        # учитываем явные переносы и длину строки
        soft_lines = 0
        for part in content.split("\n"):
            soft_lines += max(1, (len(part) + chars_per_line - 1) // chars_per_line)
        lines = max(lines, soft_lines)
    return float(min(220, max(18, lines * 15)))


def _to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value).replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _load_nomenclature_mapping() -> list[_MappingRow]:
    if not _MAPPING_FILE.exists():
        return []
    workbook = load_workbook(_MAPPING_FILE, data_only=True)
    sheet = workbook.active
    headers = {
        col_idx: _normalize(sheet.cell(1, col_idx).value)
        for col_idx in range(1, sheet.max_column + 1)
    }
    nom_col = _pick_column(headers, ["номенклатура"]) or 1
    contract_nom_col = _pick_column(headers, ["номенклатура контракта"]) or 4
    contract_col = _pick_column(headers, ["договор"]) or 3

    rows: list[_MappingRow] = []
    seen: set[str] = set()
    for row_idx in range(2, sheet.max_row + 1):
        nomenclature = _clean_text(sheet.cell(row_idx, nom_col).value)
        if not nomenclature:
            continue
        key = _normalize(nomenclature)
        # Схлопываем дубликаты по договорам — оставляем первое вхождение A
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            _MappingRow(
                nomenclature=nomenclature,
                contract_nomenclature=_clean_text(sheet.cell(row_idx, contract_nom_col).value),
                contract=_clean_text(sheet.cell(row_idx, contract_col).value),
            )
        )
    return rows


def _load_spec_sheet_names() -> list[str]:
    if not _SPECS_FILE.exists():
        return []
    workbook = load_workbook(_SPECS_FILE, read_only=True)
    return [name for name in workbook.sheetnames if _clean_text(name)]


def _match_schedule_to_nomenclatures_locally(
    products: list[str], mapping_rows: list[_MappingRow]
) -> dict[str, tuple[str | None, float]]:
    result: dict[str, tuple[str | None, float]] = {}
    for product in products:
        heuristic = _heuristic_nomenclature(product, mapping_rows)
        if heuristic:
            result[product] = (heuristic, 0.95)
            continue
        best_name: str | None = None
        best_score = 0.0
        for row in mapping_rows:
            score_a = _product_match_score(product, row.nomenclature)
            score_d = _product_match_score(product, row.contract_nomenclature) if row.contract_nomenclature else 0.0
            score = max(score_a, score_d * 0.95)
            if score > best_score:
                best_score = score
                best_name = row.nomenclature
        result[product] = (best_name, best_score)
    return result


def _heuristic_nomenclature(product: str, mapping_rows: list[_MappingRow]) -> str | None:
    text = _match_key(product)
    by_norm = {_match_key(row.nomenclature): row.nomenclature for row in mapping_rows}
    is_sokol_r = bool(re.search(r"сокол\s*-?\s*р\b", text) or "z8" in text or "z40" in text or "мини" in text)
    is_is_family = bool("ис" in text or "ист" in text or "сокол с" in text)
    is_i_family = bool(re.search(r"\bи\b", text)) and not is_is_family and not is_sokol_r

    def find_contains(*needles: str) -> str | None:
        for key, name in by_norm.items():
            if all(needle in key for needle in needles):
                return name
        return None

    # Сначала ИС / И, потом Сокол-Р — иначе «перехватчик» ломает эвристику по букве «р»
    if is_is_family and "ночь" in text:
        return find_contains("сокол с", "ночь") or find_contains("ночь", "ascent")
    if is_is_family and "день" in text:
        return find_contains("сокол с", "день")

    if is_i_family and "ночь" in text:
        if "ascent" in text or "1.01" in text:
            return find_contains("т-1.01") or find_contains("ascent", "ночь")
        for key, name in by_norm.items():
            if (
                "ночь" in key
                and "т-1.01" not in key
                and "сокол с" not in key
                and "ист" not in key
                and "ascent" not in key
            ):
                return name
        return None
    if is_i_family and "день" in text:
        if "ascent" in text or "1.01" in text:
            return find_contains("и-1.01") or find_contains("ascent", "день")
        for key, name in by_norm.items():
            if (
                "день" in key
                and "сокол с" not in key
                and "и-1.01" not in key
                and "ascent" not in key
                and "бпла" not in key
            ):
                return name
        return None

    if is_sokol_r:
        if "z8" in text or "мини" in text:
            return find_contains("мини") or find_contains("z8")
        if "z40" in text:
            return find_contains("сокол-р") or next(
                (name for key, name in by_norm.items() if "сокол" in key and "р" in key and "мини" not in key and "бпла" in key),
                None,
            )
        return find_contains("сокол-р")

    if "катапульта" in text:
        if "2.0" in text or re.search(r"v\s*2\b", text) or re.search(r"\b2\b", text):
            return find_contains("катапульта 2.0") or find_contains("катапульта 2")
        return find_contains("катапульта 1.0") or find_contains("катапульта 1")

    if "нсу" in text:
        if "2.0" in text or re.search(r"нсу\s*2", text):
            if "ascent" in text:
                return find_contains("нсу 2.0", "ascent") or find_contains("нсу 2.0")
            return next(
                (name for key, name in by_norm.items() if "нсу 2.0" in key and "ascent" not in key),
                find_contains("нсу 2.0"),
            )
        if "ascent" in text or "исполнение" in text:
            return find_contains("исполнение 2")
        # НСУ 1 без ascent → базовая 1.0, не «Исполнение 2»
        return next(
            (
                name
                for key, name in by_norm.items()
                if "нсу 1.0" in key and "исполнение" not in key and "ascent" not in key
            ),
            find_contains("нсу 1.0"),
        )

    return None


def _match_nomenclature_to_sheet(
    schedule_product: str, nomenclature: str, sheet_names: list[str]
) -> tuple[str | None, str]:
    heuristic = _heuristic_sheet(schedule_product, nomenclature, sheet_names)
    if heuristic:
        return heuristic, "эвристика листа"

    candidates = sheet_names
    # ascent только из изделия графика — mapping часто содержит ascent-варианты шире
    prefer_ascent = "ascent" in _match_key(schedule_product) or "1.01" in _match_key(schedule_product)
    scored: list[tuple[str, float]] = []
    for sheet in candidates:
        score = max(
            _product_match_score(nomenclature, sheet),
            _product_match_score(_strip_fpv_prefix(nomenclature), sheet),
            _product_match_score(schedule_product, sheet) * 0.9,
        )
        if prefer_ascent and "ascent" in _match_key(sheet):
            score += 0.08
        if not prefer_ascent and ("ascent" in _match_key(sheet) or "исполнение" in _match_key(sheet)):
            score -= 0.12
        scored.append((sheet, score))
    scored.sort(key=lambda item: item[1], reverse=True)
    if not scored or scored[0][1] < 0.45:
        return None, "лист не найден локально"
    # неоднозначность: два близких результата
    if len(scored) > 1 and scored[0][1] - scored[1][1] < 0.08 and scored[1][1] >= 0.45:
        return None, "неоднозначный выбор листа"
    return scored[0][0], f"локальный матч листа ({scored[0][1]:.2f})"


def _heuristic_sheet(schedule_product: str, nomenclature: str, sheet_names: list[str]) -> str | None:
    sched = _match_key(schedule_product)
    nom = _match_key(nomenclature)
    text = f"{sched} {nom}".strip()
    sheets = {_match_key(name): name for name in sheet_names}
    is_sokol_r = bool(re.search(r"сокол\s*-?\s*р\b", text) or "z8" in text or "z40" in text or "мини" in sched)
    is_is_family = bool("ис" in text or "ист" in text or "сокол с" in text)
    is_i_family = bool(re.search(r"\bи\b", text)) and not is_is_family and not is_sokol_r
    # ascent/исполнение берём в первую очередь из изделия графика, не из mapping
    want_ascent = "ascent" in sched or "1.01" in sched or "исполнение" in sched

    def pick(*needles: str) -> str | None:
        for key, name in sheets.items():
            if all(needle in key for needle in needles):
                return name
        return None

    if is_sokol_r:
        if "z8" in text or "мини" in text:
            return pick("мини", "z8") or pick("мини")
        if "z40" in text:
            return pick("z-40") or pick("z40")
        return pick("сокол р")

    if is_is_family and "ночь" in text:
        return pick("сокол с", "ночь") or pick("ночь", "ascent")
    if is_is_family and "день" in text:
        return pick("ис", "день")

    if is_i_family and "ночь" in text:
        if want_ascent:
            return pick("ночь", "ascent") or pick("ночь", "1.01")
        return next(
            (
                name
                for key, name in sheets.items()
                if "ночь" in key and "ascent" not in key and "сокол с" not in key and "ис" not in key
            ),
            None,
        )
    if is_i_family and "день" in text:
        if want_ascent:
            return pick("день", "ascent") or pick("день", "1.01")
        return next(
            (
                name
                for key, name in sheets.items()
                if "день" in key and "ascent" not in key and "ис" not in key and "сокол с" not in key
            ),
            None,
        )

    if "катапульта" in text:
        return pick("катапульта")
    if "нсу" in text:
        if "2.0" in sched or "2.1" in sched or re.search(r"нсу\s*2", sched):
            if want_ascent:
                return pick("нсу-2.0", "ascent") or pick("2.0", "ascent")
            return next(
                (name for key, name in sheets.items() if "нсу" in key and "2.0" in key and "ascent" not in key),
                pick("нсу-2.0"),
            )
        if want_ascent:
            return pick("исполнение 2")
        return next(
            (
                name
                for key, name in sheets.items()
                if "нсу" in key
                and ("1.0" in key or re.search(r"\b1\b", key))
                and "2.0" not in key
                and "ascent" not in key
                and "исполнение" not in key
            ),
            None,
        )
    return None


async def _match_schedule_to_nomenclatures_with_lm(
    products: list[str], mapping_rows: list[_MappingRow]
) -> dict[str, str]:
    if not products:
        return {}
    payload = _lm_settings()
    if payload is None:
        return {}
    base_url, model = payload
    options = [
        {
            "nomenclature": row.nomenclature,
            "contract_nomenclature": row.contract_nomenclature,
        }
        for row in mapping_rows
    ]
    prompt = (
        "Ты сопоставляешь изделия из графика производства с номенклатурой из таблицы сопоставления. "
        "Для каждого schedule_product выбери ровно одну nomenclature из OPTIONS (поле nomenclature). "
        "Учитывай варианты: И/ИС/ИС-Т/СОКОЛ С, день/ночь, ascent, Z8/Z40, мини, НСУ, Катапульта. "
        "Если нет подходящего — верни nomenclature=null. "
        "Верни строго JSON: "
        '{"matches":[{"schedule_product":"...","nomenclature":"...","reason":"..."}]}'
        f"\n\nSCHEDULE_PRODUCTS={json.dumps(products, ensure_ascii=False)}"
        f"\n\nOPTIONS={json.dumps(options, ensure_ascii=False)}"
    )
    try:
        data = await _post_lm_json(base_url, model, prompt, timeout=settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS)
        matches = data.get("matches")
        if not isinstance(matches, list):
            return {}
        allowed = {_normalize(row.nomenclature): row.nomenclature for row in mapping_rows}
        result: dict[str, str] = {}
        for item in matches:
            if not isinstance(item, dict):
                continue
            product = _clean_text(item.get("schedule_product"))
            nomenclature = _clean_text(item.get("nomenclature"))
            if not product or not nomenclature:
                continue
            resolved = allowed.get(_normalize(nomenclature))
            if resolved:
                result[product] = resolved
        return result
    except Exception as exc:
        logger.warning("document_analysis_agent.schedule_nomenclature_lm_failed", error=str(exc))
        return {}


async def _match_nomenclatures_to_sheets_with_lm(
    nomenclatures: list[str],
    sheet_names: list[str],
) -> dict[str, str | None]:
    """Batch LM: номенклатура → лист спецификации."""
    unique = list(dict.fromkeys(n for n in nomenclatures if n))
    if not unique or not sheet_names:
        return {}
    payload = _lm_settings()
    if payload is None:
        return {}
    base_url, model = payload
    prompt = (
        "Для каждой NOMENCLATURE выбери один sheet_name из SHEETS или null. "
        "Имя листа может быть укороченным вариантом номенклатуры. "
        'Верни строго JSON: {"matches":[{"nomenclature":"...","sheet_name":"..."|null}]}'
        f"\n\nNOMENCLATURES={json.dumps(unique, ensure_ascii=False)}"
        f"\n\nSHEETS={json.dumps(sheet_names, ensure_ascii=False)}"
    )
    try:
        data = await _post_lm_json(
            base_url, model, prompt, timeout=settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS
        )
        matches = data.get("matches")
        if not isinstance(matches, list):
            return {}
        allowed = {_normalize(name): name for name in sheet_names}
        result: dict[str, str | None] = {}
        for item in matches:
            if not isinstance(item, dict):
                continue
            nomenclature = _clean_text(item.get("nomenclature"))
            sheet_name = _clean_text(item.get("sheet_name"))
            if not nomenclature:
                continue
            result[nomenclature] = allowed.get(_normalize(sheet_name)) if sheet_name else None
        return result
    except Exception as exc:
        logger.warning("document_analysis_agent.nomenclatures_sheet_lm_failed", error=str(exc))
        return {}


async def _match_nomenclature_to_db_spec_with_lm(
    nomenclature: str, catalog: list
) -> "DbSpecCatalogEntry | None":
    from app.agents.document_analysis_agent.onec_db_sources import DbSpecCatalogEntry

    labels = [entry.label for entry in catalog]
    sheet = await _match_nomenclature_to_sheet_with_lm(nomenclature, labels)
    if not sheet:
        return None
    for entry in catalog:
        if entry.label == sheet:
            return entry
    norm = _normalize(sheet)
    for entry in catalog:
        if _normalize(entry.label) == norm:
            return entry
    return None


async def _match_nomenclature_to_sheet_with_lm(
    nomenclature: str, sheet_names: list[str]
) -> str | None:
    payload = _lm_settings()
    if payload is None:
        return None
    base_url, model = payload
    prompt = (
        "Выбери один лист спецификации (sheet_name) для данной номенклатуры. "
        "Имя листа может быть укороченным вариантом номенклатуры. "
        "Верни строго JSON: {\"sheet_name\":\"...\",\"reason\":\"...\"} "
        "или {\"sheet_name\":null}."
        f"\n\nNOMENCLATURE={json.dumps(nomenclature, ensure_ascii=False)}"
        f"\n\nSHEETS={json.dumps(sheet_names, ensure_ascii=False)}"
    )
    try:
        data = await _post_lm_json(base_url, model, prompt, timeout=settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS)
        sheet_name = _clean_text(data.get("sheet_name"))
        if not sheet_name:
            return None
        allowed = {_normalize(name): name for name in sheet_names}
        return allowed.get(_normalize(sheet_name))
    except Exception as exc:
        logger.warning("document_analysis_agent.nomenclature_sheet_lm_failed", error=str(exc))
        return None


def _product_match_score(left: str, right: str) -> float:
    a = _match_key(left)
    b = _match_key(right)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
        return 0.72 + 0.2 * (len(shorter) / max(len(longer), 1))

    a_tokens = set(a.split())
    b_tokens = set(b.split())
    if not a_tokens or not b_tokens:
        return 0.0
    noise = {"fpv", "перехватчик", "бпла", "изделие"}
    a_core = a_tokens - noise
    b_core = b_tokens - noise
    if not a_core or not b_core:
        a_core, b_core = a_tokens, b_tokens
    overlap = len(a_core & b_core) / len(a_core | b_core)

    # бонусы за ключевые маркеры
    markers = ("день", "ночь", "ascent", "ис", "ист", "z8", "z40", "мини", "нсу", "катапульта")
    marker_bonus = 0.0
    for marker in markers:
        if marker in a and marker in b:
            marker_bonus += 0.04
        elif marker in a and marker not in b:
            marker_bonus -= 0.03
        elif marker in b and marker not in a:
            marker_bonus -= 0.02

    # И vs ИС: штраф за путаницу
    if ("ис" in a or "ист" in a or "сокол с" in a) != ("ис" in b or "ист" in b or "сокол с" in b):
        if re.search(r"\bи\b", a) or re.search(r"\bи\b", b):
            marker_bonus -= 0.15

    return max(0.0, min(1.0, overlap + marker_bonus))


def _best_text_match(value: str, candidates: list[str]) -> tuple[str | None, float]:
    best_name: str | None = None
    best_score = 0.0
    for candidate in candidates:
        score = _product_match_score(value, candidate)
        if score > best_score:
            best_score = score
            best_name = candidate
    return best_name, best_score


def _strip_fpv_prefix(value: str) -> str:
    text = _clean_text(value)
    text = re.sub(r"(?i)^fpv[-\s]*перехватчик\s*", "", text).strip()
    text = re.sub(r"(?i)^бпла\s*", "", text).strip()
    return text


def _match_key(value: str) -> str:
    text = _normalize(value)
    text = text.replace("«", " ").replace("»", " ").replace('"', " ").replace("'", " ")
    text = text.replace("ё", "е")
    # нормализация вариантов ИС-Т / ИСТ / СОКОЛ С
    text = re.sub(r"\bис\s*-\s*т\b", " ист ", text)
    text = re.sub(r"\bис\s*т\b", " ист ", text)
    text = re.sub(r"\bсокол\s*с\b", " сокол с ист ", text)
    text = re.sub(r"z\s*-\s*40", "z40", text)
    text = re.sub(r"z\s*40", "z40", text)
    text = re.sub(r"z\s*8", "z8", text)
    text = re.sub(r"[^a-zа-я0-9.]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _short(value: str, limit: int = 240) -> str:
    text = _clean_text(value)
    return text if len(text) <= limit else f"{text[:limit].rstrip()}…"


def _normalize(value: Any) -> str:
    return _clean_text(value).replace("ё", "е").lower()
