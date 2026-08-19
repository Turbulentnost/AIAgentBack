"""Конвертация legacy .xls → .xlsx bytes для openpyxl."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

# xlsx/xlsm — ZIP (PK); настоящий .xls — OLE compound (D0 CF 11 E0).
_XLSX_MAGIC = b"PK"
_XLS_OLE_MAGIC = b"\xd0\xcf\x11\xe0"


def is_legacy_xls_filename(filename: str) -> bool:
    lower = filename.lower()
    return lower.endswith(".xls") and not lower.endswith((".xlsx", ".xlsm"))


def looks_like_xlsx(content: bytes) -> bool:
    return bool(content) and content.startswith(_XLSX_MAGIC)


def looks_like_ole_xls(content: bytes) -> bool:
    return bool(content) and content.startswith(_XLS_OLE_MAGIC)


def ensure_openpyxl_bytes(filename: str, content: bytes) -> bytes:
    """Вернуть содержимое, которое умеет читать openpyxl (.xlsx/.xlsm как есть, .xls → конверсия)."""
    if looks_like_xlsx(content):
        return content
    if is_legacy_xls_filename(filename) or looks_like_ole_xls(content):
        return xls_bytes_to_xlsx_bytes(content)
    return content


def xls_bytes_to_xlsx_bytes(content: bytes) -> bytes:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover
        raise ValueError(
            "Для файлов .xls нужен пакет xlrd. Установите: pip install xlrd"
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content, formatting_info=False)
    except Exception as exc:
        raise ValueError(f"Не удалось прочитать .xls: {exc}") from exc

    out = Workbook()
    default_ws = out.active
    first = True
    for sheet_index in range(book.nsheets):
        source = book.sheet_by_index(sheet_index)
        title = (source.name or f"Лист{sheet_index + 1}")[:31]
        if first:
            ws = default_ws
            ws.title = title
            first = False
        else:
            ws = out.create_sheet(title=title)

        for row_idx in range(source.nrows):
            for col_idx in range(source.ncols):
                cell = source.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate_as_datetime(cell.value, book.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(cell.value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    value = None
                ws.cell(row_idx + 1, col_idx + 1, value)

    buffer = BytesIO()
    out.save(buffer)
    return buffer.getvalue()
