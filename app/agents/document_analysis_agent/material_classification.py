"""Классификация номенклатур для условной обеспеченности изделий."""

from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

MATERIAL_KIND_REQUIRED = "required"
MATERIAL_KIND_CONSUMABLE = "consumable"
MATERIAL_KIND_WORKSHOP = "workshop"
OPTIONAL_MATERIAL_KINDS = frozenset({MATERIAL_KIND_CONSUMABLE, MATERIAL_KIND_WORKSHOP})

MATERIAL_KIND_LABELS = {
    MATERIAL_KIND_REQUIRED: "обязательная номенклатура",
    MATERIAL_KIND_CONSUMABLE: "возможно расходник",
    MATERIAL_KIND_WORKSHOP: "возможно в цехе",
}

_WS_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class MaterialClassification:
    kind: str
    label: str
    confidence: str = ""
    reason: str = ""


@dataclass(frozen=True)
class MaterialClassificationIndex:
    by_pair: dict[tuple[str, str], MaterialClassification]
    by_material: dict[str, MaterialClassification]


def normalize_material_key(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("ё", "е")
    text = text.replace('"', "").replace("«", "").replace("»", "").replace("'", "")
    return _WS_RE.sub(" ", text)


def classify_material_text(value: str | None) -> str:
    text = normalize_material_key(value)
    if not text:
        return MATERIAL_KIND_REQUIRED
    if "в цех" in text or "в цехе" in text:
        return MATERIAL_KIND_WORKSHOP
    if "расходник" in text and not text.startswith("не расходник"):
        return MATERIAL_KIND_CONSUMABLE
    return MATERIAL_KIND_REQUIRED


def _classification_priority(item: MaterialClassification) -> int:
    if item.kind == MATERIAL_KIND_CONSUMABLE:
        return 2
    if item.kind == MATERIAL_KIND_WORKSHOP:
        return 1
    return 0


def _header_map(ws: Any) -> dict[str, int]:
    result: dict[str, int] = {}
    for cell in ws[1]:
        key = normalize_material_key(cell.value)
        if key:
            result[key] = int(cell.column)
    return result


def load_material_classification_index(path: str) -> MaterialClassificationIndex:
    file_path = Path(path)
    if not file_path.exists():
        return MaterialClassificationIndex(by_pair={}, by_material={})
    return _load_material_classification_index_cached(str(file_path), file_path.stat().st_mtime_ns)


@lru_cache(maxsize=8)
def _load_material_classification_index_cached(
    path: str,
    mtime_ns: int,
) -> MaterialClassificationIndex:
    file_path = Path(path)

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = workbook["По изделиям и номенклатурам"]
    except KeyError:
        ws = workbook.worksheets[0]

    headers = _header_map(ws)
    product_col = headers.get("изделие")
    material_col = headers.get("номенклатура")
    class_col = headers.get("классификация")
    confidence_col = headers.get("уверенность")
    reason_col = headers.get("обоснование")

    if not product_col or not material_col or not class_col:
        return MaterialClassificationIndex(by_pair={}, by_material={})

    by_pair: dict[tuple[str, str], MaterialClassification] = {}
    by_material: dict[str, MaterialClassification] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        product = row[product_col - 1] if len(row) >= product_col else None
        material = row[material_col - 1] if len(row) >= material_col else None
        raw_class = row[class_col - 1] if len(row) >= class_col else None
        product_key = normalize_material_key(product)
        material_key = normalize_material_key(material)
        if not product_key or not material_key:
            continue

        kind = classify_material_text(str(raw_class or ""))
        item = MaterialClassification(
            kind=kind,
            label=MATERIAL_KIND_LABELS.get(kind, MATERIAL_KIND_LABELS[MATERIAL_KIND_REQUIRED]),
            confidence=str(row[confidence_col - 1] or "").strip()
            if confidence_col and len(row) >= confidence_col
            else "",
            reason=str(row[reason_col - 1] or "").strip()
            if reason_col and len(row) >= reason_col
            else "",
        )
        by_pair[(product_key, material_key)] = item

        current = by_material.get(material_key)
        if current is None or _classification_priority(item) > _classification_priority(current):
            by_material[material_key] = item

    return MaterialClassificationIndex(by_pair=by_pair, by_material=by_material)


def material_classification_for(
    index: MaterialClassificationIndex,
    *,
    product: str,
    material: str,
) -> MaterialClassification:
    product_key = normalize_material_key(product)
    material_key = normalize_material_key(material)
    return (
        index.by_pair.get((product_key, material_key))
        or index.by_material.get(material_key)
        or MaterialClassification(
            kind=MATERIAL_KIND_REQUIRED,
            label=MATERIAL_KIND_LABELS[MATERIAL_KIND_REQUIRED],
        )
    )


def is_optional_material_kind(kind: str | None) -> bool:
    return (kind or MATERIAL_KIND_REQUIRED) in OPTIONAL_MATERIAL_KINDS
