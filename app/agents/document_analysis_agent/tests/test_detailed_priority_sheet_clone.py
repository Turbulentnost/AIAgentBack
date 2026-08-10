import io

import openpyxl
from openpyxl.styles import PatternFill

from app.agents.document_analysis_agent.excel_service import (
    _FILL_COVER_RED,
    _clone_worksheet_into,
)


def _has_solid_fill(cell) -> bool:
    fill = cell.fill
    return bool(fill and fill.fill_type == "solid")


def test_clone_strips_data_fills_keeps_total_row_and_column() -> None:
    source_wb = openpyxl.Workbook()
    source_ws = source_wb.active
    source_ws.title = "Отчёт"
    data_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    total_fill = PatternFill(fill_type="solid", fgColor="FFCCCCCC")

    source_ws.cell(1, 4, 1)
    source_ws.cell(1, 5, "Итог").fill = total_fill
    source_ws.cell(2, 2, "Изделие A")
    source_ws.cell(2, 4, 10).fill = data_fill
    source_ws.cell(2, 5, "=SUM(D2)").fill = total_fill
    source_ws.cell(3, 2, "Итого участок").fill = total_fill
    source_ws.cell(3, 4, 10).fill = total_fill

    target_wb = openpyxl.Workbook()
    target_ws = target_wb.active
    _clone_worksheet_into(target_ws, source_ws, strip_fills_except_totals=True)

    assert not _has_solid_fill(target_ws.cell(2, 4))
    assert _has_solid_fill(target_ws.cell(2, 5))
    assert _has_solid_fill(target_ws.cell(3, 2))
    assert _has_solid_fill(target_ws.cell(3, 4))
    assert _has_solid_fill(target_ws.cell(1, 5))

    source_wb.close()
    target_wb.close()


def test_agent_fill_applied_after_strip() -> None:
    source_wb = openpyxl.Workbook()
    source_ws = source_wb.active
    source_ws.cell(2, 4, 10).fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    target_wb = openpyxl.Workbook()
    target_ws = target_wb.active
    _clone_worksheet_into(target_ws, source_ws, strip_fills_except_totals=True)
    target_ws.cell(2, 4).fill = _FILL_COVER_RED

    assert target_ws.cell(2, 4).fill.fgColor.rgb.endswith("FFC7CE")

    source_wb.close()
    target_wb.close()
