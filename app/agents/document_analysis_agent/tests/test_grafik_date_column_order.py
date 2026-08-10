import io
from datetime import date

import openpyxl

from app.agents.document_analysis_agent.temp_schedule_merge import (
    _META_HEADERS,
    _collect_grafik_date_columns,
    _reorder_grafik_date_columns,
    build_merged_schedule_preview_values,
)


def _build_sample_workbook(*dates: date) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График"
    headers = list(_META_HEADERS) + [day.isoformat() for day in dates]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    ws.cell(2, 1, "Тестовая номенклатура")
    for idx, day in enumerate(dates, start=len(_META_HEADERS) + 1):
        ws.cell(2, idx, 100)
    raw = io.BytesIO()
    wb.save(raw)
    wb.close()
    return raw.getvalue()


def test_reorder_grafik_date_columns_sorts_chronologically() -> None:
    wb = openpyxl.load_workbook(
        io.BytesIO(
            _build_sample_workbook(
                date(2026, 8, 10),
                date(2026, 8, 5),
                date(2026, 8, 20),
            )
        )
    )
    ws = wb["График"]

    mapping = _reorder_grafik_date_columns(ws)

    ordered_dates = [day for day, _ in _collect_grafik_date_columns(ws)]
    assert ordered_dates == [date(2026, 8, 5), date(2026, 8, 10), date(2026, 8, 20)]
    assert mapping[date(2026, 8, 5)] == len(_META_HEADERS) + 1
    assert mapping[date(2026, 8, 10)] == len(_META_HEADERS) + 2
    assert mapping[date(2026, 8, 20)] == len(_META_HEADERS) + 3
    assert ws.cell(2, len(_META_HEADERS) + 1).value == 100
    wb.close()


def test_reorder_grafik_date_columns_keeps_already_sorted() -> None:
    wb = openpyxl.load_workbook(
        io.BytesIO(
            _build_sample_workbook(
                date(2026, 8, 5),
                date(2026, 8, 10),
            )
        )
    )
    ws = wb["График"]
    before = _collect_grafik_date_columns(ws)

    mapping = _reorder_grafik_date_columns(ws)

    after = _collect_grafik_date_columns(ws)
    assert after == before
    assert mapping == {day: col for day, col in before}
    wb.close()


def test_preview_reflects_sorted_headers() -> None:
    wb = openpyxl.load_workbook(
        io.BytesIO(
            _build_sample_workbook(
                date(2026, 8, 15),
                date(2026, 8, 1),
            )
        )
    )
    ws = wb["График"]
    _reorder_grafik_date_columns(ws)
    raw = io.BytesIO()
    wb.save(raw)
    wb.close()

    preview = build_merged_schedule_preview_values(raw.getvalue())
    assert preview[0][len(_META_HEADERS) : len(_META_HEADERS) + 2] == [
        "2026-08-01",
        "2026-08-15",
    ]
