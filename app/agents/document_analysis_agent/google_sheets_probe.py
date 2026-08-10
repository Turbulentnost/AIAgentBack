"""TEMP: проверка доступа к Google Sheets для агента Авион. Удалить вместе с кнопкой."""

from __future__ import annotations

import io
import json
import re
import time
from typing import Any

import requests

from app.core.config import get_settings
from app.services.google_sheets_client import (
    fetch_sheet_via_api,
    get_default_spreadsheet_target,
    get_service_account_email,
    is_configured,
)

_PROBE_TIMEOUT = (10, 60)
_USER_AGENT = "AIPlatform-AveonAgent/1.0 (temp-google-sheets-probe)"


def _spreadsheet_target() -> tuple[str, str, str]:
    settings = get_settings()
    spreadsheet_id = (settings.GOOGLE_SHEETS_SPREADSHEET_ID or "").strip() or (
        "1C3SsvZ8IcK68d-nVCwOCnb7lmIy7SA03t6_wmHoJDqA"
    )
    sheet_gid = (settings.GOOGLE_SHEETS_SHEET_GID or "").strip() or "295357731"
    spreadsheet_url = (
        f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit#gid={sheet_gid}"
    )
    return spreadsheet_id, sheet_gid, spreadsheet_url


def _preview_text(raw: bytes | str | None, limit: int = 600) -> str:
    if raw is None:
        return ""
    text = raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
    text = text.replace("\r\n", "\n").strip()
    if len(text) <= limit:
        return text
    return f"{text[:limit]}… [truncated, total {len(text)} chars]"


def _attempt(
    name: str,
    url: str,
    *,
    expect_binary: bool = False,
) -> dict[str, Any]:
    started = time.perf_counter()
    record: dict[str, Any] = {
        "name": name,
        "url": url,
        "ok": False,
        "status_code": None,
        "content_type": None,
        "content_length": None,
        "elapsed_ms": None,
        "final_url": None,
        "redirects": [],
        "error": None,
        "hint": None,
        "body_preview": None,
        "parsed": None,
    }
    try:
        response = requests.get(
            url,
            timeout=_PROBE_TIMEOUT,
            headers={"User-Agent": _USER_AGENT},
            allow_redirects=True,
        )
        elapsed_ms = round((time.perf_counter() - started) * 1000)
        record["elapsed_ms"] = elapsed_ms
        record["status_code"] = response.status_code
        record["content_type"] = response.headers.get("Content-Type")
        record["content_length"] = len(response.content)
        record["final_url"] = response.url
        if response.history:
            record["redirects"] = [
                {"status": item.status_code, "url": item.url} for item in response.history
            ]

        body_lower = (response.text or "").lower()
        if response.status_code == 401:
            record["error"] = "401 Unauthorized — таблица не публичная"
            record["hint"] = "Для закрытых таблиц используйте Service Account (см. api)"
        elif response.status_code == 403:
            record["error"] = "403 Forbidden — доступ запрещён"
        elif response.status_code == 404:
            record["error"] = "404 Not Found — неверный ID таблицы или gid листа"
        elif "sign in" in body_lower or "accounts.google.com" in body_lower:
            record["error"] = "Google перенаправил на страницу входа — таблица закрыта"
            record["hint"] = "Используйте Service Account API (см. api)"
        elif "too large" in body_lower:
            record["error"] = "Google вернул «too large» — файл слишком большой для export без API"
        elif response.status_code >= 400:
            record["error"] = f"HTTP {response.status_code}"
            record["body_preview"] = _preview_text(response.content)
        elif expect_binary:
            magic = response.content[:4]
            if magic == b"PK\x03\x04":
                record["ok"] = True
                record["parsed"] = _parse_xlsx_preview(response.content)
            else:
                record["error"] = "Ответ не похож на XLSX (нет ZIP-сигнатуры PK)"
                record["body_preview"] = _preview_text(response.content)
        elif name == "gviz_json":
            parsed = _parse_gviz_json(response.text)
            if parsed.get("ok"):
                record["ok"] = True
                record["parsed"] = parsed
            else:
                record["error"] = parsed.get("error") or "Не удалось разобрать GViz JSON"
                record["body_preview"] = _preview_text(response.text)
        else:
            text = response.text or ""
            if not text.strip():
                record["error"] = "Пустой ответ"
            elif text.lstrip().startswith("<!DOCTYPE") or text.lstrip().startswith("<html"):
                record["error"] = "HTML вместо данных — нужна авторизация"
                record["body_preview"] = _preview_text(text)
            else:
                record["ok"] = True
                record["parsed"] = _parse_csv_preview(text)
    except requests.Timeout:
        record["error"] = "Timeout — Google Sheets не ответил за отведённое время"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
    except requests.RequestException as exc:
        record["error"] = f"{type(exc).__name__}: {exc}"
        record["elapsed_ms"] = round((time.perf_counter() - started) * 1000)
    return record


def _parse_csv_preview(text: str, max_rows: int = 5, max_cols: int = 8) -> dict[str, Any]:
    lines = [line for line in text.splitlines() if line.strip()]
    preview_rows: list[list[str]] = []
    for line in lines[:max_rows]:
        cells = [cell.strip().strip('"') for cell in line.split(",")]
        preview_rows.append(cells[:max_cols])
    return {
        "format": "csv",
        "row_count_estimate": len(lines),
        "preview_rows": preview_rows,
    }


def _parse_gviz_json(text: str, max_rows: int = 5) -> dict[str, Any]:
    stripped = text.strip()
    if stripped.startswith("/*"):
        stripped = re.sub(r"^/\*O_o\*/\s*", "", stripped)
        stripped = re.sub(r"^google\.visualization\.Query\.setResponse\(", "", stripped)
        if stripped.endswith(");"):
            stripped = stripped[:-2]
    try:
        payload = json.loads(stripped)
    except json.JSONDecodeError as exc:
        return {"ok": False, "error": f"JSON parse error: {exc}"}

    table = payload.get("table") or {}
    cols = [col.get("label") or col.get("id") for col in table.get("cols") or []]
    rows = []
    for row in (table.get("rows") or [])[:max_rows]:
        cells = []
        for cell in row.get("c") or []:
            if cell is None:
                cells.append("")
            else:
                cells.append(cell.get("f") if cell.get("f") is not None else cell.get("v"))
        rows.append(cells)
    return {
        "ok": True,
        "format": "gviz_json",
        "columns": cols[:12],
        "preview_rows": rows,
        "total_rows": len(table.get("rows") or []),
    }


def _parse_xlsx_preview(content: bytes, max_rows: int = 5) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"format": "xlsx", "ok": False, "error": "openpyxl not installed"}

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = workbook.sheetnames
    active = workbook.active
    preview_rows: list[list[Any]] = []
    for idx, row in enumerate(active.iter_rows(values_only=True)):
        if idx >= max_rows:
            break
        preview_rows.append(["" if cell is None else cell for cell in row[:10]])
    workbook.close()
    return {
        "format": "xlsx",
        "sheet_names": sheets,
        "active_sheet": active.title,
        "preview_rows": preview_rows,
    }


def probe_google_sheets() -> dict[str, Any]:
    """Service Account API + анонимные export URL (диагностика)."""
    spreadsheet_id, sheet_gid, spreadsheet_url = _spreadsheet_target()
    service_account_email = get_service_account_email()
    service_account_configured = is_configured()

    api_result = fetch_sheet_via_api(spreadsheet_id, sheet_gid)

    base = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
    anonymous_attempts = [
        _attempt("csv_export", f"{base}/export?format=csv&gid={sheet_gid}"),
        _attempt("xlsx_export", f"{base}/export?format=xlsx&gid={sheet_gid}", expect_binary=True),
        _attempt("gviz_json", f"{base}/gviz/tq?tqx=out:json&gid={sheet_gid}"),
        _attempt("pub_csv", f"{base}/pub?output=csv&gid={sheet_gid}"),
    ]

    attempts = [api_result, *anonymous_attempts]
    ok = bool(api_result.get("ok")) or any(item.get("ok") for item in anonymous_attempts)

    if api_result.get("ok"):
        parsed = api_result.get("parsed") or {}
        message = (
            f"Доступ через Service Account: «{parsed.get('sheet_title')}» "
            f"({parsed.get('row_count', 0)} строк)"
        )
    elif service_account_configured:
        message = api_result.get("error") or "Service Account настроен, но доступ к таблице не получен"
    else:
        message = "Service Account не настроен — см. api и recommendations"

    return {
        "ok": ok,
        "message": message,
        "spreadsheet_id": spreadsheet_id,
        "sheet_gid": sheet_gid,
        "spreadsheet_url": spreadsheet_url,
        "service_account_configured": service_account_configured,
        "service_account_email": service_account_email,
        "api": api_result,
        "attempts": attempts,
        "recommendations": _build_recommendations(api_result, anonymous_attempts, service_account_email),
        "copy_hint": (
            "Скопируйте из консоли браузера блок [Aveon TEMP Google Sheets probe] "
            "или JSON.stringify(result, null, 2) и отправьте разработчику"
        ),
    }


def _build_recommendations(
    api_result: dict[str, Any],
    anonymous_attempts: list[dict[str, Any]],
    service_account_email: str | None,
) -> list[str]:
    tips: list[str] = []

    if not is_configured():
        tips.append(
            "1. Создайте Service Account в Google Cloud, включите Google Sheets API, "
            "скачайте JSON-ключ."
        )
        tips.append(
            "2. Задайте GOOGLE_SHEETS_SERVICE_ACCOUNT_FILE=/path/to/key.json в .env backend."
        )
        tips.append(
            "3. Задайте GOOGLE_SHEETS_SPREADSHEET_ID и GOOGLE_SHEETS_SHEET_GID."
        )
    elif not api_result.get("ok") and service_account_email:
        tips.append(
            f"Выдайте таблице доступ «Просмотр» для {service_account_email} "
            "(Google Sheets → Настройки доступа)."
        )
        if api_result.get("hint"):
            tips.append(str(api_result["hint"]))

    if api_result.get("ok"):
        tips.append("Service Account работает — можно читать данные через Sheets API v4.")
    elif all(not item.get("ok") for item in anonymous_attempts):
        tips.append(
            "Анонимный export недоступен (ожидаемо для закрытых таблиц) — используйте Service Account."
        )

    return tips
