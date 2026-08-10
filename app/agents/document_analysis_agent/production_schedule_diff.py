"""Сравнение версий помесячного графика производства (оставить 2 последние, diff было/стало)."""

from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.agents.document_analysis_agent.excel_service import (
    ROLE_PRODUCTION_SCHEDULE,
    UploadedWorkbook,
    _clean_text,
    _find_production_schedule_layout,
    _is_schedule_product_name,
    _normalize,
    _safe_set_cell_value,
    _to_float,
)
from app.core.logging import get_logger

logger = get_logger(__name__)

SCHEDULE_DIFF_FILE_NAME = "график_производства_изменения.xlsx"

_DIFF_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_DIFF_FONT = Font(color="FF9C0006", bold=True)
_HEADER_FILL = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
_MEDIUM_SIDE = Side(style="medium", color="000000")
_MEDIUM_BORDER = Border(
    left=_MEDIUM_SIDE, right=_MEDIUM_SIDE, top=_MEDIUM_SIDE, bottom=_MEDIUM_SIDE
)
_THIN_SIDE = Side(style="thin", color="000000")
_THIN_BORDER = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)
_VERSION_RE = re.compile(r"версия\s*[:\-]?\s*([0-9]+(?:[.,][0-9]+)?)", re.IGNORECASE)
_NUM_RE = re.compile(r"^([0-9]+(?:[.,][0-9]+)?)$")
_CELL_REF_RE = re.compile(r"(?<![A-Z0-9])(\$?)([A-Z]{1,3})(\$?)(\d+)", re.IGNORECASE)


def _col_letters_to_index(letters: str) -> int:
    result = 0
    for char in letters.upper():
        result = result * 26 + (ord(char) - 64)
    return result


def _index_to_col_letters(index: int) -> str:
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result


def _shift_formula_rows(formula: str, insert_at: int, amount: int = 1) -> str:
    def repl(match: re.Match[str]) -> str:
        abs_col, col, abs_row, row_s = match.groups()
        row = int(row_s)
        if row >= insert_at:
            row += amount
        return f"{abs_col}{col}{abs_row}{row}"

    return _CELL_REF_RE.sub(repl, formula)


def _shift_formula_cols(formula: str, insert_at: int, amount: int = 1) -> str:
    def repl(match: re.Match[str]) -> str:
        abs_col, col, abs_row, row_s = match.groups()
        idx = _col_letters_to_index(col)
        if idx >= insert_at:
            idx += amount
        return f"{abs_col}{_index_to_col_letters(idx)}{abs_row}{row_s}"

    return _CELL_REF_RE.sub(repl, formula)


def _retarget_formula_col(formula: str, *, from_col: int, to_col: int) -> str:
    from_letters = _index_to_col_letters(from_col)
    to_letters = _index_to_col_letters(to_col)

    def repl(match: re.Match[str]) -> str:
        abs_col, col, abs_row, row_s = match.groups()
        if col.upper() == from_letters:
            return f"{abs_col}{to_letters}{abs_row}{row_s}"
        return match.group(0)

    return _CELL_REF_RE.sub(repl, formula)


_SUM_RANGE_RE = re.compile(
    r"^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$", re.IGNORECASE
)


def _formula_row_pattern(formula: str) -> tuple[str, list[int]] | None:
    text = str(formula or "").strip()
    if not text.startswith("="):
        return None
    matched = _SUM_RANGE_RE.match(text)
    if matched:
        start_row = int(matched.group(2))
        end_row = int(matched.group(4))
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        return "sum", [start_row, end_row]
    rows = [int(match.group(4)) for match in _CELL_REF_RE.finditer(text)]
    if not rows:
        return None
    rows = list(dict.fromkeys(rows))
    return "add", rows


def _make_col_formula(style: str, rows: list[int], col: int) -> str:
    letter = get_column_letter(col)
    if style == "sum" and len(rows) >= 2:
        return f"=SUM({letter}{rows[0]}:{letter}{rows[-1]})"
    if len(rows) == 1:
        return f"={letter}{rows[0]}"
    return "=" + "+".join(f"{letter}{row}" for row in rows)


def _snapshot_cell_style(cell: Any) -> dict[str, Any] | None:
    try:
        if not getattr(cell, "has_style", False):
            return None
        return {
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        }
    except Exception:
        return None


def _apply_cell_style(cell: Any, style: dict[str, Any] | None) -> None:
    if not style:
        return
    try:
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.border = style["border"]
        cell.alignment = style["alignment"]
        cell.number_format = style["number_format"]
    except Exception:
        pass


def _fill_itogo_formulas_like_example(
    sheet: Worksheet,
    *,
    name_col: int,
    qty_col: int,
    sub_row: int,
    value_cols: list[int],
) -> None:
    """Проставить формулы ИТОГО во все колонки было/стало/факт по образцу колонки B."""
    max_row = sheet.max_row or 0
    for row_idx in range(sub_row + 1, max_row + 1):
        label = _clean_text(sheet.cell(row_idx, name_col).value)
        if not _normalize(label).startswith("итого"):
            continue
        template = sheet.cell(row_idx, qty_col).value
        if not (isinstance(template, str) and template.startswith("=")):
            for col_idx in value_cols:
                value = sheet.cell(row_idx, col_idx).value
                if isinstance(value, str) and value.startswith("="):
                    template = value
                    break
        pattern = _formula_row_pattern(str(template or ""))
        if pattern is None:
            continue
        style, rows = pattern
        target_cols = [qty_col, *value_cols]
        for col_idx in target_cols:
            _safe_set_cell_value(
                sheet, row_idx, col_idx, _make_col_formula(style, rows, col_idx)
            )


def _propagate_itogo_row_fills(
    sheet: Worksheet,
    *,
    name_col: int,
    sub_row: int,
    end_col: int,
) -> None:
    """Протянуть заливку строки ИТОГО на все колонки таблицы (в т.ч. «было»)."""
    max_row = sheet.max_row or 0
    for row_idx in range(sub_row + 1, max_row + 1):
        label = _clean_text(sheet.cell(row_idx, name_col).value)
        if not _normalize(label).startswith("итого"):
            continue
        src_cell = sheet.cell(row_idx, name_col)
        src_fill = src_cell.fill
        if src_fill is None or not src_fill.fill_type:
            for col_idx in range(1, end_col + 1):
                fill = sheet.cell(row_idx, col_idx).fill
                if fill is not None and fill.fill_type:
                    src_fill = fill
                    break
        if src_fill is None or not src_fill.fill_type:
            continue
        for col_idx in range(1, end_col + 1):
            # всегда копируем — колонки «было» иначе остаются без заливки
            sheet.cell(row_idx, col_idx).fill = copy(src_fill)


def _apply_thin_borders_below_header(
    sheet: Worksheet,
    *,
    name_col: int,
    sub_row: int,
    end_col: int,
) -> None:
    """Тонкие чёрные рамки у всех ячеек таблицы ниже шапки (месяцы / было-стало)."""
    max_row = sheet.max_row or 0
    last_row = sub_row
    for row_idx in range(sub_row + 1, max_row + 1):
        name = _clean_text(sheet.cell(row_idx, name_col).value)
        if not name:
            continue
        if _is_schedule_product_name(name) or _normalize(name).startswith("итого"):
            last_row = row_idx
    if last_row <= sub_row:
        return
    thin = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for row_idx in range(sub_row + 1, last_row + 1):
        for col_idx in range(1, end_col + 1):
            sheet.cell(row_idx, col_idx).border = thin


@dataclass(frozen=True)
class ScheduleVersionInfo:
    filename: str
    version: float
    version_label: str


@dataclass
class ScheduleDiffResult:
    has_changes: bool
    changed_months: list[str] = field(default_factory=list)
    changed_cells: int = 0
    file_name: str = SCHEDULE_DIFF_FILE_NAME
    file_bytes: bytes | None = None
    old_filename: str = ""
    new_filename: str = ""
    old_version_label: str = ""
    new_version_label: str = ""


def _parse_version_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", ".")
    match = _NUM_RE.match(text)
    if match:
        try:
            return float(match.group(1).replace(",", "."))
        except ValueError:
            return None
    embedded = _VERSION_RE.search(text)
    if embedded:
        try:
            return float(embedded.group(1).replace(",", "."))
        except ValueError:
            return None
    return None


def extract_schedule_version(content: bytes) -> tuple[float, str]:
    """Достаёт номер версии из книги (ячейка «Версия» / «Версия 1.2»)."""
    workbook = load_workbook(BytesIO(content), data_only=True)
    try:
        best: float | None = None
        for sheet in workbook.worksheets:
            max_row = min(sheet.max_row or 1, 8)
            max_col = min(sheet.max_column or 1, 50)
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    raw = sheet.cell(row, col).value
                    if raw is None:
                        continue
                    text = str(raw).strip()
                    lower = text.lower()
                    if "версия" not in lower:
                        continue
                    same = _parse_version_number(raw)
                    if same is not None:
                        best = same if best is None else max(best, same)
                        continue
                    for d_col in (1, 2, -1):
                        neighbor_col = col + d_col
                        if neighbor_col < 1 or neighbor_col > max_col:
                            continue
                        parsed = _parse_version_number(sheet.cell(row, neighbor_col).value)
                        if parsed is not None:
                            best = parsed if best is None else max(best, parsed)
        if best is None:
            return 0.0, "без версии"
        label = str(int(best)) if abs(best - round(best)) < 1e-9 else f"{best:g}"
        return best, label
    finally:
        workbook.close()


def rank_production_schedules(
    files: list[tuple[str, bytes]],
) -> list[ScheduleVersionInfo]:
    ranked: list[ScheduleVersionInfo] = []
    for filename, content in files:
        version, label = extract_schedule_version(content)
        ranked.append(
            ScheduleVersionInfo(filename=filename, version=version, version_label=label)
        )
    ranked.sort(key=lambda item: (item.version, item.filename.lower()), reverse=True)
    return ranked


def select_latest_production_schedules(
    files: list[tuple[str, bytes]],
    *,
    keep: int = 2,
) -> dict[str, Any]:
    """Оставляет keep последних версий; остальные — на удаление."""
    if not files:
        return {
            "ok": True,
            "kept": [],
            "removed": [],
            "message": "Нет файлов графика производства",
        }
    ranked = rank_production_schedules(files)
    kept = ranked[:keep]
    removed = ranked[keep:]
    if len(ranked) <= 1:
        message = "Загружен один график производства — сравнение выполнится с сохранённой версией"
    elif not removed:
        message = (
            f"Оставлена последняя версия {kept[0].version_label} ({kept[0].filename})"
            if len(kept) == 1
            else f"Сравниваются версии {kept[-1].version_label} и {kept[0].version_label}"
        )
    else:
        kept_labels = ", ".join(f"{item.version_label} ({item.filename})" for item in kept)
        removed_labels = ", ".join(f"{item.version_label} ({item.filename})" for item in removed)
        message = (
            f"Оставлена последняя версия: {kept_labels}. "
            f"Удалены более старые: {removed_labels}."
        )
    return {
        "ok": True,
        "kept": [
            {
                "filename": item.filename,
                "version": item.version,
                "version_label": item.version_label,
            }
            for item in kept
        ],
        "removed": [
            {
                "filename": item.filename,
                "version": item.version,
                "version_label": item.version_label,
            }
            for item in removed
        ],
        "message": message,
    }


def _values_equal(left: float | None, right: float | None) -> bool:
    left_n = 0.0 if left is None else float(left)
    right_n = 0.0 if right is None else float(right)
    return abs(left_n - right_n) < 1e-9


def _product_plan_map(
    sheet: Worksheet,
    layout: Any,
) -> dict[str, dict[int, float | None]]:
    """normalized product → {col: qty} только для колонок «план»."""
    plan_cols = [item.col for item in layout.columns if item.metric == "план"]
    result: dict[str, dict[int, float | None]] = {}
    for row_idx in range(layout.data_start_row, (sheet.max_row or 0) + 1):
        name = _clean_text(sheet.cell(row_idx, layout.name_col).value)
        if not _is_schedule_product_name(name):
            continue
        key = _normalize(name)
        bucket = result.setdefault(key, {})
        for col in plan_cols:
            bucket[col] = _to_float(sheet.cell(row_idx, col).value)
    return result


def _find_layout_sheet(workbook: Any) -> tuple[Worksheet, Any] | None:
    chosen = None
    legacy = None
    for sheet in workbook.worksheets:
        layout = _find_production_schedule_layout(sheet)
        if layout is None:
            continue
        if layout.is_split:
            return sheet, layout
        if legacy is None:
            legacy = (sheet, layout)
    return chosen or legacy


def build_production_schedule_diff(
    old_file: tuple[str, bytes],
    new_file: tuple[str, bytes],
) -> ScheduleDiffResult:
    """Копия новой версии: шапка как в оригинале; у каждого «План» — было/стало."""
    old_name, old_bytes = old_file
    new_name, new_bytes = new_file
    _, old_label = extract_schedule_version(old_bytes)
    _, new_label = extract_schedule_version(new_bytes)
    # Порядок аргументов задаёт смысл «было/стало» (baseline из хранилища vs новая загрузка).
    # Не переставляем файлы по номеру версии в Excel — иначе baseline и upload меняются местами.

    old_wb = load_workbook(BytesIO(old_bytes), data_only=True)
    new_wb_values = load_workbook(BytesIO(new_bytes), data_only=True)
    new_wb = load_workbook(BytesIO(new_bytes))
    try:
        old_found = _find_layout_sheet(old_wb)
        new_found_values = _find_layout_sheet(new_wb_values)
        new_found = _find_layout_sheet(new_wb)
        if old_found is None or new_found is None or new_found_values is None:
            return ScheduleDiffResult(
                has_changes=False,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        old_sheet, old_layout = old_found
        new_sheet_values, new_layout_values = new_found_values
        new_sheet, new_layout = new_found

        old_plan_by_key: dict[tuple[str, str], int] = {
            (item.month, item.category): item.col
            for item in old_layout.columns
            if item.metric == "план"
        }
        new_plan_items = [item for item in new_layout.columns if item.metric == "план"]
        if not new_plan_items:
            return ScheduleDiffResult(
                has_changes=False,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        fact_orig_by_key: dict[tuple[str, str], int] = {
            (item.month, item.category): item.col
            for item in new_layout.columns
            if item.metric == "факт"
        }

        old_products = _product_plan_map(old_sheet, old_layout)
        new_products = _product_plan_map(new_sheet_values, new_layout_values)
        new_rows: dict[str, int] = {}
        for row_idx in range(new_layout.data_start_row, (new_sheet.max_row or 0) + 1):
            name = _clean_text(new_sheet.cell(row_idx, new_layout.name_col).value)
            if _is_schedule_product_name(name):
                new_rows[_normalize(name)] = row_idx

        changed_months: set[str] = set()
        cell_diffs: list[tuple[str, str, str, float | None, float | None]] = []
        # product_key, month, category, old_val, new_val
        for item in new_plan_items:
            old_col = old_plan_by_key.get((item.month, item.category))
            for product_key in new_rows:
                new_val = (
                    new_products[product_key].get(item.col)
                    if product_key in new_products
                    else None
                )
                old_val = (
                    old_products[product_key].get(old_col)
                    if old_col is not None and product_key in old_products
                    else None
                )
                if _values_equal(old_val, new_val):
                    continue
                changed_months.add(item.month)
                cell_diffs.append(
                    (product_key, item.month, item.category, old_val, new_val)
                )

        if not cell_diffs:
            return ScheduleDiffResult(
                has_changes=False,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        metric_row = new_layout.data_start_row - 1
        category_row = metric_row - 1 if new_layout.is_split else None
        month_row = (category_row - 1) if category_row else (metric_row - 1)
        plan_style_cell = new_sheet.cell(metric_row, new_plan_items[0].col)

        def _clear_merges() -> None:
            for rng in list(new_sheet.merged_cells.ranges):
                try:
                    new_sheet.unmerge_cells(str(rng))
                except Exception:
                    try:
                        new_sheet.merged_cells.ranges.remove(rng)
                    except Exception:
                        pass

        title_value = new_sheet.cell(2, 1).value
        header_top = month_row if month_row and month_row > 0 else metric_row
        name_header_value = new_sheet.cell(header_top, new_layout.name_col).value
        qty_col = new_layout.name_col + 1
        qty_header_value = new_sheet.cell(header_top, qty_col).value
        name_header_style = _snapshot_cell_style(new_sheet.cell(header_top, new_layout.name_col))
        qty_header_style = _snapshot_cell_style(new_sheet.cell(header_top, qty_col))
        title_style = _snapshot_cell_style(new_sheet.cell(2, 1))
        month_styles: dict[str, dict[str, Any] | None] = {}
        category_styles: dict[tuple[str, str], dict[str, Any] | None] = {}
        plan_styles: dict[tuple[str, str], dict[str, Any] | None] = {}
        fact_styles: dict[tuple[str, str], dict[str, Any] | None] = {}
        for item in new_plan_items:
            if month_row and month_row > 0 and item.month not in month_styles:
                month_styles[item.month] = _snapshot_cell_style(
                    new_sheet.cell(month_row, item.col)
                )
            if category_row and category_row > 0:
                category_styles[(item.month, item.category)] = _snapshot_cell_style(
                    new_sheet.cell(category_row, item.col)
                )
            plan_styles[(item.month, item.category)] = _snapshot_cell_style(
                new_sheet.cell(metric_row, item.col)
            )
            fact_orig = fact_orig_by_key.get((item.month, item.category))
            if fact_orig is not None:
                fact_styles[(item.month, item.category)] = _snapshot_cell_style(
                    new_sheet.cell(metric_row, fact_orig)
                )

        # openpyxl не сдвигает ссылки в формулах при insert_rows/cols — сохраняем и пересчитаем
        formula_records: list[dict[str, Any]] = []
        max_scan_row = new_sheet.max_row or 0
        max_scan_col = new_sheet.max_column or 1
        for row_idx in range(1, max_scan_row + 1):
            for col_idx in range(1, max_scan_col + 1):
                value = new_sheet.cell(row_idx, col_idx).value
                if isinstance(value, str) and value.startswith("="):
                    formula_records.append(
                        {"row": row_idx, "col": col_idx, "formula": value}
                    )

        _clear_merges()

        # Строка подписей «было» / «стало» между «План» и данными
        insert_at_row = metric_row + 1
        new_sheet.insert_rows(insert_at_row)
        sub_row = insert_at_row
        new_rows = {key: row + 1 for key, row in new_rows.items()}
        for rec in formula_records:
            if rec["row"] >= insert_at_row:
                rec["row"] += 1
            rec["formula"] = _shift_formula_rows(rec["formula"], insert_at_row)
        _clear_merges()

        # Для КАЖДОГО плана: вставить колонку «было» слева (справа налево)
        plan_orig_cols = sorted({item.col for item in new_plan_items}, reverse=True)
        became_by_orig: dict[int, int] = {col: col for col in plan_orig_cols}
        for orig_col in plan_orig_cols:
            new_sheet.insert_cols(orig_col)
            for key in list(became_by_orig.keys()):
                if became_by_orig[key] >= orig_col:
                    became_by_orig[key] += 1
            for rec in formula_records:
                if rec["col"] >= orig_col:
                    rec["col"] += 1
                rec["formula"] = _shift_formula_cols(rec["formula"], orig_col)
        _clear_merges()

        def shift_orig(orig_col: int) -> int:
            # Сдвиг только от вставок на оригинальных индексах ≤ orig_col
            # (нельзя накапливать pos — иначе «догоняют» более правые вставки).
            return orig_col + sum(1 for inserted in plan_orig_cols if inserted <= orig_col)

        # Пары (было, стало[, факт]) после сдвига
        plan_pairs: list[tuple[int, int, int | None, str, str]] = []
        for item in sorted(new_plan_items, key=lambda x: x.col):
            became = became_by_orig[item.col]
            was = became - 1
            fact_orig = fact_orig_by_key.get((item.month, item.category))
            fact_col = shift_orig(fact_orig) if fact_orig is not None else None
            plan_pairs.append((was, became, fact_col, item.month, item.category))

        # Пересобрать шапку месяцев / категорий / плана
        cat_labels = {
            "заказ": "Заказ",
            "опытные": "Опытные образцы",
            "склад": "Склад",
        }
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _write(row: int, col: int, value: Any):
            return _safe_set_cell_value(new_sheet, row, col, value)

        def _merge(r1: int, c1: int, r2: int, c2: int) -> None:
            if r2 < r1 or c2 < c1:
                return
            if r1 == r2 and c1 == c2:
                return
            for rng in list(new_sheet.merged_cells.ranges):
                if not (r2 < rng.min_row or r1 > rng.max_row or c2 < rng.min_col or c1 > rng.max_col):
                    try:
                        new_sheet.unmerge_cells(str(rng))
                    except Exception:
                        try:
                            new_sheet.merged_cells.ranges.remove(rng)
                        except Exception:
                            pass
            try:
                new_sheet.merge_cells(
                    start_row=r1, start_column=c1, end_row=r2, end_column=c2
                )
            except ValueError:
                pass

        _clear_merges()

        # 1) подписи без merge + стили шапки как в исходнике/примере
        for was, became, fact_col, month, category in plan_pairs:
            left = _write(metric_row, was, "План")
            _write(metric_row, became, None)
            _apply_cell_style(left, plan_styles.get((month, category)))
            left.alignment = center
            left.border = _MEDIUM_BORDER

            was_hdr = _write(sub_row, was, "было")
            became_hdr = _write(sub_row, became, "стало")
            for hdr in (was_hdr, became_hdr):
                hdr.fill = _HEADER_FILL
                hdr.font = Font(bold=True, size=10)
                hdr.alignment = center
                hdr.border = _MEDIUM_BORDER

            if fact_col is not None:
                fact_cell = _write(metric_row, fact_col, "Факт")
                _apply_cell_style(fact_cell, fact_styles.get((month, category)))
                fact_cell.alignment = center
                fact_cell.border = _MEDIUM_BORDER
                sub_fact = _write(sub_row, fact_col, None)
                _apply_cell_style(sub_fact, fact_styles.get((month, category)))
                sub_fact.border = _MEDIUM_BORDER

            if category_row and category_row > 0:
                cat_cell = _write(
                    category_row, was, cat_labels.get(category, category)
                )
                _apply_cell_style(cat_cell, category_styles.get((month, category)))
                cat_cell.alignment = center
                cat_cell.border = _MEDIUM_BORDER

            if month_row and month_row > 0:
                m_cell = _write(month_row, was, month)
                _apply_cell_style(m_cell, month_styles.get(month))
                m_cell.alignment = center
                m_cell.border = _MEDIUM_BORDER

        # 2) объединения + заливка/рамки на весь диапазон merge
        if month_row and month_row > 0:
            by_month: dict[str, list[int]] = {}
            for was, became, fact_col, month, _category in plan_pairs:
                cols = by_month.setdefault(month, [])
                cols.extend([was, became])
                if fact_col is not None:
                    cols.append(fact_col)
            for month_name, cols in by_month.items():
                _merge(month_row, min(cols), month_row, max(cols))
                for col_idx in range(min(cols), max(cols) + 1):
                    cell = new_sheet.cell(month_row, col_idx)
                    _apply_cell_style(cell, month_styles.get(month_name))
                    cell.border = _MEDIUM_BORDER
                new_sheet.cell(month_row, min(cols)).alignment = center

        if category_row and category_row > 0:
            for was, became, fact_col, month, category in plan_pairs:
                end_col = fact_col if fact_col is not None else became
                _merge(category_row, was, category_row, end_col)
                style = category_styles.get((month, category))
                for col_idx in range(was, end_col + 1):
                    cell = new_sheet.cell(category_row, col_idx)
                    _apply_cell_style(cell, style)
                    cell.border = _MEDIUM_BORDER
                new_sheet.cell(category_row, was).alignment = center

        for was, became, fact_col, month, category in plan_pairs:
            _merge(metric_row, was, metric_row, became)
            plan_style = plan_styles.get((month, category))
            for col_idx in (was, became):
                cell = new_sheet.cell(metric_row, col_idx)
                _apply_cell_style(cell, plan_style)
                cell.border = _MEDIUM_BORDER
            _write(metric_row, was, "План")
            new_sheet.cell(metric_row, was).alignment = center
            if fact_col is not None:
                _merge(metric_row, fact_col, sub_row, fact_col)
                fact_style = fact_styles.get((month, category))
                for row_idx in (metric_row, sub_row):
                    cell = new_sheet.cell(row_idx, fact_col)
                    _apply_cell_style(cell, fact_style)
                    cell.border = _MEDIUM_BORDER
                _write(metric_row, fact_col, "Факт")
                new_sheet.cell(metric_row, fact_col).alignment = center

        name_col = new_layout.name_col
        if name_header_value is not None:
            name_cell = _write(header_top, name_col, name_header_value)
            _apply_cell_style(name_cell, name_header_style)
            name_cell.alignment = Alignment(
                horizontal="center", vertical="distributed", wrap_text=True
            )
            name_cell.border = _MEDIUM_BORDER
        _merge(header_top, name_col, sub_row, name_col)
        for row_idx in range(header_top, sub_row + 1):
            cell = new_sheet.cell(row_idx, name_col)
            _apply_cell_style(cell, name_header_style)
            cell.border = _MEDIUM_BORDER
        if qty_header_value is not None:
            qty_cell = _write(header_top, qty_col, qty_header_value)
            _apply_cell_style(qty_cell, qty_header_style)
            qty_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            qty_cell.border = _MEDIUM_BORDER
        _merge(header_top, qty_col, sub_row, qty_col)
        for row_idx in range(header_top, sub_row + 1):
            cell = new_sheet.cell(row_idx, qty_col)
            _apply_cell_style(cell, qty_header_style)
            cell.border = _MEDIUM_BORDER

        # Данные: для каждого плана — было / стало; красный при расхождении
        pair_by_key = {
            (month, category): (was, became)
            for was, became, _fact, month, category in plan_pairs
        }
        for product_key, row_idx in new_rows.items():
            for item in new_plan_items:
                was_col, became_col = pair_by_key[(item.month, item.category)]
                old_col = old_plan_by_key.get((item.month, item.category))
                new_val = (
                    new_products[product_key].get(item.col)
                    if product_key in new_products
                    else None
                )
                old_val = (
                    old_products[product_key].get(old_col)
                    if old_col is not None and product_key in old_products
                    else None
                )
                if old_val is not None:
                    _write(row_idx, was_col, old_val)
                if new_val is not None:
                    _write(row_idx, became_col, new_val)
                if not _values_equal(old_val, new_val):
                    for paint_col in (was_col, became_col):
                        cell = new_sheet.cell(row_idx, paint_col)
                        cell.fill = _DIFF_FILL
                        cell.font = _DIFF_FONT

        # Стереть сдвинутые «как есть» формулы и записать пересчитанные
        clear_max_row = new_sheet.max_row or 0
        clear_max_col = new_sheet.max_column or 1
        for row_idx in range(sub_row + 1, clear_max_row + 1):
            for col_idx in range(1, clear_max_col + 1):
                cell = new_sheet.cell(row_idx, col_idx)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    _safe_set_cell_value(new_sheet, row_idx, col_idx, None)

        became_to_was = {became: was for was, became, _f, _m, _c in plan_pairs}
        for rec in formula_records:
            _safe_set_cell_value(new_sheet, rec["row"], rec["col"], rec["formula"])
            was_col = became_to_was.get(rec["col"])
            if was_col is not None:
                _safe_set_cell_value(
                    new_sheet,
                    rec["row"],
                    was_col,
                    _retarget_formula_col(rec["formula"], from_col=rec["col"], to_col=was_col),
                )

        # Как в примере: на строках ИТОГО формулы во всех колонках плана/факта
        value_cols: list[int] = []
        for was, became, fact_col, _m, _c in plan_pairs:
            value_cols.extend([was, became])
            if fact_col is not None:
                value_cols.append(fact_col)
        _fill_itogo_formulas_like_example(
            new_sheet,
            name_col=new_layout.name_col,
            qty_col=qty_col,
            sub_row=sub_row,
            value_cols=sorted(set(value_cols)),
        )

        end_col = max(
            (fact_col or became for was, became, fact_col, _m, _c in plan_pairs),
            default=new_sheet.max_column or 1,
        )
        end_col = max(int(end_col), max(value_cols) if value_cols else 1)

        width_src = 8.0
        first_pair = pair_by_key.get(
            (new_plan_items[0].month, new_plan_items[0].category)
        )
        if first_pair:
            width_src = (
                new_sheet.column_dimensions[get_column_letter(first_pair[1])].width or 8
            )
        for was_col, became_col in pair_by_key.values():
            new_sheet.column_dimensions[get_column_letter(was_col)].width = width_src
            new_sheet.column_dimensions[get_column_letter(became_col)].width = width_src

        # Восстановить заголовок графика на всю ширину
        if title_value:
            title_cell = new_sheet.cell(2, 1, title_value)
            _apply_cell_style(title_cell, title_style)
        end_col = max(int(end_col), new_sheet.max_column or 1)
        try:
            new_sheet.merge_cells(
                start_row=2, start_column=1, end_row=2, end_column=end_col
            )
        except ValueError:
            pass

        legend_col = end_col + 2
        new_sheet.cell(
            1,
            legend_col,
            f"Сравнение планов: было v{old_label} ({old_name}) → стало v{new_label} ({new_name})",
        )
        new_sheet.cell(2, legend_col, "Красным — ячейки плана с расхождением")
        new_sheet.cell(2, legend_col).fill = _DIFF_FILL

        # В самом конце: заливка ИТОГО + тонкие рамки (иначе сбиваются записью формул/стилей)
        table_end = max(
            (fact_col or became for was, became, fact_col, _m, _c in plan_pairs),
            default=end_col,
        )
        _propagate_itogo_row_fills(
            new_sheet,
            name_col=new_layout.name_col,
            sub_row=sub_row,
            end_col=int(table_end),
        )
        _apply_thin_borders_below_header(
            new_sheet,
            name_col=new_layout.name_col,
            sub_row=sub_row,
            end_col=int(table_end),
        )

        buffer = BytesIO()
        new_wb.save(buffer)
        month_order = (
            "Январь",
            "Февраль",
            "Март",
            "Апрель",
            "Май",
            "Июнь",
            "Июль",
            "Август",
            "Сентябрь",
            "Октябрь",
            "Ноябрь",
            "Декабрь",
        )
        months_ordered = sorted(
            changed_months,
            key=lambda name: (month_order.index(name) if name in month_order else 99, name),
        )
        logger.info(
            "production_schedule_diff.built",
            old=old_name,
            new=new_name,
            cells=len(cell_diffs),
            months=months_ordered,
        )
        return ScheduleDiffResult(
            has_changes=True,
            changed_months=months_ordered,
            changed_cells=len(cell_diffs),
            file_bytes=buffer.getvalue(),
            old_filename=old_name,
            new_filename=new_name,
            old_version_label=old_label,
            new_version_label=new_label,
        )
    finally:
        old_wb.close()
        new_wb_values.close()
        new_wb.close()


def _keep_latest_schedule_workbooks(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, str],
    *,
    role: str,
    keep: int = 1,
) -> tuple[list[UploadedWorkbook], dict[str, str], UploadedWorkbook | None]:
    """Оставляет только keep последних файлов роли; возвращает выбранный workbook."""
    schedule_files = [wb for wb in workbooks if role_map.get(wb.filename) == role]
    if not schedule_files:
        return workbooks, role_map, None

    if len(schedule_files) > keep:
        selection = select_latest_production_schedules(
            [(wb.filename, wb.content) for wb in schedule_files],
            keep=keep,
        )
        removed_names = {item["filename"] for item in selection["removed"]}
        if removed_names:
            workbooks = [wb for wb in workbooks if wb.filename not in removed_names]
            role_map = {
                name: role_name
                for name, role_name in role_map.items()
                if name not in removed_names
            }
            logger.info(
                "production_schedule_diff.pruned",
                role=role,
                removed=sorted(removed_names),
            )
            schedule_files = [
                wb for wb in workbooks if role_map.get(wb.filename) == role
            ]

    ranked = rank_production_schedules([(wb.filename, wb.content) for wb in schedule_files])
    if not ranked:
        return workbooks, role_map, None
    by_name = {wb.filename: wb for wb in schedule_files}
    newest = by_name.get(ranked[0].filename)
    return workbooks, role_map, newest


def compare_production_schedule_with_snapshot(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, str],
    saved_file: tuple[str, bytes] | None,
    *,
    keep: int = 1,
) -> tuple[list[UploadedWorkbook], dict[str, str], ScheduleDiffResult | None, bool]:
    """Сравнивает новую загрузку с сохранённой версией; при отсутствии базы diff не строится."""
    workbooks, role_map, newest = _keep_latest_schedule_workbooks(
        workbooks,
        role_map,
        role=ROLE_PRODUCTION_SCHEDULE,
        keep=keep,
    )
    if newest is None:
        return workbooks, role_map, None, False
    if saved_file is None:
        return workbooks, role_map, None, False

    old_name, old_bytes = saved_file
    diff = build_production_schedule_diff(
        (old_name, old_bytes),
        (newest.filename, newest.content),
    )
    return workbooks, role_map, diff, True


def prune_workbooks_to_latest_schedules(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, str],
    *,
    keep: int = 1,
    saved_file: tuple[str, bytes] | None = None,
) -> tuple[list[UploadedWorkbook], dict[str, str], ScheduleDiffResult | None, bool]:
    """Оставляет последнюю загрузку и сравнивает её с сохранённой базовой версией."""
    return compare_production_schedule_with_snapshot(
        workbooks,
        role_map,
        saved_file,
        keep=keep,
    )
