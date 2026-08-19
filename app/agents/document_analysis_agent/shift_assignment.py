"""Сменное задание менеджеру по закупкам: конкретные задачи по рискам и закупкам."""

from __future__ import annotations

import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, timedelta
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.agents.document_analysis_agent.excel_service import (
    DetailedScheduleExtract,
    LogisticsRiskBoard,
    MergedNomenclatureRow,
    _MONTH_NOMINATIVE,
)
from app.core.logging import get_logger

logger = get_logger(__name__)

SHIFT_ASSIGNMENT_FILE_NAME = "сменное_задание_закупки.xlsx"
_RISK_LEVELS_INCLUDE = frozenset({"critical", "high"})

Priority = Literal["urgent", "today", "week"]
RowKind = Literal["header", "group", "task", "empty"]

_TITLE_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_TITLE_FONT = Font(bold=True, color="FFFFFFFF", size=14)
_HEADER_FILL = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_SUBTITLE_FILL = PatternFill(start_color="FFD9EAF7", end_color="FFD9EAF7", fill_type="solid")
_SUBTITLE_FONT = Font(color="FF1F1F1F", size=11)
_URGENT_FILL = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
_TODAY_FILL = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
_WEEK_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

_PRIORITY_LABELS: dict[Priority, str] = {
    "urgent": "Срочно",
    "today": "Сегодня",
    "week": "На этой неделе",
}
_PRIORITY_ORDER: dict[Priority, int] = {"urgent": 0, "today": 1, "week": 2}
_PRIORITY_FILL: dict[Priority, PatternFill] = {
    "urgent": _URGENT_FILL,
    "today": _TODAY_FILL,
    "week": _WEEK_FILL,
}

# Стадия логистики → что происходит простыми словами
_STAGE_SITUATION: dict[str, str] = {
    "loading_dispatch": "отгрузка у поставщика",
    "msk_arrival": "прибытие в Москву",
    "customs_clearance": "таможенное оформление",
    "rostov_arrival": "прибытие в Ростов / на склад",
}

_STAGE_TASK_TYPES: dict[str, str] = {
    "loading_dispatch": "Отгрузка",
    "msk_arrival": "Логистика МСК",
    "customs_clearance": "Таможня",
    "rostov_arrival": "Логистика Ростов",
}

_TASK_TYPE_ORDER: dict[str, int] = {
    "Отгрузка": 0,
    "Логистика МСК": 1,
    "Таможня": 2,
    "Логистика Ростов": 3,
    "Необходимые закупки": 4,
}

_STAGE_SOLUTION: dict[str, tuple[str, str]] = {
    # (срочное решение, обычное решение)
    "loading_dispatch": (
        "Позвонить поставщику, подтвердить факт отгрузки и новую дату прибытия. "
        "Если ответа нет в течение часа — сообщить руководителю закупок.",
        "Позвонить или написать поставщику: когда отгрузят и когда ждать прибытие. "
        "Зафиксировать ответ в переписке.",
    ),
    "msk_arrival": (
        "Уточнить у поставщика/перевозчика, где груз и почему задержка; "
        "согласовать новую дату прибытия в Москву и записать её.",
        "Проверить у перевозчика статус по пути в Москву; "
        "если дата сдвинулась — запросить план восстановления срока.",
    ),
    "customs_clearance": (
        "Связаться с ответственным за таможню: что блокирует выпуск и когда ждать "
        "освобождение груза. При простое — эскалировать руководителю.",
        "Проверить статус оформления и комплект документов; "
        "убедиться, что выпуск ожидается в согласованный срок.",
    ),
    "rostov_arrival": (
        "Уточнить у логистики/перевозчика статус прибытия в Ростов; "
        "согласовать со складом приоритетную приёмку партии.",
        "Проверить дату прибытия в Ростов и готовность склада к приёмке; "
        "при риске срыва — согласовать ускоренную доставку.",
    ),
}


@dataclass
class ManagerTask:
    """Одно конкретное задание для менеджера по закупкам."""

    task_type: str
    priority: Priority
    problem: str
    solution: str
    nomenclature: str
    deficit_label: str
    country: str
    supplier: str
    due_label: str
    responsible_manager: str
    manager_result: str
    sort_key: tuple[Any, ...]


@dataclass
class ShiftAssignmentPreview:
    """Данные сменного задания для UI (таблица + мета)."""

    values: list[list[str]]
    row_priorities: list[Priority | None]
    row_kinds: list[RowKind]
    as_of: str
    week_period: str
    week_in_period: bool
    task_count: int
    urgent_count: int
    today_count: int
    week_count: int


@dataclass
class ShiftAssignmentBundle:
    xlsx_bytes: bytes
    preview: ShiftAssignmentPreview


_TASKS_HEADERS = [
    "№",
    "Тип задания",
    "Приоритет",
    "Проблема",
    "Что сделать",
    "Номенклатура",
    "Дефицит",
    "Страна",
    "Поставщик",
    "Крайний срок",
    "Ответственный менеджер",
    "Результат работы менеджера",
]


def _current_week_bounds(as_of: date | None = None) -> tuple[date, date]:
    as_of = as_of or date.today()
    start = as_of - timedelta(days=as_of.weekday())
    end = start + timedelta(days=6)
    return start, end


def _week_day_keys(
    detailed: DetailedScheduleExtract | None,
    as_of: date | None = None,
) -> tuple[list[str], bool, str]:
    as_of = as_of or date.today()
    week_start, week_end = _current_week_bounds(as_of)
    period_note = (
        f"{week_start.strftime('%d.%m.%Y')} — {week_end.strftime('%d.%m.%Y')}"
    )
    if detailed is None or not detailed.day_keys:
        return [], False, period_note
    keys = [
        key
        for key in detailed.day_keys
        if week_start <= date.fromisoformat(key) <= week_end
    ]
    return keys, bool(keys), period_note


def _fmt_qty(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value))}"
    return f"{value:g}"


def _fmt_deficit_label(qty: float, unit: str | None = None) -> str:
    qty_text = _fmt_qty(qty)
    unit_text = " ".join((unit or "").strip().split())
    if not unit_text:
        return qty_text
    return f"{qty_text} {unit_text}"


def _fmt_date(iso: str) -> str:
    if not iso:
        return "—"
    try:
        return date.fromisoformat(iso[:10]).strftime("%d.%m.%Y")
    except ValueError:
        return iso


def _norm_name(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def _shipment_display_name(value: str) -> str:
    text = re.sub(r"\s+", " ", (value or "").strip())
    return re.sub(
        r"\s*\(\d+\)(?:\s+[A-Za-zА-Яа-я0-9._-]{1,20})?\s*$",
        "",
        text,
    ).strip()


def _resolve_task_nomenclature(
    name: str,
    schedule_name_index: dict[str, str] | None,
) -> str | None:
    text = (name or "").strip()
    if not text:
        return None
    if schedule_name_index:
        from app.agents.document_analysis_agent.temp_schedule_merge import resolve_schedule_name

        resolved = resolve_schedule_name(text, schedule_name_index)
        if resolved:
            return resolved
        return None
    return _shipment_display_name(text) or None


def _build_country_index(merged: list[MergedNomenclatureRow]) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in merged:
        country = (row.country_of_origin or "").strip()
        if not country:
            continue
        key = _norm_name(row.nomenclature)
        if key and key not in result:
            result[key] = country
    return result


def _country_lookup_key(value: str) -> str:
    key = _norm_name(value)
    key = key.replace("(2026)", "").replace("2026", "")
    return " ".join(key.split())


def _country_for_nomenclature(
    nomenclature: str,
    country_by_name: dict[str, str],
) -> str:
    key = _norm_name(nomenclature)
    if not key:
        return ""
    exact = country_by_name.get(key)
    if exact:
        return exact

    compact_key = _country_lookup_key(nomenclature)
    for candidate_key, country in country_by_name.items():
        if not country:
            continue
        compact_candidate = _country_lookup_key(candidate_key)
        if not compact_candidate:
            continue
        shorter, longer = (
            (compact_candidate, compact_key)
            if len(compact_candidate) <= len(compact_key)
            else (compact_key, compact_candidate)
        )
        if len(shorter) >= 8 and shorter in longer:
            return country
    return ""


def _canonical_match_key(value: str) -> str:
    key = _country_lookup_key(value)
    replacements = {
        "мотор": "двигатель",
        "motor": "двигатель",
        "engine": "двигатель",
    }
    tokens = [replacements.get(token, token) for token in key.split()]
    return " ".join(tokens)


def _model_tokens(value: str) -> set[str]:
    key = _canonical_match_key(value)
    return {
        token
        for token in re.findall(r"[a-zа-я0-9]+", key)
        if any(ch.isdigit() for ch in token) or re.search(r"[a-z]", token)
    }


def _canonical_match_score(left: str, right: str) -> float:
    left_key = _canonical_match_key(left)
    right_key = _canonical_match_key(right)
    if not left_key or not right_key:
        return 0.0
    if left_key == right_key:
        return 1.0
    if left_key in right_key or right_key in left_key:
        shorter, longer = (
            (left_key, right_key) if len(left_key) <= len(right_key) else (right_key, left_key)
        )
        return 0.74 + 0.22 * (len(shorter) / max(len(longer), 1))

    left_tokens = set(left_key.split())
    right_tokens = set(right_key.split())
    if not left_tokens or not right_tokens:
        return 0.0
    overlap = len(left_tokens & right_tokens) / len(left_tokens | right_tokens)

    left_models = _model_tokens(left)
    right_models = _model_tokens(right)
    model_overlap = left_models & right_models
    if model_overlap:
        overlap += 0.25 * (len(model_overlap) / max(len(left_models | right_models), 1))
    return max(0.0, min(1.0, overlap))


def _build_merged_index(
    merged: list[MergedNomenclatureRow],
) -> tuple[dict[str, MergedNomenclatureRow], list[MergedNomenclatureRow]]:
    by_name: dict[str, MergedNomenclatureRow] = {}
    candidates: list[MergedNomenclatureRow] = []
    for row in merged:
        key = _norm_name(row.nomenclature)
        if not key:
            continue
        by_name.setdefault(key, row)
        candidates.append(row)
    return by_name, candidates


def _resolve_canonical_row(
    nomenclature: str,
    merged_by_name: dict[str, MergedNomenclatureRow],
    merged_candidates: list[MergedNomenclatureRow],
) -> MergedNomenclatureRow | None:
    key = _norm_name(nomenclature)
    if not key:
        return None
    exact = merged_by_name.get(key)
    if exact is not None:
        return exact

    best_row: MergedNomenclatureRow | None = None
    best_score = 0.0
    for row in merged_candidates:
        score = _canonical_match_score(nomenclature, row.nomenclature)
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score >= 0.72 else None


MANAGER_RUSSIA = "Аксинин Леонид"
MANAGER_CHINA = "Тищенко Надежда"
SHIFT_MANAGER_ROSTER = (MANAGER_RUSSIA, MANAGER_CHINA)
SHIFT_MANAGER_REGIONS = {
    MANAGER_RUSSIA: "Россия",
    MANAGER_CHINA: "Китай",
}
SHIFT_MANAGER_EMAILS = {
    MANAGER_RUSSIA: "aksinin.leonid@local.dev",
    MANAGER_CHINA: "tishchenko.nadezhda@local.dev",
}


def resolve_shift_manager_name(
    *,
    email: str | None = None,
    full_name: str | None = None,
) -> str | None:
    """Определяет ФИО менеджера смены по email или полному имени пользователя."""
    normalized_email = (email or "").strip().lower()
    if normalized_email:
        for name, manager_email in SHIFT_MANAGER_EMAILS.items():
            if manager_email.lower() == normalized_email:
                return name
    normalized_name = (full_name or "").strip()
    if normalized_name in SHIFT_MANAGER_ROSTER:
        return normalized_name
    return None


def _responsible_manager(country: str) -> str:
    normalized = _norm_name(country)
    if not normalized:
        return ""
    if any(token in normalized for token in ("китай", "кнр", "china", "гонконг", "hong kong")):
        return MANAGER_CHINA
    if any(token in normalized for token in ("россия", "российская федерация", "рф", "russia")):
        return MANAGER_RUSSIA
    return ""


def _logistics_priority(days_remaining: int, risk_level: str) -> Priority:
    if risk_level == "critical" or days_remaining <= 0:
        return "urgent"
    if days_remaining <= 1:
        return "today"
    return "week"


def _build_logistics_problem(
    *,
    nomenclature: str,
    supplier: str,
    quantity: float,
    stage_key: str,
    stage_label: str,
    window_end: str,
    days_remaining: int,
) -> str:
    situation = _STAGE_SITUATION.get(stage_key, stage_label.lower() or "поставка")
    qty = _fmt_qty(quantity)
    due = _fmt_date(window_end)
    who = supplier if supplier and supplier != "не указан" else "поставщик не указан"

    if days_remaining < 0:
        overdue = abs(days_remaining)
        day_word = "день" if overdue == 1 else ("дня" if overdue < 5 else "дней")
        return (
            f"Позиция «{nomenclature}» ({qty} шт., {who}): этап «{situation}» "
            f"просрочен на {overdue} {day_word} (контрольная дата была {due}). "
            f"Без действия поставка может сорвать обеспечение производства."
        )
    if days_remaining == 0:
        return (
            f"Позиция «{nomenclature}» ({qty} шт., {who}): сегодня крайний день "
            f"по этапу «{situation}» (срок {due}). Нужно подтвердить статус прямо сейчас."
        )
    day_word = "день" if days_remaining == 1 else ("дня" if days_remaining < 5 else "дней")
    return (
        f"Позиция «{nomenclature}» ({qty} шт., {who}): до конца этапа «{situation}» "
        f"осталось {days_remaining} {day_word} (срок {due}). "
        f"Есть риск опоздания — лучше уточнить статус заранее."
    )


def _build_logistics_solution(
    *,
    stage_key: str,
    supplier: str,
    priority: Priority,
) -> str:
    urgent_text, normal_text = _STAGE_SOLUTION.get(
        stage_key,
        (
            "Связаться с поставщиком, подтвердить статус и новую дату поставки; "
            "при отсутствии ответа — сообщить руководителю закупок.",
            "Связаться с поставщиком, уточнить статус поставки и зафиксировать ответ.",
        ),
    )
    text = urgent_text if priority == "urgent" else normal_text
    if supplier and supplier != "не указан":
        return f"Контакт: {supplier}. {text}"
    return f"Поставщик в данных не указан — уточните у логистики/в 1С. {text}"


def collect_logistics_tasks(
    board: LogisticsRiskBoard | None,
    merged: list[MergedNomenclatureRow] | None = None,
    country_by_name: dict[str, str] | None = None,
    schedule_name_index: dict[str, str] | None = None,
) -> list[ManagerTask]:
    if board is None:
        return []
    merged_by_name, merged_candidates = _build_merged_index(merged or [])
    country_by_name = country_by_name or {}
    tasks: list[ManagerTask] = []
    for stage in board.stages:
        task_type = _STAGE_TASK_TYPES.get(stage.key, stage.label or "Логистика")
        stage_order = _TASK_TYPE_ORDER.get(task_type, 99)
        for item in stage.items:
            if item.risk_level not in _RISK_LEVELS_INCLUDE:
                continue
            supplier = (item.supplier or "").strip() or "не указан"
            window_end = item.window_end or item.milestone_date or ""
            days_remaining = int(item.days_remaining)
            priority = _logistics_priority(days_remaining, item.risk_level)
            due = _fmt_date(window_end)
            display_name = _resolve_task_nomenclature(item.nomenclature, schedule_name_index)
            if not display_name:
                continue
            canonical_row = _resolve_canonical_row(
                display_name,
                merged_by_name,
                merged_candidates,
            )
            country = ""
            if canonical_row is not None:
                country = (canonical_row.country_of_origin or "").strip()
            if not country:
                country = _country_for_nomenclature(display_name, country_by_name)
            if supplier == "не указан":
                if canonical_row is not None:
                    canonical_supplier = (canonical_row.supplier or "").strip()
                    if canonical_supplier:
                        supplier = canonical_supplier
            tasks.append(
                ManagerTask(
                    task_type=task_type,
                    priority=priority,
                    problem=_build_logistics_problem(
                        nomenclature=display_name,
                        supplier=supplier,
                        quantity=float(item.quantity),
                        stage_key=stage.key,
                        stage_label=stage.label,
                        window_end=window_end,
                        days_remaining=days_remaining,
                    ),
                    solution=_build_logistics_solution(
                        stage_key=stage.key,
                        supplier=supplier,
                        priority=priority,
                    ),
                    nomenclature=display_name,
                    deficit_label=_fmt_deficit_label(
                        float(item.quantity),
                        canonical_row.unit if canonical_row is not None else None,
                    ),
                    country=country,
                    supplier=supplier,
                    due_label=due,
                    responsible_manager=_responsible_manager(country),
                    manager_result="",
                    sort_key=(
                        stage_order,
                        _PRIORITY_ORDER[priority],
                        days_remaining,
                        display_name.lower(),
                    ),
                )
            )
    return tasks


def _purchase_horizon_day_keys(
    detailed: DetailedScheduleExtract | None,
    as_of: date,
) -> list[str]:
    if detailed is None or not detailed.day_keys:
        return []
    _week_start, week_end = _current_week_bounds(as_of)
    return [
        key
        for key in detailed.day_keys
        if as_of <= date.fromisoformat(key) <= week_end
    ]


def _purchase_priority_by_due(due_date: date, as_of: date) -> Priority:
    days_remaining = (due_date - as_of).days
    if days_remaining <= 0:
        return "urgent"
    if days_remaining <= 1:
        return "today"
    return "week"


def _collect_daily_purchase_task(
    row: MergedNomenclatureRow,
    *,
    day_keys: list[str],
    as_of: date,
) -> ManagerTask | None:
    nomenclature = (row.nomenclature or "").strip()
    if not nomenclature:
        return None

    daily_forecast = row.daily_forecast or {}
    deficit_days = [
        (day_key, float(daily_forecast.get(day_key, 0.0)))
        for day_key in day_keys
        if float(daily_forecast.get(day_key, 0.0)) < -1e-9
    ]
    if not deficit_days:
        return None

    first_deficit_day, first_deficit_value = deficit_days[0]
    worst_day, worst_value = min(deficit_days, key=lambda item: item[1])
    need = abs(worst_value)
    due_date = date.fromisoformat(first_deficit_day)
    due_label = due_date.strftime("%d.%m.%Y")
    supplier = (row.supplier or "").strip() or "не указан"
    country = (row.country_of_origin or "").strip()
    stock_label = _fmt_qty(float(row.stock or 0.0))

    problem = (
        f"По дневному прогнозу result позиция «{nomenclature}» уходит в дефицит "
        f"{due_label} ({_fmt_qty(first_deficit_value)}). Худший дефицит до конца "
        f"текущей недели: {_fmt_qty(worst_value)} на {_fmt_date(worst_day)}. "
        f"Текущий остаток: {stock_label}. Нужно закрыть {_fmt_qty(need)}."
    )
    if supplier == "не указан":
        solution = (
            f"Найти поставщика по позиции «{nomenclature}» и согласовать поставку "
            f"минимум {_fmt_qty(need)} до {due_label}, чтобы закрыть недельный план."
        )
    else:
        solution = (
            f"Согласовать с поставщиком «{supplier}» поставку минимум {_fmt_qty(need)} "
            f"до {due_label}; зафиксировать подтверждение и дату прихода."
        )

    return ManagerTask(
        task_type="Необходимые закупки",
        priority=_purchase_priority_by_due(due_date, as_of),
        problem=problem,
        solution=solution,
        nomenclature=nomenclature,
        deficit_label=_fmt_deficit_label(need, row.unit),
        country=country,
        supplier=supplier,
        due_label=due_label,
        responsible_manager=_responsible_manager(country),
        manager_result="",
        sort_key=(
            _TASK_TYPE_ORDER["Необходимые закупки"],
            _PRIORITY_ORDER[_purchase_priority_by_due(due_date, as_of)],
            due_date,
            -need,
            nomenclature.lower(),
        ),
    )


def collect_result_purchase_tasks(
    merged: list[MergedNomenclatureRow],
    detailed: DetailedScheduleExtract | None = None,
    as_of: date | None = None,
) -> list[ManagerTask]:
    """Необходимые закупки: дневной дефицит до конца недели; fallback — месяц."""
    as_of = as_of or date.today()
    day_keys = _purchase_horizon_day_keys(detailed, as_of)
    if day_keys:
        tasks: list[ManagerTask] = []
        for row in merged:
            task = _collect_daily_purchase_task(row, day_keys=day_keys, as_of=as_of)
            if task is not None:
                tasks.append(task)
        return tasks

    current_month_label = _MONTH_NOMINATIVE[as_of.month - 1]
    month_end = date(as_of.year, as_of.month, monthrange(as_of.year, as_of.month)[1])
    due_label = month_end.strftime("%d.%m.%Y")
    days_until_month_end = (month_end - as_of).days
    tasks: list[ManagerTask] = []

    for row in merged:
        nomenclature = (row.nomenclature or "").strip()
        if not nomenclature:
            continue

        forecast_end = float(row.monthly_forecast.get(current_month_label, 0.0))
        if forecast_end >= -1e-9:
            continue

        need = abs(forecast_end)
        if days_until_month_end <= 3:
            priority: Priority = "urgent"
        elif days_until_month_end <= 10:
            priority = "today"
        else:
            priority = "week"

        supplier = (row.supplier or "").strip() or "не указан"
        country = (row.country_of_origin or "").strip()
        stock_label = _fmt_qty(float(row.stock or 0.0))

        problem = (
            f"По итоговой таблице result прогнозируемый остаток по «{nomenclature}» "
            f"на конец {current_month_label.lower()} {as_of.year} года отрицательный "
            f"({_fmt_qty(forecast_end)}). Текущий остаток: {stock_label}. "
            f"Нужно дозаказать {_fmt_qty(need)}."
        )
        if supplier == "не указан":
            solution = (
                f"Найти поставщика по позиции «{nomenclature}» и согласовать дозаказ "
                f"{_fmt_qty(need)} с поставкой до {due_label}."
            )
        else:
            solution = (
                f"Согласовать с поставщиком «{supplier}» дозаказ {_fmt_qty(need)} "
                f"с поставкой до {due_label}; зафиксировать подтверждение."
            )

        tasks.append(
            ManagerTask(
                task_type="Необходимые закупки",
                priority=priority,
                problem=problem,
                solution=solution,
                nomenclature=nomenclature,
                deficit_label=_fmt_deficit_label(need, row.unit),
                country=country,
                supplier=supplier,
                due_label=due_label,
                responsible_manager=_responsible_manager(country),
                manager_result="",
                sort_key=(
                    _TASK_TYPE_ORDER["Необходимые закупки"],
                    _PRIORITY_ORDER[priority],
                    -need,
                    nomenclature.lower(),
                ),
            )
        )
    return tasks


def _merge_and_sort_tasks(*groups: list[ManagerTask]) -> list[ManagerTask]:
    merged = [task for group in groups for task in group]
    merged.sort(key=lambda t: t.sort_key)
    return merged


def _style_title_row(ws: Any, last_col: int, title: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.fill = _TITLE_FILL
    cell.font = _TITLE_FONT
    cell.alignment = _CENTER
    for col in range(2, last_col + 1):
        c = ws.cell(1, col)
        c.fill = _TITLE_FILL
        c.font = _TITLE_FONT


def _write_header_row(ws: Any, row_idx: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row_idx, col, text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN


def _autosize(ws: Any, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_howto_sheet(
    wb: Workbook,
    *,
    as_of: date,
    week_period: str,
    week_in_period: bool,
    tasks: list[ManagerTask],
) -> None:
    ws = wb.active
    ws.title = "Как работать"
    _style_title_row(ws, 2, "Сменное задание менеджеру по закупкам")
    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value = (
        "Это список конкретных дел на смену: что случилось, что сделать, какой "
        "дефицит закрыть и кто отвечает. Идите по разделам сверху вниз: логистика, "
        "затем необходимые закупки."
    )
    sub.fill = _SUBTITLE_FILL
    sub.font = _SUBTITLE_FONT
    sub.alignment = _LEFT
    ws["B2"].fill = _SUBTITLE_FILL

    urgent_n = sum(1 for t in tasks if t.priority == "urgent")
    today_n = sum(1 for t in tasks if t.priority == "today")
    week_n = sum(1 for t in tasks if t.priority == "week")

    facts = [
        ("Дата", as_of.strftime("%d.%m.%Y")),
        ("Неделя", week_period),
        (
            "Необходимые закупки",
            "по дневному прогнозу result до конца текущей недели",
        ),
        ("Всего заданий", str(len(tasks))),
        ("Срочно", str(urgent_n)),
        ("Сегодня", str(today_n)),
        ("На этой неделе", str(week_n)),
    ]
    _write_header_row(ws, 4, ["Что смотреть", "Значение"])
    for offset, (label, value) in enumerate(facts):
        r = 5 + offset
        ws.cell(r, 1, label).border = _THIN
        ws.cell(r, 2, value).border = _THIN
        ws.cell(r, 1).alignment = _LEFT
        ws.cell(r, 2).alignment = _LEFT

    instr_row = 5 + len(facts) + 1
    ws.merge_cells(start_row=instr_row, start_column=1, end_row=instr_row, end_column=2)
    instr = ws.cell(
        instr_row,
        1,
        "Откройте лист «Задания». В каждой строке: тип задания → проблема → что сделать "
        "→ дефицит → ответственный. После звонка/письма заполните колонку "
        "«Результат работы менеджера».",
    )
    instr.fill = _SUBTITLE_FILL
    instr.font = _SUBTITLE_FONT
    instr.alignment = _LEFT
    ws.cell(instr_row, 2).fill = _SUBTITLE_FILL
    _autosize(ws, [36, 58])
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 48


def _task_row_values(task: ManagerTask, index: int) -> list[str | int]:
    return [
        index + 1,
        task.task_type,
        _PRIORITY_LABELS[task.priority],
        task.problem,
        task.solution,
        task.nomenclature,
        task.deficit_label,
        task.country,
        task.supplier,
        task.due_label,
        task.responsible_manager,
        task.manager_result,
    ]


def _grouped_tasks(tasks: list[ManagerTask]) -> list[tuple[str, list[ManagerTask]]]:
    grouped: dict[str, list[ManagerTask]] = {}
    for task in tasks:
        grouped.setdefault(task.task_type, []).append(task)
    return sorted(
        grouped.items(),
        key=lambda item: (_TASK_TYPE_ORDER.get(item[0], 99), item[0]),
    )


def build_shift_assignment_preview(
    *,
    as_of: date,
    tasks: list[ManagerTask],
    week_in_period: bool,
    week_period: str,
) -> ShiftAssignmentPreview:
    urgent_n = sum(1 for task in tasks if task.priority == "urgent")
    today_n = sum(1 for task in tasks if task.priority == "today")
    week_n = sum(1 for task in tasks if task.priority == "week")

    values: list[list[str]] = [_TASKS_HEADERS]
    row_priorities: list[Priority | None] = [None]
    row_kinds: list[RowKind] = ["header"]

    if not tasks:
        values.append(
            [
                "На сегодня срочных заданий нет: поставки в зоне риска "
                "и нехватка на неделе не найдены."
            ]
        )
        row_priorities.append(None)
        row_kinds.append("empty")
    else:
        task_index = 0
        for task_type, group_tasks in _grouped_tasks(tasks):
            values.append([task_type])
            row_priorities.append(None)
            row_kinds.append("group")
            for task in group_tasks:
                values.append([str(value) for value in _task_row_values(task, task_index)])
                row_priorities.append(task.priority)
                row_kinds.append("task")
                task_index += 1

    return ShiftAssignmentPreview(
        values=values,
        row_priorities=row_priorities,
        row_kinds=row_kinds,
        as_of=as_of.strftime("%d.%m.%Y"),
        week_period=week_period,
        week_in_period=week_in_period,
        task_count=len(tasks),
        urgent_count=urgent_n,
        today_count=today_n,
        week_count=week_n,
    )


def _write_tasks_sheet(wb: Workbook, tasks: list[ManagerTask]) -> None:
    ws = wb.create_sheet("Задания")
    headers = list(_TASKS_HEADERS)
    _style_title_row(ws, len(headers), "Задания на смену")
    _write_header_row(ws, 2, headers)

    if not tasks:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(headers))
        cell = ws.cell(
            3,
            1,
            "На сегодня срочных заданий нет: поставки в зоне риска и нехватка на неделе не найдены.",
        )
        cell.alignment = _LEFT
        cell.fill = _SUBTITLE_FILL
    else:
        row_idx = 3
        task_index = 0
        for task_type, group_tasks in _grouped_tasks(tasks):
            ws.merge_cells(
                start_row=row_idx,
                start_column=1,
                end_row=row_idx,
                end_column=len(headers),
            )
            group_cell = ws.cell(row_idx, 1, task_type)
            group_cell.fill = _SUBTITLE_FILL
            group_cell.font = Font(bold=True, color="FF1F1F1F", size=11)
            group_cell.alignment = _LEFT
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).border = _THIN
            row_idx += 1

            for task in group_tasks:
                values = _task_row_values(task, task_index)
                fill = _PRIORITY_FILL[task.priority]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row_idx, col, value)
                    cell.border = _THIN
                    cell.fill = fill
                    cell.alignment = _CENTER if col in (1, 2, 3, 7, 10) else _LEFT
                ws.row_dimensions[row_idx].height = 72
                row_idx += 1
                task_index += 1

    _autosize(ws, [5, 16, 14, 48, 48, 30, 12, 16, 20, 14, 22, 28])
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    extra_group_rows = len(_grouped_tasks(tasks)) if tasks else 0
    last_row = max(2, 2 + len(tasks) + extra_group_rows)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{last_row}"


def write_shift_assignment_xlsx(
    *,
    as_of: date,
    tasks: list[ManagerTask],
    week_in_period: bool,
    week_period: str,
) -> bytes:
    wb = Workbook()
    _write_howto_sheet(
        wb,
        as_of=as_of,
        week_period=week_period,
        week_in_period=week_in_period,
        tasks=tasks,
    )
    _write_tasks_sheet(wb, tasks)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _collect_shift_assignment_tasks(
    merged: list[MergedNomenclatureRow],
    logistics_risks: LogisticsRiskBoard | None,
    detailed: DetailedScheduleExtract | None,
    as_of: date | None = None,
    schedule_name_index: dict[str, str] | None = None,
) -> tuple[date, list[ManagerTask], bool, str]:
    as_of = as_of or date.today()
    country_by_name = _build_country_index(merged)
    logistics_tasks = collect_logistics_tasks(
        logistics_risks, merged, country_by_name, schedule_name_index
    )
    purchase_tasks = collect_result_purchase_tasks(merged, detailed, as_of)
    _, week_in_period, week_period = _week_day_keys(detailed, as_of)
    tasks = _merge_and_sort_tasks(logistics_tasks, purchase_tasks)
    return as_of, tasks, week_in_period, week_period


async def build_shift_assignment_bundle(
    merged: list[MergedNomenclatureRow],
    logistics_risks: LogisticsRiskBoard | None,
    detailed: DetailedScheduleExtract | None,
    as_of: date | None = None,
    schedule_name_index: dict[str, str] | None = None,
) -> ShiftAssignmentBundle:
    """Собирает xlsx и превью сменного задания для менеджера по закупкам."""
    as_of, tasks, week_in_period, week_period = _collect_shift_assignment_tasks(
        merged, logistics_risks, detailed, as_of, schedule_name_index
    )
    data = write_shift_assignment_xlsx(
        as_of=as_of,
        tasks=tasks,
        week_in_period=week_in_period,
        week_period=week_period,
    )
    preview = build_shift_assignment_preview(
        as_of=as_of,
        tasks=tasks,
        week_in_period=week_in_period,
        week_period=week_period,
    )
    logger.info(
        "document_analysis_agent.shift_assignment_built",
        bytes=len(data),
        tasks=len(tasks),
        week_in_period=week_in_period,
        week_period=week_period,
    )
    return ShiftAssignmentBundle(xlsx_bytes=data, preview=preview)


async def build_shift_assignment_xlsx(
    merged: list[MergedNomenclatureRow],
    logistics_risks: LogisticsRiskBoard | None,
    detailed: DetailedScheduleExtract | None,
    as_of: date | None = None,
    schedule_name_index: dict[str, str] | None = None,
) -> bytes:
    bundle = await build_shift_assignment_bundle(
        merged, logistics_risks, detailed, as_of, schedule_name_index
    )
    return bundle.xlsx_bytes


ResultEvaluationStatus = Literal["resolved", "partial", "not_resolved"]

_STATUS_ALIASES: dict[str, ResultEvaluationStatus] = {
    "resolved": "resolved",
    "partial": "partial",
    "not_resolved": "not_resolved",
    "not resolved": "not_resolved",
    "not-resolved": "not_resolved",
    "done": "resolved",
    "complete": "resolved",
    "completed": "resolved",
    "решено": "resolved",
    "выполнено": "resolved",
    "частично": "partial",
    "частичное": "partial",
    "in_progress": "partial",
    "progress": "partial",
    "не решено": "not_resolved",
    "нерешено": "not_resolved",
    "failed": "not_resolved",
    "no": "not_resolved",
}


def _normalize_result_status(raw: Any) -> ResultEvaluationStatus:
    text = str(raw or "").strip().lower().replace("_", " ")
    if text in _STATUS_ALIASES:
        return _STATUS_ALIASES[text]
    compact = text.replace(" ", "")
    if compact in _STATUS_ALIASES:
        return _STATUS_ALIASES[compact]
    if "частич" in text:
        return "partial"
    if any(token in text for token in ("не реш", "нереш", "not resolved", "not_resolved")):
        return "not_resolved"
    if any(token in text for token in ("реш", "выполн", "resolved", "done")):
        return "resolved"
    return "not_resolved"


def _count_dates_in_text(text: str) -> int:
    lowered = (text or "").lower()
    patterns = (
        r"\b\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?\b",
        r"\b\d{1,2}\s+(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)\b",
    )
    found: set[str] = set()
    for pattern in patterns:
        for match in re.finditer(pattern, lowered):
            found.add(match.group(0))
    return len(found)


def _has_delivery_quantities(text: str) -> bool:
    lowered = (text or "").lower()
    if re.search(r"\d[\d\s.,]*\s*(?:шт|штук|pcs|кол-?во)", lowered):
        return True
    return bool(re.search(r"\b\d{3,}\b", lowered))


def _mentions_split_delivery(text: str) -> bool:
    lowered = (text or "").lower()
    return any(token in lowered for token in ("парт", "раздел", "двумя", "тремя", "остат", "авари"))


def _heuristic_result_status(manager_result: str, solution: str) -> ResultEvaluationStatus:
    text = manager_result.strip().lower()
    solution_text = (solution or "").lower()
    if len(text) < 8:
        return "not_resolved"

    has_confirmation = any(
        token in text
        for token in (
            "подтверж",
            "подтверд",
            "согласован",
            "согласовали",
            "получен ответ",
            "ответ получ",
        )
    )
    has_date = _count_dates_in_text(text) >= 1
    has_action = any(token in text for token in ("связ", "позвон", "напис", "уточн", "получ"))

    if (
        _count_dates_in_text(text) >= 2
        and has_action
        and (_has_delivery_quantities(text) or _mentions_split_delivery(text))
    ):
        return "resolved"

    if has_action and has_date and (_has_delivery_quantities(text) or _mentions_split_delivery(text)):
        return "resolved"

    # Частый кейс логистики: задание просит подтвердить отгрузку и новую дату прибытия.
    # Формулировка менеджера в первом лице ("подтверждаю") тоже считается закрытием.
    if (
        "факт отгруз" in solution_text
        and "дат" in solution_text
        and "прибыт" in solution_text
        and "отгруз" in text
        and "прибыт" in text
        and has_confirmation
        and has_date
    ):
        return "resolved"

    action_tokens = (
        "позвон",
        "соглас",
        "подтверд",
        "подтверж",
        "заказ",
        "отправ",
        "отгруз",
        "уточн",
        "связ",
        "зафикс",
        "получ",
        "ответ",
        "дат",
        "прибыт",
        "поставк",
    )
    hits = sum(1 for token in action_tokens if token in text)
    solution_hits = 0
    for token in action_tokens:
        if token in solution_text and token in text:
            solution_hits += 1

    if has_confirmation and has_date and solution_hits >= 2:
        return "resolved"
    if hits >= 2 and solution_hits >= 1:
        return "resolved"
    if hits >= 1:
        return "partial"
    return "not_resolved"


def _stronger_result_status(
    left: ResultEvaluationStatus,
    right: ResultEvaluationStatus,
) -> ResultEvaluationStatus:
    rank = {"not_resolved": 0, "partial": 1, "resolved": 2}
    return left if rank[left] >= rank[right] else right


async def evaluate_manager_result(
    *,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    manager_result: str,
) -> tuple[ResultEvaluationStatus, str, str]:
    """Оценивает ответ менеджера через LM Studio: resolved / partial / not_resolved."""
    from app.agents.document_analysis_agent.excel_service import _lm_settings, _post_lm_json
    from app.core.config import settings

    manager_result = (manager_result or "").strip()
    if not manager_result:
        return "not_resolved", "Ответ менеджера пустой.", "heuristic"

    heuristic_status = _heuristic_result_status(manager_result, solution)
    prompt = (
        "Оцени, насколько ответ менеджера по закупкам соответствует заданию.\n\n"
        f"Тип задания: {task_type or '—'}\n"
        f"Номенклатура: {nomenclature or '—'}\n"
        f"Проблема: {problem or '—'}\n"
        f"Что сделать: {solution or '—'}\n"
        f"Ответ менеджера: {manager_result}\n\n"
        "Верни JSON:\n"
        '{"status":"resolved|partial|not_resolved","comment":"кратко по-русски, до 120 символов"}\n\n'
        "Правила:\n"
        "1. resolved — ответ закрывает ключевые требования задания и содержит конкретный результат: "
        "подтверждение, дату, заказ, статус поставщика или принятое решение.\n"
        "2. partial — менеджер что-то сделал, но не закрыл одно из ключевых требований "
        "(например, подтвердил контакт, но не указал дату).\n"
        "3. not_resolved — ответ пустой по смыслу, не по теме, только обещание сделать позже "
        "или явно сообщает, что задача не выполнена.\n"
        "4. Не требуй дословного совпадения с заданием. Считай формулировки в первом лице "
        "(например, «подтверждаю», «согласовали», «получили дату») валидным результатом.\n"
        "5. Если задание: подтвердить факт отгрузки и новую дату прибытия, а ответ: "
        "«факт отгрузки подтверждаю, новая дата прибытия 14 августа», это resolved.\n"
        "6. Если менеджер сообщил перенос или разделение поставки на несколько партий "
        "с конкретными датами и количествами (или «остаток/оставшаяся партия» на вторую дату) — "
        "это resolved, даже если формулировка отличается от задания.\n"
        "7. Для логистики: «связался с поставщиком» + новые даты + количества/партии = resolved.\n"
        "8. partial ставь только если нет ни одной конкретной новой даты/количества/статуса."
    )

    lm = _lm_settings()
    if lm is not None:
        base_url, model = lm
        try:
            data = await _post_lm_json(
                base_url,
                model,
                prompt,
                timeout=min(90.0, float(settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS)),
            )
            lm_status = _normalize_result_status(data.get("status"))
            status = _stronger_result_status(heuristic_status, lm_status)
            comment = str(data.get("comment") or "").strip()[:160]
            if status != lm_status and heuristic_status == "resolved":
                comment = "Ответ закрывает ключевые требования: есть подтверждение и конкретная дата."
            logger.info(
                "document_analysis_agent.shift_result_evaluated",
                status=status,
                lm_status=lm_status,
                heuristic_status=heuristic_status,
                source="lm_studio",
                nomenclature=nomenclature[:80],
            )
            return status, comment, "lm_studio"
        except Exception as exc:
            logger.warning(
                "document_analysis_agent.shift_result_eval_failed",
                error=str(exc),
                nomenclature=nomenclature[:80],
            )

    status = heuristic_status
    comment = {
        "resolved": "Ответ содержит конкретные действия и результат.",
        "partial": "Есть действия, но результат неполный.",
        "not_resolved": "Ответ слишком общий или не по заданию.",
    }[status]
    return status, comment, "heuristic"


def _merge_draft_continuation(draft: str, continuation: str) -> str:
    """Склеивает черновик менеджера и continuation от LLM."""
    draft = draft or ""
    continuation_raw = continuation or ""
    continuation = continuation_raw.strip()
    if not draft:
        return continuation[:500]
    if not continuation:
        return draft[:500]
    if continuation.startswith(draft):
        return continuation[:500]
    if draft.rstrip() and continuation.lower().startswith(draft.strip().lower()):
        return continuation[:500]

    leading_space = continuation_raw[:1].isspace()
    draft_end = draft[-1]
    continuation_start = continuation[0]
    first_token = continuation.split()[0] if continuation.split() else continuation

    if draft_end.isalnum() and continuation_start.isalnum() and not leading_space:
        completes_word = len(first_token) <= 5 and any(ch in first_token for ch in ".,;:")
        merged = f"{draft}{continuation}" if completes_word else f"{draft} {continuation}"
    elif draft_end.isspace() or leading_space or continuation_start in ".,;:!?—-)":
        merged = f"{draft} {continuation}" if leading_space and not draft_end.isspace() else f"{draft}{continuation}"
    else:
        merged = f"{draft} {continuation}"
    return merged[:500]


def _strip_prefix_overlap(draft: str, continuation: str) -> str:
    """Убирает повтор начала continuation, уже присутствующий в конце черновика."""
    if not draft or not continuation:
        return continuation
    draft_tail = draft.rstrip()
    cont = continuation.lstrip()
    if not draft_tail or not cont:
        return continuation

    draft_words = draft_tail.split()
    cont_words = cont.split()
    if draft_words and cont_words:
        for overlap in range(min(len(draft_words), len(cont_words)), 0, -1):
            left = [word.lower().strip(".,;:!?") for word in draft_words[-overlap:]]
            right = [word.lower().strip(".,;:!?") for word in cont_words[:overlap]]
            if left == right:
                rest = " ".join(cont_words[overlap:]).lstrip(" .,;:-")
                return rest or continuation

    draft_lower = draft_tail.lower()
    cont_lower = cont.lower()
    for size in range(min(len(draft_lower), len(cont_lower), 48), 2, -1):
        if draft_lower[-size:] == cont_lower[:size]:
            return cont[size:].lstrip(" .,;:-") or continuation
    return continuation


def _de_solutionize_continuation(continuation: str, solution: str) -> str:
    """Убирает явное копирование шаблона «Что сделать» в подсказке."""
    text = (continuation or "").strip()
    if not text:
        return text

    for prefix in (
        "контакт:",
        "контакт ",
        "связаться с ",
        "связаться ",
        "позвонить ",
        "проверить ",
        "уточнить ",
        "найти поставщика ",
    ):
        if text.lower().startswith(prefix):
            text = text[len(prefix) :].lstrip(" .,;:-")
            break

    if solution:
        solution_norm = " ".join(solution.lower().split())
        text_norm = " ".join(text.lower().split())
        if text_norm and len(text_norm) >= 24 and text_norm in solution_norm:
            return ""
        if solution_norm and text_norm.startswith(solution_norm[: min(40, len(solution_norm))]):
            return ""

    return text


def _build_suggest_result_prompt(
    *,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    draft_raw: str,
) -> str:
    has_draft = bool((draft_raw or "").strip())
    solution_block = solution.strip() or "—"
    problem_block = problem.strip() or "—"

    shared_context = (
        "Ты помогаешь менеджеру по закупкам сформулировать поле «Результат работы» — "
        "краткий отчёт о том, что уже сделано и какой итог, а не список поручений.\n\n"
        "=== Контекст задания (справочно) ===\n"
        f"Тип: {task_type or '—'}\n"
        f"Номенклатура: {nomenclature or '—'}\n"
        f"Проблема (ситуация): {problem_block}\n"
        f"Рекомендация системы (внутренняя инструкция): {solution_block}\n\n"
        "ВАЖНО про рекомендацию системы:\n"
        "- Это подсказка менеджеру, что нужно сделать. Её нельзя копировать в результат.\n"
        "- Запрещено переносить шаблоны: «Контакт: …», «Связаться с …», «Позвонить …», "
        "«Проверить …», перечисление шагов через точку с запятой.\n"
        "- Из рекомендации можно взять только сущности (контрагент, этап, предмет) и "
        "переформулировать их как факт/действие менеджера своими словами.\n"
    )

    if has_draft:
        return (
            f"{shared_context}"
            "=== Черновик менеджера (уже введён в поле, сохранить дословно) ===\n"
            f"«{draft_raw}»\n\n"
            "=== Твоя задача ===\n"
            "Верни JSON с полем continuation — ТОЛЬКО текст, который логично идёт "
            "сразу ПОСЛЕ черновика (без повторения черновика).\n\n"
            "Правила continuation:\n"
            "1. Приоритет №1 — смысл черновика. Продолжай мысль менеджера, а не подменяй её.\n"
            "2. Если черновик обрывается на середине слова или фразы — сначала аккуратно "
            "заверши её (например: «…связаться с» → « ответственным по таможне в …»).\n"
            "3. Не повторяй слова и фразы, которые уже есть в черновике (даже частично). "
            "Если в черновике уже «связаться с» — не начинай continuation с «связаться с».\n"
            "4. Пиши как отчёт о работе: прошедшее время, конкретика, итог. "
            "Не используй инфинитивы-поручения («связаться», «уточнить», «проверить») "
            "в начале continuation.\n"
            "5. Опираясь на проблему и смысл рекомендации, добавь 1–2 конкретных факта "
            "или статуса: с кем контактировали, что выяснили, что ожидается, нужна ли эскалация.\n"
            "6. Не выдумывай точные даты, номера документов и статусы, которых нет в "
            "черновике. Допустимы нейтральные формулировки: «ответ пока не получен», "
            "«срок уточняю», «повторю контакт».\n"
            "7. Одним связным текстом, по-русски, без markdown и кавычек, до 350 символов.\n\n"
            "Примеры ПЛОХО:\n"
            "- «связаться с. Контакт: ПАРАДИЗ ООО. Связаться с ответственным…» "
            "(копипаст рекомендации + повтор слов черновика)\n"
            "- «Позвонить поставщику, подтвердить отгрузку…» (инструкция вместо результата)\n\n"
            "Примеры ХОРОШО:\n"
            "- Черновик «я пытался связаться с» → "
            "« ответственным в ПАРАДИЗ ООО по таможне, пока без ответа — повторю звонок "
            "и при необходимости эскалирую руководителю»\n"
            "- Черновик «позвонил поставщику» → "
            "«, отгрузку подтвердили, новую дату прибытия уточняю»\n"
            "- Черновик «отправил запрос» → "
            "« в отдел логистики по статусу партии, жду ответ до конца дня»\n\n"
            'JSON: {"continuation":"..."}'
        )

    return (
        f"{shared_context}"
        "=== Твоя задача ===\n"
        "Поле результата пустое. Сформулируй готовый текст результата работы менеджера.\n\n"
        "Правила:\n"
        "1. Отчёт о выполненных действиях и итоге, а не список «что нужно сделать».\n"
        "2. Не копируй рекомендацию системы дословно — переформулируй в результат.\n"
        "3. 2–4 коротких предложения: действие → статус/ответ → следующий шаг при необходимости.\n"
        "4. Без вымышленных дат и номеров; допустимы нейтральные статусы ожидания.\n"
        "5. По-русски, без markdown и кавычек, до 500 символов.\n\n"
        'JSON: {"continuation":"полный текст результата"}'
    )


def _heuristic_continuation(draft_raw: str, solution: str, problem: str) -> str:
    draft = (draft_raw or "").strip()
    draft_lower = draft.lower()

    if draft:
        if any(token in draft_lower for token in ("связ", "звон", "писал", "написал", "обрат")):
            continuation = " ответственным, ответ пока не получен — повторю контакт и зафиксирую результат."
        elif any(token in draft_lower for token in ("уточн", "провер", "запрос")):
            continuation = " статус по позиции, ответ ожидаю — результат дополню после подтверждения."
        elif any(token in draft_lower for token in ("не знаю", "не понима", "не получ")):
            continuation = (
                " как действовать дальше: свяжусь с контрагентом по заданию и зафиксирую "
                "конкретный статус или эскалирую руководителю."
            )
        else:
            continuation = " и зафиксирую итог по заданию после получения ответа."
        return _merge_draft_continuation(draft_raw, continuation)

    if solution:
        text = solution.strip()
        replacements = (
            ("Связаться с ", "Связался с "),
            ("Позвонить ", "Позвонил "),
            ("Проверить ", "Проверил "),
            ("Уточнить ", "Уточнил "),
            ("Найти поставщика ", "Ищу поставщика "),
            ("Согласовать ", "Согласовываю "),
        )
        for src, dst in replacements:
            if text.startswith(src):
                text = dst + text[len(src) :]
                break
        if text.startswith("Контакт:"):
            text = text.replace("Контакт:", "Контактировал с", 1)
        return text[:500]

    if problem:
        return "Проверил ситуацию по позиции, результат и следующий шаг зафиксирую после ответа контрагента."
    return "Связался с контрагентом, результат уточняю."


async def suggest_manager_result(
    *,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    draft: str,
) -> tuple[str, str]:
    """Подсказка формулировки результата менеджера через LM Studio."""
    from app.agents.document_analysis_agent.excel_service import _lm_settings, _post_lm_json
    from app.core.config import settings

    draft_raw = draft or ""
    prompt = _build_suggest_result_prompt(
        task_type=task_type,
        problem=problem,
        solution=solution,
        nomenclature=nomenclature,
        draft_raw=draft_raw,
    )

    lm = _lm_settings()
    if lm is not None:
        base_url, model = lm
        try:
            data = await _post_lm_json(
                base_url,
                model,
                prompt,
                timeout=min(60.0, float(settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS)),
            )
            continuation = str(
                data.get("continuation") or data.get("suggestion") or ""
            ).strip()
            continuation = _de_solutionize_continuation(continuation, solution)
            continuation = _strip_prefix_overlap(draft_raw, continuation)
            continuation = continuation.lstrip(" .,;:-")
            if continuation:
                suggestion = _merge_draft_continuation(draft_raw, continuation)
                if suggestion:
                    logger.info(
                        "document_analysis_agent.shift_result_suggested",
                        source="lm_studio",
                        nomenclature=nomenclature[:80],
                    )
                    return suggestion, "lm_studio"
        except Exception as exc:
            logger.warning(
                "document_analysis_agent.shift_result_suggest_failed",
                error=str(exc),
                nomenclature=nomenclature[:80],
            )

    suggestion = _heuristic_continuation(draft_raw, solution, problem)
    return suggestion[:500], "heuristic"
