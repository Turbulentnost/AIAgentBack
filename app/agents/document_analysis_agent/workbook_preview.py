"""Полный табличный preview Excel для открытия файла во вкладке."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.agents.document_analysis_agent.xls_compat import ensure_openpyxl_bytes

MAX_SHEETS = 40
MAX_ROWS = 2500
MAX_COLS = 80


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%d.%m.%Y %H:%M")
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, bool):
        return "ИСТИНА" if value else "ЛОЖЬ"
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _trim_row(cells: list[str]) -> list[str]:
    while cells and cells[-1] == "":
        cells.pop()
    return cells


def _load_workbook(raw: bytes):
    try:
        return load_workbook(BytesIO(raw), data_only=True, read_only=True)
    except Exception:
        return load_workbook(BytesIO(raw), data_only=True)


def build_workbook_tab_preview(filename: str, content: bytes) -> dict[str, Any]:
    if not content:
        raise ValueError("Файл пуст")
    raw = ensure_openpyxl_bytes(filename, content)
    workbook = _load_workbook(raw)
    sheets: list[dict[str, Any]] = []
    names: list[str] = []
    try:
        names = list(workbook.sheetnames)
        for title in names[:MAX_SHEETS]:
            worksheet = workbook[title]
            values: list[list[str]] = []
            truncated_rows = False
            truncated_cols = False
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= MAX_ROWS:
                    truncated_rows = True
                    break
                if len(row) > MAX_COLS:
                    truncated_cols = True
                values.append(_trim_row([_cell_text(cell) for cell in row[:MAX_COLS]]))
            while values and not any(values[-1]):
                values.pop()
            sheets.append(
                {
                    "name": title,
                    "values": values,
                    "row_count": max(len(values) - 1, 0) if values else 0,
                    "truncated_rows": truncated_rows,
                    "truncated_cols": truncated_cols,
                }
            )
    finally:
        workbook.close()

    if not sheets:
        raise ValueError("В книге нет листов")

    return {
        "ok": True,
        "file_name": filename,
        "sheets": sheets,
        "sheet_count": len(names),
        "truncated_sheets": len(names) > MAX_SHEETS,
    }
