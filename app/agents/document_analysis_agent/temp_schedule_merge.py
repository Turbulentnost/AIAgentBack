"""TEMP(Aveon schedule merge) — объединение графика + ТАМОЖНЯ. Удалить вместе с TEMP UI."""

from __future__ import annotations

import base64
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from difflib import SequenceMatcher
from typing import Any, Iterable

import httpx
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings

_META_HEADERS = (
    "Номенклатура",
    "Изделие",
    "Заказано кол-во, шт",
    "Остаток недополученной",
    "Дата заказа",
    "Срок обработки",
    "Дата спецификации",
    "Дата платежа",
    "Логистика до МСК",
    "Логистика МСК-Ростов",
    "Сроки производства",
    "Сроки поставки в МСК",
)
_TEXT_META = (
    "Изделие",
    "Дата заказа",
    "Срок обработки",
    "Дата спецификации",
    "Дата платежа",
    "Логистика до МСК",
    "Логистика МСК-Ростов",
    "Сроки производства",
    "Сроки поставки в МСК",
)
_STOP = {
    "для",
    "шт",
    "the",
    "and",
    "with",
    "про",
    "без",
    "акб",
    "мм",
    "см",
    "кабель",
    "разъем",
    "комплект",
    "самолета",
}


def _norm_name(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[\"'`«»]", "", text)
    text = re.sub(r"\s*\(\d+\)(?:\s+[a-zа-я0-9._-]{1,20})?\s*$", "", text)
    text = re.sub(r"\s*\(2026\)\s*$", "", text)
    return text.strip()


def _display_name(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    # партия: "(1)", "(3) 3115", "(2) ..."
    text = re.sub(r"\s*\(\d+\)(?:\s+[A-Za-zА-Яа-я0-9._-]{1,20})?\s*$", "", text).strip()
    return text


def _distinctive_codes(name: str) -> frozenset[str]:
    """Уникальные коды модели из названия — без правил под конкретные товары."""
    n = _norm_name(name)
    codes = set(re.findall(r"(?<![a-z0-9])[a-z]*\d+[a-z0-9]*(?![a-z0-9])", n))
    codes |= {t for t in _tokens(n) if re.search(r"\d", t) and len(t) >= 3}
    return frozenset(codes)


def _item_signature(name: str) -> str | None:
    codes = sorted(_distinctive_codes(name))
    flags = sorted(_variant_flags(name))
    parts = list(codes) + [f"@{f}" for f in flags]
    if not parts:
        return None
    return "|".join(parts[:12])


def _polarity_markers(name: str) -> set[str]:
    n = _norm_name(name)
    markers: set[str] = set()
    if re.search(r"xt60[- ]?m|xt60m", n):
        markers.add("m")
    if re.search(r"xt60[- ]?f|xt60f", n):
        markers.add("f")
    if re.search(r"\bmale\b|[- ]m\b", n):
        markers.add("m")
    if re.search(r"\bfemale\b|[- ]f\b", n):
        markers.add("f")
    return markers


def _variant_codes(name: str) -> set[str]:
    """Коды вариантов модели (z8≠z40, mr60-m≠mr60-f)."""
    n = _norm_name(name)
    codes: set[str] = set()
    codes |= set(re.findall(r"(?<![a-z0-9])z\d+(?![a-z0-9])", n))
    codes |= set(re.findall(r"(?<![a-z0-9])x\d+[a-z0-9]*(?![a-z0-9])", n))
    codes |= set(re.findall(r"(?<![a-z0-9])mr\d+-?[mf](?![a-z0-9])", n))
    codes |= set(re.findall(r"(?<![a-z0-9])xt60-?[mf](?![a-z0-9])", n))
    codes |= set(re.findall(r"5mm-[gr]\b", n))
    return codes


def _variant_flags(name: str) -> set[str]:
    n = _norm_name(name)
    flags: set[str] = set()
    if re.search(r"(?<![a-z0-9])mini(?![a-z0-9])", n):
        flags.add("mini")
    if re.search(r"(?<![a-z0-9])pro(?![a-z0-9])", n):
        flags.add("pro")
    return flags


_KIND_TOKENS = frozenset(
    {
        "провод",
        "кабель",
        "разъем",
        "коннектор",
        "пульт",
        "передатчик",
        "контроллер",
        "регулятор",
        "мотор",
        "камера",
        "аккумулятор",
        "батарея",
        "винт",
        "гайка",
    }
)


def _product_kinds(name: str) -> set[str]:
    return _tokens(name) & _KIND_TOKENS


def _tokens(value: str) -> set[str]:
    found = re.findall(r"[a-zа-я0-9]+", _norm_name(value))
    out = set()
    for t in found:
        if t in _STOP:
            continue
        if len(t) >= 3 or re.search(r"\d", t):
            out.add(t)
    return out


def _conflicting_models(a: str, b: str) -> bool:
    """Обобщённая проверка: похожие названия, но разные коды/варианты модели."""
    va, vb = _variant_codes(a), _variant_codes(b)
    if va and vb and not (va & vb):
        return True
    score = _token_score(a, b)
    if score < 0.55:
        return False
    ca, cb = _distinctive_codes(a), _distinctive_codes(b)
    if ca and cb and not (ca & cb):
        return True
    va, vb = _variant_flags(a), _variant_flags(b)
    if ("mini" in va) ^ ("mini" in vb):
        na = re.sub(r"\bmini\b", "", _norm_name(a))
        nb = re.sub(r"\bmini\b", "", _norm_name(b))
        if SequenceMatcher(None, na, nb).ratio() > 0.88:
            return True
    sa, sb = _sig_tokens(a), _sig_tokens(b)
    brand_a = {t for t in sa if len(t) >= 5 and not re.search(r"\d", t)}
    brand_b = {t for t in sb if len(t) >= 5 and not re.search(r"\d", t)}
    if brand_a and brand_b and not (brand_a & brand_b) and score > 0.6:
        return True
    pa, pb = _polarity_markers(a), _polarity_markers(b)
    if pa and pb and pa != pb:
        return True
    return False


def _sig_tokens(value: str) -> set[str]:
    """Токены-маркеры: бренды, коды моделей."""
    toks = _tokens(value)
    return {t for t in toks if len(t) >= 4 or re.search(r"\d", t)}


def _parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _model_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        # Excel часто ломает коды вида 7012-2 в дату
        if value.year >= 2500 or value.year < 1950:
            return f"{value.year}-{value.month}"
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        if value.year >= 2500 or value.year < 1950:
            return f"{value.year}-{value.month}"
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _merge_text(*parts: Any) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for part in parts:
        text = str(part or "").strip().replace("\n", " | ")
        if not text:
            continue
        for chunk in re.split(r"\s*\|\s*", text):
            chunk = chunk.strip()
            if not chunk:
                continue
            key = chunk.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(chunk)
    return " | ".join(out)


def _parse_qty_by_dates(text: Any) -> dict[date, float]:
    result: dict[date, float] = {}
    raw = str(text or "")
    if not raw.strip():
        return result
    year = 2026
    pattern = re.compile(
        r"(\d{1,2})[./](\d{1,2})(?:[./](\d{2,4}))?\s*[-–—]\s*([\d\s\xa0]+)\s*шт",
        re.IGNORECASE,
    )
    for m in pattern.finditer(raw):
        d, mo, y, qty_s = m.groups()
        yy = int(y) if y else year
        if yy < 100:
            yy += 2000
        try:
            day = date(yy, int(mo), int(d))
        except ValueError:
            continue
        qty = _parse_number(qty_s)
        if qty:
            result[day] = result.get(day, 0.0) + qty
    return result


def _token_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    ta, tb = _tokens(a), _tokens(b)
    if not ta or not tb:
        return SequenceMatcher(None, a, b).ratio()
    jacc = len(ta & tb) / len(ta | tb)
    seq = SequenceMatcher(None, a, b).ratio()
    return max(jacc, seq)


def _is_safe_same_item(a: str, b: str) -> bool:
    """Строгое равенство товара без ложных подмен."""
    na, nb = _norm_name(a), _norm_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if _conflicting_models(a, b):
        return False
    va, vb = _variant_codes(a), _variant_codes(b)
    if va and vb and not (va & vb):
        return False
    ka, kb = _product_kinds(a), _product_kinds(b)
    if ka and kb and not (ka & kb):
        return False
    if na in nb or nb in na:
        shorter, longer = (na, nb) if len(na) <= len(nb) else (nb, na)
        sig_s, sig_l = _sig_tokens(shorter), _sig_tokens(longer)
        if sig_s and not sig_s <= sig_l:
            return False
        return _token_score(na, nb) >= 0.72

    score = _token_score(na, nb)
    if score < 0.88:
        return False
    sa, sb = _sig_tokens(a), _sig_tokens(b)
    weak = {"mkii", "акб", "без", "pro", "motor", "2026"}
    strong_a = {t for t in sa if (re.search(r"\d", t) or len(t) >= 5) and t not in weak}
    strong_b = {t for t in sb if (re.search(r"\d", t) or len(t) >= 5) and t not in weak}
    if strong_a and strong_b:
        if not (strong_a & strong_b):
            return False
        only_a = strong_a - strong_b
        only_b = strong_b - strong_a
        # один конфликтующий код модели — уже стоп
        if only_a and only_b and not (only_a & only_b):
            # если уникальные сильные токены есть с обеих сторон — подозрительно
            if len(only_a) >= 1 and len(only_b) >= 1 and score < 0.95:
                return False
        if len(only_a) >= 2 or len(only_b) >= 2:
            return False
    return True


def _prefer_name(a: str, b: str) -> str:
    """Выбрать более полное имя без потери маркеров."""
    if not a:
        return b
    if not b:
        return a
    sa, sb = _sig_tokens(a), _sig_tokens(b)
    # если один содержит все маркеры другого и длиннее — его
    if sa <= sb and len(b) >= len(a):
        return b
    if sb <= sa and len(a) >= len(b):
        return a
    # иначе более длинное с большим числом маркеров
    if len(sa) != len(sb):
        return a if len(sa) > len(sb) else b
    return a if len(a) >= len(b) else b


@dataclass
class NomRow:
    name: str
    ordered: float = 0.0
    remainder: float | None = None
    meta: dict[str, str] = field(default_factory=dict)
    schedule: dict[date, float] = field(default_factory=dict)
    sources: list[str] = field(default_factory=list)

    def add_qty(self, day: date | None, qty: float) -> None:
        if day is None or qty == 0:
            return
        self.schedule[day] = self.schedule.get(day, 0.0) + qty

    def merge_from(self, other: "NomRow") -> None:
        self.name = _prefer_name(self.name, other.name)
        self.ordered += other.ordered
        if other.remainder is not None:
            self.remainder = (self.remainder or 0.0) + other.remainder
        for key in _TEXT_META:
            self.meta[key] = _merge_text(self.meta.get(key, ""), other.meta.get(key, ""))
        for day, qty in other.schedule.items():
            self.schedule[day] = self.schedule.get(day, 0.0) + qty
        for src in other.sources:
            if src not in self.sources:
                self.sources.append(src)


@dataclass
class NameDict:
    entries: list[tuple[str, str, str]] = field(default_factory=list)
    full_by_model: dict[str, str] = field(default_factory=dict)
    full_by_norm: dict[str, str] = field(default_factory=dict)


def _build_name_dict(itc_rows: list[tuple]) -> NameDict:
    nd = NameDict()
    for row in itc_rows[2:]:
        if not row:
            continue
        poz = str(row[1] or "").strip() if len(row) > 1 else ""
        model = _model_text(row[2] if len(row) > 2 else "")
        if model.lower().startswith("спец"):
            continue
        if not poz and not model:
            continue
        full = re.sub(r"\s+", " ", f"{poz} {model}".strip())
        if not full:
            continue
        nd.entries.append((poz, model, full))
        nd.full_by_norm[_norm_name(full)] = full
        if model:
            mk = _norm_name(model)
            if mk:
                # не затирать более полное
                prev = nd.full_by_model.get(mk)
                nd.full_by_model[mk] = _prefer_name(prev or "", full)
    return nd


def _enrich_name(raw: str, nd: NameDict) -> str:
    """Только если в графике одна модель — подставить позиция+модель. Иначе имя не трогать."""
    name = _display_name(raw)
    if not name or not nd.entries:
        return name
    key = _norm_name(name)
    # точное совпадение с моделью ИТЦ
    if key in nd.full_by_model:
        full = nd.full_by_model[key]
        # обогащаем только если исходник не содержит уже позиции/бренда сверх модели
        if _sig_tokens(name) <= _sig_tokens(full) or _norm_name(name) == _norm_name(
            next((m for _p, m, _f in nd.entries if _norm_name(m) == key), name)
        ):
            return full
    for poz, model, full in nd.entries:
        if not poz or not model:
            continue
        if _norm_name(model) == key:
            return full
    return name


def _find_existing_key(name: str, by_key: dict[str, NomRow]) -> str | None:
    key = _norm_name(name)
    if key in by_key:
        return key
    for k, row in by_key.items():
        if _is_safe_same_item(name, row.name):
            return k
    fp = _item_signature(name)
    if fp:
        for k, row in by_key.items():
            if (
                _item_signature(row.name) == fp
                and _is_safe_same_item(name, row.name)
                and not _conflicting_models(name, row.name)
            ):
                return k
    return None


def build_schedule_name_index(names: Iterable[str]) -> dict[str, str]:
    """Нормализованный ключ → каноническое имя как в объединённом графике отгрузок."""
    by_key: dict[str, NomRow] = {}
    for raw in names:
        name = _display_name(str(raw or ""))
        if not name:
            continue
        existing = _find_existing_key(name, by_key)
        if existing:
            by_key[existing].name = _prefer_name(by_key[existing].name, name)
        else:
            by_key[_norm_name(name)] = NomRow(name=name)
    return {_norm_name(row.name): row.name for row in by_key.values()}


def resolve_schedule_name(name: str, index: dict[str, str]) -> str | None:
    """Сопоставляет имя с канонической номенклатурой объединённого графика."""
    text = str(name or "").strip()
    if not text or not index:
        return None
    by_key = {key: NomRow(name=value) for key, value in index.items()}
    found = _find_existing_key(text, by_key)
    if found:
        return by_key[found].name
    normalized = _norm_name(_display_name(text))
    return index.get(normalized)


def _lm_settings() -> tuple[str, str] | None:
    base_url = settings.AVEON_LM_STUDIO_BASE_URL.strip().rstrip("/")
    model = settings.AVEON_LM_STUDIO_MODEL.strip()
    if not base_url or not model:
        return None
    if "://localhost" in base_url:
        base_url = base_url.replace("://localhost", "://127.0.0.1", 1)
    return base_url, model


async def _lm_match_chunk(
    queries: list[str],
    canonical_names: list[str],
    timeout: float = 90.0,
) -> dict[str, str | None]:
    cfg = _lm_settings()
    if not cfg:
        return {}
    base_url, model = cfg
    prompt = (
        "Сопоставь queries с nomenclature. Это номенклатура поставок.\n"
        "Правила: сопоставляй ТОЛЬКО один и тот же товар. "
        "Если сомнение, разные модели (Z8/Z40, mini/не mini), разные бренды "
        "(SmallRig/GCR), разные кабели — canonical=null.\n"
        "Верни JSON: {\"matches\":[{\"query\":\"...\",\"canonical\":\"...\"|null}]}\n"
        "canonical — точная строка из nomenclature или null.\n"
        f"nomenclature={json.dumps(canonical_names, ensure_ascii=False)}\n"
        f"queries={json.dumps(queries, ensure_ascii=False)}"
    )
    try:
        async with httpx.AsyncClient(
            timeout=httpx.Timeout(timeout, connect=min(5.0, timeout)),
            trust_env=False,
        ) as client:
            response = await client.post(
                f"{base_url}/chat/completions",
                json={
                    "model": model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "Только валидный JSON. Без markdown. При сомнении null.",
                        },
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0,
                    "stream": False,
                },
            )
            response.raise_for_status()
            content = response.json()["choices"][0]["message"]["content"].strip()
            if content.startswith("```"):
                content = re.sub(r"^```(?:json)?\s*", "", content)
                content = re.sub(r"\s*```$", "", content)
            data = json.loads(content)
    except Exception:
        return {}

    canon_set = set(canonical_names)
    out: dict[str, str | None] = {}
    for item in data.get("matches") or []:
        q = str(item.get("query") or "").strip()
        c = item.get("canonical")
        if not c:
            out[q] = None
            continue
        c_str = str(c).strip()
        if c_str in canon_set and _is_safe_same_item(q, c_str):
            out[q] = c_str
        else:
            out[q] = None
    return out


async def _lm_find_duplicates(names: list[str]) -> list[tuple[str, str]]:
    """Пары дублей для склейки."""
    cfg = _lm_settings()
    if not cfg or len(names) < 2:
        return []
    base_url, model = cfg
    prompt = (
        "Найди пары названий одного и того же товара в списке names.\n"
        "Не объединяй разные модели/бренды (Z8≠Z40, mini≠обычный, SmallRig≠GCR, EOLO отдельно).\n"
        "Верни JSON: {\"pairs\":[[\"имя1\",\"имя2\"], ...]}\n"
        f"names={json.dumps(names, ensure_ascii=False)}"
    )
    try:
        async with httpx.AsyncClient(
            timeout=httpx.Timeout(120.0, connect=5.0),
            trust_env=False,
        ) as client:
            response = await client.post(
                f"{base_url}/chat/completions",
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": "Только валидный JSON. Без markdown."},
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0,
                    "stream": False,
                },
            )
            response.raise_for_status()
            content = response.json()["choices"][0]["message"]["content"].strip()
            if content.startswith("```"):
                content = re.sub(r"^```(?:json)?\s*", "", content)
                content = re.sub(r"\s*```$", "", content)
            data = json.loads(content)
    except Exception:
        return []

    name_set = set(names)
    pairs: list[tuple[str, str]] = []
    for pair in data.get("pairs") or []:
        if not isinstance(pair, (list, tuple)) or len(pair) < 2:
            continue
        a, b = str(pair[0]).strip(), str(pair[1]).strip()
        if a in name_set and b in name_set and a != b and _is_safe_same_item(a, b):
            pairs.append((a, b))
    return pairs


async def _resolve_matches(queries: list[str], by_key: dict[str, NomRow]) -> dict[str, str | None]:
    names = [row.name for row in by_key.values()]
    resolved: dict[str, str | None] = {}
    pending: list[str] = []
    for q in queries:
        local_key = _find_existing_key(q, by_key)
        if local_key:
            resolved[q] = by_key[local_key].name
        else:
            pending.append(q)
    for i in range(0, len(pending), 15):
        chunk = pending[i : i + 15]
        lm = await _lm_match_chunk(chunk, names)
        for q in chunk:
            cand = lm.get(q)
            if cand and _is_safe_same_item(q, cand):
                resolved[q] = cand
            else:
                resolved[q] = None
    return resolved


def _is_junk_name(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, (datetime, date)):
        return True
    text = str(value).strip()
    if not text:
        return True
    if _as_date(value) is not None and not re.search(r"[A-Za-zА-Яа-я]", text):
        return True
    return False


@dataclass
class SheetLayout:
    kind: str  # schedule | itc | skip
    header_row: int = 1
    data_start_row: int = 2
    product: str = ""
    name_col: int = 0
    ordered_col: int | None = 1
    remainder_col: int | None = None
    meta_cols: dict[str, int] = field(default_factory=dict)
    date_cols: list[tuple[int, date]] = field(default_factory=list)
    itc_left: dict[str, int] | None = None
    itc_batch: dict[str, int] | None = None
    confidence: float = 0.0
    reason: str = ""


_COPY_ONLY_SHEETS = frozenset({"ТАМОЖНЯ", "Реестр Заказов", "График", "Источник", "Проверки"})
_NOM_KW = ("номенклатур", "наименован", "позици", "изделие", "наимен")
_ORDER_KW = ("заказано", "кол-во", "количество", "qty", "заказ")
_REMAINDER_KW = ("остаток", "недопол", "remainder")
_META_KEYWORDS: dict[str, tuple[str, ...]] = {
    "Дата заказа": ("дата заказ",),
    "Срок обработки": ("срок обработ", "обработ"),
    "Дата спецификации": ("дата спец", "спецификац"),
    "Дата платежа": ("дата плат", "платеж"),
    "Логистика до МСК": ("логистика до", "до мск"),
    "Логистика МСК-Ростов": ("мск-ростов", "ростов"),
    "Сроки производства": ("сроки производ", "производств"),
    "Сроки поставки в МСК": ("сроки постав", "поставк в"),
}


def _cell_text(cell: Any) -> str:
    return str(cell or "").strip()


def _row_texts(row: tuple | list, limit: int = 50) -> list[str]:
    return [_cell_text(c) for c in row[:limit]]


def _sheet_preview(rows: list[tuple], max_rows: int = 28, max_cols: int = 42) -> list[list[str]]:
    out: list[list[str]] = []
    for row in rows[:max_rows]:
        cells: list[str] = []
        for cell in row[:max_cols]:
            if isinstance(cell, (datetime, date)):
                cells.append(cell.isoformat() if isinstance(cell, date) else cell.date().isoformat())
            elif isinstance(cell, float) and cell == int(cell):
                cells.append(str(int(cell)))
            else:
                cells.append(_cell_text(cell))
        out.append(cells)
    return out


def _match_meta_col(header: list[str]) -> dict[str, int]:
    meta: dict[str, int] = {}
    for idx, raw in enumerate(header):
        h = raw.lower()
        if not h:
            continue
        for label, kws in _META_KEYWORDS.items():
            if label in meta:
                continue
            if any(kw in h for kw in kws):
                meta[label] = idx
                break
    return meta


def _find_remainder_col(header: list[str], row0: list[str] | None = None) -> int | None:
    for idx, raw in enumerate(header):
        if any(kw in raw.lower() for kw in _REMAINDER_KW):
            return idx
    if row0:
        for idx, raw in enumerate(row0):
            if any(kw in raw.lower() for kw in _REMAINDER_KW):
                return idx
    return None


def _find_name_col(header: list[str]) -> int:
    for idx, raw in enumerate(header):
        if any(kw in raw.lower() for kw in _NOM_KW):
            return idx
    return 0


def _find_ordered_col(header: list[str], name_col: int) -> int | None:
    for idx, raw in enumerate(header):
        if idx == name_col:
            continue
        if any(kw in raw.lower() for kw in _ORDER_KW):
            return idx
    return 1 if len(header) > 1 and name_col != 1 else None


def _extract_date_cols(header_row: tuple | list) -> list[tuple[int, date]]:
    out: list[tuple[int, date]] = []
    for idx, cell in enumerate(header_row):
        day = _as_date(cell)
        if day:
            out.append((idx, day))
    return out


def _count_schedule_data_rows(rows: list[tuple], layout: SheetLayout) -> int:
    count = 0
    for row in rows[layout.data_start_row : layout.data_start_row + 80]:
        if not row:
            continue
        if layout.name_col >= len(row) or _is_junk_name(row[layout.name_col]):
            continue
        qty_sum = 0.0
        for idx, _ in layout.date_cols:
            if idx < len(row):
                qty_sum += _parse_number(row[idx]) or 0.0
        ordered = 0.0
        if layout.ordered_col is not None and layout.ordered_col < len(row):
            ordered = _parse_number(row[layout.ordered_col]) or 0.0
        if qty_sum > 0 or ordered > 0:
            count += 1
    return count


def _heuristic_schedule_layout(rows: list[tuple], sheet_title: str) -> SheetLayout | None:
    if len(rows) < 3:
        return None
    best: SheetLayout | None = None
    best_score = 0
    for hr in range(min(10, len(rows))):
        header = _row_texts(rows[hr])
        date_cols = _extract_date_cols(rows[hr])
        if len(date_cols) < 2:
            continue
        has_nom_kw = any(any(kw in h.lower() for kw in _NOM_KW) for h in header)
        has_order_kw = any(any(kw in h.lower() for kw in _ORDER_KW) for h in header)
        name_col = _find_name_col(header)
        ordered_col = _find_ordered_col(header, name_col)
        meta_cols = _match_meta_col(header)
        row0 = _row_texts(rows[hr - 1]) if hr > 0 else None
        remainder_col = _find_remainder_col(header, row0)
        product = ""
        if hr > 0 and rows[hr - 1]:
            product = _cell_text(rows[hr - 1][name_col]) or _cell_text(rows[hr - 1][0])
        if not product:
            product = sheet_title.strip()
        layout = SheetLayout(
            kind="schedule",
            header_row=hr,
            data_start_row=hr + 1,
            product=product,
            name_col=name_col,
            ordered_col=ordered_col,
            remainder_col=remainder_col,
            meta_cols=meta_cols,
            date_cols=date_cols,
        )
        data_rows = _count_schedule_data_rows(rows, layout)
        score = len(date_cols) * 3 + data_rows * 2 + (2 if has_nom_kw else 0) + (1 if has_order_kw else 0)
        if score > best_score and data_rows >= 1:
            best_score = score
            layout.confidence = min(0.95, 0.4 + data_rows * 0.02 + len(date_cols) * 0.03)
            layout.reason = f"heuristic: {data_rows} rows, {len(date_cols)} dates"
            best = layout
    return best


def _heuristic_itc_layout(rows: list[tuple]) -> SheetLayout | None:
    if len(rows) < 4:
        return None
    header0 = _row_texts(rows[0])
    header1 = _row_texts(rows[1]) if len(rows) > 1 else header0
    joined = " ".join(h.lower() for h in header0 + header1)
    if not any(kw in joined for kw in ("позици", "модел", "спецификац", "отгруз")):
        return None
    left: dict[str, int] = {}
    for idx, raw in enumerate(header0):
        h = raw.lower()
        if "позици" in h:
            left.setdefault("position", idx)
        elif "модел" in h:
            left.setdefault("model", idx)
        elif any(kw in h for kw in _ORDER_KW):
            left.setdefault("qty", idx)
        elif "отгруз" in h or "примерная дата" in h:
            left.setdefault("ship_text", idx)
        elif "плат" in h:
            left.setdefault("pay", idx)
        elif "остат" in h:
            left.setdefault("remain", idx)
    left.setdefault("position", 1)
    left.setdefault("model", 2)
    left.setdefault("qty", 3)
    left.setdefault("ship_text", 4)
    left.setdefault("pay", 5)
    left.setdefault("remain", 6)

    batch: dict[str, int] = {}
    for idx, raw in enumerate(header0):
        h = raw.lower()
        if "номенклатур" in h or "наимен" in h:
            batch.setdefault("name", idx)
        elif "кол" in h and "qty" not in batch:
            batch.setdefault("qty", idx)
        elif "дата" in h:
            batch.setdefault("date", idx)
        elif "коммент" in h or "примеч" in h:
            batch.setdefault("note", idx)
    if "name" not in batch:
        for idx, raw in enumerate(header0):
            if idx >= 8 and raw.strip():
                batch["name"] = idx
                break
    if "name" in batch:
        nc = batch["name"]
        batch.setdefault("qty", nc + 14)
        batch.setdefault("date", nc + 16)
        batch.setdefault("note", nc + 17)

    data_left = 0
    for row in rows[2:40]:
        if not row:
            continue
        poz = _cell_text(row[left["position"]] if len(row) > left["position"] else "")
        model = _model_text(row[left["model"]] if len(row) > left["model"] else "")
        if poz or model:
            data_left += 1
    if data_left < 2 and "name" not in batch:
        return None
    return SheetLayout(
        kind="itc",
        header_row=0,
        data_start_row=2,
        itc_left=left,
        itc_batch=batch if "name" in batch else None,
        confidence=0.7 if data_left >= 2 else 0.5,
        reason=f"heuristic itc: {data_left} spec rows",
    )


def _heuristic_skip_layout(rows: list[tuple], sheet_title: str) -> SheetLayout | None:
    if sheet_title.strip() in _COPY_ONLY_SHEETS:
        return SheetLayout(kind="skip", reason="copy-only sheet")
    if not rows:
        return SheetLayout(kind="skip", reason="empty")
    header = _row_texts(rows[0]) + _row_texts(rows[1] if len(rows) > 1 else rows[0])
    joined = " ".join(h.lower() for h in header)
    if any(kw in joined for kw in ("тамож", "реестр заказ", "тип операции", "основание")):
        return SheetLayout(kind="skip", reason="registry/customs")
    return None


async def _lm_detect_layout(sheet_title: str, preview: list[list[str]], filename: str) -> SheetLayout | None:
    cfg = _lm_settings()
    if not cfg:
        return None
    base_url, model = cfg
    prompt = (
        "Определи структуру листа Excel: график отгрузок (schedule), ИТЦ (itc) или skip.\n"
        "JSON: {\"kind\":\"schedule|itc|skip\",\"header_row\":1,\"data_start_row\":2,"
        "\"name_col\":0,\"ordered_col\":1,\"remainder_col\":null,\"product\":\"\","
        "\"meta_cols\":{\"Дата заказа\":2},"
        "\"date_cols\":[{\"col\":8,\"date\":\"2026-07-09\"}],"
        "\"itc_left\":{\"position\":1,\"model\":2,\"qty\":3,\"ship_text\":4,\"pay\":5,\"remain\":6},"
        "\"itc_batch\":{\"name\":8,\"qty\":22,\"date\":24,\"note\":25},"
        "\"confidence\":0.9,\"reason\":\"...\"}\n"
        "schedule: номенклатура + колонки дат с количествами. col — 0-based.\n"
        f"filename={filename}\nsheet={sheet_title}\n"
        f"preview={json.dumps(preview, ensure_ascii=False)}"
    )
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(60.0, connect=5.0), trust_env=False) as client:
            response = await client.post(
                f"{base_url}/chat/completions",
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": "Только валидный JSON. Без markdown."},
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0,
                    "stream": False,
                },
            )
            response.raise_for_status()
            content = response.json()["choices"][0]["message"]["content"].strip()
            if content.startswith("```"):
                content = re.sub(r"^```(?:json)?\s*", "", content)
                content = re.sub(r"\s*```$", "", content)
            data = json.loads(content)
    except Exception:
        return None

    kind = str(data.get("kind") or "skip").lower()
    if kind not in ("schedule", "itc", "skip"):
        kind = "skip"
    layout = SheetLayout(
        kind=kind,
        header_row=int(data.get("header_row") or 0),
        data_start_row=int(data.get("data_start_row") or 1),
        product=str(data.get("product") or "").strip(),
        name_col=int(data.get("name_col") or 0),
        ordered_col=data.get("ordered_col"),
        remainder_col=data.get("remainder_col"),
        confidence=float(data.get("confidence") or 0.5),
        reason=str(data.get("reason") or "lm"),
    )
    if layout.ordered_col is not None:
        layout.ordered_col = int(layout.ordered_col)
    if layout.remainder_col is not None:
        layout.remainder_col = int(layout.remainder_col)
    meta_raw = data.get("meta_cols") or {}
    if isinstance(meta_raw, dict):
        layout.meta_cols = {str(k): int(v) for k, v in meta_raw.items()}
    for item in data.get("date_cols") or []:
        if isinstance(item, dict):
            col = int(item.get("col", -1))
            day = _as_date(item.get("date"))
            if col >= 0 and day:
                layout.date_cols.append((col, day))
    if isinstance(data.get("itc_left"), dict):
        layout.itc_left = {str(k): int(v) for k, v in data["itc_left"].items()}
    if isinstance(data.get("itc_batch"), dict):
        layout.itc_batch = {str(k): int(v) for k, v in data["itc_batch"].items()}
    return layout


async def detect_sheet_layout(ws, sheet_title: str, filename: str) -> SheetLayout:
    rows = list(ws.iter_rows(values_only=True))
    title_lower = sheet_title.strip().lower()
    skip = _heuristic_skip_layout(rows, sheet_title)
    if skip and skip.kind == "skip" and sheet_title.strip() in _COPY_ONLY_SHEETS:
        return skip
    sched = _heuristic_schedule_layout(rows, sheet_title)
    itc = _heuristic_itc_layout(rows)
    if itc and ("итц" in title_lower or itc.confidence >= 0.65):
        if not sched or len(sched.date_cols) < 3:
            return itc
    if sched and len(sched.date_cols) < 2:
        sched = None
    candidates = [c for c in (sched, itc) if c]
    if skip and skip.kind == "skip" and not candidates:
        return skip
    best = max(candidates, key=lambda c: c.confidence, default=None)
    if best and best.confidence >= 0.75:
        return best
    lm = await _lm_detect_layout(sheet_title, _sheet_preview(rows), filename)
    if lm and lm.kind != "skip":
        if lm.kind == "schedule" and not lm.date_cols and sched:
            lm.date_cols = sched.date_cols
            lm.meta_cols = sched.meta_cols or lm.meta_cols
            if lm.remainder_col is None:
                lm.remainder_col = sched.remainder_col
        if lm.kind == "schedule" and _count_schedule_data_rows(rows, lm) == 0 and sched:
            return sched
        return lm
    if best:
        return best
    return skip or SheetLayout(kind="skip", reason="no pattern")


def _parse_schedule_layout(
    ws,
    layout: SheetLayout,
    nd: NameDict,
) -> tuple[list[NomRow], set[date]]:
    rows = list(ws.iter_rows(values_only=True))
    all_dates: set[date] = set(d for _, d in layout.date_cols)
    result: list[NomRow] = []
    product = layout.product or ws.title.strip()
    for row in rows[layout.data_start_row :]:
        if not row or layout.name_col >= len(row) or _is_junk_name(row[layout.name_col]):
            continue
        raw_name = _cell_text(row[layout.name_col])
        name = _enrich_name(raw_name, nd)
        ordered = 0.0
        if layout.ordered_col is not None and layout.ordered_col < len(row):
            ordered = _parse_number(row[layout.ordered_col]) or 0.0
        meta: dict[str, str] = {"Изделие": product}
        for label, col_idx in layout.meta_cols.items():
            if col_idx < len(row):
                meta[label] = _cell_text(row[col_idx])
        rem = None
        if layout.remainder_col is not None and layout.remainder_col < len(row):
            rem = _parse_number(row[layout.remainder_col])
        item = NomRow(
            name=name,
            ordered=ordered,
            remainder=rem,
            meta=meta,
            sources=[f"график:{ws.title}"],
        )
        if name != raw_name:
            item.meta["Дата заказа"] = _merge_text(
                item.meta.get("Дата заказа", ""), f"исх.имя: {raw_name}"
            )
        for idx, day in layout.date_cols:
            if idx < len(row):
                qty = _parse_number(row[idx])
                if qty:
                    item.add_qty(day, qty)
                    all_dates.add(day)
        result.append(item)
    return result, all_dates


def _col_val(row: tuple, layout_map: dict[str, int], key: str, default: int) -> Any:
    idx = layout_map.get(key, default)
    return row[idx] if len(row) > idx else None


def _parse_itc_with_layout(
    rows: list[tuple],
    layout: SheetLayout,
    nd: NameDict,
) -> tuple[list[NomRow], list[tuple[str, float, date | None, str]]]:
    left = layout.itc_left or {}
    batch_map = layout.itc_batch or {}
    left_items: list[NomRow] = []
    for row in rows[layout.data_start_row :]:
        if not row:
            continue
        poz = _cell_text(_col_val(row, left, "position", 1))
        model = _model_text(_col_val(row, left, "model", 2))
        if model.lower().startswith("спец"):
            continue
        if not poz and not model:
            continue
        full = re.sub(r"\s+", " ", f"{poz} {model}".strip())
        ordered = _parse_number(_col_val(row, left, "qty", 3))
        ship_text = _cell_text(_col_val(row, left, "ship_text", 4))
        pay = _cell_text(_col_val(row, left, "pay", 5))
        remain = _parse_number(_col_val(row, left, "remain", 6))
        item = NomRow(
            name=full,
            ordered=0.0,
            remainder=remain,
            meta={"Дата спецификации": ship_text, "Дата платежа": pay},
            sources=["итц:спецификация"],
        )
        for day, qty in _parse_qty_by_dates(ship_text).items():
            item.add_qty(day, qty)
        if ordered and not item.schedule:
            item.ordered = ordered
        left_items.append(item)

    batches: list[tuple[str, float, date | None, str]] = []
    if batch_map.get("name") is not None:
        name_idx = batch_map["name"]
        qty_idx = batch_map.get("qty", name_idx + 14)
        date_candidates = [
            batch_map.get("date"),
            name_idx + 16,
            name_idx + 15,
            name_idx + 13,
            name_idx + 11,
            name_idx + 10,
        ]
        note_idx = batch_map.get("note", name_idx + 17)
        for row in rows[layout.data_start_row :]:
            if not row:
                continue
            name = _cell_text(row[name_idx] if len(row) > name_idx else "")
            if not name and len(row) > name_idx + 13:
                name = _cell_text(row[name_idx + 13])
            if not name:
                continue
            name = _enrich_name(_display_name(name), nd)
            qty = _parse_number(row[qty_idx] if len(row) > qty_idx else None)
            if not qty:
                continue
            day = None
            for idx in date_candidates:
                if idx is None:
                    continue
                if len(row) > idx:
                    day = _as_date(row[idx])
                    if day:
                        break
            note = _cell_text(row[note_idx] if len(row) > note_idx else "")
            batches.append((name, qty, day, note))
    return left_items, batches


def _merge_rows(items: list[NomRow]) -> dict[str, NomRow]:
    by_key: dict[str, NomRow] = {}
    for item in items:
        existing = _find_existing_key(item.name, by_key)
        if existing is None:
            by_key[_norm_name(item.name)] = item
            continue
        by_key[existing].merge_from(item)
    return by_key


def _collapse_duplicates(by_key: dict[str, NomRow], pairs: list[tuple[str, str]]) -> dict[str, NomRow]:
    name_to_key = {row.name: key for key, row in by_key.items()}
    for a, b in pairs:
        ka, kb = name_to_key.get(a), name_to_key.get(b)
        if not ka or not kb or ka == kb:
            continue
        if ka not in by_key or kb not in by_key:
            continue
        by_key[ka].merge_from(by_key[kb])
        del by_key[kb]
        name_to_key[by_key[ka].name] = ka

    # по сигнатуре кодов модели
    sig_map: dict[str, str] = {}
    for key, row in list(by_key.items()):
        sig = _item_signature(row.name)
        if not sig:
            continue
        if sig not in sig_map:
            sig_map[sig] = key
            continue
        root = sig_map[sig]
        if root not in by_key or key not in by_key or root == key:
            continue
        if _conflicting_models(by_key[root].name, by_key[key].name):
            continue
        if not _is_safe_same_item(by_key[root].name, by_key[key].name):
            continue
        by_key[root].merge_from(by_key[key])
        del by_key[key]

    # локальный проход safe-same
    keys = list(by_key.keys())
    removed: set[str] = set()
    for i, ka in enumerate(keys):
        if ka in removed or ka not in by_key:
            continue
        for kb in keys[i + 1 :]:
            if kb in removed or kb not in by_key:
                continue
            if _is_safe_same_item(by_key[ka].name, by_key[kb].name):
                by_key[ka].merge_from(by_key[kb])
                del by_key[kb]
                removed.add(kb)
            else:
                sa, sb = _item_signature(by_key[ka].name), _item_signature(by_key[kb].name)
                if (
                    sa
                    and sa == sb
                    and _is_safe_same_item(by_key[ka].name, by_key[kb].name)
                    and not _conflicting_models(by_key[ka].name, by_key[kb].name)
                ):
                    by_key[ka].merge_from(by_key[kb])
                    del by_key[kb]
                    removed.add(kb)
    return by_key


def _copy_sheet_values(src_ws, dst_ws) -> None:
    for r_idx, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is not None and str(value).strip() != "":
                dst_ws.cell(r_idx, c_idx, value)


def _build_workbook(
    by_key: dict[str, NomRow],
    customs_wb: openpyxl.Workbook | None,
    stats: dict[str, Any],
    all_dates: set[date],
) -> bytes:
    dates = sorted(all_dates | {d for row in by_key.values() for d in row.schedule})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График"

    headers = list(_META_HEADERS) + [d.isoformat() for d in dates]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    meta_count = len(_META_HEADERS)
    rows_sorted = sorted(by_key.values(), key=lambda r: _norm_name(r.name))
    for r_idx, item in enumerate(rows_sorted, start=2):
        ordered = item.ordered
        schedule_sum = sum(item.schedule.values())
        if ordered <= 0 and schedule_sum > 0:
            ordered = schedule_sum
        ws.cell(r_idx, 1, item.name)
        ws.cell(r_idx, 2, item.meta.get("Изделие", ""))
        ws.cell(r_idx, 3, ordered)
        if item.remainder is not None:
            ws.cell(r_idx, 4, item.remainder)
        ws.cell(r_idx, 5, item.meta.get("Дата заказа", ""))
        ws.cell(r_idx, 6, item.meta.get("Срок обработки", ""))
        ws.cell(r_idx, 7, item.meta.get("Дата спецификации", ""))
        ws.cell(r_idx, 8, item.meta.get("Дата платежа", ""))
        ws.cell(r_idx, 9, item.meta.get("Логистика до МСК", ""))
        ws.cell(r_idx, 10, item.meta.get("Логистика МСК-Ростов", ""))
        ws.cell(r_idx, 11, item.meta.get("Сроки производства", ""))
        ws.cell(r_idx, 12, item.meta.get("Сроки поставки в МСК", ""))
        for d_idx, day in enumerate(dates):
            qty = item.schedule.get(day)
            if qty:
                ws.cell(r_idx, meta_count + 1 + d_idx, qty)

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 56
    for col in range(2, meta_count + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    src = wb.create_sheet("Источник")
    src.append(["Номенклатура", "Источники", "Сумма по датам", "Заказано", "Дельта"])
    for item in rows_sorted:
        ssum = sum(item.schedule.values())
        ordered = item.ordered if item.ordered > 0 else ssum
        src.append([item.name, " | ".join(item.sources), ssum, ordered, ordered - ssum])

    chk = wb.create_sheet("Проверки")
    for key, value in stats.items():
        chk.append(
            [key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value]
        )

    if customs_wb is not None:
        for title in ("ТАМОЖНЯ", "ИТЦ В РАБОТЕ", "Реестр Заказов"):
            if title in customs_wb.sheetnames:
                dst = wb.create_sheet(title[:31])
                _copy_sheet_values(customs_wb[title], dst)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_preview_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def build_merged_schedule_preview_values(
    raw: bytes,
    *,
    sheet_name: str = "График",
) -> list[list[str]]:
    """Табличный preview объединённого графика для UI (лист «График»)."""
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            if rows:
                continue
        rows.append([_cell_preview_value(cell) for cell in row])
    wb.close()
    return rows


_MONTHS_RU = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}


def _parse_date_value(value: Any, *, default_year: int | None = None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip().lower()
    if not text:
        return None
    iso = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b", text)
    if iso:
        try:
            return date(int(iso.group(1)), int(iso.group(2)), int(iso.group(3)))
        except ValueError:
            return None
    numeric = re.search(r"\b(\d{1,2})[./-](\d{1,2})(?:[./-](\d{2,4}))?\b", text)
    if numeric:
        year = int(numeric.group(3)) if numeric.group(3) else (default_year or date.today().year)
        if year < 100:
            year += 2000
        try:
            return date(year, int(numeric.group(2)), int(numeric.group(1)))
        except ValueError:
            return None
    ru = re.search(
        r"\b(\d{1,2})\s+"
        r"(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)"
        r"(?:\s+(\d{4}))?\b",
        text,
    )
    if ru:
        try:
            return date(
                int(ru.group(3)) if ru.group(3) else (default_year or date.today().year),
                _MONTHS_RU[ru.group(2)],
                int(ru.group(1)),
            )
        except ValueError:
            return None
    return None


def _date_iso(value: date | None) -> str:
    return value.isoformat() if value is not None else ""


def _to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _clean_number(value: float) -> int | float:
    return int(value) if value == int(value) else value


def _extract_quantity_hint(*texts: str) -> float:
    joined = "\n".join(texts).lower().replace("\u00a0", " ")
    candidates: list[float] = []
    for match in re.finditer(r"(\d[\d\s.,]*)\s*(?:шт|штук|pcs|кол-?во)", joined):
        qty = _to_number(match.group(1))
        if qty > 0:
            candidates.append(qty)
    if candidates:
        return candidates[0]
    return 0.0


def _pick_option_by_quantity(
    option_dates: dict[str, dict[str, Any]],
    quantity_hint: float,
) -> tuple[date | None, float]:
    if not option_dates:
        return None, 0.0
    if quantity_hint > 0:
        for raw_date, item in option_dates.items():
            qty = _to_number(item.get("quantity"))
            if abs(qty - quantity_hint) < 0.001:
                return _parse_date_value(raw_date), qty
        larger = [
            (raw_date, _to_number(item.get("quantity")))
            for raw_date, item in option_dates.items()
            if _to_number(item.get("quantity")) >= quantity_hint
        ]
        if larger:
            raw_date, qty = sorted(larger, key=lambda item: item[1])[0]
            return _parse_date_value(raw_date), quantity_hint if quantity_hint > 0 else qty
    if len(option_dates) == 1:
        raw_date, item = next(iter(option_dates.items()))
        return _parse_date_value(raw_date), _to_number(item.get("quantity"))
    return None, 0.0


def _extract_json_object(text: str) -> dict[str, Any]:
    try:
        data = json.loads(text)
        return data if isinstance(data, dict) else {}
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{.*\}", text, flags=re.S)
    if not match:
        return {}
    try:
        data = json.loads(match.group(0))
    except json.JSONDecodeError:
        return {}
    return data if isinstance(data, dict) else {}


async def _analyze_schedule_date_change(
    *,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    manager_result: str,
    schedule_options: list[dict[str, Any]],
) -> dict[str, Any]:
    return await _analyze_schedule_changes(
        task_type=task_type,
        problem=problem,
        solution=solution,
        nomenclature=nomenclature,
        manager_result=manager_result,
        schedule_options=schedule_options,
    )


def _parse_manager_batch_mentions(
    manager_result: str,
    *,
    default_year: int | None = None,
) -> list[dict[str, Any]]:
    """Извлекает упоминания партий (дата + количество) из текста менеджера."""
    text = (manager_result or "").replace("\u00a0", " ")
    lowered = text.lower()
    batches: list[dict[str, Any]] = []
    seen_dates: set[str] = set()

    date_patterns = (
        r"\b(\d{1,2})\s+(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)(?:\s+(\d{4}))?\b",
        r"\b(\d{1,2})[./-](\d{1,2})(?:[./-](\d{2,4}))?\b",
        r"\b(\d{4})-(\d{1,2})-(\d{1,2})\b",
    )

    for pattern in date_patterns:
        for match in re.finditer(pattern, lowered):
            parsed = _parse_date_value(match.group(0), default_year=default_year)
            if parsed is None:
                continue
            iso = parsed.isoformat()
            if iso in seen_dates:
                continue
            tail = lowered[match.end() : match.end() + 120]
            qty_match = re.search(r"(\d[\d\s.,]*)\s*(?:шт|штук|pcs|кол-?во)", tail)
            remainder = any(token in tail for token in ("остат", "оставш", "остальн"))
            quantity: float | None = _to_number(qty_match.group(1)) if qty_match else None
            if quantity is not None and quantity <= 0:
                quantity = None
            batches.append(
                {
                    "date": iso,
                    "quantity": quantity,
                    "remainder": remainder and quantity is None,
                }
            )
            seen_dates.add(iso)

    if not batches:
        return batches

    batches.sort(key=lambda item: item["date"])
    return batches


def _planned_dates_from_context(
    *,
    problem: str,
    solution: str,
    schedule_options: list[dict[str, Any]],
) -> list[date]:
    context = f"{problem}\n{solution}"
    option_dates = {
        _parse_date_value(item.get("date"))
        for item in schedule_options
        if item.get("date")
    }
    option_dates.discard(None)
    found: list[date] = []
    for match in re.finditer(
        r"\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?|\d{1,2}\s+"
        r"(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)",
        context.lower(),
    ):
        parsed = _parse_date_value(match.group(0))
        if parsed and parsed in option_dates and parsed not in found:
            found.append(parsed)
    if found:
        return found

    if schedule_options:
        best = max(
            schedule_options,
            key=lambda item: _to_number(item.get("quantity")),
        )
        parsed = _parse_date_value(best.get("date"))
        if parsed:
            return [parsed]
    return []


def _normalize_schedule_change_plan(
    raw: dict[str, Any],
    *,
    schedule_options: list[dict[str, Any]],
    problem: str,
    solution: str,
    manager_result: str,
) -> dict[str, Any]:
    if not raw.get("has_change"):
        return {"has_change": False, "comment": raw.get("comment") or "Изменение не требуется."}

    option_map = {
        str(item.get("date")): _to_number(item.get("quantity"))
        for item in schedule_options
        if item.get("date")
    }
    default_year = date.today().year
    if option_map:
        first = _parse_date_value(next(iter(option_map.keys())))
        if first:
            default_year = first.year

    remove_dates: list[str] = []
    for value in raw.get("remove_dates") or []:
        parsed = _parse_date_value(value, default_year=default_year)
        if parsed and parsed.isoformat() in option_map:
            iso = parsed.isoformat()
            if iso not in remove_dates:
                remove_dates.append(iso)

    add_batches: list[dict[str, Any]] = []
    for item in raw.get("add_batches") or []:
        if not isinstance(item, dict):
            continue
        parsed = _parse_date_value(item.get("date"), default_year=default_year)
        if not parsed:
            continue
        qty_raw = item.get("quantity")
        quantity = None if qty_raw in (None, "", 0, "0") else _to_number(qty_raw)
        remainder = bool(item.get("remainder")) or str(item.get("quantity", "")).lower() in {
            "остаток",
            "remainder",
            "rest",
        }
        add_batches.append(
            {
                "date": parsed.isoformat(),
                "quantity": quantity if quantity and quantity > 0 else None,
                "remainder": remainder and (quantity is None or quantity <= 0),
            }
        )

    # Legacy single-move fallback inside same payload
    if not add_batches and raw.get("new_date"):
        parsed_new = _parse_date_value(raw.get("new_date"), default_year=default_year)
        parsed_old = _parse_date_value(raw.get("original_date"), default_year=default_year)
        if parsed_new:
            add_batches.append(
                {
                    "date": parsed_new.isoformat(),
                    "quantity": _to_number(raw.get("quantity")) or None,
                    "remainder": False,
                }
            )
        if parsed_old and parsed_old.isoformat() in option_map:
            iso = parsed_old.isoformat()
            if iso not in remove_dates:
                remove_dates.append(iso)

    if not remove_dates:
        remove_dates = [
            day.isoformat()
            for day in _planned_dates_from_context(
                problem=problem,
                solution=solution,
                schedule_options=schedule_options,
            )
        ]

    if not add_batches:
        for mention in _parse_manager_batch_mentions(manager_result, default_year=default_year):
            add_batches.append(mention)

    if not add_batches:
        return {"has_change": False, "comment": "Не удалось определить новые партии."}

    remove_total = sum(option_map.get(iso, 0.0) for iso in remove_dates)
    explicit_total = sum(
        _to_number(batch.get("quantity"))
        for batch in add_batches
        if _to_number(batch.get("quantity")) > 0
    )
    remainder_batches = [batch for batch in add_batches if batch.get("remainder")]
    if remainder_batches:
        remainder_qty = max(remove_total - explicit_total, 0.0)
        for batch in remainder_batches:
            batch["quantity"] = remainder_qty if remainder_qty > 0 else None

    for batch in add_batches:
        if batch.get("quantity") is None and not batch.get("remainder") and remove_total > 0:
            batch["quantity"] = remove_total

    add_batches = [
        batch
        for batch in add_batches
        if batch.get("date") and (_to_number(batch.get("quantity")) > 0 or batch.get("remainder"))
    ]
    if not add_batches:
        return {"has_change": False, "comment": "Нет количеств для новых партий."}

    return {
        "has_change": True,
        "remove_dates": remove_dates,
        "add_batches": add_batches,
        "comment": str(raw.get("comment") or "График обновлён по ответу менеджера."),
    }


async def _analyze_schedule_changes(
    *,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    manager_result: str,
    schedule_options: list[dict[str, Any]],
) -> dict[str, Any]:
    from app.agents.document_analysis_agent.excel_service import _lm_settings, _post_lm_json
    from app.core.config import settings

    planned = _planned_dates_from_context(
        problem=problem,
        solution=solution,
        schedule_options=schedule_options,
    )
    planned_text = ", ".join(day.isoformat() for day in planned) or "—"

    prompt = (
        "Проанализируй ответ менеджера и определи, как обновить объединённый график отгрузок.\n"
        "Нужно вернуть только JSON.\n\n"
        f"Тип задания: {task_type or '—'}\n"
        f"Номенклатура: {nomenclature or '—'}\n"
        f"Проблема (ситуация): {problem or '—'}\n"
        f"Рекомендация системы: {solution or '—'}\n"
        f"Ответ менеджера: {manager_result or '—'}\n\n"
        f"Плановые даты в графике (контекст): {planned_text}\n"
        "Текущие даты и количества по номенклатуре в графике:\n"
        f"{json.dumps(schedule_options, ensure_ascii=False)}\n\n"
        "Верни JSON:\n"
        '{"has_change":true|false,'
        '"remove_dates":["YYYY-MM-DD"],'
        '"add_batches":[{"date":"YYYY-MM-DD","quantity":number|null,"remainder":true|false}],'
        '"comment":"кратко"}\n\n'
        "Правила:\n"
        "1. has_change=true, если менеджер сообщил новые/изменённые даты поставки, отгрузки "
        "или разделение на несколько партий.\n"
        "2. remove_dates — даты из графика, которые нужно ОЧИСТИТЬ (обычно плановая дата до "
        "изменения: из проблемы/графика). Если было 40000 на 07.08, а теперь две партии — "
        "убери 07.08.\n"
        "3. add_batches — новые партии из ответа менеджера. quantity бери из текста менеджера "
        "(например 1111 штук). Если менеджер пишет «остаток/оставшаяся партия» — quantity=null, "
        "remainder=true.\n"
        "4. Не переноси всё количество одной партией, если менеджер явно указал несколько дат.\n"
        "5. Если год не указан — используй год из плановой даты или из графика.\n"
        "6. Если менеджер указал только одну новую дату — всё равно заполни remove_dates и "
        "add_batches (не legacy original_date/new_date).\n"
        "7. Если изменений нет — has_change=false.\n\n"
        "Пример: план 40000 на 2025-08-07, ответ «две партии: 1111 шт 15 августа, остаток 30 августа»:\n"
        '{"has_change":true,"remove_dates":["2025-08-07"],'
        '"add_batches":[{"date":"2025-08-15","quantity":1111,"remainder":false},'
        '{"date":"2025-08-30","quantity":null,"remainder":true}],'
        '"comment":"Поставка разделена на две партии"}'
    )
    lm = _lm_settings()
    if lm is None:
        return {"has_change": False, "comment": "LM Studio недоступна."}
    base_url, model = lm
    try:
        data = await _post_lm_json(
            base_url,
            model,
            prompt,
            timeout=min(90.0, float(settings.AVEON_LM_STUDIO_TIMEOUT_SECONDS)),
        )
        if isinstance(data, dict):
            return _normalize_schedule_change_plan(
                data,
                schedule_options=schedule_options,
                problem=problem,
                solution=solution,
                manager_result=manager_result,
            )
    except Exception:
        return {"has_change": False, "comment": "LM Studio не смогла разобрать изменение дат."}
    return {"has_change": False, "comment": "LM Studio вернула пустой ответ."}


def _heuristic_schedule_changes(
    *,
    manager_result: str,
    problem: str,
    solution: str,
    nomenclature: str = "",
    schedule_options: list[dict[str, Any]],
) -> dict[str, Any]:
    option_map = {
        str(item.get("date")): _to_number(item.get("quantity"))
        for item in schedule_options
        if item.get("date")
    }
    default_year = date.today().year
    if option_map:
        first = _parse_date_value(next(iter(option_map.keys())))
        if first:
            default_year = first.year

    add_batches = _parse_manager_batch_mentions(manager_result, default_year=default_year)
    if len(add_batches) >= 2:
        remove_dates = [
            day.isoformat()
            for day in _planned_dates_from_context(
                problem=problem,
                solution=solution,
                schedule_options=schedule_options,
            )
        ]
        return _normalize_schedule_change_plan(
            {
                "has_change": True,
                "remove_dates": remove_dates,
                "add_batches": add_batches,
                "comment": "Поставка разделена на несколько партий по ответу менеджера.",
            },
            schedule_options=schedule_options,
            problem=problem,
            solution=solution,
            manager_result=manager_result,
        )

    legacy = _heuristic_schedule_date_change(
        manager_result=manager_result,
        problem=problem,
        solution=solution,
        nomenclature=nomenclature,
        schedule_options=schedule_options,
    )
    if legacy.get("has_change"):
        return _normalize_schedule_change_plan(
            legacy,
            schedule_options=schedule_options,
            problem=problem,
            solution=solution,
            manager_result=manager_result,
        )
    return {"has_change": False, "comment": "В ответе менеджера нет изменения графика."}


def _collect_grafik_date_columns(ws) -> list[tuple[date, int]]:
    meta_count = len(_META_HEADERS)
    entries: list[tuple[date, int]] = []
    for col in range(meta_count + 1, ws.max_column + 1):
        parsed = _parse_date_value(ws.cell(1, col).value)
        if parsed:
            entries.append((parsed, col))
    return entries


def _reorder_grafik_date_columns(ws) -> dict[date, int]:
    """Переставляет колонки дат листа «График» в хронологическом порядке."""
    meta_count = len(_META_HEADERS)
    date_entries = _collect_grafik_date_columns(ws)
    if not date_entries:
        return {}

    ordered_by_col = sorted(date_entries, key=lambda item: item[1])
    ordered_by_date = sorted(date_entries, key=lambda item: item[0])
    if [day for day, _ in ordered_by_col] == [day for day, _ in ordered_by_date]:
        return {day: col for day, col in date_entries}

    max_row = ws.max_row
    col_payload: dict[date, list[tuple[Any, Any, Any]]] = {}
    for day, col in date_entries:
        col_payload[day] = [
            (
                ws.cell(row_idx, col).value,
                ws.cell(row_idx, col).fill,
                ws.cell(row_idx, col).comment,
            )
            for row_idx in range(1, max_row + 1)
        ]

    for _, col in sorted(date_entries, key=lambda item: item[1], reverse=True):
        ws.delete_cols(col)

    new_mapping: dict[date, int] = {}
    for idx, (day, _) in enumerate(ordered_by_date):
        col = meta_count + 1 + idx
        new_mapping[day] = col
        header_cell = ws.cell(1, col, day.isoformat())
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(wrap_text=True, horizontal="center")
        for row_idx in range(2, max_row + 1):
            value, fill, comment = col_payload[day][row_idx - 1]
            cell = ws.cell(row_idx, col, value)
            if fill and fill.fill_type:
                cell.fill = fill
            if comment:
                cell.comment = comment
    return new_mapping


def _heuristic_schedule_date_change(
    *,
    manager_result: str,
    problem: str,
    solution: str,
    nomenclature: str = "",
    schedule_options: list[dict[str, Any]],
) -> dict[str, Any]:
    context = f"{problem}\n{solution}"
    option_dates = {
        str(item.get("date")): item
        for item in schedule_options
        if item.get("date") and _to_number(item.get("quantity")) > 0
    }
    new_date = _parse_date_value(manager_result)
    if new_date is None:
        return {"has_change": False, "comment": "В ответе менеджера нет новой даты."}
    original_date: date | None = None
    for match in re.finditer(
        r"\d{4}-\d{1,2}-\d{1,2}|\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?|\d{1,2}\s+"
        r"(?:января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)",
        context.lower(),
    ):
        candidate = _parse_date_value(match.group(0), default_year=new_date.year)
        if candidate and candidate.isoformat() in option_dates:
            original_date = candidate
            break
    quantity_hint = _extract_quantity_hint(problem, solution, nomenclature, manager_result)
    quantity = 0.0
    if original_date is None:
        original_date, quantity = _pick_option_by_quantity(option_dates, quantity_hint)
    if original_date is None:
        return {"has_change": False, "comment": "Не удалось выбрать исходную дату графика."}
    option = option_dates.get(original_date.isoformat()) or {}
    if quantity <= 0:
        quantity = _to_number(option.get("quantity"))
    return {
        "has_change": True,
        "original_date": original_date.isoformat(),
        "new_date": new_date.isoformat(),
        "quantity": quantity,
        "comment": "Дата перенесена по ответу менеджера.",
    }


async def apply_manager_date_change_to_schedule(
    *,
    raw: bytes,
    task_type: str,
    problem: str,
    solution: str,
    nomenclature: str,
    manager_result: str,
) -> dict[str, Any]:
    """Переносит количество в листе «График» по новой дате из ответа менеджера."""
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    if "График" not in wb.sheetnames:
        wb.close()
        return {"ok": True, "applied": False, "message": "Лист «График» не найден."}

    ws = wb["График"]
    header_values = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    date_cols: dict[date, int] = {}
    for col, value in enumerate(header_values, start=1):
        parsed = _parse_date_value(value)
        if parsed:
            date_cols[parsed] = col
    if not date_cols:
        wb.close()
        return {"ok": True, "applied": False, "message": "В графике нет колонок дат."}

    target_norm = _norm_name(nomenclature)
    row_idx: int | None = None
    best_ratio = 0.0
    strong_row_match = False
    for idx in range(2, ws.max_row + 1):
        row_name = str(ws.cell(idx, 1).value or "")
        if not row_name.strip():
            continue
        row_norm = _norm_name(row_name)
        ratio = SequenceMatcher(None, target_norm, row_norm).ratio()
        if row_norm == target_norm or target_norm in row_norm or row_norm in target_norm:
            row_idx = idx
            best_ratio = max(best_ratio, ratio)
            strong_row_match = True
            break
        if ratio > best_ratio:
            best_ratio = ratio
            row_idx = idx
    if row_idx is None or (best_ratio < 0.55 and not strong_row_match):
        wb.close()
        return {"ok": True, "applied": False, "message": "Номенклатура не найдена в графике."}

    schedule_options: list[dict[str, Any]] = []
    for day, col in sorted(date_cols.items()):
        qty = _to_number(ws.cell(row_idx, col).value)
        if qty:
            schedule_options.append({"date": day.isoformat(), "quantity": _clean_number(qty)})

    lm_result = await _analyze_schedule_changes(
        task_type=task_type,
        problem=problem,
        solution=solution,
        nomenclature=nomenclature,
        manager_result=manager_result,
        schedule_options=schedule_options,
    )
    if not lm_result.get("has_change"):
        lm_result = _heuristic_schedule_changes(
            manager_result=manager_result,
            problem=problem,
            solution=solution,
            nomenclature=nomenclature,
            schedule_options=schedule_options,
        )
    if not lm_result.get("has_change"):
        wb.close()
        return {
            "ok": True,
            "applied": False,
            "message": str(lm_result.get("comment") or "Изменение дат не найдено."),
        }

    remove_dates = [
        parsed
        for iso in lm_result.get("remove_dates") or []
        if (parsed := _parse_date_value(iso)) is not None
    ]
    add_batches = lm_result.get("add_batches") or []
    if not add_batches:
        wb.close()
        return {"ok": True, "applied": False, "message": "Не удалось определить новые партии."}

    matched_name = str(ws.cell(row_idx, 1).value or nomenclature)
    changed_cells: list[dict[str, int]] = []
    fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    removed_total = 0.0
    for remove_day in remove_dates:
        col = date_cols.get(remove_day)
        if col is None:
            continue
        qty = _to_number(ws.cell(row_idx, col).value)
        if qty <= 0:
            continue
        removed_total += qty
        ws.cell(row_idx, col).value = None
        changed_cells.append({"row": row_idx - 1, "col": col - 1})

    explicit_total = sum(
        _to_number(batch.get("quantity"))
        for batch in add_batches
        if _to_number(batch.get("quantity")) > 0
    )
    for batch in add_batches:
        if batch.get("remainder") and _to_number(batch.get("quantity")) <= 0:
            remainder_qty = max(removed_total - explicit_total, 0.0)
            if remainder_qty > 0:
                batch["quantity"] = _clean_number(remainder_qty)

    applied_batches: list[dict[str, Any]] = []
    for batch in add_batches:
        batch_day = _parse_date_value(batch.get("date"))
        quantity = _to_number(batch.get("quantity"))
        if batch_day is None or quantity <= 0:
            continue
        col = date_cols.get(batch_day)
        if col is None:
            col = ws.max_column + 1
            ws.cell(1, col, batch_day.isoformat())
            ws.cell(1, col).font = Font(bold=True)
            ws.cell(1, col).alignment = Alignment(wrap_text=True, horizontal="center")
            date_cols[batch_day] = col
        existing = _to_number(ws.cell(row_idx, col).value)
        cell = ws.cell(row_idx, col, _clean_number(existing + quantity))
        cell.fill = fill
        cell.comment = Comment(
            "Партия добавлена по результату менеджера.",
            "AI Platform",
        )
        changed_cells.append({"row": row_idx - 1, "col": col - 1})
        applied_batches.append(
            {"date": batch_day.isoformat(), "quantity": _clean_number(quantity)}
        )

    if not applied_batches:
        wb.close()
        return {"ok": True, "applied": False, "message": "Не удалось применить партии к графику."}

    date_cols = _reorder_grafik_date_columns(ws)

    raw_out = io.BytesIO()
    wb.save(raw_out)
    wb.close()
    out = raw_out.getvalue()
    preview = build_merged_schedule_preview_values(out)
    return {
        "ok": True,
        "applied": True,
        "message": str(lm_result.get("comment") or "График обновлён по ответу менеджера."),
        "file_name": "merged_schedule.xlsx",
        "file_base64": base64.b64encode(out).decode("ascii"),
        "preview_values": preview,
        "changed_cells": changed_cells,
        "change": {
            "nomenclature": matched_name,
            "remove_dates": [day.isoformat() for day in remove_dates],
            "add_batches": applied_batches,
        },
    }


async def merge_schedule_files(files: list[tuple[str, bytes]]) -> dict[str, Any]:
    google_sheets_meta: dict[str, Any] = {"included": False}
    try:
        from app.services.google_sheets_client import fetch_itc_sheet_workbook_payload

        sheets_payload = fetch_itc_sheet_workbook_payload()
        if sheets_payload:
            filename, raw = sheets_payload
            files = [*files, (filename, raw)]
            google_sheets_meta = {
                "included": True,
                "file": filename,
                "source": "google_sheets",
                "sheet": "ИТЦ В РАБОТЕ",
            }
    except Exception as exc:
        google_sheets_meta = {"included": False, "error": str(exc)}

    grafik_items: list[NomRow] = []
    all_dates: set[date] = set()
    customs_wb: openpyxl.Workbook | None = None
    itc_rows: list[tuple] | None = None
    itc_layout: SheetLayout | None = None
    ingested: list[str] = []
    layouts_log: list[dict[str, Any]] = []

    workbooks: list[tuple[str, openpyxl.Workbook]] = []
    for filename, raw in files:
        if filename.startswith("~$") or filename.lower().startswith("merged_schedule"):
            continue
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            return {"ok": False, "message": f"Не читается {filename}: {exc}", "files": []}
        if "График" in wb.sheetnames and "Источник" in wb.sheetnames:
            continue
        workbooks.append((filename, wb))

    sheet_plans: list[tuple[str, str, openpyxl.Workbook, SheetLayout]] = []
    for filename, wb in workbooks:
        for sn in wb.sheetnames:
            layout = await detect_sheet_layout(wb[sn], sn, filename)
            sheet_plans.append((filename, sn, wb, layout))
            layouts_log.append(
                {
                    "file": filename,
                    "sheet": sn,
                    "kind": layout.kind,
                    "confidence": layout.confidence,
                    "reason": layout.reason,
                    "dates": len(layout.date_cols),
                }
            )
            if layout.kind == "itc":
                customs_wb = wb
                itc_rows = list(wb[sn].iter_rows(values_only=True))
                itc_layout = layout
            elif sn.strip() in ("ТАМОЖНЯ", "ИТЦ В РАБОТЕ", "Реестр Заказов"):
                customs_wb = wb
                if sn.strip() == "ИТЦ В РАБОТЕ" and itc_rows is None:
                    itc_rows = list(wb[sn].iter_rows(values_only=True))
                    if itc_layout is None:
                        itc_layout = layout if layout.kind == "itc" else _heuristic_itc_layout(itc_rows)

    nd = _build_name_dict(itc_rows or [])

    for filename, sn, wb, layout in sheet_plans:
        if layout.kind != "schedule":
            continue
        items, dates = _parse_schedule_layout(wb[sn], layout, nd)
        grafik_items.extend(items)
        all_dates |= dates
        if filename not in ingested:
            ingested.append(filename)

    if not grafik_items and not itc_rows:
        return {
            "ok": False,
            "message": "Нет данных графика/ИТЦ в загруженных файлах",
            "files": [{"name": n} for n, _ in files],
            "layouts": layouts_log,
        }

    by_key = _merge_rows(grafik_items)
    before = len(by_key)

    left_items: list[NomRow] = []
    batches: list[tuple[str, float, date | None, str]] = []
    if itc_rows and itc_layout and itc_layout.kind == "itc":
        left_items, batches = _parse_itc_with_layout(itc_rows, itc_layout, nd)
    elif itc_rows:
        fallback = _heuristic_itc_layout(itc_rows)
        if fallback:
            left_items, batches = _parse_itc_with_layout(itc_rows, fallback, nd)

    left_merged = 0
    left_new = 0
    for item in left_items:
        existing = _find_existing_key(item.name, by_key)
        if existing:
            row = by_key[existing]
            row.name = _prefer_name(row.name, item.name)
            if item.remainder is not None and row.remainder is None:
                row.remainder = item.remainder
            for k in _TEXT_META:
                if item.meta.get(k):
                    row.meta[k] = _merge_text(row.meta.get(k, ""), item.meta.get(k, ""))
            for day, qty in item.schedule.items():
                if row.schedule.get(day, 0) == 0:
                    row.add_qty(day, qty)
                    all_dates.add(day)
            if "итц:спецификация" not in row.sources:
                row.sources.append("итц:спецификация")
            left_merged += 1
        else:
            by_key[_norm_name(item.name)] = item
            all_dates |= set(item.schedule)
            left_new += 1

    queries: list[str] = []
    seen_q: set[str] = set()
    for name, _, _, _ in batches:
        if name not in seen_q:
            seen_q.add(name)
            queries.append(name)
    matches = await _resolve_matches(queries, by_key) if queries else {}

    matched_batches = 0
    new_batches = 0
    for name, qty, day, note in batches:
        target = matches.get(name)
        if target:
            key = _find_existing_key(target, by_key) or _norm_name(target)
            if key in by_key:
                row = by_key[key]
                if "итц:партия" not in row.sources:
                    row.sources.append("итц:партия")
                if note:
                    row.meta["Дата заказа"] = _merge_text(row.meta.get("Дата заказа", ""), note)
                if day is not None:
                    # партия: не задваивать, если дата уже заполнена из графика
                    if row.schedule.get(day, 0) == 0:
                        row.add_qty(day, qty)
                    all_dates.add(day)
                matched_batches += 1
                continue
        # новая строка только если нет безопасного совпадения
        existing = _find_existing_key(name, by_key)
        if existing:
            row = by_key[existing]
            if day is not None and row.schedule.get(day, 0) == 0:
                row.add_qty(day, qty)
                all_dates.add(day)
            if "итц:партия" not in row.sources:
                row.sources.append("итц:партия")
            matched_batches += 1
        else:
            row = NomRow(name=name, ordered=qty, sources=["итц:партия"])
            if note:
                row.meta["Дата заказа"] = note
            row.add_qty(day, qty)
            if day:
                all_dates.add(day)
            by_key[_norm_name(name)] = row
            new_batches += 1

    # LM + локальная склейка дублей
    pairs = await _lm_find_duplicates([r.name for r in by_key.values()])
    by_key = _collapse_duplicates(by_key, pairs)

    for row in by_key.values():
        uniq: list[str] = []
        for s in row.sources:
            if s not in uniq:
                uniq.append(s)
        row.sources = uniq

    # контроль: все исходные имена графика должны находиться
    missing_check: list[str] = []
    for item in grafik_items:
        if _find_existing_key(item.name, by_key) is None:
            # восстановить пропуск
            by_key[_norm_name(item.name)] = item
            missing_check.append(item.name)
            all_dates |= set(item.schedule)

    stats = {
        "ingested_files": ingested,
        "grafik_rows_raw": len(grafik_items),
        "grafik_rows_merged": before,
        "nomenclature_total": len(by_key),
        "itc_left_merged": left_merged,
        "itc_left_new": left_new,
        "itc_batches": len(batches),
        "itc_batches_matched": matched_batches,
        "itc_batches_new": new_batches,
        "lm_duplicate_pairs": len(pairs),
        "restored_if_missing": missing_check,
        "date_columns": len(all_dates),
        "lm_used": bool(_lm_settings()),
        "layouts_detected": layouts_log,
        "google_sheets": google_sheets_meta,
    }

    raw_out = _build_workbook(by_key, customs_wb, stats, all_dates)
    preview_values = build_merged_schedule_preview_values(raw_out)
    return {
        "ok": True,
        "message": f"Объединено номенклатур: {len(by_key)}",
        "files": [{"name": n, "size": len(b)} for n, b in files if not n.startswith("~$")],
        "file_name": "merged_schedule.xlsx",
        "file_base64": base64.b64encode(raw_out).decode("ascii"),
        "preview_values": preview_values,
        "stats": stats,
    }
