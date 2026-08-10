"""Сравнение версий детального графика производства (2 последние; план → было/стало)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.agents.document_analysis_agent.excel_service import (
    ROLE_DETAILED_PRODUCTION_SCHEDULE,
    UploadedWorkbook,
    _clean_text,
    _find_pf_report_layout,
    _infer_detailed_sheet_year_month,
    _normalize,
    _safe_set_cell_value,
    _to_float,
)
from app.agents.document_analysis_agent.production_schedule_diff import (
    ScheduleDiffResult,
    extract_schedule_version,
    rank_production_schedules,
    select_latest_production_schedules,
)
from app.core.logging import get_logger

logger = get_logger(__name__)

DETAILED_DIFF_FILE_NAME = "детальный_график_изменения.xlsx"

_DIFF_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
_DIFF_FONT = Font(color="FF9C0006", bold=True)
_HEADER_FILL = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")


@dataclass(frozen=True)
class _DetailedPlanPair:
    plan_col: int
    fact_col: int | None
    label: str  # дата или диапазон для уведомления


def _values_equal(left: float | None, right: float | None) -> bool:
    left_n = 0.0 if left is None else float(left)
    right_n = 0.0 if right is None else float(right)
    return abs(left_n - right_n) < 1e-9


def _pair_label(pair: Any) -> str:
    days = pair.days()
    if not days:
        return ""
    if len(days) == 1:
        return days[0].strftime("%d.%m.%Y")
    return f"{days[0].strftime('%d.%m.%Y')}–{days[-1].strftime('%d.%m.%Y')}"


def infer_detailed_workbook_month(
    content: bytes,
    as_of: date | None = None,
) -> tuple[int, int] | None:
    """Год и месяц детального графика из содержимого Excel."""
    workbook = load_workbook(BytesIO(content), data_only=True)
    try:
        found = _find_detailed_pf_sheet(workbook, as_of)
        if found is None:
            return None
        _sheet, year, month, *_rest = found
        if year <= 0 or month <= 0:
            return None
        return year, month
    finally:
        workbook.close()


def _find_detailed_pf_sheet(
    workbook: Any,
    as_of: date | None = None,
) -> tuple[Worksheet, int, int, int, int, list[_DetailedPlanPair]] | None:
    """Лист детального отчёта П/ф: sheet, year, month, metric_row, name_col, pairs."""
    as_of = as_of or date.today()
    best: tuple[Worksheet, int, int, int, int, list[_DetailedPlanPair]] | None = None
    best_score = -1
    for sheet in workbook.worksheets:
        year, month = _infer_detailed_sheet_year_month(sheet, as_of.year)
        if year <= 0 or month <= 0:
            year, month = as_of.year, as_of.month
        layout = _find_pf_report_layout(sheet, year, month)
        if layout is None:
            continue
        metric_row, name_col, _stage_col, pairs = layout
        detailed_pairs = [
            _DetailedPlanPair(
                plan_col=pair.plan_col,
                fact_col=pair.fact_col,
                label=_pair_label(pair),
            )
            for pair in pairs
            if pair.plan_col > 0
        ]
        if len(detailed_pairs) < 1:
            continue
        score = len(detailed_pairs)
        # предпочтение текущего/ближайшего месяца
        if year == as_of.year and month == as_of.month:
            score += 1000
        if score > best_score:
            best_score = score
            best = (sheet, year, month, metric_row, name_col, detailed_pairs)
    return best


def _plan_values_by_row(
    sheet: Worksheet,
    *,
    metric_row: int,
    name_col: int,
    plan_cols: list[int],
) -> dict[tuple[str, int], dict[int, float | None]]:
    """(product_key, row) → {plan_col: qty}; все строки с планом (П/ф, ОТК, Склад)."""
    result: dict[tuple[str, int], dict[int, float | None]] = {}
    current_product = ""
    for row_idx in range(metric_row + 1, (sheet.max_row or 0) + 1):
        name = _clean_text(sheet.cell(row_idx, name_col).value)
        if name:
            current_product = name
        if not current_product:
            continue
        key = (_normalize(current_product), row_idx)
        bucket: dict[int, float | None] = {}
        has_any = False
        for col in plan_cols:
            val = _to_float(sheet.cell(row_idx, col).value)
            bucket[col] = val
            if val is not None:
                has_any = True
        if has_any:
            result[key] = bucket
    return result


def build_detailed_schedule_diff(
    old_file: tuple[str, bytes],
    new_file: tuple[str, bytes],
    *,
    as_of: date | None = None,
) -> ScheduleDiffResult:
    """Копия новой версии детального графика: каждый «План» → было/стало."""
    as_of = as_of or date.today()
    old_name, old_bytes = old_file
    new_name, new_bytes = new_file
    _, old_label = extract_schedule_version(old_bytes)
    _, new_label = extract_schedule_version(new_bytes)
    # baseline (old_file) всегда «было», новая загрузка (new_file) — «стало».

    old_wb = load_workbook(BytesIO(old_bytes), data_only=True)
    new_wb_values = load_workbook(BytesIO(new_bytes), data_only=True)
    new_wb = load_workbook(BytesIO(new_bytes))
    try:
        old_found = _find_detailed_pf_sheet(old_wb, as_of)
        new_found_values = _find_detailed_pf_sheet(new_wb_values, as_of)
        new_found = _find_detailed_pf_sheet(new_wb, as_of)
        if old_found is None or new_found is None or new_found_values is None:
            return ScheduleDiffResult(
                has_changes=False,
                file_name=DETAILED_DIFF_FILE_NAME,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        old_sheet, _oy, _om, old_metric, old_name_col, old_pairs = old_found
        new_sheet_values, _ny, _nm, new_metric_v, new_name_col_v, new_pairs_v = new_found_values
        new_sheet, _jy, _jm, metric_row, name_col, new_pairs = new_found

        # сопоставить пары плана по подписи даты
        old_by_label = {p.label: p.plan_col for p in old_pairs if p.label}
        new_plan_items = list(new_pairs)
        if not new_plan_items:
            return ScheduleDiffResult(
                has_changes=False,
                file_name=DETAILED_DIFF_FILE_NAME,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        old_products = _plan_values_by_row(
            old_sheet,
            metric_row=old_metric,
            name_col=old_name_col,
            plan_cols=[p.plan_col for p in old_pairs],
        )
        new_products = _plan_values_by_row(
            new_sheet_values,
            metric_row=new_metric_v,
            name_col=new_name_col_v,
            plan_cols=[p.plan_col for p in new_pairs_v],
        )

        def _ordered_by_product(
            products: dict[tuple[str, int], dict[int, float | None]],
        ) -> dict[str, list[tuple[int, dict[int, float | None]]]]:
            ordered: dict[str, list[tuple[int, dict[int, float | None]]]] = {}
            for (pk, row_idx), bucket in products.items():
                ordered.setdefault(pk, []).append((row_idx, bucket))
            for pk in ordered:
                ordered[pk].sort(key=lambda item: item[0])
            return ordered

        old_by_product = _ordered_by_product(old_products)
        new_by_product = _ordered_by_product(new_products)

        changed_labels: set[str] = set()
        cell_diffs = 0
        for item in new_plan_items:
            old_col = old_by_label.get(item.label)
            for product_key, new_rows_list in new_by_product.items():
                old_rows_list = old_by_product.get(product_key) or []
                for idx, (_row_idx, new_bucket) in enumerate(new_rows_list):
                    new_val = new_bucket.get(item.plan_col)
                    old_val = None
                    if old_col is not None and idx < len(old_rows_list):
                        old_val = old_rows_list[idx][1].get(old_col)
                    if _values_equal(old_val, new_val):
                        continue
                    changed_labels.add(item.label or f"col{item.plan_col}")
                    cell_diffs += 1

        if cell_diffs == 0:
            return ScheduleDiffResult(
                has_changes=False,
                file_name=DETAILED_DIFF_FILE_NAME,
                old_filename=old_name,
                new_filename=new_name,
                old_version_label=old_label,
                new_version_label=new_label,
            )

        date_row = metric_row - 1 if metric_row > 1 else metric_row
        plan_style_cell = new_sheet.cell(metric_row, new_plan_items[0].plan_col)

        def _clear_merges() -> None:
            for rng in list(new_sheet.merged_cells.ranges):
                try:
                    new_sheet.unmerge_cells(str(rng))
                except Exception:
                    try:
                        new_sheet.merged_cells.ranges.remove(rng)
                    except Exception:
                        pass

        title_value = new_sheet.cell(1, 1).value
        _clear_merges()

        new_sheet.insert_rows(metric_row + 1)
        sub_row = metric_row + 1
        new_by_product = {
            pk: [(row + 1, bucket) for row, bucket in rows]
            for pk, rows in new_by_product.items()
        }
        _clear_merges()

        plan_orig_cols = sorted({item.plan_col for item in new_plan_items}, reverse=True)
        became_by_orig: dict[int, int] = {col: col for col in plan_orig_cols}
        for orig_col in plan_orig_cols:
            new_sheet.insert_cols(orig_col)
            for key in list(became_by_orig.keys()):
                if became_by_orig[key] >= orig_col:
                    became_by_orig[key] += 1
        _clear_merges()

        def shift_orig(orig_col: int) -> int:
            return orig_col + sum(1 for inserted in plan_orig_cols if inserted <= orig_col)

        plan_pairs: list[tuple[int, int, int | None, str]] = []
        for item in sorted(new_plan_items, key=lambda x: x.plan_col):
            became = became_by_orig[item.plan_col]
            was = became - 1
            fact_col = shift_orig(item.fact_col) if item.fact_col else None
            plan_pairs.append((was, became, fact_col, item.label))

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def _write(row: int, col: int, value: Any):
            return _safe_set_cell_value(new_sheet, row, col, value)

        def _merge(r1: int, c1: int, r2: int, c2: int) -> None:
            if r2 < r1 or c2 < c1 or (r1 == r2 and c1 == c2):
                return
            for rng in list(new_sheet.merged_cells.ranges):
                if not (
                    r2 < rng.min_row
                    or r1 > rng.max_row
                    or c2 < rng.min_col
                    or c1 > rng.max_col
                ):
                    try:
                        new_sheet.unmerge_cells(str(rng))
                    except Exception:
                        try:
                            new_sheet.merged_cells.ranges.remove(rng)
                        except Exception:
                            pass
            try:
                new_sheet.merge_cells(
                    start_row=r1, start_column=c1, end_row=r2, end_column=c2
                )
            except ValueError:
                pass

        _clear_merges()

        # подписи
        for was, became, fact_col, label in plan_pairs:
            left = _write(metric_row, was, "План")
            _write(metric_row, became, None)
            try:
                if plan_style_cell.has_style:
                    left.font = plan_style_cell.font.copy()
                    left.fill = plan_style_cell.fill.copy()
                    left.border = plan_style_cell.border.copy()
            except Exception:
                left.font = Font(bold=True)
            left.alignment = center

            was_hdr = _write(sub_row, was, "было")
            became_hdr = _write(sub_row, became, "стало")
            for hdr in (was_hdr, became_hdr):
                hdr.fill = _HEADER_FILL
                hdr.font = Font(bold=True, size=10)
                hdr.alignment = center

            if fact_col is not None:
                fact_cell = _write(metric_row, fact_col, "Факт")
                try:
                    if plan_style_cell.has_style:
                        fact_cell.font = plan_style_cell.font.copy()
                        fact_cell.fill = plan_style_cell.fill.copy()
                        fact_cell.border = plan_style_cell.border.copy()
                except Exception:
                    fact_cell.font = Font(bold=True)
                fact_cell.alignment = center
                _write(sub_row, fact_col, None)

            if date_row and date_row > 0 and label:
                # исходная дата могла быть в merge — берём label
                d_cell = _write(date_row, was, label)
                d_cell.alignment = center

        # объединения: дата → план; факт на 2 строки; наименование через шапку
        for was, became, fact_col, _label in plan_pairs:
            end_col = fact_col if fact_col is not None else became
            if date_row and date_row > 0:
                _merge(date_row, was, date_row, end_col)
            _merge(metric_row, was, metric_row, became)
            if fact_col is not None:
                _merge(metric_row, fact_col, sub_row, fact_col)

        # «Модель / изделие» на строки метрики + было/стало
        name_header_top = date_row if date_row and date_row > 0 else metric_row
        name_header_val = new_sheet.cell(metric_row, name_col).value
        if name_header_val:
            new_sheet.cell(name_header_top, name_col, name_header_val)
        _merge(name_header_top, name_col, sub_row, name_col)
        # колонка № если есть слева от имени
        if name_col > 1:
            _merge(name_header_top, name_col - 1, sub_row, name_col - 1)

        # данные: все стадии (П/ф, ОТК, Склад) — было/стало по порядку строк изделия
        pair_by_plan_orig = {
            item.plan_col: (became_by_orig[item.plan_col] - 1, became_by_orig[item.plan_col], item.label)
            for item in new_plan_items
        }
        for product_key, new_rows_list in new_by_product.items():
            old_rows_list = old_by_product.get(product_key) or []
            for idx, (row_idx, new_bucket) in enumerate(new_rows_list):
                old_bucket = old_rows_list[idx][1] if idx < len(old_rows_list) else {}
                for item in new_plan_items:
                    was_col, became_col, label = pair_by_plan_orig[item.plan_col]
                    old_col = old_by_label.get(label) if label else None
                    new_val = new_bucket.get(item.plan_col)
                    old_val = old_bucket.get(old_col) if old_col is not None else None
                    if old_val is not None:
                        new_sheet.cell(row_idx, was_col, old_val)
                    else:
                        new_sheet.cell(row_idx, was_col, None)
                    if new_val is not None:
                        new_sheet.cell(row_idx, became_col, new_val)
                    else:
                        new_sheet.cell(row_idx, became_col, None)
                    if not _values_equal(old_val, new_val):
                        for paint_col in (was_col, became_col):
                            cell = new_sheet.cell(row_idx, paint_col)
                            cell.fill = _DIFF_FILL
                            cell.font = _DIFF_FONT

        width_src = 8.0
        if plan_pairs:
            width_src = (
                new_sheet.column_dimensions[get_column_letter(plan_pairs[0][1])].width or 8
            )
        for was_col, became_col, _f, _l in plan_pairs:
            new_sheet.column_dimensions[get_column_letter(was_col)].width = width_src
            new_sheet.column_dimensions[get_column_letter(became_col)].width = width_src

        if title_value:
            new_sheet.cell(1, 1, title_value)

        legend_col = (new_sheet.max_column or 1) + 2
        new_sheet.cell(
            1,
            legend_col,
            f"Сравнение планов: было v{old_label} ({old_name}) → стало v{new_label} ({new_name})",
        )
        new_sheet.cell(2, legend_col, "Красным — ячейки плана с расхождением")
        new_sheet.cell(2, legend_col).fill = _DIFF_FILL

        buffer = BytesIO()
        new_wb.save(buffer)
        dates_ordered = sorted(changed_labels)
        logger.info(
            "detailed_schedule_diff.built",
            old=old_name,
            new=new_name,
            cells=cell_diffs,
            dates=dates_ordered,
        )
        return ScheduleDiffResult(
            has_changes=True,
            changed_months=dates_ordered,
            changed_cells=cell_diffs,
            file_name=DETAILED_DIFF_FILE_NAME,
            file_bytes=buffer.getvalue(),
            old_filename=old_name,
            new_filename=new_name,
            old_version_label=old_label,
            new_version_label=new_label,
        )
    finally:
        old_wb.close()
        new_wb_values.close()
        new_wb.close()


def select_latest_detailed_schedules(
    files: list[tuple[str, bytes]],
    *,
    keep: int = 2,
) -> dict[str, Any]:
    result = select_latest_production_schedules(files, keep=keep)
    if len(files) <= 1:
        result["message"] = "Загружен один детальный график — сравнение выполнится с сохранённой версией"
    elif result.get("removed"):
        result["message"] = str(result["message"]).replace(
            "графика производства", "детального графика производства"
        ).replace(
            "версии графика производства", "версии детального графика"
        ).replace(
            "Оставлены 2 последние версии", "Оставлена последняя версия"
        )
    else:
        kept = result.get("kept") or []
        if len(kept) == 1:
            result["message"] = (
                f"Оставлена последняя версия детального графика "
                f"{kept[0].get('version_label')} ({kept[0].get('filename')})"
            )
        elif len(kept) == 2:
            result["message"] = (
                f"Сравниваются версии детального графика "
                f"{kept[-1].get('version_label')} и {kept[0].get('version_label')}"
            )
    return result


def compare_detailed_schedule_with_snapshot(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, str],
    saved_file: tuple[str, bytes] | None,
    *,
    keep: int = 1,
    as_of: date | None = None,
) -> tuple[list[UploadedWorkbook], dict[str, str], ScheduleDiffResult | None, bool]:
    """Сравнивает новую загрузку детального графика с сохранённой версией."""
    from app.agents.document_analysis_agent.production_schedule_diff import (
        _keep_latest_schedule_workbooks,
    )

    workbooks, role_map, newest = _keep_latest_schedule_workbooks(
        workbooks,
        role_map,
        role=ROLE_DETAILED_PRODUCTION_SCHEDULE,
        keep=keep,
    )
    if newest is None:
        return workbooks, role_map, None, False
    if saved_file is None:
        return workbooks, role_map, None, False

    old_name, old_bytes = saved_file
    diff = build_detailed_schedule_diff(
        (old_name, old_bytes),
        (newest.filename, newest.content),
        as_of=as_of,
    )
    return workbooks, role_map, diff, True


def prune_workbooks_to_latest_detailed(
    workbooks: list[UploadedWorkbook],
    role_map: dict[str, str],
    *,
    keep: int = 1,
    as_of: date | None = None,
    saved_file: tuple[str, bytes] | None = None,
) -> tuple[list[UploadedWorkbook], dict[str, str], ScheduleDiffResult | None, bool]:
    """Оставляет последнюю загрузку и сравнивает её с сохранённой базовой версией."""
    return compare_detailed_schedule_with_snapshot(
        workbooks,
        role_map,
        saved_file,
        keep=keep,
        as_of=as_of,
    )
