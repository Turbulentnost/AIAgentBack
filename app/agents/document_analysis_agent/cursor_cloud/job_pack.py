"""Сборщик пакета задания для облачного агента Cursor.

Кладёт загруженные Excel и выгрузки 1С в input/, пишет manifest.json.
"""

from __future__ import annotations

import json
import re
import shutil
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from openpyxl import Workbook

from app.agents.document_analysis_agent.cursor_cloud.analysis_result import (
    ANALYSIS_RESULT_SCHEMA_ID,
    FILE_ROLES,
)
from app.agents.document_analysis_agent.dashboard_snapshot import today_msk_iso
from app.agents.document_analysis_agent.excel_service import (
    ROLE_OTHER,
    ROLE_SPECIFICATION,
    ROLE_STOCK,
    UploadedWorkbook,
    classify_aveon_excel_files,
)
from app.core.logging import get_logger

logger = get_logger(__name__)

MANIFEST_SCHEMA_ID = "aveon.cursor.job_manifest.v1"
MANIFEST_SCHEMA_VERSION = 1
STOCK_1C_FILENAME = "stock_1c.xlsx"
SPECS_1C_FILENAME = "specs_1c.xlsx"
OUTPUT_RESULT_PATH = "output/analysis_result.json"

_JOBS_DIR = Path(__file__).resolve().parents[4] / "data" / "aveon" / "cursor_jobs"
_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._\-А-Яа-яЁё ]+")
_SCHEMA_SRC = Path(__file__).with_name("analysis_result.schema.json")


@dataclass
class ManifestFile:
    path: str
    filename: str
    role: str
    source: str
    original_filename: str = ""
    size: int = 0


@dataclass
class CursorJobPack:
    job_id: str
    job_dir: Path
    as_of: str
    manifest: dict[str, Any]
    files: list[ManifestFile] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def job_root(jobs_dir: Path | None = None) -> Path:
    return jobs_dir or _JOBS_DIR


def job_dir_for(job_id: str, jobs_dir: Path | None = None) -> Path:
    cleaned = re.sub(r"[^a-zA-Z0-9_\-]+", "_", (job_id or "").strip())[:80]
    if not cleaned:
        raise ValueError("пустой job_id")
    return job_root(jobs_dir) / cleaned


def load_job_manifest(job_id: str, jobs_dir: Path | None = None) -> dict[str, Any] | None:
    path = job_dir_for(job_id, jobs_dir) / "manifest.json"
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _safe_filename(name: str) -> str:
    raw = Path(name or "file.xlsx").name.strip() or "file.xlsx"
    cleaned = _SAFE_NAME_RE.sub("_", raw).strip(" ._") or "file.xlsx"
    if "." not in cleaned:
        cleaned = f"{cleaned}.xlsx"
    return cleaned[:180]


def _unique_name(used: set[str], filename: str) -> str:
    candidate = _safe_filename(filename)
    stem = Path(candidate).stem
    suffix = Path(candidate).suffix or ".xlsx"
    index = 2
    while candidate.lower() in used:
        candidate = f"{stem}_{index}{suffix}"
        index += 1
    used.add(candidate.lower())
    return candidate


def _write_xlsx(headers: list[str], rows: list[list[Any]], sheet_name: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _stock_index_to_xlsx(stock_index: dict[str, Any]) -> bytes:
    rows = [
        [entry.nomenclature, entry.quantity if entry.quantity is not None else 0]
        for entry in stock_index.values()
        if getattr(entry, "nomenclature", "")
    ]
    rows.sort(key=lambda item: str(item[0]))
    return _write_xlsx(["Номенклатура", "Остаток"], rows, "Остатки")


def _specs_to_xlsx(catalog: list[Any], materials: list[Any]) -> bytes:
    workbook = Workbook()
    catalog_sheet = workbook.active
    catalog_sheet.title = "Спецификации"
    catalog_sheet.append(["ref_key", "Код", "Изделие", "Описание", "Метка"])
    for entry in catalog:
        catalog_sheet.append(
            [
                getattr(entry, "ref_key", ""),
                getattr(entry, "code", ""),
                getattr(entry, "main_product_name", ""),
                getattr(entry, "description", ""),
                getattr(entry, "label", ""),
            ]
        )

    materials_sheet = workbook.create_sheet("Материалы")
    materials_sheet.append(
        ["spec_ref_key", "Изделие", "Номенклатура", "Количество", "Ед. изм.", "Код"]
    )
    for row in materials:
        materials_sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def _export_db_workbooks(db: Any) -> tuple[list[UploadedWorkbook], list[str], list[str]]:
    """Остатки и спецификации из 1С → Excel для input/."""
    from sqlalchemy import select

    from app.agents.document_analysis_agent.onec_db_sources import (
        build_stock_index_from_db,
        load_db_spec_catalog,
    )
    from app.models.onec_resource_spec import OnecResourceSpecMaterial

    extras: list[UploadedWorkbook] = []
    roles: list[str] = []
    warnings: list[str] = []

    stock_index = await build_stock_index_from_db(db)
    if stock_index:
        extras.append(
            UploadedWorkbook(filename=STOCK_1C_FILENAME, content=_stock_index_to_xlsx(stock_index))
        )
        roles.append(ROLE_STOCK)
    else:
        warnings.append("Остатки 1С пустые — stock_1c.xlsx не добавлен")

    catalog = await load_db_spec_catalog(db)
    material_rows: list[list[Any]] = []
    if catalog:
        product_by_ref = {
            entry.ref_key: entry.main_product_name or entry.label for entry in catalog
        }
        materials = (await db.execute(select(OnecResourceSpecMaterial))).scalars().all()
        for item in materials:
            name = (item.nomenclature_name or "").strip()
            if not name:
                continue
            material_rows.append(
                [
                    item.spec_ref_key,
                    product_by_ref.get(item.spec_ref_key, ""),
                    name,
                    float(item.qty or 0),
                    (item.unit or "").strip(),
                    (item.nomenclature_code or "").strip(),
                ]
            )
        extras.append(
            UploadedWorkbook(filename=SPECS_1C_FILENAME, content=_specs_to_xlsx(catalog, material_rows))
        )
        roles.append(ROLE_SPECIFICATION)
    else:
        warnings.append("Спецификации 1С пустые — specs_1c.xlsx не добавлен")

    return extras, roles, warnings


def _write_input_file(input_dir: Path, filename: str, content: bytes) -> int:
    path = input_dir / filename
    path.write_bytes(content)
    return path.stat().st_size


async def pack_aveon_cursor_job(
    workbooks: list[UploadedWorkbook],
    *,
    role_map: dict[str, str] | None = None,
    db: Any | None = None,
    user_id: UUID | str | None = None,
    as_of: date | str | None = None,
    jobs_dir: Path | None = None,
    include_db_exports: bool = True,
    job_id: str | None = None,
) -> CursorJobPack:
    """Собирает jobs/<id>/input + manifest.json. Без файлов — ошибка."""
    if not workbooks and not (include_db_exports and db is not None):
        raise ValueError("Нет файлов для пакета Cursor")

    resolved_as_of = (
        as_of.isoformat() if isinstance(as_of, date) else (as_of or today_msk_iso())
    )
    resolved_id = (job_id or uuid4().hex).strip()
    root = job_dir_for(resolved_id, jobs_dir)
    input_dir = root / "input"
    output_dir = root / "output"
    if root.exists():
        shutil.rmtree(root)
    input_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)

    uploaded_roles = dict(role_map or {})
    if workbooks and not uploaded_roles:
        uploaded_roles, _source = await classify_aveon_excel_files(workbooks, use_lm=False)

    warnings: list[str] = []
    used_names: set[str] = {STOCK_1C_FILENAME, SPECS_1C_FILENAME}
    files: list[ManifestFile] = []

    for workbook in workbooks:
        stored_name = _unique_name(used_names, workbook.filename)
        size = _write_input_file(input_dir, stored_name, workbook.content)
        role = uploaded_roles.get(workbook.filename, ROLE_OTHER)
        if role not in FILE_ROLES:
            role = ROLE_OTHER
        files.append(
            ManifestFile(
                path=f"input/{stored_name}",
                filename=stored_name,
                role=role,
                source="upload",
                original_filename=workbook.filename,
                size=size,
            )
        )

    if include_db_exports and db is not None:
        try:
            extras, extra_roles, db_warnings = await _export_db_workbooks(db)
            warnings.extend(db_warnings)
            for workbook, role in zip(extras, extra_roles, strict=True):
                size = _write_input_file(input_dir, workbook.filename, workbook.content)
                files.append(
                    ManifestFile(
                        path=f"input/{workbook.filename}",
                        filename=workbook.filename,
                        role=role,
                        source="1c",
                        original_filename=workbook.filename,
                        size=size,
                    )
                )
        except Exception as exc:
            logger.warning("cursor_job_pack.db_export_failed", error=str(exc))
            warnings.append(f"Не удалось выгрузить 1С: {exc}")

    if not files:
        shutil.rmtree(root, ignore_errors=True)
        raise ValueError("Пакет пустой: нет загруженных файлов и выгрузок 1С")

    if _SCHEMA_SRC.is_file():
        shutil.copy2(_SCHEMA_SRC, root / "analysis_result.schema.json")

    manifest = {
        "schema_id": MANIFEST_SCHEMA_ID,
        "schema_version": MANIFEST_SCHEMA_VERSION,
        "job_id": resolved_id,
        "as_of": resolved_as_of,
        "user_id": str(user_id) if user_id is not None else None,
        "result_schema": "analysis_result.schema.json",
        "result_schema_id": ANALYSIS_RESULT_SCHEMA_ID,
        "output_path": OUTPUT_RESULT_PATH,
        "files": [
            {
                "path": item.path,
                "filename": item.filename,
                "original_filename": item.original_filename,
                "role": item.role,
                "source": item.source,
                "size": item.size,
            }
            for item in files
        ],
        "warnings": warnings,
    }
    (root / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    logger.info(
        "cursor_job_pack.ready",
        job_id=resolved_id,
        files=len(files),
        warnings=len(warnings),
    )
    return CursorJobPack(
        job_id=resolved_id,
        job_dir=root,
        as_of=resolved_as_of,
        manifest=manifest,
        files=files,
        warnings=warnings,
    )
